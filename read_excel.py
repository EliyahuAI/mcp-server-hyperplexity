
import openpyxl
import os

# Construct the absolute path to the Excel file
file_name = "RPT Programs List - 08.04.25 vShared_input_input_v3_preview_enhanced.xlsx"
dir_path = "/mnt/c/Users/ellio/OneDrive - Eliyahu.AI/Desktop/src/perplexityValidator/temp_unnecessary_files"
file_path = os.path.join(dir_path, file_name)

try:
    # Load workbook to get formulas
    workbook_formulas = openpyxl.load_workbook(file_path, data_only=False)
    # Load workbook to get last calculated values
    workbook_values = openpyxl.load_workbook(file_path, data_only=True)

    print(f"Reading formulas from: {file_path}\n")

    # Iterate over all the sheets
    for sheet_name in workbook_formulas.sheetnames:
        print(f"--- Sheet: {sheet_name} ---")
        sheet_formulas = workbook_formulas[sheet_name]
        sheet_values = workbook_values[sheet_name]

        found_formulas = False
        # Iterate over all the rows and columns in the formula sheet
        for row in sheet_formulas.iter_rows():
            for cell in row:
                # Check if the cell has a formula
                if cell.data_type == 'f':
                    formula = cell.value
                    value = sheet_values[cell.coordinate].value
                    print(f"Cell {cell.coordinate}: Formula='{formula}', Value='{value}'")
                    found_formulas = True
        
        if not found_formulas:
            print("No formulas found in this sheet.")
        print("\n")


except FileNotFoundError:
    print(f"Error: File not found at '{file_path}'")
    print("Please ensure the file name and path are correct.")
except Exception as e:
    print(f"An error occurred: {e}")
