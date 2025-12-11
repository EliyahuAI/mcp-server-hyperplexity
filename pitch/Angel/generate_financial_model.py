#!/usr/bin/env python3
"""
Generate Excel Financial Model for Hyperplexity
Creates a comprehensive financial model with multiple sheets and scenarios
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_financial_model():
    """Create comprehensive financial model workbook"""

    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_assumptions_sheet(wb)
    create_revenue_model_sheet(wb)
    create_cost_model_sheet(wb)
    create_cash_flow_sheet(wb)
    create_scenarios_sheet(wb)
    create_unit_economics_sheet(wb)
    create_summary_sheet(wb)

    # Save workbook
    filename = "Hyperplexity_Financial_Model.xlsx"
    wb.save(filename)
    print(f"[SUCCESS] Created {filename}")
    return filename

def apply_header_style(cell):
    """Apply consistent header styling"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_section_header_style(cell):
    """Apply section header styling"""
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def apply_currency_format(cell):
    """Apply currency formatting"""
    cell.number_format = '$#,##0'

def apply_percentage_format(cell):
    """Apply percentage formatting"""
    cell.number_format = '0.0%'

def create_assumptions_sheet(wb):
    """Create assumptions/inputs sheet"""
    ws = wb.create_sheet("Assumptions")

    # Title
    ws['A1'] = "HYPERPLEXITY - FINANCIAL MODEL ASSUMPTIONS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    # Funding assumptions
    ws['A3'] = "FUNDING"
    apply_section_header_style(ws['A3'])
    ws['A4'] = "Raise Amount"
    ws['B4'] = 300000
    apply_currency_format(ws['B4'])
    ws['A5'] = "Valuation Cap"
    ws['B5'] = 6000000
    apply_currency_format(ws['B5'])
    ws['A6'] = "Initial Cash"
    ws['B6'] = 50000
    apply_currency_format(ws['B6'])
    ws['A7'] = "Total Starting Capital"
    ws['B7'] = "=B4+B6"
    apply_currency_format(ws['B7'])

    # D2C Assumptions
    ws['A9'] = "D2C ASSUMPTIONS"
    apply_section_header_style(ws['A9'])

    ws['A10'] = "Cost per Visit"
    ws['B10'] = 1.5
    apply_currency_format(ws['B10'])

    ws['A11'] = "Page to Preview Conversion"
    ws['B11'] = 0.10
    apply_percentage_format(ws['B11'])

    ws['A12'] = "Preview to Paid Conversion"
    ws['B12'] = 0.15
    apply_percentage_format(ws['B12'])

    ws['A13'] = "Overall Funnel Conversion"
    ws['B13'] = "=B11*B12"
    apply_percentage_format(ws['B13'])

    ws['A14'] = "Average Revenue per User (ARPU)"
    ws['B14'] = 40
    apply_currency_format(ws['B14'])

    ws['A15'] = "API Cost % of Revenue"
    ws['B15'] = 0.33
    apply_percentage_format(ws['B15'])

    ws['A16'] = "Gross Margin %"
    ws['B16'] = "=1-B15"
    apply_percentage_format(ws['B16'])

    ws['A17'] = "Customer Retention (24mo)"
    ws['B17'] = 0.60
    apply_percentage_format(ws['B17'])

    ws['A18'] = "Target CAC"
    ws['B18'] = "=B10/(B11*B12)"
    apply_currency_format(ws['B18'])

    ws['A19'] = "LTV (24 months)"
    ws['B19'] = "=B14*24*B17"
    apply_currency_format(ws['B19'])

    ws['A20'] = "LTV/CAC Ratio"
    ws['B20'] = "=B19/B18"
    ws['B20'].number_format = '0.0x'

    ws['A21'] = "Payback Period (months)"
    ws['B21'] = "=B18/(B14*(1-B15))"
    ws['B21'].number_format = '0.0'

    # B2B Assumptions
    ws['A23'] = "B2B ASSUMPTIONS"
    apply_section_header_style(ws['A23'])

    ws['A24'] = "Average Setup Fee"
    ws['B24'] = 25000
    apply_currency_format(ws['B24'])

    ws['A25'] = "Average Monthly MRR"
    ws['B25'] = 8000
    apply_currency_format(ws['B25'])

    ws['A26'] = "B2B Gross Margin %"
    ws['B26'] = 0.72
    apply_percentage_format(ws['B26'])

    # Monthly Ad Spend Schedule
    ws['A28'] = "MONTHLY AD SPEND SCHEDULE"
    apply_section_header_style(ws['A28'])

    months = ['Month 1', 'Month 2', 'Month 3', 'Month 4', 'Month 5', 'Month 6']
    ad_spend = [15000, 18000, 20000, 20000, 20000, 20000]

    for i, (month, spend) in enumerate(zip(months, ad_spend)):
        ws[f'A{29+i}'] = month
        ws[f'B{29+i}'] = spend
        apply_currency_format(ws[f'B{29+i}'])

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

def create_revenue_model_sheet(wb):
    """Create revenue projections sheet"""
    ws = wb.create_sheet("Revenue Model")

    # Title
    ws['A1'] = "REVENUE MODEL - AGGRESSIVE CASE"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    # Headers
    headers = ['Metric', 'Month 1', 'Month 2', 'Month 3', 'Month 4', 'Month 5', 'Month 6']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_header_style(cell)

    # D2C Section
    ws['A5'] = "D2C BUSINESS"
    apply_section_header_style(ws['A5'])

    # D2C data - starting values
    d2c_customers = [100, 250, 400, 550, 700, 850]
    d2c_arpu = 40

    row = 6
    ws[f'A{row}'] = "D2C Paying Customers"
    for col, customers in enumerate(d2c_customers, 2):
        ws.cell(row=row, column=col).value = customers

    row = 7
    ws[f'A{row}'] = "D2C ARPU"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = d2c_arpu
        apply_currency_format(ws.cell(row=row, column=col))

    row = 8
    ws[f'A{row}'] = "D2C MRR"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"={get_column_letter(col)}6*{get_column_letter(col)}7"
        apply_currency_format(ws.cell(row=row, column=col))

    # B2B Section
    ws['A10'] = "B2B BUSINESS"
    apply_section_header_style(ws['A10'])

    row = 11
    ws[f'A{row}'] = "B2B Contracts"
    b2b_contracts = [0, 0, 1, 1, 2, 2]
    for col, contracts in enumerate(b2b_contracts, 2):
        ws.cell(row=row, column=col).value = contracts

    row = 12
    ws[f'A{row}'] = "B2B MRR per Contract"
    b2b_mrr_values = [0, 0, 5000, 8000, 6000, 9000]  # Average per contract
    for col, mrr in enumerate(b2b_mrr_values, 2):
        ws.cell(row=row, column=col).value = mrr
        apply_currency_format(ws.cell(row=row, column=col))

    row = 13
    ws[f'A{row}'] = "B2B Total MRR"
    b2b_total_mrr = [0, 0, 5000, 8000, 12000, 18000]
    for col, mrr in enumerate(b2b_total_mrr, 2):
        ws.cell(row=row, column=col).value = mrr
        apply_currency_format(ws.cell(row=row, column=col))

    # Total Revenue
    ws['A15'] = "TOTAL REVENUE"
    apply_section_header_style(ws['A15'])

    row = 16
    ws[f'A{row}'] = "Total MRR"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"={get_column_letter(col)}8+{get_column_letter(col)}13"
        apply_currency_format(ws.cell(row=row, column=col))
        ws.cell(row=row, column=col).font = Font(bold=True)

    row = 17
    ws[f'A{row}'] = "Cumulative Revenue"
    for col in range(2, 8):
        if col == 2:
            ws.cell(row=row, column=col).value = f"={get_column_letter(col)}16"
        else:
            ws.cell(row=row, column=col).value = f"={get_column_letter(col-1)}17+{get_column_letter(col)}16"
        apply_currency_format(ws.cell(row=row, column=col))

    # Growth metrics
    ws['A19'] = "GROWTH METRICS"
    apply_section_header_style(ws['A19'])

    row = 20
    ws[f'A{row}'] = "MoM Growth %"
    for col in range(2, 8):
        if col == 2:
            ws.cell(row=row, column=col).value = "N/A"
        else:
            ws.cell(row=row, column=col).value = f"=({get_column_letter(col)}16-{get_column_letter(col-1)}16)/{get_column_letter(col-1)}16"
            apply_percentage_format(ws.cell(row=row, column=col))

    row = 21
    ws[f'A{row}'] = "New D2C Customers"
    for col in range(2, 8):
        if col == 2:
            ws.cell(row=row, column=col).value = f"={get_column_letter(col)}6"
        else:
            ws.cell(row=row, column=col).value = f"={get_column_letter(col)}6-{get_column_letter(col-1)}6"
        ws.cell(row=row, column=col).number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14

def create_cost_model_sheet(wb):
    """Create cost structure sheet"""
    ws = wb.create_sheet("Cost Model")

    # Title
    ws['A1'] = "COST MODEL - MONTHLY EXPENSES"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    # Headers
    headers = ['Category', 'Month 1', 'Month 2', 'Month 3', 'Month 4', 'Month 5', 'Month 6']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_header_style(cell)

    # Team costs
    ws['A5'] = "TEAM"
    apply_section_header_style(ws['A5'])

    row = 6
    ws[f'A{row}'] = "Founder Salary"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = 12500
        apply_currency_format(ws.cell(row=row, column=col))

    row = 7
    ws[f'A{row}'] = "Marketing Support"
    costs = [4000, 4000, 5000, 5500, 6000, 6000]
    for col, cost in enumerate(costs, 2):
        ws.cell(row=row, column=col).value = cost
        apply_currency_format(ws.cell(row=row, column=col))

    row = 8
    ws[f'A{row}'] = "Direct Hire (BD/Sales)"
    costs = [0, 0, 10000, 10000, 10000, 10000]
    for col, cost in enumerate(costs, 2):
        ws.cell(row=row, column=col).value = cost
        apply_currency_format(ws.cell(row=row, column=col))

    row = 9
    ws[f'A{row}'] = "Total Team"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}8)"
        apply_currency_format(ws.cell(row=row, column=col))
        ws.cell(row=row, column=col).font = Font(bold=True)

    # Operations
    ws['A11'] = "OPERATIONS"
    apply_section_header_style(ws['A11'])

    row = 12
    ws[f'A{row}'] = "API Costs"
    costs = [3000, 4000, 6000, 8000, 10000, 12000]
    for col, cost in enumerate(costs, 2):
        ws.cell(row=row, column=col).value = cost
        apply_currency_format(ws.cell(row=row, column=col))

    row = 13
    ws[f'A{row}'] = "Infrastructure"
    costs = [1000, 1500, 2000, 2500, 2500, 3000]
    for col, cost in enumerate(costs, 2):
        ws.cell(row=row, column=col).value = cost
        apply_currency_format(ws.cell(row=row, column=col))

    row = 14
    ws[f'A{row}'] = "Other OpEx"
    costs = [5500, 5500, 6000, 6500, 6500, 7000]
    for col, cost in enumerate(costs, 2):
        ws.cell(row=row, column=col).value = cost
        apply_currency_format(ws.cell(row=row, column=col))

    row = 15
    ws[f'A{row}'] = "Total Operations"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"=SUM({get_column_letter(col)}12:{get_column_letter(col)}14)"
        apply_currency_format(ws.cell(row=row, column=col))
        ws.cell(row=row, column=col).font = Font(bold=True)

    # Marketing
    ws['A17'] = "MARKETING"
    apply_section_header_style(ws['A17'])

    row = 18
    ws[f'A{row}'] = "Ad Spend"
    ad_spend = [15000, 18000, 20000, 20000, 20000, 20000]
    for col, spend in enumerate(ad_spend, 2):
        ws.cell(row=row, column=col).value = spend
        apply_currency_format(ws.cell(row=row, column=col))

    # Total Monthly Burn
    ws['A20'] = "TOTAL MONTHLY COSTS"
    apply_section_header_style(ws['A20'])

    row = 21
    ws[f'A{row}'] = "Total Monthly Burn"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"={get_column_letter(col)}9+{get_column_letter(col)}15+{get_column_letter(col)}18"
        apply_currency_format(ws.cell(row=row, column=col))
        ws.cell(row=row, column=col).font = Font(bold=True, color="FF0000")

    # Cost breakdown %
    ws['A23'] = "COST BREAKDOWN %"
    apply_section_header_style(ws['A23'])

    row = 24
    ws[f'A{row}'] = "Team % of Total"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"={get_column_letter(col)}9/{get_column_letter(col)}21"
        apply_percentage_format(ws.cell(row=row, column=col))

    row = 25
    ws[f'A{row}'] = "Operations % of Total"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"={get_column_letter(col)}15/{get_column_letter(col)}21"
        apply_percentage_format(ws.cell(row=row, column=col))

    row = 26
    ws[f'A{row}'] = "Marketing % of Total"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"={get_column_letter(col)}18/{get_column_letter(col)}21"
        apply_percentage_format(ws.cell(row=row, column=col))

    # Column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14

def create_cash_flow_sheet(wb):
    """Create cash flow and runway analysis"""
    ws = wb.create_sheet("Cash Flow")

    # Title
    ws['A1'] = "CASH FLOW & RUNWAY ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    # Headers
    headers = ['Metric', 'Month 1', 'Month 2', 'Month 3', 'Month 4', 'Month 5', 'Month 6']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_header_style(cell)

    # Revenue
    ws['A5'] = "REVENUE"
    apply_section_header_style(ws['A5'])

    row = 6
    ws[f'A{row}'] = "Monthly Revenue"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"='Revenue Model'!{get_column_letter(col)}16"
        apply_currency_format(ws.cell(row=row, column=col))

    # Costs
    ws['A8'] = "COSTS"
    apply_section_header_style(ws['A8'])

    row = 9
    ws[f'A{row}'] = "Monthly Costs"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"='Cost Model'!{get_column_letter(col)}21"
        apply_currency_format(ws.cell(row=row, column=col))

    # Net burn
    ws['A11'] = "NET POSITION"
    apply_section_header_style(ws['A11'])

    row = 12
    ws[f'A{row}'] = "Monthly Net"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"={get_column_letter(col)}6-{get_column_letter(col)}9"
        apply_currency_format(ws.cell(row=row, column=col))
        # Red if negative
        ws.cell(row=row, column=col).font = Font(bold=True, color="FF0000")

    row = 13
    ws[f'A{row}'] = "Cumulative Cash Burn"
    for col in range(2, 8):
        if col == 2:
            ws.cell(row=row, column=col).value = f"={get_column_letter(col)}12"
        else:
            ws.cell(row=row, column=col).value = f"={get_column_letter(col-1)}13+{get_column_letter(col)}12"
        apply_currency_format(ws.cell(row=row, column=col))

    row = 14
    ws[f'A{row}'] = "Cumulative Revenue"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"='Revenue Model'!{get_column_letter(col)}17"
        apply_currency_format(ws.cell(row=row, column=col))

    # Runway
    ws['A16'] = "RUNWAY ANALYSIS"
    apply_section_header_style(ws['A16'])

    row = 17
    ws[f'A{row}'] = "Starting Capital"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = "=Assumptions!B7"
        apply_currency_format(ws.cell(row=row, column=col))

    row = 18
    ws[f'A{row}'] = "Cash Remaining"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"={get_column_letter(col)}17+{get_column_letter(col)}13"
        apply_currency_format(ws.cell(row=row, column=col))
        ws.cell(row=row, column=col).font = Font(bold=True)

    row = 19
    ws[f'A{row}'] = "Months of Runway Remaining"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f"=ABS({get_column_letter(col)}18/{get_column_letter(col)}12)"
        ws.cell(row=row, column=col).number_format = '0.0'
        # Red if < 3 months
        ws.cell(row=row, column=col).font = Font(bold=True)

    row = 20
    ws[f'A{row}'] = "Cash Flow Positive?"
    for col in range(2, 8):
        ws.cell(row=row, column=col).value = f'=IF({get_column_letter(col)}12>=0,"YES","NO")'
        ws.cell(row=row, column=col).font = Font(bold=True)

    # Key metrics
    ws['A22'] = "KEY METRICS"
    apply_section_header_style(ws['A22'])

    row = 23
    ws[f'A{row}'] = "Burn Rate (Avg)"
    ws['B23'] = "=AVERAGE(B9:G9)"
    apply_currency_format(ws['B23'])
    ws.merge_cells('B23:G23')

    row = 24
    ws[f'A{row}'] = "Total Capital Deployed"
    ws['B24'] = "=G13"
    apply_currency_format(ws['B24'])
    ws.merge_cells('B24:G24')

    row = 25
    ws[f'A{row}'] = "Final MRR"
    ws['B25'] = "=G6"
    apply_currency_format(ws['B25'])
    ws.merge_cells('B25:G25')

    row = 26
    ws[f'A{row}'] = "Total Revenue Generated"
    ws['B26'] = "=G14"
    apply_currency_format(ws['B26'])
    ws.merge_cells('B26:G26')

    # Column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14

def create_scenarios_sheet(wb):
    """Create scenario comparison sheet"""
    ws = wb.create_sheet("Scenarios")

    # Title
    ws['A1'] = "SCENARIO ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    # Month 6 comparison
    ws['A3'] = "MONTH 6 OUTCOMES"
    apply_section_header_style(ws['A3'])
    ws.merge_cells('A3:D3')

    # Headers
    headers = ['Metric', 'Best Case', 'Aggressive (Base)', 'Conservative']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        apply_header_style(cell)

    # Scenarios data
    scenarios = {
        'D2C Customers': [1100, 850, 50],
        'B2B Contracts': [3, 2, 1],
        'Total MRR': [65000, 52000, 7000],
        'Monthly Burn': [75000, 70500, 45000],
        'Months Runway': [1.6, 1.7, 3.8],
    }

    row = 5
    for metric, values in scenarios.items():
        ws[f'A{row}'] = metric
        for col, value in enumerate(values, 2):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if 'MRR' in metric or 'Burn' in metric:
                apply_currency_format(cell)
            elif 'Runway' in metric:
                cell.number_format = '0.0'
        row += 1

    # Decision matrix
    ws['A12'] = "DECISION MATRIX"
    apply_section_header_style(ws['A12'])
    ws.merge_cells('A12:D12')

    # Headers
    headers = ['Outcome', 'MRR Range', 'Next Step', 'Funding Need']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col)
        cell.value = header
        apply_header_style(cell)

    decisions = [
        ['Success', '$52K+', 'Series A raise', '$3-5M'],
        ['Strong', '$35-50K', 'Strong bridge round', '$500K-1M+'],
        ['Moderate', '$20-35K', 'Bridge round, optimize', '$500K'],
        ['Pivot', '<$20K', 'Pivot to B2B focus', '$300-500K'],
    ]

    row = 14
    for decision in decisions:
        for col, value in enumerate(decision, 1):
            ws.cell(row=row, column=col).value = value
        row += 1

    # Conversion sensitivity
    ws['A20'] = "CONVERSION RATE SENSITIVITY"
    apply_section_header_style(ws['A20'])
    ws.merge_cells('A20:E20')

    # Headers
    ws['A21'] = "Page→Preview %"
    ws['B21'] = "Preview→Paid %"
    ws['C21'] = "Overall Funnel %"
    ws['D21'] = "Est. CAC"
    ws['E21'] = "Month 6 Customers"

    for col in range(1, 6):
        apply_header_style(ws.cell(row=21, column=col))

    # Sensitivity scenarios
    sensitivity = [
        [0.08, 0.12, 0.0096, 156, 680],
        [0.10, 0.15, 0.0150, 100, 850],  # Base case
        [0.12, 0.18, 0.0216, 69, 1020],
        [0.15, 0.20, 0.0300, 50, 1300],
    ]

    row = 22
    for data in sensitivity:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if col <= 3:
                apply_percentage_format(cell)
            elif col == 4:
                apply_currency_format(cell)
            # Highlight base case
            if row == 23:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

def create_unit_economics_sheet(wb):
    """Create unit economics analysis"""
    ws = wb.create_sheet("Unit Economics")

    # Title
    ws['A1'] = "UNIT ECONOMICS ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')

    # D2C Unit Economics
    ws['A3'] = "D2C UNIT ECONOMICS"
    apply_section_header_style(ws['A3'])
    ws.merge_cells('A3:C3')

    metrics = [
        ['ARPU (Monthly)', '=Assumptions!B14', True],
        ['API Cost per Customer', '=Assumptions!B14*Assumptions!B15', True],
        ['Gross Profit per Customer', '=B5-B6', True],
        ['Gross Margin %', '=B7/B5', False],
        ['', '', False],
        ['Cost per Visit', '=Assumptions!B10', True],
        ['Page→Preview Conversion', '=Assumptions!B11', False],
        ['Preview→Paid Conversion', '=Assumptions!B12', False],
        ['Overall Funnel %', '=Assumptions!B13', False],
        ['CAC', '=Assumptions!B18', True],
        ['', '', False],
        ['LTV (24 months)', '=Assumptions!B19', True],
        ['LTV/CAC Ratio', '=Assumptions!B20', False],
        ['Payback Period (months)', '=Assumptions!B21', False],
        ['', '', False],
        ['Assessment', '', False],
    ]

    row = 4
    for metric, formula, is_currency in metrics:
        ws[f'A{row}'] = metric
        if formula:
            ws[f'B{row}'] = formula
            if is_currency:
                apply_currency_format(ws[f'B{row}'])
            else:
                if 'Ratio' in metric:
                    ws[f'B{row}'].number_format = '0.0x'
                elif 'months' in metric:
                    ws[f'B{row}'].number_format = '0.0'
                else:
                    apply_percentage_format(ws[f'B{row}'])
        row += 1

    ws['A20'] = "Target: LTV/CAC > 3x"
    ws['B20'] = "=IF(B16>3,'PASS','NEEDS IMPROVEMENT')"
    ws['B20'].font = Font(bold=True)

    ws['A21'] = "Target: Payback < 12 months"
    ws['B21'] = "=IF(B17<12,'PASS','NEEDS IMPROVEMENT')"
    ws['B21'].font = Font(bold=True)

    # B2B Unit Economics
    ws['A24'] = "B2B UNIT ECONOMICS"
    apply_section_header_style(ws['A24'])
    ws.merge_cells('A24:C24')

    b2b_metrics = [
        ['Average Setup Fee', '=Assumptions!B24', True],
        ['Average Monthly MRR', '=Assumptions!B25', True],
        ['Year 1 Revenue', '=B26+B27*12', True],
        ['Gross Margin %', '=Assumptions!B26', False],
        ['', '', False],
        ['Sales & Marketing CAC', 25000, True],
        ['', '', False],
        ['3-Year LTV', '=B27*36*B29', True],
        ['LTV/CAC Ratio', '=B33/B31', False],
        ['Payback Period (months)', '=B31/(B27*B29)', False],
    ]

    row = 25
    for metric, formula, is_currency in b2b_metrics:
        ws[f'A{row}'] = metric
        if formula:
            if isinstance(formula, str) and formula.startswith('='):
                ws[f'B{row}'] = formula
            else:
                ws[f'B{row}'] = formula
            if is_currency:
                apply_currency_format(ws[f'B{row}'])
            else:
                if 'Ratio' in metric:
                    ws[f'B{row}'].number_format = '0.0x'
                elif 'months' in metric or 'Period' in metric:
                    ws[f'B{row}'].number_format = '0.0'
                else:
                    apply_percentage_format(ws[f'B{row}'])
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

def create_summary_sheet(wb):
    """Create executive summary dashboard"""
    ws = wb.create_sheet("Executive Summary", 0)  # Make it first sheet

    # Title
    ws['A1'] = "HYPERPLEXITY - FINANCIAL MODEL"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws.merge_cells('A1:D1')

    ws['A2'] = "6-Month Aggressive Growth Plan"
    ws['A2'].font = Font(size=12)
    ws.merge_cells('A2:D2')

    # Key Highlights
    ws['A4'] = "KEY METRICS - MONTH 6"
    apply_section_header_style(ws['A4'])
    ws.merge_cells('A4:D4')

    highlights = [
        ['Total MRR', "='Revenue Model'!G16", True],
        ['D2C Customers', "='Revenue Model'!G6", False],
        ['B2B Contracts', "='Revenue Model'!G11", False],
        ['Monthly Burn', "='Cost Model'!G21", True],
        ['Cash Remaining', "='Cash Flow'!G18", True],
        ['Months Runway', "='Cash Flow'!G19", False],
    ]

    row = 5
    for metric, formula, is_currency in highlights:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        ws[f'B{row}'].font = Font(bold=True, size=12)
        if is_currency:
            apply_currency_format(ws[f'B{row}'])
        elif 'Months' in metric:
            ws[f'B{row}'].number_format = '0.0'
        else:
            ws[f'B{row}'].number_format = '#,##0'
        row += 1

    # Funding
    ws['A13'] = "FUNDING"
    apply_section_header_style(ws['A13'])
    ws.merge_cells('A13:D13')

    funding = [
        ['Raise Amount', "=Assumptions!B4", True],
        ['Valuation Cap', "=Assumptions!B5", True],
        ['Initial Cash', "=Assumptions!B6", True],
        ['Total Capital', "=Assumptions!B7", True],
        ['Capital Deployed (6mo)', "='Cash Flow'!B24", True],
        ['Capital Efficiency', "='Cash Flow'!B26/'Cash Flow'!B24", False],
    ]

    row = 14
    for metric, formula, is_currency in funding:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        if is_currency:
            apply_currency_format(ws[f'B{row}'])
        else:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1

    # Unit Economics
    ws['A21'] = "UNIT ECONOMICS"
    apply_section_header_style(ws['A21'])
    ws.merge_cells('A21:D21')

    unit_econ = [
        ['D2C ARPU', "=Assumptions!B14", True],
        ['D2C CAC', "=Assumptions!B18", True],
        ['D2C LTV/CAC', "=Assumptions!B20", False],
        ['D2C Payback (months)', "=Assumptions!B21", False],
        ['', '', False],
        ['B2B Avg MRR', "=Assumptions!B25", True],
        ['B2B Gross Margin', "=Assumptions!B26", False],
    ]

    row = 22
    for metric, formula, is_currency in unit_econ:
        ws[f'A{row}'] = metric
        if formula:
            ws[f'B{row}'] = formula
            if is_currency:
                apply_currency_format(ws[f'B{row}'])
            elif 'Ratio' in metric:
                ws[f'B{row}'].number_format = '0.0x'
            elif 'months' in metric:
                ws[f'B{row}'].number_format = '0.0'
            else:
                apply_percentage_format(ws[f'B{row}'])
        row += 1

    # Decision Point
    ws['A31'] = "MONTH 6 DECISION POINT"
    apply_section_header_style(ws['A31'])
    ws.merge_cells('A31:D31')

    ws['A32'] = "Target: $52K+ MRR to justify Series A"
    ws.merge_cells('A32:D32')

    ws['A33'] = "Status:"
    ws['B33'] = "=IF('Revenue Model'!G16>=52000,'ON TRACK - SERIES A READY',IF('Revenue Model'!G16>=35000,'STRONG - BRIDGE ROUND',IF('Revenue Model'!G16>=20000,'MODERATE - NEEDS OPTIMIZATION','PIVOT REQUIRED')))"
    ws['B33'].font = Font(bold=True, size=11)
    ws.merge_cells('B33:D33')

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

if __name__ == "__main__":
    create_financial_model()
