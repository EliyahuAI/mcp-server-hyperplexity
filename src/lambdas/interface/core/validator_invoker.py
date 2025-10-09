"""
Handles invoking the core validator Lambda function.
"""
import io
import json
import logging
import os
import tempfile
import time
from openpyxl import Workbook, load_workbook
import csv
import boto3
import math
from pathlib import Path

from row_key_utils import generate_row_key
from schema_validator_simplified import SimplifiedSchemaValidator
from interface_lambda.reporting.markdown_report import create_markdown_table_from_results
from shared_table_parser import S3TableParser

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3_client = boto3.client('s3')
lambda_client = boto3.client('lambda', config=boto3.session.Config(
    read_timeout=900,  # 15 minutes read timeout for validation lambda calls
    connect_timeout=60  # 1 minute connect timeout
))

def is_validation_response_complete(response_payload, expected_row_count, is_preview=False):
    """
    Check if validation response indicates complete processing.

    Args:
        response_payload: The parsed response from validation lambda
        expected_row_count: Number of rows we expected to be processed
        is_preview: Whether this is a preview validation (less strict checking)

    Returns:
        tuple: (is_complete: bool, reason: str, validation_results: dict)
    """
    try:
        # Check for basic response structure
        if not response_payload or not isinstance(response_payload, dict):
            return False, "No response or invalid response structure", None

        # Check HTTP status code
        status_code = response_payload.get('statusCode', 200)
        if status_code != 200:
            error_msg = "HTTP error"
            if status_code == 500:
                body = response_payload.get('body', {})
                if isinstance(body, dict):
                    error_msg = body.get('error', body.get('errorMessage', body.get('message', 'Unknown validation error')))
                    # If still unknown, log the response structure
                    if error_msg == 'Unknown validation error':
                        logger.error(f"[COMPLETENESS_CHECK] HTTP 500 - response structure: {response_payload}")
                        error_msg = f"Internal server error - response keys: {list(body.keys()) if body else 'empty body'}"
                else:
                    error_msg = str(body)
            return False, f"HTTP {status_code}: {error_msg}", None

        # Extract response body
        body = response_payload.get('body', {})
        if not isinstance(body, dict):
            return False, "Invalid response body structure", None

        # Check success flag
        if not body.get('success', False):
            error_msg = body.get('error', 'Unknown validation error')
            # Try to get more detailed error information
            if error_msg == 'Unknown validation error':
                # Log the actual response structure for debugging
                logger.error(f"[COMPLETENESS_CHECK] Unknown error - response body structure: {body}")
                # Try alternative error message locations
                if 'message' in body:
                    error_msg = body['message']
                elif 'errorMessage' in body:
                    error_msg = body['errorMessage']
                elif 'errorType' in body:
                    error_msg = f"{body['errorType']}: {body.get('errorMessage', 'No details')}"
                else:
                    error_msg = f"Unknown validation error - response keys: {list(body.keys())}"
            return False, f"Validation failed: {error_msg}", None

        # Extract data section
        data = body.get('data', {})
        if not isinstance(data, dict):
            return False, "Missing or invalid data section", None

        # Extract validation results
        validation_results = data.get('rows', {})
        if not validation_results or not isinstance(validation_results, dict):
            return False, "No validation results in response", None

        # Extract metadata
        metadata = body.get('metadata', {})
        if not isinstance(metadata, dict):
            return False, "Missing or invalid metadata section", None

        # Check row count completion (different logic for preview vs full validation)
        if is_preview:
            # For preview, we're more flexible - just check that we got some results
            if len(validation_results) == 0:
                return False, "Preview returned no validation results", None
            # Preview is complete if we got any results
            return True, "Preview validation complete", validation_results
        else:
            # For full validation, check completed vs total rows
            completed_rows = metadata.get('completed_rows', 0)
            total_rows = metadata.get('total_rows', expected_row_count)

            logger.info(f"[COMPLETENESS_CHECK] Completed: {completed_rows}, Expected: {expected_row_count}, Total in metadata: {total_rows}")

            # If we have explicit row counts, use them
            if completed_rows < expected_row_count:
                return False, f"Partial processing: {completed_rows}/{expected_row_count} rows completed", validation_results

            # Check that we have validation results for the expected number of rows
            actual_results_count = len(validation_results)
            if actual_results_count < expected_row_count:
                return False, f"Incomplete results: {actual_results_count}/{expected_row_count} rows returned", validation_results

        # Check for essential metadata fields
        required_metadata = ['processing_time', 'cache_hits', 'cache_misses']
        missing_metadata = [field for field in required_metadata if field not in metadata]
        if missing_metadata:
            logger.warning(f"[COMPLETENESS_CHECK] Missing metadata fields: {missing_metadata} (proceeding anyway)")

        # Check for token usage (indicates actual API calls were made)
        token_usage = metadata.get('token_usage', {})
        if not token_usage and not is_preview:
            logger.warning(f"[COMPLETENESS_CHECK] No token usage found (possible cached-only processing)")

        # Enhanced Excel validation - check for expected sheets and structure
        if not is_preview:
            # For full validation, verify the enhanced Excel should contain expected sheets
            enhanced_metrics = metadata.get('enhanced_metrics', {})
            if enhanced_metrics:
                # Check if enhanced_metrics contains expected structure
                validation_calls = enhanced_metrics.get('validation_calls_by_provider', {})
                if not validation_calls:
                    logger.warning(f"[COMPLETENESS_CHECK] No validation calls found in enhanced metrics")

                # Check for QC data if QC was expected
                qc_results = data.get('qc_results', {})
                qc_metrics = data.get('qc_metrics', {})
                if qc_results or qc_metrics:
                    logger.info(f"[COMPLETENESS_CHECK] QC data found: {len(qc_results)} results, {qc_metrics.get('total_fields_reviewed', 0)} fields reviewed")

        # All checks passed
        return True, "Validation complete", validation_results

    except Exception as e:
        logger.error(f"[COMPLETENESS_CHECK] Error checking validation completeness: {e}")
        return False, f"Completeness check error: {str(e)}", None

def _invoke_validator_with_retry(lambda_client, function_name, payload, logger, max_retries=2):
    """
    Invoke validator lambda with retry logic for timeouts.
    Since the validation lambda caches results, retries are safe and efficient.
    """
    import botocore.exceptions
    
    # Get session info for tracking
    session_id = payload.get('session_id', 'unknown')
    payload_size_kb = len(json.dumps(payload).encode('utf-8')) / 1024
    
    logger.info(f"[RETRY_TRACKER] Starting validation lambda call for session {session_id}")
    logger.info(f"[RETRY_TRACKER] Function: {function_name}, Payload size: {payload_size_kb:.1f}KB, Max retries: {max_retries}")
    
    for attempt in range(max_retries + 1):
        attempt_start_time = time.time()
        try:
            logger.info(f"[RETRY_TRACKER] ATTEMPT {attempt + 1}/{max_retries + 1} - Starting lambda invoke at {time.strftime('%H:%M:%S')}")
            
            response = lambda_client.invoke(
                FunctionName=function_name,
                InvocationType='RequestResponse',
                Payload=json.dumps(payload)
            )
            
            elapsed_time = time.time() - attempt_start_time
            logger.info(f"[RETRY_TRACKER] SUCCESS - Attempt {attempt + 1} completed in {elapsed_time:.2f}s")
            
            # Log success without interfering with the response stream
            logger.info(f"[RETRY_TRACKER] Lambda invoke completed successfully")
            
            # Check for response payload size issues (HTTP 413 - Request Entity Too Large)
            try:
                # First, check if we can read the response at all
                response_data = response['Payload'].read()
                response_size = len(response_data)
                logger.info(f"[RESPONSE_SIZE] Lambda response payload size: {response_size/1024/1024:.2f} MB")

                # Check for AWS Lambda runtime errors in the response
                if response_size == 0:
                    logger.error(f"[RESPONSE_ERROR] Empty response from lambda - possible runtime failure")
                    raise Exception("Empty response from validation lambda - possible runtime error")

                # Try to decode as JSON to check for malformed responses
                try:
                    response_str = response_data.decode('utf-8')
                    if "Failed to post invocation response" in response_str:
                        logger.error(f"[RESPONSE_ERROR] Lambda runtime failed to post response - likely 413 (Request Too Large)")
                        raise Exception("Lambda response too large - validation data exceeds 6MB limit")
                    # Additional check for empty response that might indicate 413 error
                    if response_size > 0 and response_size < 100:  # Very small responses might be error messages
                        if "413" in response_str or "Request Entity Too Large" in response_str:
                            logger.error(f"[RESPONSE_ERROR] Detected 413 error in small response: {response_str[:200]}")
                            raise Exception("Lambda response too large - validation data exceeds 6MB limit")
                except UnicodeDecodeError:
                    logger.error(f"[RESPONSE_ERROR] Response contains invalid UTF-8 - possible corruption")
                    raise Exception("Corrupted response from validation lambda")

                # Reset the stream for normal processing
                response['Payload'] = io.BytesIO(response_data)

            except Exception as read_error:
                logger.error(f"[RESPONSE_ERROR] Failed to validate response: {str(read_error)}")
                raise Exception(f"Invalid response from validation lambda: {str(read_error)}")

            return response

        except botocore.exceptions.ReadTimeoutError as e:
            elapsed_time = time.time() - attempt_start_time
            logger.error(f"[RETRY_TRACKER] READ_TIMEOUT - Attempt {attempt + 1} failed after {elapsed_time:.2f}s")
            logger.error(f"[RETRY_TRACKER] Read timeout details: {str(e)}")
            
            if attempt < max_retries:
                wait_time = 5 * (attempt + 1)  # Progressive backoff: 5s, 10s
                logger.warning(f"[RETRY_TRACKER] RETRYING - Will retry in {wait_time}s (attempt {attempt + 2}/{max_retries + 1})")
                logger.info(f"[RETRY_TRACKER] Note: Validation lambda uses caching, so retry should be faster")
                time.sleep(wait_time)
            else:
                logger.error(f"[RETRY_TRACKER] FINAL_FAILURE - All {max_retries + 1} attempts failed due to read timeouts")
                logger.error(f"[RETRY_TRACKER] Total elapsed time across all attempts: {time.time() - (attempt_start_time - elapsed_time):.2f}s")
                raise
                
        except botocore.exceptions.ConnectTimeoutError as e:
            elapsed_time = time.time() - attempt_start_time
            logger.error(f"[RETRY_TRACKER] CONNECT_TIMEOUT - Attempt {attempt + 1} failed after {elapsed_time:.2f}s")
            logger.error(f"[RETRY_TRACKER] Connection timeout details: {str(e)}")
            
            if attempt < max_retries:
                wait_time = 3 * (attempt + 1)  # Shorter backoff for connection issues
                logger.warning(f"[RETRY_TRACKER] RETRYING - Will retry in {wait_time}s (attempt {attempt + 2}/{max_retries + 1})")
                time.sleep(wait_time)
            else:
                logger.error(f"[RETRY_TRACKER] FINAL_FAILURE - All {max_retries + 1} attempts failed due to connection timeouts")
                raise
                
        except Exception as e:
            elapsed_time = time.time() - attempt_start_time
            logger.error(f"[RETRY_TRACKER] NON_TIMEOUT_ERROR - Attempt {attempt + 1} failed after {elapsed_time:.2f}s")
            logger.error(f"[RETRY_TRACKER] Error type: {type(e).__name__}")
            logger.error(f"[RETRY_TRACKER] Error details: {str(e)}")
            logger.error(f"[RETRY_TRACKER] NOT_RETRYING - Non-timeout errors are not retried")
            raise

def invoke_validator_lambda(excel_s3_key, config_s3_key, max_rows, batch_size, S3_CACHE_BUCKET, VALIDATOR_LAMBDA_NAME, preview_first_row=False, preview_max_rows=5, sequential_call=None, session_id=None, update_callback=None, special_request=None):
    """Invoke the core validator Lambda with Excel data."""
    
    function_start_time = time.time()
    logger.info(">>> ENTER invoke_validator_lambda <<<")
    logger.info(f">>> Parameters: excel_s3_key={excel_s3_key}, preview={preview_first_row}, sequential_call={sequential_call}, special_request={special_request} <<<")
    logger.info(f"[RETRY_TRACKER] FUNCTION_START - invoke_validator_lambda called at {time.strftime('%H:%M:%S')} for session {session_id}")
    logger.info(f"[RETRY_TRACKER] Lambda timeouts configured: read_timeout=900s, connect_timeout=60s")
    
    # Handle special requests (like config generation)
    if special_request:
        logger.info(f"Processing special request: {special_request.get('action', 'unknown')}")
        return _handle_special_request(special_request, excel_s3_key, S3_CACHE_BUCKET, VALIDATOR_LAMBDA_NAME)
    
    try:
        logger.info(f"Starting invoke_validator_lambda - preview_first_row: {preview_first_row}")
        logger.info(f"Excel S3 key: {excel_s3_key}")
        logger.info(f"Config S3 key: {config_s3_key}")

        # Get accurate row count using shared_table_parser (robust empty row handling)
        logger.info(f"Using shared_table_parser to get accurate row count")
        table_parser = S3TableParser()
        parsed_data = table_parser.parse_s3_table(S3_CACHE_BUCKET, excel_s3_key)
        accurate_total_rows = parsed_data['total_rows']
        logger.info(f"Accurate row count from shared_table_parser: {accurate_total_rows}")

        # Download Excel file from S3
        excel_response = s3_client.get_object(Bucket=S3_CACHE_BUCKET, Key=excel_s3_key)
        excel_content = excel_response['Body'].read()
        
        # Download config file from S3
        config_response = s3_client.get_object(Bucket=S3_CACHE_BUCKET, Key=config_s3_key)
        config_data = json.loads(config_response['Body'].read().decode('utf-8'))
        
        # Preprocess config to support both 'column' and 'name' fields
        if 'validation_targets' in config_data and isinstance(config_data['validation_targets'], list):
            for target in config_data['validation_targets']:
                # Ensure both 'column' and 'name' fields exist for compatibility
                if 'column' in target and 'name' not in target:
                    target['name'] = target['column']
                elif 'name' in target and 'column' not in target:
                    target['column'] = target['name']
                # Keep both fields for backward compatibility
        
        # Detect file type and load data accordingly
        # Check if this is a CSV file and convert to Excel format if needed
        original_file_type = "Excel"
        
        # First check if it's a valid Excel file by checking ZIP signature
        if excel_content.startswith(b'PK'):
            # This is likely a ZIP-based file (Excel .xlsx)
            logger.info("Detected Excel file format (ZIP signature)")
            original_file_type = "Excel"
        else:
            # Not a ZIP file, try to process as CSV
            try:
                # Try multiple encodings to handle different CSV file encodings
                encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-16']
                text_content = None
                used_encoding = None
                
                for encoding in encodings_to_try:
                    try:
                        text_content = excel_content.decode(encoding)
                        used_encoding = encoding
                        break
                    except UnicodeDecodeError:
                        continue
                
                if text_content is None:
                    raise ValueError(f"Could not decode file with any of the tried encodings: {encodings_to_try}")
                    
                logger.info(f"Successfully decoded file using {used_encoding} encoding")
                # Simple heuristic: if it contains commas, treat as CSV
                if ',' in text_content:
                    original_file_type = "CSV"
                    logger.info("Detected CSV file format - converting to Excel")
                    
                    # Parse CSV content
                    csv_reader = csv.reader(io.StringIO(text_content))
                    csv_rows = list(csv_reader)
                    
                    if not csv_rows:
                        raise ValueError("CSV file is empty")
                    
                    # Convert CSV to Excel format in memory
                    workbook = Workbook()
                    worksheet = workbook.active
                    worksheet.title = "Data"
                    
                    # Write CSV data to Excel worksheet
                    for row_idx, csv_row in enumerate(csv_rows, 1):
                        for col_idx, cell_value in enumerate(csv_row, 1):
                            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    # Save Excel workbook to bytes
                    excel_buffer = io.BytesIO()
                    workbook.save(excel_buffer)
                    excel_content = excel_buffer.getvalue()
                    excel_buffer.close()
                    
                    logger.info(f"Converted CSV with {len(csv_rows)} rows to Excel format")
                else:
                    # Can decode as UTF-8 but no commas - might be old Excel format or other file
                    logger.error(f"File doesn't appear to be CSV or Excel format. Content starts with: {text_content[:100]}...")
                    raise ValueError("File format not supported - must be Excel (.xlsx) or CSV")
            except UnicodeDecodeError:
                # Can't decode as UTF-8 and not a ZIP file - unsupported format
                logger.error("File is binary but not a valid Excel file (no ZIP signature)")
                logger.error(f"File starts with bytes: {excel_content[:20]}")
                raise ValueError("File format not supported - file appears to be binary but not a valid Excel file")
        
        # Now process as Excel file (whether original or converted from CSV)
        logger.info(f"Processing {original_file_type} file as Excel format")
        workbook = load_workbook(io.BytesIO(excel_content))
        
        # Log available sheet names
        logger.info(f"Available sheets in Excel: {workbook.sheetnames}")
        
        # Select the appropriate sheet
        if 'Results' in workbook.sheetnames:
            worksheet = workbook['Results']
            logger.info(f"Using 'Results' sheet as data source")
        elif len(workbook.sheetnames) > 0:
            worksheet = workbook[workbook.sheetnames[0]]
            logger.info(f"Using first sheet '{worksheet.title}' as data source")
        else:
            worksheet = workbook.active
            logger.info(f"Using active sheet: {worksheet.title}")
        
        # Read headers from first row
        headers = []
        for cell in worksheet[1]:
            header = cell.value
            if header:
                headers.append(str(header).strip())
            else:
                headers.append('')
        
        logger.info(f"Headers found: {headers}")
        logger.info(f"Total columns: {len(headers)}")

        # Process data rows
        rows = []
        # Use accurate row count from shared_table_parser (handles empty rows correctly)
        total_rows = accurate_total_rows
        logger.info(f"Using accurate row count: {total_rows} (worksheet.max_row was {worksheet.max_row})")

        # Limit rows for preview mode or max_rows
        if preview_first_row:
            # For preview mode, we want to load all rows up to preview_max_rows
            # so the validator can process them sequentially and use caching
            max_process_rows = min(preview_max_rows, total_rows)
        else:
            max_process_rows = min(max_rows, total_rows) if max_rows else total_rows
        
        logger.info(f"Processing {max_process_rows} rows from {original_file_type} (total available: {total_rows})")
        
        # Get ID fields from config for row key generation
        id_fields = []
        logger.info("Searching for ID fields in validation targets...")
        for i, target in enumerate(config_data.get('validation_targets', [])):
            importance = target.get('importance', '')
            logger.info(f"Target {i}: importance='{importance}', column='{target.get('column')}', name='{target.get('name')}'")
            if importance.upper() == 'ID':
                # Support both 'name' and 'column' fields
                field_name = target.get('name') or target.get('column')
                if field_name:
                    id_fields.append(field_name)
                    logger.info(f"Found ID field: {field_name}")
                else:
                    logger.warning(f"ID target found but no field name: {target}")
            else:
                logger.debug(f"Target '{target.get('column')}' has importance '{importance}' (not ID)")
        
        # If no ID fields found, try to use SimplifiedSchemaValidator to determine primary keys
        if not id_fields:
            try:
                validator = SimplifiedSchemaValidator(config_data)
                id_fields = validator.primary_key
                logger.info(f"Using primary keys from SimplifiedSchemaValidator: {id_fields}")
            except ImportError:
                logger.warning("SimplifiedSchemaValidator not available in deployment package")
            except Exception as e:
                logger.warning(f"Could not use SimplifiedSchemaValidator: {e}")
        
        logger.info(f"ID fields for row key generation: {id_fields}")
        
        # Process data rows using Excel format (works for both original Excel and converted CSV)
        # Extract all rows first, then apply hybrid hashing
        excel_rows = []
        for row_idx in range(2, min(2 + max_process_rows, worksheet.max_row + 1)):
            row_data = {}

            # Extract cell values
            for col_idx, header in enumerate(headers):
                if header:  # Skip empty headers
                    cell_value = worksheet.cell(row=row_idx, column=col_idx + 1).value
                    row_data[header] = str(cell_value) if cell_value is not None else ""

            # Debug: Log first row's data to see what we're getting
            if row_idx == 2:  # First data row
                logger.info(f"First row data extracted: {json.dumps(row_data)}")
                logger.info(f"Looking for ID fields: {id_fields}")
                for id_field in id_fields:
                    logger.info(f"  {id_field}: '{row_data.get(id_field, 'NOT FOUND')}'")

            excel_rows.append(row_data)

        # HYBRID ROW KEY GENERATION (matches async validation flow)
        # 1. Try ID-field hashing first (for history matching)
        # 2. For duplicates, use full-row hashing (to distinguish them)
        id_hash_counts = {}  # Track duplicate ID hashes

        # First pass: Generate ID-field hashes and detect duplicates
        for row_idx, row_data in enumerate(excel_rows):
            # Generate ID-field hash
            id_hash = generate_row_key(row_data, primary_keys=id_fields if id_fields else None)

            if id_hash not in id_hash_counts:
                id_hash_counts[id_hash] = []
            id_hash_counts[id_hash].append(row_idx)

        # Second pass: Assign final row keys (ID-hash or full-row hash for duplicates)
        for row_idx, row_data in enumerate(excel_rows):
            id_hash = generate_row_key(row_data, primary_keys=id_fields if id_fields else None)

            # If this ID hash appears multiple times, use full-row hash
            if len(id_hash_counts[id_hash]) > 1:
                row_key = generate_row_key(row_data, primary_keys=None)  # Full-row hash
                logger.debug(f"Row {row_idx}: Duplicate ID detected, using full-row hash: {row_key[:8]}...")
            else:
                row_key = id_hash  # Use ID-field hash
                logger.debug(f"Row {row_idx}: Using ID-field hash: {row_key[:8]}...")

            row_data['_row_key'] = row_key
            rows.append(row_data)

        # Log duplicate summary
        duplicate_id_count = sum(1 for count_list in id_hash_counts.values() if len(count_list) > 1)
        total_duplicate_rows = sum(len(count_list) for count_list in id_hash_counts.values() if len(count_list) > 1)
        if duplicate_id_count > 0:
            logger.info(f"[PREVIEW_HYBRID_HASH] Found {duplicate_id_count} duplicate ID groups ({total_duplicate_rows} total rows)")
            logger.info(f"[PREVIEW_HYBRID_HASH] Duplicate rows will use full-row hashing")
        else:
            logger.info(f"[PREVIEW_HYBRID_HASH] No duplicate IDs detected, all rows using ID-field hashing")
        
        # Show sample row data to understand what keys will be generated
        if rows and len(rows) > 0:
            sample_row = rows[0]
            logger.info("Sample row data for first row:")
            for id_field in id_fields:
                value = sample_row.get(id_field, 'NOT FOUND')
                logger.info(f"  {id_field}: {value}")
        
        # Load validation history using new extract_validation_history method
        validation_history = {}
        if not preview_first_row:  # Load validation history for all files (now all are Excel format)
            try:
                # Use S3TableParser to extract validation history from Updated Values sheet
                # Pass parsed rows data so row keys match (includes deduplication)
                parsed_rows_data = {'data': rows}  # rows already have _row_key
                history_data = table_parser.extract_validation_history(
                    S3_CACHE_BUCKET,
                    excel_s3_key,
                    parsed_data=parsed_rows_data
                )

                # Extract the validation_history dict from the returned structure
                original_validation_history = history_data.get('validation_history', {})

                logger.info(f"Loaded validation history for {len(original_validation_history)} row keys from Excel")

                if original_validation_history:
                    validation_history = original_validation_history

                    # Log matching summary
                    matched_count = 0
                    for row in rows:
                        payload_key = row.get('_row_key', '')
                        if payload_key and payload_key in validation_history:
                            matched_count += 1

                    logger.info(f"Matched {matched_count} out of {len(rows)} rows with validation history")

                    if matched_count == 0 and len(validation_history) > 0:
                        logger.warning("No rows matched with validation history - may be using different primary keys")
                        logger.info(f"Current primary keys: {id_fields}")

            except Exception as e:
                logger.warning(f"Could not load validation history: {e}")
                import traceback
                traceback.print_exc()
                validation_history = {}
        
        # Initialize aggregated metadata
        aggregated_metadata = {
            'total_rows': total_rows,
            'token_usage': {
                'total_tokens': 0, 'total_cost': 0.0, 'api_calls': 0, 'cached_calls': 0,
                'by_provider': {
                    'perplexity': {'prompt_tokens': 0, 'completion_tokens': 0, 'total_tokens': 0, 'calls': 0, 'input_cost': 0.0, 'output_cost': 0.0, 'total_cost': 0.0},
                    'anthropic': {'input_tokens': 0, 'output_tokens': 0, 'cache_creation_tokens': 0, 'cache_read_tokens': 0, 'total_tokens': 0, 'calls': 0, 'input_cost': 0.0, 'output_cost': 0.0, 'total_cost': 0.0}
                },
                'by_model': {}
            }
        }
        
        if preview_first_row:
            total_rows_to_send = min(sequential_call, preview_max_rows, len(rows)) if sequential_call else min(preview_max_rows, len(rows))
            logger.info(f"Preview mode: sending {total_rows_to_send} rows to validator")
            
            preview_rows = rows[:total_rows_to_send]
            
            payload = {
                "test_mode": True, 
                "config": config_data,
                "validation_data": {
                    "rows": preview_rows,
                    "batch_info": {
                        "current_batch": 1,
                        "total_batches": 1,
                        "batch_start_idx": 0,
                        "batch_end_idx": len(preview_rows),
                        "total_dataset_size": total_rows  # Use actual total rows from Excel, not just preview rows
                    }
                }, 
                "validation_history": {},
                "session_id": session_id,
                "is_preview": True,
                "total_rows": total_rows  # Also add as top-level field for clarity
            }
            
            try:
                validation_start_time = time.time()
                logger.info(f"[RETRY_TRACKER] PREVIEW_MODE - Starting validation call for {len(preview_rows)} rows")
                response = _invoke_validator_with_retry(lambda_client, VALIDATOR_LAMBDA_NAME, payload, logger)
                validation_processing_time = time.time() - validation_start_time
                logger.info(f"[RETRY_TRACKER] PREVIEW_COMPLETE - Validator Lambda processing took {validation_processing_time:.2f} seconds")

                response_payload = json.loads(response['Payload'].read().decode('utf-8'))

                # ========== VALIDATION COMPLETENESS DETECTION - PREVIEW ==========
                is_complete, completeness_reason, extracted_results = is_validation_response_complete(
                    response_payload, preview_max_rows, is_preview=True
                )

                if not is_complete:
                    logger.error(f"[PREVIEW_INCOMPLETE] Preview validation incomplete: {completeness_reason}")
                    # For preview, we're more tolerant - return what we got with a warning
                    logger.warning(f"[PREVIEW_INCOMPLETE] Continuing with incomplete preview results")

                logger.info(f"[PREVIEW_COMPLETE] Preview validation completeness: {completeness_reason}")

                # Check for validation lambda errors (now handled by completeness check too)
                if isinstance(response_payload, dict) and response_payload.get('statusCode') == 500:
                    error_body = response_payload.get('body', {})
                    error_message = error_body.get('error', 'Unknown validation error') if isinstance(error_body, dict) else str(error_body)
                    logger.error(f"[VALIDATION_ERROR] Validation lambda returned error: {error_message}")
                    raise Exception(f"Validation failed: {error_message}")

                validation_results, metadata, qc_results, qc_metrics = None, None, None, None
                if isinstance(response_payload, dict):
                    body = response_payload.get('body', {})
                    if isinstance(body, dict):
                        data = body.get('data', {})
                        validation_results = data.get('rows')
                        # Extract QC data from the data section
                        qc_results = data.get('qc_results', {})
                        qc_metrics = data.get('qc_metrics', {})
                        metadata = body.get('metadata', {})
                        logger.info(f"[QC_EXTRACT_DEBUG] Extracted QC data - results: {len(qc_results)}, metrics fields: {qc_metrics.get('total_fields_reviewed', 0) if qc_metrics else 0}")
                    if 'validation_results' in response_payload:
                        validation_results = response_payload['validation_results']

                if metadata:
                    aggregated_metadata.update(metadata)

                new_row_processed = None
                if validation_results:
                    for row_key in sorted(validation_results.keys(), key=lambda x: int(x) if x.isdigit() else -1, reverse=True):
                        row_data = validation_results[row_key]
                        if isinstance(row_data, dict) and '_raw_responses' in row_data:
                            if any(not r.get('is_cached', True) for r in row_data['_raw_responses'].values()):
                                new_row_processed = int(row_key) + 1
                                break
                
                preview_complete = (new_row_processed is None or new_row_processed >= preview_max_rows or new_row_processed >= len(rows))
                
                result_response = {
                    'total_rows': total_rows, 'validation_results': validation_results, 'metadata': aggregated_metadata,
                    'total_processed_rows': len(validation_results) if validation_results else 0,
                    'new_row_number': new_row_processed, 'preview_complete': preview_complete,
                    'qc_results': qc_results,  # Pass QC results
                    'qc_metrics': qc_metrics   # Pass QC metrics
                }
                if isinstance(response_payload, dict):
                    result_response.update(response_payload)
                return result_response
                
            except Exception as e:
                logger.error(f"[RETRY_TRACKER] PREVIEW_FAILED - Final error after all retries: {str(e)}")
                logger.warning(f"Validator Lambda timeout or error in preview mode: {str(e)}")
                return {'status': 'timeout', 'total_rows': total_rows, 'validation_results': {}, 'metadata': aggregated_metadata, 'note': f'Validation timed out after retries. Error: {str(e)}'}
        else:
            # For normal mode, send full dataset to validation lambda and let enhanced batch manager handle batching
            logger.info(f"Sending {len(rows)} rows to validation lambda for enhanced batch processing")
            
            if update_callback:
                update_callback(session_id, "PROCESSING", 0, f"Starting enhanced batch processing of {len(rows)} rows.", 0, None, 0, 1)

            # Send full dataset to validation lambda - let enhanced batch manager determine optimal batching
            payload = {
                "test_mode": False, 
                "config": config_data, 
                "validation_data": {
                    "rows": rows,  # Send ALL rows at once
                    "max_rows": max_rows,
                    "batch_size": batch_size,  # Pass through the batch_size (may be None)
                    "total_dataset_size": len(rows)
                }, 
                "validation_history": validation_history, 
                "session_id": session_id
            }
            
            try:
                logger.info(f"[RETRY_TRACKER] FULL_MODE - Starting validation call for {len(rows)} rows")
                response = _invoke_validator_with_retry(lambda_client, VALIDATOR_LAMBDA_NAME, payload, logger)
                response_payload = json.loads(response['Payload'].read().decode('utf-8'))

                # ========== VALIDATION COMPLETENESS DETECTION - FULL VALIDATION ==========
                is_complete, completeness_reason, extracted_results = is_validation_response_complete(
                    response_payload, len(rows), is_preview=False
                )

                logger.info(f"[FULL_VALIDATION_COMPLETE] Validation completeness: {completeness_reason}")

                if not is_complete:
                    logger.error(f"[FULL_VALIDATION_INCOMPLETE] Full validation incomplete: {completeness_reason}")
                    # For full validation, this is a critical failure - we should delegate to async if possible
                    # Return incomplete status so Smart Delegation System can decide what to do
                    return {
                        'status': 'incomplete',
                        'total_rows': len(rows),
                        'validation_results': extracted_results if extracted_results else {},
                        'metadata': aggregated_metadata,
                        'completeness_reason': completeness_reason,
                        'incomplete_validation': True
                    }

                # Check for validation lambda errors (now handled by completeness check too)
                if isinstance(response_payload, dict) and response_payload.get('statusCode') == 500:
                    error_body = response_payload.get('body', {})
                    error_message = error_body.get('error', 'Unknown validation error') if isinstance(error_body, dict) else str(error_body)
                    logger.error(f"[VALIDATION_ERROR] Validation lambda returned error: {error_message}")
                    raise Exception(f"Validation failed: {error_message}")

                validation_results, metadata, qc_results, qc_metrics = None, None, None, None
                if isinstance(response_payload, dict):
                    body = response_payload.get('body', {})
                    if isinstance(body, dict):
                       data = body.get('data', {})
                       validation_results = data.get('rows')
                       # Extract QC data from the data section
                       qc_results = data.get('qc_results', {})
                       qc_metrics = data.get('qc_metrics', {})
                       metadata = body.get('metadata', {})
                       logger.info(f"[QC_EXTRACT_DEBUG] Full mode - Extracted QC data - results: {len(qc_results) if qc_results else 0}, metrics fields: {qc_metrics.get('total_fields_reviewed', 0) if qc_metrics else 0}")
                    if 'validation_results' in response_payload:
                       validation_results = response_payload['validation_results']
                       metadata = response_payload.get('metadata', {})

                # Merge the metadata from the single response
                if metadata:
                    if 'token_usage' in metadata:
                        aggregated_metadata['token_usage'] = metadata['token_usage']
                    
                    # Copy other metadata fields
                    for key, value in metadata.items():
                        if key != 'token_usage':
                            aggregated_metadata[key] = value
                
                # Use the validation results directly (no need to reindex since we sent all rows at once)
                all_validation_results = validation_results if validation_results else {}
                    
            except Exception as e:
                logger.error(f"[RETRY_TRACKER] FULL_MODE_FAILED - Final error after all retries: {str(e)}")
                logger.error(f"Error processing full dataset: {str(e)}")
                return {'status': 'error', 'total_rows': total_rows, 'validation_results': {}, 'metadata': aggregated_metadata, 'error': str(e)}
            
            # Update callback with completion
            if update_callback:
                processed_rows_count = len(all_validation_results)
                percent_complete = 100
                verbose_status = "Validation enhanced batch processing completed"
                update_callback(session_id, "COMPLETED", processed_rows_count, verbose_status, percent_complete, None, 1, 1)

            logger.info(f"Enhanced batch processing completed. Total results: {len(all_validation_results)}")
            total_function_time = time.time() - function_start_time
            logger.info(f"[RETRY_TRACKER] FUNCTION_COMPLETE - Total function execution time: {total_function_time:.2f}s")
            return {
                'total_rows': total_rows,
                'validation_results': all_validation_results,
                'metadata': aggregated_metadata,
                'status': 'completed',
                'qc_results': qc_results,  # Pass QC results from full validation
                'qc_metrics': qc_metrics   # Pass QC metrics from full validation
            }
            
    except Exception as e:
        total_function_time = time.time() - function_start_time
        logger.error(f"[RETRY_TRACKER] FUNCTION_ERROR - Function failed after {total_function_time:.2f}s")
        logger.error(f"Error invoking validator Lambda: {str(e)}")
        import traceback
        traceback.print_exc()
        raise 


def _handle_special_request(special_request, excel_s3_key, S3_CACHE_BUCKET, VALIDATOR_LAMBDA_NAME):
    """Handle special requests like config generation that use the validation lambda's Claude integration"""
    
    action = special_request.get('action')
    
    if action == 'generate_config':
        logger.info("Handling config generation request")
        
        try:
            # Build config generation payload for validation lambda
            payload = {
                "config_generation_request": True,
                "table_analysis": special_request.get('table_analysis'),
                "generation_mode": special_request.get('generation_mode', 'automatic'),
                "conversation_id": special_request.get('conversation_id'),
                "user_message": special_request.get('user_message', ''),
                "session_id": special_request.get('session_id'),
                "excel_s3_key": excel_s3_key,
                "s3_cache_bucket": S3_CACHE_BUCKET
            }
            
            logger.info(f"Calling validation lambda for config generation with session_id: {special_request.get('session_id')}")
            logger.info(f"[RETRY_TRACKER] CONFIG_GEN - Starting config generation call")
            
            # Call validation lambda
            response = _invoke_validator_with_retry(lambda_client, VALIDATOR_LAMBDA_NAME, payload, logger)

            response_payload = json.loads(response['Payload'].read().decode('utf-8'))

            # Check for validation lambda errors
            if isinstance(response_payload, dict) and response_payload.get('statusCode') == 500:
                error_body = response_payload.get('body', {})
                error_message = error_body.get('error', 'Unknown validation error') if isinstance(error_body, dict) else str(error_body)
                logger.error(f"[VALIDATION_ERROR] Config generation lambda returned error: {error_message}")
                raise Exception(f"Config generation failed: {error_message}")

            logger.info(f"Config generation response received from validation lambda")

            # Extract the result from validation lambda response
            if isinstance(response_payload, dict):
                body = response_payload.get('body', {})
                if isinstance(body, dict):
                    return body
                else:
                    # Direct response format
                    return response_payload
            
            return response_payload
            
        except Exception as e:
            logger.error(f"Error in config generation request: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': f'Config generation failed: {str(e)}'
            }
    
    else:
        logger.error(f"Unknown special request action: {action}")
        return {
            'success': False,
            'error': f'Unknown special request action: {action}'
        }