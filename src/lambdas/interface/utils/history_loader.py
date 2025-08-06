"""
Fallback validation history loader using openpyxl.

This module provides a pandas-free implementation for loading validation history
from the 'Details' sheet of an Excel file.
"""

import logging
from datetime import datetime
import openpyxl

logger = logging.getLogger()
logger.setLevel(logging.INFO)


def load_validation_history_from_excel(excel_path):
    """Load validation history from Excel using openpyxl (pandas-free implementation)."""
    logger.info(f"Using fallback validation history loader for: {excel_path}")
    try:
        # Import row key utilities if available
        # No sanitization needed for hash-based row keys
        
        # Load workbook with openpyxl
        logger.info(f"Loading workbook with openpyxl from: {excel_path}")
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
        
        # Log available sheet names
        sheet_names = workbook.sheetnames
        logger.info(f"Available sheets: {sheet_names}")
        
        # Check for Details sheet
        if 'Details' not in sheet_names:
            logger.info("No Details worksheet found in Excel file")
            workbook.close()
            return {}
        
        # Load Details worksheet
        details_sheet = workbook['Details']
        logger.info("Found Details worksheet")
        
        # Get headers from first row
        headers = []
        for cell in details_sheet[1]:
            if cell.value:
                headers.append(str(cell.value))
            else:
                headers.append('')
        
        logger.info(f"Details sheet headers: {headers}")
        
        # Create mapping for ID columns (handle ID: prefix for backwards compatibility)
        id_column_mapping = {}
        for header in headers:
            if header.startswith('ID:'):
                # Map ID:ColumnName to ColumnName for backwards compatibility
                clean_name = header[3:]  # Remove 'ID:' prefix
                id_column_mapping[clean_name] = header
        
        logger.info(f"ID column mapping (for backwards compatibility): {id_column_mapping}")
        
        # Find column indices
        col_indices = {
            'row_key': None,
            'column': None,
            'value': None,
            'confidence': None,
            'quote': None,
            'sources': None,
            'timestamp': None
        }
        
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if 'row key' in header_lower:
                col_indices['row_key'] = idx
            elif header_lower == 'column':
                col_indices['column'] = idx
            elif 'validated value' in header_lower:
                col_indices['value'] = idx
            elif header_lower == 'confidence':
                col_indices['confidence'] = idx
            elif header_lower == 'quote':
                col_indices['quote'] = idx
            elif header_lower == 'sources':
                col_indices['sources'] = idx
            elif header_lower == 'timestamp':
                col_indices['timestamp'] = idx
        
        logger.info(f"Column indices found: {col_indices}")
        
        # Check if we have minimum required columns
        if col_indices['row_key'] is None or col_indices['column'] is None:
            logger.warning("Required columns (Row Key, Column) not found in Details sheet")
            workbook.close()
            return {}
        
        # Convert to validation_history structure
        validation_history = {}
        row_count = 0
        
        # Iterate through rows (skip header)
        for row_idx, row in enumerate(details_sheet.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Extract values with safe indexing
                row_key = str(row[col_indices['row_key']] or '') if col_indices['row_key'] < len(row) else ''
                column = str(row[col_indices['column']] or '') if col_indices['column'] < len(row) else ''
                
                # Skip empty rows
                if not row_key or not column:
                    continue
                
                # Get other values safely
                value = str(row[col_indices['value']] or '') if col_indices['value'] and col_indices['value'] < len(row) else ''
                confidence = str(row[col_indices['confidence']] or '') if col_indices['confidence'] and col_indices['confidence'] < len(row) else ''
                quote = str(row[col_indices['quote']] or '') if col_indices['quote'] and col_indices['quote'] < len(row) else ''
                sources_str = str(row[col_indices['sources']] or '') if col_indices['sources'] and col_indices['sources'] < len(row) else ''
                timestamp = str(row[col_indices['timestamp']] or '') if col_indices['timestamp'] and col_indices['timestamp'] < len(row) else ''
                
                # Use the row key as-is (no conversion needed for hash-based keys)
                sanitized_row_key = row_key
                
                # Initialize structures
                if sanitized_row_key not in validation_history:
                    validation_history[sanitized_row_key] = {}
                
                if column not in validation_history[sanitized_row_key]:
                    validation_history[sanitized_row_key][column] = []
                
                # Parse sources
                sources = []
                if sources_str and sources_str != 'N/A' and sources_str.strip():
                    sources = [s.strip() for s in sources_str.split(';') if s.strip()]
                
                # Default timestamp if missing
                if not timestamp or timestamp == 'N/A' or timestamp == 'nan':
                    timestamp = datetime.utcnow().isoformat()
                
                # Create history entry
                history_entry = {
                    'timestamp': timestamp,
                    'value': value,
                    'confidence_level': confidence,
                    'quote': quote if quote != 'N/A' else '',
                    'sources': sources
                }
                
                validation_history[sanitized_row_key][column].append(history_entry)
                row_count += 1
                
                # Log first few entries for debugging
                if row_count <= 3:
                    logger.info(f"Sample entry {row_count}: row_key='{sanitized_row_key}', column='{column}', value='{value}'")
                
            except Exception as row_error:
                logger.warning(f"Error processing row {row_idx}: {row_error}")
                continue
        
        workbook.close()
        logger.info(f"Loaded validation history from Details worksheet for {len(validation_history)} row keys ({row_count} entries)")
        
        # Log sample keys for debugging
        if validation_history:
            sample_keys = list(validation_history.keys())[:3]
            logger.info(f"Sample validation history keys: {sample_keys}")
        
        return validation_history
        
    except Exception as e:
        logger.error(f"Error loading validation history from Excel: {e}")
        import traceback
        logger.error(traceback.format_exc())
    return {} 