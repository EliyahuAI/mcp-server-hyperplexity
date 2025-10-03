"""
Handles the main background processing for validation tasks.
This is where the core orchestration happens for a validation run after
it has been triggered.
"""
import json
import logging
import os
import time
from datetime import datetime, timezone, timedelta
import math
import io
import tempfile
from typing import Dict

import boto3

# Initialize AWS clients
s3_client = boto3.client('s3')
ses_client = boto3.client('ses', region_name='us-east-1')
sqs_client = boto3.client('sqs')
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Import DynamoDB functions at module level
# Define dummy functions first as fallback
def update_processing_metrics(*args, **kwargs): 
    logger.warning("DynamoDB not available - processing metrics not updated")
def track_email_delivery(*args, **kwargs): 
    logger.warning("DynamoDB not available - email delivery not tracked")
def track_user_request(*args, **kwargs): 
    logger.warning("DynamoDB not available - user request not tracked")
def update_run_status(**kwargs): 
    logger.warning("DynamoDB not available - run status not updated")
def create_run_record(*args, **kwargs):
    logger.warning("DynamoDB not available - run record not created")

DYNAMODB_AVAILABLE = False

try:
    import sys
    import os
    sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', 'shared'))
    from dynamodb_schemas import update_processing_metrics, track_email_delivery, track_user_request, update_run_status, create_run_record
    from dynamodb_schemas import save_delegation_context, update_async_progress, load_delegation_context, mark_async_completion
    DYNAMODB_AVAILABLE = True
    logger.debug("DynamoDB functions imported successfully at module level")
except ImportError as e:
    logger.error(f"Failed to import dynamodb_schemas at module level: {e}")
    # Dummy functions are already defined above

# Common final processing will be implemented with cleaner bifurcation approach

# Global variable for the API Gateway Management client
api_gateway_management_client = None

# ========== SMART DELEGATION SYSTEM CONFIGURATION ==========
# Configurable sync timeout limit - jobs exceeding this will be delegated to async processing
MAX_SYNC_INVOCATION_TIME_MINUTES = float(os.environ.get('MAX_SYNC_INVOCATION_TIME', '10.0'))  # Normal: 10 minutes
VALIDATOR_SAFETY_BUFFER_MINUTES = float(os.environ.get('VALIDATOR_SAFETY_BUFFER', '3.0'))

# SQS Queue Names for async processing
ASYNC_VALIDATOR_QUEUE = os.environ.get('ASYNC_VALIDATOR_QUEUE', 'perplexity-validator-async-queue')
INTERFACE_COMPLETION_QUEUE = os.environ.get('INTERFACE_COMPLETION_QUEUE', 'perplexity-validator-completion-queue')

logger.debug(f"[DELEGATION_CONFIG] MAX_SYNC_INVOCATION_TIME: {MAX_SYNC_INVOCATION_TIME_MINUTES} minutes")
logger.debug(f"[DELEGATION_CONFIG] ASYNC_VALIDATOR_QUEUE: {ASYNC_VALIDATOR_QUEUE}")
logger.debug(f"[DELEGATION_CONFIG] INTERFACE_COMPLETION_QUEUE: {INTERFACE_COMPLETION_QUEUE}")

def send_validation_failure_alert(session_id, email, error_type, error_details, session_data=None):
    """
    Send high-priority alert emails for validation failures

    Args:
        session_id: Session identifier
        email: User's email address
        error_type: Type of error (e.g., "Response Too Large", "Timeout", etc.)
        error_details: Detailed error information
        session_data: Optional session metadata for context
    """
    try:
        # Prepare session information
        session_info = session_data or {}
        timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')

        # Email to user
        user_subject = "🟥 URGENT: Validation Processing Issue - We're On It!"
        user_body = f"""
Dear Valued Customer,

We've encountered a technical issue while processing your validation request and wanted to notify you immediately.

SESSION DETAILS:
- Session ID: {session_id}
- Time: {timestamp}
- Issue Type: {error_type}
- Your Email: {email}

WHAT HAPPENED:
{error_details}

WHAT WE'RE DOING:
✅ Our technical team has been automatically alerted
✅ We're investigating the issue immediately
✅ Your data and session information are safely stored
✅ We will resolve this and complete your validation

NEXT STEPS:
- You will receive an update within 2 hours
- No action needed on your part
- Your account will not be charged for failed processing

If you have any questions or concerns, please don't hesitate to reach out to us directly.

Best regards,
The Hyperplexity Team

---
Technical Reference: {session_id}
Support: eliyahu@eliyahu.ai
"""

        # Email to admin (Eliyahu)
        admin_subject = f"🟥 CRITICAL: Validation Lambda Failure - {error_type}"
        admin_body = f"""
CRITICAL VALIDATION SYSTEM FAILURE

SESSION: {session_id}
USER: {email}
TIME: {timestamp}
ERROR TYPE: {error_type}

ERROR DETAILS:
{error_details}

SESSION DATA:
{json.dumps(session_info, indent=2) if session_info else 'No session data available'}

IMMEDIATE ACTION REQUIRED:
1. Investigate validation lambda response size/timeout issues
2. Check for data processing bottlenecks
3. Consider implementing chunked processing
4. Follow up with user within 2 hours

SYSTEM HEALTH CHECK NEEDED:
- Lambda timeout configurations
- Response payload size limits
- Memory allocation
- Batch processing logic

This is an automated alert from the validation system.
"""

        # Send user email
        try:
            ses_client.send_email(
                Source='eliyahu@eliyahu.ai',  # Use verified email address
                Destination={'ToAddresses': [email]},
                Message={
                    'Subject': {'Data': user_subject, 'Charset': 'UTF-8'},
                    'Body': {'Text': {'Data': user_body, 'Charset': 'UTF-8'}}
                }
            )
            logger.debug(f"[ALERT_EMAIL] User alert sent to {email}")
        except Exception as e:
            logger.error(f"[ALERT_EMAIL] Failed to send user alert: {e}")

        # Send admin email (high priority)
        try:
            ses_client.send_email(
                Source='eliyahu@eliyahu.ai',  # Use verified email address
                Destination={'ToAddresses': ['eliyahu@eliyahu.ai']},
                Message={
                    'Subject': {'Data': admin_subject, 'Charset': 'UTF-8'},
                    'Body': {'Text': {'Data': admin_body, 'Charset': 'UTF-8'}}
                }
            )
            logger.debug(f"[ALERT_EMAIL] Admin alert sent to eliyahu@eliyahu.ai")
        except Exception as e:
            logger.error(f"[ALERT_EMAIL] Failed to send admin alert: {e}")

    except Exception as e:
        logger.error(f"[ALERT_EMAIL] Failed to send validation failure alerts: {e}")

def send_preview_failure_alert(session_id, email, error_type, error_details, session_data=None):
    """
    Send lower-priority alert emails for preview failures with CSV guidance

    Args:
        session_id: Session identifier
        email: User's email address
        error_type: Type of error
        error_details: Detailed error information
        session_data: Optional session metadata for context
    """
    try:
        # Prepare session information
        session_info = session_data or {}
        timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')

        # Email to user (informative, not urgent)
        user_subject = "🟥 Preview Processing Issue - Troubleshooting Guide"
        user_body = f"""
Hi there,

We encountered an issue while processing your data preview and wanted to let you know right away.

SESSION DETAILS:
- Session ID: {session_id}
- Time: {timestamp}
- Issue: {error_type}

WHAT HAPPENED:
{error_details}

TROUBLESHOOTING STEPS:
Please visit our troubleshooting guide for step-by-step solutions:
https://eliyahu.ai/troubleshooting

This guide contains the most up-to-date recommendations for resolving common processing issues, including file format suggestions and table structure requirements.

We've queued this issue for our team to review, but following the troubleshooting guide will likely resolve it immediately.

If you continue to have issues after following the troubleshooting guide, please don't hesitate to reach out.

Best regards,
The Hyperplexity Team

---
Technical Reference: {session_id}
Support: eliyahu@eliyahu.ai
"""

        # Email to admin (lower priority - informational)
        admin_subject = f"Preview Processing Issue - {error_type}"
        admin_body = f"""
Preview Processing Issue (Lower Priority)

SESSION: {session_id}
USER: {email}
TIME: {timestamp}
ERROR TYPE: {error_type}

ERROR DETAILS:
{error_details}

SESSION DATA:
{json.dumps(session_info, indent=2) if session_info else 'No session data available'}

SUGGESTED INVESTIGATION:
- Check if table format is non-standard
- Look for multiple tables or complex formatting
- Review if column headers are properly positioned
- Consider if Excel-specific features are causing issues

USER GUIDANCE PROVIDED:
- Directed to troubleshooting guide
- Referenced dynamic recommendations
- No urgency communicated

This is a preview issue, not full validation failure.
"""

        # Send user email
        try:
            ses_client.send_email(
                Source='eliyahu@eliyahu.ai',  # Use verified email address
                Destination={'ToAddresses': [email]},
                Message={
                    'Subject': {'Data': user_subject, 'Charset': 'UTF-8'},
                    'Body': {'Text': {'Data': user_body, 'Charset': 'UTF-8'}}
                }
            )
            logger.debug(f"[PREVIEW_ALERT] User guidance sent to {email}")
        except Exception as e:
            logger.error(f"[PREVIEW_ALERT] Failed to send user guidance: {e}")

        # Send admin email (informational priority)
        try:
            ses_client.send_email(
                Source='eliyahu@eliyahu.ai',  # Use verified email address
                Destination={'ToAddresses': ['eliyahu@eliyahu.ai']},
                Message={
                    'Subject': {'Data': admin_subject, 'Charset': 'UTF-8'},
                    'Body': {'Text': {'Data': admin_body, 'Charset': 'UTF-8'}}
                }
            )
            logger.debug(f"[PREVIEW_ALERT] Admin notification sent")
        except Exception as e:
            logger.error(f"[PREVIEW_ALERT] Failed to send admin notification: {e}")

    except Exception as e:
        logger.error(f"[PREVIEW_ALERT] Failed to send preview failure alerts: {e}")

def _calculate_cost_estimated(token_usage: Dict, metadata: Dict) -> float:
    """
    Calculate estimated cost without caching benefits for projections.
    This estimates what the cost would be if no caching were available.
    
    Args:
        token_usage: Token usage data from validation lambda
        metadata: Additional metadata from validation results
        
    Returns:
        Estimated cost without caching benefits
    """
    try:
        # Extract actual cost and cached call information
        actual_cost = token_usage.get('total_cost', 0.0)
        total_calls = token_usage.get('api_calls', 0)
        cached_calls = token_usage.get('cached_calls', 0)
        
        # If no caching occurred, estimated cost equals actual cost
        if cached_calls == 0 or total_calls == 0:
            logger.debug(f"[COST_CALC] No caching detected - estimated cost equals actual: ${actual_cost:.6f}")
            return actual_cost
        
        # Calculate cache hit rate
        cache_hit_rate = cached_calls / max(1, total_calls)
        non_cached_calls = total_calls - cached_calls
        
        # Estimate cost per non-cached call
        if non_cached_calls > 0:
            cost_per_non_cached_call = actual_cost / non_cached_calls
            # Estimated cost if all calls were non-cached
            estimated_cost = cost_per_non_cached_call * total_calls
            
            logger.debug(f"[COST_CALC] Cache analysis - Hit rate: {cache_hit_rate:.2%}, "
                        f"Cost per call: ${cost_per_non_cached_call:.6f}, "
                        f"Estimated without cache: ${estimated_cost:.6f}")
            
            return estimated_cost
        else:
            # All calls were cached - use average cost estimation
            # This is a fallback case for when cached costs aren't tracked properly
            logger.warning(f"[COST_CALC] All {total_calls} calls were cached - using fallback estimation")
            
            # Fallback: assume average cost per call based on token usage
            total_tokens = token_usage.get('total_tokens', 0)
            if total_tokens > 0:
                # Use rough estimation based on tokens (typical cost per 1000 tokens)
                estimated_cost_per_1k_tokens = 0.003  # Conservative estimate
                estimated_cost = (total_tokens / 1000) * estimated_cost_per_1k_tokens
                return max(actual_cost, estimated_cost)  # Use higher of actual or estimated
            
            return actual_cost  # Final fallback
            
    except Exception as e:
        logger.error(f"[COST_CALC] Error calculating estimated cost without cache: {e}")
        # Fallback to actual cost with small buffer for safety
        return token_usage.get('total_cost', 0.0) * 1.2  # 20% buffer for estimation errors

def _apply_domain_multiplier_with_validation(email: str, base_cost: float, session_id: str = None) -> Dict[str, float]:
    """
    Apply domain multiplier with comprehensive validation and audit logging.
    
    Args:
        email: User email address
        base_cost: Base cost before multiplier
        session_id: Session ID for audit trail
        
    Returns:
        Dictionary with multiplier details and final quoted cost
    """
    try:
        from dynamodb_schemas import get_domain_multiplier
        from decimal import Decimal, ROUND_UP
        
        # Extract domain with validation
        if not email or '@' not in email:
            logger.error(f"[MULTIPLIER_ERROR] Invalid email format: {email}")
            return {
                'multiplier': 1.0,
                'domain': 'invalid',
                'base_cost': base_cost,
                'quoted_cost': max(2.0, math.ceil(base_cost)),  # $2 minimum, rounded up
                'error': 'invalid_email'
            }
        
        domain = email.split('@')[-1].lower().strip()
        if not domain:
            logger.error(f"[MULTIPLIER_ERROR] Empty domain extracted from email: {email}")
            domain = 'unknown'
        
        # Get multiplier with validation
        multiplier = float(get_domain_multiplier(domain))
        
        # Validate multiplier is reasonable
        if multiplier <= 0 or multiplier > 100:  # Sanity check
            logger.error(f"[MULTIPLIER_ERROR] Unreasonable multiplier {multiplier} for domain {domain}")
            multiplier = 5.0  # Default fallback
        
        # Calculate quoted cost with multiplier, rounding, and minimum
        cost_with_multiplier = base_cost * multiplier
        quoted_cost = max(2.0, math.ceil(cost_with_multiplier))  # $2 minimum, rounded up to nearest dollar
        
        # Audit logging
        logger.debug(f"[MULTIPLIER_AUDIT] Domain: {domain}, Base: ${base_cost:.6f}, "
                   f"Multiplier: {multiplier}x, With multiplier: ${cost_with_multiplier:.6f}, "
                   f"Final quoted: ${quoted_cost:.2f}, Session: {session_id}")
        
        return {
            'multiplier': multiplier,
            'domain': domain,
            'base_cost': base_cost,
            'cost_with_multiplier': cost_with_multiplier,
            'quoted_cost': quoted_cost,
            'minimum_applied': quoted_cost == 2.0,
            'rounding_applied': quoted_cost > cost_with_multiplier
        }
        
    except Exception as e:
        logger.error(f"[MULTIPLIER_ERROR] Failed to apply domain multiplier for {email}: {e}")
        # Safe fallback
        return {
            'multiplier': 5.0,  # Default multiplier
            'domain': 'error',
            'base_cost': base_cost,
            'quoted_cost': max(2.0, math.ceil(base_cost * 5.0)),
            'error': str(e)
        }

def calculate_discount(session_id: str, config_version: int, quoted_validation_cost: float) -> float:
    """
    Calculate discount based on session type and configuration version.

    Business Rules:
    - Demo sessions (session_id contains 'demo') with v1 configs get 100% discount
    - All other sessions get 0 discount

    Args:
        session_id: Session identifier
        config_version: Configuration version number (1, 2, etc.)
        quoted_validation_cost: The quoted cost before discount

    Returns:
        Discount amount (0.0 or quoted_validation_cost for full discount)
    """
    try:
        # Check if this is a demo session with v1 config
        is_demo = 'demo' in session_id.lower() if session_id else False
        is_v1_config = (config_version == 1) or (str(config_version) == '1')

        if is_demo and is_v1_config:
            discount = float(quoted_validation_cost)
            logger.info(f"[DISCOUNT] Demo session with v1 config - applying 100% discount: ${discount:.2f}")
            logger.info(f"[DISCOUNT]   session_id='{session_id}', config_version={config_version}")
            return discount
        else:
            logger.debug(f"[DISCOUNT] No discount applied - is_demo={is_demo}, is_v1={is_v1_config}")
            return 0.0

    except Exception as e:
        logger.error(f"[DISCOUNT_ERROR] Failed to calculate discount: {e}")
        return 0.0  # Safe fallback - no discount on error

def _create_fallback_preview_excel(validation_results, config_data, input_filename, is_full=False):
    """Create a basic Excel file using openpyxl when xlsxwriter is not available."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation Results"
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Get unique column names from validation results
        all_columns = set()
        for result in validation_results:
            all_columns.update(result.keys())
        
        # Sort columns for consistent ordering
        column_list = sorted(list(all_columns))
        
        # Write headers
        for col_idx, column in enumerate(column_list, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data rows
        for row_idx, result in enumerate(validation_results, 2):
            for col_idx, column in enumerate(column_list, 1):
                value = result.get(column, "")
                # Handle complex objects by converting to string
                if isinstance(value, (dict, list)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx in range(1, len(column_list) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 15
        
        # Add metadata sheet
        meta_ws = wb.create_sheet("Metadata")
        meta_ws.append(["Key", "Value"])
        meta_ws.append(["Original File", input_filename])
        meta_ws.append(["Validation Type", "Full Validation" if is_full else "Preview"])
        meta_ws.append(["Timestamp", datetime.now().isoformat()])
        meta_ws.append(["Total Rows", len(validation_results) if validation_results else 0])
        meta_ws.append(["Total Columns", len(column_list)])
        
        if config_data:
            config_version = config_data.get('storage_metadata', {}).get('version', 1)
            meta_ws.append(["Config Version", config_version])
        
        # Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error creating fallback Excel: {str(e)}")
        return None

def _get_api_gateway_management_client(context):
    """Initializes the API Gateway Management client."""
    global api_gateway_management_client
    if api_gateway_management_client is None:
        # The endpoint URL must be retrieved from the context of the WebSocket API request
        # For the background handler, it must be passed in via an environment variable.
        endpoint_url = os.environ.get('WEBSOCKET_API_URL')
        if not endpoint_url:
            logger.error("WEBSOCKET_API_URL environment variable not set.")
            return None
        
        # Convert WebSocket URL (wss://) to HTTPS endpoint for API Gateway Management
        if endpoint_url.startswith('wss://'):
            endpoint_url = endpoint_url.replace('wss://', 'https://')
            logger.debug(f"Converted WebSocket URL to HTTPS endpoint: {endpoint_url}")
        
        # Ensure the endpoint URL includes the stage (prod)
        if not endpoint_url.endswith('/prod'):
            endpoint_url = endpoint_url.rstrip('/') + '/prod'
            logger.debug(f"Added /prod stage to endpoint URL: {endpoint_url}")
        
        api_gateway_management_client = boto3.client('apigatewaymanagementapi', endpoint_url=endpoint_url)
    return api_gateway_management_client

def _send_balance_update(session_id: str, balance_data: dict):
    """Send balance update via WebSocket."""
    from dynamodb_schemas import get_connection_by_session, remove_websocket_connection
    
    connection_id = get_connection_by_session(session_id)
    if connection_id:
        try:
            client = _get_api_gateway_management_client(None)
            if client:
                client.post_to_connection(
                    ConnectionId=connection_id,
                    Data=json.dumps(balance_data).encode('utf-8')
                )
                logger.debug(f"Balance update sent to WebSocket {connection_id}")
        except client.exceptions.GoneException:
            logger.warning(f"WebSocket connection {connection_id} is stale. Removing from DB.")
            remove_websocket_connection(connection_id)
        except Exception as e:
            logger.error(f"Failed to send balance update to WebSocket {connection_id}: {e}")

def _update_progress(session_id: str, status: str, processed_rows: int, verbose_status: str, percent_complete: int, total_rows: int = None, current_batch: int = None, total_batches: int = None):
    """Callback function to update DynamoDB and push WebSocket messages."""
    from dynamodb_schemas import update_run_status, get_connection_by_session, remove_websocket_connection, find_existing_run_key
    
    
    # 1. Update DynamoDB
    run_key = find_existing_run_key(session_id)
    if run_key:
        update_run_status(session_id=session_id, run_key=run_key, status=status,
            processed_rows=processed_rows, verbose_status=verbose_status,
            percent_complete=percent_complete)

    # 2. Push update over WebSocket
    connection_id = get_connection_by_session(session_id)
    if connection_id:
        try:
            client = _get_api_gateway_management_client(None) # Context is not available here
            if client:
                payload = {
                    'status': status, 'processed_rows': processed_rows,
                    'verbose_status': verbose_status, 'percent_complete': percent_complete
                }
                # Include total_rows if provided (mainly for completion messages)
                if total_rows is not None:
                    payload['total_rows'] = total_rows
                # Include batch information if provided
                if current_batch is not None:
                    payload['current_batch'] = current_batch
                if total_batches is not None:
                    payload['total_batches'] = total_batches
                    
                client.post_to_connection(
                    ConnectionId=connection_id,
                    Data=json.dumps(payload).encode('utf-8')
                )
        except client.exceptions.GoneException:
            logger.warning(f"WebSocket connection {connection_id} is stale. Removing from DB.")
            remove_websocket_connection(connection_id)
        except Exception as e:
            logger.error(f"Failed to post to WebSocket connection {connection_id}: {e}")

# Environment variables will be needed here, or passed as arguments.
def get_unified_bucket():
    """Get the unified S3 bucket name from environment"""
    return os.environ.get('S3_UNIFIED_BUCKET', 'hyperplexity-storage')

S3_UNIFIED_BUCKET = get_unified_bucket()
S3_RESULTS_BUCKET = os.environ.get('S3_RESULTS_BUCKET', 'perplexity-results')
VALIDATOR_LAMBDA_NAME = os.environ.get('VALIDATOR_LAMBDA_NAME', 'perplexity-validator')


def verify_async_validation_trigger(session_id, sqs_message_id, run_key=None, timeout_seconds=30):
    """
    Verify that the async validation lambda gets triggered within the timeout period.
    Returns True if triggered successfully, False otherwise.
    """
    import time

    try:
        logger.debug(f"[VALIDATION_TRIGGER_CHECK] Monitoring DynamoDB for async validation start (session: {session_id}, timeout: {timeout_seconds}s)...")

        # The validation lambda updates DynamoDB run status to 'ASYNC_PROCESSING_STARTED' when it begins
        # We'll monitor for this status change

        # Use the run_key passed as parameter
        if not run_key:
            # If no run_key provided, we need to get the latest run for this session
            logger.debug(f"[VALIDATION_TRIGGER_CHECK] No run_key provided, will query for latest run")
        else:
            logger.debug(f"[VALIDATION_TRIGGER_CHECK] Using run_key: {run_key}")

        check_interval = 2  # Check every 2 seconds
        max_checks = timeout_seconds // check_interval

        for check_num in range(max_checks):
            try:
                # Check DynamoDB for the run status
                dynamodb = boto3.resource('dynamodb')
                table = dynamodb.Table('perplexity-validator-runs')

                if run_key:
                    # Direct lookup if we have the run_key
                    response = table.get_item(
                        Key={
                            'session_id': session_id,
                            'run_key': run_key
                        }
                    )

                    if 'Item' in response:
                        status = response['Item'].get('status', '')
                        verbose_status = response['Item'].get('verbose_status', '')

                        logger.debug(f"[VALIDATION_TRIGGER_CHECK] DynamoDB status: {status}, verbose: {verbose_status}")

                        # Check for any async processing status (initial or continuation)
                        if (status == 'ASYNC_PROCESSING_STARTED' or
                            status.startswith('ASYNC_CONTINUATION_') or
                            'async' in status.lower()):

                            # Extract continuation info if available
                            continuation_chain = response['Item'].get('continuation_chain', [])
                            if continuation_chain:
                                latest_continuation = continuation_chain[-1]
                                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Found async status: {status}")
                                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Latest continuation: #{latest_continuation.get('continuation_number', 0)} started at {latest_continuation.get('started_at', 'unknown')}")
                            else:
                                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Found ASYNC_PROCESSING_STARTED status in DynamoDB!")

                            _send_websocket_message(session_id, {
                                'type': 'status_update',
                                'status': '[SUCCESS] Async validation lambda confirmed active via DynamoDB',
                                'session_id': session_id
                            })

                            return True
                else:
                    # Query for the latest run for this session
                    response = table.query(
                        KeyConditionExpression='session_id = :sid',
                        ExpressionAttributeValues={
                            ':sid': session_id
                        },
                        ScanIndexForward=False,  # Get most recent first
                        Limit=1
                    )

                    if response.get('Items'):
                        item = response['Items'][0]
                        status = item.get('status', '')
                        verbose_status = item.get('verbose_status', '')
                        found_run_key = item.get('run_key', '')

                        logger.debug(f"[VALIDATION_TRIGGER_CHECK] Latest run - status: {status}, run_key: {found_run_key}")

                        # Check for any async processing status (initial or continuation)
                        if (status == 'ASYNC_PROCESSING_STARTED' or
                            status.startswith('ASYNC_CONTINUATION_') or
                            'async' in status.lower()):

                            # Extract continuation info if available
                            continuation_chain = item.get('continuation_chain', [])
                            if continuation_chain:
                                latest_continuation = continuation_chain[-1]
                                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Found async status in latest run: {status}")
                                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Latest continuation: #{latest_continuation.get('continuation_number', 0)} started at {latest_continuation.get('started_at', 'unknown')}")
                            else:
                                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Found ASYNC_PROCESSING_STARTED in latest run!")

                            _send_websocket_message(session_id, {
                                'type': 'status_update',
                                'status': '[SUCCESS] Async validation lambda confirmed active via DynamoDB',
                                'session_id': session_id
                            })

                            return True

                # Also still check S3 as a fallback (but for early files, not final ones)
                # Check if the validation results are starting to appear in S3
                s3_client = boto3.client('s3')

                # Check for the validation results file that the lambda creates
                # The validation lambda saves to: sessions/{session_id}/complete_validation_results.json
                results_key = f"sessions/{session_id}/complete_validation_results.json"
                unified_bucket = get_unified_bucket()

                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Checking S3 for {results_key} in bucket {unified_bucket}")

                try:
                    # Try to get the results file that the validation lambda creates
                    response = s3_client.head_object(
                        Bucket=unified_bucket,
                        Key=results_key
                    )

                    # If the file exists, the validation lambda has started and saved initial results
                    logger.debug(f"[VALIDATION_TRIGGER_CHECK] Found validation results file in S3 - validation lambda is active!")
                    logger.debug(f"[VALIDATION_TRIGGER_CHECK] File size: {response.get('ContentLength', 0)} bytes, Last modified: {response.get('LastModified')}")

                    # Also send WebSocket update
                    _send_websocket_message(session_id, {
                        'type': 'status_update',
                        'status': '[SUCCESS] Async validation lambda triggered and processing',
                        'session_id': session_id
                    })

                    return True

                except s3_client.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == '404':
                        # File doesn't exist yet, continue checking
                        logger.debug(f"[VALIDATION_TRIGGER_CHECK] Results file not found yet (check {check_num + 1})")
                    else:
                        logger.debug(f"[VALIDATION_TRIGGER_CHECK] S3 check error: {e}")

                # Also check for any session files as a backup indicator
                session_prefix = f"sessions/{session_id}/"
                try:
                    response = s3_client.list_objects_v2(
                        Bucket=unified_bucket,
                        Prefix=session_prefix,
                        MaxKeys=1
                    )

                    if response.get('KeyCount', 0) > 0:
                        logger.debug(f"[VALIDATION_TRIGGER_CHECK] Found session files in S3 - validation lambda has started!")
                        logger.debug(f"[VALIDATION_TRIGGER_CHECK] First file: {response['Contents'][0]['Key']}")

                        _send_websocket_message(session_id, {
                            'type': 'status_update',
                            'status': '[SUCCESS] Async validation lambda triggered (session files detected)',
                            'session_id': session_id
                        })

                        return True

                except Exception as list_error:
                    logger.debug(f"[VALIDATION_TRIGGER_CHECK] S3 list error: {list_error}")

                # Also check CloudWatch metrics as a backup
                validator_lambda_name = os.environ.get('VALIDATOR_LAMBDA_NAME', 'perplexity-validator')
                cloudwatch = boto3.client('cloudwatch')

                end_time = datetime.now(timezone.utc)
                start_time = end_time - timedelta(minutes=2)

                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Checking CloudWatch metrics for {validator_lambda_name}")
                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Time window: {start_time} to {end_time}")

                response = cloudwatch.get_metric_statistics(
                    Namespace='AWS/Lambda',
                    MetricName='Invocations',
                    Dimensions=[{'Name': 'FunctionName', 'Value': validator_lambda_name}],
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=60,
                    Statistics=['Sum']
                )

                datapoints = response.get('Datapoints', [])
                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Got {len(datapoints)} datapoints from CloudWatch")
                for dp in datapoints:
                    logger.debug(f"[VALIDATION_TRIGGER_CHECK]   {dp['Timestamp']}: {dp['Sum']} invocations")

                recent_invocations = sum([point['Sum'] for point in datapoints])
                if recent_invocations > 0:
                    logger.debug(f"[VALIDATION_TRIGGER_CHECK] Found {recent_invocations} recent lambda invocations via metrics")

                    # Give it a moment for S3 files to appear
                    if check_num >= 2:  # After 4+ seconds
                        _send_websocket_message(session_id, {
                            'type': 'status_update',
                            'status': '[SUCCESS] Async validation lambda triggered (confirmed via metrics)',
                            'session_id': session_id
                        })
                        return True

            except Exception as check_error:
                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Check error (attempt {check_num + 1}): {check_error}")

            # Wait before next check
            if check_num < max_checks - 1:
                time.sleep(check_interval)
                logger.debug(f"[VALIDATION_TRIGGER_CHECK] Waiting {check_interval}s before next check...")

        # Timeout reached without finding validation lambda activity
        logger.error(f"[VALIDATION_TRIGGER_CHECK] Async validation lambda failed to start for session {session_id} within {timeout_seconds}s")
        logger.warning(f"[VALIDATION_TRIGGER_CHECK] No validation lambda activity detected for session {session_id} within {timeout_seconds}s")

        # Check if lambda actually failed by checking CloudWatch errors
        try:
            cloudwatch = boto3.client('cloudwatch')
            validator_lambda_name = os.environ.get('VALIDATOR_LAMBDA_NAME', 'perplexity-validator')

            end_time = datetime.now(timezone.utc)
            start_time = end_time - timedelta(minutes=5)

            error_response = cloudwatch.get_metric_statistics(
                Namespace='AWS/Lambda',
                MetricName='Errors',
                Dimensions=[{'Name': 'FunctionName', 'Value': validator_lambda_name}],
                StartTime=start_time,
                EndTime=end_time,
                Period=60,
                Statistics=['Sum']
            )

            total_errors = sum(dp['Sum'] for dp in error_response['Datapoints'])
            if total_errors > 0:
                logger.error(f"[VALIDATION_FAILURE] Detected {total_errors} lambda errors in last 5 minutes - validation likely failed")
                status_msg = f'[ERROR] Async validation lambda failed - {total_errors} errors detected'
            else:
                logger.debug(f"[VALIDATION_TRIGGER_CHECK] No lambda errors detected - validator may not have been triggered")
                status_msg = f'[WARNING] Async validation not detected after {timeout_seconds}s - falling back to sync'

        except Exception as metric_error:
            logger.debug(f"[VALIDATION_TRIGGER_CHECK] Could not check CloudWatch errors: {metric_error}")
            status_msg = f'[WARNING] Async validation not detected after {timeout_seconds}s - falling back to sync'

        _send_websocket_message(session_id, {
            'type': 'status_update',
            'status': status_msg,
            'session_id': session_id
        })

        return False

    except Exception as e:
        logger.error(f"[VALIDATION_TRIGGER_CHECK] Error verifying async validation trigger: {e}")
        return False


def handle_async_completion_in_background_handler(event, context):
    """
    Handle async completion within the background handler.
    Load results from S3, reconstruct variables, and continue to common completion code.
    """
    try:
        session_id = event.get('session_id')
        results_s3_key = event.get('results_s3_key')

        if not session_id or not results_s3_key:
            logger.error("[ASYNC_COMPLETION] Missing required fields: session_id or results_s3_key")
            return {'statusCode': 400, 'body': json.dumps({'error': 'Missing required fields'})}

        logger.debug(f"[ASYNC_COMPLETION] Loading results for session {session_id} from {results_s3_key}")

        # Load validation results from S3
        try:
            # The results_s3_key should be just the key path (e.g., "sessions/xxx/complete_validation_results.json")
            # The bucket should come from the environment or event
            s3_bucket = event.get('S3_UNIFIED_BUCKET') or S3_UNIFIED_BUCKET
            s3_key = results_s3_key

            logger.debug(f"[ASYNC_COMPLETION] Loading from bucket={s3_bucket}, key={s3_key}")

            response = s3_client.get_object(Bucket=s3_bucket, Key=s3_key)
            async_validation_results = json.loads(response['Body'].read().decode('utf-8'))
            logger.debug(f"[ASYNC_COMPLETION] Loaded validation results: {len(str(async_validation_results))} bytes")

            # DEBUG: Log the actual structure of async results to understand what we're working with
            logger.debug(f"[ASYNC_COMPLETION] Async results type: {type(async_validation_results)}")
            if isinstance(async_validation_results, dict):
                logger.debug(f"[ASYNC_COMPLETION] Async results top-level keys: {list(async_validation_results.keys())}")
                # Show a sample of the structure
                preview = str(async_validation_results)[:500] + "..." if len(str(async_validation_results)) > 500 else str(async_validation_results)
                logger.debug(f"[ASYNC_COMPLETION] Async results structure preview: {preview}")

            # Async results from S3 have a different format than sync results
            # Sync expects: {statusCode, body: {data: {rows: ...}}}
            # Async has: {validation_results: {...}, token_usage: {...}, metadata: {...}}
            # We need to wrap async results to match sync format for the common path

            if isinstance(async_validation_results, dict) and 'validation_results' in async_validation_results:
                # Wrap async results in sync format
                # CRITICAL: In S3, enhanced_metrics is at top level, but sync expects it INSIDE metadata
                metadata_with_enhanced = async_validation_results.get('metadata', {}).copy()
                if 'enhanced_metrics' in async_validation_results:
                    metadata_with_enhanced['enhanced_metrics'] = async_validation_results['enhanced_metrics']
                    # Debug: Verify by_provider data is present
                    aggregated_metrics = async_validation_results['enhanced_metrics'].get('aggregated_metrics', {})
                    providers = aggregated_metrics.get('providers', {})
                    logger.debug(f"[ASYNC_COMPLETION] Added enhanced_metrics to metadata for sync compatibility")
                    logger.debug(f"[ASYNC_COMPLETION] by_provider data present: {list(providers.keys()) if providers else 'NONE'}")

                validation_results = {
                    'statusCode': 200,
                    'body': {
                        'data': {
                            'rows': async_validation_results['validation_results'],
                            # CRITICAL: QC results and metrics must be in body.data to match sync format
                            'qc_results': async_validation_results.get('qc_results', {}),
                            'qc_metrics': async_validation_results.get('qc_metrics', {})
                        },
                        'token_usage': async_validation_results.get('token_usage', {}),
                        'metadata': metadata_with_enhanced
                    },
                    # Preserve the original structure for other uses
                    'validation_results': async_validation_results['validation_results'],
                    'token_usage': async_validation_results.get('token_usage', {}),
                    'metadata': metadata_with_enhanced,
                    # Also preserve at top level for backward compatibility
                    'qc_results': async_validation_results.get('qc_results', {}),
                    'qc_metrics': async_validation_results.get('qc_metrics', {})
                }
                logger.debug(f"[ASYNC_COMPLETION] Wrapped async results in sync format for common processing path")
                if async_validation_results.get('qc_results'):
                    logger.debug(f"[ASYNC_COMPLETION] QC results preserved: {len(async_validation_results['qc_results'])} rows")
                if async_validation_results.get('qc_metrics'):
                    logger.debug(f"[ASYNC_COMPLETION] QC metrics preserved: {async_validation_results['qc_metrics'].get('total_fields_reviewed', 0)} fields reviewed")
            else:
                # Fallback if structure is unexpected
                validation_results = async_validation_results
                logger.warning(f"[ASYNC_COMPLETION] Unexpected async results structure, using as-is")

        except Exception as e:
            logger.error(f"[ASYNC_COMPLETION] Failed to load results from s3://{s3_bucket}/{s3_key}: {e}")
            return {'statusCode': 404, 'body': json.dumps({'error': f'Could not load results: {str(e)}'})}

        # CRITICAL: Check if validation failed before proceeding
        # Failed validations should NOT charge the customer
        validation_failed = False
        failure_reason = None

        if isinstance(async_validation_results, dict):
            # Check for explicit failure status
            status = async_validation_results.get('status', '')
            if status in ['FAILED', 'FAILED_INCOMPLETE', 'FAILED_NO_PROGRESS']:
                validation_failed = True
                failure_reason = status
                logger.error(f"[BILLING_PROTECTION] Validation marked as {status} - will NOT charge customer")

            # Check for validation_error field
            if async_validation_results.get('validation_error'):
                validation_failed = True
                failure_reason = async_validation_results.get('validation_error')
                logger.error(f"[BILLING_PROTECTION] Validation error detected: {failure_reason} - will NOT charge customer")

            # Check if incomplete (rows processed < total rows)
            validation_results_data = async_validation_results.get('validation_results', {})
            if isinstance(validation_results_data, dict):
                rows_processed = len(validation_results_data)
                # Get total rows from continuation_metadata if available, otherwise from validation_data in context
                total_rows = async_validation_results.get('continuation_metadata', {}).get('total_rows', 0)
                if total_rows > 0 and rows_processed < total_rows:
                    validation_failed = True
                    failure_reason = f"Incomplete validation: {rows_processed}/{total_rows} rows processed"
                    logger.error(f"[BILLING_PROTECTION] {failure_reason} - will NOT charge customer")

        # Load delegation context from DynamoDB
        try:
            delegation_context = load_delegation_context(session_id)
            if not delegation_context:
                logger.error(f"[ASYNC_COMPLETION] No delegation context found for session {session_id}")
                return {'statusCode': 404, 'body': json.dumps({'error': 'No delegation context found'})}

            logger.debug(f"[ASYNC_COMPLETION] Loaded delegation context for session {session_id}")

        except Exception as e:
            logger.error(f"[ASYNC_COMPLETION] Failed to load delegation context: {e}")
            return {'statusCode': 500, 'body': json.dumps({'error': f'Failed to load context: {str(e)}'})}

        # Reconstruct the event and variables as if sync processing had just completed
        request_context = delegation_context.get('request_context', {})
        file_locations = delegation_context.get('file_locations', {})
        processing_metadata = delegation_context.get('processing_metadata', {})

        # Create reconstructed event for the common completion code
        reconstructed_event = {
            'session_id': session_id,
            'run_key': request_context.get('run_key'),
            'email': request_context.get('email'),
            'email_address': request_context.get('email'),
            'preview_mode': request_context.get('preview_mode', False),
            'excel_s3_key': file_locations.get('excel_s3_key'),
            'config_s3_key': file_locations.get('config_s3_key'),
            'results_key': file_locations.get('results_key'),
            'total_rows': request_context.get('total_rows'),
            'reference_pin': request_context.get('reference_pin'),
            'timestamp': request_context.get('timestamp'),
            'background_processing': True,
            'async_completion_mode': True,  # Flag to indicate this is async completion
            # Restore additional critical variables
            'clean_session_id': request_context.get('clean_session_id', session_id),
            'preview_email': request_context.get('preview_email', False),
            'email_folder': request_context.get('email_folder', 'Full Validation'),
            'batch_size': request_context.get('batch_size'),
            'max_rows': request_context.get('max_rows'),
            'config_data': request_context.get('config_data'),
            # Restore processing metadata for Excel generation
            'table_data': processing_metadata.get('table_data'),
            'config_version': processing_metadata.get('config_version', 'v1'),
            'S3_UNIFIED_BUCKET': processing_metadata.get('S3_UNIFIED_BUCKET', S3_UNIFIED_BUCKET)
        }

        logger.debug(f"[ASYNC_COMPLETION] Reconstructed event for session {session_id}, proceeding to common completion")

        # Add validation results to event and process with main handler
        reconstructed_event['_validation_results_from_async'] = validation_results
        reconstructed_event['_skip_validation_call'] = True

        # CRITICAL: Pass validation_failed flag to main handler to prevent billing
        if validation_failed:
            reconstructed_event['_validation_failed'] = True
            reconstructed_event['_validation_failure_reason'] = failure_reason
            logger.error(f"[ASYNC_COMPLETION] Validation failed - passing failure info to main handler: {failure_reason}")

        logger.debug(f"[ASYNC_COMPLETION] Calling main handler for completion processing")

        # Process the reconstructed event through the main handler (without async completion check)
        return handle_main_processing(reconstructed_event, context)

    except Exception as e:
        logger.error(f"[ASYNC_COMPLETION] Failed to handle async completion: {e}")
        return {'statusCode': 500, 'body': json.dumps({'error': str(e)})}



def handle(event, context):
    """Handle background processing for both normal and preview mode validation."""
    # ========== ASYNC COMPLETION CHECK ==========
    # Check if this is async completion (validation lambda finished and we need to complete the job)
    # Check both old field name and new message_type for compatibility
    if (event.get('async_completion') or
        event.get('message_type') == 'ASYNC_COMPLETION_REQUEST' or
        event.get('message_type') == 'ASYNC_VALIDATION_COMPLETE'):
        logger.debug(f"[ASYNC_COMPLETION] Processing async completion request (message_type={event.get('message_type')}, async_completion={event.get('async_completion', False)})")
        return handle_async_completion_in_background_handler(event, context)

    # Normal processing
    return handle_main_processing(event, context)


def handle_main_processing(event, context):
    """Main processing logic for background validation."""
    try:

        from ..core.validator_invoker import invoke_validator_lambda
        from ..reporting.markdown_report import create_markdown_table_from_results

        # Initialize S3 client for file operations
        import boto3
        s3_client = boto3.client('s3')

        try:
            from excel_report_qc_unified import create_qc_enhanced_excel_for_interface
            EXCEL_ENHANCEMENT_AVAILABLE = True
        except ImportError:
            try:
                from ..reporting.excel_report_new import create_enhanced_excel_with_validation, EXCEL_ENHANCEMENT_AVAILABLE
                # Fallback function that maps to QC interface
                def create_qc_enhanced_excel_for_interface(table_data, validation_results, config_data, session_id, validated_sheet_name=None):
                    return create_enhanced_excel_with_validation(table_data, validation_results, config_data, session_id, validated_sheet_name)
            except ImportError:
                EXCEL_ENHANCEMENT_AVAILABLE = False
                def create_qc_enhanced_excel_for_interface(*args, **kwargs): return None
        
        try:
            from email_sender import send_validation_results_email
            EMAIL_SENDER_AVAILABLE = True
        except ImportError:
            EMAIL_SENDER_AVAILABLE = False

        logger.info("--- Background Handler Started ---")
        
        # Record when background handler actually starts processing
        background_start_time = datetime.now(timezone.utc).isoformat()
        
        # Check if this is a config generation request
        if event.get('request_type') == 'config_generation':
            logger.info("Detected config generation request, forwarding to validation lambda")
            return handle_config_generation(event, context)
        
        # Extract parameters from event, ensuring correct types
        # Check both preview_mode and request_type for compatibility
        is_preview = event.get('preview_mode', False) or event.get('request_type') == 'preview'
        session_id = event['session_id']
        timestamp = event.get('timestamp')

        # Always use a clean session ID format: session_YYYY-MM-DDTHH_MM_SS_XXXXXX
        # Ensure session ID has proper prefix (moved up to fix UnboundLocalError)
        clean_session_id = session_id
        if not clean_session_id.startswith('session_'):
            clean_session_id = f"session_{clean_session_id}"

        reference_pin = event.get('reference_pin', '000000')

        # If reference_pin is 'preview' or similar, extract it from clean_session_id
        if reference_pin == 'preview' or not reference_pin or reference_pin == '000000':
            # Extract reference pin from clean session ID
            if '_' in clean_session_id:
                reference_pin = clean_session_id.split('_')[-1]
            else:
                reference_pin = clean_session_id[:6]
        excel_s3_key = event['excel_s3_key']
        config_s3_key = event['config_s3_key']
        email = event.get('email', 'unknown@example.com').lower().strip()
        preview_email = event.get('preview_email', False)
        # Extract parameters early before using them
        try:
            max_rows_str = event.get('max_rows')
            max_rows = int(max_rows_str) if max_rows_str else 1000
            
            batch_size_str = event.get('batch_size')
            batch_size = int(batch_size_str) if batch_size_str else None  # Let enhanced batch manager determine optimal size
        except (ValueError, TypeError):
            logger.warning("Invalid max_rows or batch_size, using defaults.")
            max_rows = 1000
            batch_size = None  # Let enhanced batch manager determine optimal size

        run_key = event.get('run_key')  # Extract run_key from SQS message
        
        # Handle case where run_key is missing (backward compatibility)
        if not run_key:
            logger.warning(f"No run_key found in SQS message for session {session_id}, attempting to find existing run")
            # Try to find existing run record first to avoid duplicates
            try:
                from dynamodb_schemas import find_existing_run_key
                run_key = find_existing_run_key(session_id)
                
                if run_key:
                    logger.debug(f"Found existing run_key for session {session_id}: {run_key}")
                else:
                    # Only create new run record if none exists
                    run_type_initial = "Preview" if is_preview else "Validation"
                    run_key = create_run_record(session_id=session_id, email=email, total_rows=0, batch_size=batch_size or 10, run_type=run_type_initial)
                    logger.debug(f"Created new run_key for backward compatibility: {run_key}")
            except Exception as e:
                logger.error(f"Failed to find/create run_key for backward compatibility: {e}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
                run_key = None
        
        # Always use a clean session ID format: session_YYYY-MM-DDTHH_MM_SS_XXXXXX
        # Ensure session ID has proper prefix
        clean_session_id = session_id
        if not clean_session_id.startswith('session_'):
            clean_session_id = f"session_{clean_session_id}"
            
        logger.debug(f"Background handler called for session {session_id}, clean_session_id={clean_session_id}, is_preview={is_preview}")

        email_folder = event.get('email_folder', 'default')
        email_address = event.get('email_address')
        if email_address:
            email_address = email_address.lower().strip()
        
        # Initialize processing_time early to prevent "referenced before assignment" error
        processing_time = 0.0
        
        # Create a wrapper function for run status updates with this session's run_key
        def update_run_status_for_session(**kwargs):
            if run_key:
                try:
                    return update_run_status(session_id=session_id, run_key=run_key, **kwargs)
                except Exception as e:
                    logger.error(f"Failed to update run status for session {session_id}, run_key {run_key}: {e}")
                    logger.error(f"Status update kwargs: {kwargs}")
                    return None
            else:
                logger.warning(f"Skipping run status update due to missing run_key for session {session_id}: {kwargs}")
                return None
        
        # After a job starts
        run_type_initial = "Preview" if is_preview else "Validation"
        update_run_status_for_session(status='PROCESSING', run_type=run_type_initial)

        if is_preview:
            preview_max_rows = event.get('preview_max_rows', 5)
            sequential_call_num = event.get('sequential_call')
            logger.debug(f"Background preview processing for session {session_id}")
            
            # Send initial WebSocket progress update - interface setup range 0-5%
            _send_websocket_message_deduplicated(session_id, {
                'type': 'preview_progress',
                'progress': 1,  # Interface setup: 0-5% range
                'status': f'🚀 Starting preview validation for {preview_max_rows} rows...',
                'session_id': session_id,
                'preview_max_rows': preview_max_rows
            }, "preview_progress_start")
            
            # Initial status update for preview - use 0-5% range for interface setup
            update_run_status_for_session(
                status='PROCESSING', 
                run_type="Preview",
                verbose_status=f"Starting preview for {preview_max_rows} rows...",
                percent_complete=2,  # Interface setup: 0-5% range
                batch_size=preview_max_rows
            )

            # Send file retrieval progress update - interface setup range 0-5%
            _send_websocket_message_deduplicated(session_id, {
                'type': 'preview_progress',
                'progress': 3,  # Interface setup: 0-5% range
                'status': '📁 Retrieving Excel file and configuration...',
                'session_id': session_id
            }, "preview_progress_files")
            
            # Use unified storage to get files for preview
            from ..core.unified_s3_manager import UnifiedS3Manager
            storage_manager = UnifiedS3Manager()
            
            # Get files from unified storage using clean session ID
            excel_content, actual_excel_s3_key = storage_manager.get_excel_file(email, clean_session_id)
            config_data, actual_config_s3_key = storage_manager.get_latest_config(email, clean_session_id)
            
            if not excel_content or not config_data:
                logger.error(f"Failed to retrieve files from unified storage for preview session {clean_session_id}")

                # Send error WebSocket update
                error_payload = {
                    'type': 'preview_failed',
                    'progress': 100,
                    'status': '❌ Failed to retrieve files for preview',
                    'session_id': session_id,
                    'error': 'Files not found'
                }
                _send_websocket_message_deduplicated(session_id, error_payload, "preview_failed_files")

                update_run_status_for_session(
                    status='FAILED',
                    run_type="Preview",
                    verbose_status="Failed to retrieve files for preview.",
                    percent_complete=100,
                    processed_rows=0,
                    batch_size=10,
                    eliyahu_cost=0.0,
                    quoted_validation_cost=0.0,
                    estimated_validation_eliyahu_cost=0.0,
                    time_per_row_seconds=0.0
                )

                # Update session_info.json with failure status and correct config version
                try:
                    # Try to get config version even if config_data is None
                    config_version = 1
                    config_id = None
                    if config_data and isinstance(config_data, dict):
                        config_version = config_data.get('storage_metadata', {}).get('version', 1)
                        config_id = config_data.get('storage_metadata', {}).get('config_id')

                    storage_manager.update_session_results(
                        email=email,
                        session_id=clean_session_id,
                        operation_type="preview",
                        config_id=config_id or f"{clean_session_id}_config_v{config_version}",
                        version=config_version,
                        run_key=run_key,
                        status="failed",
                        completed_at=datetime.now(timezone.utc).isoformat(),
                        frontend_payload=error_payload
                    )
                    logger.debug(f"[SESSION_TRACKING] Updated session_info.json with preview failure (files not found)")
                except Exception as e:
                    logger.error(f"Failed to update session_info.json for preview failure: {e}")

                return {'statusCode': 500, 'body': json.dumps({'status': 'failed', 'error': 'Files not found'})}

            # Parse table_data ONCE at the beginning (don't re-parse later for Excel)
            from shared_table_parser import S3TableParser
            from row_key_utils import generate_row_key
            table_parser = S3TableParser()
            table_data = table_parser.parse_s3_table(storage_manager.bucket_name, actual_excel_s3_key, extract_formulas=True)

            # Add row keys to table_data using hybrid hashing (same logic as async validation)
            if table_data and isinstance(table_data, dict) and 'data' in table_data:
                excel_rows = table_data['data']

                # Extract ID fields from config
                id_fields = []
                for target in config_data.get('validation_targets', []):
                    if target.get('importance', '').upper() == 'ID':
                        field_name = target.get('name') or target.get('column')
                        if field_name:
                            id_fields.append(field_name)

                logger.debug(f"[PREVIEW_TABLE_DATA] Using ID fields for hybrid hashing: {id_fields}")

                # Hybrid hashing: ID-field for unique rows, full-row for duplicates
                id_hash_counts = {}

                # First pass: detect duplicates
                for row_idx, row_data in enumerate(excel_rows):
                    id_hash = generate_row_key(row_data, primary_keys=id_fields if id_fields else None)
                    if id_hash not in id_hash_counts:
                        id_hash_counts[id_hash] = []
                    id_hash_counts[id_hash].append(row_idx)

                # Second pass: assign row keys
                for row_idx, row_data in enumerate(excel_rows):
                    id_hash = generate_row_key(row_data, primary_keys=id_fields if id_fields else None)

                    if len(id_hash_counts[id_hash]) > 1:
                        row_key = generate_row_key(row_data, primary_keys=None)  # Full-row hash
                    else:
                        row_key = id_hash  # ID-field hash

                    row_data['_row_key'] = row_key

                logger.info(f"[PREVIEW_TABLE_DATA] Added row keys to {len(excel_rows)} rows in table_data")

            # Send validation start progress update - interface setup complete, AI work begins (5-90% reserved for validation lambda)
            _send_websocket_message_deduplicated(session_id, {
                'type': 'preview_progress',
                'progress': 5,  # Interface setup complete: hand off to validation lambda for 5-90%
                'status': f'🔍 Running AI validation on {preview_max_rows} sample rows...',
                'session_id': session_id
            }, "preview_progress_validation")
            
            # Update start_time to when background handler actually begins validation processing
            if run_key:
                update_run_status_for_session(
                    status='IN_PROGRESS',
                    run_type="Preview",
                    verbose_status=f"Running AI validation on {preview_max_rows} sample rows...",
                    start_time=background_start_time,  # Use background handler start time
                    percent_complete=5
                )
            
            try:
                validation_results = invoke_validator_lambda(
                    actual_excel_s3_key, actual_config_s3_key, max_rows, batch_size, S3_UNIFIED_BUCKET, VALIDATOR_LAMBDA_NAME,
                    preview_first_row=True, preview_max_rows=preview_max_rows, sequential_call=sequential_call_num,
                    session_id=session_id
                )

                # Validate preview results (less strict than full validation)
                if not validation_results or not isinstance(validation_results, dict):
                    raise Exception("Preview validation returned empty or invalid results")

                logger.debug(f"[PREVIEW_SUCCESS] Received preview validation results")

            except Exception as preview_error:
                error_msg = str(preview_error)
                logger.error(f"[PREVIEW_FAILURE] Preview validation error: {error_msg}")

                # Determine error type
                error_type = "Preview Processing Error"
                if "response too large" in error_msg.lower() or "413" in error_msg:
                    error_type = "Preview Response Too Large"
                elif "timeout" in error_msg.lower():
                    error_type = "Preview Timeout"
                elif "empty or invalid results" in error_msg.lower():
                    error_type = "Preview Empty Response"

                # Send lower-priority preview error emails
                send_preview_failure_alert(session_id, email_address, error_type, error_msg, {
                    'session_id': session_id,
                    'excel_s3_key': actual_excel_s3_key,
                    'config_s3_key': actual_config_s3_key,
                    'preview_max_rows': preview_max_rows,
                    'timestamp': datetime.now(timezone.utc).isoformat()
                })

                # Update run status to failed
                update_run_status_for_session(
                    status='FAILED',
                    run_type='Preview',
                    verbose_status=f'Preview processing failed: {error_type}',
                    percent_complete=100
                )

                # Send failure notification via WebSocket
                error_payload = {
                    'type': 'preview_failed',
                    'session_id': session_id,
                    'progress': 100,
                    'status': f'⚠️ Preview failed: {error_type}',
                    'error': error_type.lower().replace(' ', '_'),
                    'message': 'Preview encountered an issue. Please try saving as CSV or check table format.'
                }
                _send_websocket_message(session_id, error_payload)

                # Update session_info.json with failure status and correct config version
                try:
                    config_version = config_data.get('storage_metadata', {}).get('version', 1)
                    config_id = config_data.get('storage_metadata', {}).get('config_id')

                    storage_manager.update_session_results(
                        email=email,
                        session_id=clean_session_id,
                        operation_type="preview",
                        config_id=config_id or f"{clean_session_id}_config_v{config_version}",
                        version=config_version,
                        run_key=run_key,
                        status="failed",
                        completed_at=datetime.now(timezone.utc).isoformat(),
                        frontend_payload=error_payload
                    )
                    logger.debug(f"[SESSION_TRACKING] Updated session_info.json with preview failure ({error_type})")
                except Exception as e:
                    logger.error(f"Failed to update session_info.json for preview failure: {e}")

                # Return error response
                return {
                    'statusCode': 500,
                    'body': json.dumps({
                        'status': 'failed',
                        'error': error_type,
                        'message': 'Preview processing failed. Consider trying CSV format.',
                        'session_id': session_id
                    })
                }

            # Only continue if preview was successful
            
            # Send validation completion progress update - interface final processing begins (90-100% range)
            _send_websocket_message_deduplicated(session_id, {
                'type': 'preview_progress',
                'progress': 90,  # Interface final processing: 90-100% range
                'status': '✅ AI validation completed, processing results...',
                'session_id': session_id
            }, "preview_progress_validation_complete")
            
            # Send results processing progress update - interface final processing (90-100% range)
            _send_websocket_message_deduplicated(session_id, {
                'type': 'preview_progress',
                'progress': 92,  # Interface final processing: 90-100% range
                'status': '📊 Analyzing results and generating estimates...',
                'session_id': session_id
            }, "preview_progress_analysis")
            
            # Final status update for preview
            if validation_results and validation_results.get('validation_results'):
                # Extract metadata first
                real_results = validation_results.get('validation_results', {})

                # Extract QC data if present (it's at the top level of validation_results, not in metadata)
                qc_results = validation_results.get('qc_results', {})
                qc_metrics_data = validation_results.get('qc_metrics', {})
                logger.debug(f"[QC_DEBUG] Extracted QC data - results: {len(qc_results)} rows, metrics: {qc_metrics_data.get('total_fields_reviewed', 0)} fields reviewed")

                # Merge QC data into real_results for display purposes (QC values take precedence)
                if qc_results:
                    logger.debug(f"[QC_MERGE] Merging QC data into preview results for display")
                    logger.debug(f"[QC_MERGE_DEBUG] QC results structure: {list(qc_results.keys())[:3]}")
                    logger.debug(f"[QC_MERGE_DEBUG] real_results keys sample: {list(real_results.keys())[:3]}")

                    # Create mapping from hash keys (QC) to numeric keys (validation results)
                    # QC uses hash keys, validation results use numeric string keys
                    qc_hash_keys = list(qc_results.keys())
                    validation_numeric_keys = list(real_results.keys())

                    # Map QC hash keys to validation numeric keys by position
                    for i, qc_hash_key in enumerate(qc_hash_keys):
                        if i < len(validation_numeric_keys):
                            validation_key = validation_numeric_keys[i]
                            row_qc_data = qc_results[qc_hash_key]

                            logger.debug(f"[QC_MERGE_DEBUG] Mapping QC hash key {qc_hash_key} -> validation key {validation_key}")

                            if validation_key in real_results:
                                logger.debug(f"[QC_MERGE_DEBUG] Row {validation_key}: QC fields = {list(row_qc_data.keys())}")
                                for field_name, field_qc_data in row_qc_data.items():
                                    logger.debug(f"[QC_MERGE_DEBUG] Field {field_name}: QC data keys = {list(field_qc_data.keys()) if isinstance(field_qc_data, dict) else 'not dict'}")
                                    logger.debug(f"[QC_MERGE_DEBUG] Field {field_name}: qc_applied = {field_qc_data.get('qc_applied') if isinstance(field_qc_data, dict) else 'N/A'}")

                                    if isinstance(field_qc_data, dict) and (field_qc_data.get('qc_applied') is True or field_qc_data.get('qc_applied') == 'Yes'):
                                        # Since QC is now comprehensive, always use QC values when available
                                        qc_entry = field_qc_data.get('qc_entry', '')
                                        qc_confidence = field_qc_data.get('qc_confidence', '')
                                        logger.debug(f"[QC_MERGE_DEBUG] Field {field_name}: has qc_entry = {bool(qc_entry)}, has qc_confidence = {bool(qc_confidence)}")

                                        # Always use QC entry and confidence when available (comprehensive QC)
                                        if qc_entry and str(qc_entry).strip():
                                            real_results[validation_key][field_name]['value'] = qc_entry
                                            logger.debug(f"[QC_MERGE] {field_name}: Using QC entry value: {qc_entry}")

                                        if qc_confidence and str(qc_confidence).strip():
                                            real_results[validation_key][field_name]['confidence_level'] = qc_confidence
                                            logger.debug(f"[QC_MERGE] {field_name}: Using QC confidence: {qc_confidence}")
                logger.debug(f"[EXCEL_DEBUG] validation_results keys: {list(validation_results.keys())}")
                logger.debug(f"[EXCEL_DEBUG] real_results type: {type(real_results)}, keys: {list(real_results.keys()) if isinstance(real_results, dict) else 'N/A'}")
                if isinstance(real_results, dict) and real_results:
                    sample_key = list(real_results.keys())[0]
                    sample_value = real_results[sample_key]
                    logger.debug(f"[EXCEL_DEBUG] Sample validation entry - Key: {sample_key}, Value type: {type(sample_value)}")
                    if isinstance(sample_value, dict):
                        logger.debug(f"[EXCEL_DEBUG] Sample validation entry - Value keys: {list(sample_value.keys())}")
                        for field_name, field_data in sample_value.items():
                            if isinstance(field_data, dict):
                                logger.debug(f"[EXCEL_DEBUG] Field '{field_name}' - keys: {list(field_data.keys())}")
                                confidence = field_data.get('confidence') or field_data.get('confidence_level')
                                logger.debug(f"[EXCEL_DEBUG] Field '{field_name}' - confidence: {confidence}")
                            break  # Just show one field sample
                total_rows = validation_results.get('total_rows', 1)
                metadata = validation_results.get('metadata', {})
                total_rows_processed = validation_results.get('total_processed_rows', 1)
                token_usage = metadata.get('token_usage', {})
                # Initialize variables early to avoid scope issues
                quoted_full_cost = None
                eliyahu_cost = 0.0
                cost_estimated = 0.0
                multiplier = 1.0
                estimated_time_per_row = 0.0
                total_estimated_processing_time = 0.0
                estimated_total_time_seconds = None
                time_per_row = 0.0
                email_domain = email.split('@')[-1] if '@' in email else 'unknown'
                # ========== ENHANCED AI_API_CLIENT INTEGRATION ==========
                # Extract enhanced cost/time data from validation lambda's ai_client aggregation
                enhanced_metrics = metadata.get('enhanced_metrics', {})
                # validation_metrics is nested inside enhanced_metrics
                validation_metrics = enhanced_metrics.get('validation_metrics', {})
                logger.debug(f"[METADATA_DEBUG] enhanced_metrics keys: {list(enhanced_metrics.keys()) if enhanced_metrics else 'None'}")
                logger.debug(f"[METADATA_DEBUG] validation_metrics extracted: {validation_metrics}")

                # Log what background handler received from validation lambda (CALL COUNTS DEBUGGING)
                validation_calls_received = {}
                qc_calls_received = 0
                if enhanced_metrics and enhanced_metrics.get('aggregated_metrics'):
                    providers = enhanced_metrics['aggregated_metrics'].get('providers', {})
                    for provider, data in providers.items():
                        if not data.get('is_metadata_only', False):  # Exclude metadata-only providers
                            validation_calls_received[provider] = data.get('calls', 0)
                        else:
                            logger.debug(f"[BACKGROUND_HANDLER_RECEIVED] Excluding metadata-only provider '{provider}' with {data.get('calls', 0)} calls")

                if qc_metrics_data:
                    qc_calls_received = qc_metrics_data.get('total_qc_calls', 0)

                logger.debug(f"[BACKGROUND_HANDLER_RECEIVED] Received call counts from validation lambda:")
                logger.debug(f"[BACKGROUND_HANDLER_RECEIVED]   Validation calls by provider: {validation_calls_received}")
                logger.debug(f"[BACKGROUND_HANDLER_RECEIVED]   QC calls total: {qc_calls_received}")
                logger.debug(f"[BACKGROUND_HANDLER_RECEIVED]   Grand total calls: {sum(validation_calls_received.values()) + qc_calls_received}")

                # Extract enhanced models parameter from validation response (it's inside enhanced_metrics)
                logger.debug(f"[MODELS_DEBUG] metadata keys available: {list(metadata.keys()) if metadata else 'No metadata'}")
                if enhanced_metrics:
                    logger.debug(f"[MODELS_DEBUG] enhanced_metrics keys available: {list(enhanced_metrics.keys())}")
                    enhanced_models_parameter = enhanced_metrics.get('enhanced_models_parameter', {})
                else:
                    enhanced_models_parameter = {}
                logger.debug(f"[MODELS_DEBUG] enhanced_models_parameter from enhanced_metrics: {list(enhanced_models_parameter.keys()) if enhanced_models_parameter else 'EMPTY'}")
                if enhanced_models_parameter:
                    logger.debug(f"[MODELS_DEBUG] enhanced_models_parameter has {len(enhanced_models_parameter)} search groups")

                # DEBUG: Log the exact structure of the received enhanced_metrics
                logger.debug(f"[DATA_FLOW_DEBUG] Received enhanced_metrics structure: {json.dumps(enhanced_metrics, indent=2)}")
                
                # DEBUG: Log what we received from validation lambda
                logger.debug(f"[ENHANCED_DEBUG] enhanced_metrics available: {bool(enhanced_metrics)}")
                if enhanced_metrics:
                    logger.debug(f"[ENHANCED_DEBUG] enhanced_metrics keys: {list(enhanced_metrics.keys())}")
                    logger.debug(f"[ENHANCED_DEBUG] aggregated_metrics available: {bool(enhanced_metrics.get('aggregated_metrics'))}")
                    if enhanced_metrics.get('aggregated_metrics'):
                        aggregated_data = enhanced_metrics['aggregated_metrics']
                        logger.debug(f"[ENHANCED_DEBUG] aggregated_data keys: {list(aggregated_data.keys())}")
                        totals = aggregated_data.get('totals', {})
                        logger.debug(f"[ENHANCED_DEBUG] totals keys: {list(totals.keys())}")
                        logger.debug(f"[ENHANCED_DEBUG] total_cost_actual: {totals.get('total_cost_actual', 'NOT_FOUND')}")
                        logger.debug(f"[ENHANCED_DEBUG] total_cost_estimated: {totals.get('total_cost_estimated', 'NOT_FOUND')}")
                        providers = aggregated_data.get('providers', {})
                        logger.debug(f"[ENHANCED_DEBUG] providers: {list(providers.keys())}")
                else:
                    logger.warning(f"[ENHANCED_DEBUG] No enhanced_metrics found in metadata. Metadata keys: {list(metadata.keys())}")
                
                if enhanced_metrics and enhanced_metrics.get('aggregated_metrics'):
                    # Extract three-tier cost data from enhanced aggregation
                    aggregated_data = enhanced_metrics['aggregated_metrics']
                    totals = aggregated_data.get('totals', {})
                    
                    # Tier 1: Eliyahu Cost (actual cost paid with caching benefits)
                    eliyahu_cost = totals.get('total_cost_actual', 0.0)

                    # Tier 2: Estimated cost without cache (for full validation projections)
                    cost_estimated = totals.get('total_cost_estimated', 0.0)

                    # Extract provider-specific measures
                    providers = aggregated_data.get('providers', {})
                    perplexity_eliyahu_cost = providers.get('perplexity', {}).get('cost_actual', 0.0)
                    anthropic_eliyahu_cost = providers.get('anthropic', {}).get('cost_actual', 0.0)
                    
                    # Use provider sum as fallback if total is wrong (ensures consistency)
                    provider_cost_sum = perplexity_eliyahu_cost + anthropic_eliyahu_cost
                    logger.debug(f"[COST_DEBUG] eliyahu_cost=${eliyahu_cost:.6f}, provider_sum=${provider_cost_sum:.6f}, perplexity=${perplexity_eliyahu_cost:.6f}, anthropic=${anthropic_eliyahu_cost:.6f}")
                    
                    # Use provider sum if it's different from the total (handles both 0 and other mismatches)
                    if provider_cost_sum > 0 and abs(eliyahu_cost - provider_cost_sum) > 0.000001:
                        logger.debug(f"[COST_CORRECTION] Using provider sum ${provider_cost_sum:.6f} instead of total ${eliyahu_cost:.6f}")
                        eliyahu_cost = provider_cost_sum
                    
                    # Extract API call counts
                    perplexity_calls = providers.get('perplexity', {}).get('calls', 0)
                    anthropic_calls = providers.get('anthropic', {}).get('calls', 0)
                    
                    # Extract token counts
                    perplexity_tokens = providers.get('perplexity', {}).get('tokens', 0)
                    anthropic_tokens = providers.get('anthropic', {}).get('tokens', 0)
                    
                    # Extract both actual and estimated time data
                    # Actual time (with caching benefits) for reporting what actually happened
                    total_processing_time = totals.get('total_actual_processing_time', 0.0)
                    # Estimated time (without caching) for projecting onto non-cached rows
                    # Use direct calculation over all rows if available, otherwise divide total by rows
                    if 'avg_estimated_row_processing_time' in totals:
                        estimated_time_per_row = totals.get('avg_estimated_row_processing_time')
                        time_calculation_method = "direct calculation over all rows"
                    else:
                        estimated_time_per_row = totals.get('total_estimated_processing_time', 0.0) / max(1, total_rows_processed)
                        time_calculation_method = "total time divided by rows (fallback)"
                    total_estimated_time_seconds = totals.get('total_estimated_processing_time', 0.0)
                    
                    logger.debug(f"[ENHANCED_COSTS] Using ai_client aggregated data - Actual: ${eliyahu_cost:.6f}, No cache: ${cost_estimated:.6f}")
                    logger.debug(f"[ENHANCED_PROVIDERS] Perplexity: ${perplexity_eliyahu_cost:.6f} ({perplexity_calls} calls), Anthropic: ${anthropic_eliyahu_cost:.6f} ({anthropic_calls} calls)")
                    logger.debug(f"[ENHANCED_TIME] Actual time: {total_processing_time:.3f}s, Estimated time per row: {estimated_time_per_row:.3f}s ({time_calculation_method}), Total estimated: {total_estimated_time_seconds:.3f}s")
                    
                    # Debug timing extraction from totals
                    logger.debug(f"[TIME_DEBUG] totals timing keys: {[k for k in totals.keys() if 'time' in k.lower()]}")
                    logger.debug(f"[TIME_DEBUG] total_actual_processing_time from totals: {totals.get('total_actual_processing_time', 'NOT_FOUND')}")
                    logger.debug(f"[TIME_DEBUG] total_estimated_processing_time from totals: {totals.get('total_estimated_processing_time', 'NOT_FOUND')}")
                else:
                    # Fallback to legacy token_usage for backward compatibility
                    eliyahu_cost = token_usage.get('total_cost', 0.0)
                    cost_estimated = _calculate_cost_estimated(token_usage, metadata)
                    
                    # Extract legacy provider-specific data
                    by_provider = token_usage.get('by_provider', {})
                    perplexity_data = by_provider.get('perplexity', {})
                    anthropic_data = by_provider.get('anthropic', {})
                    
                    perplexity_eliyahu_cost = perplexity_data.get('total_cost', 0.0)
                    anthropic_eliyahu_cost = anthropic_data.get('total_cost', 0.0)
                    perplexity_calls = perplexity_data.get('api_calls', 0)
                    anthropic_calls = anthropic_data.get('api_calls', 0) 
                    perplexity_tokens = perplexity_data.get('total_tokens', 0)
                    anthropic_tokens = anthropic_data.get('total_tokens', 0)
                    
                    # Legacy time calculation - assume processing_time is actual time with cache benefits
                    total_processing_time = metadata.get('processing_time', 0.0)  # Actual time
                    total_estimated_time_seconds = total_processing_time * 1.2  # Rough estimate: 20% slower without cache
                    estimated_time_per_row = total_estimated_time_seconds / max(1, total_rows_processed)
                    
                    logger.warning(f"[ENHANCED_COSTS] Falling back to legacy token_usage - enhanced_metrics not available")
                
                # ========== ENHANCED VALIDATION ESTIMATES ==========
                # For both preview and validation operations, use enhanced estimates from ai_api_client
                logger.debug(f"[ESTIMATES_CHECK] is_preview: {is_preview}, enhanced_metrics available: {bool(enhanced_metrics)}")
                if enhanced_metrics:
                    logger.debug(f"[ESTIMATES_CHECK] enhanced_metrics keys: {list(enhanced_metrics.keys())}")
                    logger.debug(f"[ESTIMATES_CHECK] full_validation_estimates available: {bool(enhanced_metrics.get('full_validation_estimates'))}")
                
                if enhanced_metrics and enhanced_metrics.get('full_validation_estimates'):
                    estimates = enhanced_metrics['full_validation_estimates']
                    total_estimates = estimates.get('total_estimates', {})

                    # DEBUG: Log the entire estimates object to see what we're working with
                    logger.debug(f"[ESTIMATES_DEBUG] Full validation estimates object: {json.dumps(estimates, indent=2)}")
                    
                    # Extract batch timing analysis for proper time scaling
                    batch_timing = estimates.get('batch_timing_analysis', {})
                    timing_estimates = estimates.get('timing_estimates', {})
                    per_provider_estimates = estimates.get('per_provider_estimates', {})
                    
                    # Use properly scaled estimates from validation lambda calculations
                    # estimated_validation_eliyahu_cost: Raw eliyahu cost for full table (no multiplier, no cache)
                    # This comes from the validation lambda's scaling: preview_cost * (total_rows / preview_rows)
                    estimated_total_cost_raw = total_estimates.get('estimated_total_cost_estimated', cost_estimated * total_rows / max(1, total_rows_processed))
                    
                    # estimated_validation_time: Use batch-based scaling for accurate time projection
                    # The validation lambda calculates this using proper batch architecture
                    logger.debug(f"[TIME_DEBUG] timing_estimates: {timing_estimates}")
                    logger.debug(f"[TIME_DEBUG] batch_timing.estimated_total_time_for_full_validation: {batch_timing.get('estimated_total_time_for_full_validation')}")
                    logger.debug(f"[TIME_DEBUG] total_estimates.estimated_total_processing_time: {total_estimates.get('estimated_total_processing_time')}")
                    estimated_total_time_seconds = (timing_estimates.get('total_estimated_time_seconds') or
                                                   batch_timing.get('estimated_total_time_for_full_validation') or
                                                   total_estimates.get('estimated_total_processing_time', 0.0))

                    logger.debug(f"[TIME_DEBUG] estimated_total_time_seconds: {estimated_total_time_seconds}")
                    logger.debug(f"[TIME_DEBUG] Converting to minutes: {estimated_total_time_seconds / 60:.1f} minutes")
                    logger.debug(f"[TIME_DEBUG] timing_estimates.total_estimated_time_seconds: {timing_estimates.get('total_estimated_time_seconds', 'NOT_SET')}")
                    logger.debug(f"[TIME_DEBUG] batch_timing.estimated_total_time_for_full_validation: {batch_timing.get('estimated_total_time_for_full_validation', 'NOT_SET')}")
                    logger.debug(f"[TIME_DEBUG] total_estimates.estimated_total_processing_time: {total_estimates.get('estimated_total_processing_time', 'NOT_SET')}")
                    
                    # Extract per-provider costs from new structure if available
                    if per_provider_estimates:
                        logger.debug(f"[ENHANCED_PROVIDER_DATA] Found per-provider estimates: {list(per_provider_estimates.keys())}")
                        
                        # Override provider costs with enhanced data
                        perplexity_data = per_provider_estimates.get('perplexity', {})
                        anthropic_data = per_provider_estimates.get('anthropic', {})
                        
                        if perplexity_data:
                            perplexity_per_row_estimated_cost = perplexity_data.get('per_row_estimated_cost', 0)
                            perplexity_total_actual_cost = perplexity_data.get('total_cost_actual', 0)
                        
                        if anthropic_data:
                            anthropic_per_row_estimated_cost = anthropic_data.get('per_row_estimated_cost', 0)
                            anthropic_total_actual_cost = anthropic_data.get('total_cost_actual', 0)
                    
                    # Debug: Log what we actually received
                    logger.debug(f"[ENHANCED_ESTIMATES_DEBUG] batch_timing keys: {list(batch_timing.keys())}")
                    logger.debug(f"[ENHANCED_ESTIMATES_DEBUG] total_estimates keys: {list(total_estimates.keys())}")
                    logger.debug(f"[ENHANCED_ESTIMATES_DEBUG] estimated_total_time_for_full_validation: {batch_timing.get('estimated_total_time_for_full_validation', 'NOT_FOUND')}")
                    logger.debug(f"[ENHANCED_ESTIMATES_DEBUG] estimated_total_cost_estimated: {total_estimates.get('estimated_total_cost_estimated', 'NOT_FOUND')}")
                    
                    logger.debug(f"[ENHANCED_ESTIMATES] ✅ Using ai_client full validation estimates:")
                    logger.info(f"  - Eliyahu cost (no cache, scaled): ${estimated_total_cost_raw:.6f}")
                    logger.info(f"  - Validation time (batch-based): {estimated_total_time_seconds:.3f}s ({estimated_total_time_seconds/60:.1f} minutes)")
                    logger.info(f"  - Batch analysis: {batch_timing.get('estimated_batches_for_full_table', 0)} estimated batches")
                    logger.info(f"  - Preview batch size: {batch_timing.get('preview_average_batch_size', 'N/A')}")
                    logger.info(f"  - Target batch size: {batch_timing.get('target_full_validation_batch_size', 'N/A')}")
                    logger.info(f"  - Avg batch time: {batch_timing.get('estimated_time_per_batch', 0):.3f}s")
                else:
                    # Fallback to manual scaling calculations for non-preview or legacy data
                    estimated_total_cost_raw = None  # Will be calculated later in manual scaling section
                    estimated_total_time_seconds = None  # Will be calculated later in manual scaling section
                    logger.debug(f"[ENHANCED_ESTIMATES] ❌ No enhanced estimates available - will use manual scaling")

                # Validate cost calculations
                if eliyahu_cost < 0 or cost_estimated < 0:
                    logger.error(f"[COST_ERROR] Invalid negative costs - Actual: ${eliyahu_cost:.6f}, Estimated: ${cost_estimated:.6f}")
                    eliyahu_cost = max(0.0, eliyahu_cost)
                    cost_estimated = max(0.0, cost_estimated)
                
                if eliyahu_cost > cost_estimated:
                    logger.warning(f"[COST_WARNING] Actual cost ${eliyahu_cost:.6f} > estimated ${cost_estimated:.6f} - possible pricing issue")
                
                logger.debug(f"[COST_DEBUG] Three-tier costs - Actual: ${eliyahu_cost:.6f}, Estimated: ${cost_estimated:.6f}")
                total_tokens = token_usage.get('total_tokens', 0)
                total_api_calls = token_usage.get('api_calls', 0)
                total_cached_calls = token_usage.get('cached_calls', 0)
                logger.debug(f"[METADATA_DEBUG] metadata keys: {list(metadata.keys()) if metadata else 'None'}")
                logger.debug(f"[METADATA_DEBUG] Looking for validation_metrics in metadata")
                
                # Initialize balance variables (preview doesn't charge)
                initial_balance = 0
                final_balance = 0
                multiplier = 1.0
                charged_cost = 0.0  # Preview is free
                
                # ========== HARDENED DOMAIN MULTIPLIER SYSTEM ==========
                # Apply domain multiplier with comprehensive validation and audit trail
                try:
                    import sys
                    import os
                    sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', 'shared'))
                    from dynamodb_schemas import track_preview_cost, track_api_usage_detailed, check_user_balance
                    from decimal import Decimal
                    
                    # Use scaled eliyahu cost for multiplier calculation if available, otherwise preview cost
                    cost_for_multiplier = estimated_total_cost_raw if estimated_total_cost_raw is not None else cost_estimated
                    
                    logger.debug(f"[MULTIPLIER_DEBUG] Calling _apply_domain_multiplier_with_validation with cost: ${cost_for_multiplier:.6f}")
                    logger.debug(f"[MULTIPLIER_DEBUG] - Using {'scaled eliyahu cost' if estimated_total_cost_raw is not None else 'preview cost'}")
                    multiplier_result = _apply_domain_multiplier_with_validation(email, cost_for_multiplier, session_id)
                    multiplier = multiplier_result['multiplier']
                    domain = multiplier_result['domain']
                    
                    # Cost with multiplier and business logic applied (what user will pay)
                    cost_with_multiplier = multiplier_result['cost_with_multiplier']
                    quoted_full_cost = multiplier_result['quoted_cost']

                    logger.debug(f"[MULTIPLIER_DEBUG] Domain multiplier result: domain={domain}, multiplier={multiplier}, "
                              f"cost_with_multiplier=${cost_with_multiplier:.6f}, quoted_cost=${quoted_full_cost:.2f}")

                    # ========== DISCOUNT CALCULATION ==========
                    # Calculate discount based on session type and config version
                    config_version = config_data.get('storage_metadata', {}).get('version', 1) if config_data else 1
                    discount = calculate_discount(session_id, config_version, quoted_full_cost)
                    effective_cost = max(0.0, quoted_full_cost - discount)

                    logger.info(f"[DISCOUNT] Discount calculation:")
                    logger.info(f"[DISCOUNT]   quoted_cost=${quoted_full_cost:.2f}, discount=${discount:.2f}, effective_cost=${effective_cost:.2f}")
                    logger.info(f"[DISCOUNT]   session_id={session_id}, config_version={config_version}")
                    logger.info(f"[DISCOUNT]   current_balance=${current_balance:.2f}, sufficient={float(current_balance) >= effective_cost}, credits_needed=${max(0, effective_cost - float(current_balance)):.2f}")
                    
                    # Validation and audit logging
                    if 'error' in multiplier_result:
                        logger.error(f"[MULTIPLIER_ERROR] Preview domain multiplier error: {multiplier_result['error']}")
                    
                    logger.debug(f"[COST_AUDIT] Preview processed {total_rows_processed} rows - "
                               f"Preview actual: ${eliyahu_cost:.6f}, Preview estimated: ${cost_estimated:.6f}, "
                               f"Scaled eliyahu cost: ${cost_for_multiplier:.6f}, "
                               f"Domain: {domain}, Multiplier: {multiplier}x, Quoted: ${quoted_full_cost:.2f}")
                    
                    # Get account balance for tracking (no charges for preview)
                    initial_balance = check_user_balance(email)
                    final_balance = initial_balance  # No change for preview
                    balance_error_occurred = False  # Preview doesn't charge
                    charged_amount = 0  # No charges for preview
                    
                    # Track preview cost (actual cost for your tracking, estimated cost for user display)
                    track_preview_cost(session_id, email, Decimal(str(cost_estimated)), Decimal(str(multiplier)), total_tokens)
                    
                    # Provider tracking now handled by enhanced metrics from ai_client
                    logger.debug(f"[ENHANCED_METRICS] Preview provider metrics tracked in enhanced data structure")
                            
                    
                except Exception as e:
                    logger.error(f"Error applying domain multiplier in preview: {e}")
                    multiplier = 1.0
                    # Fallback if multiplier lookup fails - apply default multiplier properly
                    fallback_cost_with_multiplier = cost_estimated * multiplier
                    quoted_full_cost = max(2.0, math.ceil(fallback_cost_with_multiplier))
                    logger.warning(f"[MULTIPLIER_FALLBACK] Using fallback multiplier {multiplier} -> quoted_cost: ${quoted_full_cost:.2f}")
                
                # ========== SIMPLIFIED TIME/BATCH CALCULATIONS ==========
                # Extract simple batch/timing info for backward compatibility
                # All cost calculations now come from enhanced ai_client data
                
                # Get basic processing time for backward compatibility
                processing_time = total_estimated_processing_time if total_estimated_processing_time > 0 else metadata.get('processing_time', 0.0)
                
                # Simple batch size extraction for display purposes
                effective_batch_size = batch_size or 3  # Small batch for testing continuations
                per_model_batch_stats = metadata.get('per_model_batch_stats', {})
                if per_model_batch_stats:
                    actual_batch_sizes = per_model_batch_stats.get('model_batch_sizes', {})
                    if actual_batch_sizes:
                        effective_batch_size = min(actual_batch_sizes.values())
                
                # Simple fallback time estimates if enhanced data not available
                if estimated_total_time_seconds is None:
                    # Use simple per-row calculation
                    time_per_row_fallback = processing_time / max(1, total_rows_processed)
                    estimated_total_time_seconds = time_per_row_fallback * total_rows
                
                # Calculate time per row for full validation (time_per_row_seconds field)
                # First try to get from enhanced timing estimates, then fall back to calculation
                logger.debug(f"[TIME_DEBUG_FINAL] Final estimated_total_time_seconds before payload: {estimated_total_time_seconds}")
                if estimated_total_time_seconds:
                    logger.debug(f"[TIME_DEBUG_FINAL] Will be converted to minutes: {estimated_total_time_seconds / 60:.1f}")
                else:
                    logger.debug(f"[TIME_DEBUG_FINAL] Will be converted to minutes: None")
                if enhanced_metrics and enhanced_metrics.get('full_validation_estimates'):
                    timing_estimates = enhanced_metrics['full_validation_estimates'].get('timing_estimates', {})
                    time_per_row = timing_estimates.get('time_per_row_seconds', 0.0)
                    if time_per_row == 0.0:
                        time_per_row = estimated_total_time_seconds / max(1, total_rows) if estimated_total_time_seconds and total_rows else 0.0
                else:
                    time_per_row = estimated_total_time_seconds / max(1, total_rows) if estimated_total_time_seconds and total_rows else 0.0
                
                # Simple batch count for display
                total_batches = math.ceil(total_rows / effective_batch_size) if effective_batch_size > 0 else 1
                
                logger.debug(f"[SIMPLIFIED_TIMING] Processing time: {processing_time:.3f}s, Est. total: {estimated_total_time_seconds:.3f}s")
                logger.debug(f"[SIMPLIFIED_BATCH] Batch size: {effective_batch_size}, Total batches: {total_batches}")
                
                # All cost/time calculations now come from enhanced estimates - no manual scaling
                if estimated_total_cost_raw is None:
                    logger.warning(f"[ENHANCED_DATA] Missing estimated_total_cost_raw from enhanced estimates - using fallback")
                    scaling_factor = total_rows / max(1, total_rows_processed)
                    estimated_total_cost_raw = cost_estimated * scaling_factor
                    logger.debug(f"[SCALING_DEBUG] Manual scaling: ${cost_estimated:.6f} × ({total_rows}/{total_rows_processed}) = ${estimated_total_cost_raw:.6f}")
                
                # Quoted cost comes from enhanced estimates with business logic applied
                if quoted_full_cost is None:
                    logger.warning(f"[ENHANCED_DATA] Missing quoted_full_cost from enhanced estimates - applying business logic")
                    raw_quoted_cost = cost_estimated * multiplier * (total_rows / max(1, total_rows_processed))
                    quoted_full_cost = max(2.0, math.ceil(raw_quoted_cost))  # Add $2 minimum charge, rounded up
                    logger.debug(f"[ENHANCED_DATA] Calculated fallback quoted_full_cost: ${quoted_full_cost:.2f} (base: ${cost_estimated:.6f}, multiplier: {multiplier}, scaling: {total_rows}/{total_rows_processed})")
                
                estimated_total_tokens = total_tokens * (total_rows / max(1, total_rows_processed))
                
                # DEBUG: Log enhanced cost data
                logger.debug(f"[ENHANCED_COST] Final estimates:")
                logger.debug(f"[ENHANCED_COST]   eliyahu_cost: ${eliyahu_cost:.6f}")
                logger.debug(f"[ENHANCED_COST]   estimated_total_cost_raw: ${estimated_total_cost_raw:.6f}")
                logger.debug(f"[ENHANCED_COST]   quoted_full_cost (with business logic): ${quoted_full_cost:.2f}")
                logger.debug(f"[ENHANCED_COST]   estimated_total_time_seconds: {estimated_total_time_seconds:.3f}s")

                # Get the most recent account balance right before calculating sufficient_balance
                # This ensures we capture any recent credit additions from the frontend
                logger.debug(f"[BALANCE_CHECK] Refreshing balance for {email} before sufficient_balance calculation")
                current_balance = check_user_balance(email)
                logger.debug(f"[BALANCE_CHECK] Current balance: {current_balance}, Quoted full cost: {quoted_full_cost}")

                # Create the markdown table for the response
                markdown_table = create_markdown_table_from_results(real_results, 3, actual_config_s3_key, S3_UNIFIED_BUCKET, qc_results)
                
                preview_payload = {
                    "status": "COMPLETED", "session_id": session_id,
                    "markdown_table": markdown_table, "total_rows": total_rows,
                    "total_processed_rows": total_rows_processed,
                    "preview_processing_time": processing_time,
                    "estimated_total_processing_time": estimated_total_time_seconds,
                    "estimated_validation_time_minutes": round(estimated_total_time_seconds / 60, 1),
                    "actual_batch_size": effective_batch_size,
                    "estimated_validation_batches": total_batches,
                    "cost_estimates": {
                        "preview_cost": charged_cost,  # What user pays for preview (0)
                        "preview_tokens": total_tokens,
                        "estimated_total_tokens": estimated_total_tokens,
                        "per_row_time": time_per_row
                    },
                    "token_usage": token_usage,
                    "validation_metrics": validation_metrics,
                    "account_info": {
                        "current_balance": float(current_balance) if current_balance is not None else 0,
                        "sufficient_balance": float(current_balance if current_balance is not None else 0) >= effective_cost,
                        "credits_needed": max(0, effective_cost - (float(current_balance) if current_balance is not None else 0)),
                        "domain_multiplier": float(multiplier),
                        "email_domain": email_domain
                    }
                }
                
                # Extract provider costs and totals: per-row estimated costs and total run actual costs
                perplexity_per_row_estimated_cost = 0
                anthropic_per_row_estimated_cost = 0
                perplexity_total_actual_cost = 0
                anthropic_total_actual_cost = 0
                totals = {}  # Initialize totals for use later
                
                if enhanced_metrics and enhanced_metrics.get('aggregated_metrics'):
                    providers = enhanced_metrics['aggregated_metrics'].get('providers', {})
                    totals = enhanced_metrics['aggregated_metrics'].get('totals', {})  # Extract totals here
                    
                    # Extract per-row ESTIMATED costs (what it would cost per row without caching)
                    if total_rows_processed > 0:
                        perplexity_per_row_estimated_cost = providers.get('perplexity', {}).get('cost_estimated', 0) / total_rows_processed
                        anthropic_per_row_estimated_cost = providers.get('anthropic', {}).get('cost_estimated', 0) / total_rows_processed
                    
                    # Extract total run ACTUAL costs (what was actually paid for the entire run)
                    perplexity_total_actual_cost = providers.get('perplexity', {}).get('cost_actual', 0)
                    anthropic_total_actual_cost = providers.get('anthropic', {}).get('cost_actual', 0)
                else:
                    # Fallback to legacy token_usage if enhanced data not available
                    by_provider = token_usage.get('by_provider', {})
                    perplexity_data = by_provider.get('perplexity', {})
                    anthropic_data = by_provider.get('anthropic', {})
                    
                    # Legacy calculation
                    perplexity_total_actual_cost = float(perplexity_data.get('total_cost', 0))
                    anthropic_total_actual_cost = float(anthropic_data.get('total_cost', 0))
                    perplexity_per_row_estimated_cost = perplexity_total_actual_cost / total_rows_processed if total_rows_processed > 0 else 0
                    anthropic_per_row_estimated_cost = anthropic_total_actual_cost / total_rows_processed if total_rows_processed > 0 else 0
                    
                    # Create fallback totals for legacy path
                    totals = {
                        'total_cost_estimated': perplexity_total_actual_cost + anthropic_total_actual_cost,
                        'total_calls': perplexity_data.get('calls', 0) + anthropic_data.get('calls', 0)
                    }
                
                # Add provider costs to cost_estimates
                preview_payload['cost_estimates']['perplexity_per_row_estimated_cost'] = perplexity_per_row_estimated_cost  # Per-row cost without caching
                preview_payload['cost_estimates']['anthropic_per_row_estimated_cost'] = anthropic_per_row_estimated_cost  # Per-row cost without caching
                preview_payload['cost_estimates']['perplexity_total_actual_cost'] = perplexity_total_actual_cost  # Total actual cost paid for entire run
                preview_payload['cost_estimates']['anthropic_total_actual_cost'] = anthropic_total_actual_cost  # Total actual cost paid for entire run
                
                # Add enhanced timing data if available
                if enhanced_metrics and enhanced_metrics.get('full_validation_estimates'):
                    timing_estimates = enhanced_metrics['full_validation_estimates'].get('timing_estimates', {})
                    
                    # Add timing fields to cost_estimates for frontend
                    preview_payload['cost_estimates']['actual_processing_time_seconds'] = timing_estimates.get('actual_processing_time_seconds', 0.0)
                    preview_payload['cost_estimates']['actual_time_per_batch_seconds'] = timing_estimates.get('actual_time_per_batch_seconds', 0.0)
                    
                    logger.debug(f"[ENHANCED_TIMING] Added timing data - processing: {timing_estimates.get('actual_processing_time_seconds', 0):.3f}s, "
                               f"per batch: {timing_estimates.get('actual_time_per_batch_seconds', 0):.3f}s")
                
                # Add key frontend-expected fields to cost_estimates object (frontend looks for them here)
                preview_payload['cost_estimates'].update({
                    "quoted_validation_cost": quoted_full_cost,  # What user will pay for full validation (rounded up)
                    "estimated_validation_eliyahu_cost": estimated_total_cost_raw,  # Raw eliyahu cost estimate for full table (no multiplier)
                    "discount": discount,  # Discount amount applied
                    "effective_cost": effective_cost,  # Cost after discount (what user actually pays)
                    "estimated_total_processing_time": estimated_total_time_seconds,  # Keep for backward compatibility
                    "estimated_validation_time": estimated_total_time_seconds, # Explicitly pass the full validation time estimate
                    "total_provider_cost_estimated": totals.get('total_cost_estimated', 0.0),  # Total estimated cost across all providers
                    "total_provider_calls": totals.get('total_calls', 0) + qc_metrics_data.get('total_qc_calls', 0) if qc_metrics_data else totals.get('total_calls', 0)  # Total calls including QC
                })

                # Add fields at top level for backward compatibility and easy access
                preview_payload.update({
                    "quoted_validation_cost": quoted_full_cost,  # What user will pay for full validation (rounded up)
                    "estimated_validation_eliyahu_cost": estimated_total_cost_raw,  # Raw eliyahu cost estimate for full table (no multiplier)
                    "discount": discount,  # Discount amount
                    "effective_cost": effective_cost  # Cost after discount
                })
                
                # Generate enhanced Excel download link and add to preview payload
                enhanced_download_url = None
                logger.info("Generating enhanced Excel for preview mode")
                
                try:
                    # Get excel content and config for enhanced Excel generation
                    from ..core.unified_s3_manager import UnifiedS3Manager
                    storage_manager = UnifiedS3Manager()
                    excel_content, excel_s3_key = storage_manager.get_excel_file(email, clean_session_id)
                    config_data, config_s3_key = storage_manager.get_latest_config(email, clean_session_id)
                    
                    if excel_content and config_data:
                        input_filename = excel_s3_key.split('/')[-1]
                        
                        # Calculate summary data for potential email
                        all_fields = set()
                        confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
                        original_confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
                        
                        for row_data in real_results.values():
                            for field_name, field_data in row_data.items():
                                if isinstance(field_data, dict):
                                    all_fields.add(field_name)
                                    
                                    # Count updated confidence levels
                                    if 'confidence_level' in field_data:
                                        conf_level = field_data.get('confidence_level', 'UNKNOWN')
                                        if conf_level in confidence_counts:
                                            confidence_counts[conf_level] += 1
                                    
                                    # Count original confidence levels
                                    if 'original_confidence' in field_data:
                                        original_conf = field_data.get('original_confidence')
                                        if original_conf and str(original_conf).upper() in original_confidence_counts:
                                            original_confidence_counts[str(original_conf).upper()] += 1
                        
                        # Create enhanced Excel if available
                        enhanced_excel_content = None
                        
                        if EXCEL_ENHANCEMENT_AVAILABLE:
                            try:
                                # Reuse table_data that was parsed earlier with row keys already added
                                validated_sheet = table_data.get('metadata', {}).get('sheet_name') if isinstance(table_data, dict) else None

                                logger.debug(f"[PREVIEW_EXCEL] Reusing table_data with {len(table_data.get('data', []))} rows (row keys already added)")

                                excel_buffer = create_qc_enhanced_excel_for_interface(
                                    table_data, validation_results, config_data, session_id, validated_sheet_name=validated_sheet
                                )
                                
                                if excel_buffer:
                                    enhanced_excel_content = excel_buffer.getvalue()
                                else:
                                    logger.error("Enhanced Excel creation failed - excel_buffer is None")
                            except Exception as e:
                                logger.error(f"Error creating enhanced Excel for preview: {str(e)}")
                                import traceback
                                logger.error(traceback.format_exc())
                        else:
                            # Fallback: Create basic Excel using openpyxl when xlsxwriter is not available
                            try:
                                logger.debug("[DEBUG] xlsxwriter not available, creating fallback Excel using openpyxl")
                                enhanced_excel_content = _create_fallback_preview_excel(real_results, config_data, input_filename)
                                logger.debug(f"[DEBUG] Fallback Excel created. Size: {len(enhanced_excel_content) if enhanced_excel_content else 0} bytes")
                            except Exception as e:
                                logger.error(f"Error creating fallback Excel for preview: {str(e)}")
                                import traceback
                                logger.error(traceback.format_exc())
                        
                        # Store enhanced Excel in versioned results folder and create download link
                        logger.debug(f"[DEBUG] About to store enhanced Excel. enhanced_excel_content size: {len(enhanced_excel_content) if enhanced_excel_content else 0}")
                        if enhanced_excel_content:
                            try:
                                # Get version from config
                                config_version = config_data.get('storage_metadata', {}).get('version', 1)
                                enhanced_filename = f"{os.path.splitext(input_filename)[0]}_v{config_version}_preview_enhanced.xlsx"
                                logger.debug(f"[DEBUG] Storing enhanced Excel with filename: {enhanced_filename}")
                                
                                # Store enhanced Excel in versioned results folder
                                enhanced_result = storage_manager.store_enhanced_files(
                                    email, clean_session_id, config_version,
                                    enhanced_excel_content, None,
                                    result_type='preview'
                                )
                                
                                if enhanced_result['success']:
                                    logger.debug(f"[DEBUG] Enhanced Excel stored in results folder: {enhanced_result['stored_files']}")
                                    
                                    # Also create public download link for immediate download
                                    enhanced_download_url = storage_manager.create_public_download_link(
                                        enhanced_excel_content, 
                                        enhanced_filename,
                                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                    )
                                    logger.debug(f"[DEBUG] Enhanced Excel download link created successfully: {enhanced_download_url}")
                                else:
                                    logger.error(f"[DEBUG] Failed to store enhanced Excel in results folder: {enhanced_result.get('error')}")
                                    enhanced_download_url = None
                                    
                            except Exception as e:
                                logger.error(f"Failed to store enhanced Excel: {e}")
                                import traceback
                                logger.error(traceback.format_exc())
                        else:
                            logger.error("[DEBUG] No enhanced Excel content available - cannot store or create download link")
                except Exception as e:
                    logger.error(f"Error during enhanced Excel generation for preview: {e}")
                    import traceback
                    logger.error(traceback.format_exc())

                # Add enhanced Excel download URL to preview payload
                preview_payload["enhanced_download_url"] = enhanced_download_url
                logger.debug(f"[DEBUG] Added enhanced_download_url to preview_payload: {enhanced_download_url}")
                
                # Get account balance and multiplier for preview tracking
                try:
                    import sys
                    import os
                    sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', 'shared'))
                    from dynamodb_schemas import check_user_balance, get_domain_multiplier
                    
                    current_balance = check_user_balance(email)
                    email_domain = email.split('@')[-1] if '@' in email else 'unknown'
                    multiplier = float(get_domain_multiplier(email_domain))
                except Exception as e:
                    logger.warning(f"Could not get account info for preview: {e}")
                    current_balance = None
                    multiplier = 1.0
                
                # Create search group models tracking string for preview - one entry per search group with validated columns count
                search_groups_models = []
                search_groups_count = validation_metrics.get('search_groups_count', 0)
                enhanced_context_groups = validation_metrics.get('high_context_search_groups_count', 0)
                claude_groups = validation_metrics.get('claude_search_groups_count', 0)
                regular_perplexity_groups = search_groups_count - enhanced_context_groups - claude_groups
                validated_columns_count = validation_metrics.get('validated_columns_count', 0)
                
                # Count actual validation targets per search group from config
                validation_targets = config_data.get('validation_targets', [])
                columns_per_search_group = {}
                
                # Count validation targets assigned to each search group (excluding ID and IGNORED)
                for target in validation_targets:
                    importance = target.get('importance', '').upper()
                    if importance not in ["ID", "IGNORED"]:
                        search_group_id = target.get('search_group', 0)
                        columns_per_search_group[search_group_id] = columns_per_search_group.get(search_group_id, 0) + 1
                
                # Build search group models by iterating through actual search groups in config
                search_groups_list = config_data.get('search_groups', [])
                anthropic_default_web_searches = config_data.get('anthropic_max_web_searches_default', 3)
                
                # Iterate through each search group to get exact models and settings
                for search_group in search_groups_list:
                    search_group_id = search_group.get('group_id', 0)
                    columns_for_this_group = columns_per_search_group.get(search_group_id, 0)

                    # CRITICAL FIX: Skip search groups with no validation targets
                    # This handles:
                    # 1. Non-uniform group numbering (gaps like 1,2,3,5)
                    # 2. Groups with only IGNORED/ID columns (no CRITICAL columns)
                    if columns_for_this_group == 0:
                        logger.debug(f"[SEARCH_GROUP_SKIP] Skipping group {search_group_id} - no validation targets")
                        continue

                    model = search_group.get('model', 'sonar-pro')
                    search_context = search_group.get('search_context', 'low')
                    
                    # Build model display name
                    if 'claude' in model.lower() or 'anthropic' in model.lower():
                        # Get anthropic_max_web_searches from this specific search group, or fall back to default
                        max_web_searches = search_group.get('anthropic_max_web_searches', anthropic_default_web_searches)
                        if search_context == 'high':
                            model_display = f"{model} ({max_web_searches}) (high context) X {columns_for_this_group}"
                        else:
                            model_display = f"{model} ({max_web_searches}) X {columns_for_this_group}"
                    else:
                        # Perplexity models
                        if search_context == 'high':
                            model_display = f"{model} (high context) X {columns_for_this_group}"
                        else:
                            model_display = f"{model} X {columns_for_this_group}"
                    
                    search_groups_models.append(model_display)
                
                # Get configuration ID for tracking
                configuration_id = config_data.get('storage_metadata', {}).get('config_id', 'unknown')
                if not configuration_id or configuration_id == 'unknown':
                    # Fallback to generation metadata
                    configuration_id = config_data.get('generation_metadata', {}).get('config_id', 'unknown')
                
                # Extract consolidated fields from preview payload
                logger.debug(f"[TIME_SAVE_DEBUG] preview_payload keys: {list(preview_payload.keys())}")
                logger.debug(f"[TIME_SAVE_DEBUG] estimated_validation_time_minutes in preview_payload: {preview_payload.get('estimated_validation_time_minutes', 'NOT_FOUND')}")

                eliyahu_cost = preview_payload.get('cost_estimates', {}).get('preview_cost', 0.0)  # Actual cost incurred
                quoted_validation_cost_value = preview_payload.get('cost_estimates', {}).get('quoted_validation_cost', 0.0)  # What user will pay for full
                estimated_validation_eliyahu_cost_value = preview_payload.get('cost_estimates', {}).get('estimated_validation_eliyahu_cost', 0.0)  # Raw eliyahu cost estimate
                time_per_row = preview_payload.get('cost_estimates', {}).get('per_row_time', 0.0)  # Time per row estimate
                estimated_time_minutes = preview_payload.get('estimated_validation_time_minutes', 0.0)  # Total estimated time in minutes

                logger.debug(f"[TIME_SAVE_DEBUG] extracted estimated_time_minutes: {estimated_time_minutes}")
                logger.debug(f"[TIME_SAVE_DEBUG] Conversion check: {estimated_time_minutes} minutes = {estimated_time_minutes * 60:.1f} seconds")
                logger.debug(f"[TIME_SAVE_DEBUG] Source estimated_total_time_seconds was: {preview_payload.get('estimated_total_processing_time', 'NOT_IN_PAYLOAD')}")

                # Debug total_provider_calls calculation
                total_validation_calls = 0
                total_qc_calls = 0
                if enhanced_metrics and enhanced_metrics.get('aggregated_metrics'):
                    providers = enhanced_metrics['aggregated_metrics'].get('providers', {})
                    for provider, data in providers.items():
                        # QC is tracked separately in qc_metrics_data, exclude QC_Costs provider
                        if provider != 'QC_Costs':
                            total_validation_calls += data.get('calls', 0)
                    logger.debug(f"[PROVIDER_CALLS_DEBUG] From enhanced_metrics - validation_calls: {total_validation_calls}")
                else:
                    # Fallback to token_usage
                    for provider in ['perplexity', 'anthropic']:
                        provider_data = token_usage.get('by_provider', {}).get(provider, {})
                        total_validation_calls += provider_data.get('api_calls', 0)
                    logger.debug(f"[PROVIDER_CALLS_DEBUG] From token_usage - validation_calls: {total_validation_calls}")

                # Add QC calls from qc_metrics_data
                if qc_metrics_data:
                    qc_calls_from_metrics = qc_metrics_data.get('total_qc_calls', 0)
                    logger.debug(f"[PROVIDER_CALLS_DEBUG] QC calls from qc_metrics_data: {qc_calls_from_metrics}")
                    total_qc_calls = max(total_qc_calls, qc_calls_from_metrics)  # Use max to avoid double-counting

                # Calculate grand total (QC calls now included in validation_calls via anthropic provider)
                total_provider_calls_override = total_validation_calls
                logger.debug(f"[PROVIDER_CALLS_DEBUG] TOTAL calls (QC included in validation): {total_provider_calls_override}, QC calls (debug): {total_qc_calls}")
                batch_size_used = preview_payload.get('actual_batch_size', 10)  # Actual batch size used for preview
                
                # Use enhanced provider metrics from ai_client aggregation if available
                provider_metrics_for_db = {}
                total_rows_processed = len(validation_results.get('validation_results', {})) if validation_results else 1
                
                if enhanced_metrics and enhanced_metrics.get('aggregated_metrics'):
                    # Use properly aggregated enhanced metrics from validation lambda
                    providers = enhanced_metrics['aggregated_metrics'].get('providers', {})

                    for provider, provider_data in providers.items():
                        # Skip QC_Costs as it's now in separate qc_metrics field
                        if provider == 'QC_Costs':
                            continue
                        provider_metrics_for_db[provider] = {
                            'calls': provider_data.get('calls', 0),
                            'tokens': provider_data.get('tokens', 0),
                            'cost_actual': provider_data.get('cost_actual', 0.0),
                            'cost_estimated': provider_data.get('cost_estimated', 0.0),
                            'processing_time': provider_data.get('time_actual', 0.0),
                            'cache_hit_tokens': provider_data.get('cache_hit_tokens', 0),
                            'cost_per_row_actual': provider_data.get('cost_actual', 0.0) / total_rows_processed if total_rows_processed > 0 else 0,
                            'cost_per_row_estimated': provider_data.get('cost_estimated', 0.0) / total_rows_processed if total_rows_processed > 0 else 0,
                            'time_per_row_actual': provider_data.get('time_actual', 0.0) / total_rows_processed if total_rows_processed > 0 else 0,
                            'time_per_row_estimated': provider_data.get('time_estimated', 0.0) / total_rows_processed if total_rows_processed > 0 else 0,
                            'cache_efficiency_percent': provider_data.get('cache_efficiency_percent', 0.0)
                        }
                    
                    logger.debug(f"[PROVIDER_METRICS] Using enhanced aggregated metrics for provider_metrics_for_db")

                # Create separate QC metrics for database (not mixed with providers)
                qc_metrics_for_db = {}
                if qc_metrics_data:
                    # Calculate per-row QC metrics similar to provider metrics
                    qc_rows_processed = qc_metrics_data.get('total_rows_processed', total_rows_processed)
                    qc_cost_actual = qc_metrics_data.get('total_qc_cost', 0.0)
                    qc_cost_estimated = qc_metrics_data.get('total_qc_cost_estimated', 0.0)
                    qc_time_actual = qc_metrics_data.get('total_qc_time_actual', 0.0)
                    qc_time_estimated = qc_metrics_data.get('total_qc_time_estimated', 0.0)

                    qc_metrics_for_db = {
                        'enabled': True,
                        'total_fields_reviewed': qc_metrics_data.get('total_fields_reviewed', 0),
                        'total_fields_modified': qc_metrics_data.get('total_fields_modified', 0),
                        'total_qc_cost': qc_cost_actual,
                        'total_qc_cost_estimated': qc_cost_estimated,
                        'total_qc_calls': qc_metrics_data.get('total_qc_calls', 0),
                        'total_qc_tokens': qc_metrics_data.get('total_qc_tokens', 0),
                        'qc_models_used': qc_metrics_data.get('qc_models_used', []),
                        'confidence_lowered_count': qc_metrics_data.get('confidence_lowered_count', 0),
                        'values_replaced_count': qc_metrics_data.get('values_replaced_count', 0),
                        # Add per-row metrics
                        'cost_per_row_actual': qc_cost_actual / qc_rows_processed if qc_rows_processed > 0 else 0.0,
                        'cost_per_row_estimated': qc_cost_estimated / qc_rows_processed if qc_rows_processed > 0 else 0.0,
                        'time_actual': qc_time_actual,
                        'time_estimated': qc_time_estimated,
                        'time_per_row_actual': qc_time_actual / qc_rows_processed if qc_rows_processed > 0 else 0.0,
                        'time_per_row_estimated': qc_time_estimated / qc_rows_processed if qc_rows_processed > 0 else 0.0,
                        # Add qc_by_column if available
                        'qc_by_column': qc_metrics_data.get('qc_by_column', {})
                    }
                    logger.debug(f"[QC_METRICS] Created QC metrics for DB with qc_by_column: {len(qc_metrics_for_db.get('qc_by_column', {}))} columns")

                # The models field already contains all validation model info via enhanced_models_parameter
                # QC model info is in qc_metrics_for_db['qc_models_used']
                logger.debug(f"[MODELS_SUMMARY] Validation models in enhanced_models_parameter: {len(enhanced_models_parameter)} search groups")
                if qc_metrics_for_db:
                    logger.debug(f"[QC_MODELS] QC models used: {qc_metrics_for_db.get('qc_models_used', [])}")
                else:
                    # Fallback to legacy token_usage conversion
                    by_provider = token_usage.get('by_provider', {})
                    
                    for provider, provider_usage in by_provider.items():
                        if provider_usage and isinstance(provider_usage, dict):
                            # Extract metrics from existing token_usage structure
                            actual_cost = provider_usage.get('total_cost', 0.0)
                            total_tokens = provider_usage.get('total_tokens', 0)
                            api_calls = provider_usage.get('api_calls', 0)
                            cached_calls = provider_usage.get('cached_calls', 0)
                            
                            # Estimate cost without cache benefits (approximate 8x multiplier for time, varies for cost)
                            cache_efficiency = cached_calls / max(api_calls, 1) if api_calls > 0 else 0
                            cost_estimated = actual_cost * (1 + (cache_efficiency * 1.5))  # Conservative estimate
                            
                            provider_metrics_for_db[provider] = {
                                'calls': api_calls,
                                'tokens': total_tokens,
                                'cost_actual': actual_cost,
                                'cost_estimated': cost_estimated,
                                'processing_time': processing_time * (api_calls / max(sum(p.get('api_calls', 0) for p in by_provider.values()), 1)),
                                'cache_hit_tokens': cached_calls * (total_tokens / max(api_calls, 1)) if api_calls > 0 else 0,
                                'cost_per_row_actual': actual_cost / total_rows_processed if total_rows_processed > 0 else 0,
                                'cost_per_row_estimated': cost_estimated / total_rows_processed if total_rows_processed > 0 else 0,
                                'time_per_row_actual': (processing_time * (api_calls / max(sum(p.get('api_calls', 0) for p in by_provider.values()), 1))) / total_rows_processed if total_rows_processed > 0 else 0,
                                'cache_efficiency_percent': (1 - (actual_cost / max(cost_estimated, 0.000001))) * 100
                            }
                    
                    logger.warning(f"[PROVIDER_METRICS] Falling back to legacy token_usage conversion for provider_metrics_for_db")
                
                # Record completion time for background handler
                background_end_time = datetime.now(timezone.utc).isoformat()
                
                # Calculate actual background handler processing time
                start_time = datetime.fromisoformat(background_start_time.replace('Z', '+00:00'))
                end_time = datetime.fromisoformat(background_end_time.replace('Z', '+00:00'))
                background_processing_time_seconds = (end_time - start_time).total_seconds()
                logger.debug(f"[PREVIEW_TIMING] Background handler processing time: {background_processing_time_seconds:.3f}s")
                
                # Debug QC search groups calculation
                base_claude_groups = validation_metrics.get('claude_search_groups_count', 0)
                qc_has_calls = qc_metrics_data and qc_metrics_data.get('total_qc_calls', 0) > 0
                qc_calls_count = qc_metrics_data.get('total_qc_calls', 0) if qc_metrics_data else 0
                corrected_claude_groups = base_claude_groups + (1 if qc_has_calls else 0)
                logger.debug(f"[QC_SEARCH_GROUPS_DEBUG] base_claude_groups: {base_claude_groups}, qc_has_calls: {qc_has_calls}, qc_calls_count: {qc_calls_count}, corrected: {corrected_claude_groups}")

                # Preserve original validation metrics for database storage BEFORE modifying for frontend
                original_validation_metrics = validation_metrics.copy()

                # Update preview_payload validation_metrics for WebSocket logging
                # Frontend expects: perplexityGroups = searchGroups - claudeGroups, so adjust accordingly
                total_search_groups = validation_metrics.get('search_groups_count', 0)  # 4 total validation groups
                perplexity_only_groups = total_search_groups - base_claude_groups  # 4 - 1 = 3 perplexity groups
                claude_groups_for_frontend = base_claude_groups + (1 if qc_has_calls else 0)  # Include QC as fake Claude group (1 + 1 = 2)
                # Add fake group to total so frontend math works: perplexity = total - claude = 5 - 2 = 3
                total_groups_for_frontend = total_search_groups + (1 if qc_has_calls else 0)  # Add fake QC group to total

                if 'validation_metrics' not in preview_payload:
                    preview_payload['validation_metrics'] = {}
                preview_payload['validation_metrics'].update({
                    "validated_columns_count": validation_metrics.get('validated_columns_count', 0),
                    "search_groups_count": total_groups_for_frontend,  # Total groups so frontend math works
                    "claude_search_groups_count": claude_groups_for_frontend  # Claude + QC groups
                })

                # Create minimal frontend payload with only consumed fields
                frontend_payload = {
                    "markdown_table": preview_payload.get("markdown_table", ""),
                    "enhanced_download_url": preview_payload.get("enhanced_download_url"),
                    "total_rows": preview_payload.get("total_rows", 0),
                    "cost_estimates": {
                        "quoted_validation_cost": quoted_full_cost,
                        "discount": discount,
                        "effective_cost": effective_cost,
                        "estimated_validation_time": estimated_total_time_seconds
                    },
                    "validation_metrics": {
                        "validated_columns_count": validation_metrics.get('validated_columns_count', 0),
                        "search_groups_count": total_groups_for_frontend,  # Total groups so frontend math works
                        "claude_search_groups_count": claude_groups_for_frontend  # Claude + QC groups
                    },
                    "account_info": {
                        "current_balance": float(current_balance) if current_balance else 0,
                        "sufficient_balance": float(current_balance) >= effective_cost if current_balance else False,
                        "credits_needed": max(0, effective_cost - (float(current_balance) if current_balance else 0))
                    }
                }
                
                # Update DynamoDB with the minimal frontend payload
                update_run_status_for_session(status='COMPLETED',
                    run_type="Preview",
                    verbose_status="Preview complete. Results available.",
                    percent_complete=100,
                    processed_rows=len(validation_results.get('validation_results', {})) if validation_results else 0,
                    total_rows=total_rows,  # Actual total rows in the table
                    preview_data=frontend_payload,  # Send minimal frontend payload
                    account_current_balance=float(current_balance) if current_balance else 0,
                    account_sufficient_balance="n/a",
                    account_credits_needed="n/a",
                    account_domain_multiplier=float(multiplier),
                    models=json.dumps(enhanced_models_parameter) if enhanced_models_parameter else json.dumps({}),
                    input_table_name=input_filename,
                    configuration_id=configuration_id,
                    batch_size=batch_size_used,
                    eliyahu_cost=eliyahu_cost,  # Actual cost paid (with caching benefits)
                    quoted_validation_cost=quoted_full_cost,  # What user will pay for full validation (with multiplier, rounding, $2 min)
                    discount=discount,  # Discount applied
                    estimated_validation_eliyahu_cost=estimated_total_cost_raw,  # Raw cost estimate for full table without caching benefit
                    time_per_row_seconds=time_per_row,
                    estimated_validation_time_minutes=estimated_time_minutes,
                    end_time=background_end_time,  # Use background handler completion time
                    run_time_s=background_processing_time_seconds,  # Actual background handler processing time
                    provider_metrics=provider_metrics_for_db,  # Enhanced provider-specific metrics
                    qc_metrics=qc_metrics_for_db if qc_metrics_for_db else None,  # Separate QC metrics (not mixed with providers)
                    total_provider_calls=total_provider_calls_override  # Override with correct total including QC
                )

                # Log what we're saving to database
                logger.debug(f"[DB_SAVE_DEBUG] Saving to database:")
                logger.debug(f"[DB_SAVE_DEBUG]   estimated_validation_time_minutes: {estimated_time_minutes}")
                logger.debug(f"[DB_SAVE_DEBUG]   total_provider_calls: {total_provider_calls_override}")
                logger.debug(f"[DB_SAVE_DEBUG]   qc_metrics: {qc_metrics_for_db}")
                if qc_metrics_for_db and 'qc_by_column' in qc_metrics_for_db:
                    logger.debug(f"[DB_SAVE_DEBUG]   qc_by_column columns: {list(qc_metrics_for_db['qc_by_column'].keys())}")
                    logger.debug(f"[DB_SAVE_DEBUG]   qc_by_column data: {qc_metrics_for_db['qc_by_column']}")
                logger.debug(f"[DB_SAVE_DEBUG]   provider_metrics keys: {list(provider_metrics_for_db.keys()) if provider_metrics_for_db else 'None'}")
                
                # Track enhanced user metrics for preview
                logger.debug(f"[USER_TRACKING] Tracking preview request for email: {email}")
                logger.debug(f"[VALIDATION_METRICS_DEBUG] validation_metrics: {validation_metrics}")
                logger.debug(f"[VALIDATION_METRICS_DEBUG] validated_columns_count: {validation_metrics.get('validated_columns_count', 'KEY_NOT_FOUND')}")
                try:
                    track_result = track_user_request(
                    email=email,
                    request_type='preview',
                    tokens_used=total_tokens,
                    cost_usd=charged_cost,  # Legacy field
                    perplexity_tokens=token_usage.get('by_provider', {}).get('perplexity', {}).get('total_tokens', 0),
                    perplexity_cost=token_usage.get('by_provider', {}).get('perplexity', {}).get('total_cost', 0),
                    anthropic_tokens=token_usage.get('by_provider', {}).get('anthropic', {}).get('total_tokens', 0),
                    anthropic_cost=token_usage.get('by_provider', {}).get('anthropic', {}).get('total_cost', 0),
                    # Enhanced metrics
                    rows_processed=total_rows_processed,
                    total_rows=validation_results.get('total_rows', 0),
                    columns_validated=original_validation_metrics.get('validated_columns_count', 0),
                    search_groups=original_validation_metrics.get('search_groups_count', 0),
                    high_context_search_groups=original_validation_metrics.get('enhanced_context_search_groups_count', 0),
                    claude_calls=original_validation_metrics.get('claude_search_groups_count', 0),
                    eliyahu_cost=eliyahu_cost,  # Actual cost paid
                    estimated_cost=cost_estimated,  # Raw cost estimate without caching
                    quoted_validation_cost=quoted_full_cost,  # This is the scaled full table quote
                    charged_cost=0.0,  # Preview doesn't charge
                    total_api_calls=total_api_calls,
                    total_cached_calls=total_cached_calls
                    )
                    logger.debug(f"[USER_TRACKING] Preview tracking result: {track_result}")
                except Exception as e:
                    logger.error(f"[USER_TRACKING] Failed to track preview request: {e}")
                    import traceback
                    logger.error(f"[USER_TRACKING] Traceback: {traceback.format_exc()}")
                
                # Send WebSocket notification with preview results
                from dynamodb_schemas import get_connection_by_session, remove_websocket_connection
                logger.info(f"Looking up WebSocket connection for session {session_id}")
                connection_id = get_connection_by_session(session_id)
                logger.info(f"Found WebSocket connection_id: {connection_id}")
                
                if connection_id:
                    try:
                        client = _get_api_gateway_management_client(context)
                        logger.info(f"API Gateway management client initialized: {client is not None}")
                        
                        if client:
                            websocket_payload = {
                                'status': 'COMPLETED',
                                'percent_complete': 100,
                                'verbose_status': 'Preview complete. Results available.',
                                'preview_data': preview_payload
                            }
                            
                            # DEBUG: Log the cost estimates being sent to frontend
                            cost_estimates = preview_payload.get('cost_estimates', {})
                            logger.debug(f"[COST_DEBUG] WebSocket payload cost_estimates: {cost_estimates}")
                            logger.debug(f"[COST_DEBUG] quoted_validation_cost: {cost_estimates.get('quoted_validation_cost')}")

                            # Log call counts being sent to frontend (CALL COUNTS DEBUGGING)
                            total_provider_calls_ws = cost_estimates.get('total_provider_calls', 0)
                            validation_metrics_ws = preview_payload.get('validation_metrics', {})
                            search_groups_count_ws = validation_metrics_ws.get('search_groups_count', 0)
                            claude_search_groups_count_ws = validation_metrics_ws.get('claude_search_groups_count', 0)
                            calculated_perplexity_groups = search_groups_count_ws - claude_search_groups_count_ws

                            logger.debug(f"[WEBSOCKET_SENDING] Sending call counts to frontend:")
                            logger.debug(f"[WEBSOCKET_SENDING]   total_provider_calls: {total_provider_calls_ws}")
                            logger.debug(f"[WEBSOCKET_SENDING]   search_groups_count (total): {search_groups_count_ws}")
                            logger.debug(f"[WEBSOCKET_SENDING]   claude_search_groups_count: {claude_search_groups_count_ws}")
                            logger.debug(f"[WEBSOCKET_SENDING]   frontend will calculate perplexity_groups: {calculated_perplexity_groups}")
                            logger.debug(f"[WEBSOCKET_SENDING]   payload contains preview_data with {len(preview_payload)} top-level fields")

                            logger.info(f"Sending WebSocket payload to connection {connection_id}: {len(json.dumps(websocket_payload))} bytes")
                            
                            client.post_to_connection(
                                ConnectionId=connection_id,
                                Data=json.dumps(websocket_payload).encode('utf-8')
                            )
                            logger.debug(f"Successfully sent preview completion notification to WebSocket {connection_id}")
                        else:
                            logger.error("API Gateway management client is None - cannot send WebSocket message")
                    except client.exceptions.GoneException:
                        logger.warning(f"WebSocket connection {connection_id} is stale. Removing from DB.")
                        remove_websocket_connection(connection_id)
                    except Exception as e:
                        logger.error(f"❌ Failed to post preview completion to WebSocket {connection_id}: {e}")
                        import traceback
                        logger.error(traceback.format_exc())
                else:
                    logger.warning(f"No WebSocket connection found for session {session_id} - cannot send notification")
                
                # Enhanced Excel generation has been moved earlier in the function (before WebSocket message)
                # to ensure enhanced_download_url is available in the preview payload
                
                # Store preview results in versioned results folder using unified storage
                config_version = 1
                config_id = "unknown"
                try:
                    existing_config, latest_config_key = storage_manager.get_latest_config(email, clean_session_id)
                    if existing_config and existing_config.get('storage_metadata', {}):
                        config_version = existing_config['storage_metadata'].get('version', 1)
                        config_id = existing_config['storage_metadata'].get('config_id', 'unknown')
                except Exception as e:
                    logger.warning(f"Could not determine config version for preview: {e}")
                
                # Send final storage progress update - interface final processing (90-100% range)
                _send_websocket_message_deduplicated(session_id, {
                    'type': 'preview_progress',
                    'progress': 98,  # Interface final processing: 90-100% range
                    'status': '💾 Storing preview results...',
                    'session_id': session_id
                }, "preview_progress_storage")
                
                # Update session info with account data for preview
                account_info = {
                    'initial_balance': float(initial_balance) if initial_balance else 0,
                    'final_balance': float(final_balance) if final_balance else 0,
                    'amount_charged': 0,  # Previews are free
                    'domain_multiplier': float(multiplier),
                    'eliyahu_cost': float(eliyahu_cost),  # Your actual expense
                    'estimated_cost': float(cost_estimated),  # What it would cost without caching
                    # Removed preview_abandoned and insufficient_balance_encountered - unnecessary fields
                    'processing_type': 'preview'
                }
                _update_session_info_with_account_data(email, clean_session_id, account_info)
                
                # Frontend payload already created above for DynamoDB
                
                # Store preview results in unified storage (full payload for processing)
                result = storage_manager.store_results(
                    email, clean_session_id, config_version, preview_payload, 'preview'
                )
                
                if result['success']:
                    logger.info(f"Preview results stored in versioned folder: {result['s3_key']}")
                    
                    # Update session tracking with minimal frontend payload for UX analysis
                    try:
                        success = storage_manager.update_session_results(
                            email=email,
                            session_id=clean_session_id,
                            operation_type="preview",
                            config_id=config_id,
                            version=config_version,
                            run_key=run_key,
                            results_path=result['s3_key'],  # Path to preview results JSON
                            frontend_payload=frontend_payload  # Minimal frontend payload for UX tracking
                        )
                        
                        if success:
                            logger.info(f"✅ Updated session_info.json with preview completion")
                        else:
                            logger.error(f"❌ Failed to update session_info.json with preview completion")
                    except Exception as e:
                        logger.error(f"Failed to update preview session tracking: {e}")
                else:
                    logger.error(f"Failed to store preview results: {result.get('error')}")
                
                # Send final completion progress update - interface final processing complete (100%)
                _send_websocket_message_deduplicated(session_id, {
                    'type': 'preview_progress',
                    'progress': 100,  # Interface final processing complete: 100%
                    'status': '🎉 Preview completed successfully!',
                    'session_id': session_id
                }, "preview_progress_final")

                # No longer need to save a separate S3 object for previews
            else:
                # Send failure progress update
                error_payload = {
                    'type': 'preview_failed',
                    'progress': 100,
                    'status': '❌ Preview failed to generate results',
                    'session_id': session_id,
                    'error': 'No validation results returned'
                }
                _send_websocket_message_deduplicated(session_id, error_payload, "preview_failed_no_results")

                # Update session info for failed preview
                try:
                    from dynamodb_schemas import check_user_balance, get_domain_multiplier
                    initial_balance = check_user_balance(email)
                    email_domain = email.split('@')[-1] if '@' in email else 'unknown'
                    multiplier = float(get_domain_multiplier(email_domain))

                    account_info = {
                        'initial_balance': float(initial_balance) if initial_balance else 0,
                        'final_balance': float(initial_balance) if initial_balance else 0,
                        'amount_charged': 0,  # Previews are free
                        'domain_multiplier': float(multiplier),
                        'raw_cost': 0,  # No processing happened
                        # Removed preview_abandoned and insufficient_balance_encountered - unnecessary fields
                        'processing_type': 'preview_failed'
                    }
                    _update_session_info_with_account_data(email, clean_session_id, account_info)
                except Exception as e:
                    logger.error(f"Failed to update session info for failed preview: {e}")

                update_run_status_for_session(status='FAILED',
                    run_type="Preview",
                    verbose_status="Preview failed to generate results.",
                    percent_complete=100,
                    processed_rows=0,
                    batch_size=10,  # Default batch size
                    eliyahu_cost=0.0,
                    time_per_row_seconds=0.0)

                # Update session_info.json with failure status and correct config version
                try:
                    storage_manager.update_session_results(
                        email=email,
                        session_id=clean_session_id,
                        operation_type="preview",
                        config_id=config_id,
                        version=config_version,
                        run_key=run_key,
                        status="failed",
                        completed_at=datetime.now(timezone.utc).isoformat(),
                        frontend_payload=error_payload
                    )
                    logger.debug(f"[SESSION_TRACKING] Updated session_info.json with preview failure (no results)")
                except Exception as e:
                    logger.error(f"Failed to update session_info.json for preview failure: {e}")
            
            # Return early for preview mode to prevent fallthrough to normal processing
            return {'statusCode': 200, 'body': json.dumps({'status': 'preview_completed', 'session_id': session_id})}

        else:
            # Normal mode processing
            results_key = event.get('results_key')
            if not results_key:
                # Generate results_key if not provided (e.g., for async completion)
                # No longer creating ZIP - just track the results folder
                results_key = f"results/{session_id}/validation_complete"
                logger.info(f"Generated results_key for session {session_id}: {results_key}")
            logger.debug(f"Background normal processing for session {session_id}")

            # SECURITY: Check balance BEFORE starting expensive full validation processing
            from dynamodb_schemas import check_user_balance, get_domain_multiplier
            from ..core.unified_s3_manager import UnifiedS3Manager

            try:
                current_balance = check_user_balance(email)
                logger.debug(f"[SECURITY] Full validation balance check for {email}: ${current_balance:.6f}")

                # Get estimated cost from preview data (what was quoted to user)
                storage_manager_temp = UnifiedS3Manager()
                preview_data = storage_manager_temp.get_latest_preview_results(email, clean_session_id)

                if preview_data and preview_data.get('cost_estimates'):
                    estimated_cost = preview_data['cost_estimates'].get('quoted_validation_cost', 0.01)
                    logger.debug(f"[SECURITY] Using quoted cost from preview: ${estimated_cost:.6f}")
                else:
                    # Fallback: calculate estimated cost using domain multiplier
                    email_domain = email.split('@')[-1] if '@' in email else 'unknown'
                    multiplier = get_domain_multiplier(email_domain)
                    estimated_cost = max(0.01, 0.01 * float(multiplier))  # Minimum cost with domain multiplier
                    logger.warning(f"[SECURITY] No preview cost found, using fallback: ${estimated_cost:.6f} (domain: {email_domain}, multiplier: {multiplier})")

                # Reject if insufficient balance
                if current_balance is None or current_balance < estimated_cost:
                    logger.error(f"[SECURITY] BLOCKING full validation - insufficient balance: ${current_balance:.6f} < ${estimated_cost:.6f}")

                    # Send insufficient balance error via WebSocket
                    _send_websocket_message(session_id, {
                        'type': 'validation_failed',
                        'session_id': session_id,
                        'progress': 100,
                        'status': '❌ Insufficient balance for full validation',
                        'error': 'insufficient_balance',
                        'current_balance': float(current_balance) if current_balance else 0,
                        'estimated_cost': float(estimated_cost),
                        'credits_needed': max(0, estimated_cost - (current_balance or 0))
                    })

                    # Update run status to failed
                    update_run_status_for_session(
                        status='FAILED',
                        run_type='Validation',
                        verbose_status=f'Insufficient balance: ${current_balance:.2f} < ${estimated_cost:.2f}',
                        percent_complete=100
                    )

                    return {
                        'statusCode': 402,  # Payment Required
                        'body': json.dumps({
                            'status': 'failed',
                            'error': 'insufficient_balance',
                            'current_balance': float(current_balance) if current_balance else 0,
                            'estimated_cost': float(estimated_cost),
                            'session_id': session_id
                        })
                    }

                logger.debug(f"[SECURITY] ✅ Balance sufficient for full validation: ${current_balance:.6f} >= ${estimated_cost:.6f}")

            except ImportError:
                logger.warning("[SECURITY] DynamoDB not available - skipping balance check (development mode)")
            except Exception as e:
                error_msg = str(e)
                logger.error(f"[SECURITY] Critical error in balance check before full validation: {error_msg}")

                # Classify the error type
                error_type = "Balance Check System Failure"
                if "get_latest_preview_results" in error_msg or "get_preview_data" in error_msg:
                    error_type = "Preview Data Access Error"
                elif "UnifiedS3Manager" in error_msg:
                    error_type = "Storage System Error"
                elif "check_user_balance" in error_msg:
                    error_type = "Balance Verification Error"

                # Prepare session data for alert
                session_data = {
                    'session_id': session_id,
                    'email': email,
                    'clean_session_id': clean_session_id,
                    'error_location': 'pre_validation_balance_check',
                    'system_component': 'UnifiedS3Manager/DynamoDB',
                    'timestamp': datetime.now(timezone.utc).isoformat()
                }

                # Send critical system failure alert
                send_validation_failure_alert(session_id, email, error_type, error_msg, session_data)

                # Update run status to failed
                update_run_status_for_session(
                    status='FAILED',
                    run_type='Validation',
                    verbose_status=f'Pre-validation system check failed: {error_type}',
                    percent_complete=100
                )

                # Send failure notification via WebSocket
                _send_websocket_message(session_id, {
                    'type': 'validation_failed',
                    'session_id': session_id,
                    'progress': 100,
                    'status': f'❌ System error: {error_type}',
                    'error': error_type.lower().replace(' ', '_'),
                    'message': 'Technical system error encountered. Our team has been notified and will resolve this promptly.'
                })

                # DO NOT continue processing - return error immediately
                return {
                    'statusCode': 500,
                    'body': json.dumps({
                        'status': 'failed',
                        'error': error_type,
                        'message': 'System error during pre-validation checks. Processing halted for safety.',
                        'session_id': session_id
                    })
                }

            # Use unified storage to get Excel and config files
            from ..core.unified_s3_manager import UnifiedS3Manager
            storage_manager = UnifiedS3Manager()
            
            # Get files from unified storage using clean session ID
            excel_content, actual_excel_s3_key = storage_manager.get_excel_file(email, clean_session_id)
            config_data, actual_config_s3_key = storage_manager.get_latest_config(email, clean_session_id)
            
            if not excel_content:
                logger.error(f"Failed to retrieve Excel file from unified storage for session {clean_session_id}")
                raise Exception(f"Excel file not found in unified storage for session {clean_session_id}")
            
            if not config_data:
                logger.error(f"Failed to retrieve config file from unified storage for session {clean_session_id}")
                raise Exception(f"Config file not found in unified storage for session {clean_session_id}")
            
            logger.info(f"Retrieved files from unified storage - Excel: {actual_excel_s3_key}, Config: {actual_config_s3_key}")
            
            # Use shared table parser to get row count (handles both CSV and Excel)
            from shared_table_parser import S3TableParser
            table_parser = S3TableParser()
            
            # Parse table to get structure and row count
            table_data = table_parser.parse_s3_table(storage_manager.bucket_name, actual_excel_s3_key, extract_formulas=True)
            total_rows_in_file = len(table_data.get('rows', []))
            rows_to_process = min(max_rows, total_rows_in_file) if max_rows else total_rows_in_file

            # The actual batching loop
            all_validation_results = {}
            processed_rows_count = 0

            # Ensure batch_size has a default value if None
            effective_batch_size = batch_size if batch_size is not None else 3  # Small batch for testing continuations
            total_batches = (rows_to_process + effective_batch_size - 1) // effective_batch_size
            
            update_run_status_for_session( status='PROCESSING', run_type="Validation", verbose_status=f"Validation preparing to process {rows_to_process} rows in {total_batches} batches.", batch_size=effective_batch_size)

            # DISABLED: Fake simulation replaced with real validation lambda invocation
            logger.debug(f"[DEBUG] Fake batch processing loop DISABLED for session {session_id}")
            logger.debug(f"[DEBUG] Should invoke validation lambda with {rows_to_process} rows in {total_batches} batches")
            
            # TODO: Replace with actual validation lambda invocation
            # This simulation was creating fake "first batch -> final batch" jumps
            
            # Interface setup for full validation - use 0-5% range  
            # Update start_time to when background handler actually begins validation processing
            update_run_status_for_session(status='PROCESSING',
                processed_rows=0,
                percent_complete=2,  # Interface setup: 0-5% range
                verbose_status=f"Validation starting full processing for {rows_to_process} rows...",
                start_time=background_start_time,  # Use background handler start time
                run_type="Validation")
            
            # The real processing should happen via validation lambda invocation
            # which will send proper WebSocket progress updates
            
            # Original simulation loop commented out:
            """
            for i in range(total_batches):
                start_row = i * batch_size
                end_row = min(start_row + batch_size, rows_to_process)
                
                # In a real scenario, you'd slice the data and send only the batch
                # Here, we'll just simulate the progress update
                
                # SIMULATE a call to a batch processor
                time.sleep(1) # Simulate work
                
                processed_rows_count = end_row
                percent_complete = int((processed_rows_count / rows_to_process) * 100)
                verbose_status = f"Processing batch {i + 1} of {total_batches} ({processed_rows_count}/{rows_to_process} rows)..."
                
                update_run_status_for_session(status='PROCESSING',
                    processed_rows=processed_rows_count,
                    percent_complete=percent_complete,
                    verbose_status=verbose_status)
            """
            
            # After the loop, we would have the final results.
            # For now, we will call the original invoker to get the final result in one go,
            # as refactoring it to return per-batch results is a larger change.

            # ========== GENERATE COMPLETE VALIDATION PAYLOAD (SHARED BY SYNC AND ASYNC) ==========
            # Skip payload generation for async completion - we already have the results
            if event.get('async_completion_mode'):
                logger.debug(f"[PAYLOAD_GENERATION] Skipping payload generation for async completion - results already available")
                complete_validation_payload = {}  # Not needed for async completion
            else:
                # This payload generation is used by sync processing and initial async delegation
                logger.debug(f"[PAYLOAD_GENERATION] Creating complete validation payload for {rows_to_process} rows")

                # Get parsed rows from the table data (S3TableParser uses 'data' field)
                excel_rows = table_data.get('data', [])
                if max_rows:
                    excel_rows = excel_rows[:max_rows]

                logger.debug(f"[PAYLOAD_GENERATION] Processing {len(excel_rows)} rows for validation payload")

                # Get ID fields from config for row key generation (same logic as sync validation)
                id_fields = []
                logger.debug(f"[PAYLOAD_GENERATION] Searching for ID fields in validation targets...")
                for i, target in enumerate(config_data.get('validation_targets', [])):
                    importance = target.get('importance', '')
                    if importance.lower() == 'id':
                        column = target.get('column', '')
                        if column:
                            id_fields.append(column)
                            logger.debug(f"[PAYLOAD_GENERATION] Found ID field: {column}")

                # Fallback to SimplifiedSchemaValidator primary keys if no explicit ID fields
                if not id_fields:
                    try:
                        from simplified_schema_validator import SimplifiedSchemaValidator
                        validator = SimplifiedSchemaValidator(config_data)
                        id_fields = validator.primary_key
                        logger.debug(f"[PAYLOAD_GENERATION] Using primary keys from SimplifiedSchemaValidator: {id_fields}")
                    except Exception as e:
                        logger.warning(f"[PAYLOAD_GENERATION] Could not use SimplifiedSchemaValidator: {e}")
                        # Use default fallback ID fields for demo
                        id_fields = ['Ticker_Symbol', 'Asset_ID']
                        logger.debug(f"[PAYLOAD_GENERATION] Using fallback ID fields: {id_fields}")

                logger.debug(f"[PAYLOAD_GENERATION] ID fields for row key generation: {id_fields}")

                # Add pre-computed row keys to each row using HYBRID hashing:
                # 1. Try ID-field hashing first (for history matching)
                # 2. For duplicates, use full-row hashing (to distinguish them)
                from row_key_utils import generate_row_key
                processed_rows = []
                id_hash_counts = {}  # Track duplicate ID hashes

                # First pass: Generate ID-field hashes and detect duplicates
                for row_idx, row_data in enumerate(excel_rows):
                    # Generate ID-field hash
                    id_hash = generate_row_key(row_data, primary_keys=id_fields if id_fields else None)

                    if id_hash not in id_hash_counts:
                        id_hash_counts[id_hash] = []
                    id_hash_counts[id_hash].append(row_idx)

                # Second pass: Assign final row keys (ID-hash or full-row hash for duplicates)
                for row_idx, row_data in enumerate(excel_rows):
                    id_hash = generate_row_key(row_data, primary_keys=id_fields if id_fields else None)

                    # If this ID hash appears multiple times, use full-row hash
                    if len(id_hash_counts[id_hash]) > 1:
                        row_key = generate_row_key(row_data, primary_keys=None)  # Full-row hash
                        logger.debug(f"[PAYLOAD_GENERATION] Row {row_idx}: Duplicate ID detected, using full-row hash: {row_key[:8]}...")
                    else:
                        row_key = id_hash  # Use ID-field hash
                        logger.debug(f"[PAYLOAD_GENERATION] Row {row_idx}: Using ID-field hash: {row_key[:8]}...")

                    # Add row key to row data
                    row_data['_row_key'] = row_key
                    processed_rows.append(row_data)

                # Log duplicate summary
                duplicate_id_count = sum(1 for count_list in id_hash_counts.values() if len(count_list) > 1)
                total_duplicate_rows = sum(len(count_list) for count_list in id_hash_counts.values() if len(count_list) > 1)
                if duplicate_id_count > 0:
                    logger.info(f"[PAYLOAD_GENERATION] Found {duplicate_id_count} duplicate ID groups ({total_duplicate_rows} total rows)")
                    logger.info(f"[PAYLOAD_GENERATION] Duplicate rows will use full-row hashing (history matching disabled for these)")
                else:
                    logger.info(f"[PAYLOAD_GENERATION] No duplicate IDs detected, all rows using ID-field hashing")

                logger.debug(f"[PAYLOAD_GENERATION] Added pre-computed row keys to {len(processed_rows)} rows")

                # CRITICAL: Update table_data with row keys so Excel report can access them
                if table_data and isinstance(table_data, dict) and 'data' in table_data:
                    table_data['data'] = processed_rows  # Replace with rows that include _row_key
                    logger.debug(f"[PAYLOAD_GENERATION] Updated table_data with processed rows (including _row_key)")

                # Load validation history if available and we have Excel content (matching invoke_validator_lambda logic)
                validation_history = {}
                try:
                    # Import validation history loader
                    from interface_lambda.utils.history_loader import load_validation_history_from_excel

                    # Save Excel content to a temporary file for history extraction
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                        tmp_file.write(excel_content)
                        tmp_file_path = tmp_file.name

                    original_validation_history = load_validation_history_from_excel(tmp_file_path)

                    logger.debug(f"[PAYLOAD_GENERATION] Loaded validation history for {len(original_validation_history)} row keys from Excel")

                    if original_validation_history:
                        validation_history = original_validation_history

                        # Log matching summary
                        matched_count = 0
                        for row in processed_rows:
                            payload_key = row.get('_row_key', '')
                            if payload_key and payload_key in validation_history:
                                matched_count += 1

                        logger.debug(f"[PAYLOAD_GENERATION] Matched {matched_count} out of {len(processed_rows)} rows with validation history")

                        if matched_count == 0 and len(validation_history) > 0:
                            logger.warning("[PAYLOAD_GENERATION] No rows matched with validation history - may be using different primary keys")
                            logger.debug(f"[PAYLOAD_GENERATION] Current primary keys: {id_fields}")

                    # Clean up temp file
                    try:
                        import os
                        os.unlink(tmp_file_path)
                    except Exception as cleanup_error:
                        logger.warning(f"[PAYLOAD_GENERATION] Failed to cleanup temp file: {cleanup_error}")

                except Exception as e:
                    logger.warning(f"[PAYLOAD_GENERATION] Failed to load validation history: {e}")
                    validation_history = {}

                # Create the complete validation payload (used by both sync and async)
                complete_validation_payload = {
                    "test_mode": False,
                    "config": config_data,  # Embedded config data
                    "validation_data": {
                        "rows": processed_rows,  # All Excel data with pre-computed row keys
                        "max_rows": max_rows,
                        "batch_size": batch_size,
                        "total_dataset_size": len(processed_rows)
                    },
                    "validation_history": validation_history,  # Properly loaded validation history
                    "session_id": session_id
                }

                logger.debug(f"[PAYLOAD_GENERATION] Complete validation payload created with {len(processed_rows)} rows")

            # ========== SMART DELEGATION SYSTEM DECISION POINT ==========
            # Check if we should delegate to async processing based on estimated time
            # IMPORTANT: Never delegate previews to async - they must run synchronously
            estimated_minutes = 0.0
            should_delegate = False

            # URGENT FIX: Skip delegation if this is async completion mode (prevents infinite loop)
            if event.get('async_completion_mode'):
                logger.debug(f"[DELEGATION_DECISION] Async completion mode detected - FORCING synchronous completion to prevent infinite loop")
                should_delegate = False
            # Skip delegation for previews - they must always run synchronously
            elif is_preview:
                logger.debug(f"[DELEGATION_DECISION] Preview mode detected - forcing synchronous processing")
                should_delegate = False
            else:
                # Try to get the estimated processing time from stored preview data
                try:
                    # Retrieve the time estimate from stored preview data
                    storage_manager_temp = UnifiedS3Manager()
                    preview_data = storage_manager_temp.get_latest_preview_results(email, clean_session_id)

                    estimated_total_time_seconds = None
                    if preview_data:
                        # Try multiple possible field names where time estimates might be stored
                        time_estimate_fields = [
                        'estimated_validation_time',           # Seconds - most direct field
                        'estimated_total_processing_time',     # Seconds - backup field
                        'estimated_total_time_seconds'         # Seconds - legacy field
                    ]

                        for field_name in time_estimate_fields:
                            if field_name in preview_data and preview_data[field_name]:
                                estimated_total_time_seconds = preview_data[field_name]
                                logger.debug(f"[DELEGATION_DECISION] Found time estimate in field '{field_name}': {estimated_total_time_seconds}s")
                                break

                        # If we didn't find seconds, try minutes fields
                        if not estimated_total_time_seconds:
                            minutes_fields = [
                            'estimated_validation_time_minutes'  # Minutes - convert to seconds
                        ]

                            for field_name in minutes_fields:
                                if field_name in preview_data and preview_data[field_name]:
                                    estimated_total_time_seconds = preview_data[field_name] * 60.0
                                    logger.debug(f"[DELEGATION_DECISION] Found time estimate in field '{field_name}': {preview_data[field_name]}min = {estimated_total_time_seconds}s")
                                    break

                        # Also try nested locations for backwards compatibility
                        if not estimated_total_time_seconds:
                            nested_locations = [
                            ('processing_estimates', 'estimated_total_time_seconds'),
                            ('cost_estimates', 'estimated_validation_time')
                        ]

                            for parent_key, child_key in nested_locations:
                                if preview_data.get(parent_key) and preview_data[parent_key].get(child_key):
                                    estimated_total_time_seconds = preview_data[parent_key][child_key]
                                    logger.debug(f"[DELEGATION_DECISION] Found time estimate in nested location '{parent_key}.{child_key}': {estimated_total_time_seconds}s")
                                    break

                    if estimated_total_time_seconds and estimated_total_time_seconds > 0:
                        estimated_minutes = estimated_total_time_seconds / 60.0
                        should_delegate = estimated_minutes > MAX_SYNC_INVOCATION_TIME_MINUTES

                        logger.debug(f"[DELEGATION_DECISION] Estimated processing time: {estimated_minutes:.1f} minutes")
                        logger.debug(f"[DELEGATION_DECISION] Sync timeout limit: {MAX_SYNC_INVOCATION_TIME_MINUTES:.1f} minutes")
                        logger.debug(f"[DELEGATION_DECISION] Should delegate: {should_delegate}")
                    else:
                        logger.warning(f"[DELEGATION_DECISION] No valid time estimate found in preview data, using sync processing")
                        logger.debug(f"[DELEGATION_DECISION] Available preview data keys: {list(preview_data.keys()) if preview_data else 'None'}")
                        if preview_data:
                            # Log some field values to help debug
                            debug_fields = ['estimated_validation_time', 'estimated_total_processing_time', 'estimated_validation_time_minutes']
                            for field in debug_fields:
                                if field in preview_data:
                                    logger.debug(f"[DELEGATION_DECISION] {field}: {preview_data[field]}")
                except Exception as e:
                    logger.error(f"[DELEGATION_DECISION] Error checking delegation criteria: {e}")

            if should_delegate:
                # ========== ASYNC DELEGATION WORKFLOW ==========
                logger.debug(f"[DELEGATION] Delegating to async processing (estimated {estimated_minutes:.1f}min > {MAX_SYNC_INVOCATION_TIME_MINUTES:.1f}min)")

                try:
                    # Import WebSocket connection function
                    from dynamodb_schemas import get_connection_by_session

                    # Get WebSocket connection ID for context preservation
                    connection_id = get_connection_by_session(session_id) if session_id else None

                    # Prepare delegation context with all necessary data
                    delegation_context = {
                        'request_context': {
                            'session_id': session_id,
                            'run_key': run_key,
                            'email': email,
                            'websocket_connection_id': connection_id,
                            'start_time': background_start_time,  # Use background_start_time instead of undefined start_time
                            'config_data': config_data,
                            'total_rows': total_rows_in_file,
                            'max_rows': max_rows,
                            'batch_size': batch_size,
                            # Add critical variables for async completion
                            'reference_pin': reference_pin,
                            'preview_mode': is_preview,
                            'preview_email': preview_email,
                            'timestamp': timestamp,
                            'clean_session_id': clean_session_id,
                            'email_folder': event.get('email_folder', 'Full Validation' if not is_preview else 'Preview')  # Add email folder info
                        },
                        'file_locations': {
                            'excel_s3_key': excel_s3_key,
                            'config_s3_key': config_s3_key,
                            'validation_history': None,  # Set to None for now - validation_history variables not defined in this scope
                            'results_key': None  # Will be updated after validation completes
                        },
                        'preview_estimates': {
                            'estimated_total_time_seconds': estimated_total_time_seconds,
                            'estimated_minutes': estimated_minutes,
                            'cost_estimated': cost_estimated if 'cost_estimated' in locals() else 0.0,
                            'multiplier': multiplier if 'multiplier' in locals() else 1.0
                        },
                        'processing_metadata': {
                            'S3_UNIFIED_BUCKET': S3_UNIFIED_BUCKET,
                            'VALIDATOR_LAMBDA_NAME': VALIDATOR_LAMBDA_NAME,
                            'table_data': table_data,  # Store table_data for Excel generation
                            'config_version': config_data.get('storage_metadata', {}).get('version', 1) if config_data else 1  # Track config version
                        }
                    }

                    # Save delegation context to DynamoDB
                    delegation_success = save_delegation_context(
                        session_id=session_id,
                        run_key=run_key,
                        context_data=delegation_context,
                        estimated_minutes=estimated_minutes,
                        sync_timeout_minutes=MAX_SYNC_INVOCATION_TIME_MINUTES,
                        reason=f"estimated_{estimated_minutes:.1f}min_exceeds_{MAX_SYNC_INVOCATION_TIME_MINUTES:.1f}min_sync_limit"
                    )

                    if delegation_success:
                        # ========== USE PRE-GENERATED VALIDATION PAYLOAD ==========
                        # Use the complete validation payload that was already generated
                        logger.debug(f"[DELEGATION] Using pre-generated validation payload for async processing")

                        # Check if we have valid payload with rows
                        payload_rows = complete_validation_payload.get('validation_data', {}).get('rows', [])
                        if len(payload_rows) == 0:
                            logger.error(f"[DELEGATION] CRITICAL: Pre-generated payload has no rows!")
                            logger.error(f"[DELEGATION] This will cause async validation to fail")
                            should_delegate = False  # Fall back to sync processing
                        else:
                            logger.debug(f"[DELEGATION] Pre-generated payload contains {len(payload_rows)} rows with pre-computed keys")

                            # Extract config version for S3 path construction
                            config_version = config_data.get('storage_metadata', {}).get('version', 1) if config_data else 1

                            # Construct results path for continuation chain
                            domain = email.split('@')[-1].lower().strip() if email and '@' in email else 'unknown'
                            email_prefix = email.split('@')[0].replace('.', '_').replace('+', '_plus_')[:20] if email and '@' in email else 'unknown'
                            results_path = f"results/{domain}/{email_prefix}/{session_id}/v{config_version}_results"

                            # Add async-specific fields to the pre-generated payload
                            async_payload = complete_validation_payload.copy()
                            async_payload.update({
                                "async_delegation_request": True,  # Flag to indicate async processing
                                "run_key": run_key,
                                "S3_UNIFIED_BUCKET": S3_UNIFIED_BUCKET,
                                "VALIDATOR_LAMBDA_NAME": VALIDATOR_LAMBDA_NAME,
                                "email": email,  # Include for S3 path construction
                                "config_version": config_version,  # Include for S3 path construction
                                "results_path": results_path  # Include to avoid reconstruction
                            })

                        if should_delegate and len(payload_rows) > 0:  # Only proceed if we have valid payload with rows

                            # Store complete payload in S3 for async validator
                            payload_s3_key = f"async_payloads/{session_id}/{run_key}/complete_validation_payload.json"

                            # s3_client already initialized at function start
                            s3_client.put_object(
                                Bucket=S3_UNIFIED_BUCKET,
                                Key=payload_s3_key,
                                Body=json.dumps(async_payload, default=str),
                                ContentType='application/json'
                            )

                            logger.debug(f"[DELEGATION] Stored complete payload in S3: {payload_s3_key}")

                            # Trigger async validator via direct Lambda invocation
                            async_payload_event = {
                                'message_type': 'ASYNC_VALIDATION_REQUEST',
                                'session_id': session_id,
                                'run_key': run_key,
                                'async_delegation_request': True,
                                'complete_payload_s3_key': payload_s3_key,  # Key to complete sync-compatible payload
                                'S3_UNIFIED_BUCKET': S3_UNIFIED_BUCKET,
                                'VALIDATOR_LAMBDA_NAME': VALIDATOR_LAMBDA_NAME,
                                'results_path': results_path,  # Pass through for continuation chain
                                'config_version': config_version,  # Pass through for S3 path construction
                                'email': email,  # Pass through for S3 path construction
                                # Legacy fields for backward compatibility
                                'excel_s3_key': excel_s3_key,
                                'config_s3_key': config_s3_key,
                                'max_rows': max_rows,
                                'batch_size': batch_size,
                                'validation_history': None
                            }

                            try:
                                # Initialize Lambda client for async invocation
                                import boto3
                                lambda_client = boto3.client('lambda')

                                # Log critical delegation parameters before invoking
                                logger.debug(f"[DELEGATION] Invoking async validator with:")
                                logger.debug(f"[DELEGATION]   FunctionName: {VALIDATOR_LAMBDA_NAME}")
                                logger.debug(f"[DELEGATION]   session_id: {session_id}")
                                logger.debug(f"[DELEGATION]   run_key: {run_key}")
                                logger.debug(f"[DELEGATION]   complete_payload_s3_key: {payload_s3_key}")
                                logger.debug(f"[DELEGATION]   S3_UNIFIED_BUCKET: {S3_UNIFIED_BUCKET}")
                                logger.debug(f"[DELEGATION]   results_path: {results_path}")
                                logger.debug(f"[DELEGATION]   config_version: {config_version}")
                                logger.debug(f"[DELEGATION]   email: {email}")
                                logger.debug(f"[DELEGATION] Full payload keys: {list(async_payload_event.keys())}")

                                # Direct Lambda invocation (async)
                                response = lambda_client.invoke(
                                    FunctionName=VALIDATOR_LAMBDA_NAME,
                                    InvocationType='Event',  # Async invocation
                                    Payload=json.dumps(async_payload_event, default=str)
                                )

                                logger.debug(f"[DELEGATION] Successfully triggered async validator via direct invocation, status: {response['StatusCode']}")

                                # Check if invocation was successful (StatusCode 202 for async invocation)
                                if response['StatusCode'] == 202:
                                    logger.debug(f"[DELEGATION] Async validation lambda successfully invoked")

                                    # Send success notification via WebSocket
                                    try:
                                        _send_websocket_message(session_id, {
                                            'type': 'delegation_success',
                                            'status': 'DELEGATED_TO_ASYNC',
                                            'message': f'Job delegated to async processing (estimated {estimated_minutes:.1f} minutes)',
                                            'estimated_minutes': estimated_minutes,
                                            'info': 'Your validation is now running in the background. You will receive updates as processing progresses.'
                                        })
                                        logger.debug(f"[DELEGATION] Sent delegation success notification via WebSocket for session {session_id}")
                                    except Exception as websocket_error:
                                        logger.error(f"[DELEGATION] Failed to send WebSocket success notification: {websocket_error}")

                                    # Send final message with timeout information for frontend
                                    max_expected_runtime = max(estimated_minutes * 1.5, 30)  # 1.5x estimated or 30min minimum
                                    try:
                                        _send_websocket_message(session_id, {
                                            'type': 'async_delegation_complete',
                                            'status': 'ASYNC_PROCESSING',
                                            'message': f'Validation running in background (est. {estimated_minutes:.1f}min)',
                                            'estimated_minutes': estimated_minutes,
                                            'max_expected_minutes': max_expected_runtime,
                                            'timeout_warning': f'If no completion message is received within {max_expected_runtime:.0f} minutes, the validator may have failed',
                                            'instructions': 'This interface will close now. Check your email for completion notification or refresh the page later to check status.'
                                        })
                                        logger.debug(f"[DELEGATION] Sent final delegation message with {max_expected_runtime:.0f}min timeout warning")
                                    except Exception as websocket_error:
                                        logger.error(f"[DELEGATION] Failed to send final delegation message: {websocket_error}")

                                    # Background handler job is done - async validator will take over
                                    logger.debug(f"[DELEGATION] Background handler completed delegation for session {session_id}")
                                    return

                                else:
                                    logger.error(f"[DELEGATION] Async validation lambda invocation failed with status: {response['StatusCode']}")
                                    logger.error(f"[DELEGATION] This indicates a problem with lambda permissions or configuration")

                                    # Send failure notification via WebSocket
                                    try:
                                        _send_websocket_message(session_id, {
                                            'type': 'validation_error',
                                            'status': 'FAILED',
                                            'error': 'Async validation system failed to start',
                                            'details': f'Lambda invocation returned status code {response["StatusCode"]}',
                                            'troubleshooting': 'Check lambda permissions and configuration',
                                            'retry_suggestion': 'Please try your validation again or contact support'
                                        })
                                        logger.debug(f"[DELEGATION] Sent failure notification via WebSocket for session {session_id}")
                                    except Exception as websocket_error:
                                        logger.error(f"[DELEGATION] Failed to send WebSocket failure notification: {websocket_error}")

                                    # Continue with sync processing as fallback
                                    logger.warning(f"[DELEGATION] Async delegation failed, will fall back to sync processing")
                                    should_delegate = False

                            except Exception as sqs_error:
                                logger.error(f"[DELEGATION] Failed to send SQS message: {sqs_error}")
                                should_delegate = False  # Fall back to sync processing
                    else:
                        logger.error(f"[DELEGATION] Failed to save context, falling back to sync processing")
                        should_delegate = False

                except Exception as e:
                    logger.error(f"[DELEGATION] Delegation failed, falling back to sync processing: {e}")
                    should_delegate = False

            if not should_delegate:
                # ========== SYNCHRONOUS PROCESSING (CURRENT BEHAVIOR) ==========
                logger.debug(f"[DELEGATION] Using synchronous processing (estimated {estimated_minutes:.1f}min <= {MAX_SYNC_INVOCATION_TIME_MINUTES:.1f}min)")

            try:
                # Check if this is async completion with pre-loaded results
                if event.get('_skip_validation_call') and event.get('_validation_results_from_async'):
                    logger.debug("[ASYNC_COMPLETION] Using pre-loaded validation results from async completion")
                    validation_results = event['_validation_results_from_async']
                else:
                    # Use the same pre-generated payload for sync validation
                    logger.debug("[SYNC_VALIDATION] Using pre-generated validation payload for sync processing")

                    # Call validation lambda directly with the complete payload
                    import boto3
                    lambda_client = boto3.client('lambda')

                    # Use environment variable directly to get correct lambda name
                    validator_lambda_name = os.environ.get('VALIDATOR_LAMBDA_NAME', 'perplexity-validator')
                    logger.debug(f"[SYNC_VALIDATION] Invoking validation lambda: {validator_lambda_name}")

                    # Add critical fields to sync payload (same as async payload)
                    sync_payload = complete_validation_payload.copy()
                    sync_payload.update({
                        "run_key": run_key,
                        "S3_UNIFIED_BUCKET": S3_UNIFIED_BUCKET,
                        "VALIDATOR_LAMBDA_NAME": VALIDATOR_LAMBDA_NAME,
                        "email": email,
                        "config_version": config_version,
                        "results_path": results_path
                    })

                    # Debug payload before sending
                    config_targets_count = len(sync_payload.get('config', {}).get('validation_targets', []))
                    rows_count = len(sync_payload.get('validation_data', {}).get('rows', []))
                    logger.debug(f"[SYNC_VALIDATION] Payload debug - config targets: {config_targets_count}, rows: {rows_count}")
                    logger.debug(f"[SYNC_VALIDATION] Added fields - email: {email}, config_version: {config_version}, results_path: {results_path}")

                    response = lambda_client.invoke(
                        FunctionName=validator_lambda_name,
                        InvocationType='RequestResponse',
                        Payload=json.dumps(sync_payload, default=str)
                    )

                    validation_results = json.loads(response['Payload'].read().decode('utf-8'))

                # Check if sync validation returned incomplete status and should be delegated to async
                if validation_results and validation_results.get('status') == 'incomplete':
                    logger.warning(f"[SYNC_INCOMPLETE] Synchronous validation incomplete: {validation_results.get('completeness_reason', 'Unknown reason')}")
                    # This case should trigger async delegation through the existing Smart Delegation System logic
                    # We'll handle this as a validation error and let the error handling decide what to do
                    raise Exception(f"Sync validation incomplete: {validation_results.get('completeness_reason', 'Unknown reason')}")

                # Validate that we got complete results
                if not validation_results or not isinstance(validation_results, dict):
                    raise Exception("Validation lambda returned empty or invalid results")

                # Check for expected data structure - validation lambda returns {statusCode, body: {data: {rows: ...}}}
                if 'body' not in validation_results:
                    raise Exception("Validation lambda response missing body field")

                body = validation_results.get('body', {})
                if 'data' not in body:
                    raise Exception("Validation lambda response body missing data field")

                data = body.get('data', {})
                if 'rows' not in data:
                    raise Exception("Validation lambda response data missing rows field")

                # Check if we have meaningful data - enhanced detection for 413 errors
                val_results = data.get('rows', {})
                results_count = len(val_results) if isinstance(val_results, dict) else 0

                logger.debug(f"[VALIDATION_DEBUG] val_results type: {type(val_results)}, content preview: {str(val_results)[:200] if val_results else 'Empty'}")
                logger.debug(f"[VALIDATION_DEBUG] rows_to_process: {rows_to_process}, results_count: {results_count}")

                # Primary check: completely empty results for non-empty dataset
                if not val_results and rows_to_process > 0:
                    raise Exception("Validation lambda returned completely empty results for non-empty dataset")

                # Secondary check: zero results count (catches 413 payload errors)
                if rows_to_process > 0 and results_count == 0:
                    raise Exception(f"Validation lambda processed {rows_to_process} rows but returned 0 results - suspected payload size limit exceeded (HTTP 413)")

                # Tertiary check: suspiciously small results relative to input
                if rows_to_process > 10 and results_count < (rows_to_process * 0.1):  # Less than 10% results returned
                    logger.warning(f"[VALIDATION_WARNING] Suspiciously low result count: {results_count} results from {rows_to_process} rows")

                logger.debug(f"[VALIDATION_SUCCESS] Received complete validation results: {results_count} rows processed")

            except Exception as validation_error:
                error_msg = str(validation_error)
                logger.error(f"[VALIDATION_FAILURE] Critical validation error: {error_msg}")

                # Determine error type for better classification
                error_type = "Unknown Error"
                if "response too large" in error_msg.lower() or "413" in error_msg:
                    if "suspected" in error_msg.lower():
                        error_type = "Response Too Large (413) Suspected"
                    else:
                        error_type = "Response Too Large (413)"
                elif "timeout" in error_msg.lower():
                    error_type = "Lambda Timeout"
                elif "empty or invalid results" in error_msg.lower():
                    error_type = "Empty Response"
                elif "missing" in error_msg.lower() and ("body" in error_msg.lower() or "data" in error_msg.lower() or "rows" in error_msg.lower()):
                    error_type = "Malformed Response"

                # Prepare comprehensive session data for alert
                session_data = {
                    'session_id': session_id,
                    'excel_s3_key': excel_s3_key,
                    'config_s3_key': config_s3_key,
                    'max_rows': max_rows,
                    'batch_size': batch_size,
                    'rows_to_process': rows_to_process,
                    'is_preview': is_preview,
                    'timestamp': datetime.now(timezone.utc).isoformat()
                }

                # Send alert emails
                send_validation_failure_alert(session_id, email, error_type, error_msg, session_data)

                # Update run status to failed
                update_run_status_for_session(
                    status='FAILED',
                    run_type='Validation',
                    verbose_status=f'Validation processing failed: {error_type}',
                    percent_complete=100
                )

                # Send failure notification via WebSocket
                _send_websocket_message(session_id, {
                    'type': 'error',  # Use generic 'error' type that frontend definitely handles
                    'session_id': session_id,
                    'progress': 100,
                    'status': f'❌ Validation failed: {error_type}',
                    'error': f'Validation failed: {error_type}',  # Full error message for frontend
                    'error_type': error_type.lower().replace(' ', '_'),  # Specific error type for debugging
                    'message': 'Technical issue encountered. Our team has been notified and will resolve this promptly.',
                    'validation_failed': True  # Additional flag for validation-specific handling if needed
                })

                # Return error response - DO NOT continue processing
                return {
                    'statusCode': 500,
                    'body': json.dumps({
                        'status': 'failed',
                        'error': error_type,
                        'message': 'Validation processing failed. Alerts have been sent.',
                        'session_id': session_id
                    })
                }

            # Only continue if validation was successful
        
        # Initialize QC variables outside conditional blocks to ensure they're always available
        qc_results = {}
        qc_metrics = {}

        # After invoke_validator_lambda returns
        if validation_results:
            # Extract processed rows count from the correct structure
            body = validation_results.get('body', {})
            data = body.get('data', {})
            processed_rows_count = len(data.get('rows', {}))
            run_type_processing = "Preview" if is_preview else "Validation"
            update_run_status_for_session( status='PROCESSING', run_type=run_type_processing, processed_rows=processed_rows_count)

        # Check for results based on the correct response structure
        logger.error(f"[HAS_RESULTS_DEBUG] validation_results exists: {validation_results is not None}")
        if validation_results:
            logger.error(f"[HAS_RESULTS_DEBUG] validation_results type: {type(validation_results)}")
            logger.error(f"[HAS_RESULTS_DEBUG] validation_results keys: {list(validation_results.keys()) if isinstance(validation_results, dict) else 'not dict'}")
            logger.error(f"[HAS_RESULTS_DEBUG] 'body' in validation_results: {'body' in validation_results if isinstance(validation_results, dict) else 'not dict'}")

            if 'body' in validation_results:
                body_check = validation_results.get('body', {})
                logger.error(f"[HAS_RESULTS_DEBUG] body type: {type(body_check)}")
                logger.error(f"[HAS_RESULTS_DEBUG] body keys: {list(body_check.keys()) if isinstance(body_check, dict) else 'not dict'}")
                logger.error(f"[HAS_RESULTS_DEBUG] 'data' in body: {'data' in body_check if isinstance(body_check, dict) else 'not dict'}")

                if 'data' in body_check:
                    data_check = body_check.get('data', {})
                    logger.error(f"[HAS_RESULTS_DEBUG] data type: {type(data_check)}")
                    logger.error(f"[HAS_RESULTS_DEBUG] data keys: {list(data_check.keys()) if isinstance(data_check, dict) else 'not dict'}")
                    logger.error(f"[HAS_RESULTS_DEBUG] 'rows' in data: {'rows' in data_check if isinstance(data_check, dict) else 'not dict'}")

        has_results = (validation_results and
                      'body' in validation_results and
                      'data' in validation_results.get('body', {}) and
                      'rows' in validation_results.get('body', {}).get('data', {}))

        logger.error(f"[HAS_RESULTS_DEBUG] Final has_results evaluation: {has_results}")

        if has_results:
            body = validation_results.get('body', {})
            data = body.get('data', {})
            real_results = data.get('rows', {})
            total_rows = body.get('metadata', {}).get('total_rows', 1)
            metadata = body.get('metadata', {})
            token_usage = metadata.get('token_usage', {})

            # Debug the full validation_results structure to find QC data
            logger.error(f"[VALIDATION_STRUCTURE] validation_results top-level keys: {list(validation_results.keys()) if isinstance(validation_results, dict) else 'not dict'}")
            logger.error(f"[VALIDATION_STRUCTURE] body keys: {list(body.keys()) if isinstance(body, dict) else 'not dict'}")
            logger.error(f"[VALIDATION_STRUCTURE] data keys: {list(data.keys()) if isinstance(data, dict) else 'not dict'}")
            logger.error(f"[VALIDATION_STRUCTURE] metadata keys: {list(metadata.keys()) if isinstance(metadata, dict) else 'not dict'}")

            # Check for QC data in all possible locations
            qc_in_validation_results = validation_results.get('qc_results')
            qc_in_body = body.get('qc_results')
            qc_in_data = data.get('qc_results')
            qc_in_metadata = metadata.get('qc_results')

            logger.error(f"[QC_LOCATION_DEBUG] QC in validation_results: {qc_in_validation_results is not None}")
            logger.error(f"[QC_LOCATION_DEBUG] QC in body: {qc_in_body is not None}")
            logger.error(f"[QC_LOCATION_DEBUG] QC in data: {qc_in_data is not None}")
            logger.error(f"[QC_LOCATION_DEBUG] QC in metadata: {qc_in_metadata is not None}")

            # Extract QC data - check both data level and top level
            logger.error(f"[QC_EXTRACT_VALIDATION] data structure keys: {list(data.keys()) if isinstance(data, dict) else 'not dict'}")
            logger.error(f"[QC_EXTRACT_VALIDATION] Looking for qc_results in data: {'qc_results' in data if isinstance(data, dict) else 'N/A'}")

            # QC data can be in data section (as expected) or top level (actual location)
            qc_results = data.get('qc_results', {})
            qc_metrics = data.get('qc_metrics', {})

            # If not found in data section, check top level of validation_results
            if not qc_results and validation_results:
                qc_results = validation_results.get('qc_results', {})
                qc_metrics = validation_results.get('qc_metrics', {})
                logger.error(f"[QC_EXTRACT_VALIDATION] QC data not in data section, found at top level: {bool(qc_results)}")

            logger.error(f"[QC_EXTRACT_VALIDATION] Final extracted qc_results: {type(qc_results)} with {len(qc_results) if isinstance(qc_results, dict) else 'N/A'} items")
            logger.error(f"[QC_EXTRACT_VALIDATION] Final extracted qc_metrics: {type(qc_metrics)} with content: {qc_metrics}")
            if qc_results:
                logger.debug(f"[QC_MERGE_FULL] Merging QC data into full validation results for display")
                logger.debug(f"[QC_MERGE_FULL_DEBUG] QC results structure: {list(qc_results.keys())[:3]}")
                logger.debug(f"[QC_MERGE_FULL_DEBUG] real_results keys sample: {list(real_results.keys())[:3]}")

                # Create mapping from hash keys (QC) to numeric keys (validation results)
                qc_hash_keys = list(qc_results.keys())
                validation_numeric_keys = list(real_results.keys())

                # Map QC hash keys to validation numeric keys by position
                for i, qc_hash_key in enumerate(qc_hash_keys):
                    if i < len(validation_numeric_keys):
                        validation_key = validation_numeric_keys[i]
                        row_qc_data = qc_results[qc_hash_key]

                        logger.debug(f"[QC_MERGE_FULL_DEBUG] Mapping QC hash key {qc_hash_key} -> validation key {validation_key}")

                        if validation_key in real_results:
                            logger.debug(f"[QC_MERGE_FULL_DEBUG] Row {validation_key}: QC fields = {list(row_qc_data.keys())}")
                            for field_name, field_qc_data in row_qc_data.items():
                                logger.debug(f"[QC_MERGE_FULL_DEBUG] Field {field_name}: QC data keys = {list(field_qc_data.keys()) if isinstance(field_qc_data, dict) else 'not dict'}")
                                logger.debug(f"[QC_MERGE_FULL_DEBUG] Field {field_name}: qc_applied = {field_qc_data.get('qc_applied') if isinstance(field_qc_data, dict) else 'N/A'}")

                                if isinstance(field_qc_data, dict) and (field_qc_data.get('qc_applied') is True or field_qc_data.get('qc_applied') == 'Yes'):
                                    # Since QC is now comprehensive, always use QC values when available
                                    qc_entry = field_qc_data.get('qc_entry', '')
                                    qc_confidence = field_qc_data.get('qc_confidence', '')
                                    logger.debug(f"[QC_MERGE_FULL_DEBUG] Field {field_name}: has qc_entry = {bool(qc_entry)}, has qc_confidence = {bool(qc_confidence)}")

                                    # Always use QC entry and confidence when available (comprehensive QC)
                                    if qc_entry and str(qc_entry).strip():
                                        real_results[validation_key][field_name]['value'] = qc_entry
                                        logger.debug(f"[QC_MERGE_FULL] {field_name}: Using QC entry value: {qc_entry}")

                                    if qc_confidence and str(qc_confidence).strip():
                                        real_results[validation_key][field_name]['confidence_level'] = qc_confidence
                                        logger.debug(f"[QC_MERGE_FULL] {field_name}: Using QC confidence: {qc_confidence}")
            # Extract enhanced metrics from validation response
            enhanced_metrics = metadata.get('enhanced_metrics', {})
            # validation_metrics is nested inside enhanced_metrics
            validation_metrics = enhanced_metrics.get('validation_metrics', {})

            # Extract enhanced models parameter from enhanced_metrics (not top-level metadata)
            enhanced_models_parameter = enhanced_metrics.get('enhanced_models_parameter', {}) if enhanced_metrics else {}
            
            # Apply domain multiplier to raw costs and handle billing
            try:
                import sys
                import os
                sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', 'shared'))
                from dynamodb_schemas import get_domain_multiplier, track_api_usage_detailed, deduct_from_balance, check_user_balance
                from decimal import Decimal
                
                email_domain = email.split('@')[-1] if '@' in email else 'unknown'
                multiplier = float(get_domain_multiplier(email_domain))
                
                # ========== HARDENED THREE-TIER COST SYSTEM (FULL VALIDATION) ==========
                # Extract cost data from validation lambda's centralized cost calculation
                eliyahu_cost = token_usage.get('total_cost', 0.0)  # Actual cost paid (includes caching benefits)
                
                # Calculate estimated cost without caching benefit for consistency
                cost_estimated = _calculate_cost_estimated(token_usage, metadata)
                
                # Validate cost calculations
                if eliyahu_cost < 0 or cost_estimated < 0:
                    logger.error(f"[COST_ERROR] Full validation invalid negative costs - Actual: ${eliyahu_cost:.6f}, Estimated: ${cost_estimated:.6f}")
                    eliyahu_cost = max(0.0, eliyahu_cost)
                    cost_estimated = max(0.0, cost_estimated)
                
                if eliyahu_cost > cost_estimated:
                    logger.warning(f"[COST_WARNING] Full validation actual cost ${eliyahu_cost:.6f} > estimated ${cost_estimated:.6f}")
                
                # Apply hardened domain multiplier for fallback calculation
                multiplier_result = _apply_domain_multiplier_with_validation(email, cost_estimated, session_id)
                multiplier = multiplier_result['multiplier']
                
                # For full validation, use the quoted_full_cost from preview (what user was promised to pay)
                # This ensures users pay exactly what was quoted in preview, regardless of actual full validation costs
                charged_cost = multiplier_result['quoted_cost']  # Fallback if preview cost not found
                logger.debug(f"[COST_AUDIT] Full validation fallback - Actual: ${eliyahu_cost:.6f}, "
                           f"Estimated: ${cost_estimated:.6f}, Quoted fallback: ${charged_cost:.2f}")
                logger.warning(f"[BILLING_PATH] Using FALLBACK calculation instead of preview quoted cost!")
                
                try:
                    # Try to get the quoted cost from preview results stored in S3
                    # Try multiple config versions since preview and full validation might use different versions
                    session_path = storage_manager.get_session_path(email, clean_session_id)
                    logger.debug(f"BILLING DEBUG: Using email='{email}', clean_session_id='{clean_session_id}' for preview lookup")
                    
                    preview_data = None
                    preview_results_key = None
                    
                    # Try to find the latest config version first
                    latest_version = 1
                    try:
                        existing_config, latest_config_key = storage_manager.get_latest_config(email, clean_session_id)
                        if existing_config and existing_config.get('storage_metadata', {}).get('version'):
                            latest_version = existing_config['storage_metadata']['version']
                            logger.debug(f"BILLING DEBUG: Found latest config version: {latest_version}")
                    except Exception as e:
                        logger.warning(f"Could not determine latest config version: {e}")
                    
                    # Try versions from latest down to v1
                    for version in range(latest_version, 0, -1):
                        preview_results_key = f"{session_path}v{version}_results/preview_results.json"
                        logger.debug(f"BILLING DEBUG: Trying preview results at s3://{storage_manager.bucket_name}/{preview_results_key}")
                        
                        try:
                            response = storage_manager.s3_client.get_object(
                                Bucket=storage_manager.bucket_name,
                                Key=preview_results_key
                            )
                            preview_data = json.loads(response['Body'].read().decode('utf-8'))
                            logger.debug(f"BILLING DEBUG: Found preview results at version {version}")
                            break
                        except storage_manager.s3_client.exceptions.NoSuchKey:
                            logger.debug(f"BILLING DEBUG: No preview results found at version {version}")
                            continue
                    
                    if not preview_data:
                        raise Exception("No preview results found in any config version")
                    
                    # Try both locations for preview quoted cost
                    preview_quoted_cost = None
                    if preview_data:
                        # First try account_info (nested location)
                        preview_quoted_cost = preview_data.get('account_info', {}).get('quoted_validation_cost')
                        
                        # If not found, try top level (direct location)  
                        if not preview_quoted_cost:
                            preview_quoted_cost = preview_data.get('quoted_validation_cost')
                    
                    if preview_quoted_cost:
                        charged_cost = float(preview_quoted_cost)
                        logger.debug(f"[BILLING_PATH] ✅ Using preview quoted cost for full validation: ${charged_cost:.6f}")
                    else:
                        logger.warning(f"[BILLING_PATH] ❌ Preview quoted cost not found in results - preview_data keys: {list(preview_data.keys()) if preview_data else 'None'}")
                        if preview_data and 'account_info' in preview_data:
                            logger.warning(f"[BILLING_PATH] account_info keys: {list(preview_data['account_info'].keys())}")
                        logger.warning("[BILLING_PATH] Using calculated fallback cost")
                        
                except Exception as e:
                    logger.warning(f"Failed to retrieve preview quoted cost: {e}, using calculated cost")
                
                logger.debug(f"[BILLING_SUMMARY] Full validation costs:")
                logger.info(f"  - Eliyahu (actual): ${eliyahu_cost:.6f}")
                logger.info(f"  - Estimated (current): ${cost_estimated:.6f}")
                logger.info(f"  - Multiplier: {multiplier}x")
                logger.info(f"  - User Charged (before discount): ${charged_cost:.6f}")
                logger.info(f"  - Fallback would be: ${multiplier_result['quoted_cost']:.2f}")

                # ========== DISCOUNT CALCULATION FOR FULL VALIDATION ==========
                # Calculate discount - try to get from preview first, otherwise recalculate
                discount = 0.0
                effective_cost = charged_cost

                try:
                    # Try to get discount from preview data
                    if preview_data:
                        preview_discount = preview_data.get('discount', None)
                        if preview_discount is not None:
                            discount = float(preview_discount)
                            logger.info(f"[DISCOUNT] Using discount from preview: ${discount:.2f}")
                        else:
                            # Discount not in preview, recalculate
                            config_data_for_discount, _ = storage_manager.get_latest_config(email, clean_session_id)
                            config_version = config_data_for_discount.get('storage_metadata', {}).get('version', 1) if config_data_for_discount else 1
                            discount = calculate_discount(session_id, config_version, charged_cost)
                            logger.info(f"[DISCOUNT] Recalculated discount for full validation: ${discount:.2f}")
                    else:
                        # No preview data, recalculate discount
                        config_data_for_discount, _ = storage_manager.get_latest_config(email, clean_session_id)
                        config_version = config_data_for_discount.get('storage_metadata', {}).get('version', 1) if config_data_for_discount else 1
                        discount = calculate_discount(session_id, config_version, charged_cost)
                        logger.info(f"[DISCOUNT] Calculated discount (no preview data): ${discount:.2f}")

                    # Calculate effective cost after discount
                    effective_cost = max(0.0, charged_cost - discount)
                    logger.info(f"[DISCOUNT] Full validation: quoted=${charged_cost:.2f}, discount=${discount:.2f}, effective=${effective_cost:.2f}")

                except Exception as e:
                    logger.error(f"[DISCOUNT_ERROR] Failed to calculate discount for full validation: {e}")
                    discount = 0.0
                    effective_cost = charged_cost

                # Update charged_cost to effective_cost for actual billing
                charged_amount = effective_cost
                logger.info(f"BILLING DEBUG: is_preview={is_preview}, charged_cost={charged_cost:.2f}, discount={discount:.2f}, effective_cost={effective_cost:.2f}, session_id={session_id}")

                # ========== EARLY COMPLETENESS CHECK BEFORE CHARGING ==========
                # Check validation completeness BEFORE charging the user
                if not is_preview and validation_results:
                    logger.debug(f"[EARLY_COMPLETENESS_CHECK] Checking validation completeness BEFORE charging user")

                    # Extract validation results for completeness check
                    # Validator returns {statusCode, body: {data: {rows}, metadata}}
                    body = validation_results.get('body', {}) if validation_results else {}
                    data = body.get('data', {})
                    final_validation_results = data.get('rows', {})
                    expected_row_count = total_rows_in_file
                    actual_results_count = len(final_validation_results) if isinstance(final_validation_results, dict) else 0

                    logger.debug(f"[EARLY_COMPLETENESS_CHECK] Expected rows: {expected_row_count}, Actual results: {actual_results_count}")
                    logger.debug(f"[EARLY_COMPLETENESS_CHECK] validation_results keys: {list(validation_results.keys()) if validation_results else 'None'}")
                    logger.debug(f"[EARLY_COMPLETENESS_CHECK] body keys: {list(body.keys()) if body else 'None'}")

                    # Check for basic completeness
                    is_complete = True
                    completeness_issues = []

                    # Check row count completeness
                    if actual_results_count < expected_row_count:
                        is_complete = False
                        completeness_issues.append(f"Missing results: {actual_results_count}/{expected_row_count} rows")

                    # Check for meaningful data in results
                    if actual_results_count == 0:
                        is_complete = False
                        completeness_issues.append("No validation results found")

                    # Check metadata completeness
                    if validation_results:
                        metadata = body.get('metadata', {})
                        if not metadata or not isinstance(metadata, dict):
                            is_complete = False
                            completeness_issues.append("Missing or invalid metadata")

                    # If validation is incomplete, stop here BEFORE charging
                    if not is_complete:
                        logger.error(f"[EARLY_COMPLETENESS_CHECK] ❌ Validation results are INCOMPLETE: {'; '.join(completeness_issues)}")
                        logger.error(f"[EARLY_COMPLETENESS_CHECK] 🚫 STOPPING - will NOT charge user or send email")

                        # Send error message to WebSocket
                        _send_websocket_message_deduplicated(session_id, {
                            'type': 'validation_error',
                            'status': 'FAILED',
                            'error': f"Validation incomplete: {'; '.join(completeness_issues)}",
                            'session_id': session_id
                        }, "validation_incomplete_error")

                        # Update DynamoDB status to FAILED
                        update_run_status(
                            session_id=session_id,
                            run_key=run_key,
                            status='FAILED',
                            error_message=f"Validation incomplete: {'; '.join(completeness_issues)}",
                            verbose_status="Validation failed - incomplete results"
                        )

                        # Return early to prevent charging and email
                        return {
                            'statusCode': 500,
                            'body': json.dumps({
                                'error': 'Validation incomplete',
                                'details': completeness_issues,
                                'session_id': session_id
                            })
                        }
                    else:
                        logger.debug(f"[EARLY_COMPLETENESS_CHECK] ✅ Validation results are COMPLETE - proceeding with billing")

                # Initialize billing variables for later use (after email success)
                initial_balance = check_user_balance(email)
                final_balance = initial_balance
                balance_error_occurred = False
                # CRITICAL FIX: Use charged_cost (actual amount) instead of 0
                charged_amount = charged_cost
                logger.debug(f"[BILLING_INIT] charged_amount set to charged_cost: ${charged_amount:.2f}")
                        
                
            except Exception as e:
                logger.error(f"Error applying domain multiplier in full validation: {e}")
                multiplier = 1.0
                charged_cost = eliyahu_cost
            
            if is_preview:
                logger.info(f"Got preview validation results for {session_id}, storing for polling.")
                
                # Extract metrics for this section  
                total_cost = charged_cost  # Use charged cost for display
                total_tokens = token_usage.get('total_tokens', 0)
                total_rows_processed = validation_results.get('total_processed_rows', len(real_results) if real_results else 0)
                
                # Create the markdown table for the response
                markdown_table = create_markdown_table_from_results(real_results, 3, actual_config_s3_key, S3_UNIFIED_BUCKET, qc_results)
                
                # --- Start of Complex Estimations Logic ---
                
                # Get actual batch size used during validation (same logic as preview)
                per_model_batch_stats = metadata.get('per_model_batch_stats', {})
                effective_batch_size = 3  # Small batch for testing continuations
                
                if per_model_batch_stats:
                    actual_batch_sizes = per_model_batch_stats.get('model_batch_sizes', {})
                    if actual_batch_sizes:
                        # Use minimum batch size (what the validator actually used)
                        effective_batch_size = min(actual_batch_sizes.values())
                        logger.info(f"📊 Using actual batch size from full validation: {effective_batch_size} (from models: {actual_batch_sizes})")
                    else:
                        logger.warning(f"📊 per_model_batch_stats available but no model_batch_sizes found")
                elif batch_size and batch_size > 0:
                    effective_batch_size = batch_size
                    logger.info(f"📊 Using configured batch size for full validation: {effective_batch_size}")
                else:
                    logger.warning(f"📊 No batch size data available for full validation, using default: {effective_batch_size}")
                
                # Use batch timing info from the validator if available, otherwise fallback
                batch_timing = metadata.get('batch_timing', {})
                validator_processing_time = metadata.get('processing_time', 0.0)
                time_per_batch = 20.0 # Default fallback
                time_per_row = 4.0 # Default fallback

                # Extract timing data from enhanced metrics or use fallback
                processing_time = 0.0  # Initialize
                
                logger.debug(f"[TIMING_DEBUG] batch_timing available: {bool(batch_timing)}, validator_processing_time: {validator_processing_time}")
                if batch_timing:
                    logger.debug(f"[TIMING_DEBUG] batch_timing keys: {list(batch_timing.keys())}")
                
                if batch_timing and batch_timing.get('total_batch_time_seconds', 0.0) > 0:
                    processing_time = batch_timing.get('total_batch_time_seconds', 0.0)
                    time_per_batch = batch_timing.get('average_batch_time_seconds', time_per_batch)
                    time_per_row = batch_timing.get('average_time_per_row_seconds', time_per_row)
                    logger.debug(f"[TIMING_DEBUG] Using batch_timing: processing_time={processing_time:.3f}s")
                elif validator_processing_time > 0:
                    processing_time = validator_processing_time
                    if total_rows_processed > 0:
                        time_per_row = processing_time / total_rows_processed
                        time_per_batch = time_per_row * effective_batch_size
                    logger.debug(f"[TIMING_DEBUG] Using validator_processing_time: processing_time={processing_time:.3f}s")
                else:
                    # Additional fallback - calculate from start/end times if available
                    start_time_str = validation_results.get('start_time')
                    end_time_str = validation_results.get('end_time') 
                    if start_time_str and end_time_str:
                        try:
                            start_time = datetime.fromisoformat(start_time_str.replace('Z', '+00:00'))
                            end_time = datetime.fromisoformat(end_time_str.replace('Z', '+00:00'))
                            processing_time = (end_time - start_time).total_seconds()
                            if total_rows_processed > 0:
                                time_per_row = processing_time / total_rows_processed
                                time_per_batch = time_per_row * effective_batch_size
                            logger.debug(f"[TIMING_DEBUG] Calculated from start/end times: processing_time={processing_time:.3f}s")
                        except Exception as e:
                            logger.warning(f"[TIMING_DEBUG] Failed to calculate from start/end times: {e}")
                            processing_time = 0.0
                    else:
                        logger.warning(f"[TIMING_DEBUG] No valid timing data found - processing_time will be 0")
                        processing_time = 0.0
                
                # Extract per-row cost from enhanced data (not manual calculation)
                per_row_cost_full = eliyahu_cost / total_rows_processed if total_rows_processed > 0 else eliyahu_cost
                
                # Simple batch count for display
                total_batches = math.ceil(total_rows / effective_batch_size) if total_rows > 0 else 0

                preview_payload = {
                    "status": "preview_completed",
                    "session_id": session_id,
                    "reference_pin": reference_pin,
                    "markdown_table": markdown_table,
                    "total_rows": total_rows,
                    "total_processed_rows": total_rows_processed,
                    "preview_processing_time": processing_time,
                    "estimated_total_processing_time": estimated_total_time_seconds,
                    "estimated_validation_time_minutes": round(estimated_total_time_seconds / 60, 1),
                    "actual_batch_size": effective_batch_size,
                    "estimated_validation_batches": total_batches,
                    "enhanced_download_url": enhanced_download_url,  # Download link for enhanced Excel
                    "cost_estimates": {
                        "preview_cost": charged_cost,  # What user pays for preview (0)
                        "preview_tokens": total_tokens,
                        "estimated_total_tokens": estimated_total_tokens,
                        "per_row_time": time_per_row
                    },
                    "token_usage": token_usage,
                    "validation_metrics": validation_metrics
                }


            else: # Full processing
                logger.info(f"Got full validation results for {session_id}, creating enhanced ZIP.")
                
                # Send post-validation progress update - processing results (90-95% range)
                _send_websocket_message_deduplicated(session_id, {
                    'type': 'progress_update',
                    'progress': 92,
                    'message': '📊 Processing validation results and generating files...',
                    'status': 'Processing validation results and generating files...',
                    'session_id': session_id
                }, "full_validation_processing")
                
                # Load Excel content
                excel_response = s3_client.get_object(Bucket=storage_manager.bucket_name, Key=excel_s3_key)
                excel_content = excel_response['Body'].read()

                # Load or restore config_data
                if event.get('async_completion_mode') and event.get('config_data'):
                    config_data = event['config_data']
                    logger.debug(f"[ASYNC_COMPLETION] Using restored config_data from delegation context")
                else:
                    # Load config from S3 for sync processing or if not available from context
                    config_response = s3_client.get_object(Bucket=storage_manager.bucket_name, Key=config_s3_key)
                    config_data = json.loads(config_response['Body'].read().decode('utf-8'))
                    logger.info(f"Loaded config data from S3")

                # Try to get original filename from S3 object metadata first
                try:
                    s3_response = s3_client.head_object(Bucket=S3_UNIFIED_BUCKET, Key=excel_s3_key)
                    s3_metadata = s3_response.get('Metadata', {})
                    original_filename = s3_metadata.get('original_filename')

                    logger.debug(f"[FILENAME_METADATA] S3 key: {excel_s3_key}")
                    logger.debug(f"[FILENAME_METADATA] S3 metadata keys: {list(s3_metadata.keys())}")
                    logger.debug(f"[FILENAME_METADATA] original_filename from metadata: {original_filename}")

                    if original_filename:
                        input_filename = original_filename
                        logger.debug(f"[FILENAME_SOURCE] Using original filename from S3 metadata: {input_filename}")
                    else:
                        # Fallback to S3 key filename
                        input_filename = excel_s3_key.split('/')[-1]
                        logger.warning(f"[FILENAME_SOURCE] No original filename in S3 metadata, using S3 key: {input_filename}")
                except Exception as e:
                    # Fallback to S3 key filename if metadata read fails
                    input_filename = excel_s3_key.split('/')[-1]
                    logger.error(f"[FILENAME_SOURCE] Failed to read S3 metadata: {e}, using S3 key: {input_filename}")
                
                # Try to get the config lambda filename from metadata, fallback to S3 key
                config_filename = config_s3_key.split('/')[-1]
                if 'generation_metadata' in config_data:
                    # First try config_lambda_filename (set by interface lambda)
                    if 'config_lambda_filename' in config_data['generation_metadata']:
                        config_filename = config_data['generation_metadata']['config_lambda_filename']
                        logger.info(f"Using config lambda filename from metadata: {config_filename}")
                    # Then try saved_filename (set by config lambda)
                    elif 'saved_filename' in config_data['generation_metadata']:
                        config_filename = config_data['generation_metadata']['saved_filename']
                        logger.info(f"Using saved filename from metadata: {config_filename}")

                # Get structured table data for enhanced Excel creation
                # Check if table_data was restored from async completion context
                if event.get('async_completion_mode') and event.get('table_data'):
                    table_data = event['table_data']
                    logger.debug(f"[ASYNC_COMPLETION] Using restored table_data from delegation context")
                else:
                    # Parse table data for sync processing or if not available from context
                    try:
                        from shared_table_parser import S3TableParser
                        table_parser = S3TableParser()
                        table_data = table_parser.parse_s3_table(storage_manager.bucket_name, excel_s3_key, extract_formulas=True)
                        logger.info(f"Parsed table data for enhanced Excel creation: {type(table_data)}")
                    except Exception as e:
                        logger.error(f"Failed to parse table data for Excel: {e}")
                        table_data = None

                # Determine config version and ID BEFORE creating enhanced Excel
                # This fixes "local variable 'config_version' referenced before assignment" error
                config_version = 1
                config_id = "unknown"
                try:
                    existing_config, latest_config_key = storage_manager.get_latest_config(email, clean_session_id)
                    if existing_config and existing_config.get('storage_metadata', {}):
                        config_version = existing_config['storage_metadata'].get('version', 1)
                        config_id = existing_config['storage_metadata'].get('config_id', 'unknown')
                    elif latest_config_key:
                        # Fallback: try to extract version from filename
                        config_filename = latest_config_key.split('/')[-1]
                        if config_filename.startswith('config_v') and '_' in config_filename:
                            version_part = config_filename.split('_')[1]  # config_v{N}_{source}.json
                            if version_part.startswith('v'):
                                config_version = int(version_part[1:])
                except Exception as e:
                    logger.warning(f"Could not determine config version for full results: {e}")

                # Create enhanced Excel directly (no ZIP needed)
                enhanced_excel_content = None
                if table_data and EXCEL_ENHANCEMENT_AVAILABLE:
                    try:
                        validated_sheet = table_data.get('metadata', {}).get('sheet_name') if isinstance(table_data, dict) else None

                        # Create validation_results structure that includes QC data at top level for Excel function
                        excel_validation_results = {
                            'validation_results': real_results,  # The actual row data
                            'qc_results': qc_results,  # QC data extracted earlier
                            'qc_metrics': qc_metrics   # QC metrics extracted earlier
                        }

                        logger.debug(f"[QC_EXCEL_FULL] Passing QC data to Excel: {len(qc_results) if qc_results else 0} QC rows")
                        logger.error(f"[QC_SCOPE_DEBUG] qc_results defined: {qc_results is not None}, has data: {bool(qc_results)}")
                        logger.error(f"[QC_SCOPE_DEBUG] qc_metrics defined: {qc_metrics is not None}, has data: {bool(qc_metrics)}")
                        logger.error(f"[QC_SCOPE_DEBUG] qc_results type: {type(qc_results)}, content: {qc_results}")
                        if qc_results:
                            logger.error(f"[QC_SCOPE_DEBUG] QC results keys sample: {list(qc_results.keys())[:3]}")
                        else:
                            logger.error(f"[QC_SCOPE_DEBUG] QC results is empty or None")

                        # Debug the exact structure being passed to Excel
                        logger.error(f"[QC_PASS_DEBUG] excel_validation_results structure: {list(excel_validation_results.keys())}")
                        logger.error(f"[QC_PASS_DEBUG] excel_validation_results['qc_results'] type: {type(excel_validation_results.get('qc_results'))}")
                        logger.error(f"[QC_PASS_DEBUG] excel_validation_results['qc_results'] content: {excel_validation_results.get('qc_results')}")

                        excel_buffer = create_qc_enhanced_excel_for_interface(
                            table_data, excel_validation_results, config_data, session_id,
                            validated_sheet_name=validated_sheet
                        )
                        if excel_buffer:
                            enhanced_excel_content = excel_buffer.getvalue()
                            logger.info(f"Created enhanced Excel: {len(enhanced_excel_content)} bytes")

                            # Save enhanced Excel to S3 for potential future use
                            try:
                                session_path = storage_manager.get_session_path(email, clean_session_id)
                                enhanced_excel_key = f"{session_path}v{config_version}_results/enhanced_validation.xlsx"
                                s3_client.put_object(
                                    Bucket=storage_manager.bucket_name,
                                    Key=enhanced_excel_key,
                                    Body=enhanced_excel_content,
                                    ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                                logger.info(f"Saved enhanced Excel to S3: {enhanced_excel_key}")
                            except Exception as save_e:
                                logger.warning(f"Failed to save enhanced Excel to S3: {save_e}")
                        else:
                            logger.error("Enhanced Excel creation failed - excel_buffer is None")
                    except Exception as e:
                        logger.error(f"Error creating enhanced Excel: {str(e)}")
                        import traceback
                        logger.error(traceback.format_exc())

                # Note: config_version and config_id were already determined above before creating enhanced Excel
                # This prevents "local variable 'config_version' referenced before assignment" error

                # Store validation results in versioned folder
                validation_data = {
                    'validation_results': real_results,
                    'total_rows': total_rows,
                    'metadata': metadata,
                    'session_id': session_id,
                    'reference_pin': reference_pin
                }
                
                result = storage_manager.store_results(
                    email, clean_session_id, config_version, validation_data, 'validation'
                )
                
                # Update session info with account data for full validation
                account_info = {
                    'initial_balance': float(initial_balance) if initial_balance else 0,
                    'final_balance': float(final_balance) if final_balance else 0,
                    'amount_charged': float(charged_amount),
                    'domain_multiplier': float(multiplier),
                    'eliyahu_cost': float(eliyahu_cost),
                    # Removed preview_abandoned and insufficient_balance_encountered - unnecessary fields
                    'processing_type': 'full_validation'
                }
                _update_session_info_with_account_data(email, clean_session_id, account_info)
                
                if result['success']:
                    logger.info(f"Full validation results stored in versioned folder: {result['s3_key']}")
                    validation_results_path = result['s3_key']
                else:
                    logger.error(f"Failed to store full validation results: {result.get('error')}")
                    validation_results_path = None
                
                # Store enhanced Excel directly (no ZIP extraction needed)
                enhanced_excel_path = None
                if enhanced_excel_content:
                    try:
                        enhanced_result = storage_manager.store_enhanced_files(
                            email, clean_session_id, config_version,
                            enhanced_excel_content, None  # No summary text needed
                        )

                        if enhanced_result['success']:
                            logger.info(f"Stored enhanced Excel: {enhanced_result['stored_files']}")
                            # Extract enhanced Excel path if available
                            if enhanced_result.get('stored_files'):
                                for file_path in enhanced_result['stored_files']:
                                    if 'enhanced.xlsx' in file_path:
                                        enhanced_excel_path = file_path
                                        break
                        else:
                            logger.error(f"Failed to store enhanced Excel: {enhanced_result.get('error')}")
                    except Exception as e:
                        logger.error(f"Failed to store enhanced Excel: {e}")
                
                # Update session tracking with validation results and enhanced Excel paths
                if validation_results_path:
                    try:
                        success = storage_manager.update_session_results(
                            email=email,
                            session_id=clean_session_id,
                            operation_type="validation",
                            config_id=config_id,
                            version=config_version,
                            run_key=run_key,
                            results_path=validation_results_path,
                            enhanced_excel_path=enhanced_excel_path
                        )
                        
                        if success:
                            logger.info(f"✅ Updated session_info.json with validation completion")
                        else:
                            logger.error(f"❌ Failed to update session_info.json with validation completion")
                    except Exception as e:
                        logger.error(f"Failed to update validation session tracking: {e}")
                else:
                    logger.warning(f"No validation results path available, skipping session tracking")

                if EMAIL_SENDER_AVAILABLE and email_address:
                    # Calculate summary data for the email
                    # First, determine which fields are ID fields (should not be counted as "validated")
                    id_fields = set()
                    if config_data and 'validation_targets' in config_data:
                        for target in config_data['validation_targets']:
                            if target.get('importance', '').upper() in ['ID', 'IGNORED']:
                                id_fields.add(target.get('column'))
                    
                    validated_fields = set()  # Only count fields that were actually validated
                    confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
                    original_confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
                    
                    for row_data in real_results.values():
                        for field_name, field_data in row_data.items():
                            if isinstance(field_data, dict):
                                # Only count as validated if it's not an ID/IGNORED field
                                if field_name not in id_fields:
                                    validated_fields.add(field_name)
                                
                                # Count updated confidence levels (only for validated fields)
                                if 'confidence_level' in field_data and field_name not in id_fields:
                                    conf_level = field_data.get('confidence_level', 'UNKNOWN')
                                    if conf_level in confidence_counts:
                                        confidence_counts[conf_level] += 1
                                
                                # Count original confidence levels (only for validated fields)
                                if 'original_confidence' in field_data and field_name not in id_fields:
                                    original_conf = field_data.get('original_confidence')
                                    if original_conf and str(original_conf).upper() in original_confidence_counts:
                                        original_confidence_counts[str(original_conf).upper()] += 1
                    
                    summary_data = {
                        'total_rows': len(real_results) if real_results else 0,
                        'fields_validated': list(validated_fields),  # Only count actually validated fields (exclude ID/IGNORED)
                        'confidence_distribution': confidence_counts,
                        'original_confidence_distribution': original_confidence_counts
                    }
                    
                    logger.info(f"Email summary: {len(validated_fields)} validated fields (excluding {len(id_fields)} ID/ignored fields): {sorted(validated_fields)}")
                    
                    # Enhanced Excel was already created above in the sync validation flow
                    # No need to load from S3 since we have it in memory
                    logger.info(f"Using enhanced Excel created during validation: {len(enhanced_excel_content) if enhanced_excel_content else 0} bytes")
                    
                    # Ensure enhanced_excel_content is bytes or None (not other types)
                    safe_enhanced_excel_content = enhanced_excel_content if isinstance(enhanced_excel_content, (bytes, type(None))) else None
                    
                    # Extract API call counts from enhanced_metrics (authoritative source)
                    # FIX: Use aggregated_metrics.providers instead of token_usage.by_provider
                    perplexity_calls = 0
                    anthropic_calls = 0

                    if enhanced_metrics and enhanced_metrics.get('aggregated_metrics'):
                        providers = enhanced_metrics['aggregated_metrics'].get('providers', {})
                        perplexity_calls = providers.get('perplexity', {}).get('calls', 0)
                        anthropic_calls = providers.get('anthropic', {}).get('calls', 0)
                        logger.debug(f"[API_CALLS_EXTRACT] From enhanced_metrics: Perplexity={perplexity_calls}, Anthropic={anthropic_calls}")
                    else:
                        # Fallback to token_usage if enhanced_metrics not available
                        by_provider = token_usage.get('by_provider', {})
                        perplexity_calls = by_provider.get('perplexity', {}).get('calls', 0)
                        anthropic_calls = by_provider.get('anthropic', {}).get('calls', 0)
                        logger.warning(f"[API_CALLS_EXTRACT] Falling back to token_usage: Perplexity={perplexity_calls}, Anthropic={anthropic_calls}")

                    # Extract QC call counts from already-extracted qc_metrics variable (defined around line 3416)
                    # Using the already-extracted variable instead of re-fetching to avoid stale data
                    qc_calls = 0
                    if qc_metrics and isinstance(qc_metrics, dict):
                        qc_calls = qc_metrics.get('total_qc_calls', 0)
                        logger.debug(f"[RECEIPT_QC_CALLS] Extracted QC calls for receipt: {qc_calls}")
                    else:
                        logger.warning(f"[RECEIPT_QC_CALLS] qc_metrics not available or not a dict: {type(qc_metrics)}")
                    
                    # Extract original table name (remove _input suffix if present)
                    # This is used for receipt and email display - should be clean and user-friendly
                    logger.error(f"[FILENAME_CLEAN_START] input_filename: '{input_filename}'")

                    original_table_name = input_filename
                    if input_filename and '_input' in input_filename:
                        original_table_name = input_filename.replace('_input', '')
                        logger.debug(f"[FILENAME_CLEAN] Removed _input: '{original_table_name}'")

                    # Remove excel_ prefix if present (from S3 key fallback)
                    if original_table_name and original_table_name.startswith('excel_'):
                        original_table_name = original_table_name[6:]  # Remove 'excel_' prefix
                        logger.debug(f"[FILENAME_CLEAN] Removed excel_ prefix: '{original_table_name}'")

                    # Ensure .xlsx extension
                    if original_table_name and not original_table_name.endswith('.xlsx'):
                        # Remove any existing extension and add .xlsx
                        base_name = original_table_name.rsplit('.', 1)[0] if '.' in original_table_name else original_table_name
                        original_table_name = f"{base_name}.xlsx"
                        logger.debug(f"[FILENAME_CLEAN] Added .xlsx extension: '{original_table_name}'")

                    logger.error(f"[FILENAME_CLEAN_FINAL] original_table_name: '{original_table_name}'")
                    
                    # Extract config_id from config metadata (will be added later after config_id is calculated)
                    # This is a placeholder - config_id will be added to billing_info after line 1404
                    
                    # Prepare billing information for receipt
                    billing_info = {
                        'amount_charged': charged_amount,
                        'eliyahu_cost': eliyahu_cost,
                        'actual_cost': eliyahu_cost,  # Add actual cost from DynamoDB
                        'multiplier': multiplier,
                        'initial_balance': float(initial_balance) if initial_balance else 0,
                        'final_balance': float(final_balance) if final_balance else 0,
                        # Add receipt-specific data
                        'perplexity_api_calls': perplexity_calls,
                        'anthropic_api_calls': anthropic_calls,
                        'qc_api_calls': qc_calls,  # Add QC calls for receipt
                        'rows_processed': processed_rows_count,
                        'table_name': original_table_name,
                        'columns_validated_count': len(summary_data.get('fields_validated', []))
                    }
                    
                    # Store enhanced Excel in versioned results folder and create download link
                    enhanced_download_url = None
                    if safe_enhanced_excel_content:
                        try:
                            # Get version from config
                            config_version = config_data.get('storage_metadata', {}).get('version', 1)
                            enhanced_filename = f"{os.path.splitext(input_filename)[0]}_v{config_version}_full_enhanced.xlsx"
                            
                            # Store enhanced Excel in versioned results folder
                            enhanced_result = storage_manager.store_enhanced_files(
                                email, clean_session_id, config_version, 
                                safe_enhanced_excel_content, None
                            )
                            
                            if enhanced_result['success']:
                                logger.info(f"Enhanced Excel stored in results folder for full validation: {enhanced_result['stored_files']}")
                                
                                # Also create public download link for immediate download
                                enhanced_download_url = storage_manager.create_public_download_link(
                                    safe_enhanced_excel_content, 
                                    enhanced_filename,
                                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                                logger.info(f"Created enhanced Excel download link for full results: {enhanced_download_url}")
                            else:
                                logger.error(f"Failed to store enhanced Excel in results folder for full validation: {enhanced_result.get('error')}")
                                
                        except Exception as e:
                            logger.error(f"Failed to store enhanced Excel for full validation: {e}")
                    
                    # Extract config_id from config metadata
                    config_id = config_data.get('storage_metadata', {}).get('config_id', 'N/A')
                    
                    # Add config_id to billing_info for receipt
                    billing_info['config_id'] = config_id
                    
                    # Get actual processing time from DynamoDB run record instead of metadata
                    actual_processing_time = 0
                    try:
                        from dynamodb_schemas import get_run_status
                        run_status = get_run_status(session_id, run_key)
                        if run_status:
                            # Use actual processing time if available, otherwise calculate from start/end times
                            if 'actual_processing_time_seconds' in run_status:
                                actual_processing_time = float(run_status['actual_processing_time_seconds'])
                            elif 'start_time' in run_status and 'end_time' in run_status:
                                start_time = datetime.fromisoformat(run_status['start_time'].replace('Z', '+00:00'))
                                end_time = datetime.fromisoformat(run_status['end_time'].replace('Z', '+00:00'))
                                actual_processing_time = (end_time - start_time).total_seconds()
                        logger.info(f"Using actual processing time for email: {actual_processing_time:.2f} seconds")
                    except Exception as e:
                        logger.warning(f"Could not get actual processing time for email, using metadata fallback: {e}")
                        actual_processing_time = metadata.get('processing_time', 0)
                    
                    # CRITICAL: Check for validation failures BEFORE sending success email
                    validation_failed = event.get('_validation_failed', False)
                    failure_reason = event.get('_validation_failure_reason', 'Unknown failure')

                    if validation_failed:
                        logger.error(f"[VALIDATION_FAILURE_DETECTED] Validation failed: {failure_reason}")
                        logger.error(f"[VALIDATION_FAILURE_DETECTED] Will NOT send success email or charge customer")

                        # Send failure notification via WebSocket
                        _send_websocket_message_deduplicated(session_id, {
                            'type': 'validation_failed',
                            'session_id': session_id,
                            'progress': 100,
                            'status': f'[ERROR] Validation failed: {failure_reason}',
                            'error': failure_reason
                        })

                        # Mark status as FAILED in DynamoDB
                        update_run_status_for_session(
                            status='FAILED',
                            error_message=failure_reason,
                            verbose_status=f"[ERROR] Validation failed: {failure_reason}",
                            validation_incomplete=True
                        )

                        # Send failure email to user (not success email!)
                        try:
                            from email_sender import send_validation_failure_alert
                            send_validation_failure_alert(
                                session_id=session_id,
                                email=email_address,
                                error_type='VALIDATION_FAILURE',
                                error_msg=failure_reason,
                                session_data={
                                    'session_id': session_id,
                                    'reference_pin': reference_pin,
                                    'email': email_address
                                }
                            )
                            logger.debug(f"[FAILURE_EMAIL] Sent failure notification to {email_address}")
                        except Exception as email_error:
                            logger.error(f"[FAILURE_EMAIL] Failed to send failure email: {email_error}")

                        # Early return - do not proceed with success email or billing
                        logger.error(f"[VALIDATION_FAILURE_DETECTED] Exiting without sending success email or billing")
                        # Continue to completion tracking but skip email/billing
                    else:
                        # Validation succeeded - proceed with success email
                        logger.debug(f"[VALIDATION_SUCCESS] Validation succeeded - proceeding with success email and billing")

                        # Send email progress update - sending results (95-98% range)
                        _send_websocket_message_deduplicated(session_id, {
                            'type': 'progress_update',
                            'progress': 96,
                            'message': '[SUCCESS] Sending validation results to your email...',
                            'status': 'Sending validation results to your email...',
                            'session_id': session_id
                        }, "full_validation_email")

                        # Convert any Decimal objects to float for JSON serialization
                        def convert_decimals(obj):
                            if isinstance(obj, dict):
                                return {k: convert_decimals(v) for k, v in obj.items()}
                            elif isinstance(obj, list):
                                return [convert_decimals(item) for item in obj]
                            elif hasattr(obj, '__class__') and obj.__class__.__name__ == 'Decimal':
                                return float(obj)
                            return obj

                        safe_config_data = convert_decimals(config_data)

                        # Create clean filenames for email attachments
                        # Use original_table_name (already cleaned above) for enhanced Excel filename
                        base_table_name = os.path.splitext(original_table_name)[0] if original_table_name else "Table"
                        # Use "_Validated" suffix
                        enhanced_excel_filename_for_email = f"{base_table_name}_Validated.xlsx"
                        input_filename_for_email = original_table_name or input_filename  # Use clean name for display

                        logger.error(f"[EMAIL_FILENAMES] Base: '{base_table_name}', Enhanced: '{enhanced_excel_filename_for_email}', Input: '{input_filename_for_email}'")

                        # DEBUG: Log billing_info before sending email to verify it's populated
                        logger.error(f"[EMAIL_BILLING_DEBUG] billing_info keys: {list(billing_info.keys()) if billing_info else 'None'}")
                        if billing_info:
                            logger.error(f"[EMAIL_BILLING_DEBUG] amount_charged: {billing_info.get('amount_charged')}")
                            logger.error(f"[EMAIL_BILLING_DEBUG] perplexity_api_calls: {billing_info.get('perplexity_api_calls')}")
                            logger.error(f"[EMAIL_BILLING_DEBUG] anthropic_api_calls: {billing_info.get('anthropic_api_calls')}")
                            logger.error(f"[EMAIL_BILLING_DEBUG] qc_api_calls: {billing_info.get('qc_api_calls')}")
                            logger.error(f"[EMAIL_BILLING_DEBUG] table_name: {billing_info.get('table_name')}")

                        email_result = send_validation_results_email(
                            email_address=email_address, excel_content=excel_content,
                            config_content=json.dumps(safe_config_data, indent=2).encode('utf-8'),
                            enhanced_excel_content=safe_enhanced_excel_content,
                            input_filename=input_filename_for_email, config_filename=config_filename,
                            enhanced_excel_filename=enhanced_excel_filename_for_email,
                            session_id=session_id, summary_data=summary_data,
                            processing_time=actual_processing_time,
                            reference_pin=reference_pin, metadata=metadata, preview_email=preview_email,
                            billing_info=billing_info,
                            config_id=config_id
                        )

                        # ========== BILLING AFTER SUCCESSFUL EMAIL DELIVERY ==========
                        # Only charge the user AFTER the email has been successfully sent
                        if email_result and email_result.get('success', False):
                            logger.debug(f"[BILLING_SECURITY] [SUCCESS] Email delivered successfully - proceeding with billing")

                            # For full validation, deduct from account balance
                            # IMPORTANT: Use effective_cost (after discount) for actual billing
                            if not is_preview and effective_cost > 0:
                                logger.info(f"BILLING: Proceeding with charge for {email}: ${effective_cost:.2f} (quoted: ${charged_cost:.2f}, discount: ${discount:.2f}), run_key: {run_key}")
                                # Check if user has sufficient balance
                                current_balance = check_user_balance(email)
                                if current_balance is not None and current_balance >= Decimal(str(effective_cost)):
                                    # Deduct from balance with run_key protection against duplicate charges
                                    deduct_success = deduct_from_balance(
                                        email=email,
                                        amount=Decimal(str(effective_cost)),  # Use effective_cost (after discount)
                                        session_id=session_id,
                                        description=f"Full validation - {len(real_results) if real_results else 0} rows processed" + (f" (${discount:.2f} discount applied)" if discount > 0 else ""),
                                        raw_cost=Decimal(str(eliyahu_cost)),
                                        multiplier=Decimal(str(multiplier)),
                                        run_key=run_key  # CRITICAL: Prevents duplicate charges for same run_key
                                    )
                                    if deduct_success:
                                        final_balance = check_user_balance(email)
                                        charged_amount = effective_cost
                                        logger.debug(f"[BILLING_SUCCESS] [SUCCESS] Successfully charged ${effective_cost:.2f} - new balance: ${final_balance:.6f}")
                                        # Send balance update via WebSocket
                                        _send_balance_update(session_id, {
                                            'type': 'balance_update',
                                            'new_balance': float(final_balance) if final_balance else 0,
                                            'transaction': {
                                                'amount': -float(effective_cost),
                                                'description': f"Full validation - {len(real_results) if real_results else 0} rows processed",
                                                'eliyahu_cost': float(eliyahu_cost),
                                                'multiplier': float(multiplier),
                                                'discount': float(discount)
                                            }
                                        })
                                    else:
                                        logger.error(f"[BILLING_ERROR] [ERROR] Failed to deduct ${effective_cost:.2f} from {email} balance")
                                        balance_error_occurred = True
                                else:
                                    logger.warning(f"[BILLING_ERROR] [ERROR] Insufficient balance for {email}: {current_balance} < ${effective_cost:.2f}")
                                    balance_error_occurred = True
                            else:
                                logger.debug(f"[BILLING_SKIP] Skipping charge - is_preview={is_preview}, effective_cost={effective_cost}")

                            # Track detailed API usage for each provider
                            by_provider = token_usage.get('by_provider', {})
                            for provider, provider_usage in by_provider.items():
                                if provider_usage and isinstance(provider_usage, dict):
                                    usage_data = {
                                        'api_calls': provider_usage.get('api_calls', 0),
                                        'cached_calls': provider_usage.get('cached_calls', 0),
                                        'total_tokens': provider_usage.get('total_tokens', 0),
                                        'cost': provider_usage.get('total_cost', 0.0) * multiplier,  # Apply multiplier
                                        'eliyahu_cost': provider_usage.get('total_cost', 0.0),  # Store eliyahu cost too
                                        'multiplier_applied': multiplier
                                    }

                                    # Add provider-specific token fields
                                    if provider == 'perplexity':
                                        usage_data.update({
                                            'prompt_tokens': provider_usage.get('prompt_tokens', 0),
                                            'completion_tokens': provider_usage.get('completion_tokens', 0)
                                        })
                                    elif provider == 'anthropic':
                                        usage_data.update({
                                            'input_tokens': provider_usage.get('input_tokens', 0),
                                            'output_tokens': provider_usage.get('output_tokens', 0),
                                            'cache_tokens': provider_usage.get('cache_tokens', 0)
                                        })
                        else:
                            logger.error(f"[BILLING_SECURITY] [ERROR] Email delivery failed - BLOCKING ALL CHARGES to protect user")
                            logger.error(f"[BILLING_SECURITY] Email result: {email_result}")
                            balance_error_occurred = True  # Mark as error so no charges occur

                            # Send error notification instead of success
                            error_type = email_result.get('error', 'unknown')
                            error_message = email_result.get('message', 'Email delivery failed')

                            # Check if this is a validation failure (empty Details sheet)
                            if error_type == 'validation_failed':
                                logger.error(f"[VALIDATION_FAILURE] Enhanced Excel validation failed - will send failure notification to frontend")
                                _send_websocket_message_deduplicated(session_id, {
                                    'type': 'validation_failed',
                                    'session_id': session_id,
                                    'error': 'Validation processing error - enhanced results appear invalid',
                                    'message': 'Validation processing encountered an error. You have not been charged. Please contact support.',
                                    'status': 'FAILED'
                                }, "validation_failure_notification")
                            else:
                                # Email sending failed for other reasons
                                logger.error(f"[EMAIL_FAILURE] Email sending failed: {error_message}")
                                _send_websocket_message_deduplicated(session_id, {
                                    'type': 'email_failed',
                                    'session_id': session_id,
                                    'error': error_type,
                                    'message': 'Failed to send results email. Please contact support.',
                                    'status': 'EMAIL_FAILED'
                                }, "email_failure_notification")

                            # Skip the success completion notification below
                            logger.error(f"[COMPLETION_BLOCKED] Not sending success completion due to email/validation failure")

                    # Only send success completion if email was delivered successfully
                    if email_result and email_result.get('success', False):
                        # Send final processing progress update - finalizing (98-100% range)
                        _send_websocket_message_deduplicated(session_id, {
                            'type': 'progress_update',
                            'progress': 99,
                            'message': '✅ Finalizing download links and completion...',
                            'status': 'Finalizing download links and completion...',
                            'session_id': session_id
                        }, "full_validation_finalizing")

                        # Send final completion notification with download URLs
                        processed_rows_count = len(validation_results.get('validation_results', {}))
                        total_rows_in_file = validation_results.get('total_rows', processed_rows_count)

                        # No longer creating ZIP - enhanced Excel is stored directly in S3
                        # Download URLs will be provided for the enhanced Excel file directly
                        zip_download_url = None  # Kept for backward compatibility but set to None
                    
                        # Send completion with download URLs (consistent with preview completion)
                        completion_payload = {
                            'status': 'COMPLETED',
                            'processed_rows': processed_rows_count,
                            'total_rows': total_rows_in_file,
                            'verbose_status': 'Validation complete. Results should be in your inbox shortly.',
                            'percent_complete': 100  # Interface final processing complete: 100%
                        }
                    
                        # Add download URLs if available
                        if enhanced_download_url:
                            completion_payload['enhanced_download_url'] = enhanced_download_url
                        if zip_download_url:
                            completion_payload['download_url'] = zip_download_url
                    
                        logger.info(f"Sending completion WebSocket message with download URLs: enhanced={bool(enhanced_download_url)}, zip={bool(zip_download_url)}")
                        _send_websocket_message(session_id, completion_payload)
                    
                        # Get enhanced results S3 key (points to enhanced validation results directory)
                        enhanced_results_s3_key = results_key  # Default to ZIP file if enhanced path not found
                        try:
                            # Try to get the enhanced results directory path instead of ZIP file
                            config_version = config_data.get('storage_metadata', {}).get('version', 1)
                            session_path = storage_manager.get_session_path(email, clean_session_id)
                            enhanced_results_s3_key = f"{session_path}v{config_version}_results/validation_results_enhanced.xlsx"
                            logger.info(f"Using enhanced results S3 key: {enhanced_results_s3_key}")
                        except Exception as e:
                            logger.warning(f"Could not determine enhanced results path, using ZIP path: {e}")
                    
                        # Update status with both ZIP file and enhanced Excel download URLs
                        status_update_data = {
                            'session_id': session_id,
                            'run_key': run_key,
                            'status': 'COMPLETED', 
                            'results_s3_key': enhanced_results_s3_key  # Points to enhanced validation results directory
                        }
                        if enhanced_download_url:
                            status_update_data['enhanced_download_url'] = enhanced_download_url
                        if zip_download_url:
                            status_update_data['download_url'] = zip_download_url
                    
                        # Get and add the actual batch size used during validation
                        per_model_batch_stats = metadata.get('per_model_batch_stats', {})
                        effective_batch_size = 3  # Small batch for testing continuations
                    
                        if per_model_batch_stats:
                            actual_batch_sizes = per_model_batch_stats.get('model_batch_sizes', {})
                            if actual_batch_sizes:
                                effective_batch_size = min(actual_batch_sizes.values())
                                logger.info(f"📊 Final status update batch size from validation: {effective_batch_size}")
                        elif batch_size and batch_size > 0:
                            effective_batch_size = batch_size
                    
                        if effective_batch_size:
                            status_update_data['batch_size'] = effective_batch_size
                    
                        # Add token/cost data to runs table for full runs (similar to preview runs)
                        # This ensures CSV exports have complete token metrics for both preview and full runs
                        full_run_data = {
                            "actual_batch_size": effective_batch_size,
                            "estimated_validation_batches": math.ceil(total_rows_in_file / effective_batch_size) if effective_batch_size > 0 else 0,
                            "cost_estimates": {
                                "preview_cost": 0  # Full runs don't have preview cost
                            },
                            "token_usage": {
                                "total_tokens": token_usage.get('total_tokens', 0),
                                "by_provider": token_usage.get('by_provider', {}),
                                "api_calls": token_usage.get('api_calls', 0),
                                "cached_calls": token_usage.get('cached_calls', 0),
                                "total_cost": eliyahu_cost
                            },
                            "validation_metrics": validation_metrics
                        }
                        status_update_data['preview_data'] = full_run_data
                    
                        # Add account tracking fields
                        status_update_data['account_current_balance'] = float(final_balance) if final_balance else 0
                        status_update_data['account_sufficient_balance'] = "n/a"
                        status_update_data['account_credits_needed'] = "n/a" 
                        status_update_data['account_domain_multiplier'] = float(multiplier)
                    
                        # Create search group models tracking string - one entry per search group with validated columns count
                        search_groups_models = []
                        # validation_metrics should already be defined from metadata.get('validation_metrics', {})
                        search_groups_count = validation_metrics.get('search_groups_count', 0)
                        enhanced_context_groups = validation_metrics.get('high_context_search_groups_count', 0)
                        claude_groups = validation_metrics.get('claude_search_groups_count', 0)
                        regular_perplexity_groups = search_groups_count - enhanced_context_groups - claude_groups
                        validated_columns_count = validation_metrics.get('validated_columns_count', 0)
                    
                        # Count actual validation targets per search group from config
                        validation_targets = config_data.get('validation_targets', [])
                        columns_per_search_group = {}
                    
                        # Count validation targets assigned to each search group (excluding ID and IGNORED)
                        for target in validation_targets:
                            importance = target.get('importance', '').upper()
                            if importance not in ["ID", "IGNORED"]:
                                search_group_id = target.get('search_group', 0)
                                columns_per_search_group[search_group_id] = columns_per_search_group.get(search_group_id, 0) + 1
                    
                        # Build search group models by iterating through actual search groups in config
                        search_groups_list = config_data.get('search_groups', [])
                        anthropic_default_web_searches = config_data.get('anthropic_max_web_searches_default', 3)
                    
                        # Iterate through each search group to get exact models and settings
                        for search_group in search_groups_list:
                            search_group_id = search_group.get('group_id', 0)
                            columns_for_this_group = columns_per_search_group.get(search_group_id, 0)
                            model = search_group.get('model', 'sonar-pro')
                            search_context = search_group.get('search_context', 'low')
                        
                            # Build model display name
                            if 'claude' in model.lower() or 'anthropic' in model.lower():
                                # Get anthropic_max_web_searches from this specific search group, or fall back to default
                                max_web_searches = search_group.get('anthropic_max_web_searches', anthropic_default_web_searches)
                                if search_context == 'high':
                                    model_display = f"{model} ({max_web_searches}) (high context) X {columns_for_this_group}"
                                else:
                                    model_display = f"{model} ({max_web_searches}) X {columns_for_this_group}"
                            else:
                                # Perplexity models
                                if search_context == 'high':
                                    model_display = f"{model} (high context) X {columns_for_this_group}"
                                else:
                                    model_display = f"{model} X {columns_for_this_group}"
                        
                            search_groups_models.append(model_display)
                    
                        status_update_data['models'] = json.dumps(enhanced_models_parameter) if enhanced_models_parameter else json.dumps({})
                    
                        # Add input table name and configuration ID
                        status_update_data['input_table_name'] = input_filename
                    
                        # Get configuration ID for tracking
                        configuration_id = config_data.get('storage_metadata', {}).get('config_id', 'unknown')
                        if not configuration_id or configuration_id == 'unknown':
                            # Fallback to generation metadata
                            configuration_id = config_data.get('generation_metadata', {}).get('config_id', 'unknown')
                    
                        # If still no config ID, generate one from session and version info
                        if not configuration_id or configuration_id == 'unknown':
                            version = config_data.get('storage_metadata', {}).get('version') or config_data.get('generation_metadata', {}).get('version', 'v1')
                            configuration_id = f"{session_id}_{version}_config"
                            logger.debug(f"[CONFIG_ID] Generated configuration_id: {configuration_id}")
                    
                        status_update_data['configuration_id'] = configuration_id
                    
                        # Add total rows count
                        status_update_data['total_rows'] = total_rows_in_file
                    
                        # Update verbose status to start with 'Validation' for full validations
                        status_update_data['verbose_status'] = "Validation complete. Results should be in your inbox shortly."
                    
                        # Add consolidated fields according to new schema
                        status_update_data['run_type'] = "Validation"
                        # ========== HARDENED THREE-TIER COST FIELDS ==========
                        status_update_data['eliyahu_cost'] = eliyahu_cost  # Actual cost paid for full validation (includes caching benefits)
                        status_update_data['quoted_validation_cost'] = charged_cost  # What user would pay without discount (with multiplier, rounding, $2 min)
                        status_update_data['discount'] = discount  # Discount applied
                        # Note: effective_cost = charged_cost - discount is what user actually paid
                    
                        # NOTE: Do NOT overwrite estimated_validation_eliyahu_cost for full validations
                        # This field should preserve the preview estimate for accuracy comparison
                        # Log the comparison between preview estimate and actual full validation cost
                        try:
                            # Try to retrieve the preview estimate from the existing run record
                            run_table = boto3.resource('dynamodb', region_name=os.environ.get('AWS_REGION', 'us-east-1')).Table('perplexity-validator-runs')
                            existing_run = run_table.get_item(Key={'session_id': session_id, 'run_key': run_key})
                        
                            if 'Item' in existing_run and 'estimated_validation_eliyahu_cost' in existing_run['Item']:
                                preview_estimate = float(existing_run['Item']['estimated_validation_eliyahu_cost'])
                                actual_full_cost_estimated = cost_estimated
                                estimate_accuracy = ((preview_estimate - actual_full_cost_estimated) / preview_estimate * 100) if preview_estimate > 0 else 0
                            
                                logger.debug(f"[COST_COMPARISON] Preview estimated: ${preview_estimate:.6f} | "
                                          f"Actual full cost (no cache): ${actual_full_cost_estimated:.6f} | "
                                          f"Estimate accuracy: {estimate_accuracy:.1f}% | User charged: ${charged_cost:.2f}")
                            else:
                                logger.debug(f"[COST_COMPARISON] No preview estimate found | "
                                          f"Actual full cost (no cache): ${cost_estimated:.6f} | "
                                          f"User charged: ${charged_cost:.2f}")
                        except Exception as e:
                            logger.warning(f"[COST_COMPARISON] Could not retrieve preview estimate for comparison: {e}")
                            logger.debug(f"[COST_COMPARISON] Actual full cost (no cache): ${cost_estimated:.6f} | "
                                      f"User charged: ${charged_cost:.2f}")
                    
                        # Calculate actual validation processing time from DynamoDB record times
                        validation_processing_time = 0.0
                    
                        # Get the current run's start time from DynamoDB
                        try:
                            from dynamodb_schemas import get_run_status
                            current_run = get_run_status(session_id, run_key)
                            if current_run and 'start_time' in current_run:
                                start_time_str = current_run['start_time']
                                end_time_str = datetime.now(timezone.utc).isoformat()  # Current time as end time
                            
                                logger.debug(f"[VALIDATION_TIMING_DEBUG] DynamoDB start_time: {start_time_str}, end_time: {end_time_str}")
                            
                                start_time = datetime.fromisoformat(start_time_str.replace('Z', '+00:00'))
                                end_time = datetime.fromisoformat(end_time_str.replace('Z', '+00:00'))
                                validation_processing_time = (end_time - start_time).total_seconds()
                                logger.debug(f"[VALIDATION_TIMING_DEBUG] Calculated from DynamoDB times: validation_processing_time={validation_processing_time:.3f}s")
                            else:
                                logger.warning(f"[VALIDATION_TIMING_DEBUG] Could not get start_time from DynamoDB run record")
                        except Exception as db_error:
                            logger.warning(f"[VALIDATION_TIMING_DEBUG] Failed to get timing from DynamoDB: {db_error}")
                        
                            # Fallback to validation_results times if available
                            start_time_str = validation_results.get('start_time')
                            end_time_str = validation_results.get('end_time') 
                        
                            logger.debug(f"[VALIDATION_TIMING_DEBUG] Fallback start_time: {start_time_str}, end_time: {end_time_str}")
                        
                            if start_time_str and end_time_str:
                                try:
                                    start_time = datetime.fromisoformat(start_time_str.replace('Z', '+00:00'))
                                    end_time = datetime.fromisoformat(end_time_str.replace('Z', '+00:00'))
                                    validation_processing_time = (end_time - start_time).total_seconds()
                                    logger.debug(f"[VALIDATION_TIMING_DEBUG] Calculated from start/end times: validation_processing_time={validation_processing_time:.3f}s")
                                except Exception as e:
                                    logger.warning(f"[VALIDATION_TIMING_DEBUG] Failed to calculate from start/end times: {e}")
                                    # Fallback to other timing sources only if start/end calculation fails
                                    batch_timing = metadata.get('batch_timing', {})
                                    validator_processing_time = metadata.get('processing_time', 0.0)
                                
                                    if batch_timing and batch_timing.get('total_batch_time_seconds', 0.0) > 0:
                                        validation_processing_time = batch_timing.get('total_batch_time_seconds', 0.0)
                                        logger.debug(f"[VALIDATION_TIMING_DEBUG] Fallback to batch_timing: validation_processing_time={validation_processing_time:.3f}s")
                                    elif validator_processing_time > 0:
                                        validation_processing_time = validator_processing_time
                                        logger.debug(f"[VALIDATION_TIMING_DEBUG] Fallback to validator_processing_time: validation_processing_time={validation_processing_time:.3f}s")
                                    else:
                                        logger.warning(f"[VALIDATION_TIMING_DEBUG] No valid timing data found - validation_processing_time will be 0")
                                        validation_processing_time = 0.0
                            else:
                                logger.warning(f"[VALIDATION_TIMING_DEBUG] No start/end times available - using fallback timing sources")
                                # Fallback when start/end times are missing
                                batch_timing = metadata.get('batch_timing', {})
                                validator_processing_time = metadata.get('processing_time', 0.0)
                            
                                if batch_timing and batch_timing.get('total_batch_time_seconds', 0.0) > 0:
                                    validation_processing_time = batch_timing.get('total_batch_time_seconds', 0.0)
                                    logger.debug(f"[VALIDATION_TIMING_DEBUG] Fallback to batch_timing: validation_processing_time={validation_processing_time:.3f}s")
                                elif validator_processing_time > 0:
                                    validation_processing_time = validator_processing_time
                                    logger.debug(f"[VALIDATION_TIMING_DEBUG] Fallback to validator_processing_time: validation_processing_time={validation_processing_time:.3f}s")
                                else:
                                    logger.warning(f"[VALIDATION_TIMING_DEBUG] No valid timing data found - validation_processing_time will be 0")
                                    validation_processing_time = 0.0
                    
                        # Do NOT update estimated_validation_eliyahu_cost - preserve the preview estimate
                        status_update_data['time_per_row_seconds'] = validation_processing_time / processed_rows_count if processed_rows_count > 0 else 0.0  # Actual time per row
                        status_update_data['run_time_s'] = validation_processing_time  # Actual validation run time in seconds
                        status_update_data['actual_processing_time_seconds'] = validation_processing_time  # Same as run_time_s for compatibility
                        status_update_data['actual_time_per_batch_seconds'] = validation_processing_time  # For single batch validation, same as total time
                        logger.debug(f"[TIMING_DEBUG] Set actual_processing_time_seconds={validation_processing_time:.3f}, actual_time_per_batch_seconds={validation_processing_time:.3f}")
                        status_update_data['percent_complete'] = 100  # Mark validation as 100% complete
                        # estimated_validation_time_minutes will be calculated from actual processing time automatically in update_run_status
                    
                        # Use enhanced provider metrics from ai_client aggregation if available, otherwise fallback to token_usage
                        provider_metrics_for_db = {}
                    
                        if enhanced_metrics and enhanced_metrics.get('aggregated_metrics'):
                            # Use properly aggregated enhanced metrics from validation lambda
                            providers = enhanced_metrics['aggregated_metrics'].get('providers', {})

                            for provider, provider_data in providers.items():
                                # Skip QC_Costs as it's now in separate qc_metrics field
                                if provider == 'QC_Costs':
                                    continue
                                provider_metrics_for_db[provider] = {
                                    'calls': provider_data.get('calls', 0),
                                    'tokens': provider_data.get('tokens', 0),
                                    'cost_actual': provider_data.get('cost_actual', 0.0),
                                    'cost_estimated': provider_data.get('cost_estimated', 0.0),
                                    'processing_time': provider_data.get('time_actual', 0.0),
                                    'cache_hit_tokens': provider_data.get('cache_hit_tokens', 0),
                                    'cost_per_row_actual': provider_data.get('cost_actual', 0.0) / processed_rows_count if processed_rows_count > 0 else 0,
                                    'cost_per_row_estimated': provider_data.get('cost_estimated', 0.0) / processed_rows_count if processed_rows_count > 0 else 0,
                                    'time_per_row_actual': provider_data.get('time_actual', 0.0) / processed_rows_count if processed_rows_count > 0 else 0,
                                    'time_per_row_estimated': provider_data.get('time_estimated', 0.0) / processed_rows_count if processed_rows_count > 0 else 0,
                                    'cache_efficiency_percent': provider_data.get('cache_efficiency_percent', 0.0)
                                }
                        
                            logger.debug(f"[VALIDATION_PROVIDER_METRICS] Using enhanced aggregated metrics for provider_metrics_for_db")
                        else:
                            # Fallback to legacy token_usage conversion
                            by_provider = token_usage.get('by_provider', {})
                        
                            for provider, provider_usage in by_provider.items():
                                if provider_usage and isinstance(provider_usage, dict):
                                    # Extract metrics from existing token_usage structure
                                    actual_cost = provider_usage.get('total_cost', 0.0)
                                    total_tokens = provider_usage.get('total_tokens', 0)
                                    api_calls = provider_usage.get('api_calls', 0)
                                    cached_calls = provider_usage.get('cached_calls', 0)
                                
                                    # Estimate cost without cache benefits (approximate based on cache efficiency)
                                    cache_efficiency = cached_calls / max(api_calls, 1) if api_calls > 0 else 0
                                    cost_estimated = actual_cost * (1 + (cache_efficiency * 1.5))  # Conservative estimate
                                
                                    provider_time = processing_time * (api_calls / max(sum(p.get('api_calls', 0) for p in by_provider.values()), 1))
                                    provider_metrics_for_db[provider] = {
                                        'calls': api_calls,
                                        'tokens': total_tokens,
                                        'cost_actual': actual_cost,
                                        'cost_estimated': cost_estimated,
                                        'processing_time': provider_time,
                                        'cache_hit_tokens': cached_calls * (total_tokens / max(api_calls, 1)) if api_calls > 0 else 0,
                                        'cost_per_row_actual': actual_cost / processed_rows_count if processed_rows_count > 0 else 0,
                                        'cost_per_row_estimated': cost_estimated / processed_rows_count if processed_rows_count > 0 else 0,
                                        'time_per_row_actual': provider_time / processed_rows_count if processed_rows_count > 0 else 0,
                                        'time_per_row_estimated': provider_time / processed_rows_count if processed_rows_count > 0 else 0,  # In fallback, use same as actual
                                        'cache_efficiency_percent': (1 - (actual_cost / max(cost_estimated, 0.000001))) * 100
                                    }
                        
                            logger.warning(f"[VALIDATION_PROVIDER_METRICS] Falling back to legacy token_usage conversion for provider_metrics_for_db")
                    
                        status_update_data['provider_metrics'] = provider_metrics_for_db

                        # Extract QC metrics from validation results if available
                        qc_metrics_data = validation_results.get('qc_metrics', {})
                        if qc_metrics_data:
                            logger.debug(f"[QC_DB_STORAGE] Found QC metrics in validation results: {list(qc_metrics_data.keys())}")
                            status_update_data['qc_metrics'] = qc_metrics_data
                        else:
                            logger.debug(f"[QC_DB_STORAGE] No QC metrics found in validation results")

                        # Calculate total_provider_calls from all providers (QC calls now included in anthropic provider)
                        total_provider_calls_override = sum(
                            provider_data.get('calls', 0)
                            for provider_name, provider_data in provider_metrics_for_db.items()
                            if not provider_data.get('is_metadata_only', False)  # Exclude QC_Costs metadata-only provider
                        )
                        status_update_data['total_provider_calls'] = total_provider_calls_override

                        # Debug logging to verify call counting
                        anthropic_calls = provider_metrics_for_db.get('anthropic', {}).get('calls', 0)
                        perplexity_calls = provider_metrics_for_db.get('perplexity', {}).get('calls', 0)
                        qc_calls_debug = qc_metrics_data.get('total_qc_calls', 0) if qc_metrics_data else 0
                        logger.debug(f"[FULL_VALIDATION_CALLS] Anthropic calls (incl QC): {anthropic_calls}, Perplexity calls: {perplexity_calls}, QC calls (debug): {qc_calls_debug}, Total: {total_provider_calls_override}")

                        # Record completion time for background handler
                        background_end_time = datetime.now(timezone.utc).isoformat()
                    
                        # Calculate actual background handler processing time
                        start_time = datetime.fromisoformat(background_start_time.replace('Z', '+00:00'))
                        end_time = datetime.fromisoformat(background_end_time.replace('Z', '+00:00'))
                        background_processing_time_seconds = (end_time - start_time).total_seconds()
                        logger.debug(f"[VALIDATION_TIMING] Background handler processing time: {background_processing_time_seconds:.3f}s")
                    
                        # Override timing with actual background handler processing time
                        status_update_data['end_time'] = background_end_time  # Use background handler completion time
                        status_update_data['run_time_s'] = background_processing_time_seconds  # Actual background handler processing time
                        status_update_data['actual_processing_time_seconds'] = background_processing_time_seconds  # Override with background handler time
                        status_update_data['actual_time_per_batch_seconds'] = background_processing_time_seconds  # Override with background handler time
                        logger.debug(f"[TIMING_OVERRIDE] Override timing fields with background handler time: {background_processing_time_seconds:.3f}s")

                        # ========== DETAILED VALIDATION COMPLETENESS CHECK ==========
                        # This is a detailed check that happens AFTER successful billing
                        # The initial check before billing prevents charges for incomplete results
                        logger.debug(f"[FINAL_COMPLETENESS_CHECK] Performing detailed validation result completeness check")

                        # Extract validation results for completeness check
                        final_validation_results = validation_results.get('validation_results', {}) if validation_results else {}
                        expected_row_count = total_rows_in_file
                        actual_results_count = len(final_validation_results) if isinstance(final_validation_results, dict) else 0

                        logger.debug(f"[FINAL_COMPLETENESS_CHECK] Expected rows: {expected_row_count}, Actual results: {actual_results_count}")

                        # Check for basic completeness
                        is_complete = True
                        completeness_issues = []

                        # Check row count completeness
                        if actual_results_count < expected_row_count:
                            is_complete = False
                            completeness_issues.append(f"Missing results: {actual_results_count}/{expected_row_count} rows")

                        # Check for meaningful data in results
                        if actual_results_count == 0:
                            is_complete = False
                            completeness_issues.append("No validation results found")

                        # Check for expected data structure in a sample of results
                        if final_validation_results and isinstance(final_validation_results, dict):
                            # Check first few results for proper structure
                            sample_keys = list(final_validation_results.keys())[:3]
                            for row_key in sample_keys:
                                row_data = final_validation_results.get(row_key, {})
                                if not isinstance(row_data, dict):
                                    is_complete = False
                                    completeness_issues.append(f"Invalid row data structure for row {row_key}")
                                    break

                                # Check if row has validation results (not just empty dict)
                                if not any(isinstance(v, dict) and 'value' in v for v in row_data.values()):
                                    is_complete = False
                                    completeness_issues.append(f"Row {row_key} missing validation field data")
                                    break

                        # Check for expected sheets in enhanced Excel (if available)
                        try:
                            if enhanced_excel_content:
                                import openpyxl
                                import io
                                wb = openpyxl.load_workbook(io.BytesIO(enhanced_excel_content), read_only=True)
                                expected_sheets = ['Updated Values', 'Original Values', 'Details']  # Correct expected sheets
                                missing_sheets = [sheet for sheet in expected_sheets if sheet not in wb.sheetnames]
                                if missing_sheets:
                                    is_complete = False
                                    completeness_issues.append(f"Enhanced Excel missing sheets: {missing_sheets}")
                                wb.close()
                        except Exception as excel_check_error:
                            logger.warning(f"[FINAL_COMPLETENESS_CHECK] Could not verify enhanced Excel structure: {excel_check_error}")

                        # Check metadata completeness
                        if validation_results:
                            metadata = validation_results.get('metadata', {})
                            if not metadata or not isinstance(metadata, dict):
                                is_complete = False
                                completeness_issues.append("Missing or invalid metadata")
                            else:
                                # Check for essential metadata fields
                                essential_metadata = ['processing_time', 'completed_rows']
                                missing_metadata = [field for field in essential_metadata if field not in metadata]
                                if missing_metadata:
                                    completeness_issues.append(f"Missing metadata fields: {missing_metadata}")

                        # Log completeness results
                        if is_complete:
                            logger.debug(f"[FINAL_COMPLETENESS_CHECK] ✅ Validation results are COMPLETE")
                            # Continue with normal completion flow
                            update_run_status(**status_update_data)
                        else:
                            logger.error(f"[FINAL_COMPLETENESS_CHECK] ❌ Validation results are INCOMPLETE: {'; '.join(completeness_issues)}")
                            logger.error(f"[FINAL_COMPLETENESS_CHECK] 🚫 STOPPING completion flow - will not charge user or send email")

                            # Mark as FAILED instead of COMPLETED
                            status_update_data['status'] = 'FAILED'
                            status_update_data['error_message'] = f"Validation incomplete: {'; '.join(completeness_issues)}"
                            status_update_data['verbose_status'] = f"❌ Validation failed: {'; '.join(completeness_issues)}"
                            status_update_data['completeness_issues'] = completeness_issues
                            status_update_data['validation_incomplete'] = True

                            # Update status as FAILED
                            update_run_status(**status_update_data)

                            # Send error notification via WebSocket
                            error_payload = {
                                'type': 'validation_failed',
                                'session_id': session_id,
                                'progress': 100,
                                'status': f'❌ Validation failed: {completeness_issues[0] if completeness_issues else "Incomplete results"}',
                                'error': f"Validation incomplete: {'; '.join(completeness_issues)}"
                            }
                            _send_websocket_message_deduplicated(session_id, error_payload)

                            # Update session_info.json with failure status and correct config version
                            try:
                                storage_manager.update_session_results(
                                    email=email,
                                    session_id=clean_session_id,
                                    operation_type="validation",
                                    config_id=config_id,
                                    version=config_version,
                                    run_key=run_key,
                                    status="failed",
                                    completed_at=datetime.now(timezone.utc).isoformat(),
                                    frontend_payload=error_payload
                                )
                                logger.debug(f"[SESSION_TRACKING] Updated session_info.json with validation failure (incomplete results)")
                            except Exception as e:
                                logger.error(f"Failed to update session_info.json for validation failure: {e}")

                            # Return early - do NOT continue with billing, email, or completion
                            return {'statusCode': 500, 'body': json.dumps({
                                'status': 'validation_failed',
                                'error': f"Validation incomplete: {'; '.join(completeness_issues)}",
                                'session_id': session_id
                            })}
                    
                        # Track enhanced user metrics for full validation
                        logger.debug(f"[USER_TRACKING] Tracking full validation request for email: {email}")
                        try:
                            track_result = track_user_request(
                            email=email,
                            request_type='full',
                            tokens_used=token_usage.get('total_tokens', 0),
                            cost_usd=charged_cost,  # Legacy field
                            perplexity_tokens=token_usage.get('by_provider', {}).get('perplexity', {}).get('total_tokens', 0),
                            perplexity_cost=token_usage.get('by_provider', {}).get('perplexity', {}).get('total_cost', 0),
                            anthropic_tokens=token_usage.get('by_provider', {}).get('anthropic', {}).get('total_tokens', 0),
                            anthropic_cost=token_usage.get('by_provider', {}).get('anthropic', {}).get('total_cost', 0),
                            # Enhanced metrics
                            rows_processed=processed_rows_count,
                            total_rows=total_rows_in_file,
                            columns_validated=validation_metrics.get('validated_columns_count', 0),
                            search_groups=validation_metrics.get('search_groups_count', 0),
                            high_context_search_groups=validation_metrics.get('enhanced_context_search_groups_count', 0),
                            claude_calls=validation_metrics.get('claude_search_groups_count', 0),
                            eliyahu_cost=eliyahu_cost,  # Actual cost paid
                            estimated_cost=cost_estimated,  # Raw cost estimate without caching
                            quoted_validation_cost=charged_cost,  # This is the quoted cost before discount
                            charged_cost=effective_cost,  # Full validation charges user the effective cost (after discount)
                            total_api_calls=token_usage.get('api_calls', 0),
                            total_cached_calls=token_usage.get('cached_calls', 0)
                            )
                            logger.debug(f"[USER_TRACKING] Full validation tracking result: {track_result}")
                        except Exception as e:
                            logger.error(f"[USER_TRACKING] Failed to track full validation request: {e}")
                            import traceback
                            logger.error(f"[USER_TRACKING] Traceback: {traceback.format_exc()}")
        
        else: # No results
             logger.error(f"[DEBUG] No validation results returned from validator for session {session_id}")
             logger.error(f"[DEBUG] is_preview = {is_preview}")
             run_type_for_failed = "Preview" if is_preview else "Validation"
             logger.error(f"[DEBUG] About to update run status to FAILED for {run_type_for_failed}")
             update_run_status_for_session(status='FAILED', 
                run_type=run_type_for_failed,
                error_message="No validation results returned", 
                verbose_status="Failed to process results.", 
                percent_complete=100,
                processed_rows=0,
                batch_size=10,  # Default batch size
                eliyahu_cost=0.0,
                time_per_row_seconds=0.0)
             logger.error(f"[DEBUG] Completed run status update to FAILED")
             # Handle empty results case for preview if needed
             logger.error(f"[DEBUG] Checking if is_preview: {is_preview}")
             if is_preview:
                logger.error(f"[DEBUG] Entering preview handling for empty results")
                # Even for empty results, store in versioned folder
                config_version = 1
                config_id = None
                try:
                    logger.debug(f"[DEBUG] Getting latest config for email={email}, session={clean_session_id}")
                    config_data, latest_config_key = storage_manager.get_latest_config(email, clean_session_id)
                    logger.debug(f"[DEBUG] Got latest_config_key: {latest_config_key}")
                    if config_data:
                        config_version = config_data.get('storage_metadata', {}).get('version', 1)
                        config_id = config_data.get('storage_metadata', {}).get('config_id')
                    elif latest_config_key:
                        config_filename = latest_config_key.split('/')[-1]
                        if config_filename.startswith('v') and '_' in config_filename:
                            config_version = int(config_filename.split('_')[0][1:])
                    logger.debug(f"[DEBUG] Determined config_version: {config_version}")
                except Exception as e:
                    logger.warning(f"[DEBUG] Exception getting config version: {e}")
                    pass

                logger.debug(f"[DEBUG] About to store preview results with config_version={config_version}")
                result = storage_manager.store_results(
                    email, clean_session_id, config_version,
                    {"status": "preview_completed", "note": "No results"}, 'preview'
                )
                logger.debug(f"[DEBUG] Store results completed: {result}")

                if result['success']:
                    preview_results_key = result['s3_key']
                    logger.debug(f"[DEBUG] Store results successful, s3_key: {preview_results_key}")
                else:
                    logger.warning(f"[DEBUG] Store results failed, using fallback")
                    # Fallback
                    preview_results_key = f"preview_results/{email_folder}/{session_id}.json"
                    s3_client.put_object(
                        Bucket=S3_RESULTS_BUCKET, Key=preview_results_key,
                        Body=json.dumps({"status": "preview_completed", "note": "No results"}),
                        ContentType='application/json'
                    )

                # Update session_info.json with failure status and correct config version
                try:
                    error_payload = {
                        'type': 'preview_failed',
                        'status': '❌ Preview failed - no results',
                        'session_id': session_id,
                        'error': 'No validation results returned'
                    }

                    storage_manager.update_session_results(
                        email=email,
                        session_id=clean_session_id,
                        operation_type="preview",
                        config_id=config_id or f"{clean_session_id}_config_v{config_version}",
                        version=config_version,
                        run_key=run_key,
                        status="failed",
                        completed_at=datetime.now(timezone.utc).isoformat(),
                        frontend_payload=error_payload
                    )
                    logger.debug(f"[SESSION_TRACKING] Updated session_info.json with preview failure (no results)")
                except Exception as e:
                    logger.error(f"Failed to update session_info.json for preview failure: {e}")

                logger.debug(f"[DEBUG] About to return preview_completed response for session {session_id}")
                return {'statusCode': 200, 'body': json.dumps({'status': 'preview_completed', 'session_id': session_id})}
             else:
                # Return error response for non-preview failures (validation)
                logger.debug(f"[DEBUG] About to return background_failed response for non-preview session {session_id}")

                # Update session_info.json with failure status and correct config version
                try:
                    # Try to get config version and config_id
                    config_version = 1
                    config_id = None
                    try:
                        config_data, _ = storage_manager.get_latest_config(email, clean_session_id)
                        if config_data:
                            config_version = config_data.get('storage_metadata', {}).get('version', 1)
                            config_id = config_data.get('storage_metadata', {}).get('config_id')
                    except Exception as cfg_err:
                        logger.warning(f"Could not get config for failed validation: {cfg_err}")

                    error_payload = {
                        'type': 'validation_failed',
                        'status': '❌ Validation failed - no results',
                        'session_id': session_id,
                        'error': 'No validation results returned'
                    }

                    storage_manager.update_session_results(
                        email=email,
                        session_id=clean_session_id,
                        operation_type="validation",
                        config_id=config_id or f"{clean_session_id}_config_v{config_version}",
                        version=config_version,
                        run_key=run_key,
                        status="failed",
                        completed_at=datetime.now(timezone.utc).isoformat(),
                        frontend_payload=error_payload
                    )
                    logger.debug(f"[SESSION_TRACKING] Updated session_info.json with validation failure (no results)")
                except Exception as e:
                    logger.error(f"Failed to update session_info.json for validation failure: {e}")

                return {'statusCode': 500, 'body': json.dumps({
                    'status': 'background_failed',
                    'error': 'No validation results returned',
                    'session_id': session_id
                })}
            
        return {'statusCode': 200, 'body': json.dumps({'status': 'background_completed', 'session_id': session_id})}

    except Exception as e:
        logger.error(f"Critical error in background processing for session {event.get('session_id', 'unknown')}: {str(e)}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        traceback.print_exc()
        if DYNAMODB_AVAILABLE:
            # Determine run type from event or default to unknown
            is_preview = event.get('preview_mode', False) or event.get('request_type') == 'preview'
            is_config = event.get('request_type') == 'config_generation'
            run_type_for_error = "Config Generation" if is_config else ("Preview" if is_preview else "Validation")
            # Try to get run_key from event, if not available create one for error tracking
            session_id_for_error = event.get('session_id')
            run_key_for_error = event.get('run_key')
            
            if not run_key_for_error and session_id_for_error:
                try:
                    # Create a new run record for this error if one doesn't exist
                    run_key_for_error = create_run_record(
                        session_id=session_id_for_error, 
                        email=event.get('email', 'unknown'), 
                        total_rows=0, 
                        batch_size=1, 
                        run_type=run_type_for_error
                    )
                    logger.info(f"Created run_key for error tracking: {run_key_for_error}")
                except Exception as create_error:
                    logger.error(f"Failed to create run_key for error tracking: {create_error}")
                    run_key_for_error = None
            
            if run_key_for_error:
                update_run_status(
                    session_id=session_id_for_error,
                    run_key=run_key_for_error,
                    status='FAILED',
                    run_type=run_type_for_error,
                    error_message=str(e),
                    processed_rows=0,
                    batch_size=1 if is_config else 10,
                    eliyahu_cost=0.0,
                    time_per_row_seconds=0.0
                )
            else:
                logger.warning(f"Cannot update run status for failed session {session_id_for_error} - no run_key available")

            # Send error notification to frontend via WebSocket
            if session_id_for_error:
                try:
                    if is_config:
                        error_message = {
                            'type': 'config_generation_failed',
                            'session_id': session_id_for_error,
                            'success': False,
                            'error': str(e),
                            'progress': 100
                        }
                    else:
                        error_message = {
                            'type': 'validation_failed' if not is_preview else 'preview_failed',
                            'session_id': session_id_for_error,
                            'progress': 100,
                            'status': f'❌ {run_type_for_error} failed: {str(e)[:100]}{"..." if len(str(e)) > 100 else ""}',
                            'error': str(e)
                        }

                    _send_websocket_message_deduplicated(session_id_for_error, error_message)
                    logger.info(f"Sent error notification via WebSocket for session {session_id_for_error}")
                except Exception as ws_error:
                    logger.error(f"Failed to send error notification via WebSocket: {ws_error}")

                # Send email notification if EMAIL_SENDER_AVAILABLE
                if EMAIL_SENDER_AVAILABLE:
                    try:
                        email_for_error = event.get('email', 'unknown')
                        if email_for_error != 'unknown':
                            send_validation_results_email(
                                email=email_for_error,
                                session_id=session_id_for_error,
                                status='FAILED',
                                error_message=str(e),
                                is_preview=is_preview,
                                is_config=is_config
                            )
                            logger.info(f"Sent error notification via email to {email_for_error}")
                        else:
                            logger.warning(f"Cannot send email notification - no email address in event")
                    except Exception as email_error:
                        logger.error(f"Failed to send error notification via email: {email_error}")
        return {'statusCode': 500, 'body': json.dumps({'status': 'background_failed', 'error': str(e)})}

def handle_config_generation(event, context):
    """Handle config generation requests by forwarding to config lambda."""
    try:
        import time
        execution_id = f"{context.aws_request_id if context else 'no-context'}_{int(time.time() * 1000)}"
        logger.debug(f"[CONFIG_GEN_START] {execution_id} - Handling config generation request for session {event.get('session_id')}")
        session_id = event.get('session_id')
        config_email = event.get('email', '').lower().strip()
        
        # Record when background handler actually starts processing
        background_start_time = datetime.now(timezone.utc).isoformat()
        
        # Create initial config generation run record
        if DYNAMODB_AVAILABLE and config_email and session_id:
            try:
                # Get input table name from event
                input_table_name = "unknown_table"
                if event.get('excel_s3_key'):
                    input_table_name = event['excel_s3_key'].split('/')[-1]
                elif event.get('table_name'):
                    input_table_name = event.get('table_name')
                
                # Get total rows if available
                total_rows = event.get('total_rows', 0)
                
                logger.debug(f"[CONFIG_RUN_TRACKING] Creating config generation run record for session {session_id}")
                run_key = create_run_record(session_id, config_email, total_rows, 1, "Config Generation")  # batch_size=1 for config generation
                logger.debug(f"[CONFIG_RUN_TRACKING] Config generation run_key: {run_key}")
                
                # Update start_time to when background handler actually begins processing
                update_run_status(session_id=session_id, run_key=run_key, status='IN_PROGRESS',
                    run_type="Config Generation",
                    verbose_status="Configuration generation starting with AI analysis...",
                    percent_complete=5,
                    processed_rows=0,
                    total_rows=total_rows,
                    input_table_name=input_table_name,
                    account_current_balance=0,  # Will be updated later
                    account_sufficient_balance="n/a",
                    start_time=background_start_time,  # Use background handler start time
                    account_credits_needed="n/a",
                    account_domain_multiplier=1.0,  # Config generation typically doesn't use domain multiplier
                    models="TBD",  # Will be updated after AI processing
                    batch_size=1,  # Config generation batch size
                    eliyahu_cost=0.0,  # Will be updated after completion
                    time_per_row_seconds=0.0  # Will be updated after completion
                )
            except Exception as e:
                logger.error(f"[CONFIG_RUN_TRACKING] Failed to create config run record: {e}")
        
        # Send initial progress update - interface setup (0-5% range)
        _send_websocket_message_deduplicated(session_id, {
            'type': 'config_generation_progress',
            'progress': 1,  # Interface setup: 0-5% range
            'status': '🚀 Starting AI configuration generation...',
            'session_id': session_id
        }, "config_progress_start")
        
        # Send analysis progress update - interface setup (0-5% range)
        _send_websocket_message_deduplicated(session_id, {
            'type': 'config_generation_progress', 
            'progress': 3,  # Interface setup: 0-5% range
            'status': '🔍 Analyzing table structure and data patterns...',
            'session_id': session_id
        }, "config_progress_analysis")
        
        # Invoke the config lambda directly - interface setup complete (5% handoff to config lambda for 5-90%)
        _send_websocket_message_deduplicated(session_id, {
            'type': 'config_generation_progress',
            'progress': 5,  # Interface setup complete: hand off to config lambda for 5-90%
            'status': '🤖 Developing configuration with AI...',
            'session_id': session_id
        }, "config_progress_ai_invoke")
        
        response = invoke_config_lambda(event)
        
        logger.info(f"Config generation response: {response}")
        logger.info(f"Response keys: {list(response.keys()) if isinstance(response, dict) else 'Not a dict'}")
        
        # Send AI processing completion update - interface final processing begins (90-100% range)
        _send_websocket_message_deduplicated(session_id, {
            'type': 'config_generation_progress',
            'progress': 90,  # Interface final processing: 90-100% range
            'status': '✨ AI configuration generated successfully!',
            'session_id': session_id
        }, "config_progress_ai_complete")
        
        if response.get('success'):
            # Get the S3 key and create download URL
            config_s3_key = response.get('config_s3_key') or response.get('generated_config_s3_key')
            # Config lambda returns 'ai_summary'
            ai_summary = response.get('ai_summary', response.get('ai_response', ''))
            download_url = response.get('config_download_url', response.get('download_url', ''))
            
            logger.info(f"Extracted fields: config_s3_key={config_s3_key}, ai_summary_length={len(ai_summary) if ai_summary else 0}, download_url={download_url}")
            
            # Send storage progress update - interface final processing (90-100% range)
            _send_websocket_message_deduplicated(session_id, {
                'type': 'config_generation_progress',
                'progress': 95,  # Interface final processing: 90-100% range
                'status': '💾 Storing configuration in unified storage...',
                'session_id': session_id
            }, "config_progress_storage")
            
            # Use download URL from config lambda response, or create one if not provided
            if config_s3_key and not download_url:
                download_url = create_config_download_url(config_s3_key)
                logger.info(f"Created fallback download URL: {download_url}")
            elif download_url:
                logger.info(f"Using download URL from config lambda: {download_url}")
            
            # Send final completion progress - interface final processing complete (100%)
            _send_websocket_message_deduplicated(session_id, {
                'type': 'config_generation_progress',
                'progress': 100,  # Interface final processing complete: 100%
                'status': '🎉 Configuration generation complete!',
                'session_id': session_id
            }, "config_progress_complete")
            
            # Track config generation metrics
            config_cost = response.get('cost_info', {}).get('total_cost', 0.0)
            config_tokens = response.get('cost_info', {}).get('total_tokens', 0)
            
            # Get and normalize email for tracking
            config_email = event.get('email', '').lower().strip()
            if config_email:  # Only track if email is not empty
                logger.debug(f"[USER_TRACKING] Tracking config generation for email: {config_email}")
                try:
                    track_result = track_user_request(
                    email=config_email,
                request_type='config',
                tokens_used=config_tokens,
                cost_usd=config_cost,  # Legacy field
                config_cost=config_cost,
                # Config generation typically uses Perplexity or Anthropic
                perplexity_tokens=response.get('cost_info', {}).get('perplexity_tokens', 0),
                perplexity_cost=response.get('cost_info', {}).get('perplexity_cost', 0.0),
                anthropic_tokens=response.get('cost_info', {}).get('anthropic_tokens', 0),
                anthropic_cost=response.get('cost_info', {}).get('anthropic_cost', 0.0),
                eliyahu_cost=config_cost,  # Config generation costs are our actual costs
                estimated_cost=config_cost,
                quoted_validation_cost=0.0,  # Config generation is currently free
                charged_cost=0.0,  # Config generation is currently free
                total_api_calls=response.get('cost_info', {}).get('api_calls', 0),
                total_cached_calls=response.get('cost_info', {}).get('cached_calls', 0)
                    )
                    logger.debug(f"[USER_TRACKING] Config tracking result: {track_result}")
                except Exception as e:
                    logger.error(f"[USER_TRACKING] Failed to track config generation: {e}")
                    import traceback
                    logger.error(f"[USER_TRACKING] Traceback: {traceback.format_exc()}")
            else:
                logger.warning("Config generation completed but no email provided for user tracking")
            
            # Update config generation run record with completion data
            if DYNAMODB_AVAILABLE and config_email and session_id:
                try:
                    logger.debug(f"[CONFIG_RUN_TRACKING] Updating config generation run record with completion data")
                    
                    # Get the actual model used from config lambda response
                    models = response.get('model_used', 'unknown')
                    
                    # Get configuration ID  
                    configuration_id = "unknown"
                    if response.get('updated_config'):
                        configuration_id = response['updated_config'].get('generation_metadata', {}).get('config_id', 'unknown')
                    
                    # Extract config version from response
                    config_version = 1
                    if response.get('updated_config'):
                        config_version = response['updated_config'].get('generation_metadata', {}).get('version', 1)
                        logger.info(f"Extracted config version {config_version} from generation_metadata")
                    
                    # Extract cost_info from response
                    cost_info = response.get('cost_info', {})
                    
                    # Build preview_data for config generation
                    config_data = {
                        "config_type": "ai_generated",
                        "ai_models_used": models,
                        "processing_time_seconds": response.get('processing_time', 0),
                        "cost_estimates": {
                            "total_cost": config_cost,
                            "per_generation_cost": config_cost
                        },
                        "token_usage": {
                            "total_tokens": config_tokens,
                            "by_provider": {
                                "perplexity": {
                                    "total_tokens": cost_info.get('perplexity_tokens', 0),
                                    "total_cost": cost_info.get('perplexity_cost', 0.0),
                                    "calls": cost_info.get('perplexity_calls', 0)
                                },
                                "anthropic": {
                                    "total_tokens": cost_info.get('anthropic_tokens', 0),
                                    "total_cost": cost_info.get('anthropic_cost', 0.0), 
                                    "calls": cost_info.get('anthropic_calls', 0)
                                }
                            }
                        },
                        "generation_metadata": {
                            "config_version": config_version,
                            "config_s3_key": config_s3_key,
                            "download_url": download_url,
                            "ai_summary_length": len(ai_summary) if ai_summary else 0,
                            "clarifying_questions_count": len(response.get('clarifying_questions', '').split('\n')) if response.get('clarifying_questions') else 0
                        }
                    }
                    
                    # Record completion time for background handler
                    background_end_time = datetime.now(timezone.utc).isoformat()
                    
                    # Calculate timing metrics for config generation from background handler processing time
                    start_time = datetime.fromisoformat(background_start_time.replace('Z', '+00:00'))
                    end_time = datetime.fromisoformat(background_end_time.replace('Z', '+00:00'))
                    processing_time_seconds = (end_time - start_time).total_seconds()
                    logger.debug(f"[CONFIG_TIMING] Background handler processing time: {processing_time_seconds:.3f}s")
                    
                    time_per_config = processing_time_seconds  # For config generation, it's time per config
                    processing_time_minutes = processing_time_seconds / 60.0
                    
                    # Use enhanced data from config lambda if available, otherwise fallback to token_usage
                    provider_metrics_for_db = {}
                    config_enhanced_data = response.get('enhanced_data', {})
                    
                    # Get actual costs from config response (more reliable than enhanced_data for config)
                    config_cost_actual = response.get('eliyahu_cost', 0.0)
                    config_cost_estimated = response.get('estimated_cost', config_cost_actual)
                    
                    if config_enhanced_data and config_enhanced_data.get('provider_metrics'):
                        # Use enhanced data from config lambda (single call structure)
                        for provider, provider_data in config_enhanced_data['provider_metrics'].items():
                            provider_metrics_for_db[provider] = {
                                'calls': provider_data.get('calls', 1),
                                'tokens': provider_data.get('tokens', 0),
                                'cost_actual': config_cost_actual,  # Use actual cost from response
                                'cost_estimated': config_cost_estimated,  # Use estimated cost from response
                                'processing_time': processing_time_seconds,  # Use actual processing time
                                'cache_hit_tokens': provider_data.get('cache_hit_tokens', 0),
                                'cost_per_row_actual': config_cost_actual,  # For config, "per row" is per config
                                'cost_per_row_estimated': config_cost_estimated,
                                'time_per_row_actual': processing_time_seconds,  # Use actual processing time
                                'cache_efficiency_percent': provider_data.get('cache_efficiency_percent', 0.0)
                            }
                        
                        logger.debug(f"[CONFIG_PROVIDER_METRICS] Using enhanced data from config lambda")
                    else:
                        # Fallback to legacy token_usage conversion
                        token_usage_data = config_data.get('token_usage', {}).get('by_provider', {})
                        
                        for provider, provider_usage in token_usage_data.items():
                            if provider_usage and isinstance(provider_usage, dict):
                                # Extract metrics from config token_usage structure
                                actual_cost = provider_usage.get('total_cost', 0.0)
                                total_tokens = provider_usage.get('total_tokens', 0)
                                api_calls = provider_usage.get('calls', 0)
                                
                                # For config generation, assume minimal caching benefit (conservative estimate)
                                cache_multiplier = 1.1  # 10% increase for non-cached config operations
                                cost_estimated = actual_cost * cache_multiplier
                                
                                provider_metrics_for_db[provider] = {
                                    'calls': api_calls,
                                    'tokens': total_tokens,
                                    'cost_actual': actual_cost,
                                    'cost_estimated': cost_estimated,
                                    'processing_time': processing_time_seconds * (api_calls / max(sum(p.get('calls', 0) for p in token_usage_data.values()), 1)) if api_calls > 0 else 0,
                                    'cache_hit_tokens': 0,  # Config generation typically doesn't benefit from significant caching
                                    'cost_per_row_actual': actual_cost,  # For config, "per row" is per config
                                    'cost_per_row_estimated': cost_estimated,
                                    'time_per_row_actual': processing_time_seconds * (api_calls / max(sum(p.get('calls', 0) for p in token_usage_data.values()), 1)) if api_calls > 0 else processing_time_seconds,
                                    'cache_efficiency_percent': ((cost_estimated - actual_cost) / max(cost_estimated, 0.000001)) * 100
                                }
                        
                        logger.warning(f"[CONFIG_PROVIDER_METRICS] Falling back to legacy token_usage conversion")
                    
                    update_run_status(session_id=session_id, run_key=run_key, status='COMPLETED',
                        run_type="Config Generation",
                        verbose_status="Configuration generation completed successfully.",
                        percent_complete=100,
                        processed_rows=1,  # One config generated
                        preview_data=config_data,
                        models=models,
                        configuration_id=configuration_id,
                        results_s3_key=config_s3_key,  # Points to generated config file
                        account_current_balance=0,  # Config generation is free
                        account_sufficient_balance="n/a",
                        account_credits_needed="n/a",
                        end_time=background_end_time,  # Use background handler completion time
                        account_domain_multiplier=1.0,
                        batch_size=1,  # Config generation batch size
                        eliyahu_cost=config_cost,  # Actual cost for config generation
                        quoted_validation_cost=0.0,  # Config generation is free to users
                        estimated_validation_eliyahu_cost=None,  # Config generation doesn't use validation cost estimates
                        time_per_row_seconds=None,  # Config generation doesn't use validation timing fields
                        estimated_validation_time_minutes=None,  # Config generation doesn't use validation timing fields
                        run_time_s=processing_time_seconds,  # Actual config generation time in seconds
                        provider_metrics=provider_metrics_for_db  # Enhanced provider-specific metrics
                    )
                    logger.debug(f"[CONFIG_RUN_TRACKING] Successfully updated config generation run record")
                except Exception as e:
                    logger.error(f"[CONFIG_RUN_TRACKING] Failed to update config run record: {e}")
                    import traceback
                    logger.error(f"[CONFIG_RUN_TRACKING] Traceback: {traceback.format_exc()}")
            
            # Extract version from the actual config
            config_version = 1
            if response.get('updated_config'):
                config_version = response['updated_config'].get('generation_metadata', {}).get('version', 1)
                logger.info(f"Extracted config version {config_version} from generation_metadata")
            
            # Send success message via WebSocket
            websocket_message = {
                'type': 'config_generation_complete',
                'session_id': session_id,
                'success': True,
                'download_url': download_url,
                'ai_summary': ai_summary,
                'ai_response': ai_summary,  # For backward compatibility
                'config_s3_key': config_s3_key,
                'config_version': config_version,  # Explicit version from config
                'clarifying_questions': response.get('clarifying_questions', ''),
                'clarification_urgency': response.get('clarification_urgency', 0.0)
            }
            
            logger.info(f"WebSocket message to send: {json.dumps(websocket_message, indent=2)}")
            
            # Store the generated config in unified storage for future refinements
            try:
                from ..core.unified_s3_manager import UnifiedS3Manager
                
                storage_manager = UnifiedS3Manager()
                email = event.get('email', '').lower().strip() if event.get('email') else ''
                original_session_id = event.get('original_session_id') or session_id
                
                if email and original_session_id and response.get('updated_config'):
                    # CRITICAL: Add clarifying questions and conversation log to config before storing
                    updated_config = response['updated_config'].copy()

                    # Add clarifying questions to config
                    config_clarifying_questions = response.get('clarifying_questions', '')
                    clarification_urgency = response.get('clarification_urgency', 0.0)
                    if config_clarifying_questions:
                        updated_config['clarifying_questions'] = config_clarifying_questions
                        updated_config['clarification_urgency'] = clarification_urgency
                        print(f"🔍 BACKGROUND_DEBUG: Added clarifying questions to config")

                    # Add conversation entry to config_change_log
                    if 'config_change_log' not in updated_config:
                        updated_config['config_change_log'] = []

                    # Create conversation entry for this interaction
                    conversation_entry = {
                        'timestamp': datetime.now().isoformat(),
                        'action': 'unified_generation',
                        'session_id': original_session_id,
                        'instructions': event.get('instructions', ''),
                        'clarifying_questions': config_clarifying_questions,
                        'clarification_urgency': clarification_urgency,
                        'reasoning': response.get('reasoning', ''),
                        'ai_summary': response.get('ai_summary', ''),
                        'technical_ai_summary': response.get('technical_ai_summary', ''),
                        'version': updated_config.get('generation_metadata', {}).get('version', 1),
                        'model_used': response.get('model_used', 'unknown'),
                        'config_filename': response.get('config_filename', ''),
                        'entry_type': 'ai_response'
                    }

                    updated_config['config_change_log'].append(conversation_entry)
                    print(f"🔍 BACKGROUND_DEBUG: Added conversation entry to config_change_log (total entries: {len(updated_config['config_change_log'])})")

                    # Store config in unified storage using versioning system
                    from ..actions.generate_config_unified import store_config_with_versioning
                    storage_result = store_config_with_versioning(
                        email, original_session_id, updated_config,
                        source='ai_generated'
                    )
                    logger.info(f"Stored config in unified storage: {storage_result}")
                    
                    # If the config lambda didn't provide a working download URL, create one from unified storage
                    if not download_url and storage_result.get('success'):
                        try:
                            config_version = storage_result.get('version', 1)
                            unified_download_url = storage_manager.create_public_download_link(
                                response['updated_config'], f"config_v{config_version}_{original_session_id}.json"
                            )
                            if unified_download_url:
                                download_url = unified_download_url
                                websocket_message['download_url'] = download_url
                                logger.info(f"Created fallback download URL from unified storage: {download_url}")
                        except Exception as download_error:
                            logger.error(f"Failed to create download URL from unified storage: {download_error}")
                    
            except Exception as storage_error:
                logger.error(f"Failed to store config in unified storage: {storage_error}")
                # Don't fail the whole operation for this
            
            # Send via WebSocket with deduplication
            try:
                logger.debug(f"[CONFIG_COMPLETION] About to send config_generation_complete for {session_id} with download_url: {websocket_message.get('download_url', 'None')}")
                _send_websocket_message_deduplicated(session_id, websocket_message, "config_generation_complete")
                logger.debug(f"[CONFIG_COMPLETION] Sent config_generation_complete for {session_id}")
            except Exception as ws_error:
                logger.error(f"Failed to send WebSocket message: {ws_error}")
        else:
            # Update config generation run record with failure
            if DYNAMODB_AVAILABLE and config_email and session_id:
                try:
                    logger.debug(f"[CONFIG_RUN_TRACKING] Updating config generation run record with failure")
                    # run_key might not be available in this scope, so get it
                    try:
                        from dynamodb_schemas import find_existing_run_key
                        error_run_key = find_existing_run_key(session_id)
                    except:
                        error_run_key = None

                    if error_run_key:
                        update_run_status(session_id=session_id, run_key=error_run_key, status='FAILED',
                        run_type="Config Generation",
                        verbose_status=f"Configuration generation failed: {response.get('error', 'Unknown error')}",
                        percent_complete=0,
                        processed_rows=0,
                        error_message=response.get('error', 'Config generation failed'),
                        account_current_balance=0,
                        account_sufficient_balance="n/a",
                        account_credits_needed="n/a",
                        account_domain_multiplier=1.0,
                        models="Config generation failed",
                        batch_size=1,
                        eliyahu_cost=0.0,
                        estimated_validation_eliyahu_cost=0.0,
                        time_per_row_seconds=0.0
                        )
                    else:
                        logger.warning(f"[CONFIG_RUN_TRACKING] Could not update run status - no run_key found for session {session_id}")
                    logger.debug(f"[CONFIG_RUN_TRACKING] Successfully updated config run record with failure")
                except Exception as e:
                    logger.error(f"[CONFIG_RUN_TRACKING] Failed to update config run record with failure: {e}")
            
            # Send error message via WebSocket
            websocket_message = {
                'type': 'config_generation_failed',
                'session_id': session_id,
                'success': False,
                'error': response.get('error', 'Unknown error')
            }
            
            try:
                _send_websocket_message_deduplicated(session_id, websocket_message, "config_generation_failed")
            except Exception as ws_error:
                logger.error(f"Failed to send WebSocket error message: {ws_error}")
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'status': 'config_generation_completed',
                'session_id': session_id,
                'response': response
            })
        }
        
    except Exception as e:
        logger.error(f"Config generation error for session {event.get('session_id')}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Update config generation run record with exception failure
        session_id = event.get('session_id')
        config_email = event.get('email', '').lower().strip()
        if DYNAMODB_AVAILABLE and config_email and session_id:
            try:
                logger.debug(f"[CONFIG_RUN_TRACKING] Updating config generation run record with exception failure")
                # run_key might not be available in this scope, so get it
                try:
                    from dynamodb_schemas import find_existing_run_key
                    exception_run_key = find_existing_run_key(session_id)
                except:
                    exception_run_key = None

                if exception_run_key:
                    update_run_status(session_id=session_id, run_key=exception_run_key, status='FAILED',
                    run_type="Config Generation",
                    verbose_status=f"Configuration generation failed with exception: {str(e)}",
                    percent_complete=0,
                    processed_rows=0,
                    error_message=str(e),
                    account_current_balance=0,
                    account_sufficient_balance="n/a",
                    account_credits_needed="n/a",
                    account_domain_multiplier=1.0,
                    models="Config generation exception",
                    batch_size=1,
                    eliyahu_cost=0.0,
                    quoted_validation_cost=0.0,
                    estimated_validation_eliyahu_cost=0.0,
                    time_per_row_seconds=0.0
                    )
                else:
                    logger.warning(f"[CONFIG_RUN_TRACKING] Could not update run status - no run_key found for session {session_id}")
                logger.debug(f"[CONFIG_RUN_TRACKING] Successfully updated config generation run record with exception failure")
            except Exception as update_error:
                logger.error(f"[CONFIG_RUN_TRACKING] Failed to update config run record with exception: {update_error}")
        
        # Send error via WebSocket
        try:
            _send_websocket_message_deduplicated(event.get('session_id'), {
                'type': 'config_generation_failed',
                'session_id': event.get('session_id'),
                'success': False,
                'error': str(e)
            }, "config_generation_error")
        except:
            pass
        
        return {
            'statusCode': 500,
            'body': json.dumps({
                'status': 'config_generation_failed',
                'error': str(e),
                'session_id': event.get('session_id')
            })
        }

def invoke_config_lambda(event: Dict) -> Dict:
    """Invoke the config lambda function."""
    try:
        import boto3
        from botocore.config import Config
        
        # Configure longer timeouts for Opus processing
        config = Config(
            read_timeout=900,  # 15 minutes read timeout  
            connect_timeout=60,  # 1 minute connect timeout
            retries={'max_attempts': 1}  # Don't retry on timeout
        )
        
        lambda_client = boto3.client('lambda', config=config)
        config_lambda_name = os.environ.get('CONFIG_LAMBDA_NAME', 'perplexity-validator-config')
        
        logger.info(f"Invoking config lambda: {config_lambda_name}")
        
        # Always try to retrieve existing config and validation results for config generation
        if event.get('email') and event.get('session_id'):
            try:
                logger.info(f"BACKGROUND_HANDLER: Retrieving existing config and validation results for {event.get('email')}/{event.get('session_id')}")

                # Import and use UnifiedS3Manager to get both config and validation results
                from ..core.unified_s3_manager import UnifiedS3Manager
                storage_manager = UnifiedS3Manager()

                # Try to get existing configuration
                if not event.get('existing_config'):
                    logger.info("BACKGROUND_HANDLER: No existing_config in event, trying to retrieve from storage")
                    try:
                        existing_config = storage_manager.get_latest_config(
                            email=event.get('email'),
                            session_id=event.get('session_id')
                        )
                        if existing_config:
                            event['existing_config'] = existing_config
                            logger.info(f"BACKGROUND_HANDLER: Successfully retrieved existing config!")
                            logger.info(f"Config keys: {list(existing_config.keys())}")
                        else:
                            logger.info(f"BACKGROUND_HANDLER: No existing config found")
                    except Exception as config_error:
                        logger.warning(f"BACKGROUND_HANDLER: Exception retrieving existing config: {config_error}")

                # Try to get validation results
                if not event.get('latest_validation_results'):
                    logger.info("BACKGROUND_HANDLER: No validation results in event, trying to retrieve from storage")
                    latest_validation_results = storage_manager.get_latest_validation_results(
                        email=event.get('email'),
                        session_id=event.get('session_id')
                    )

                    if latest_validation_results:
                        event['latest_validation_results'] = latest_validation_results
                        logger.info(f"BACKGROUND_HANDLER: Successfully added validation results to config lambda payload!")
                        logger.info(f"Validation results keys: {list(latest_validation_results.keys())}")
                        if 'markdown_table' in latest_validation_results:
                            logger.info(f"markdown_table size: {len(latest_validation_results['markdown_table'])}")
                    else:
                        logger.info(f"BACKGROUND_HANDLER: No validation results found")

            except Exception as e:
                logger.warning(f"BACKGROUND_HANDLER: Exception retrieving config/validation data: {e}")
                import traceback
                logger.warning(f"Traceback: {traceback.format_exc()}")
        
        # Debug what we're sending to config lambda
        logger.info(f"CONFIG_LAMBDA_PAYLOAD_DEBUG: Event keys being sent: {list(event.keys())}")
        logger.info(f"CONFIG_LAMBDA_PAYLOAD_DEBUG: Has existing_config: {bool(event.get('existing_config'))}")
        logger.info(f"CONFIG_LAMBDA_PAYLOAD_DEBUG: Has latest_validation_results: {bool(event.get('latest_validation_results'))}")
        if 'email' in event and 'session_id' in event:
            logger.info(f"CONFIG_LAMBDA_PAYLOAD_DEBUG: email={event.get('email')}, session_id={event.get('session_id')}")
        
        # Check payload size for potential issues
        try:
            payload_json = json.dumps(event)
            payload_size_mb = len(payload_json.encode('utf-8')) / (1024 * 1024)
            logger.info(f"CONFIG_LAMBDA_PAYLOAD_SIZE: {payload_size_mb:.2f} MB")
            if payload_size_mb > 5.5:
                logger.warning(f"CONFIG_LAMBDA_PAYLOAD_SIZE: Large payload detected! Size: {payload_size_mb:.2f} MB (limit is 6MB)")
                # Log size breakdown if payload is large
                if event.get('latest_validation_results'):
                    vr_size = len(json.dumps(event['latest_validation_results']).encode('utf-8')) / (1024 * 1024)
                    logger.info(f"latest_validation_results contributes: {vr_size:.2f} MB")
        except Exception as size_check_error:
            logger.warning(f"Could not check payload size: {size_check_error}")
        
        # Strip generation_metadata from existing_config to prevent cache pollution
        clean_existing_config = event.get('existing_config')
        if clean_existing_config and 'generation_metadata' in clean_existing_config:
            clean_existing_config = clean_existing_config.copy()
            logger.info("BACKGROUND_HANDLER: Stripping generation_metadata from existing_config for config lambda to prevent cache issues")
            del clean_existing_config['generation_metadata']

        # Construct proper payload for config lambda (extract only the fields it expects)
        config_lambda_payload = {
            'table_analysis': event.get('table_analysis'),
            'existing_config': clean_existing_config,
            'instructions': event.get('instructions', 'Generate an optimal configuration for this data validation scenario'),
            'session_id': event.get('session_id', 'unknown'),
            'latest_validation_results': event.get('latest_validation_results'),
            'conversation_history': event.get('conversation_history', []),
            # Add table data sources that config lambda can use if table_analysis is missing
            'excel_s3_key': event.get('excel_s3_key'),
            'csv_s3_key': event.get('csv_s3_key'),
            'table_data': event.get('table_data')
        }

        logger.info(f"BACKGROUND_HANDLER: Constructed config lambda payload with keys: {list(config_lambda_payload.keys())}")
        logger.info(f"BACKGROUND_HANDLER: existing_config in payload: {bool(config_lambda_payload.get('existing_config'))}")
        logger.info(f"BACKGROUND_HANDLER: table_analysis in payload: {bool(config_lambda_payload.get('table_analysis'))}")
        logger.info(f"BACKGROUND_HANDLER: excel_s3_key in payload: {bool(config_lambda_payload.get('excel_s3_key'))}")
        logger.info(f"BACKGROUND_HANDLER: csv_s3_key in payload: {bool(config_lambda_payload.get('csv_s3_key'))}")
        logger.info(f"BACKGROUND_HANDLER: table_data in payload: {bool(config_lambda_payload.get('table_data'))}")

        # Invoke the lambda
        response = lambda_client.invoke(
            FunctionName=config_lambda_name,
            InvocationType='RequestResponse',
            Payload=json.dumps(config_lambda_payload)
        )
        
        # Parse response
        payload = json.loads(response['Payload'].read())
        
        if response['StatusCode'] == 200:
            # Parse the body if it's a string
            if isinstance(payload, dict) and 'body' in payload:
                body = payload['body']
                if isinstance(body, str):
                    return json.loads(body)
                else:
                    return body
            else:
                return payload
        else:
            logger.error(f"Config lambda invocation failed: {payload}")
            return {'success': False, 'error': 'Config lambda invocation failed'}
            
    except Exception as e:
        logger.error(f"Failed to invoke config lambda: {str(e)}")
        return {'success': False, 'error': f'Failed to invoke config lambda: {str(e)}'}

def create_config_download_url(s3_key: str) -> str:
    """Create S3 download URL for the generated configuration."""
    try:
        import boto3
        
        s3_client = boto3.client('s3')
        bucket_name = os.environ.get('S3_UNIFIED_BUCKET', 'perplexity-cache')
        
        # Generate presigned URL (valid for 1 hour)
        download_url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket_name, 'Key': s3_key},
            ExpiresIn=3600
        )
        
        logger.info(f"Created download URL for S3 key: {s3_key}")
        return download_url
        
    except Exception as e:
        logger.error(f"Failed to create config download URL: {str(e)}")
        return ""

# WebSocket message deduplication cache
_websocket_message_cache = {}
# Special cache for config generation completion messages
_config_completion_cache = {}

def _send_websocket_message_deduplicated(session_id: str, message: Dict, message_type: str = None):
    """Send a WebSocket message to a session with deduplication."""
    import hashlib
    import time
    
    try:
        current_time = time.time()
        
        # Special deduplication for config generation completion messages
        if message_type in ['config_generation_complete', 'config_generation_failed']:
            completion_key = f"{session_id}:{message_type}"
            if completion_key in _config_completion_cache:
                last_completion = _config_completion_cache[completion_key]
                if current_time - last_completion < 60:  # 60 second window for completion messages
                    logger.info(f"🚫 DEDUPLICATED config completion message for {session_id} (type: {message_type}) - last sent {current_time - last_completion:.1f}s ago")
                    return
            _config_completion_cache[completion_key] = current_time
            logger.debug(f"[DEDUP_CACHE] Cached config completion for {session_id}:{message_type}")
            
            # Clean old completion cache entries
            expired_completion_keys = [k for k, v in _config_completion_cache.items() if current_time - v > 120]
            for key in expired_completion_keys:
                del _config_completion_cache[key]
        
        # Create message fingerprint for general deduplication
        message_content = json.dumps(message, sort_keys=True)
        message_hash = hashlib.md5(f"{session_id}:{message_content}".encode()).hexdigest()
        
        # Check if this exact message was sent recently 
        # Use longer deduplication window for completion messages
        dedup_window = 30 if message_type in ['config_generation_complete', 'config_generation_failed'] else 5
        cache_key = f"{session_id}:{message_hash}"
        
        if cache_key in _websocket_message_cache:
            last_sent = _websocket_message_cache[cache_key]
            if current_time - last_sent < dedup_window:
                logger.info(f"🚫 DEDUPLICATED WebSocket message for {session_id} (type: {message_type}, window: {dedup_window}s)")
                return
        
        # Clean old cache entries (older than 60 seconds)
        expired_keys = [k for k, v in _websocket_message_cache.items() if current_time - v > 60]
        for key in expired_keys:
            del _websocket_message_cache[key]
        
        from dynamodb_schemas import get_connections_for_session
        
        connections = get_connections_for_session(session_id)
        if not connections:
            logger.warning(f"No WebSocket connections found for session {session_id}")
            return
        
        # Get API Gateway Management client
        client = _get_api_gateway_management_client(None)
        if not client:
            logger.error("Failed to get API Gateway Management client")
            return
        
        # Send to all connections
        sent_count = 0
        for connection_id in connections:
            try:
                client.post_to_connection(
                    ConnectionId=connection_id,
                    Data=json.dumps(message)
                )
                sent_count += 1
                # Reduce individual connection logging - already logged in summary
            except Exception as e:
                logger.error(f"Failed to send to connection {connection_id}: {e}")
                # Remove stale connection
                try:
                    from dynamodb_schemas import remove_websocket_connection
                    remove_websocket_connection(connection_id)
                except:
                    pass
        
        # Cache this message to prevent duplicates
        _websocket_message_cache[cache_key] = current_time
        # Only log WebSocket messages for important events or when multiple connections
        if sent_count > 1 or message_type in ['preview_progress_final', 'validation_complete', 'error']:
            logger.info(f"📤 WebSocket message sent to {sent_count} connections for {session_id} (type: {message_type})")
                    
    except Exception as e:
        logger.error(f"Failed to send WebSocket message: {e}")
        raise

def _send_websocket_message(session_id: str, message: Dict):
    """Legacy WebSocket sender - delegates to deduplicated version."""
    _send_websocket_message_deduplicated(session_id, message, "legacy")

def _update_session_info_with_account_data(email: str, session_id: str, account_info: Dict):
    """Update session_info.json with account information using unified session management."""
    try:
        from ..core.unified_s3_manager import UnifiedS3Manager
        
        storage_manager = UnifiedS3Manager()
        
        # Use unified session management to preserve version tracking data
        session_info = storage_manager.load_session_info(email, session_id)
        
        # Update with account information (preserves all existing data)
        session_info['account_info'] = account_info
        session_info['last_updated'] = datetime.now().isoformat()
        
        # Save using unified session management to preserve structure
        success = storage_manager.save_session_info(email, session_id, session_info)
        
        if success:
            logger.info(f"Updated session_info.json with account data for {session_id} (preserving version tracking)")
        else:
            logger.error(f"Failed to save session_info.json with account data for {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to update session_info with account data: {e}")