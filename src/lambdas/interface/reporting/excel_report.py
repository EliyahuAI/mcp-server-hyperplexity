"""
Functions for creating enhanced Excel reports with validation results.
"""
import io
import logging
from datetime import datetime
import sys
from pathlib import Path

# Add the project root to the Python path
ROOT_DIR = Path(__file__).resolve().parents[4]
sys.path.append(str(ROOT_DIR))

import openpyxl

# xlsxwriter is an optional dependency, so we handle its import gracefully.
try:
    import xlsxwriter
    EXCEL_ENHANCEMENT_AVAILABLE = True
except ImportError:
    EXCEL_ENHANCEMENT_AVAILABLE = False

from src.shared.schema_validator_simplified import SimplifiedSchemaValidator
from src.shared.row_key_utils import generate_row_key

logger = logging.getLogger()
logger.setLevel(logging.INFO)

def safe_for_excel(value):
    """Convert value to Excel-safe format, handling control characters but NOT XML escaping."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Check for NaN without pandas
        if value != value:  # NaN check
            return ""
        # Check for infinity
        if isinstance(value, float) and (value == float('inf') or value == float('-inf')):
            return ""
    
    # Convert to string for processing
    value_str = str(value)
    
    # First, handle control characters that are illegal in XML
    # Replace all control characters except tab (9), newline (10), and carriage return (13)
    cleaned = []
    for char in value_str:
        code = ord(char)
        if code < 32 and code not in (9, 10, 13):
            # Replace illegal control characters with space
            cleaned.append(' ')
        elif code > 127 and code < 160:
            # Replace non-breaking spaces and other problematic high ASCII
            cleaned.append(' ')
        elif code == 8232 or code == 8233:
            # Replace Unicode line/paragraph separators with regular newlines
            cleaned.append('\n')
        else:
            cleaned.append(char)
    value_str = ''.join(cleaned)
    
    # IMPORTANT: Do NOT escape XML characters here!
    # xlsxwriter handles XML escaping internally.
    # Double-escaping causes corruption.
    
    # Handle Excel's cell content limit
    if len(value_str) > 32767:
        return value_str[:32700] + "..."
    
    return value_str

def create_enhanced_excel_with_validation(excel_file_content, validation_results, config_data, session_id, skip_history=False):
    """Create color-coded Excel file with validation results and history tracking.
    
    Args:
        excel_file_content: Original Excel file content
        validation_results: Validation results from the validator
        config_data: Configuration data
        session_id: Session ID for tracking
        skip_history: If True, skip loading existing Details sheet (for fallback mode)
    """
    if not EXCEL_ENHANCEMENT_AVAILABLE:
        logger.warning("Enhanced Excel not available, skipping Excel creation")
        return None
        
    try:
        # Create Excel buffer
        excel_buffer = io.BytesIO()
        
        # Load original Excel data
        workbook = openpyxl.load_workbook(io.BytesIO(excel_file_content))
        
        # Select the appropriate sheet (same logic as invoke_validator_lambda)
        if 'Results' in workbook.sheetnames:
            worksheet = workbook['Results']
            logger.info(f"Using 'Results' sheet as data source for enhanced Excel creation")
        elif len(workbook.sheetnames) > 0:
            worksheet = workbook[workbook.sheetnames[0]]
            logger.info(f"Using first sheet '{worksheet.title}' as data source for enhanced Excel creation")
        else:
            worksheet = workbook.active
            logger.info(f"Using active sheet: {worksheet.title}")
        
        # Get headers and data
        headers = [cell.value for cell in worksheet[1]]
        
        # Convert to list of dictionaries 
        rows_data = []
        for row_idx in range(2, worksheet.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers):
                if header:
                    cell_value = worksheet.cell(row=row_idx, column=col_idx + 1).value
                    row_data[header] = str(cell_value) if cell_value is not None else ""
            rows_data.append(row_data)
        
        # Get ID fields from config for proper row key generation
        id_fields = []
        for target in config_data.get('validation_targets', []):
            if target.get('importance', '').upper() == 'ID':
                # Support both 'name' and 'column' fields
                field_name = target.get('name') or target.get('column')
                if field_name:
                    id_fields.append(field_name)
        
        # If no ID fields found, try to use SimplifiedSchemaValidator to determine primary keys
        if not id_fields:
            try:
                validator = SimplifiedSchemaValidator(config_data)
                id_fields = validator.primary_key
                logger.info(f"Using primary keys from SimplifiedSchemaValidator: {id_fields}")
            except ImportError:
                logger.warning("SimplifiedSchemaValidator not available in deployment package")
            except Exception as e:
                logger.warning(f"Could not use SimplifiedSchemaValidator: {e}")
        
        # Generate row keys for each row
        row_keys = []
        for row_data in rows_data:
            # We need to decide where to get generate_row_key from. For now, let's assume a fallback.
            # A better solution would be to pass it in or have it in a shared util.
            try:
                row_key = generate_row_key(row_data, id_fields)
            except ImportError:
                # Fallback to simple join
                key_columns = id_fields if id_fields else list(headers)[:3]
                row_key = "||".join([str(row_data.get(col, "")) for col in key_columns])
            row_keys.append(row_key)
        
        # Load existing Details entries from the original Excel (for history preservation)
        existing_details = []
        details_sheet_exists = False
        try:
            if 'Details' in workbook.sheetnames:
                details_sheet_exists = True
                details_worksheet = workbook['Details']
                
                # Read existing details headers
                details_headers = [cell.value for cell in details_worksheet[1]]
                
                # Read existing details data
                for row_idx in range(2, details_worksheet.max_row + 1):
                    detail_row = {}
                    for col_idx, header in enumerate(details_headers):
                        if header:
                            cell_value = details_worksheet.cell(row=row_idx, column=col_idx + 1).value
                            detail_row[header] = cell_value
                    
                    # Mark existing entries as "History"
                    if detail_row:
                        # If the row already has a 'New' column, preserve its value unless it's 'New'
                        if 'New' in detail_row and detail_row['New'] == 'New':
                            detail_row['New'] = 'History'
                        elif 'New' not in detail_row:
                            detail_row['New'] = 'History'
                        existing_details.append(detail_row)
                
                logger.info(f"Loaded {len(existing_details)} existing detail entries from Excel")
        except Exception as e:
            logger.warning(f"Could not load existing Details sheet: {e}")
        
        # Create Excel with xlsxwriter for advanced formatting
        with xlsxwriter.Workbook(excel_buffer, {'strings_to_urls': False, 'nan_inf_to_errors': True}) as workbook:
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'top',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            
            confidence_formats = {
                'HIGH': workbook.add_format({'bold': True, 'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'italic': True, 'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }
            
            # Create Results worksheet
            results_sheet = workbook.add_worksheet('Results')
            
            # Write headers
            for col_idx, col_name in enumerate(headers):
                results_sheet.write(0, col_idx, col_name, header_format)
                results_sheet.set_column(col_idx, col_idx, 20)  # Set column width
            
            # Write data rows with formatting
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row - try multiple key formats
                row_validation_data = None
                
                # First try the actual row key
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                # Then try position-based keys for backwards compatibility
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                for col_idx, col_name in enumerate(headers):
                    if not col_name:
                        continue
                        
                    cell_value = row_data.get(col_name, '')
                    validated_value = cell_value
                    confidence_level = None
                    quote = None
                    sources = None
                    
                    # Check if this column has validation results
                    if row_validation_data and isinstance(row_validation_data, dict):
                        if col_name in row_validation_data and isinstance(row_validation_data[col_name], dict):
                            field_data = row_validation_data[col_name]
                            validated_value = field_data.get('value', cell_value)
                            confidence_level = field_data.get('confidence_level')
                            quote = field_data.get('quote')
                            sources = field_data.get('sources', [])
                    
                    # Choose format based on confidence level (but not for ID fields)
                    cell_format = None
                    if col_name not in id_fields and confidence_level and confidence_level.upper() in confidence_formats:
                        cell_format = confidence_formats[confidence_level.upper()]
                    
                    # Write cell value
                    results_sheet.write(row_idx + 1, col_idx, safe_for_excel(validated_value or ''), cell_format)
                    
                    # Add comment if we have quote or sources
                    if quote or sources:
                        comment_parts = []
                        if quote and str(quote).strip() and quote != 'N/A':
                            comment_parts.append(f'Quote: "{quote}"')
                        if sources and isinstance(sources, list) and sources:
                            sources_text = ', '.join(sources[:3])  # Limit to first 3 sources
                            comment_parts.append(f'Sources: {sources_text}')
                        
                        if comment_parts:
                            comment_text = '\n\n'.join(comment_parts)
                            try:
                                results_sheet.write_comment(row_idx + 1, col_idx, comment_text, 
                                                          {'width': 300, 'height': 150})
                            except Exception as e:
                                logger.warning(f"Could not add comment: {e}")
            
            # Create Details worksheet with proper structure matching local version
            details_sheet = workbook.add_worksheet('Details')
            
            # Build detail headers dynamically to include ID fields
            detail_headers = ["Row Key", "Identifier"]
            
            # Add ID field columns without "ID:" prefix
            for id_field in id_fields:
                detail_headers.append(id_field)
            
            # Add the rest of the standard columns
            detail_headers.extend(["Column", "Original Value", "Validated Value", 
                            "Confidence", "Quote", "Sources", "Explanation", "Update Required", 
                            "Substantially Different", "Consistent with Model", "Model", "Timestamp", "New"])
            
            for col_idx, header in enumerate(detail_headers):
                details_sheet.write(0, col_idx, header, header_format)
            
            # Set column widths dynamically
            col_idx = 0
            details_sheet.set_column(col_idx, col_idx, 15)  # Row Key
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 30)  # Identifier
            col_idx += 1
            
            # ID field columns
            for _ in id_fields:
                details_sheet.set_column(col_idx, col_idx, 20)
                col_idx += 1
            
            # Standard columns
            details_sheet.set_column(col_idx, col_idx, 25)  # Column
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 30)  # Original Value
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 30)  # Validated Value
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 15)  # Confidence
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 50)  # Quote
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 40)  # Sources
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 60)  # Explanation
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 15)  # Update Required
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 20)  # Substantially Different
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 25)  # Consistent with Model
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 25)  # Model
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 20)  # Timestamp
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 10)  # New
            
            detail_row = 1
            current_timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            
            # Track which row keys have been processed to avoid duplicates
            processed_row_keys = set()
            
            # First, write NEW validation results
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row
                row_validation_data = None
                
                # First try the actual row key
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                # Then try position-based keys for backwards compatibility
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                if row_validation_data and isinstance(row_validation_data, dict):
                    # Create identifier from ID fields
                    identifier_parts = []
                    for id_field in id_fields:
                        if id_field in row_data:
                            identifier_parts.append(f"{id_field}: {row_data[id_field]}")
                    identifier = ", ".join(identifier_parts) if identifier_parts else f"Row {row_idx + 1}"
                    
                    for field_name, field_data in row_validation_data.items():
                        if isinstance(field_data, dict) and 'confidence_level' in field_data:
                            # Write detail row with all columns
                            col_idx = 0
                            details_sheet.write(detail_row, col_idx, safe_for_excel(row_key))  # Row Key
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(identifier))  # Identifier
                            col_idx += 1
                            
                            # Write ID field values
                            for id_field in id_fields:
                                value = row_data.get(id_field, '')
                                details_sheet.write(detail_row, col_idx, safe_for_excel(str(value)))
                                col_idx += 1
                            
                            # Write standard columns
                            details_sheet.write(detail_row, col_idx, safe_for_excel(field_name))  # Column
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(row_data.get(field_name, ''))))  # Original value
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('value', ''))))  # Validated value
                            col_idx += 1
                            
                            confidence = field_data.get('confidence_level', '')
                            confidence_format = confidence_formats.get(confidence.upper()) if confidence else None
                            details_sheet.write(detail_row, col_idx, safe_for_excel(confidence), confidence_format)
                            col_idx += 1
                            
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('quote', ''))))
                            col_idx += 1
                            sources_text = ', '.join(field_data.get('sources', []))
                            details_sheet.write(detail_row, col_idx, safe_for_excel(sources_text))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('explanation', ''))))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('update_required', ''))))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('substantially_different', ''))))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('consistent_with_model_knowledge', ''))))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('model', ''))))  # Model
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(current_timestamp))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, 'New')  # Mark all current validations as "New"
                            
                            # Track this combination as processed
                            processed_row_keys.add((row_key, field_name))
                            detail_row += 1
            
            # Then, append existing HISTORICAL details
            for existing_detail in existing_details:
                # Check if this is a duplicate of a new entry
                existing_row_key = existing_detail.get('Row Key', '')
                existing_column = existing_detail.get('Column', '')
                
                if (existing_row_key, existing_column) not in processed_row_keys:
                    # Write the historical entry
                    col_idx = 0
                    
                    # Row Key
                    details_sheet.write(detail_row, col_idx, safe_for_excel(str(existing_detail.get('Row Key', ''))))
                    col_idx += 1
                    
                    # Identifier
                    details_sheet.write(detail_row, col_idx, safe_for_excel(str(existing_detail.get('Identifier', ''))))
                    col_idx += 1
                    
                    # ID field values - extract from existing detail or try to reconstruct
                    for id_field in id_fields:
                        # Check for ID field with or without "ID:" prefix (for backwards compatibility)
                        value = ''
                        
                        # First try without prefix (new format)
                        if id_field in existing_detail:
                            value = existing_detail.get(id_field, '')
                        # Then try with prefix (old format)
                        elif f"ID:{id_field}" in existing_detail:
                            value = existing_detail.get(f"ID:{id_field}", '')
                        else:
                            # Try to extract from Row Key if not present
                            # This is a fallback for very old format Details sheets
                            if existing_row_key and '||' in existing_row_key:
                                # Try to parse the row key (assuming it follows the pattern)
                                key_parts = existing_row_key.split('||')
                                # This is approximate - we can't always reconstruct perfectly
                                id_field_index = id_fields.index(id_field) if id_field in id_fields else -1
                                if 0 <= id_field_index < len(key_parts):
                                    value = key_parts[id_field_index]
                        
                        details_sheet.write(detail_row, col_idx, safe_for_excel(str(value)))
                        col_idx += 1
                    
                    # Standard columns
                    standard_columns = ['Column', 'Original Value', 'Validated Value', 'Confidence', 
                                      'Quote', 'Sources', 'Explanation', 'Update Required', 
                                      'Substantially Different', 'Consistent with Model', 'Model', 'Timestamp', 'New']
                    
                    for col_name in standard_columns:
                        value = existing_detail.get(col_name, '')
                        
                        # Apply confidence formatting if applicable
                        if col_name == 'Confidence' and value and value.upper() in confidence_formats:
                            details_sheet.write(detail_row, col_idx, safe_for_excel(value), confidence_formats[value.upper()])
                        else:
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(value) if value is not None else ''))
                        
                        col_idx += 1
                    
                    detail_row += 1
            
            logger.info(f"Wrote {detail_row - 1} total detail entries (new + historical)")
            
            # Create Reasons worksheet
            reasons_sheet = workbook.add_worksheet('Reasons')
            reasons_headers = ["Row Key", "Field", "Explanation", "Update Required", "Substantially Different"]
            
            for col_idx, header in enumerate(reasons_headers):
                reasons_sheet.write(0, col_idx, header, header_format)
                reasons_sheet.set_column(col_idx, col_idx, 30)
            
            reasons_row = 1
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row
                row_validation_data = None
                
                # First try the actual row key
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                # Then try position-based keys for backwards compatibility
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                if row_validation_data and isinstance(row_validation_data, dict):
                    for field_name, field_data in row_validation_data.items():
                        if isinstance(field_data, dict) and 'explanation' in field_data:
                            reasons_sheet.write(reasons_row, 0, safe_for_excel(row_key))  # Use actual row key
                            reasons_sheet.write(reasons_row, 1, safe_for_excel(field_name))
                            reasons_sheet.write(reasons_row, 2, safe_for_excel(str(field_data.get('explanation', ''))))
                            reasons_sheet.write(reasons_row, 3, safe_for_excel(str(field_data.get('update_required', ''))))
                            reasons_sheet.write(reasons_row, 4, safe_for_excel(str(field_data.get('substantially_different', ''))))
                            
                            reasons_row += 1
        
        excel_buffer.seek(0)
        return excel_buffer
        
    except Exception as e:
        logger.error(f"Error creating enhanced Excel: {str(e)}")
        return None 