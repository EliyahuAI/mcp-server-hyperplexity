"""
Functions for creating enhanced Excel reports with validation results (3-sheet structure).
"""
import io
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

# xlsxwriter is an optional dependency, so we handle its import gracefully.
try:
    import xlsxwriter
    EXCEL_ENHANCEMENT_AVAILABLE = True
except ImportError:
    EXCEL_ENHANCEMENT_AVAILABLE = False

from schema_validator_simplified import SimplifiedSchemaValidator
from row_key_utils import generate_row_key

logger = logging.getLogger()
logger.setLevel(logging.INFO)

def safe_for_excel(value):
    """Convert value to Excel-safe format, handling control characters but NOT XML escaping."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Check for NaN without pandas
        if value != value:  # NaN check
            return ""
        # Check for infinity
        if isinstance(value, float) and (value == float('inf') or value == float('-inf')):
            return ""
    
    # Convert to string for processing
    value_str = str(value)
    
    # First, handle control characters that are illegal in XML
    # Replace all control characters except tab (9), newline (10), and carriage return (13)
    cleaned = []
    for char in value_str:
        code = ord(char)
        if code < 32 and code not in (9, 10, 13):
            # Replace illegal control characters with space
            cleaned.append(' ')
        elif code > 127 and code < 160:
            # Replace non-breaking spaces and other problematic high ASCII
            cleaned.append(' ')
        elif code == 8232 or code == 8233:
            # Replace Unicode line/paragraph separators with regular newlines
            cleaned.append('\n')
        else:
            cleaned.append(char)
    value_str = ''.join(cleaned)
    
    # Handle Excel's cell content limit
    if len(value_str) > 32767:
        return value_str[:32700] + "..."
    
    return value_str

def is_null_confidence(confidence):
    """Check if confidence should be treated as null (no coloring)."""
    if confidence is None:
        return True
    confidence_str = str(confidence).strip()
    return confidence_str == '' or confidence_str == '-' or confidence_str.lower() == 'null'

def get_confidence_format(confidence, format_dict):
    """Get confidence format, handling null values properly."""
    if is_null_confidence(confidence):
        return None
    confidence_str = str(confidence).strip().upper()
    return format_dict.get(confidence_str)

def should_update_value(original_confidence, validation_confidence):
    """
    Determine if original value should be updated based on confidence comparison.
    Only update when validation confidence is higher than or equal to original confidence.
    
    Confidence hierarchy: HIGH > MEDIUM > LOW > None
    Special case: If original has no confidence (blank/empty values), any validation confidence is better.
    """
    # If we have no validation confidence, never update
    if is_null_confidence(validation_confidence):
        return False
    
    # If original has no confidence but we have validation confidence, always update
    # (any confidence is better than no confidence for blank/empty values)
    if is_null_confidence(original_confidence):
        return True
    
    # Define confidence hierarchy (higher number = higher confidence)
    confidence_levels = {
        'HIGH': 3,
        'MEDIUM': 2, 
        'LOW': 1
    }
    
    original_level = confidence_levels.get(str(original_confidence).strip().upper(), 0)
    validation_level = confidence_levels.get(str(validation_confidence).strip().upper(), 0)
    
    # Only update if validation confidence is higher than or equal to original
    return validation_level >= original_level

def create_enhanced_excel_with_validation(excel_data, validation_results, config_data, session_id, skip_history=False, validated_sheet_name=None):
    """Create 3-sheet Excel file with validation results.
    
    Args:
        excel_data: Structured table data from shared_table_parser (replaces excel_file_content)
        validation_results: Validation results from the validator
        config_data: Configuration data
        session_id: Session ID for tracking
        skip_history: If True, skip loading existing Details sheet (for fallback mode)
        validated_sheet_name: Name of the sheet that was actually validated (from metadata)
    """
    if not EXCEL_ENHANCEMENT_AVAILABLE:
        logger.warning("Enhanced Excel not available, skipping Excel creation")
        return None
        
    try:
        # Create Excel buffer
        excel_buffer = io.BytesIO()
        
        # Debug logging for excel_data type and content
        logger.info(f"Excel data type: {type(excel_data)}")
        logger.info(f"Excel data is dict: {isinstance(excel_data, dict)}")
        if isinstance(excel_data, dict):
            logger.info(f"Excel data keys: {list(excel_data.keys())}")
        else:
            logger.info(f"Excel data length: {len(excel_data) if hasattr(excel_data, '__len__') else 'N/A'}")
        
        # Extract structured data (already parsed by shared_table_parser)
        if isinstance(excel_data, dict) and excel_data.get('column_names') and excel_data.get('data'):
            headers = excel_data.get('column_names', [])
            rows_data = excel_data.get('data', [])
            # Handle sheet name for both CSV and Excel files
            metadata = excel_data.get('metadata', {})
            file_type = metadata.get('file_type', 'unknown')
            
            if file_type == 'csv':
                actual_sheet_name = 'CSV Data'  # CSV files don't have sheet names
            else:
                actual_sheet_name = validated_sheet_name or metadata.get('sheet_name', 'Unknown')
            
            logger.info(f"Using structured data from validated source: '{actual_sheet_name}' (file_type: {file_type})")
        else:
            # Fallback for backward compatibility (if raw content is still passed)
            logger.warning(f"Excel data is not structured dict, attempting fallback with raw data processing")
            if not isinstance(excel_data, (bytes, io.BytesIO)):
                logger.error(f"Excel data is neither structured dict nor bytes/BytesIO, cannot process. Type: {type(excel_data)}")
                return None
            
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(excel_data) if isinstance(excel_data, bytes) else excel_data)
            except Exception as openpyxl_error:
                logger.error(f"Failed to load Excel data with openpyxl: {openpyxl_error}")
                return None
            
            # Use validated_sheet_name if provided, otherwise fall back to old logic
            if validated_sheet_name and validated_sheet_name in workbook.sheetnames:
                worksheet = workbook[validated_sheet_name]
                logger.info(f"Using validated sheet '{validated_sheet_name}' for enhanced Excel creation")
            elif 'Results' in workbook.sheetnames:
                worksheet = workbook['Results']
                logger.info(f"Using 'Results' sheet as data source for enhanced Excel creation")
            elif len(workbook.sheetnames) > 0:
                worksheet = workbook[workbook.sheetnames[0]]
                logger.info(f"Using first sheet '{worksheet.title}' as data source for enhanced Excel creation")
            else:
                worksheet = workbook.active
                logger.info(f"Using active sheet: {worksheet.title}")
            
            # Get headers and data
            headers = [cell.value for cell in worksheet[1]]
            
            # Convert to list of dictionaries 
            rows_data = []
            for row_idx in range(2, worksheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    if header:
                        cell_value = worksheet.cell(row=row_idx, column=col_idx + 1).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                rows_data.append(row_data)
        
        # Get ID fields from config for proper row key generation
        id_fields = []
        for target in config_data.get('validation_targets', []):
            if target.get('importance', '').upper() == 'ID':
                field_name = target.get('name') or target.get('column')
                if field_name:
                    id_fields.append(field_name)
        
        # Generate row keys for each row
        row_keys = []
        for row_data in rows_data:
            try:
                from row_key_utils import generate_row_key
                row_key = generate_row_key(row_data, id_fields)
            except ImportError:
                # Fallback to simple join
                key_columns = id_fields if id_fields else list(headers)[:3]
                row_key = "||".join([str(row_data.get(col, "")) for col in key_columns])
            row_keys.append(row_key)
        
        # Load existing Details entries from the original Excel (for history preservation)
        existing_details = []
        details_sheet_exists = False
        if not skip_history and not isinstance(excel_data, dict):
            # Only try to load existing details if we have raw Excel data, not structured data
            try:
                if 'Details' in workbook.sheetnames:
                    details_sheet_exists = True
                    details_worksheet = workbook['Details']
                    
                    # Read existing details headers
                    details_headers = [cell.value for cell in details_worksheet[1]]
                    
                    # Read existing details data
                    for row_idx in range(2, details_worksheet.max_row + 1):
                        detail_row = {}
                        for col_idx, header in enumerate(details_headers):
                            if header:
                                cell_value = details_worksheet.cell(row=row_idx, column=col_idx + 1).value
                                detail_row[header] = cell_value
                        
                        # Mark existing entries as "History"
                        if detail_row:
                            if 'New' in detail_row and detail_row['New'] == 'New':
                                detail_row['New'] = 'History'
                            elif 'New' not in detail_row:
                                detail_row['New'] = 'History'
                            existing_details.append(detail_row)
                    
                    logger.info(f"Loaded {len(existing_details)} existing detail entries from Excel")
            except Exception as e:
                logger.warning(f"Could not load existing Details sheet: {e}")
        
        # Create Excel with xlsxwriter for advanced formatting
        with xlsxwriter.Workbook(excel_buffer, {'strings_to_urls': False, 'nan_inf_to_errors': True}) as workbook:
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'top',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            
            # Original confidence formats (HIGH/MEDIUM/LOW) - use same colors as validation confidence
            original_confidence_formats = {
                'HIGH': workbook.add_format({'bold': True, 'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'italic': True, 'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }
            
            # Validation confidence formats (HIGH/MEDIUM/LOW)
            validation_confidence_formats = {
                'HIGH': workbook.add_format({'bold': True, 'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'italic': True, 'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }
            
            # SHEET 1: Updated Values
            updated_sheet = workbook.add_worksheet('Updated Values')
            
            # Headers for updated values sheet (no additional columns)
            updated_headers = headers
            for col_idx, col_name in enumerate(updated_headers):
                updated_sheet.write(0, col_idx, col_name, header_format)
                updated_sheet.set_column(col_idx, col_idx, 20)
            
            # Write all rows to updated sheet (not just rows with changes)
            updated_row_idx = 1
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row
                row_validation_data = None
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                # Write all rows (not just those with updates)
                if True:  # Process all rows
                    # Write updated values for this row
                    
                    for col_idx, col_name in enumerate(headers):
                        original_value = row_data.get(col_name, '')
                        updated_value = original_value
                        
                        if row_validation_data and col_name in row_validation_data:
                            field_data = row_validation_data[col_name]
                            if isinstance(field_data, dict):
                                original_confidence = field_data.get('original_confidence')
                                validation_confidence = field_data.get('confidence_level', field_data.get('confidence', ''))
                                
                                # Only update if validation confidence is higher than original confidence
                                # Now properly handles case where original has no confidence (blank values)
                                if should_update_value(original_confidence, validation_confidence):
                                    updated_value = field_data.get('value', original_value)
                        
                        # Apply validation confidence formatting to updated values
                        validation_confidence = None
                        if row_validation_data and col_name in row_validation_data:
                            field_data = row_validation_data[col_name]
                            if isinstance(field_data, dict):
                                validation_confidence = field_data.get('confidence_level', field_data.get('confidence', ''))
                        
                        cell_format = get_confidence_format(validation_confidence, validation_confidence_formats)
                        updated_sheet.write(updated_row_idx, col_idx, safe_for_excel(updated_value), cell_format)
                        
                        # Add comment with original value and reasoning (same as Original Values sheet)
                        comment_text = None
                        if row_validation_data and col_name in row_validation_data:
                            field_data = row_validation_data[col_name]
                            if isinstance(field_data, dict):
                                original_value = row_data.get(col_name, '')
                                validated_value = field_data.get('value', '')
                                reasoning = field_data.get('reasoning', '')
                                
                                # Create comment with original value and reasoning
                                if validated_value != original_value or reasoning:
                                    comment_parts = []
                                    if validated_value != original_value:
                                        comment_parts.append(f'Original Value: {original_value}')
                                    if reasoning:
                                        comment_parts.append(f'Supporting Information: {reasoning}')
                                    
                                    if comment_parts:
                                        comment_text = '\\n\\n'.join(comment_parts)
                        
                        # Add comment if needed
                        if comment_text:
                            try:
                                updated_sheet.write_comment(updated_row_idx, col_idx, comment_text,
                                                           {'width': 300, 'height': 150})
                            except Exception as e:
                                logger.warning(f"Could not add comment to Updated Values sheet: {e}")
                    
                    # No additional columns to write (removed Original Value and Supporting Information)
                    
                    updated_row_idx += 1
            
            # SHEET 2: Original Values
            original_sheet = workbook.add_worksheet('Original Values')
            
            # Write headers
            for col_idx, col_name in enumerate(headers):
                original_sheet.write(0, col_idx, col_name, header_format)
                original_sheet.set_column(col_idx, col_idx, 20)
            
            # Write original data with confidence-based coloring and comments
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row
                row_validation_data = None
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                for col_idx, col_name in enumerate(headers):
                    if not col_name:
                        continue
                        
                    original_value = row_data.get(col_name, '')
                    cell_format = None
                    comment_text = None
                    
                    # Check if this column has validation results
                    if row_validation_data and isinstance(row_validation_data, dict):
                        if col_name in row_validation_data and isinstance(row_validation_data[col_name], dict):
                            field_data = row_validation_data[col_name]
                            original_confidence = field_data.get('original_confidence')
                            validated_value = field_data.get('value', '')
                            reasoning = field_data.get('reasoning', '')
                            
                            # Apply original confidence color (only if confidence should be colored)
                            cell_format = get_confidence_format(original_confidence, original_confidence_formats)
                            
                            # Create comment with updated value and reasoning
                            if validated_value != original_value or reasoning:
                                comment_parts = []
                                if validated_value != original_value:
                                    comment_parts.append(f'Updated Value: {validated_value}')
                                if reasoning:
                                    comment_parts.append(f'Supporting Information: {reasoning}')
                                
                                if comment_parts:
                                    comment_text = '\\n\\n'.join(comment_parts)
                    
                    # Write original value
                    original_sheet.write(row_idx + 1, col_idx, safe_for_excel(original_value), cell_format)
                    
                    # Add comment if needed
                    if comment_text:
                        try:
                            original_sheet.write_comment(row_idx + 1, col_idx, comment_text,
                                                       {'width': 300, 'height': 150})
                        except Exception as e:
                            logger.warning(f"Could not add comment: {e}")
            
            # SHEET 3: Details (comprehensive view)
            details_sheet = workbook.add_worksheet('Details')
            
            # Build detail headers dynamically to include ID fields
            detail_headers = ["Row Key", "Identifier"]
            
            # Add ID field columns
            for id_field in id_fields:
                detail_headers.append(id_field)
            
            # Add the rest of the standard columns (new order as requested)
            detail_headers.extend(["Column", "Original Value", "Original Confidence", "Validated Value", 
                            "Validation Confidence", "Reasoning", "Sources", 
                            "Explanation", "Consistent with Model", "Model", "Timestamp", "New"])
            
            for col_idx, header in enumerate(detail_headers):
                details_sheet.write(0, col_idx, header, header_format)
                details_sheet.set_column(col_idx, col_idx, 20)
            
            detail_row = 1
            current_timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            
            # Track which row keys have been processed to avoid duplicates
            processed_row_keys = set()
            
            # Write NEW validation results to details sheet
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row
                row_validation_data = None
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                if row_validation_data and isinstance(row_validation_data, dict):
                    # Create identifier from ID fields
                    identifier_parts = []
                    for id_field in id_fields:
                        if id_field in row_data:
                            identifier_parts.append(f"{id_field}: {row_data[id_field]}")
                    identifier = ", ".join(identifier_parts) if identifier_parts else f"Row {row_idx + 1}"
                    
                    for field_name, field_data in row_validation_data.items():
                        if isinstance(field_data, dict) and 'confidence' in field_data:
                            col_idx = 0
                            details_sheet.write(detail_row, col_idx, safe_for_excel(row_key))  # Row Key
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(identifier))  # Identifier
                            col_idx += 1
                            
                            # Write ID field values
                            for id_field in id_fields:
                                value = row_data.get(id_field, '')
                                details_sheet.write(detail_row, col_idx, safe_for_excel(str(value)))
                                col_idx += 1
                            
                            # Write standard columns in new order: Column, Original Value, Original Confidence, Validated Value, Validation Confidence
                            details_sheet.write(detail_row, col_idx, safe_for_excel(field_name))  # Column
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(row_data.get(field_name, ''))))  # Original value
                            col_idx += 1
                            
                            # Original confidence
                            original_confidence = field_data.get('original_confidence', '')
                            original_confidence_str = str(original_confidence) if original_confidence else ''
                            original_format = get_confidence_format(original_confidence, original_confidence_formats)
                            details_sheet.write(detail_row, col_idx, safe_for_excel(original_confidence_str), original_format)
                            col_idx += 1
                            
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('value', ''))))  # Validated value
                            col_idx += 1
                            
                            # Validation confidence
                            validation_confidence = field_data.get('confidence_level', field_data.get('confidence', ''))  # Try both field names
                            validation_confidence_str = str(validation_confidence) if validation_confidence else ''
                            confidence_format = get_confidence_format(validation_confidence, validation_confidence_formats)
                            details_sheet.write(detail_row, col_idx, safe_for_excel(validation_confidence_str), confidence_format)
                            col_idx += 1
                            
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('reasoning', ''))))  # Reasoning
                            col_idx += 1
                            sources_text = ', '.join(field_data.get('sources', []))
                            details_sheet.write(detail_row, col_idx, safe_for_excel(sources_text))  # Sources
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('explanation', ''))))  # Explanation
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('consistent_with_model_knowledge', ''))))  # Consistent with Model
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('model', ''))))  # Model
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(current_timestamp))  # Timestamp
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, 'New')  # Mark as new
                            
                            # Track this combination as processed
                            processed_row_keys.add((row_key, field_name))
                            detail_row += 1
            
            # Then, append existing HISTORICAL details
            for existing_detail in existing_details:
                # Check if this is a duplicate of a new entry
                existing_row_key = existing_detail.get('Row Key', '')
                existing_column = existing_detail.get('Column', '')
                
                if (existing_row_key, existing_column) not in processed_row_keys:
                    # Write the historical entry
                    col_idx = 0
                    
                    # Row Key
                    details_sheet.write(detail_row, col_idx, safe_for_excel(str(existing_detail.get('Row Key', ''))))
                    col_idx += 1
                    
                    # Identifier
                    details_sheet.write(detail_row, col_idx, safe_for_excel(str(existing_detail.get('Identifier', ''))))
                    col_idx += 1
                    
                    # ID field values - extract from existing detail or try to reconstruct
                    for id_field in id_fields:
                        value = existing_detail.get(id_field, '')
                        details_sheet.write(detail_row, col_idx, safe_for_excel(str(value)))
                        col_idx += 1
                    
                    # Standard columns (map old field names to new ones)
                    standard_columns = ['Column', 'Original Value', 'Validated Value', 'Validation Confidence', 
                                      'Original Confidence', 'Reasoning', 'Sources', 'Explanation', 
                                      'Consistent with Model', 'Model', 'Timestamp', 'New']
                    
                    # Handle field name mapping
                    field_mapping = {
                        'Confidence': 'Validation Confidence',
                        'Quote': 'Reasoning'
                    }
                    
                    for col_name in standard_columns:
                        # Try the new column name first, then check mappings
                        value = existing_detail.get(col_name, '')
                        if not value and col_name in field_mapping.values():
                            # Try the old field name
                            old_name = [k for k, v in field_mapping.items() if v == col_name]
                            if old_name:
                                value = existing_detail.get(old_name[0], '')
                        
                        # Apply confidence formatting if applicable
                        if 'Confidence' in col_name and col_name != 'Original Confidence':
                            # Validation confidence
                            confidence_format = get_confidence_format(value, validation_confidence_formats)
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(value) if value is not None else ''), confidence_format)
                        elif col_name == 'Original Confidence':
                            # Original confidence
                            confidence_format = get_confidence_format(value, original_confidence_formats)
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(value) if value is not None else ''), confidence_format)
                        else:
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(value) if value is not None else ''))
                        
                        col_idx += 1
                    
                    detail_row += 1
            
            logger.info(f"Created 3-sheet Excel with {detail_row - 1} total detail entries (new + historical)")
        
        excel_buffer.seek(0)
        return excel_buffer
        
    except Exception as e:
        logger.error(f"Error creating enhanced Excel: {str(e)}")
        import traceback
        logger.error(f"Enhanced Excel creation traceback: {traceback.format_exc()}")
        return None