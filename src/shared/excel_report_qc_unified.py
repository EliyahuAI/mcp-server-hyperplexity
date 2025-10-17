"""
Functions for creating enhanced Excel reports with validation results (3-sheet structure).
"""
import io
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

# xlsxwriter is an optional dependency, so we handle its import gracefully.
try:
    import xlsxwriter
    EXCEL_ENHANCEMENT_AVAILABLE = True
except ImportError:
    EXCEL_ENHANCEMENT_AVAILABLE = False

from schema_validator_simplified import SimplifiedSchemaValidator
from row_key_utils import generate_row_key

logger = logging.getLogger()
logger.setLevel(logging.INFO)


def safe_for_excel(value, preserve_formulas=False):
    """Convert value to Excel-safe format, handling control characters, Unicode bullets, and complex QC content.

    Args:
        value: The value to make Excel-safe
        preserve_formulas: If True, don't escape formulas (for IGNORE columns that should retain original formulas)
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Check for NaN without pandas
        if value != value:  # NaN check
            return ""
        # Check for infinity
        if isinstance(value, float) and (value == float('inf') or value == float('-inf')):
            return ""

    # Convert to string for processing
    value_str = str(value)

    # --- Unified Formula Handling ---
    if value_str.startswith('='):
        # If it's an external formula, always escape it as a string with a note.
        if '[' in value_str:
            return f"(external source) {value_str}"

        # If we are preserving internal formulas, return them as is (with length check).
        if preserve_formulas:
            if len(value_str) > 32767:
                logger.warning(f"Formula exceeds Excel cell limit ({len(value_str)} chars): {value_str[:100]}...")
                return value_str[:32767]
            return value_str
        
        # Otherwise (not preserving internal formulas), escape with a single quote.
        return "'" + value_str

    # Handle very long content first - be more conservative for QC data
    if len(value_str) > 10000:
        # For very long QC content, truncate more aggressively
        value_str = value_str[:10000] + "... [Content truncated for Excel compatibility]"

    # Replace problematic Unicode characters commonly found in QC content
    replacements = {
        '•': '- ',     # Replace Unicode bullet with ASCII dash
        '–': '-',      # Replace en-dash with regular dash
        '—': '-',      # Replace em-dash with regular dash
        ''': "'",      # Replace smart quote
        ''': "'",      # Replace smart quote
        '"': '"',      # Replace smart quote
        '"': '"',      # Replace smart quote
        '…': '...',    # Replace ellipsis
        '\u00a0': ' ', # Replace non-breaking space
        '\u2022': '- ', # Replace bullet point
        '\u2013': '-', # Replace en dash
        '\u2014': '-', # Replace em dash
        '\u2018': "'", # Replace left single quote
        '\u2019': "'", # Replace right single quote
        '\u201c': '"', # Replace left double quote
        '\u201d': '"', # Replace right double quote
        '\u2026': '...', # Replace horizontal ellipsis
    }

    for unicode_char, replacement in replacements.items():
        value_str = value_str.replace(unicode_char, replacement)

    # Clean control characters that are illegal in XML
    # Replace all control characters except tab (9), newline (10), and carriage return (13)
    cleaned = []
    for char in value_str:
        code = ord(char)
        if code < 32 and code not in (9, 10, 13):
            # Replace illegal control characters with space
            cleaned.append(' ')
        elif code > 127 and code < 160:
            # Replace problematic high ASCII
            cleaned.append(' ')
        elif code == 8232 or code == 8233:
            # Replace Unicode line/paragraph separators with regular newlines
            cleaned.append('\n')
        elif code > 65535:
            # Replace characters outside BMP (Basic Multilingual Plane) that can cause Excel issues
            cleaned.append('?')
        else:
            cleaned.append(char)
    value_str = ''.join(cleaned)

    # Normalize excessive whitespace and newlines that can cause Excel issues
    import re
    # Replace multiple consecutive newlines with maximum of 2
    value_str = re.sub(r'\n{3,}', '\n\n', value_str)
    # Replace multiple consecutive spaces with single space
    value_str = re.sub(r' {3,}', ' ', value_str)

    # Final Excel cell content limit check
    if len(value_str) > 32767:
        return value_str[:32700] + "... [Excel limit reached]"

    return value_str

def is_null_confidence(confidence):
    """Check if confidence should be treated as null (no coloring)."""
    if confidence is None:
        return True
    confidence_str = str(confidence).strip()
    return confidence_str == '' or confidence_str == '-' or confidence_str.lower() == 'null'

def get_confidence_format(confidence, format_dict):
    """Get confidence format, handling null values properly."""
    if is_null_confidence(confidence):
        return None
    confidence_str = str(confidence).strip().upper()
    return format_dict.get(confidence_str)

def get_qc_confidence_format(qc_data, qc_confidence_formats):
    """Get QC format based on confidence level (same styling as validation)."""
    if not qc_data or not isinstance(qc_data, dict):
        return None

    # Check for QC confidence first, then fall back to updated confidence
    qc_confidence = qc_data.get('qc_confidence') or qc_data.get('updated_confidence', '')
    if is_null_confidence(qc_confidence):
        return None
    confidence_str = str(qc_confidence).strip().upper()
    return qc_confidence_formats.get(confidence_str)

def should_update_value(original_confidence, validation_confidence):
    """
    Determine if original value should be updated based on confidence comparison.
    Only update when validation confidence is higher than or equal to original confidence.
    
    Confidence hierarchy: HIGH > MEDIUM > LOW > None
    Special case: If original has no confidence (blank/empty values), any validation confidence is better.
    """
    # If we have no validation confidence, never update
    if is_null_confidence(validation_confidence):
        return False
    
    # If original has no confidence but we have validation confidence, always update
    # (any confidence is better than no confidence for blank/empty values)
    if is_null_confidence(original_confidence):
        return True
    
    # Define confidence hierarchy (higher number = higher confidence)
    confidence_levels = {
        'HIGH': 3,
        'MEDIUM': 2, 
        'LOW': 1
    }
    
    original_level = confidence_levels.get(str(original_confidence).strip().upper(), 0)
    validation_level = confidence_levels.get(str(validation_confidence).strip().upper(), 0)
    
    # Only update if validation confidence is higher than or equal to original
    return validation_level >= original_level

def calculate_confidence_distribution(validation_results, qc_results):
    """
    Calculate percentage distribution of confidence levels.
    Uses QC-adjusted confidences when QC was applied.

    Args:
        validation_results: Dictionary of validation results keyed by row hash
        qc_results: Optional dictionary of QC results keyed by row hash

    Returns:
        Tuple of (original_str, updated_str) with percentage distributions
        e.g., ("L: 15%, M: 45%, H: 40%", "L: 5%, M: 35%, H: 60%")
    """
    original_counts = {'LOW': 0, 'MEDIUM': 0, 'HIGH': 0}
    updated_counts = {'LOW': 0, 'MEDIUM': 0, 'HIGH': 0}
    total_fields = 0

    for row_key, row_results in validation_results.items():
        if not isinstance(row_results, dict):
            continue

        for field, field_data in row_results.items():
            if not isinstance(field_data, dict):
                continue

            total_fields += 1

            # Use QC-adjusted confidence if available
            if qc_results and row_key in qc_results and field in qc_results[row_key]:
                qc_field = qc_results[row_key][field]
                if isinstance(qc_field, dict):
                    original_conf = qc_field.get('qc_original_confidence', field_data.get('original_confidence'))
                    updated_conf = qc_field.get('qc_confidence', field_data.get('confidence_level'))
                else:
                    original_conf = field_data.get('original_confidence')
                    updated_conf = field_data.get('confidence_level')
            else:
                original_conf = field_data.get('original_confidence')
                updated_conf = field_data.get('confidence_level')

            # Count confidences (handling various confidence formats)
            if original_conf:
                original_conf_upper = str(original_conf).strip().upper()
                if original_conf_upper in original_counts:
                    original_counts[original_conf_upper] += 1

            if updated_conf:
                updated_conf_upper = str(updated_conf).strip().upper()
                if updated_conf_upper in updated_counts:
                    updated_counts[updated_conf_upper] += 1

    # Format as percentages (handle division by zero)
    if total_fields > 0:
        original_str = f"L: {original_counts['LOW']/total_fields*100:.0f}%, M: {original_counts['MEDIUM']/total_fields*100:.0f}%, H: {original_counts['HIGH']/total_fields*100:.0f}%"
        updated_str = f"L: {updated_counts['LOW']/total_fields*100:.0f}%, M: {updated_counts['MEDIUM']/total_fields*100:.0f}%, H: {updated_counts['HIGH']/total_fields*100:.0f}%"
    else:
        original_str = "L: 0%, M: 0%, H: 0%"
        updated_str = "L: 0%, M: 0%, H: 0%"

    return original_str, updated_str

def create_validation_record_sheet(workbook, header_format, validation_results, qc_results,
                                   session_id, config_s3_key, rows_data, headers,
                                   existing_validation_record=None, is_preview=False, run_key=None):
    """
    Create Validation Record sheet with run-level metadata.

    Args:
        workbook: xlsxwriter workbook object
        header_format: Format for headers
        validation_results: Validation results dictionary
        qc_results: QC results dictionary (optional)
        session_id: Session ID
        config_s3_key: S3 key for configuration file
        rows_data: List of row dictionaries
        headers: List of column headers
        existing_validation_record: Existing validation record data (list of dicts)
        is_preview: Whether this is a preview run
        run_key: DynamoDB run key (e.g., #Preview_34t345345346346)

    Returns:
        Validation Record worksheet object
    """
    from datetime import datetime, timezone
    import time

    # Create the worksheet
    validation_record_sheet = workbook.add_worksheet('Validation Record')

    # Define headers
    record_headers = [
        'Run_Number', 'Run_Time', 'Session_ID', 'Configuration_ID',
        'Run_Key', 'Rows', 'Columns', 'Original_Confidences', 'Updated_Confidences'
    ]

    # Write headers
    for col_idx, header in enumerate(record_headers):
        validation_record_sheet.write(0, col_idx, header, header_format)
        validation_record_sheet.set_column(col_idx, col_idx, 20)

    # Calculate confidence distributions
    original_conf_str, updated_conf_str = calculate_confidence_distribution(
        validation_results, qc_results
    )

    # Generate run metadata
    run_time = datetime.now(timezone.utc).isoformat()
    # Use the actual run_key from DynamoDB or fallback to session_id if not provided
    if not run_key:
        run_key = f"{session_id}_{int(time.time())}"
        logger.warning(f"[VALIDATION_RECORD] No run_key provided, using fallback: {run_key}")
    total_rows = len(rows_data)
    total_columns = len(headers)

    # Determine run number and whether to overwrite
    run_number = 1
    should_overwrite = False

    if existing_validation_record and len(existing_validation_record) > 0:
        # Check if we should overwrite Run_Number 1 (preview -> full validation)
        first_record = existing_validation_record[0]
        if (first_record.get('Run_Number') == 1 and
            first_record.get('Session_ID') == session_id and
            not is_preview):
            # Full validation following a preview - overwrite run 1
            run_number = 1
            should_overwrite = True
            logger.info(f"[VALIDATION_RECORD] Overwriting Run_Number 1 (preview -> full validation)")
        elif is_preview:
            # Preview run - always overwrite run 1 if it exists
            run_number = 1
            should_overwrite = True
            logger.info(f"[VALIDATION_RECORD] Overwriting Run_Number 1 (preview)")
        else:
            # Append as new run
            max_run_number = max([r.get('Run_Number', 0) for r in existing_validation_record], default=0)
            run_number = max_run_number + 1
            logger.info(f"[VALIDATION_RECORD] Appending new run with Run_Number {run_number}")

    # Write validation record entries
    row_idx = 1

    if should_overwrite and existing_validation_record:
        # Skip the first record (will be replaced by new run)
        records_to_write = existing_validation_record[1:]
    else:
        records_to_write = existing_validation_record or []

    # Write new run first
    validation_record_sheet.write(row_idx, 0, run_number)
    validation_record_sheet.write(row_idx, 1, run_time)
    validation_record_sheet.write(row_idx, 2, session_id)
    validation_record_sheet.write(row_idx, 3, config_s3_key or '')
    validation_record_sheet.write(row_idx, 4, run_key)
    validation_record_sheet.write(row_idx, 5, total_rows)
    validation_record_sheet.write(row_idx, 6, total_columns)
    validation_record_sheet.write(row_idx, 7, original_conf_str)
    validation_record_sheet.write(row_idx, 8, updated_conf_str)
    row_idx += 1

    # Write historical records
    for record in records_to_write:
        validation_record_sheet.write(row_idx, 0, record.get('Run_Number', ''))
        validation_record_sheet.write(row_idx, 1, record.get('Run_Time', ''))
        validation_record_sheet.write(row_idx, 2, record.get('Session_ID', ''))
        validation_record_sheet.write(row_idx, 3, record.get('Configuration_ID', ''))
        validation_record_sheet.write(row_idx, 4, record.get('Run_Key', ''))
        validation_record_sheet.write(row_idx, 5, record.get('Rows', ''))
        validation_record_sheet.write(row_idx, 6, record.get('Columns', ''))
        validation_record_sheet.write(row_idx, 7, record.get('Original_Confidences', ''))
        validation_record_sheet.write(row_idx, 8, record.get('Updated_Confidences', ''))
        row_idx += 1

    logger.info(f"[VALIDATION_RECORD] Created Validation Record sheet with {row_idx - 1} runs")
    return validation_record_sheet

def create_enhanced_excel_with_validation(excel_data, validation_results, config_data, session_id, skip_history=False, validated_sheet_name=None, qc_results=None, config_s3_key=None, run_key=None):
    """Create 3-sheet Excel file with validation results and optional QC data.

    Args:
        excel_data: Structured table data from shared_table_parser (replaces excel_file_content)
        validation_results: Validation results from the validator
        config_data: Configuration data
        session_id: Session ID for tracking
        skip_history: If True, skip loading existing Details sheet (for fallback mode)
        validated_sheet_name: Name of the sheet that was actually validated (from metadata)
        qc_results: Optional QC results to integrate into the Excel file
        config_s3_key: S3 key for the configuration file (for Validation Record)
        run_key: DynamoDB run key (e.g., #Preview_34t345345346346)
    """
    logger.debug("🔥🔥🔥 EXCEL REPORT QC UNIFIED VERSION IS RUNNING 🔥🔥🔥")

    # DEBUG: Log QC results structure for key mapping investigation
    if qc_results:
        logger.debug(f"[QC_EXCEL_DEBUG] QC results provided with {len(qc_results)} entries")
        logger.debug(f"[QC_EXCEL_DEBUG] QC keys sample: {list(qc_results.keys())[:3]}")
        logger.debug(f"[QC_EXCEL_DEBUG] Validation results keys sample: {list(validation_results.keys())[:3] if isinstance(validation_results, dict) else 'Not a dict'}")

        # CREATE QC-TO-EXCEL POSITIONAL MAPPING
        # Both QC and Excel use hash keys, but they might be different hashes
        # Map by position: first QC result -> first Excel row, second QC -> second Excel row, etc.
        qc_to_excel_mapping = {}
        excel_to_qc_mapping = {}

        if qc_results:
            qc_keys = list(qc_results.keys())
            logger.debug(f"[QC_KEY_MAPPING] QC has {len(qc_keys)} entries")
            logger.debug(f"[QC_KEY_MAPPING] QC keys sample: {qc_keys[:3]}")

            # We'll create the Excel-to-QC mapping when we process rows
            # For now just store the QC keys in order
            qc_keys_ordered = qc_keys
        else:
            qc_keys_ordered = []
    else:
        logger.debug("[QC_EXCEL_DEBUG] No QC results provided to Excel function")
        qc_to_validation_mapping = {}

    # Helper function to find QC data for an Excel row by position
    def get_qc_data_for_row(excel_row_key, row_position):
        """Find QC data for an Excel row using direct hash key lookup."""
        if not qc_results:
            return None

        # Handle None row keys (can happen for partial validation/preview)
        if excel_row_key is None:
            logger.debug(f"[QC_LOOKUP_SKIPPED] Row {row_position} has no row key (partial validation/preview)")
            return None

        # Direct lookup using hash-based keys (should work now with extracted keys)
        if excel_row_key in qc_results:
            logger.debug(f"[QC_LOOKUP_SUCCESS] Direct match for row {row_position}: {excel_row_key[:8]}...")
            return qc_results[excel_row_key]

        # If direct lookup fails, check if this is expected (preview) or unexpected (validation with missing QC)
        # We need validation_results context to make this distinction
        logger.debug(f"[QC_LOOKUP_FAILED] No QC data found for row {row_position}, Excel key: {excel_row_key[:8]}...")
        return None

    if not EXCEL_ENHANCEMENT_AVAILABLE:
        logger.warning("Enhanced Excel not available, skipping Excel creation")
        return None
        
    try:
        # Create Excel buffer
        excel_buffer = io.BytesIO()
        
        # Debug logging for excel_data type and content
        logger.info(f"Excel data type: {type(excel_data)}")
        logger.info(f"Excel data is dict: {isinstance(excel_data, dict)}")
        if isinstance(excel_data, dict):
            logger.info(f"Excel data keys: {list(excel_data.keys())}")
        else:
            logger.info(f"Excel data length: {len(excel_data) if hasattr(excel_data, '__len__') else 'N/A'}")
        
        # Extract structured data (already parsed by shared_table_parser)
        if isinstance(excel_data, dict) and excel_data.get('column_names') and excel_data.get('data'):
            headers = excel_data.get('column_names', [])
            rows_data = excel_data.get('data', [])
            # Get formula data if available (for restoring original formulas)
            formula_data = excel_data.get('formulas', []) or excel_data.get('formula_data', [])
            logger.info(f"Formula data available: {len(formula_data)} rows")
            # Handle sheet name for both CSV and Excel files
            metadata = excel_data.get('metadata', {})
            has_external_links = metadata.get('has_external_links', False)
            file_type = metadata.get('file_type', 'unknown')
            
            if file_type == 'csv':
                actual_sheet_name = 'CSV Data'  # CSV files don't have sheet names
            else:
                actual_sheet_name = validated_sheet_name or metadata.get('sheet_name', 'Unknown')
            
            logger.info(f"Using structured data from validated source: '{actual_sheet_name}' (file_type: {file_type})")
        else:
            # Fallback for backward compatibility (if raw content is still passed)
            logger.warning(f"Excel data is not structured dict, attempting fallback with raw data processing")
            if not isinstance(excel_data, (bytes, io.BytesIO)):
                logger.error(f"Excel data is neither structured dict nor bytes/BytesIO, cannot process. Type: {type(excel_data)}")
                return None
            
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(excel_data) if isinstance(excel_data, bytes) else excel_data)
            except Exception as openpyxl_error:
                logger.error(f"Failed to load Excel data with openpyxl: {openpyxl_error}")
                return None
            
            # Use validated_sheet_name if provided, otherwise fall back to old logic
            if validated_sheet_name and validated_sheet_name in workbook.sheetnames:
                worksheet = workbook[validated_sheet_name]
                logger.info(f"Using validated sheet '{validated_sheet_name}' for enhanced Excel creation")
            elif 'Results' in workbook.sheetnames:
                worksheet = workbook['Results']
                logger.info(f"Using 'Results' sheet as data source for enhanced Excel creation")
            elif len(workbook.sheetnames) > 0:
                worksheet = workbook[workbook.sheetnames[0]]
                logger.info(f"Using first sheet '{worksheet.title}' as data source for enhanced Excel creation")
            else:
                worksheet = workbook.active
                logger.info(f"Using active sheet: {worksheet.title}")
            
            # Get headers and data
            headers = [cell.value for cell in worksheet[1]]
            
            # Convert to list of dictionaries 
            rows_data = []
            for row_idx in range(2, worksheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    if header:
                        cell_value = worksheet.cell(row=row_idx, column=col_idx + 1).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                rows_data.append(row_data)
        
        # Extract ID fields for Details sheet (not used for row key generation anymore)
        id_fields = []
        # Build column importance mapping for formula preservation
        column_importance = {}
        for target in config_data.get('validation_targets', []):
            field_name = target.get('name') or target.get('column')
            if field_name:
                importance = target.get('importance', '').upper()
                column_importance[field_name] = importance
                if importance == 'ID':
                    id_fields.append(field_name)

        def should_preserve_formulas(column_name):
            """Check if formulas should be preserved for this column (IGNORED and ID columns)."""
            importance = column_importance.get(column_name, '').upper()
            return importance in ['IGNORED', 'ID']

        def should_apply_coloring(column_name):
            """Check if confidence-based coloring should be applied (only for validated columns)."""
            importance = column_importance.get(column_name, '').upper()
            return importance not in ['IGNORED', 'ID']

        def get_original_value_with_formulas(row_idx, col_name, default_value):
            """Get the original value, preferring formula over calculated value for IGNORED/ID columns."""
            # If the sheet has external links, never restore formulas, always use the calculated value.
            if has_external_links:
                return default_value

            if should_preserve_formulas(col_name) and formula_data and row_idx < len(formula_data):
                # Check if this row/column has a formula
                row_formulas = formula_data[row_idx]
                if isinstance(row_formulas, dict) and col_name in row_formulas:
                    formula_info = row_formulas[col_name]
                    if isinstance(formula_info, dict) and 'formula' in formula_info:
                        original_formula = formula_info['formula']
                        logger.info(f"Restoring formula for {col_name}: {original_formula}")
                        return original_formula
            return default_value
        
        # CRITICAL: Extract row keys from rows_data if available (pre-computed by interface lambda)
        # Otherwise fall back to extracting from validation_results keys
        row_keys = []

        # First, try to extract _row_key from rows_data (if rows include it)
        if rows_data and isinstance(rows_data[0], dict):
            # DEBUG: Show what keys are in the first row
            first_row_keys = list(rows_data[0].keys())
            logger.debug(f"[ROW_KEY_EXTRACT] First row has keys: {first_row_keys[:10]}")
            logger.debug(f"[ROW_KEY_EXTRACT] Checking for '_row_key': {'_row_key' in rows_data[0]}")

            if '_row_key' in rows_data[0]:
                logger.debug(f"[ROW_KEY_EXTRACT] Extracting pre-computed row keys from rows_data")
                for row_idx, row_data in enumerate(rows_data):
                    row_key = row_data.get('_row_key')
                    if row_key:
                        row_keys.append(row_key)
                    else:
                        logger.error(f"[ROW_KEY_EXTRACT] Row {row_idx} missing _row_key!")
                logger.debug(f"[ROW_KEY_EXTRACT] Extracted {len(row_keys)} pre-computed row keys from rows_data")
                logger.debug(f"[ROW_KEY_EXTRACT] Sample keys: {row_keys[:3]}")
            else:
                logger.warning(f"[ROW_KEY_EXTRACT] No _row_key in rows_data[0], keys present: {first_row_keys[:10]}")
                # Fall through to next section
        else:
            logger.warning(f"[ROW_KEY_EXTRACT] rows_data is empty or first row is not a dict")

        # CRITICAL ERROR: No row keys found
        if not row_keys:
            logger.error(f"[ROW_KEY_EXTRACT] CRITICAL FAILURE: No _row_key found in rows_data!")
            logger.error(f"[ROW_KEY_EXTRACT] This means table_data was not parsed with id_fields or _row_key was stripped")
            logger.error(f"[ROW_KEY_EXTRACT] Excel creation will fail - row keys are REQUIRED for matching validation results")
            logger.error(f"[ROW_KEY_EXTRACT] Validation is parallel/unordered - positional matching is IMPOSSIBLE")

            # Generate None keys to trigger validation failure
            row_keys = [None] * len(rows_data)

            # Log diagnostic info
            if isinstance(validation_results, dict):
                logger.error(f"[ROW_KEY_EXTRACT] Validation results available: {len(validation_results)} keys")
                logger.error(f"[ROW_KEY_EXTRACT] Sample validation keys: {list(validation_results.keys())[:3]}")
            logger.error(f"[ROW_KEY_EXTRACT] Rows data count: {len(rows_data)}")
            logger.error(f"[ROW_KEY_EXTRACT] This Excel will have empty Details sheet")

        # CRITICAL: Verify Excel will show all input rows (including duplicates)
        unique_hashes = set(row_keys)
        duplicate_count = len(row_keys) - len(unique_hashes)
        if duplicate_count > 0:
            logger.debug(f"[ROW_KEY_DUPLICATE] Found {duplicate_count} duplicate rows (same hash)")
            logger.debug(f"[ROW_KEY_DUPLICATE] Excel will show all {len(row_keys)} rows, duplicates will share validation results")
        else:
            logger.debug(f"[ROW_KEY_DUPLICATE] All {len(row_keys)} rows are unique")

        # Verify hash matching with validation results
        if isinstance(validation_results, dict) and validation_results:
            available_validation_keys = set(validation_results.keys())

            # Count how many UNIQUE hashes have validation results
            unique_validated = unique_hashes & available_validation_keys

            # Count how many TOTAL rows will have validation results (including duplicates)
            total_rows_with_validation = sum(1 for key in row_keys if key in available_validation_keys)

            logger.debug(f"[ROW_KEY_MATCH] {len(unique_validated)} unique hashes have validation results")
            logger.debug(f"[ROW_KEY_MATCH] {total_rows_with_validation}/{len(row_keys)} total Excel rows will display validation results")
            logger.debug(f"[ROW_KEY_MATCH] Matching keys sample: {list(unique_validated)[:3]}")

            # Check for partial validation (expected in previews/continuations)
            if len(unique_validated) < len(unique_hashes):
                unvalidated_unique = len(unique_hashes) - len(unique_validated)
                total_unvalidated = len(row_keys) - total_rows_with_validation
                logger.debug(f"[ROW_KEY_MATCH] {unvalidated_unique} unique hashes not validated")
                logger.debug(f"[ROW_KEY_MATCH] {total_unvalidated} total rows without validation (expected for preview/partial)")

            # Detect potential hash mismatch issues
            if len(unique_validated) == 0 and len(available_validation_keys) > 0:
                logger.error(f"[ROW_KEY_MATCH] HASH MISMATCH: 0 matching keys despite having {len(available_validation_keys)} validation results!")
                logger.error(f"[ROW_KEY_MATCH] Excel keys sample: {row_keys[:3]}")
                logger.error(f"[ROW_KEY_MATCH] Validation keys sample: {list(available_validation_keys)[:3]}")
        else:
            logger.warning("[ROW_KEY_MATCH] No validation_results to match against")
        

        # Log QC vs Excel key comparison for debugging
        if qc_results:
            qc_keys_sample = list(qc_results.keys())[:3]
            excel_keys_sample = row_keys[:3]
            logger.debug(f"[KEY_MATCH_DEBUG] QC keys sample: {qc_keys_sample}")
            logger.debug(f"[KEY_MATCH_DEBUG] Excel keys sample: {excel_keys_sample}")

            # Check for exact matches
            matching_keys = set(row_keys) & set(qc_results.keys())
            logger.debug(f"[KEY_MATCH_DEBUG] Found {len(matching_keys)} exact key matches out of {len(row_keys)} Excel rows and {len(qc_results)} QC rows")

            if matching_keys:
                logger.error(f"[KEY_MATCH_SUCCESS] Key matching working correctly!")
                logger.error(f"[KEY_MATCH_SUCCESS] Sample matching keys: {list(matching_keys)[:2]}")
            else:
                logger.error(f"[KEY_MATCH_FAILURE] No matching keys found - QC data will not be applied to Excel")
                logger.error(f"[KEY_MATCH_FAILURE] All QC keys: {list(qc_results.keys())}")
                logger.error(f"[KEY_MATCH_FAILURE] All Excel keys: {row_keys}")
        else:
            logger.debug(f"[KEY_MATCH_DEBUG] No QC results provided to Excel function")
        
        # Load existing Details entries and Validation Record from the original Excel (for history preservation)
        existing_details = []
        existing_validation_record = []
        details_sheet_exists = False
        if not skip_history and not isinstance(excel_data, dict):
            # Only try to load existing details/validation record if we have raw Excel data, not structured data
            try:
                if 'Details' in workbook.sheetnames:
                    details_sheet_exists = True
                    details_worksheet = workbook['Details']

                    # Read existing details headers
                    details_headers = [cell.value for cell in details_worksheet[1]]

                    # Read existing details data
                    for row_idx in range(2, details_worksheet.max_row + 1):
                        detail_row = {}
                        for col_idx, header in enumerate(details_headers):
                            if header:
                                cell_value = details_worksheet.cell(row=row_idx, column=col_idx + 1).value
                                detail_row[header] = cell_value

                        # Mark existing entries as "History"
                        if detail_row:
                            if 'New' in detail_row and detail_row['New'] == 'New':
                                detail_row['New'] = 'History'
                            elif 'New' not in detail_row:
                                detail_row['New'] = 'History'
                            existing_details.append(detail_row)

                    logger.info(f"Loaded {len(existing_details)} existing detail entries from Excel")

                # Load existing Validation Record
                if 'Validation Record' in workbook.sheetnames:
                    validation_record_worksheet = workbook['Validation Record']

                    # Read validation record headers
                    validation_record_headers = [cell.value for cell in validation_record_worksheet[1]]

                    # Read existing validation record data
                    for row_idx in range(2, validation_record_worksheet.max_row + 1):
                        record_row = {}
                        for col_idx, header in enumerate(validation_record_headers):
                            if header:
                                cell_value = validation_record_worksheet.cell(row=row_idx, column=col_idx + 1).value
                                record_row[header] = cell_value

                        if record_row:
                            existing_validation_record.append(record_row)

                    logger.info(f"Loaded {len(existing_validation_record)} existing validation record entries from Excel")
            except Exception as e:
                logger.warning(f"Could not load existing Details/Validation Record sheets: {e}")
        
        # Create Excel with xlsxwriter for advanced formatting
        with xlsxwriter.Workbook(excel_buffer, {'strings_to_urls': False, 'nan_inf_to_errors': True}) as workbook:
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'top',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            
            # Original confidence formats (HIGH/MEDIUM/LOW) - no italics, no bolding (for when values stayed same)
            original_confidence_formats = {
                'HIGH': workbook.add_format({'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }

            # Original confidence formats with BOLD (no italics) - for when values changed
            original_confidence_formats_bold = {
                'HIGH': workbook.add_format({'bold': True, 'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'bold': True, 'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'bold': True, 'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }

            # Updated confidence formats with BOLD only (no italics) - for when values changed
            validation_confidence_formats_bold = {
                'HIGH': workbook.add_format({'bold': True, 'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'bold': True, 'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'bold': True, 'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }

            # Updated confidence formats without bold - for when values stayed the same
            validation_confidence_formats = {
                'HIGH': workbook.add_format({'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }

            # QC applied formats with BOLD only (no italics) - for when QC changed values
            qc_confidence_formats_bold = {
                'HIGH': workbook.add_format({'bold': True, 'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'bold': True, 'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'bold': True, 'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }

            # QC applied formats without bold - for when QC didn't change values
            qc_confidence_formats = {
                'HIGH': workbook.add_format({'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }
            
            # SHEET 1: Updated Values
            updated_sheet = workbook.add_worksheet('Updated Values')
            
            # Headers for updated values sheet (no additional columns)
            updated_headers = headers
            for col_idx, col_name in enumerate(updated_headers):
                updated_sheet.write(0, col_idx, col_name, header_format)
                updated_sheet.set_column(col_idx, col_idx, 20)
            
            # Write all rows to updated sheet (not just rows with changes)
            updated_row_idx = 1
            logger.debug(f"[EXCEL_ROW_DEBUG] Starting to process {len(rows_data)} rows")
            logger.debug(f"[EXCEL_ROW_DEBUG] row_keys sample: {row_keys[:3] if row_keys else 'None'}")
            logger.debug(f"[EXCEL_ROW_DEBUG] validation_results keys sample: {list(validation_results.keys())[:3] if isinstance(validation_results, dict) else 'Not a dict'}")

            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row
                row_validation_data = None
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                # Write all rows (not just those with updates)
                if True:  # Process all rows
                    # Write updated values for this row
                    
                    for col_idx, col_name in enumerate(headers):
                        # Get original value, preferring formula for IGNORED/ID columns
                        base_value = row_data.get(col_name, '')
                        original_value = get_original_value_with_formulas(row_idx, col_name, base_value)
                        updated_value = original_value
                        qc_applied = False

                        # For IGNORED/ID columns, always keep original value (including formulas)
                        # Only process validation/QC for validated columns
                        if should_apply_coloring(col_name):
                            # Check for QC value first (highest priority)
                            row_qc_data = get_qc_data_for_row(row_key, row_idx)
                            if row_qc_data and col_name in row_qc_data:
                                field_qc_data = row_qc_data[col_name]
                                if isinstance(field_qc_data, dict) and field_qc_data.get('qc_applied', False):
                                    updated_value = field_qc_data.get('qc_entry', original_value)
                                    qc_applied = True
                                    logger.debug(f"[QC_APPLIED] Row {row_key}, Col {col_name}: Using QC value '{updated_value}'")
                                else:
                                    logger.error(f"[QC_NOT_APPLIED] Row {row_key}, Col {col_name}: QC available but not applied - qc_applied={field_qc_data.get('qc_applied') if isinstance(field_qc_data, dict) else 'N/A'}")
                            else:
                                # Check if this is validation data (should have QC) vs preview data (expected no QC)
                                row_validation_data = validation_results.get(row_key) if isinstance(validation_results, dict) else None
                                if row_validation_data and col_name in row_validation_data:
                                    # We have validation data but no QC - this is noteworthy
                                    logger.info(f"[QC_MISSING] Row {row_key}, Col {col_name}: Validation data exists but no QC data found")
                                else:
                                    # No validation data either - this is expected for previews
                                    logger.debug(f"[QC_MISSING] Row {row_key}, Col {col_name}: No QC data found (expected for preview)")

                            # If no QC applied, use validation value
                            if not qc_applied and row_validation_data and col_name in row_validation_data:
                                field_data = row_validation_data[col_name]
                                if isinstance(field_data, dict):
                                    original_confidence = field_data.get('original_confidence')
                                    validation_confidence = field_data.get('confidence_level', field_data.get('confidence', ''))

                                    # Only update if validation confidence is higher than original confidence
                                    # Now properly handles case where original has no confidence (blank values)
                                    if should_update_value(original_confidence, validation_confidence):
                                        updated_value = field_data.get('value', original_value)
                        # For IGNORED/ID columns, updated_value remains the same as original_value (including formulas)
                        
                        # Apply appropriate formatting (QC takes priority over validation confidence)
                        # Skip coloring for IGNORED and ID columns
                        if should_apply_coloring(col_name):
                            # Determine if value changed to decide on bold formatting
                            value_changed = (updated_value != original_value)

                            if qc_applied:
                                # Get QC confidence format - use bold if value changed, regular if not
                                row_qc_data = get_qc_data_for_row(row_key, row_idx)
                                if row_qc_data and col_name in row_qc_data:
                                    field_qc_data = row_qc_data[col_name]
                                    format_dict = qc_confidence_formats_bold if value_changed else qc_confidence_formats
                                    cell_format = get_qc_confidence_format(field_qc_data, format_dict)
                                if not cell_format:
                                    # Fallback to generic format if no confidence found
                                    format_dict = qc_confidence_formats_bold if value_changed else qc_confidence_formats
                                    cell_format = format_dict.get('MEDIUM')  # Default QC format
                            else:
                                validation_confidence = None
                                if row_validation_data and col_name in row_validation_data:
                                    field_data = row_validation_data[col_name]
                                    if isinstance(field_data, dict):
                                        validation_confidence = field_data.get('confidence_level', field_data.get('confidence', ''))
                                # Use bold if value changed, regular if not
                                format_dict = validation_confidence_formats_bold if value_changed else validation_confidence_formats
                                cell_format = get_confidence_format(validation_confidence, format_dict)
                        else:
                            cell_format = None  # No coloring for IGNORED and ID columns

                        # Use xlsxwriter's write_formula method for preserved formulas to ensure proper Excel format
                        if should_preserve_formulas(col_name) and str(updated_value).startswith('='):
                            try:
                                updated_sheet.write_formula(updated_row_idx, col_idx, str(updated_value), cell_format)
                            except Exception as e:
                                logger.warning(f"Failed to write formula, falling back to text: {e}")
                                updated_sheet.write(updated_row_idx, col_idx, safe_for_excel(updated_value, should_preserve_formulas(col_name)), cell_format)
                        else:
                            updated_sheet.write(updated_row_idx, col_idx, safe_for_excel(updated_value, should_preserve_formulas(col_name)), cell_format)
                        
                        # Add comment with original value and reasoning (Updated Values sheet format)
                        comment_text = None
                        if row_validation_data and col_name in row_validation_data:
                            field_data = row_validation_data[col_name]
                            if isinstance(field_data, dict):
                                original_value = row_data.get(col_name, '')
                                validated_value = field_data.get('value', '')
                                reasoning = field_data.get('reasoning', '')

                                # Create comment with original value, confidence, and citations
                                comment_parts = []
                                # ALWAYS show what appears in Original Values sheet (not just when different)
                                # Get original confidence (use QC-adjusted if available)
                                original_confidence = field_data.get('original_confidence', '')
                                row_qc_data = get_qc_data_for_row(row_key, row_idx)
                                if row_qc_data and col_name in row_qc_data:
                                    field_qc_data = row_qc_data[col_name]
                                    if isinstance(field_qc_data, dict):
                                        qc_original_confidence = field_qc_data.get('qc_original_confidence')
                                        if qc_original_confidence and str(qc_original_confidence).strip():
                                            original_confidence = qc_original_confidence

                                # Lead with Original Value and confidence
                                if original_confidence:
                                    comment_parts.append(f'Original Value: {original_value} ({original_confidence} Confidence)')
                                else:
                                    comment_parts.append(f'Original Value: {original_value}')

                                # Add Key Citation
                                key_citation = None
                                row_qc_data = get_qc_data_for_row(row_key, row_idx)
                                if row_qc_data and col_name in row_qc_data:
                                    field_qc_data = row_qc_data[col_name]
                                    if isinstance(field_qc_data, dict):
                                        qc_citations = field_qc_data.get('qc_citations', '')
                                        if qc_citations and str(qc_citations).strip():
                                            key_citation = qc_citations

                                # If no QC citation, use first validation citation
                                if not key_citation:
                                    citations = field_data.get('citations', [])
                                    if citations and len(citations) > 0:
                                        first_citation = citations[0]
                                        cite_text = first_citation.get('title', 'Source')
                                        cite_url = first_citation.get('url', '')
                                        cite_snippet = first_citation.get('cited_text', '')
                                        if cite_snippet and len(cite_snippet) > 150:
                                            cite_snippet = cite_snippet[:150] + "..."
                                        key_citation = f"{cite_text}"
                                        if cite_snippet:
                                            key_citation += f" - {cite_snippet}"
                                        if cite_url:
                                            key_citation += f" ({cite_url})"

                                if key_citation:
                                    # Ensure key citation doesn't have problematic newlines
                                    clean_citation = key_citation.replace('\n', ' ').replace('\r', ' ')
                                    comment_parts.append(f'Key Citation: {clean_citation}')

                                # Add Sources section
                                citations = field_data.get('citations', [])
                                if citations:
                                    citation_texts = []
                                    for i, citation in enumerate(citations, 1):
                                        cite_text = f"[{i}] {citation.get('title', 'Untitled')}"
                                        cite_url = citation.get('url', '')
                                        cite_snippet = citation.get('cited_text', '')
                                        # Format: [{#}] {Title}: "{quote}" (URL)
                                        if cite_snippet:
                                            cite_text += f": \"{cite_snippet}\""
                                        if cite_url:
                                            cite_text += f" ({cite_url})"
                                        citation_texts.append(cite_text)

                                    if citation_texts:
                                        comment_parts.append(f"Sources:\n" + "\n".join(citation_texts))

                                # Create comment only if we have meaningful content
                                if comment_parts:
                                    comment_text = '\n\n'.join(comment_parts)
                        
                        # Add comment if needed
                        if comment_text:
                            try:
                                updated_sheet.write_comment(updated_row_idx, col_idx, comment_text,
                                                           {'width': 360, 'height': 582})
                            except Exception as e:
                                logger.warning(f"Could not add comment to Updated Values sheet: {e}")
                    
                    # No additional columns to write (removed Original Value and Supporting Information)
                    
                    updated_row_idx += 1

            # CRITICAL VERIFICATION: Ensure all input rows were written to Excel
            total_input_rows = len(rows_data)
            total_output_rows = updated_row_idx - 1  # Subtract header row
            logger.debug(f"[EXCEL_VERIFICATION] Updated Values sheet: Wrote {total_output_rows} rows (input had {total_input_rows} rows)")
            if total_output_rows == total_input_rows:
                logger.debug(f"[EXCEL_VERIFICATION] ✅ SUCCESS: All {total_input_rows} input rows written to Excel")
            else:
                logger.error(f"[EXCEL_VERIFICATION] ❌ MISMATCH: Input {total_input_rows} rows but wrote {total_output_rows} rows")

            # SHEET 2: Original Values
            original_sheet = workbook.add_worksheet('Original Values')
            
            # Write headers
            for col_idx, col_name in enumerate(headers):
                original_sheet.write(0, col_idx, col_name, header_format)
                original_sheet.set_column(col_idx, col_idx, 20)
            
            # Write original data with confidence-based coloring and comments
            logger.debug(f"[ORIG_SHEET_DEBUG] Starting Original Values sheet write. Total rows: {len(rows_data)}")
            original_rows_processed = 0

            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                try:
                    # Get validation results for this row
                    row_validation_data = None
                    if row_key and row_key in validation_results:
                        row_validation_data = validation_results[row_key]
                        # DEBUG: Log which validation data we got for this row
                        if row_idx < 3:  # Only log first few
                            logger.debug(f"[ORIG_SHEET_DEBUG] Row {row_idx} (Conference: {row_data.get('Conference', 'N/A')}): Using validation data for key {row_key[:8]}...")
                    # REMOVED fallback to positional lookup - causes wrong row matching!
                    # The fallback to str(row_idx) and row_idx was causing random row mismatches
                    # because validation_results dictionary order doesn't match Excel row order

                    for col_idx, col_name in enumerate(headers):
                        if not col_name:
                            continue

                        # Get original value, preferring formula for IGNORED/ID columns
                        base_value = row_data.get(col_name, '')
                        original_value = get_original_value_with_formulas(row_idx, col_name, base_value)
                        cell_format = None
                        comment_text = None
                        validated_value = None  # Reset for each column to prevent carryover from previous column

                        # Check if this column has validation results
                        field_data = None  # Initialize field_data to None
                        if row_validation_data and isinstance(row_validation_data, dict):
                            if col_name in row_validation_data and isinstance(row_validation_data[col_name], dict):
                                field_data = row_validation_data[col_name]
                                original_confidence = field_data.get('original_confidence')

                            # Determine the actual updated value using the SAME logic as Updated Values sheet
                            # This needs to match lines 746-777 exactly
                            actual_updated_value = original_value  # Start with original

                            if should_apply_coloring(col_name):
                                # Check for QC value first (highest priority)
                                row_qc_data_for_value = get_qc_data_for_row(row_key, row_idx)
                                if row_qc_data_for_value and col_name in row_qc_data_for_value:
                                    field_qc_data_for_value = row_qc_data_for_value[col_name]
                                    if isinstance(field_qc_data_for_value, dict) and field_qc_data_for_value.get('qc_applied', False):
                                        actual_updated_value = field_qc_data_for_value.get('qc_entry', original_value)

                                # If no QC applied, use validation value with confidence check
                                elif field_data:
                                    validation_original_confidence = field_data.get('original_confidence')
                                    validation_confidence = field_data.get('confidence_level', field_data.get('confidence', ''))

                                    # Only update if validation confidence is higher than original confidence
                                    if should_update_value(validation_original_confidence, validation_confidence):
                                        actual_updated_value = field_data.get('value', original_value)

                            # Keep validated_value for legacy code that might use it
                            validated_value = actual_updated_value

                            reasoning = field_data.get('reasoning', '')

                            # DEBUG: Log the validated_value we just extracted
                            if col_name == 'Start Date':
                                logger.debug(f"[ORIG_SHEET_DEBUG] Row {row_idx} {col_name}: validated_value='{validated_value}', original_value='{original_value}', qc_applied={field_data.get('qc_applied', False)}")

                            # Check for QC original confidence override
                            row_qc_data = get_qc_data_for_row(row_key, row_idx)
                            if row_qc_data and isinstance(row_qc_data, dict) and col_name in row_qc_data:
                                field_qc_data = row_qc_data[col_name]
                                if isinstance(field_qc_data, dict):
                                    qc_original_confidence = field_qc_data.get('qc_original_confidence')
                                    if qc_original_confidence and str(qc_original_confidence).strip():
                                        original_confidence = qc_original_confidence

                            # Apply original confidence color (only if confidence should be colored)
                            # Skip coloring for IGNORED and ID columns
                            if should_apply_coloring(col_name):
                                # Determine if value changed to decide on bold formatting
                                # Compare original value with the actual updated value
                                try:
                                    value_changed = (validated_value != original_value) if validated_value is not None else False
                                except Exception as e:
                                    # Handle type comparison issues gracefully
                                    logger.warning(f"Could not compare values for {col_name}: {e}")
                                    value_changed = False

                                # Use bold format if value changed, regular format if not
                                format_dict = original_confidence_formats_bold if value_changed else original_confidence_formats
                                cell_format = get_confidence_format(original_confidence, format_dict)
                            else:
                                cell_format = None  # No coloring for IGNORED and ID columns
                            
                            # Create comment with updated value and confidence (Original Values sheet format)
                            # DEBUG: Log validated_value right before building comment
                            if col_name == 'Start Date':
                                logger.info(f"[COMMENT_BUILD_DEBUG] Row {row_idx} {col_name}: About to build comment with validated_value='{validated_value}'")

                            comment_parts = []
                            # ALWAYS show what appears in Updated Values sheet (not just when different)
                            # Get updated confidence (use QC-adjusted if available)
                            updated_confidence = field_data.get('confidence_level', field_data.get('confidence', ''))
                            row_qc_data = get_qc_data_for_row(row_key, row_idx)
                            if row_qc_data and col_name in row_qc_data:
                                field_qc_data = row_qc_data.get(col_name)
                                if isinstance(field_qc_data, dict):
                                    qc_confidence = field_qc_data.get('qc_confidence')
                                    if qc_confidence and str(qc_confidence).strip():
                                        updated_confidence = qc_confidence

                            # Lead with Updated Value and confidence
                            if updated_confidence:
                                comment_parts.append(f'Updated Value: {validated_value} ({updated_confidence} Confidence)')
                            else:
                                comment_parts.append(f'Updated Value: {validated_value}')

                            # Add Key Citation
                            key_citation = None
                            row_qc_data = get_qc_data_for_row(row_key, row_idx)
                            if row_qc_data and col_name in row_qc_data:
                                field_qc_data = row_qc_data.get(col_name)
                                if isinstance(field_qc_data, dict):
                                    qc_citations = field_qc_data.get('qc_citations', '')
                                    if qc_citations and str(qc_citations).strip():
                                        key_citation = qc_citations

                            # If no QC citation, use first validation citation
                            if not key_citation:
                                citations = field_data.get('citations', [])
                                if citations and len(citations) > 0:
                                    first_citation = citations[0]
                                    cite_text = first_citation.get('title', 'Source')
                                    cite_url = first_citation.get('url', '')
                                    cite_snippet = first_citation.get('cited_text', '')
                                    if cite_snippet and len(cite_snippet) > 150:
                                        cite_snippet = cite_snippet[:150] + "..."
                                    key_citation = f"{cite_text}"
                                    if cite_snippet:
                                        key_citation += f" - {cite_snippet}"
                                    if cite_url:
                                        key_citation += f" ({cite_url})"

                            if key_citation:
                                # Ensure key citation doesn't have problematic newlines
                                clean_citation = key_citation.replace('\n', ' ').replace('\r', ' ')
                                comment_parts.append(f'Key Citation: {clean_citation}')

                            # Add Sources section
                            citations = field_data.get('citations', [])
                            if citations:
                                citation_texts = []
                                for i, citation in enumerate(citations, 1):
                                    cite_text = f"[{i}] {citation.get('title', 'Untitled')}"
                                    cite_url = citation.get('url', '')
                                    cite_snippet = citation.get('cited_text', '')
                                    # Format: [{#}] {Title}: "{quote}" (URL)
                                    if cite_snippet:
                                        cite_text += f": \"{cite_snippet}\""
                                    if cite_url:
                                        cite_text += f" ({cite_url})"
                                    citation_texts.append(cite_text)

                                if citation_texts:
                                    comment_parts.append(f"Sources:\n" + "\n".join(citation_texts))

                            # Create comment only if we have meaningful content
                            if comment_parts:
                                comment_text = '\n\n'.join(comment_parts)

                        # Write original value - use xlsxwriter's write_formula method for preserved formulas
                        if should_preserve_formulas(col_name) and str(original_value).startswith('='):
                            try:
                                original_sheet.write_formula(row_idx + 1, col_idx, str(original_value), cell_format)
                            except Exception as e:
                                logger.warning(f"Failed to write formula, falling back to text: {e}")
                                original_sheet.write(row_idx + 1, col_idx, safe_for_excel(original_value, should_preserve_formulas(col_name)), cell_format)
                        else:
                            original_sheet.write(row_idx + 1, col_idx, safe_for_excel(original_value, should_preserve_formulas(col_name)), cell_format)

                        # Add comment if needed
                        if comment_text:
                            try:
                                # DEBUG: Log the comment text being written
                                if col_name == 'Start Date':
                                    logger.info(f"[COMMENT_WRITE_DEBUG] Row {row_idx} {col_name}: Writing comment at Excel row {row_idx + 1}: {comment_text[:200]}...")

                                original_sheet.write_comment(row_idx + 1, col_idx, comment_text,
                                                           {'width': 360, 'height': 582})
                            except Exception as e:
                                logger.warning(f"Could not add comment: {e}")

                    original_rows_processed += 1
                except Exception as row_error:
                    logger.error(f"[ORIG_SHEET_ERROR] Failed to write Original Values row {row_idx} (key: {str(row_key)[:20] if row_key else 'None'}): {row_error}")
                    import traceback
                    logger.error(f"[ORIG_SHEET_ERROR] Traceback: {traceback.format_exc()}")
                    # Continue with next row instead of breaking entire Excel creation
                    continue

            logger.debug(f"[ORIG_SHEET_DEBUG] Original Values sheet write complete. Processed {original_rows_processed} rows")

            # SHEET 3: Validation Record (run-level metadata)
            # Use config_s3_key parameter if provided, otherwise try to extract from config_data
            if not config_s3_key:
                config_s3_key = config_data.get('config_s3_key', '') if isinstance(config_data, dict) else ''

            # Determine if this is a preview run (heuristic: check if we have partial validation)
            is_preview = False
            if isinstance(validation_results, dict) and validation_results:
                validated_row_count = len([k for k in row_keys if k in validation_results])
                total_row_count = len(row_keys)
                # If less than 50% of rows validated, consider it a preview
                if total_row_count > 0 and validated_row_count < (total_row_count * 0.5):
                    is_preview = True
                    logger.info(f"[VALIDATION_RECORD] Detected preview run ({validated_row_count}/{total_row_count} rows validated)")

            validation_record_sheet = create_validation_record_sheet(
                workbook=workbook,
                header_format=header_format,
                validation_results=validation_results,
                qc_results=qc_results,
                session_id=session_id,
                config_s3_key=config_s3_key,
                rows_data=rows_data,
                headers=headers,
                existing_validation_record=existing_validation_record,
                is_preview=is_preview,
                run_key=run_key
            )

            # SHEET 4: Details (comprehensive view)
            details_sheet = workbook.add_worksheet('Details')
            
            # Build detail headers dynamically to include ID fields
            detail_headers = ["Row Key", "Identifier"]
            
            # Add ID field columns
            for id_field in id_fields:
                detail_headers.append(id_field)
            
            # Add the rest of the standard columns with QC integration
            # Debug flag to isolate QC column issues
            INCLUDE_QC_COLUMNS = True  # Set to False to test if QC columns cause corruption

            if INCLUDE_QC_COLUMNS:
                detail_headers.extend(["Column", "Original Value", "Updated Value", "QC Value",
                                "QC Original Confidence", "QC Updated Confidence", "QC Confidence",
                                "QC Applied", "QC Reasoning", "Update Importance", "QC Sources", "QC Citations",
                                "Final Value", "Reasoning", "Sources", "Citations",
                                "Explanation", "Consistent with Model", "Model", "Timestamp", "New"])
            else:
                detail_headers.extend(["Column", "Original Value", "Original Confidence", "Validated Value",
                                "Validation Confidence", "Final Value", "Reasoning", "Sources", "Citations",
                                "Explanation", "Consistent with Model", "Model", "Timestamp", "New"])
            
            for col_idx, header in enumerate(detail_headers):
                details_sheet.write(0, col_idx, header, header_format)
                details_sheet.set_column(col_idx, col_idx, 20)
            
            detail_row = 1
            current_timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            
            # Track which row keys have been processed to avoid duplicates
            processed_row_keys = set()
            
            # Write NEW validation results to details sheet
            logger.debug(f"[DETAILS_DEBUG] Starting Details sheet data write. Total rows: {len(rows_data)}, Validation results keys: {len(validation_results) if isinstance(validation_results, dict) else 'N/A'}")

            # DEBUG: Show sample keys from both sides to diagnose mismatch
            if isinstance(validation_results, dict):
                validation_keys_sample = list(validation_results.keys())[:3]
                row_keys_sample = [str(k)[:20] for k in row_keys[:3]]
                logger.debug(f"[KEY_MISMATCH_DEBUG] Validation results keys (first 3): {validation_keys_sample}")
                logger.debug(f"[KEY_MISMATCH_DEBUG] Row keys from excel (first 3): {row_keys_sample}")

            details_rows_written = 0

            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                try:
                    # Get validation results for this row
                    row_validation_data = None
                    if row_key in validation_results:
                        row_validation_data = validation_results[row_key]
                    elif str(row_idx) in validation_results:
                        row_validation_data = validation_results[str(row_idx)]
                    elif row_idx in validation_results:
                        row_validation_data = validation_results[row_idx]

                    if not row_validation_data:
                        if row_idx < 3:  # Log first few misses
                            logger.warning(f"[DETAILS_DEBUG] Row {row_idx} (key: {str(row_key)[:20] if row_key else 'None'}): No validation data found")
                        continue

                    if not isinstance(row_validation_data, dict):
                        logger.warning(f"[DETAILS_DEBUG] Row {row_idx}: validation data is not a dict: {type(row_validation_data)}")
                        continue

                    # Create identifier from ID fields
                    identifier_parts = []
                    for id_field in id_fields:
                        if id_field in row_data:
                            identifier_parts.append(f"{id_field}: {row_data[id_field]}")
                    identifier = ", ".join(identifier_parts) if identifier_parts else f"Row {row_idx + 1}"

                    for field_name, field_data in row_validation_data.items():
                        if isinstance(field_data, dict) and 'confidence' in field_data:
                            details_rows_written += 1
                            col_idx = 0
                            details_sheet.write(detail_row, col_idx, safe_for_excel(row_key))  # Row Key
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(identifier))  # Identifier
                            col_idx += 1
                            
                            # Write ID field values
                            for id_field in id_fields:
                                value = row_data.get(id_field, '')
                                details_sheet.write(detail_row, col_idx, safe_for_excel(str(value)))
                                col_idx += 1
                            
                            # Write standard columns in new order: Column, Original Value, Updated Value, QC Value
                            details_sheet.write(detail_row, col_idx, safe_for_excel(field_name))  # Column
                            col_idx += 1
                            # Get original value with formulas for IGNORED/ID columns
                            base_original_value = str(row_data.get(field_name, ''))
                            original_value_with_formula = get_original_value_with_formulas(row_idx, field_name, base_original_value)
                            # Write original value with formula preservation
                            if should_preserve_formulas(field_name) and str(original_value_with_formula).startswith('='):
                                try:
                                    details_sheet.write_formula(detail_row, col_idx, str(original_value_with_formula))
                                except Exception as e:
                                    logger.warning(f"Failed to write formula, falling back to text: {e}")
                                    details_sheet.write(detail_row, col_idx, safe_for_excel(original_value_with_formula, should_preserve_formulas(field_name)))
                            else:
                                details_sheet.write(detail_row, col_idx, safe_for_excel(original_value_with_formula, should_preserve_formulas(field_name)))
                            col_idx += 1

                            # Updated Value (pre-QC result) - use preserved pre-QC value if available
                            if field_data.get('qc_applied') and 'pre_qc_value' in field_data:
                                updated_value = str(field_data.get('pre_qc_value', ''))
                                updated_confidence = field_data.get('pre_qc_confidence', '')
                            else:
                                updated_value = str(field_data.get('value', ''))
                                updated_confidence = field_data.get('confidence_level', field_data.get('confidence', ''))

                            # Write updated value with formula preservation
                            if should_preserve_formulas(field_name) and str(updated_value).startswith('='):
                                try:
                                    details_sheet.write_formula(detail_row, col_idx, str(updated_value))
                                except Exception as e:
                                    logger.warning(f"Failed to write formula, falling back to text: {e}")
                                    details_sheet.write(detail_row, col_idx, safe_for_excel(updated_value, should_preserve_formulas(field_name)))
                            else:
                                details_sheet.write(detail_row, col_idx, safe_for_excel(updated_value, should_preserve_formulas(field_name)))
                            col_idx += 1

                            # Extract QC data properly
                            qc_applied = False
                            qc_value = ''
                            qc_confidence = ''
                            qc_reasoning = ''
                            qc_sources = []
                            qc_citations = ''
                            qc_original_confidence = ''
                            qc_updated_confidence = ''

                            # Final value starts as updated value
                            final_value = str(field_data.get('value', ''))

                            row_qc_data = get_qc_data_for_row(row_key, row_idx)
                            if row_qc_data and field_name in row_qc_data:
                                field_qc_data = row_qc_data[field_name]
                                if isinstance(field_qc_data, dict):
                                        qc_applied = field_qc_data.get('qc_applied', False)
                                        logger.debug(f"[QC_EXCEL_EXTRACT_DEBUG] {field_name}: qc_applied={qc_applied}, available_keys={list(field_qc_data.keys())}")
                                        if qc_applied:
                                            # QC was applied - extract QC's proposed values with proper null handling and Excel safety
                                            raw_qc_value = field_qc_data.get('qc_entry') or ''
                                            qc_value = safe_for_excel(str(raw_qc_value))
                                            qc_confidence = str(field_qc_data.get('qc_confidence') or '')
                                            raw_qc_reasoning = field_qc_data.get('qc_reasoning') or ''
                                            qc_reasoning = safe_for_excel(str(raw_qc_reasoning))

                                            # Handle QC sources with null safety
                                            raw_qc_sources = field_qc_data.get('qc_sources', [])
                                            if isinstance(raw_qc_sources, list):
                                                qc_sources = [safe_for_excel(str(s)) for s in raw_qc_sources if s is not None and str(s).strip()]
                                            else:
                                                qc_sources = []

                                            # Safely extract QC citations with enhanced cleaning
                                            raw_qc_citations = field_qc_data.get('qc_citations') or ''
                                            qc_citations = safe_for_excel(raw_qc_citations)

                                            qc_original_confidence = str(field_qc_data.get('qc_original_confidence') or '')
                                            qc_updated_confidence = str(field_qc_data.get('qc_updated_confidence') or '')

                                            # Final value logic: always use QC value when available (since QC is now comprehensive)
                                            if qc_value and str(qc_value).strip():
                                                # QC provided a replacement value
                                                final_value = qc_value
                                            else:
                                                # QC only changed confidence, keep the updated value
                                                final_value = str(field_data.get('value', ''))
                                            logger.debug(f"[QC_EXCEL_EXTRACT_DEBUG] {field_name}: QC extracted - value='{qc_value}', confidence='{qc_confidence}'")

                            # Use QC confidence format for QC Value when QC applied
                            # Skip coloring for IGNORED and ID columns
                            if should_apply_coloring(field_name):
                                qc_format = None
                                if qc_applied and qc_confidence:
                                    qc_format = get_confidence_format(qc_confidence, qc_confidence_formats)
                                if not qc_format and qc_applied:
                                    qc_format = qc_confidence_formats.get('MEDIUM')  # Default QC format
                            else:
                                qc_format = None  # No coloring for IGNORED and ID columns

                            try:
                                # Write QC value with formula preservation
                                if should_preserve_formulas(field_name) and str(qc_value).startswith('='):
                                    try:
                                        details_sheet.write_formula(detail_row, col_idx, str(qc_value), qc_format)
                                    except Exception as e:
                                        logger.warning(f"Failed to write QC formula, falling back to text: {e}")
                                        details_sheet.write(detail_row, col_idx, safe_for_excel(qc_value, should_preserve_formulas(field_name)), qc_format)
                                else:
                                    details_sheet.write(detail_row, col_idx, safe_for_excel(qc_value, should_preserve_formulas(field_name)), qc_format)  # QC Value
                                logger.debug(f"[EXCEL_WRITE_DEBUG] QC Value written successfully for {field_name}")
                            except Exception as e:
                                logger.error(f"[EXCEL_WRITE_ERROR] QC Value write failed for {field_name}: {e}")
                                details_sheet.write(detail_row, col_idx, '')  # Write empty on error
                            col_idx += 1

                            # QC Original Confidence column (only populated when QC changes original confidence)
                            # Skip coloring for IGNORED and ID columns
                            if should_apply_coloring(field_name):
                                qc_original_confidence_format = get_confidence_format(qc_original_confidence, qc_confidence_formats) if qc_original_confidence else None
                            else:
                                qc_original_confidence_format = None
                            details_sheet.write(detail_row, col_idx, safe_for_excel(qc_original_confidence), qc_original_confidence_format)  # QC Original Confidence
                            col_idx += 1

                            # QC Updated Confidence column (only populated when QC changes updated confidence)
                            # Skip coloring for IGNORED and ID columns
                            if should_apply_coloring(field_name):
                                qc_updated_confidence_format = get_confidence_format(qc_updated_confidence, qc_confidence_formats) if qc_updated_confidence else None
                            else:
                                qc_updated_confidence_format = None
                            details_sheet.write(detail_row, col_idx, safe_for_excel(qc_updated_confidence), qc_updated_confidence_format)  # QC Updated Confidence
                            col_idx += 1

                            # QC Confidence column
                            try:
                                # Skip coloring for IGNORED and ID columns
                                if should_apply_coloring(field_name):
                                    qc_confidence_format = get_confidence_format(qc_confidence, qc_confidence_formats) if qc_confidence else None
                                else:
                                    qc_confidence_format = None
                                details_sheet.write(detail_row, col_idx, safe_for_excel(qc_confidence), qc_confidence_format)  # QC Confidence
                                logger.debug(f"[EXCEL_WRITE_DEBUG] QC Confidence written successfully for {field_name}")
                            except Exception as e:
                                logger.error(f"[EXCEL_WRITE_ERROR] QC Confidence write failed for {field_name}: {e}")
                                details_sheet.write(detail_row, col_idx, '')  # Write empty on error
                            col_idx += 1

                            details_sheet.write(detail_row, col_idx, 'Yes' if qc_applied else 'No')  # QC Applied
                            col_idx += 1

                            try:
                                details_sheet.write(detail_row, col_idx, safe_for_excel(qc_reasoning))  # QC Reasoning
                                logger.debug(f"[EXCEL_WRITE_DEBUG] QC Reasoning written successfully for {field_name}")
                            except Exception as e:
                                logger.error(f"[EXCEL_WRITE_ERROR] QC Reasoning write failed for {field_name}: {e}")
                                details_sheet.write(detail_row, col_idx, '')  # Write empty on error
                            col_idx += 1

                            # Update Importance column
                            update_importance = ''
                            if row_qc_data and field_name in row_qc_data:
                                field_qc_data = row_qc_data[field_name]
                                if isinstance(field_qc_data, dict):
                                    update_importance = str(field_qc_data.get('update_importance', '0'))
                            details_sheet.write(detail_row, col_idx, safe_for_excel(update_importance))  # Update Importance
                            col_idx += 1

                            # QC Sources column with safe joining
                            try:
                                # Filter out empty sources and ensure all are strings
                                safe_qc_sources = [str(s).strip() for s in qc_sources if s and str(s).strip()]
                                qc_sources_str = '; '.join(safe_qc_sources) if safe_qc_sources else ''
                            except Exception as e:
                                logger.warning(f"QC sources processing failed: {e}")
                                qc_sources_str = ''
                            details_sheet.write(detail_row, col_idx, safe_for_excel(qc_sources_str))  # QC Sources
                            col_idx += 1

                            # QC Citations column
                            details_sheet.write(detail_row, col_idx, safe_for_excel(qc_citations))  # QC Citations
                            col_idx += 1

                            # Final Value should use QC confidence format if QC applied, otherwise updated confidence format
                            # Skip coloring for IGNORED and ID columns
                            if should_apply_coloring(field_name):
                                updated_confidence_format = get_confidence_format(updated_confidence, validation_confidence_formats)
                                final_format = qc_format if qc_applied else updated_confidence_format
                            else:
                                final_format = None  # No coloring for IGNORED and ID columns
                            # Write final value with formula preservation
                            if should_preserve_formulas(field_name) and str(final_value).startswith('='):
                                try:
                                    details_sheet.write_formula(detail_row, col_idx, str(final_value), final_format)
                                except Exception as e:
                                    logger.warning(f"Failed to write final formula, falling back to text: {e}")
                                    details_sheet.write(detail_row, col_idx, safe_for_excel(final_value, should_preserve_formulas(field_name)), final_format)
                            else:
                                details_sheet.write(detail_row, col_idx, safe_for_excel(final_value, should_preserve_formulas(field_name)), final_format)  # Final Value
                            col_idx += 1

                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('reasoning', ''))))  # Reasoning
                            col_idx += 1

                            # Sources column - blend validation and QC sources when QC is applied
                            try:
                                validation_sources = field_data.get('sources', []) or []
                                if qc_applied and qc_sources:
                                    # Combine validation sources with QC sources (avoid duplicates)
                                    all_sources = [str(s) for s in validation_sources if s]  # Clean validation sources
                                    for qc_source in qc_sources:
                                        if qc_source and str(qc_source) not in all_sources:
                                            all_sources.append(str(qc_source))
                                    sources_text = ', '.join(all_sources)
                                else:
                                    clean_validation_sources = [str(s) for s in validation_sources if s]
                                    sources_text = ', '.join(clean_validation_sources)
                            except Exception as e:
                                logger.warning(f"Sources blending failed: {e}")
                                sources_text = ''
                            details_sheet.write(detail_row, col_idx, safe_for_excel(sources_text))  # Sources
                            col_idx += 1
                            
                            # Citations column - blend validation and QC citations when QC is applied
                            try:
                                citations = field_data.get('citations', []) or []
                                citation_texts = []

                                # Add validation citations first
                                if citations:
                                    for i, citation in enumerate(citations, 1):
                                        if isinstance(citation, dict):
                                            cite_text = f"[{i}] {citation.get('title', 'Untitled')}"
                                            cite_url = citation.get('url', '')
                                            cite_snippet = citation.get('cited_text', '')
                                            # Format: [{#}] {Title}: "{quote}" (URL)
                                            if cite_snippet:
                                                # Show full citation content - no truncation
                                                cite_text += f": \"{cite_snippet}\""
                                            if cite_url:
                                                cite_text += f" ({cite_url})"
                                            citation_texts.append(cite_text)

                                # Add QC citations if QC is applied and has citations
                                if qc_applied and qc_citations and str(qc_citations).strip():
                                    if citation_texts:
                                        citation_texts.append("")  # Add blank line separator
                                    citation_texts.append("--- QC CITATIONS ---")
                                    citation_texts.append(str(qc_citations).strip())

                                citations_text = '\n'.join(citation_texts) if citation_texts else ''
                            except Exception as e:
                                logger.warning(f"Citations blending failed: {e}")
                                citations_text = ''
                            details_sheet.write(detail_row, col_idx, safe_for_excel(citations_text))  # Citations
                            col_idx += 1
                            
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('explanation', ''))))  # Explanation
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('consistent_with_model_knowledge', ''))))  # Consistent with Model
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('model', ''))))  # Model
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(current_timestamp))  # Timestamp
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, 'New')  # Mark as new

                            # Track this combination as processed
                            processed_row_keys.add((row_key, field_name))
                            detail_row += 1
                except Exception as row_error:
                    logger.error(f"[DETAILS_ERROR] Failed to write Details row {row_idx} (key: {str(row_key)[:20] if row_key else 'None'}): {row_error}")
                    import traceback
                    logger.error(f"[DETAILS_ERROR] Traceback: {traceback.format_exc()}")
                    # Continue with next row instead of breaking entire Excel creation
                    continue

            logger.debug(f"[DETAILS_DEBUG] Details sheet write complete. Wrote {details_rows_written} field entries across {detail_row - 1} detail rows")
            
            # Then, append existing HISTORICAL details
            for existing_detail in existing_details:
                # Check if this is a duplicate of a new entry
                existing_row_key = existing_detail.get('Row Key', '')
                existing_column = existing_detail.get('Column', '')
                
                if (existing_row_key, existing_column) not in processed_row_keys:
                    # Write the historical entry
                    col_idx = 0
                    
                    # Row Key
                    details_sheet.write(detail_row, col_idx, safe_for_excel(str(existing_detail.get('Row Key', ''))))
                    col_idx += 1
                    
                    # Identifier
                    details_sheet.write(detail_row, col_idx, safe_for_excel(str(existing_detail.get('Identifier', ''))))
                    col_idx += 1
                    
                    # ID field values - extract from existing detail or try to reconstruct
                    for id_field in id_fields:
                        value = existing_detail.get(id_field, '')
                        details_sheet.write(detail_row, col_idx, safe_for_excel(str(value)))
                        col_idx += 1
                    
                    # Standard columns (map old field names to new ones)
                    standard_columns = ['Column', 'Original Value', 'Updated Value', 'QC Value',
                                      'QC Original Confidence', 'QC Updated Confidence', 'QC Confidence',
                                      'QC Applied', 'QC Reasoning', 'Update Importance', 'QC Sources', 'QC Citations',
                                      'Final Value', 'Reasoning', 'Sources', 'Citations', 'Explanation',
                                      'Consistent with Model', 'Model', 'Timestamp', 'New']

                    # Handle field name mapping for backward compatibility
                    field_mapping = {
                        'Validated Value': 'Updated Value',
                        'Validation Confidence': 'QC Updated Confidence',
                        'Confidence': 'QC Updated Confidence',
                        'Quote': 'Reasoning'
                    }
                    
                    for col_name in standard_columns:
                        # Try the new column name first, then check mappings
                        value = existing_detail.get(col_name, '')
                        if not value and col_name in field_mapping.values():
                            # Try the old field name
                            old_name = [k for k, v in field_mapping.items() if v == col_name]
                            if old_name:
                                value = existing_detail.get(old_name[0], '')
                        
                        # Apply confidence formatting if applicable
                        # Skip coloring for IGNORED and ID columns (use existing_column as the field name)
                        if col_name in ['QC Original Confidence', 'QC Updated Confidence', 'QC Confidence'] and should_apply_coloring(existing_column):
                            # QC confidence formatting (only for validated columns)
                            confidence_format = get_confidence_format(value, qc_confidence_formats)
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(value) if value is not None else ''), confidence_format)
                        else:
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(value) if value is not None else ''))
                        
                        col_idx += 1
                    
                    detail_row += 1
            
            logger.info(f"Created 3-sheet Excel with {detail_row - 1} total detail entries (new + historical)")
        
        excel_buffer.seek(0)
        return excel_buffer
        
    except Exception as e:
        logger.error(f"Error creating enhanced Excel: {str(e)}")
        import traceback
        logger.error(f"Enhanced Excel creation traceback: {traceback.format_exc()}")
        return None


def create_qc_enhanced_excel_for_interface(
    table_data: any,
    validation_results: dict,
    config_data: dict,
    session_id: str,
    validated_sheet_name: str = None,
    config_s3_key: str = None,
    run_key: str = None
):
    """
    QC-enhanced Excel creation that integrates with interface lambda.

    This function extracts QC results from validation_results and calls
    the unified Excel creation function with QC support.

    Args:
        table_data: Structured table data from shared_table_parser
        validation_results: Validation results (may contain qc_results)
        config_data: Configuration data
        session_id: Session ID
        validated_sheet_name: Name of validated sheet
        config_s3_key: S3 key for the configuration file (for Validation Record)
        run_key: DynamoDB run key (e.g., #Preview_34t345345346346)

    Returns:
        BytesIO buffer with Excel content, or None if creation failed
    """
    logger.error("🚀🚀🚀 QC ENHANCED INTERFACE FUNCTION IS RUNNING 🚀🚀🚀")
    try:
        # Extract QC results and actual validation results from the full response
        qc_results = None
        actual_validation_results = validation_results

        if isinstance(validation_results, dict):
            # QC results are at the top level
            qc_results = validation_results.get('qc_results')

            # Actual validation results might be nested under 'validation_results'
            if 'validation_results' in validation_results:
                actual_validation_results = validation_results['validation_results']
                logger.debug(f"[QC_EXCEL_DEBUG] Using nested validation_results structure")

            logger.debug(f"[QC_EXTRACT_DEBUG] Raw qc_results from validation_results.get('qc_results'): {type(qc_results)} = {qc_results}")
            logger.debug(f"[QC_EXCEL_DEBUG] Extracted QC results: {qc_results is not None}")
            if qc_results:
                logger.debug(f"[QC_EXCEL_DEBUG] QC results keys: {list(qc_results.keys())}")
                logger.debug(f"[QC_EXCEL_DEBUG] QC results sample: {list(qc_results.values())[:1] if qc_results else 'None'}")
            else:
                logger.debug(f"[QC_EXCEL_DEBUG] No QC results found in validation_results")
                logger.debug(f"[QC_EXCEL_DEBUG] validation_results keys: {list(validation_results.keys()) if isinstance(validation_results, dict) else 'Not a dict'}")
                # Let's see what's actually in the qc_results key
                raw_qc = validation_results.get('qc_results', 'KEY_NOT_FOUND')
                logger.debug(f"[QC_EXTRACT_DEBUG] Raw content of 'qc_results' key: {raw_qc}")
                logger.debug(f"[QC_EXTRACT_DEBUG] Type of raw qc_results: {type(raw_qc)}")
                if isinstance(raw_qc, dict):
                    logger.debug(f"[QC_EXTRACT_DEBUG] Raw qc_results has {len(raw_qc)} items")
                    if raw_qc:
                        logger.debug(f"[QC_EXTRACT_DEBUG] Sample raw qc key: {list(raw_qc.keys())[0] if raw_qc else 'No keys'}")

        # Call the unified Excel creation function with the actual validation results
        excel_buffer = create_enhanced_excel_with_validation(
            excel_data=table_data,
            validation_results=actual_validation_results,
            config_data=config_data,
            session_id=session_id,
            skip_history=False,
            validated_sheet_name=validated_sheet_name,
            qc_results=qc_results,
            config_s3_key=config_s3_key,
            run_key=run_key
        )

        if excel_buffer:
            logger.info(f"Created QC-enhanced Excel for session {session_id}")

        return excel_buffer

    except Exception as e:
        logger.error(f"Error creating QC-enhanced Excel for interface: {str(e)}")

        # Fallback to standard Excel creation without QC
        try:
            return create_enhanced_excel_with_validation(
                excel_data=table_data,
                validation_results=validation_results,
                config_data=config_data,
                session_id=session_id,
                skip_history=False,
                validated_sheet_name=validated_sheet_name,
                qc_results=None,
                config_s3_key=config_s3_key,
                run_key=run_key
            )
        except Exception as fallback_error:
            logger.error(f"Fallback Excel creation also failed: {str(fallback_error)}")
            return None