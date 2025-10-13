"""
QC-Enhanced Excel Report Generator

Extends the existing Excel reporting functionality to handle QC results.
Creates enhanced Excel files with QC data integrated into Results, Details, and Reasons sheets.
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

import openpyxl

# xlsxwriter is an optional dependency, so we handle its import gracefully.
try:
    import xlsxwriter
    EXCEL_ENHANCEMENT_AVAILABLE = True
except ImportError:
    EXCEL_ENHANCEMENT_AVAILABLE = False

from qc_excel_formatter import (
    get_qc_enhanced_detail_headers,
    format_qc_result_for_excel,
    get_qc_enhanced_column_widths,
    write_qc_enhanced_detail_row,
    create_qc_formats
)

logger = logging.getLogger()

def safe_for_excel(value):
    """Convert value to Excel-safe format, handling control characters but NOT XML escaping."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Check for NaN without pandas
        if value != value:  # NaN check
            return ""
        # Check for infinity
        if isinstance(value, float) and (value == float('inf') or value == float('-inf')):
            return ""

    # Convert to string for processing
    value_str = str(value)

    # First, handle control characters that are illegal in XML
    # Replace all control characters except tab (9), newline (10), and carriage return (13)
    cleaned = []
    for char in value_str:
        code = ord(char)
        if code < 32 and code not in (9, 10, 13):
            # Replace illegal control characters with space
            cleaned.append(' ')
        elif code > 127 and code < 160:
            # Replace non-breaking spaces and other problematic high ASCII
            cleaned.append(' ')
        elif code == 8232 or code == 8233:
            # Replace line separator and paragraph separator
            cleaned.append(' ')
        else:
            cleaned.append(char)

    return ''.join(cleaned)

def create_qc_enhanced_excel_with_validation(
    excel_file_content: bytes,
    validation_results: Dict[str, Any],
    qc_results: Optional[Dict[str, Dict[str, Any]]] = None,
    config_data: Dict[str, Any] = None,
    session_id: str = "",
    skip_history: bool = False
):
    """
    Create QC-enhanced Excel file with validation results and QC data.

    Args:
        excel_file_content: Original Excel file content
        validation_results: Validation results from the validator
        qc_results: QC results merged with validation results (from QC integration)
        config_data: Configuration data
        session_id: Session ID for tracking
        skip_history: If True, skip loading existing Details sheet

    Returns:
        Excel file content as bytes, or None if creation failed
    """
    if not EXCEL_ENHANCEMENT_AVAILABLE:
        logger.warning("Enhanced Excel not available, skipping QC Excel creation")
        return None

    try:
        # Create Excel buffer
        excel_buffer = io.BytesIO()

        # Load original Excel data
        workbook = openpyxl.load_workbook(io.BytesIO(excel_file_content))

        # Select the appropriate sheet
        if 'Results' in workbook.sheetnames:
            worksheet = workbook['Results']
            logger.info(f"Using 'Results' sheet as data source for QC-enhanced Excel creation")
        elif len(workbook.sheetnames) > 0:
            worksheet = workbook[workbook.sheetnames[0]]
            logger.info(f"Using first sheet '{worksheet.title}' as data source for QC-enhanced Excel creation")
        else:
            worksheet = workbook.active

        # Get headers and data
        headers = [cell.value for cell in worksheet[1]]

        # Convert to list of dictionaries
        rows_data = []
        for row_idx in range(2, worksheet.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers):
                if header:
                    cell_value = worksheet.cell(row=row_idx, column=col_idx + 1).value
                    row_data[header] = str(cell_value) if cell_value is not None else ""
            rows_data.append(row_data)

        # Get ID fields from config for proper row key generation
        id_fields = []
        if config_data:
            for target in config_data.get('validation_targets', []):
                if target.get('importance', '').upper() == 'ID':
                    # Support both 'name' and 'column' fields
                    field_name = target.get('name') or target.get('column')
                    if field_name:
                        id_fields.append(field_name)

        logger.info(f"QC-Enhanced Excel: Processing {len(rows_data)} rows with {len(headers)} columns, ID fields: {id_fields}")

        # Generate row keys for matching validation results
        from row_key_utils import generate_row_key
        row_keys = []
        for row_data in rows_data:
            try:
                row_key = generate_row_key(row_data, id_fields)
                row_keys.append(row_key)
            except Exception as e:
                logger.warning(f"Failed to generate row key for row {len(row_keys)}: {e}")
                row_keys.append(f"row_{len(row_keys)}")

        # Create new workbook with QC enhancements
        with xlsxwriter.Workbook(excel_buffer, {'strings_to_urls': False, 'nan_inf_to_errors': True}) as workbook:

            # Create QC-enhanced formats
            formats = create_qc_formats(workbook)
            header_format = formats['header_format']
            italic_format = formats['italic_format']
            qc_applied_format = formats['qc_applied_format']

            # 1. Create Results sheet with QC values
            results_sheet = workbook.add_worksheet('Results')

            # Write headers
            for col_idx, col_name in enumerate(headers):
                results_sheet.write(0, col_idx, col_name, header_format)
                results_sheet.set_column(col_idx, col_idx, 20)  # Set column width

            # Write data rows with QC values where applicable
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                for col_idx, col_name in enumerate(headers):
                    original_value = row_data.get(col_name, '')

                    # Check if QC has a value for this field
                    qc_value = original_value
                    qc_applied = False

                    if qc_results and row_key in qc_results:
                        row_qc_data = qc_results[row_key]
                        if col_name in row_qc_data:
                            field_qc_data = row_qc_data[col_name]
                            qc_applied = field_qc_data.get('qc_applied', False)
                            if qc_applied:
                                qc_value = field_qc_data.get('qc_entry', original_value)

                    # Write value with appropriate formatting
                    cell_format = qc_applied_format if qc_applied else None
                    results_sheet.write(row_idx + 1, col_idx, safe_for_excel(qc_value), cell_format)

            # 2. Create QC-Enhanced Details sheet
            details_sheet = workbook.add_worksheet('Details')

            # Build QC-enhanced detail headers
            detail_headers = get_qc_enhanced_detail_headers(id_fields)

            for col_idx, header in enumerate(detail_headers):
                details_sheet.write(0, col_idx, header, header_format)

            # Set column widths
            base_widths = get_qc_enhanced_column_widths()
            col_idx = 0
            details_sheet.set_column(col_idx, col_idx, base_widths[0])  # Row Key
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, base_widths[1])  # Identifier
            col_idx += 1

            # ID field columns
            for _ in id_fields:
                details_sheet.set_column(col_idx, col_idx, 20)
                col_idx += 1

            # Remaining columns
            for width in base_widths[2:]:
                details_sheet.set_column(col_idx, col_idx, width)
                col_idx += 1

            detail_row = 1
            current_timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

            # Write QC-enhanced detail rows
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Create identifier from ID fields
                identifier_parts = []
                for id_field in id_fields:
                    if id_field in row_data:
                        identifier_parts.append(f"{id_field}: {row_data[id_field]}")
                identifier = ", ".join(identifier_parts) if identifier_parts else f"Row {row_idx + 1}"

                # Get validation and QC results for this row
                row_validation_data = None
                row_qc_data = {}

                # Get validation results
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]

                # Get QC results
                if qc_results and row_key in qc_results:
                    row_qc_data = qc_results[row_key]

                if row_validation_data and isinstance(row_validation_data, dict):
                    for field_name, field_data in row_validation_data.items():
                        if isinstance(field_data, dict) and 'confidence_level' in field_data:

                            # Get QC data for this field if available
                            field_qc_data = row_qc_data.get(field_name, {})

                            # Format the result for Excel
                            original_value = row_data.get(field_name, '')

                            # If we have QC data, use it; otherwise create from validation data
                            if field_qc_data:
                                formatted_result = format_qc_result_for_excel(
                                    merged_qc_result=field_qc_data,
                                    field_name=field_name,
                                    original_value=original_value
                                )
                            else:
                                # Create QC format from validation data (no QC applied)
                                formatted_result = {
                                    'column': field_name,
                                    'original_value': original_value,
                                    'updated_value': field_data.get('value', ''),
                                    'qc_value': field_data.get('value', ''),
                                    'qc_applied': 'No',
                                    'qc_action': 'No Change',
                                    'qc_reasoning': '',
                                    'final_confidence': field_data.get('confidence_level', ''),
                                    'original_confidence': field_data.get('original_confidence', ''),
                                    'quote': field_data.get('reasoning', ''),
                                    'sources': ', '.join(field_data.get('sources', [])),
                                    'explanation': field_data.get('explanation', ''),
                                    'update_required': 'Yes' if field_data.get('value', '') != original_value else 'No',
                                    'substantially_different': 'No',
                                    'consistent_with_model': 'Yes'
                                }

                            # Get ID field values
                            id_field_values = [str(row_data.get(id_field, '')) for id_field in id_fields]

                            # Write the detail row
                            write_qc_enhanced_detail_row(
                                worksheet=details_sheet,
                                row_num=detail_row,
                                row_key=row_key,
                                identifier=identifier,
                                id_field_values=id_field_values,
                                formatted_result=formatted_result,
                                model_name=field_data.get('model', 'Unknown'),
                                timestamp=current_timestamp,
                                formats=formats
                            )

                            detail_row += 1

            # 3. Create Reasons sheet (enhanced with QC reasoning)
            reasons_sheet = workbook.add_worksheet('Reasons')
            reasons_headers = ["Row Key", "Field", "Explanation", "QC Applied", "QC Reasoning", "Update Required", "Substantially Different"]

            for col_idx, header in enumerate(reasons_headers):
                reasons_sheet.write(0, col_idx, header, header_format)
                reasons_sheet.set_column(col_idx, col_idx, 30)

            reason_row = 1
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation and QC results for this row
                row_validation_data = validation_results.get(row_key, validation_results.get(str(row_idx), validation_results.get(row_idx, {})))
                row_qc_data = qc_results.get(row_key, {}) if qc_results else {}

                if isinstance(row_validation_data, dict):
                    for field_name, field_data in row_validation_data.items():
                        if isinstance(field_data, dict) and 'confidence_level' in field_data:
                            # Get QC data for this field
                            field_qc_data = row_qc_data.get(field_name, {})
                            qc_applied = field_qc_data.get('qc_applied', False)

                            reasons_sheet.write(reason_row, 0, safe_for_excel(row_key))
                            reasons_sheet.write(reason_row, 1, safe_for_excel(field_name))
                            reasons_sheet.write(reason_row, 2, safe_for_excel(field_data.get('explanation', '')))
                            reasons_sheet.write(reason_row, 3, 'Yes' if qc_applied else 'No')
                            reasons_sheet.write(reason_row, 4, safe_for_excel(field_qc_data.get('qc_reasoning', '')))

                            original_value = row_data.get(field_name, '')
                            final_value = field_qc_data.get('qc_entry', field_data.get('value', '')) if qc_applied else field_data.get('value', '')

                            reasons_sheet.write(reason_row, 5, 'Yes' if final_value != original_value else 'No')
                            reasons_sheet.write(reason_row, 6, 'Yes' if qc_applied else 'No')

                            reason_row += 1

        excel_buffer.seek(0)
        result = excel_buffer.read()
        logger.info(f"Created QC-enhanced Excel file with {len(rows_data)} data rows")
        return result

    except Exception as e:
        logger.error(f"Error creating QC-enhanced Excel: {str(e)}")
        return None