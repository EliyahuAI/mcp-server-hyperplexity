#!/usr/bin/env python3
"""
Consolidated S3 Table Parser
Unified parser for CSV/Excel files from S3 used by both interface and validation code
"""

import boto3
import csv
import json
import logging
import os
import tempfile
import time
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
from functools import wraps
from botocore.exceptions import ClientError

logger = logging.getLogger(__name__)

def retry_s3_operation(max_attempts=3, initial_delay=2.0, backoff_factor=2.0):
    """
    Retry decorator for S3 operations with exponential backoff.

    Args:
        max_attempts: Maximum number of retry attempts (default: 3)
        initial_delay: Initial delay in seconds before first retry (default: 2.0)
        backoff_factor: Multiplier for delay between retries (default: 2.0)
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            delay = initial_delay
            last_exception = None

            for attempt in range(max_attempts):
                try:
                    return func(*args, **kwargs)
                except ClientError as e:
                    error_code = e.response.get('Error', {}).get('Code', '')

                    # Only retry on 404 (NoSuchKey) errors
                    if error_code == '404' or error_code == 'NoSuchKey':
                        last_exception = e

                        if attempt < max_attempts - 1:
                            logger.warning(
                                f"S3 operation failed with 404 (attempt {attempt + 1}/{max_attempts}). "
                                f"Retrying in {delay:.1f} seconds..."
                            )
                            time.sleep(delay)
                            delay *= backoff_factor
                        else:
                            logger.error(
                                f"S3 operation failed after {max_attempts} attempts. "
                                f"File may not exist or is still being created."
                            )
                    else:
                        # For other errors, don't retry - re-raise immediately
                        raise
                except Exception as e:
                    # For non-ClientError exceptions, don't retry
                    raise

            # If we exhausted all retries, raise the last exception
            if last_exception:
                raise last_exception

        return wrapper
    return decorator

# Try to import openpyxl
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel files won't be supported")

class S3TableParser:
    """Unified S3 table-to-JSON parser used by both interface and validation code"""

    def __init__(self, enable_cleaning_log: bool = True, output_dir: str = None):
        self.s3_client = boto3.client('s3')
        self.logger = logging.getLogger(__name__)
        self.enable_cleaning_log = enable_cleaning_log
        self.output_dir = output_dir if output_dir is not None else (
            '/tmp/results' if os.environ.get('AWS_LAMBDA_FUNCTION_NAME') else 'results'
        )
        self.cleaning_logger = None

        if self.enable_cleaning_log:
            try:
                from table_cleaning_logger import TableCleaningLogger
                # Pass None to let TableCleaningLogger auto-detect Lambda environment
                self.cleaning_logger = TableCleaningLogger(None)
            except ImportError:
                self.logger.debug("TableCleaningLogger not available, cleaning operations won't be logged")

    def _normalize_column_name(self, col_name: str) -> str:
        """
        Normalize column names by converting special Unicode characters to ASCII equivalents.
        This prevents mismatches when AI generates configs (AI often loses special chars).
        """
        if not col_name:
            return col_name

        # Replace special Unicode dashes with regular hyphen
        normalized = col_name.replace('\u2013', '-')  # en dash -> hyphen
        normalized = normalized.replace('\u2014', '-')  # em dash -> hyphen
        normalized = normalized.replace('\u2212', '-')  # minus sign -> hyphen

        # Replace special Unicode quotes with regular quotes
        normalized = normalized.replace('\u2018', "'")  # left single quote
        normalized = normalized.replace('\u2019', "'")  # right single quote
        normalized = normalized.replace('\u201C', '"')  # left double quote
        normalized = normalized.replace('\u201D', '"')  # right double quote

        # Replace special spaces
        normalized = normalized.replace('\u00A0', ' ')  # non-breaking space -> space
        normalized = normalized.replace('\u2009', ' ')  # thin space -> space

        return normalized

    @retry_s3_operation(max_attempts=3, initial_delay=2.0, backoff_factor=2.0)
    def _check_s3_object_exists(self, bucket: str, key: str) -> bool:
        """
        Check if an S3 object exists with retry logic.
        Wrapper method to apply retry decorator.
        """
        self.s3_client.head_object(Bucket=bucket, Key=key)
        return True

    def _get_preferred_file_key(self, bucket: str, key: str) -> Tuple[str, bool]:
        """
        Check if a focused/cleaned version exists and return the preferred key.
        Looks for the cleaned file in the SAME DIRECTORY as the original file.

        Args:
            bucket: S3 bucket name
            key: Original S3 object key (e.g., results/domain/user/session/file.csv)

        Returns:
            Tuple of (preferred_key, is_focused_version)
        """
        # Extract directory and filename from key
        key_parts = key.rsplit('/', 1)
        if len(key_parts) == 2:
            directory, filename = key_parts
        else:
            directory = ""
            filename = key

        base_name = Path(filename).stem
        extension = Path(filename).suffix

        # Build cleaned file paths in the SAME directory as the original
        focused_patterns = []
        if directory:
            # Check for cleaned versions in the same directory
            focused_patterns = [
                f"{directory}/{base_name}_cleaned{extension}",
                f"{directory}/{base_name}_focused{extension}",
            ]
        else:
            # If no directory, check in root
            focused_patterns = [
                f"{base_name}_cleaned{extension}",
                f"{base_name}_focused{extension}",
            ]

        # Check for focused/cleaned versions with reduced retries
        for focused_key in focused_patterns:
            try:
                # Use head_object directly without retry for existence check
                self.s3_client.head_object(Bucket=bucket, Key=focused_key)
                self.logger.info(f"Found cleaned version: {focused_key}")
                return focused_key, True
            except ClientError as e:
                error_code = e.response.get('Error', {}).get('Code', '')
                if error_code == '404' or error_code == 'NoSuchKey':
                    continue
                else:
                    # For non-404 errors, raise
                    raise

        # No focused version found, use original
        self.logger.debug(f"No cleaned version found for {key}, using original")
        return key, False

    @retry_s3_operation(max_attempts=3, initial_delay=2.0, backoff_factor=2.0)
    def _download_s3_object(self, bucket: str, key: str) -> bytes:
        """
        Download S3 object with retry logic.
        Wrapper method to apply retry decorator.
        """
        response = self.s3_client.get_object(Bucket=bucket, Key=key)
        return response['Body'].read()

    def parse_s3_table(self, bucket: str, key: str, sheet_name: Optional[str] = None,
                      extract_formulas: bool = False, id_fields: Optional[List[str]] = None,
                      use_focused: bool = True, save_cleaned: bool = True) -> Dict[str, Any]:
        """
        Parse Excel/CSV from S3 into standard JSON format

        Args:
            bucket: S3 bucket name
            key: S3 object key
            sheet_name: Excel sheet name (optional, uses first sheet if not specified)
            extract_formulas: Whether to extract Excel formulas with descriptions (default: False)
            id_fields: List of ID field names for row key generation (optional)
            use_focused: Whether to prefer focused/cleaned versions if available (default: True)
            save_cleaned: Whether to save cleaned version if cleaning is performed (default: True)

        Returns:
            Structured table data with metadata, optionally including formulas.
            If id_fields provided, each row will include a 'row_key' field.
        """
        try:
            # CRITICAL: Check for cached parsed JSON first to ensure row key consistency
            # This prevents regenerating row keys with different id_fields
            parsed_cache_key = self._get_parsed_cache_key(key)
            cached_result = self._try_load_parsed_cache(bucket, parsed_cache_key, id_fields)
            if cached_result:
                self.logger.info(f"[SUCCESS] Using cached parsed table from {parsed_cache_key}")
                return cached_result

            # Check for focused version if enabled
            actual_key = key
            is_focused = False
            if use_focused:
                actual_key, is_focused = self._get_preferred_file_key(bucket, key)

            # Download file from S3 (with retry logic)
            self.logger.info(f"Downloading table from S3: s3://{bucket}/{actual_key}")
            file_content = self._download_s3_object(bucket, actual_key)

            # Keep track of original content if we'll be cleaning
            original_content = file_content if save_cleaned and not is_focused else None

            # Determine file type from key
            filename = actual_key.split('/')[-1]

            # Initialize cleaning log if needed
            if self.cleaning_logger and not is_focused:
                original_shape = self._get_file_shape(file_content, filename)
                self.cleaning_logger.start_file_cleaning(filename,
                                                        'csv' if filename.endswith('.csv') else 'excel',
                                                        original_shape,
                                                        sheet_name=sheet_name)

            # Parse the file
            if filename.endswith('.csv'):
                result = self._parse_csv_content(file_content, filename, id_fields)
            elif filename.endswith(('.xlsx', '.xls')):
                if not OPENPYXL_AVAILABLE:
                    raise ImportError("openpyxl is required for Excel file support")
                result = self._parse_excel_content(file_content, filename, sheet_name, extract_formulas, id_fields)
            else:
                raise ValueError(f"Unsupported file format: {filename}")

            # Add metadata about focused version
            result['metadata']['used_focused_version'] = is_focused
            result['metadata']['original_key'] = key if is_focused else None

            # Finalize cleaning log and save files if we did cleaning
            if self.cleaning_logger and not is_focused and save_cleaned:
                final_shape = (result['total_rows'], result['total_columns'])
                self.cleaning_logger.finalize_cleaning(final_shape)

                # Save the cleaning log
                log_path = self.cleaning_logger.save_log()
                result['metadata']['cleaning_log_path'] = log_path

                # If we have original content and made changes, save both versions
                if original_content:
                    # For Excel files, we need to regenerate the cleaned content
                    # since we modified it in memory
                    if filename.endswith(('.xlsx', '.xls')):
                        # Create a cleaned version using the parsed data
                        # Note: This is simplified - in production we'd want to
                        # preserve the actual cleaned Excel file
                        self.logger.info("Note: Excel file cleaning saved to log only")
                        # TODO: Implement Excel file saving if needed
                    else:
                        # For CSV files, regenerate from cleaned data and upload to S3
                        cleaned_csv = self._generate_csv_from_data(result)
                        saved_files = self.cleaning_logger.save_original_and_cleaned(
                            original_content=original_content,
                            cleaned_content=cleaned_csv.encode('utf-8'),
                            base_filename=filename,
                            save_original_copy=True
                        )
                        result['metadata']['saved_files'] = saved_files
                        self.logger.info(f"Saved cleaned CSV files locally: {saved_files}")

                        # Upload cleaned version to S3 in same directory as original
                        try:
                            # Extract directory from original key
                            key_parts = key.rsplit('/', 1)
                            if len(key_parts) == 2:
                                directory = key_parts[0]
                                cleaned_s3_key = f"{directory}/{Path(filename).stem}_cleaned{Path(filename).suffix}"
                            else:
                                cleaned_s3_key = f"{Path(filename).stem}_cleaned{Path(filename).suffix}"

                            # Upload cleaned CSV to S3
                            self.s3_client.put_object(
                                Bucket=bucket,
                                Key=cleaned_s3_key,
                                Body=cleaned_csv.encode('utf-8'),
                                ContentType='text/csv'
                            )
                            self.logger.info(f"[SUCCESS] Uploaded cleaned file to S3: s3://{bucket}/{cleaned_s3_key}")
                            result['metadata']['s3_cleaned_key'] = cleaned_s3_key
                        except Exception as e:
                            self.logger.error(f"[ERROR] Failed to upload cleaned file to S3: {str(e)}")

            # Save parsed JSON to cache for future use
            # This ensures row keys are never regenerated with different id_fields
            self._save_parsed_cache(bucket, parsed_cache_key, result, id_fields)

            return result

        except Exception as e:
            self.logger.error(f"Failed to parse S3 table {bucket}/{key}: {str(e)}")
            raise

    def _get_parsed_cache_key(self, key: str) -> str:
        """
        Generate the cache key for a parsed table.

        Args:
            key: Original S3 key (e.g., 'uploads/table.xlsx')

        Returns:
            Cache key with _parsed.json suffix (e.g., 'uploads/table_parsed.json')
        """
        # Split the key into directory and filename
        if '/' in key:
            directory, filename = key.rsplit('/', 1)
        else:
            directory = ''
            filename = key

        # Remove file extension and add _parsed.json
        base_name = Path(filename).stem
        cache_filename = f"{base_name}_parsed.json"

        # Reconstruct the full cache key
        if directory:
            cache_key = f"{directory}/{cache_filename}"
        else:
            cache_key = cache_filename

        return cache_key

    def _try_load_parsed_cache(self, bucket: str, cache_key: str, id_fields: Optional[List[str]]) -> Optional[Dict]:
        """
        Attempt to load cached parsed JSON from S3.

        Args:
            bucket: S3 bucket name
            cache_key: Cache key (with _parsed.json suffix)
            id_fields: ID fields that should match the cache

        Returns:
            Cached parsed data if valid, None otherwise
        """
        try:
            # Get the original source file key (remove _parsed.json suffix)
            if cache_key.endswith('_parsed.json'):
                # Reconstruct original key - need to determine file type
                # Try common extensions in order of likelihood
                base_key = cache_key[:-len('_parsed.json')]

                # Check if any of these files exist
                for ext in ['.xlsx', '.csv', '.xls']:
                    original_key = f"{base_key}{ext}"
                    try:
                        # Get metadata for both files
                        cache_metadata = self.s3_client.head_object(Bucket=bucket, Key=cache_key)
                        original_metadata = self.s3_client.head_object(Bucket=bucket, Key=original_key)

                        cache_modified = cache_metadata['LastModified']
                        original_modified = original_metadata['LastModified']

                        # Check if cache is newer than source
                        if cache_modified < original_modified:
                            self.logger.info(f"[CACHE_SKIP] Cache is older than source file. Cache: {cache_modified}, Source: {original_modified}")
                            return None

                        # Cache is valid, proceed to load it
                        break
                    except ClientError as e:
                        if e.response.get('Error', {}).get('Code') in ['404', 'NoSuchKey']:
                            continue  # Try next extension
                        raise
                else:
                    # No original file found
                    self.logger.warning(f"[CACHE_SKIP] Could not find original source file for cache validation")
                    return None

            # Try to load the cached file
            self.logger.debug(f"[CACHE_CHECK] Attempting to load cache from s3://{bucket}/{cache_key}")
            response = self.s3_client.get_object(Bucket=bucket, Key=cache_key)
            cached_data = json.loads(response['Body'].read())

            # Validate that the cache has the expected structure
            if 'metadata' not in cached_data or 'data' not in cached_data:
                self.logger.warning(f"[CACHE_SKIP] Cache missing required structure (metadata/data)")
                return None

            # Validate that id_fields match what was used to generate the cache
            cached_id_fields = cached_data.get('metadata', {}).get('id_fields')

            # Handle None vs empty list comparison
            if id_fields != cached_id_fields:
                self.logger.warning(
                    f"[CACHE_SKIP] ID fields mismatch. "
                    f"Requested: {id_fields}, Cached: {cached_id_fields}"
                )
                return None

            # Verify that rows have _row_key if id_fields were used
            if id_fields and cached_data.get('data'):
                first_row = cached_data['data'][0]
                if '_row_key' not in first_row:
                    self.logger.warning(f"[CACHE_SKIP] Cache missing _row_key despite having id_fields")
                    return None

            self.logger.info(
                f"[CACHE_HIT] Loaded cached parsed table with {len(cached_data.get('data', []))} rows "
                f"and id_fields={id_fields}"
            )
            return cached_data

        except ClientError as e:
            error_code = e.response.get('Error', {}).get('Code', '')
            if error_code in ['404', 'NoSuchKey']:
                self.logger.debug(f"[CACHE_MISS] No cache found at s3://{bucket}/{cache_key}")
            else:
                self.logger.warning(f"[CACHE_ERROR] Failed to load cache: {str(e)}")
            return None
        except Exception as e:
            self.logger.warning(f"[CACHE_ERROR] Failed to load/validate cache: {str(e)}")
            return None

    def _save_parsed_cache(self, bucket: str, cache_key: str, parsed_data: Dict, id_fields: Optional[List[str]]) -> None:
        """
        Save parsed JSON to S3 cache for future use.

        Args:
            bucket: S3 bucket name
            cache_key: Cache key (with _parsed.json suffix)
            parsed_data: Parsed table data to cache
            id_fields: ID fields used for row key generation
        """
        try:
            # Add id_fields to metadata for validation on cache load
            if 'metadata' not in parsed_data:
                parsed_data['metadata'] = {}
            parsed_data['metadata']['id_fields'] = id_fields

            # Serialize to JSON
            cache_content = json.dumps(parsed_data, indent=2, ensure_ascii=False)

            # Upload to S3
            self.s3_client.put_object(
                Bucket=bucket,
                Key=cache_key,
                Body=cache_content.encode('utf-8'),
                ContentType='application/json'
            )

            self.logger.info(
                f"[CACHE_SAVE] Saved parsed table cache to s3://{bucket}/{cache_key} "
                f"with {len(parsed_data.get('data', []))} rows and id_fields={id_fields}"
            )

        except Exception as e:
            # Don't fail the entire parsing operation if cache save fails
            self.logger.warning(f"[CACHE_ERROR] Failed to save cache: {str(e)}")

    def _generate_csv_from_data(self, parsed_data: Dict[str, Any]) -> str:
        """Generate CSV content from parsed data structure."""
        import io
        import csv

        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        headers = parsed_data['column_names']
        writer.writerow(headers)

        # Write data rows
        for row in parsed_data['data']:
            row_values = []
            for header in headers:
                # Skip internal fields like _row_key
                if not header.startswith('_'):
                    value = row.get(header, '')
                    row_values.append(value)
            writer.writerow(row_values)

        return output.getvalue()

    def _get_file_shape(self, file_content: bytes, filename: str) -> Tuple[int, int]:
        """Get the shape (rows, columns) of a file for logging."""
        try:
            if filename.endswith('.csv'):
                text_content = file_content.decode('utf-8-sig')
                reader = csv.reader(text_content.splitlines())
                rows = list(reader)
                return (len(rows), len(rows[0]) if rows else 0)
            else:
                # Excel file
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp.write(file_content)
                    tmp_path = tmp.name

                workbook = load_workbook(tmp_path, read_only=True)
                worksheet = workbook.active
                shape = (worksheet.max_row, worksheet.max_column)
                workbook.close()
                os.unlink(tmp_path)
                return shape
        except:
            return (0, 0)
    
    def _parse_csv_content(self, file_content: bytes, filename: str, id_fields: Optional[List[str]] = None) -> Dict[str, Any]:
        """Parse CSV content into structured format"""
        try:
            # Try multiple encodings to handle different CSV file encodings (expanded list)
            encodings_to_try = [
                'utf-8', 'utf-8-sig',  # UTF-8 with and without BOM
                'latin-1', 'cp1252', 'iso-8859-1',  # Windows/Western European
                'utf-16', 'utf-16le', 'utf-16be',  # UTF-16 variants
                'cp437', 'cp850',  # DOS/OEM encodings
                'ascii'  # Basic ASCII as last resort
            ]
            text_content = None
            used_encoding = None
            
            for encoding in encodings_to_try:
                try:
                    text_content = file_content.decode(encoding)
                    used_encoding = encoding
                    break
                except UnicodeDecodeError:
                    continue
            
            if text_content is None:
                raise ValueError(f"Could not decode CSV file with any of the tried encodings: {encodings_to_try}")
                
            print(f"Successfully decoded CSV using {used_encoding} encoding")
            
            # Parse CSV with standard method first
            try:
                reader = csv.reader(text_content.splitlines())
                rows = list(reader)
                
                if not rows:
                    raise ValueError("Empty CSV file")
                    
                # Quick validation - check if we got reasonable parsing
                if not self._is_valid_csv_parse(rows):
                    raise ValueError("Standard CSV parsing produced poor results, trying fallback")
                    
            except (csv.Error, ValueError) as e:
                self.logger.warning(f"Standard CSV parsing failed: {e}, trying robust fallback")
                # Fallback to robust parsing with delimiter detection
                rows = self._parse_csv_with_robust_fallback(text_content)
            
            # Find actual table start with robust detection
            table_start_row, headers = self._find_table_start_csv(rows)
            data_rows = rows[table_start_row + 1:]
            
            # Clean data rows
            data_rows = self._clean_data_rows_csv(data_rows, len(headers))
            
            # Clean headers
            clean_headers = [self._normalize_column_name(str(header).strip()) if header else f"Column_{i+1}"
                           for i, header in enumerate(headers)]

            # Import row key generation function
            generate_row_key = None
            try:
                from row_key_utils import generate_row_key
            except ImportError:
                self.logger.warning("row_key_utils not available - row keys will not be generated")

            # Convert to structured format
            structured_data = []

            # If id_fields provided, do two-pass with deduplication check
            if id_fields and generate_row_key:
                # First pass: collect all rows and detect duplicates
                temp_rows = []
                id_hash_counts = {}

                for row in data_rows:
                    # Skip completely empty rows only - keep all data rows even if sparse
                    non_empty_count = sum(1 for cell in row if cell and str(cell).strip())
                    if non_empty_count == 0:
                        continue  # Skip completely empty rows

                    # After headers are found, keep ALL data rows (even sparse ones)
                    row_dict = {}
                    for i, header in enumerate(clean_headers):
                        cell_value = row[i] if i < len(row) else ""
                        row_dict[header] = str(cell_value).strip() if cell_value else ""
                    temp_rows.append(row_dict)

                    # Generate ID-based hash to detect duplicates
                    id_hash = generate_row_key(row_dict, primary_keys=id_fields)
                    if id_hash not in id_hash_counts:
                        id_hash_counts[id_hash] = 0
                    id_hash_counts[id_hash] += 1

                # Second pass: assign row keys (ID hash for unique, full hash for duplicates)
                for row_dict in temp_rows:
                    id_hash = generate_row_key(row_dict, primary_keys=id_fields)

                    if id_hash_counts[id_hash] > 1:
                        # Duplicate detected - use full row hash
                        row_key = generate_row_key(row_dict, primary_keys=None)
                    else:
                        # Unique ID - use ID hash
                        row_key = id_hash

                    row_dict['_row_key'] = row_key
                    structured_data.append(row_dict)

                # Log duplicate summary
                duplicate_count = sum(1 for count in id_hash_counts.values() if count > 1)
                if duplicate_count > 0:
                    self.logger.info(f"CSV: Found {duplicate_count} duplicate ID groups, using full-row hashing for those")
            else:
                # No id_fields - just use full row hash for all rows
                for row in data_rows:
                    # Skip completely empty rows only - keep all data rows even if sparse
                    non_empty_count = sum(1 for cell in row if cell and str(cell).strip())
                    if non_empty_count == 0:
                        continue  # Skip completely empty rows

                    # After headers are found, keep ALL data rows (even sparse ones)
                    row_dict = {}
                    for i, header in enumerate(clean_headers):
                        cell_value = row[i] if i < len(row) else ""
                        row_dict[header] = str(cell_value).strip() if cell_value else ""

                    # Generate full row hash (AFTER building complete row dict)
                    if generate_row_key:
                        row_key = generate_row_key(row_dict, primary_keys=None)
                        row_dict['_row_key'] = row_key

                    structured_data.append(row_dict)
            
            return {
                'filename': filename,
                'total_rows': len(structured_data),
                'total_columns': len(clean_headers),
                'column_names': clean_headers,
                'data': structured_data,
                'metadata': {
                    'file_type': 'csv',
                    'source': 's3',
                    'sample_rows': min(5, len(structured_data))
                }
            }
            
        except Exception as e:
            self.logger.error(f"Failed to parse CSV content: {str(e)}")
            raise
    
    def _parse_excel_content(self, file_content: bytes, filename: str, sheet_name: Optional[str] = None, extract_formulas: bool = False, id_fields: Optional[List[str]] = None) -> Dict[str, Any]:
        """Parse Excel content into structured format"""
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_file_path = temp_file.name

            try:
                # Load workbook (can't use read_only if we need formulas or to delete rows)
                # When we need to preserve formulas through row deletion, we need write access
                workbook = load_workbook(temp_file_path, read_only=False, data_only=not extract_formulas)

                # Log available sheets if we have a logger
                if self.cleaning_logger:
                    self.cleaning_logger.set_available_sheets(workbook.sheetnames)

                # Select worksheet
                if sheet_name:
                    if sheet_name not in workbook.sheetnames:
                        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
                    worksheet = workbook[sheet_name]
                else:
                    # Default behavior: try "Updated Values" sheet first, then fall back to first sheet
                    if "Updated Values" in workbook.sheetnames:
                        worksheet = workbook["Updated Values"]
                        sheet_name = "Updated Values"
                        print(f"Using 'Updated Values' sheet by default")
                    else:
                        # Use first sheet (workbook.active might not be the first sheet)
                        worksheet = workbook[workbook.sheetnames[0]]
                        sheet_name = workbook.sheetnames[0]
                        print(f"No 'Updated Values' sheet found, using first sheet: '{sheet_name}'")

                # Import advanced table detector if available
                use_advanced_detection = False
                try:
                    from excel_table_detector import ExcelTableDetector
                    detector = ExcelTableDetector()
                    use_advanced_detection = True
                    self.logger.info("Using advanced Excel table detection")
                except ImportError:
                    self.logger.debug("Advanced table detector not available, using basic detection")

                # Read data with robust table detection
                if extract_formulas:
                    # Get both cell objects and values for formula extraction
                    all_rows, all_cell_objects = self._extract_rows_with_formulas(worksheet)
                else:
                    all_rows = list(worksheet.iter_rows(values_only=True))
                    all_cell_objects = None

                if not all_rows:
                    raise ValueError("Empty Excel worksheet")

                # Advanced table boundary detection if available
                table_info = None
                if use_advanced_detection:
                    tables = detector.detect_table_boundaries(worksheet, start_row=0)
                    if tables:
                        # First, look for tables with ID columns - these are most likely the real data tables
                        id_column_tables = []
                        for table in tables:
                            if table['table_type'] == 'data':
                                # Check if headers contain ID columns
                                has_id_col = False
                                for header in table.get('headers', []):
                                    if header and any(pattern in str(header).upper() for pattern in
                                                     ['_ID', ' ID', 'ID ', 'PRODUCT_ID', 'COMPANY_ID',
                                                      'CUSTOMER_ID', 'ITEM_ID', 'ORDER_ID', 'USER_ID']):
                                        has_id_col = True
                                        break
                                if has_id_col:
                                    id_column_tables.append(table)

                        # Prefer tables with ID columns
                        if id_column_tables:
                            table_info = id_column_tables[0]
                            self.logger.info(
                                f"Found data table with ID columns: rows {table_info['start_row']}-{table_info['end_row']}, "
                                f"has_summary={table_info['has_summary']}"
                            )
                        else:
                            # Fallback to first data table
                            for table in tables:
                                if table['table_type'] == 'data':
                                    table_info = table
                                    self.logger.info(
                                        f"Found data table: rows {table['start_row']}-{table['end_row']}, "
                                        f"has_summary={table['has_summary']}"
                                    )
                                    break

                        if not table_info and tables:
                            # No data table found, use the first table
                            table_info = tables[0]

                # Find actual table start (skip empty rows and detect headers)
                # Initialize header_column_indices to avoid uninitialized variable error
                header_column_indices = None

                if table_info:
                    # Use detected boundaries
                    table_start_row = table_info['start_row'] - 1  # Convert to 0-based
                    headers = table_info['headers']
                    header_column_indices = table_info.get('header_column_indices', None)

                    # Extract only data rows (exclude summary if present)
                    if table_info['has_summary'] and table_info['summary_start_row']:
                        # summary_start_row is 1-based from detector, convert to 0-based
                        summary_row_idx = table_info['summary_start_row'] - 1
                        data_rows = all_rows[table_start_row + 1:summary_row_idx]
                        self.logger.info(f"Table info: start_row={table_info['start_row']} (1-based), summary_start_row={table_info['summary_start_row']} (1-based)")
                        self.logger.info(f"Extracting data rows from index {table_start_row + 1} to {summary_row_idx - 1} (0-based)")
                        self.logger.info(f"Slice: all_rows[{table_start_row + 1}:{summary_row_idx}]")
                        self.logger.info(f"Data rows extracted: {len(data_rows)} rows")
                        # Debug: show what's in each row
                        for i, row in enumerate(data_rows[:3]):  # Show first 3 rows
                            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
                            self.logger.debug(f"  Row {i}: {non_empty} non-empty cells, first cell: {row[0] if row else 'empty'}")
                    else:
                        # end_row is 1-based from detector and represents the last data row
                        data_end_idx = table_info['end_row'] - 1  # Convert to 0-based
                        data_rows = all_rows[table_start_row + 1:data_end_idx + 1]
                        self.logger.info(f"Table info: start_row={table_info['start_row']} (1-based), end_row={table_info['end_row']} (1-based)")
                        self.logger.info(f"Extracting data rows from index {table_start_row + 1} to {data_end_idx} (0-based)")
                        self.logger.info(f"Slice: all_rows[{table_start_row + 1}:{data_end_idx + 1}]")
                        self.logger.info(f"Data rows extracted: {len(data_rows)} rows")
                else:
                    # Fallback to basic detection
                    table_start_row, headers = self._find_table_start(all_rows)
                    data_rows = all_rows[table_start_row + 1:]

                # Clean data rows with enhanced filtering if detector available
                if use_advanced_detection:
                    # Note: filter_data_rows expects the first row to be headers, so we prepend headers
                    rows_with_headers = [headers] + data_rows
                    filtered_rows, filter_metadata = detector.filter_data_rows(
                        rows_with_headers, headers, preserve_summary=False
                    )
                    # Remove the header row that we added
                    data_rows = filtered_rows
                    self.logger.info(
                        f"Advanced filtering: removed {filter_metadata['removed_empty']} empty, "
                        f"{filter_metadata['removed_sparse']} sparse, {filter_metadata['removed_metadata']} metadata rows"
                    )
                else:
                    # Use basic cleaning
                    data_rows = self._clean_data_rows(data_rows, len(headers))
                
                # Track original column count before trimming
                original_column_count = len(headers) if not hasattr(headers[0], '__iter__') else len(headers)

                # Clean headers and remove empty header columns
                # First identify which columns have headers
                non_empty_header_indices = []
                clean_headers = []
                for i, header in enumerate(headers):
                    if header is not None and str(header).strip():
                        non_empty_header_indices.append(i)
                        clean_headers.append(self._normalize_column_name(str(header).strip()))

                # If we removed empty header columns, we need to filter data accordingly
                if len(non_empty_header_indices) < len(headers):
                    self.logger.info(f"Removing {len(headers) - len(non_empty_header_indices)} columns without headers")

                    # Log column removal if we have a cleaning logger
                    if self.cleaning_logger:
                        removed_indices = [i for i in range(len(headers)) if i not in non_empty_header_indices]
                        removed_headers = [headers[i] if i < len(headers) else "" for i in removed_indices]
                        self.cleaning_logger.log_column_removal(
                            column_indices=removed_indices,
                            column_headers=removed_headers,
                            removal_reason="No header content"
                        )

                    # Filter data_rows to only include columns with headers
                    filtered_data_rows = []
                    for row in data_rows:
                        filtered_row = [row[i] if i < len(row) else None for i in non_empty_header_indices]
                        filtered_data_rows.append(filtered_row)
                    data_rows = filtered_data_rows

                # Check if we trimmed columns and are extracting formulas
                columns_trimmed = original_column_count > len(clean_headers)
                if extract_formulas and columns_trimmed:
                    self.logger.warning(
                        f"Column trimming detected ({original_column_count} -> {len(clean_headers)}) "
                        f"while extracting formulas. Formula references to trimmed columns will be marked as 'Outside table range'"
                    )

                # Import row key generation function
                generate_row_key = None
                try:
                    from row_key_utils import generate_row_key
                except ImportError:
                    self.logger.warning("row_key_utils not available - row keys will not be generated")

                # Convert to structured format
                structured_data = []
                formula_data = [] if extract_formulas else None

                # First pass: extract all row data and formulas
                temp_rows = []
                temp_formulas = []

                # If we have header column indices from the detector, use them to properly align data
                if header_column_indices and table_info:
                    self.logger.info(f"Using header column indices for data alignment: {header_column_indices}")
                    self.logger.info(f"Processing {len(data_rows)} data rows")
                    for row_idx, row in enumerate(data_rows):
                        # Skip completely empty rows only - keep all data rows even if sparse
                        non_empty_count = sum(1 for cell in row if cell is not None and str(cell).strip())
                        if row_idx < 3:  # Debug first 3 rows
                            self.logger.debug(f"  Data row {row_idx}: non_empty={non_empty_count}, first 4 cells: {row[:4] if row else 'empty'}")
                        if non_empty_count == 0:
                            # Log empty row removal if we have a cleaning logger
                            if self.cleaning_logger:
                                self.cleaning_logger.log_row_removal(
                                    row_indices=[table_start_row + 1 + row_idx],
                                    rows_content=[[str(c) if c else "" for c in row]],
                                    removal_reason="Completely empty row",
                                    row_type='empty'
                                )
                            continue  # Skip completely empty rows

                        # After headers are found, keep ALL data rows (even sparse ones)
                        row_dict = {}
                        formula_dict = {} if extract_formulas else None

                        # Map data from original columns to headers using indices
                        for header_idx, header in enumerate(clean_headers):
                            if header_idx < len(header_column_indices):
                                original_col_idx = header_column_indices[header_idx]
                                cell_value = row[original_col_idx] if original_col_idx < len(row) else None
                            else:
                                cell_value = None

                            row_dict[header] = str(cell_value).strip() if cell_value is not None else ""

                            # Extract formula if requested
                            if extract_formulas and all_cell_objects:
                                actual_row_idx = table_start_row + 1 + row_idx
                                if actual_row_idx < len(all_cell_objects) and header_idx < len(header_column_indices):
                                    original_col_idx = header_column_indices[header_idx]
                                    if original_col_idx < len(all_cell_objects[actual_row_idx]):
                                        cell_obj = all_cell_objects[actual_row_idx][original_col_idx]
                                        formula_info = self._extract_formula_info(cell_obj, clean_headers, worksheet)
                                        if formula_info:
                                            formula_dict[header] = formula_info

                        temp_rows.append(row_dict)
                        if extract_formulas:
                            temp_formulas.append(formula_dict if formula_dict else {})
                else:
                    # Fallback to original logic if no header indices available
                    for row_idx, row in enumerate(data_rows):
                        # Skip completely empty rows only - keep all data rows even if sparse
                        non_empty_count = sum(1 for cell in row if cell is not None and str(cell).strip())
                        if non_empty_count == 0:
                            # Log empty row removal if we have a cleaning logger
                            if self.cleaning_logger:
                                self.cleaning_logger.log_row_removal(
                                    row_indices=[table_start_row + 1 + row_idx],
                                    rows_content=[[str(c) if c else "" for c in row]],
                                    removal_reason="Completely empty row",
                                    row_type='empty'
                                )
                            continue  # Skip completely empty rows

                        # After headers are found, keep ALL data rows (even sparse ones)
                        row_dict = {}
                        formula_dict = {} if extract_formulas else None

                        for i, header in enumerate(clean_headers):
                            cell_value = row[i] if i < len(row) else None
                            row_dict[header] = str(cell_value).strip() if cell_value is not None else ""

                            # Extract formula if requested
                            if extract_formulas and all_cell_objects:
                                actual_row_idx = table_start_row + 1 + row_idx
                                if actual_row_idx < len(all_cell_objects) and i < len(all_cell_objects[actual_row_idx]):
                                    cell_obj = all_cell_objects[actual_row_idx][i]
                                    formula_info = self._extract_formula_info(cell_obj, clean_headers, worksheet)
                                    if formula_info:
                                        formula_dict[header] = formula_info

                        temp_rows.append(row_dict)
                        if extract_formulas:
                            temp_formulas.append(formula_dict if formula_dict else {})

                # Second pass: assign row keys with deduplication if id_fields provided
                if id_fields and generate_row_key:
                    # Detect duplicates
                    id_hash_counts = {}
                    for row_dict in temp_rows:
                        id_hash = generate_row_key(row_dict, primary_keys=id_fields)
                        if id_hash not in id_hash_counts:
                            id_hash_counts[id_hash] = 0
                        id_hash_counts[id_hash] += 1

                    # Assign row keys
                    for row_dict in temp_rows:
                        id_hash = generate_row_key(row_dict, primary_keys=id_fields)

                        if id_hash_counts[id_hash] > 1:
                            # Duplicate detected - use full row hash
                            row_key = generate_row_key(row_dict, primary_keys=None)
                        else:
                            # Unique ID - use ID hash
                            row_key = id_hash

                        row_dict['_row_key'] = row_key
                        structured_data.append(row_dict)

                    # Log duplicate summary
                    duplicate_count = sum(1 for count in id_hash_counts.values() if count > 1)
                    if duplicate_count > 0:
                        self.logger.info(f"Excel: Found {duplicate_count} duplicate ID groups, using full-row hashing for those")

                    # Add formulas if extracted
                    if extract_formulas:
                        formula_data = temp_formulas
                elif generate_row_key:
                    # No id_fields - just use full row hash for all rows
                    for row_dict in temp_rows:
                        row_key = generate_row_key(row_dict, primary_keys=None)
                        row_dict['_row_key'] = row_key
                        structured_data.append(row_dict)

                    # Add formulas if extracted
                    if extract_formulas:
                        formula_data = temp_formulas
                else:
                    # No row key generation available
                    structured_data = temp_rows
                    if extract_formulas:
                        formula_data = temp_formulas
                
                # --- New Logic: Detect external links ---
                has_external_links = False
                if extract_formulas:
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.data_type == 'f' and cell.value and '[' in str(cell.value):
                                has_external_links = True
                                self.logger.info("External link detected in formulas.")
                                break
                        if has_external_links:
                            break
                
                workbook.close()
                
                result = {
                    'filename': filename,
                    'total_rows': len(structured_data),
                    'total_columns': len(clean_headers),
                    'column_names': clean_headers,
                    'data': structured_data,
                    'metadata': {
                        'file_type': 'excel',
                        'sheet_name': sheet_name,
                        'source': 's3',
                        'sample_rows': min(5, len(structured_data)),
                        'has_external_links': has_external_links
                    }
                }

                # Add formula data if extracted
                if extract_formulas and formula_data:
                    result['formulas'] = formula_data
                    result['metadata']['has_formulas'] = True
                    result['metadata']['formula_count'] = sum(len(row_formulas) for row_formulas in formula_data)

                return result
                
            finally:
                # Clean up temp file
                os.unlink(temp_file_path)
                
        except Exception as e:
            self.logger.error(f"Failed to parse Excel content: {str(e)}")
            raise
    
    def get_table_sample(self, bucket: str, key: str, max_rows: int = 10) -> Dict[str, Any]:
        """Get a sample of the table data without loading the entire file"""
        try:
            full_data = self.parse_s3_table(bucket, key)
            
            # Return sample
            sample_data = full_data['data'][:max_rows]
            
            return {
                'filename': full_data['filename'],
                'total_rows': full_data['total_rows'],
                'total_columns': full_data['total_columns'],
                'column_names': full_data['column_names'],
                'sample_data': sample_data,
                'metadata': full_data['metadata']
            }
            
        except Exception as e:
            self.logger.error(f"Failed to get table sample: {str(e)}")
            raise
    
    def analyze_table_structure(self, bucket: str, key: str, extract_formulas: bool = False) -> Dict[str, Any]:
        """Analyze table structure for AI config generation"""
        try:
            # Get sample data - extract formulas if requested for Excel files
            should_extract_formulas = extract_formulas and (key.endswith('.xlsx') or key.endswith('.xls'))

            if should_extract_formulas:
                # Use full parse with formula extraction for Excel files
                full_data = self.parse_s3_table(bucket, key, extract_formulas=True)
                sample = {
                    'filename': full_data['filename'],
                    'total_rows': full_data['total_rows'],
                    'total_columns': full_data['total_columns'],
                    'column_names': full_data['column_names'],
                    'sample_data': full_data['data'][:20],  # First 20 rows
                    'metadata': full_data['metadata']
                }
                # Include formula data if available
                formula_data = full_data.get('formulas', [])
            else:
                # Use regular sample method for faster analysis
                sample = self.get_table_sample(bucket, key, max_rows=20)
                formula_data = []

            # Analyze each column
            column_analysis = {}
            for col_name in sample['column_names']:
                # Get values for this column
                values = [row.get(col_name, '') for row in sample['sample_data']]
                non_empty_values = [v for v in values if v.strip()]

                column_analysis[col_name] = {
                    'data_type': self._infer_data_type(non_empty_values),
                    'non_null_count': len(non_empty_values),
                    'unique_count': len(set(non_empty_values)),
                    'sample_values': sorted(list(set(non_empty_values)))[:5],  # Up to 5 unique samples, deterministically sorted
                    'fill_rate': len(non_empty_values) / len(values) if values else 0.0
                }

            result = {
                'basic_info': {
                    'filename': sample['filename'],
                    'shape': [sample['total_rows'], sample['total_columns']],
                    'column_names': sample['column_names'],
                    'sample_rows_analyzed': len(sample['sample_data'])
                },
                'column_analysis': column_analysis,
                'domain_info': self._infer_domain_info(sample),
                'metadata': sample['metadata']
            }

            # Add formula data if extracted
            if should_extract_formulas and formula_data:
                result['formula_data'] = formula_data
                # Ensure metadata indicates formulas were found
                result['metadata']['has_formulas'] = True
                result['metadata']['formula_count'] = sum(len(row_formulas) for row_formulas in formula_data)
                self.logger.info(f"Formula extraction: {result['metadata']['formula_count']} formulas found in {len(formula_data)} rows")

            return result
            
        except Exception as e:
            self.logger.error(f"Failed to analyze table structure: {str(e)}")
            raise
    
    def _infer_data_type(self, values: List[str]) -> str:
        """Infer data type from sample values"""
        if not values:
            return "Unknown"
        
        # Check for numbers
        numeric_count = 0
        for value in values:
            try:
                float(value)
                numeric_count += 1
            except ValueError:
                pass
        
        if numeric_count / len(values) > 0.8:
            return "Numeric"
        
        # Check for dates (simple heuristic)
        date_keywords = ['date', 'time', '/', '-', '20', '19']
        date_count = sum(1 for value in values if any(kw in value.lower() for kw in date_keywords))
        
        if date_count / len(values) > 0.5:
            return "Date"
        
        return "Text"
    
    def _infer_domain_info(self, sample: Dict[str, Any]) -> Dict[str, Any]:
        """Infer domain information from table structure"""
        column_names = [name.lower() for name in sample['column_names']]
        
        # Domain keywords
        domain_scores = {}
        
        # Biotech/pharma keywords
        biotech_keywords = ['target', 'indication', 'phase', 'compound', 'drug', 'protein', 'gene']
        biotech_score = sum(1 for col in column_names if any(kw in col for kw in biotech_keywords))
        if biotech_score > 0:
            domain_scores['biotech'] = biotech_score
        
        # Competitive intelligence keywords
        ci_keywords = ['company', 'competitor', 'product', 'market', 'revenue', 'stage']
        ci_score = sum(1 for col in column_names if any(kw in col for kw in ci_keywords))
        if ci_score > 0:
            domain_scores['competitive_intelligence'] = ci_score
        
        # Financial keywords
        financial_keywords = ['price', 'cost', 'revenue', 'profit', 'budget', 'finance']
        financial_score = sum(1 for col in column_names if any(kw in col for kw in financial_keywords))
        if financial_score > 0:
            domain_scores['financial'] = financial_score
        
        # Determine likely domain
        if domain_scores:
            likely_domain = max(domain_scores, key=domain_scores.get)
            confidence = domain_scores[likely_domain] / len(column_names)
        else:
            likely_domain = "general"
            confidence = 0.1
        
        return {
            'likely_domain': likely_domain,
            'domain_scores': domain_scores,
            'confidence': confidence
        }
    
    def _find_table_start(self, all_rows):
        """Find the actual start of the table data by detecting headers intelligently."""
        # First, find the maximum column count in the data
        max_cols = 0
        for row in all_rows[:20]:  # Check first 20 rows for max columns
            if row and any(cell is not None and str(cell).strip() for cell in row):
                max_cols = max(max_cols, len(row))

        # Look for rows containing ID columns - these are definitive headers
        for row_idx, row in enumerate(all_rows):
            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            # Check if this row contains ID columns (strong indicator of headers)
            has_id_column = False
            for cell in row:
                if cell is not None:
                    cell_str = str(cell).strip().upper()
                    # Look for ID patterns in column names
                    if any(pattern in cell_str for pattern in ['_ID', ' ID', 'ID ',
                                                                'PRODUCT_ID', 'COMPANY_ID', 'CUSTOMER_ID',
                                                                'ITEM_ID', 'ORDER_ID', 'USER_ID']):
                        has_id_column = True
                        break

            if has_id_column:
                # This is definitely a header row - ID columns must be in headers
                self.logger.debug(f"Found ID column in row {row_idx}, treating as header row")
                headers = self._trim_trailing_empty(row)

                # Log header detection if we have a cleaning logger
                if self.cleaning_logger:
                    id_columns_found = [str(cell).strip() for cell in row
                                       if cell and any(p in str(cell).upper() for p in
                                                      ['_ID', ' ID', 'ID ', 'PRODUCT_ID', 'COMPANY_ID'])]
                    self.cleaning_logger.log_header_detection(
                        method='id_column',
                        header_row=row_idx,
                        headers=[str(h).strip() if h else "" for h in headers],
                        id_columns_found=id_columns_found
                    )

                return row_idx, headers

            # Count non-empty cells
            non_empty_count = sum(1 for cell in row if cell is not None and str(cell).strip())

            # Apply sparse row filtering ONLY for header detection
            # Skip rows that are too sparse to be headers (less than 80% filled)
            if max_cols > 0 and non_empty_count < max_cols * 0.8:
                # This row is sparse - check if it's significantly incomplete
                if non_empty_count < max(2, max_cols * 0.3):
                    self.logger.debug(f"Skipping sparse Excel row {row_idx} as potential header: {non_empty_count}/{max_cols} cells filled")

                    # Log metadata row removal if we have a cleaning logger
                    if self.cleaning_logger:
                        row_content = [str(cell) if cell else "" for cell in row]
                        self.cleaning_logger.log_row_removal(
                            row_indices=[row_idx],
                            rows_content=[row_content],
                            removal_reason=f"Sparse metadata row ({non_empty_count}/{max_cols} cells filled)",
                            row_type='metadata'
                        )

                    continue

            # Consider this a header row if it has at least 2 non-empty cells
            # and the next few rows seem to contain data
            if non_empty_count >= 2:
                # Check for metadata patterns that disqualify this as headers
                has_metadata_pattern = False
                for cell in row:
                    if cell is not None and str(cell).strip():
                        cell_lower = str(cell).strip().lower()
                        if any(pattern in cell_lower for pattern in
                              ['department:', 'quarter:', 'generated', 'report',
                               'confidential', 'version', 'internal use']):
                            has_metadata_pattern = True
                            self.logger.debug(f"Row {row_idx} has metadata pattern: {str(cell).strip()}")
                            break

                # Skip rows with metadata patterns
                if has_metadata_pattern:
                    continue

                # Check if this looks like a data table start
                if self._looks_like_header_row(row, all_rows, row_idx):
                    # Trim trailing empty columns from headers
                    headers = self._trim_trailing_empty(row)
                    return row_idx, headers

        # Fallback: use first non-empty row as headers
        for row_idx, row in enumerate(all_rows):
            if any(cell is not None and str(cell).strip() for cell in row):
                headers = self._trim_trailing_empty(row)
                return row_idx, headers

        # Ultimate fallback
        return 0, all_rows[0] if all_rows else []
    
    def _looks_like_header_row(self, current_row, all_rows, row_idx):
        """Determine if this row looks like a header by examining following rows."""
        # If this is the last row, it's probably not a header
        if row_idx >= len(all_rows) - 1:
            return False
        
        # Check if there are data rows after this potential header
        data_rows_found = 0
        for check_idx in range(row_idx + 1, min(row_idx + 4, len(all_rows))):
            check_row = all_rows[check_idx]
            if any(cell is not None and str(cell).strip() for cell in check_row):
                data_rows_found += 1
        
        # Consider it a header if there's at least one data row following
        return data_rows_found > 0
    
    def _trim_trailing_empty(self, row):
        """Remove trailing empty cells from a row."""
        # Find the last non-empty cell
        last_content_idx = -1
        for i, cell in enumerate(row):
            if cell is not None and str(cell).strip():
                last_content_idx = i
        
        # Return row up to the last content
        if last_content_idx >= 0:
            return row[:last_content_idx + 1]
        else:
            return row[:1] if row else []  # Keep at least one column
    
    def _clean_data_rows(self, data_rows, header_count):
        """Clean data rows by removing empty trailing rows and trimming to header count."""
        cleaned_rows = []
        
        for row in data_rows:
            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            
            # Trim row to match header count (or pad if shorter)
            if len(row) < header_count:
                # Pad with empty strings
                cleaned_row = list(row) + [''] * (header_count - len(row))
            else:
                # Trim to header count
                cleaned_row = row[:header_count]
            
            cleaned_rows.append(cleaned_row)
        
        return cleaned_rows
    
    def _find_table_start_csv(self, rows):
        """Find table start for CSV files (similar logic but handles string data)."""
        # First, find the maximum column count in the data
        max_cols = 0
        for row in rows[:20]:  # Check first 20 rows for max columns
            if row and not str(row[0]).strip().startswith('#'):
                max_cols = max(max_cols, len(row))

        for row_idx, row in enumerate(rows):
            # Skip completely empty rows
            if not any(str(cell).strip() for cell in row):
                continue

            # Skip rows that look like comments (first cell starts with #)
            if row and str(row[0]).strip().startswith('#'):
                self.logger.debug(f"Skipping comment row {row_idx}: {str(row[0])[:50]}...")
                continue

            # Check if this row contains ID columns (strong indicator of headers)
            has_id_column = False
            for cell in row:
                if cell:
                    cell_str = str(cell).strip().upper()
                    # Look for ID patterns in column names
                    if any(pattern in cell_str for pattern in ['_ID', ' ID', 'ID ',
                                                                'PRODUCT_ID', 'COMPANY_ID', 'CUSTOMER_ID',
                                                                'ITEM_ID', 'ORDER_ID', 'USER_ID']):
                        has_id_column = True
                        break

            if has_id_column:
                # This is definitely a header row - ID columns must be in headers
                self.logger.debug(f"Found ID column in CSV row {row_idx}, treating as header row")
                headers = self._trim_trailing_empty_csv(row)
                return row_idx, headers

            # Count non-empty cells
            non_empty_count = sum(1 for cell in row if str(cell).strip())

            # Apply sparse row filtering ONLY for header detection
            # Skip rows that are too sparse to be headers (less than 80% filled)
            if max_cols > 0 and non_empty_count < max_cols * 0.8:
                # This row is sparse - check if it's significantly incomplete
                if non_empty_count < max(2, max_cols * 0.3):
                    self.logger.debug(f"Skipping sparse row {row_idx} as potential header: {non_empty_count}/{max_cols} cells filled")
                    continue

            # Consider this a header row if it has at least 2 non-empty cells
            if non_empty_count >= 2:
                # Check if this looks like a data table start
                if self._looks_like_header_row_csv(row, rows, row_idx):
                    # Trim trailing empty columns from headers
                    headers = self._trim_trailing_empty_csv(row)
                    return row_idx, headers

        # Fallback: use first non-empty, non-comment row as headers
        for row_idx, row in enumerate(rows):
            # Skip comment rows in fallback too
            if row and str(row[0]).strip().startswith('#'):
                continue
            if any(str(cell).strip() for cell in row):
                headers = self._trim_trailing_empty_csv(row)
                return row_idx, headers

        # Ultimate fallback
        return 0, rows[0] if rows else []
    
    def _looks_like_header_row_csv(self, current_row, rows, row_idx):
        """CSV version of header detection."""
        if row_idx >= len(rows) - 1:
            return False
        
        # Check if there are data rows after this potential header
        data_rows_found = 0
        for check_idx in range(row_idx + 1, min(row_idx + 4, len(rows))):
            check_row = rows[check_idx]
            if any(str(cell).strip() for cell in check_row):
                data_rows_found += 1
        
        return data_rows_found > 0
    
    def _trim_trailing_empty_csv(self, row):
        """CSV version of trailing empty cell removal."""
        last_content_idx = -1
        for i, cell in enumerate(row):
            if str(cell).strip():
                last_content_idx = i
        
        if last_content_idx >= 0:
            return row[:last_content_idx + 1]
        else:
            return row[:1] if row else []
    
    def _clean_data_rows_csv(self, data_rows, header_count):
        """CSV version of data row cleaning."""
        cleaned_rows = []
        
        for row in data_rows:
            # Skip completely empty rows
            if not any(str(cell).strip() for cell in row):
                continue
            
            # Trim or pad to match header count
            if len(row) < header_count:
                cleaned_row = list(row) + [''] * (header_count - len(row))
            else:
                cleaned_row = row[:header_count]
            
            cleaned_rows.append(cleaned_row)
        
        return cleaned_rows
    
    def _is_valid_csv_parse(self, rows):
        """Check if CSV parsing produced reasonable results."""
        if not rows or len(rows) < 2:
            return False
        
        # Check if first few rows have similar column counts
        header_count = len(rows[0])
        if header_count < 2:  # Need at least 2 columns to be useful
            return False
        
        # Check first few data rows
        consistent_rows = 0
        for i in range(1, min(4, len(rows))):
            row_count = len(rows[i])
            # Allow some flexibility (within 1-2 columns)
            if abs(row_count - header_count) <= 2:
                consistent_rows += 1
        
        # If most rows have similar column counts, parsing is probably good
        return consistent_rows >= min(2, len(rows) - 1)
    
    def _parse_csv_with_robust_fallback(self, text_content):
        """Robust CSV parsing with delimiter detection for non-standard files."""
        lines = text_content.splitlines()
        if not lines:
            return []
        
        # Try to detect delimiter
        delimiter = self._detect_delimiter(text_content)
        self.logger.info(f"Detected delimiter: '{delimiter}'")
        
        # Parse with detected delimiter
        try:
            if delimiter:
                reader = csv.reader(lines, delimiter=delimiter)
                rows = list(reader)
                
                # Validate this parsing
                if self._is_valid_csv_parse(rows):
                    return rows
        except csv.Error as e:
            self.logger.warning(f"Failed with detected delimiter '{delimiter}': {e}")
        
        # Fallback to manual splitting if CSV reader fails
        return self._manual_csv_parse(lines)
    
    def _detect_delimiter(self, text_content):
        """Detect the most likely delimiter in the CSV file."""
        # Common delimiters to try (in order of preference)
        delimiters = [',', ';', '\t', '|', ':', ' ']
        
        # Sample first few lines for analysis
        sample_lines = text_content.splitlines()[:10]
        sample_text = '\n'.join(sample_lines)
        
        best_delimiter = None
        best_score = 0
        
        for delimiter in delimiters:
            try:
                # Use csv.Sniffer to help detect
                sniffer = csv.Sniffer()
                if sniffer.sniff(sample_text, delimiters=delimiter):
                    # Count occurrences and consistency
                    score = self._score_delimiter(sample_lines, delimiter)
                    if score > best_score:
                        best_score = score
                        best_delimiter = delimiter
            except csv.Error:
                # If sniffer fails, try manual scoring
                score = self._score_delimiter(sample_lines, delimiter)
                if score > best_score:
                    best_score = score
                    best_delimiter = delimiter
        
        return best_delimiter or ','  # Default to comma
    
    def _score_delimiter(self, lines, delimiter):
        """Score how well a delimiter works for parsing."""
        if not lines:
            return 0
        
        # Count delimiter occurrences per line
        counts = []
        for line in lines[:5]:  # Check first 5 lines
            if line.strip():  # Skip empty lines
                count = line.count(delimiter)
                if count > 0:  # Only count lines that have the delimiter
                    counts.append(count)
        
        if len(counts) < 2:
            return 0
        
        # Score based on consistency of delimiter count
        avg_count = sum(counts) / len(counts)
        variance = sum((count - avg_count) ** 2 for count in counts) / len(counts)
        
        # Good delimiters should have:
        # 1. At least 1 occurrence per line on average
        # 2. Low variance (consistent across lines)
        if avg_count >= 1:
            consistency_score = max(0, 10 - variance)  # Lower variance = higher score
            frequency_score = min(avg_count, 10)  # Cap at 10 to prevent over-weighting
            return consistency_score + frequency_score
        
        return 0
    
    def _manual_csv_parse(self, lines):
        """Manual CSV parsing for really problematic files."""
        rows = []
        
        # Try common delimiters in order
        test_delimiters = [',', ';', '\t', '|', ':', ' ']
        
        for delimiter in test_delimiters:
            test_rows = []
            for line in lines[:5]:  # Test first 5 lines
                if line.strip():
                    parts = [part.strip().strip('"\'') for part in line.split(delimiter)]
                    if len(parts) > 1:  # Must split into multiple parts
                        test_rows.append(parts)
            
            # If we got reasonable results, use this delimiter for all lines
            if len(test_rows) >= 2 and self._check_manual_parse_consistency(test_rows):
                self.logger.info(f"Using manual parsing with delimiter: '{delimiter}'")
                # Parse all lines with this delimiter
                for line in lines:
                    if line.strip():
                        parts = [part.strip().strip('"\'') for part in line.split(delimiter)]
                        rows.append(parts)
                return rows
        
        # Ultimate fallback: split by whitespace
        self.logger.warning("Using whitespace splitting as ultimate fallback")
        for line in lines:
            if line.strip():
                # Split by any whitespace and filter empty parts
                parts = [part for part in line.split() if part]
                if parts:
                    rows.append(parts)
        
        return rows
    
    def _check_manual_parse_consistency(self, rows):
        """Check if manually parsed rows have consistent structure."""
        if len(rows) < 2:
            return False
        
        # Check if column counts are reasonably consistent
        col_counts = [len(row) for row in rows]
        avg_cols = sum(col_counts) / len(col_counts)
        
        # Allow some variance but not too much
        consistent_count = sum(1 for count in col_counts if abs(count - avg_cols) <= 1)
        return consistent_count >= len(rows) * 0.7  # 70% of rows should be consistent

    def _extract_rows_with_formulas(self, worksheet):
        """Extract both values and cell objects from worksheet for formula analysis"""
        all_rows = []
        all_cell_objects = []

        for row in worksheet.iter_rows():
            value_row = []
            cell_row = []

            for cell in row:
                value_row.append(cell.value)
                cell_row.append(cell)

            all_rows.append(tuple(value_row))
            all_cell_objects.append(cell_row)

        return all_rows, all_cell_objects

    def _extract_formula_info(self, cell_obj, column_headers, worksheet):
        """Extract detailed formula information with column references"""
        if not cell_obj:
            return None

        # Check if cell has a formula - openpyxl stores formulas in the 'value' attribute
        # when data_only=False, and the data_type will be 'f' for formulas
        formula = None

        # Try different ways to get the formula
        if hasattr(cell_obj, 'data_type') and cell_obj.data_type == 'f':
            formula = cell_obj.value
        elif hasattr(cell_obj, 'value') and str(cell_obj.value).startswith('='):
            formula = cell_obj.value
        elif hasattr(cell_obj, 'formula') and cell_obj.formula:
            formula = cell_obj.formula

        if not formula or not str(formula).startswith('='):
            return None

        try:
            formula_str = str(formula)

            # Extract cell references and convert to column names
            referenced_columns = self._extract_column_references(formula_str, column_headers, worksheet)

            # Determine formula type
            formula_type = self._classify_formula(formula_str)

            # Get the calculated value - if data_only=False, we need to evaluate or get cached result
            calculated_value = "Not calculated"
            if hasattr(cell_obj, 'displayed_value') and cell_obj.displayed_value is not None:
                calculated_value = str(cell_obj.displayed_value)
            elif hasattr(cell_obj, '_value') and cell_obj._value is not None:
                calculated_value = str(cell_obj._value)
            else:
                # For formula cells, the value might be stored separately
                # We'll show the formula instead
                calculated_value = f"Formula result: {formula_str}"

            return {
                'formula': formula_str,
                'formula_type': formula_type,
                'referenced_columns': referenced_columns,
                'calculated_value': calculated_value,
                'description': self._generate_formula_description(formula_str, referenced_columns, formula_type)
            }

        except Exception as e:
            self.logger.warning(f"Error extracting formula info: {e}")
            return {
                'formula': str(formula),
                'formula_type': 'unknown',
                'referenced_columns': [],
                'calculated_value': str(cell_obj.value),
                'description': f"Formula: {formula}"
            }

    def _extract_column_references(self, formula_str, column_headers, worksheet):
        """Extract column references from formula and map to column names"""
        import re

        # Find all cell references (like A1, B2, C3:D10, etc.)
        cell_ref_pattern = r'([A-Z]+)(\d+)'
        matches = re.findall(cell_ref_pattern, formula_str)

        referenced_columns = []
        seen_columns = set()

        for col_letter, row_num in matches:
            # Convert column letter to index
            col_idx = self._column_letter_to_index(col_letter)

            # Map to column name if within our headers
            if 0 <= col_idx < len(column_headers):
                column_name = column_headers[col_idx]
                if column_name not in seen_columns:
                    referenced_columns.append({
                        'column_name': column_name,
                        'column_index': col_idx,
                        'excel_column': col_letter,
                        'example_cell': f"{col_letter}{row_num}"
                    })
                    seen_columns.add(column_name)
            else:
                # Column is outside our data range - this can happen when:
                # 1. Formula references columns beyond the data
                # 2. We trimmed trailing empty columns but formulas still reference them
                # 3. Formula references external sheets/files
                referenced_columns.append({
                    'column_name': f"Column_{col_letter}",
                    'column_index': col_idx,
                    'excel_column': col_letter,
                    'example_cell': f"{col_letter}{row_num}",
                    'note': "Outside table range (may be trimmed column)"
                })

        return referenced_columns

    def _column_letter_to_index(self, col_letter):
        """Convert Excel column letter(s) to 0-based index"""
        result = 0
        for char in col_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    def _classify_formula(self, formula_str):
        """Classify the type of Excel formula"""
        formula_upper = formula_str.upper()

        # Common function patterns
        if 'SUM(' in formula_upper:
            return 'sum'
        elif 'AVERAGE(' in formula_upper or 'AVG(' in formula_upper:
            return 'average'
        elif 'COUNT(' in formula_upper:
            return 'count'
        elif 'IF(' in formula_upper:
            return 'conditional'
        elif 'VLOOKUP(' in formula_upper or 'HLOOKUP(' in formula_upper or 'INDEX(' in formula_upper:
            return 'lookup'
        elif 'CONCATENATE(' in formula_upper or '&' in formula_str:
            return 'text'
        elif any(func in formula_upper for func in ['MIN(', 'MAX(', 'MEDIAN(']):
            return 'statistical'
        elif any(func in formula_upper for func in ['DATE(', 'TODAY(', 'NOW(']):
            return 'date'
        elif any(func in formula_upper for func in ['MOD(', 'ROUND(', 'ABS(', 'SQRT(', 'POWER(', 'FLOOR(', 'CEILING(']):
            return 'arithmetic'
        elif any(op in formula_str for op in ['+', '-', '*', '/']):
            return 'arithmetic'
        else:
            return 'other'

    def _generate_formula_description(self, formula_str, referenced_columns, formula_type):
        """Generate a human-readable description of the formula"""
        if not referenced_columns:
            return f"Formula: {formula_str}"

        # Replace cell references with column names
        formula_with_column_names = formula_str
        for ref in referenced_columns:
            # Replace Excel column references (like B2, B3, etc.) with column name
            import re
            pattern = rf"{ref['excel_column']}\d+"
            formula_with_column_names = re.sub(pattern, ref['column_name'], formula_with_column_names)

        return f"Formula: {formula_with_column_names}"

        # Add specific column details if there are only a few referenced columns
        if len(referenced_columns) <= 3:
            details = []
            for ref in referenced_columns:
                if 'note' in ref:
                    details.append(f"{ref['column_name']} ({ref['note']})")
                else:
                    details.append(f"{ref['column_name']} (Excel column {ref['excel_column']})")

            if details:
                base_description += f" - Details: {'; '.join(details)}"

        return base_description

    def extract_validation_history(self, bucket: str, key: str, parsed_data: Dict[str, Any] = None) -> Dict:
        """
        Extract validation history from Updated Values sheet + Validation Record.

        Args:
            bucket: S3 bucket name
            key: S3 object key for previously validated Excel file
            parsed_data: Already-parsed data with row keys (optional, will parse if not provided)

        Returns:
            {
                'validation_history': {
                    row_key: {
                        column: {
                            'prior_value': str,
                            'prior_confidence': str,
                            'prior_timestamp': str,
                            'original_value': str,
                            'original_confidence': str,
                            'original_key_citation': str,
                            'original_sources': list[str],
                            'original_timestamp': str
                        }
                    }
                },
                'file_timestamp': str
            }
        """
        try:
            self.logger.info(f"Extracting validation history from s3://{bucket}/{key}")

            # Parse the file if not already parsed (to get row keys with proper deduplication)
            if not parsed_data:
                self.logger.info("No parsed_data provided, parsing file to get row keys")
                # Note: We don't have id_fields here, so this will use full row hash
                # This is a fallback - callers should provide parsed_data
                parsed_data = self.parse_s3_table(bucket, key, sheet_name="Updated Values")

            # Extract row keys from parsed data
            parsed_rows = parsed_data.get('data', [])
            if not parsed_rows:
                self.logger.warning("No data rows in parsed data")
                return {
                    'validation_history': {},
                    'file_timestamp': ''
                }

            self.logger.info(f"Using row keys from {len(parsed_rows)} parsed rows")

            # Load timestamps from Validation Record sheet
            timestamps = self._load_validation_timestamps(bucket, key)

            # Create temporary file to load the Excel workbook (with retry logic)
            file_content = self._download_s3_object(bucket, key)

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_file_path = temp_file.name

            try:
                # Load workbook without read_only to access comments
                if not OPENPYXL_AVAILABLE:
                    raise ImportError("openpyxl is required for Excel history extraction")

                workbook = load_workbook(temp_file_path, data_only=True)

                # Check if Updated Values sheet exists
                if 'Updated Values' not in workbook.sheetnames:
                    self.logger.warning(f"No 'Updated Values' sheet found in {key}")
                    return {
                        'validation_history': {},
                        'file_timestamp': timestamps.get('file_timestamp', '')
                    }

                worksheet = workbook['Updated Values']

                # Read header row
                rows = list(worksheet.iter_rows(values_only=False))
                if not rows:
                    self.logger.warning(f"Empty Updated Values sheet in {key}")
                    return {
                        'validation_history': {},
                        'file_timestamp': timestamps.get('file_timestamp', '')
                    }

                # Find header row
                header_row_idx = 0
                headers = []
                for idx, row in enumerate(rows):
                    row_values = [cell.value for cell in row]
                    if any(val and str(val).strip() for val in row_values):
                        headers = [self._normalize_column_name(str(cell.value).strip()) if cell.value else f"Column_{i+1}"
                                   for i, cell in enumerate(row)]
                        header_row_idx = idx
                        break

                # Extract validation history from data rows using parsed row keys
                validation_history = {}

                # Match parsed rows to Excel rows by index
                excel_data_start_idx = header_row_idx + 1
                for parsed_idx, parsed_row in enumerate(parsed_rows):
                    excel_row_idx = excel_data_start_idx + parsed_idx

                    # Check if we have a corresponding Excel row
                    if excel_row_idx >= len(rows):
                        self.logger.warning(f"Parsed row {parsed_idx} has no corresponding Excel row")
                        break

                    excel_row = rows[excel_row_idx]

                    # Get row key from parsed data (already has deduplication applied)
                    row_key = parsed_row.get('_row_key')
                    if not row_key:
                        self.logger.warning(f"Parsed row {parsed_idx} has no _row_key, skipping")
                        continue

                    if row_key not in validation_history:
                        validation_history[row_key] = {}

                    # Process each cell in the Excel row to extract comments
                    for col_idx, cell in enumerate(excel_row):
                        if col_idx >= len(headers):
                            break

                        column_name = headers[col_idx]
                        cell_value = cell.value

                        # Skip empty cells
                        if not cell_value or not str(cell_value).strip():
                            continue

                        # Extract comment if it exists
                        comment_text = ""
                        if cell.comment:
                            comment_text = cell.comment.text if hasattr(cell.comment, 'text') else str(cell.comment)

                        # Parse comment to extract historical data
                        parsed_comment = {}
                        if comment_text:
                            parsed_comment = self._parse_validation_comment(comment_text)

                        # Build field history
                        # IMPORTANT: Updated Values sheet cell = Original/Current (INPUT)
                        #            Updated Values sheet comment "Original Value:" = Prior (from Original Values sheet)
                        field_history = {
                            'prior_value': parsed_comment.get('original_value', ''),  # From comment's "Original Value:" field
                            'prior_confidence': parsed_comment.get('original_confidence', ''),
                            'prior_timestamp': timestamps.get('prior_timestamp', ''),
                            'original_value': str(cell_value).strip(),  # Updated Values sheet cell = INPUT
                            'original_confidence': parsed_comment.get('original_confidence', ''),
                            'original_key_citation': parsed_comment.get('key_citation', ''),
                            'original_sources': parsed_comment.get('sources', []),  # URLs for backward compatibility
                            'original_sources_full': parsed_comment.get('sources_full', []),  # Full citation text
                            'original_timestamp': timestamps.get('original_timestamp', '')
                        }

                        validation_history[row_key][column_name] = field_history

                self.logger.info(f"Extracted history using parsed row keys, found {len(validation_history)} rows with history")

                workbook.close()

                self.logger.info(f"Extracted history for {len(validation_history)} rows")

                return {
                    'validation_history': validation_history,
                    'file_timestamp': timestamps.get('file_timestamp', '')
                }

            finally:
                # Clean up temp file
                os.unlink(temp_file_path)

        except Exception as e:
            self.logger.error(f"Failed to extract validation history: {str(e)}")
            # Return empty history instead of raising - treat as new file
            return {
                'validation_history': {},
                'file_timestamp': ''
            }

    def _parse_validation_comment(self, comment_text: str) -> Dict:
        """
        Parse structured validation comment from Updated Values sheet.

        Example input:
            Original Value: ABC Corp (MEDIUM Confidence)

            Key Citation: Company website (https://...)

            Sources:
            [1] Title (URL): "snippet"
            [2] Title (URL): "snippet"

        Returns:
            {
                'original_value': 'ABC Corp',
                'original_confidence': 'MEDIUM',
                'key_citation': 'Company website (https://...)',
                'sources': ['https://...', 'https://...']
            }
        """
        result = {}

        if not comment_text:
            return result

        lines = comment_text.split('\n')

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            # Parse Original Value with confidence
            if line.startswith('Original Value:'):
                # Extract: "Original Value: ABC Corp (MEDIUM Confidence)"
                content = line.replace('Original Value:', '').strip()
                if '(' in content and content.endswith('Confidence)'):
                    # Split by last occurrence of '('
                    last_paren = content.rfind('(')
                    value = content[:last_paren].strip()
                    conf = content[last_paren+1:].replace('Confidence)', '').strip()
                    result['original_value'] = value
                    result['original_confidence'] = conf
                else:
                    result['original_value'] = content

            # Parse Key Citation
            elif line.startswith('Key Citation:'):
                result['key_citation'] = line.replace('Key Citation:', '').strip()

            # Parse Sources section
            elif line.startswith('Sources:'):
                sources = []
                sources_full = []  # Keep full citation text
                i += 1
                while i < len(lines):
                    source_line = lines[i].strip()
                    if not source_line:
                        break
                    # Keep the full source line for complete citation
                    sources_full.append(source_line)
                    # Also extract just the URL for backward compatibility
                    if '(' in source_line and ')' in source_line:
                        url_start = source_line.find('(')
                        url_end = source_line.find(')', url_start)
                        url = source_line[url_start+1:url_end]
                        sources.append(url)
                    i += 1
                result['sources'] = sources
                result['sources_full'] = sources_full  # Full citation text
                continue

            i += 1

        return result

    def _load_validation_timestamps(self, bucket: str, key: str) -> Dict:
        """
        Read Validation Record sheet to get timestamps.

        Returns:
            {
                'original_timestamp': str,  # Run_Number=1, Run_Time field
                'prior_timestamp': str,     # Last run, Run_Time field
                'file_timestamp': str       # Only for internal tracking, NOT used in validation
            }
        """
        try:
            # Download file to read Validation Record sheet (with retry logic)
            file_content = self._download_s3_object(bucket, key)

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_file_path = temp_file.name

            try:
                if not OPENPYXL_AVAILABLE:
                    raise ImportError("openpyxl is required for timestamp extraction")

                workbook = load_workbook(temp_file_path, read_only=True, data_only=True)

                # Check if Validation Record sheet exists
                if 'Validation Record' not in workbook.sheetnames:
                    self.logger.info(f"No 'Validation Record' sheet found, returning empty timestamps to avoid cache break")
                    workbook.close()
                    return {
                        'original_timestamp': '',
                        'prior_timestamp': '',
                        'file_timestamp': ''
                    }

                worksheet = workbook['Validation Record']
                rows = list(worksheet.iter_rows(values_only=True))

                if not rows or len(rows) < 2:
                    self.logger.warning(f"Empty or invalid Validation Record sheet, returning empty timestamps")
                    workbook.close()
                    return {
                        'original_timestamp': '',
                        'prior_timestamp': '',
                        'file_timestamp': ''
                    }

                # Parse header row to find column indices
                header_row = rows[0]
                run_number_idx = None
                run_time_idx = None

                for idx, header in enumerate(header_row):
                    if header and 'Run_Number' in str(header):
                        run_number_idx = idx
                    elif header and 'Run_Time' in str(header):
                        run_time_idx = idx

                if run_number_idx is None or run_time_idx is None:
                    self.logger.warning(f"Could not find Run_Number or Run_Time columns in Validation Record")
                    workbook.close()
                    return {
                        'original_timestamp': '',
                        'prior_timestamp': '',
                        'file_timestamp': ''
                    }

                # Extract timestamps
                # The most recent run (highest run number) is the "original" (current input)
                # An earlier run (if exists) is the "prior"
                original_timestamp = ''
                prior_timestamp = ''

                # Find the highest run number (most recent = original)
                max_run_number = 0
                for row in rows[1:]:
                    if len(row) > run_number_idx and row[run_number_idx]:
                        try:
                            run_num = int(row[run_number_idx])
                            if run_num > max_run_number:
                                max_run_number = run_num
                                if len(row) > run_time_idx and row[run_time_idx]:
                                    original_timestamp = str(row[run_time_idx])
                        except (ValueError, TypeError):
                            continue

                # Find a prior timestamp (any run before the most recent)
                # Look for the second-highest run number if multiple runs exist
                if max_run_number > 1:
                    # There are earlier runs, find the second most recent
                    second_highest = 0
                    for row in rows[1:]:
                        if len(row) > run_number_idx and row[run_number_idx]:
                            try:
                                run_num = int(row[run_number_idx])
                                # Find the highest run that's less than max
                                if run_num < max_run_number and run_num > second_highest:
                                    second_highest = run_num
                                    if len(row) > run_time_idx and row[run_time_idx]:
                                        prior_timestamp = str(row[run_time_idx])
                            except (ValueError, TypeError):
                                continue

                workbook.close()

                return {
                    'original_timestamp': original_timestamp,
                    'prior_timestamp': prior_timestamp,
                    'file_timestamp': ''  # Not used - here for backward compatibility only
                }

            finally:
                # Clean up temp file
                os.unlink(temp_file_path)

        except Exception as e:
            self.logger.error(f"Failed to load validation timestamps: {str(e)}")
            # Return empty timestamps to avoid cache break
            # DO NOT use S3 LastModified - it changes on every access and breaks caching
            return {
                'original_timestamp': '',
                'prior_timestamp': '',
                'file_timestamp': ''
            }

# Global instance for easy imports
s3_table_parser = S3TableParser()