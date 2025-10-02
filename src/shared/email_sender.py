#!/usr/bin/env python3
"""
Email sender module for Perplexity Validator results
Sends enhanced validation results via Amazon SES
"""
import boto3
from botocore.exceptions import ClientError
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from io import BytesIO
import json
import os

def get_excel_features_text():
    """Common language about Excel file features used in all validation emails."""
    return ""  # Removed to avoid duplication in email body
from datetime import datetime
import re
import logging

# Using PyMuPDF for PDF generation (no PIL dependencies, works in Lambda)

logger = logging.getLogger(__name__)

# Email configuration
SENDER = "eliyahu@eliyahu.ai"
BCC_ADDRESS = "ppp@eliyahu.ai"  # For tracking/analytics
CHARSET = "UTF-8"


def generate_simple_text_receipt(session_id: str, email: str, amount: float, 
                                transaction_details: dict) -> bytes:
    """Generate a simple text-based receipt when PDF generation is not available"""
    from datetime import datetime
    
    receipt_date = datetime.now().strftime("%B %d, %Y at %I:%M %p UTC")
    
    # Extract transaction details with defaults
    rows_processed = transaction_details.get('rows_processed', 0)
    fields_validated = transaction_details.get('columns_validated_count', 0)
    perplexity_calls = transaction_details.get('perplexity_api_calls', 0)
    claude_calls = transaction_details.get('anthropic_api_calls', 0)
    qc_calls = transaction_details.get('qc_api_calls', 0)

    # NOTE: anthropic_api_calls already includes QC calls, don't add again
    total_claude_calls = claude_calls
    table_name = transaction_details.get('table_name', transaction_details.get('input_filename', 'N/A'))
    config_id = transaction_details.get('config_id', 'N/A')
    
    receipt_text = f"""
======================================================================
                    HYPERPLEXITY.AI TABLE RESEARCH
                           PAYMENT RECEIPT
======================================================================

Receipt Date:          {receipt_date}
Session ID:            {session_id}
Customer Email:        {email}
Service:               Table Validation

{f'Input Table:            {table_name}' if table_name and table_name != 'N/A' else ''}
{f'Configuration Code:     {config_id}' if config_id and config_id != 'N/A' else ''}

----------------------------------------------------------------------
                            SERVICE DETAILS
----------------------------------------------------------------------

Rows Processed:        {rows_processed:,}
Columns Validated:      {fields_validated:,}
Perplexity API Calls:  {perplexity_calls:,}
Claude API Calls:      {total_claude_calls:,}

----------------------------------------------------------------------
                               TOTAL
----------------------------------------------------------------------

Total Charged:         ${amount:.2f}

======================================================================

Thank you for using Hyperplexity!
For support, contact: eliyahu@eliyahu.ai
This receipt is for your records.

======================================================================
    """.strip()
    
    logger.info(f"Generated simple text receipt for session {session_id}: ${amount:.2f}")
    return receipt_text.encode('utf-8')


# Alias for backward compatibility with test files
def generate_receipt_pdf(session_id: str, email: str, amount: float, transaction_details: dict) -> bytes:
    """Alias for generate_receipt_pdf_html for backward compatibility with tests"""
    return generate_receipt_pdf_html(session_id, email, amount, transaction_details)

def generate_receipt_pdf_html(session_id: str, email: str, amount: float,
                              transaction_details: dict) -> bytes:
    """Generate PDF receipt using PyMuPDF - works without PIL dependencies"""
    try:
        import fitz  # PyMuPDF
        logger.info("PyMuPDF imported successfully for PDF generation")
    except ImportError as import_e:
        logger.error(f"PyMuPDF not available: {import_e}")
        return generate_simple_text_receipt(session_id, email, amount, transaction_details)

    try:
        # --- Page setup: US Letter, not A4 ---
        inch = 72
        page_width, page_height = 8.5 * inch, 11 * inch  # 612 × 792
        doc = fitz.open()
        page = doc.new_page(width=page_width, height=page_height)

        # --- Margins & typography ---
        left_margin   = 1.5 * inch
        right_margin  = page_width - 1.5 * inch
        top_margin    = 0.35 * inch  # Logo position: 0.25 + 0.1 = 0.35 inch from top
        bottom_margin = 1.00 * inch

        # Font names: use built-in Helvetica family
        FN_REG = "helv"
        FN_BOLD = "hebo"  # Helvetica Bold
        FS_H1 = 18        # Reduced from 20
        FS_H2 = 14        # Reduced from 16
        FS_LABEL = 10     # Reduced from 11
        FS_VALUE = 10     # Reduced from 11
        FS_SECTION = 12   # Reduced from 14
        FS_FOOTER = 9     # Reduced from 10

        # helpers
        def text_w(s: str, size: float, font: str = FN_REG) -> float:
            return fitz.get_text_length(s, fontname=font, fontsize=size)

        def center_line(y: float, text: str, size: float, font: str = FN_BOLD):
            # Draw centered text using insert_text with proper centering
            text_width = fitz.get_text_length(text, fontname=font, fontsize=size)
            x_centered = (page_width - text_width) / 2
            page.insert_text((x_centered, y), text, fontsize=size, fontname=font)

        def wrap_long_text(text: str, max_width: float, fontsize: float) -> list:
            """Break long text into multiple lines that fit within max_width"""
            if not text:
                return [text]
            
            # Check if text fits on one line
            text_width = fitz.get_text_length(text, fontname=FN_REG, fontsize=fontsize)
            if text_width <= max_width:
                return [text]
            
            # Split long text into multiple lines
            words = text.split()
            lines = []
            current_line = ""
            
            for word in words:
                test_line = current_line + (" " if current_line else "") + word
                test_width = fitz.get_text_length(test_line, fontname=FN_REG, fontsize=fontsize)
                
                if test_width <= max_width:
                    current_line = test_line
                else:
                    if current_line:
                        lines.append(current_line)
                        current_line = word
                    else:
                        # Single word is too long, break it
                        lines.append(word)
            
            if current_line:
                lines.append(current_line)
            
            return lines

        def draw_label_value(y: float, label: str, value: str) -> float:
            # Label at left margin (bold)
            page.insert_text((left_margin, y), label, fontsize=FS_LABEL, fontname=FN_BOLD)
            
            # Calculate available width for value (from middle to right margin)
            available_width = right_margin - (left_margin + 2.5 * inch)  # Leave space after label
            
            # Wrap long values into multiple lines
            value_lines = wrap_long_text(value, available_width, FS_VALUE)
            
            # Draw each line of the value
            line_y = y
            for line in value_lines:
                value_width = fitz.get_text_length(line, fontname=FN_REG, fontsize=FS_VALUE)
                x_right = right_margin - value_width
                page.insert_text((x_right, line_y), line, fontsize=FS_VALUE, fontname=FN_REG)
                line_y += 0.15 * inch  # Smaller line spacing for wrapped text
            
            # Return position after all lines (add extra space if multi-line)
            extra_space = 0.1 * inch if len(value_lines) > 1 else 0
            return line_y + 0.25 * inch + extra_space

        y = top_margin

        # --- Logo (optional) ---
        logo_paths = [
            "/var/task/hyperplexity-logo-2.png", "./hyperplexity-logo-2.png",
            "./frontend/hyperplexity-logo-2.png", "../deployment/package/hyperplexity-logo-2.png",
            "../../deployment/package/hyperplexity-logo-2.png", "/opt/hyperplexity-logo-2.png",
            "/var/task/EliyahuLogo_NoText_Crop.png", "./EliyahuLogo_NoText_Crop.png",
            "./src/lambdas/config/EliyahuLogo_NoText_Crop.png", "../config/EliyahuLogo_NoText_Crop.png",
            "../EliyahuLogo_NoText_Crop.png", "../../EliyahuLogo_NoText_Crop.png",
            "/tmp/hyperplexity-logo-2.png", "/tmp/EliyahuLogo_NoText_Crop.png"
        ]
        logo_drawn = False
        for p in logo_paths:
            if os.path.exists(p):
                try:
                    with open(p, "rb") as f:
                        data = f.read()
                    size = 1.5 * inch
                    x = (page_width - size) / 2
                    rect = fitz.Rect(x, y, x + size, y + size)
                    page.insert_image(rect, stream=data)
                    y = rect.br.y + 0.4 * inch
                    logo_drawn = True
                    break
                except Exception as e:
                    logger.warning(f"Logo load failed for {p}: {e}")
        if not logo_drawn:
            y += 0.5 * inch  # give some space if no logo

        # --- Header / Title (centered) - final positioning ---
        y += 0.1 * inch  # Move text up 0.1 inches: was +0.2, now +0.1
        center_line(y, "Hyperplexity.AI Table Research", FS_H1, FN_BOLD)
        y += 0.35 * inch
        center_line(y, "Payment Receipt", FS_H2, FN_BOLD)
        y += 0.80 * inch

        # --- Info rows ---
        receipt_date = datetime.now().strftime("%B %d, %Y at %I:%M %p UTC")
        y = draw_label_value(y, "Receipt Date:", receipt_date)
        y = draw_label_value(y, "Session ID:", session_id)
        y = draw_label_value(y, "Customer Email:", email)
        y = draw_label_value(y, "Service:", "Table Validation")

        table_name = transaction_details.get('table_name', transaction_details.get('input_filename', ''))
        if table_name:
            y = draw_label_value(y, "Input Table:", table_name)

        config_id = transaction_details.get('config_id', '')
        if config_id:
            y = draw_label_value(y, "Configuration Code:", config_id)

        # --- Service details ---
        y += 0.4 * inch
        page.insert_text((left_margin, y), "Service Details", fontsize=FS_SECTION, fontname=FN_BOLD)
        y += 0.30 * inch

        rp = transaction_details.get('rows_processed', 0)
        cv = transaction_details.get('columns_validated_count', 0)
        pa = transaction_details.get('perplexity_api_calls', 0)
        ca = transaction_details.get('anthropic_api_calls', 0)
        # NOTE: anthropic_api_calls already includes QC calls, don't add qc_api_calls again

        y = draw_label_value(y, "Rows Processed:", f"{rp:,}")
        y = draw_label_value(y, "Columns Validated:", f"{cv:,}")
        y = draw_label_value(y, "Perplexity API Calls:", f"{pa:,}")
        y = draw_label_value(y, "Claude API Calls:", f"{ca:,}")

        # --- Separator line ---
        y += 0.2 * inch
        page.draw_line(fitz.Point(left_margin, y), fitz.Point(right_margin, y), width=1)
        y += 0.30 * inch

        # --- Total ---
        page.insert_text((left_margin, y), "Total Charged:", fontsize=FS_SECTION, fontname=FN_BOLD)
        total_str = f"${amount:.2f}"
        total_width = fitz.get_text_length(total_str, fontname=FN_BOLD, fontsize=FS_SECTION)
        x_total_right = right_margin - total_width
        page.insert_text((x_total_right, y), total_str, fontsize=FS_SECTION, fontname=FN_BOLD)

        # --- Footer (centered, anchored by bottom margin) ---
        footer_lines = [
            "Thank you for using Hyperplexity!",
            "For support, contact: eliyahu@eliyahu.ai",
            "This receipt is for your records."
        ]
        fy = page_height - bottom_margin - (len(footer_lines) * 15)  # Start higher
        for line in footer_lines:
            line_width = fitz.get_text_length(line, fontname=FN_REG, fontsize=FS_FOOTER)
            x_centered = (page_width - line_width) / 2
            page.insert_text((x_centered, fy), line, fontsize=FS_FOOTER, fontname=FN_REG)
            fy += 15  # Move down for next line

        pdf_bytes = doc.tobytes()  # reliable across PyMuPDF versions
        doc.close()
        logger.info(f"PyMuPDF PDF generation successful! Size: {len(pdf_bytes):,} bytes")
        return pdf_bytes

    except Exception as e:
        logger.error(f"Error generating PyMuPDF PDF receipt: {e}")
        return generate_simple_text_receipt(session_id, email, amount, transaction_details)



def generate_receipt(session_id: str, email: str, amount: float, raw_cost: float, 
                    multiplier: float, transaction_details: dict) -> str:
    """Legacy HTML receipt function - kept for compatibility"""
    try:
        receipt_date = datetime.now().strftime("%B %d, %Y at %I:%M %p UTC")
        receipt_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
                .receipt {{ max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .header {{ text-align: center; border-bottom: 2px solid #4a90e2; padding-bottom: 20px; margin-bottom: 30px; }}
                .logo {{ font-size: 24px; font-weight: bold; color: #4a90e2; margin-bottom: 5px; }}
                .receipt-title {{ font-size: 18px; color: #666; }}
                .receipt-info {{ margin-bottom: 30px; }}
                .info-row {{ display: flex; justify-content: space-between; margin: 8px 0; }}
                .label {{ font-weight: bold; color: #333; }}
                .value {{ color: #666; }}
                .cost-breakdown {{ background: #f8f9fa; padding: 20px; border-radius: 5px; margin: 20px 0; }}
                .cost-row {{ display: flex; justify-content: space-between; margin: 5px 0; }}
                .total-row {{ border-top: 2px solid #ddd; padding-top: 10px; font-weight: bold; font-size: 16px; }}
                .footer {{ text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; color: #888; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="receipt">
                <div class="header">
                    <div class="logo">HYPERPLEXITY VALIDATION</div>
                    <div class="receipt-title">Payment Receipt</div>
                </div>
                
                <div class="receipt-info">
                    <div class="info-row">
                        <span class="label">Receipt Date:</span>
                        <span class="value">{receipt_date}</span>
                    </div>
                    <div class="info-row">
                        <span class="label">Session ID:</span>
                        <span class="value">{session_id}</span>
                    </div>
                    <div class="info-row">
                        <span class="label">Customer Email:</span>
                        <span class="value">{email}</span>
                    </div>
                    <div class="info-row">
                        <span class="label">Service:</span>
                        <span class="value">Table Validation</span>
                    </div>
                </div>
                
                <div class="cost-breakdown">
                    <h3 style="margin-top: 0; color: #333;">Cost Breakdown</h3>
                    <div class="cost-row">
                        <span>Raw Processing Cost:</span>
                        <span>${raw_cost:.2f}</span>
                    </div>
                    <div class="cost-row">
                        <span>Domain Multiplier:</span>
                        <span>{multiplier:.1f}x</span>
                    </div>
                    <div class="cost-row">
                        <span>Rows Processed:</span>
                        <span>{transaction_details.get('rows_processed', 'N/A')}</span>
                    </div>
                    <div class="cost-row total-row">
                        <span>Total Charged:</span>
                        <span>${amount:.2f}</span>
                    </div>
                </div>
                
                <div class="footer">
                    <p>Thank you for using Hyperplexity!</p>
                    <p>For support, contact: eliyahu@eliyahu.ai</p>
                    <p>This receipt is for your records.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return receipt_content.strip()
        
    except Exception as e:
        logger.error(f"Error generating receipt: {e}")
        # Return simple text receipt as fallback
        return f"""
HYPERPLEXITY VALIDATION RECEIPT
==============================
Session: {session_id}
Date: {datetime.now().isoformat()}
Email: {email}

Service: Table Validation
Raw Processing Cost: ${raw_cost:.2f}
Domain Multiplier: {multiplier}x
Total Charged: ${amount:.2f}

Thank you for using Hyperplexity!
        """.strip()


def send_credit_confirmation_email(email_address: str, amount_purchased: float, new_balance: float, transaction_id: str) -> dict:
    """Send confirmation email for credit purchase."""
    try:
        # Create message
        message = MIMEMultipart()
        message["From"] = SENDER
        message["To"] = email_address
        message["Subject"] = f"💳 Credits Added - ${amount_purchased:.2f} - Hyperplexity"
        
        # Create email body
        purchase_date = datetime.now().strftime("%B %d, %Y at %I:%M %p UTC")
        
        body_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
                .email-container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .header {{ background: linear-gradient(135deg, #4a90e2, #357abd); color: white; padding: 30px; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 24px; }}
                .content {{ padding: 30px; }}
                .purchase-info {{ background: #f8f9fa; padding: 20px; border-radius: 5px; margin: 20px 0; }}
                .info-row {{ display: flex; justify-content: space-between; margin: 10px 0; }}
                .label {{ font-weight: bold; color: #333; }}
                .value {{ color: #666; }}
                .balance-highlight {{ background: #e8f5e8; padding: 15px; border-radius: 5px; text-align: center; margin: 20px 0; }}
                .balance-amount {{ font-size: 24px; font-weight: bold; color: #27ae60; }}
                .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="email-container">
                <div class="header">
                    <h1>🎉 Credits Successfully Added!</h1>
                </div>
                
                <div class="content">
                    <p>Thank you for your purchase! Your credits have been successfully added to your Hyperplexity account.</p>
                    
                    <div class="purchase-info">
                        <h3 style="margin-top: 0; color: #333;">Purchase Details</h3>
                        <div class="info-row">
                            <span class="label">Purchase Date:</span>
                            <span class="value">{purchase_date}</span>
                        </div>
                        <div class="info-row">
                            <span class="label">Amount Purchased:</span>
                            <span class="value">${amount_purchased:.2f}</span>
                        </div>
                        <div class="info-row">
                            <span class="label">Transaction ID:</span>
                            <span class="value">{transaction_id}</span>
                        </div>
                    </div>
                    
                    <div class="balance-highlight">
                        <p style="margin: 0; color: #333;">Your Current Balance:</p>
                        <div class="balance-amount">${new_balance:.2f}</div>
                    </div>
                    
                    <p>Your credits are ready to use! You can now process table validations and the costs will be automatically deducted from your account balance.</p>
                    
                    <p><strong>What's Next?</strong></p>
                    <ul>
                        <li>Upload your Excel or CSV files for validation</li>
                        <li>Generate AI-powered configuration files</li>
                        <li>Monitor your usage and balance in real-time</li>
                    </ul>
                </div>
                
                <div class="footer">
                    <p>Thank you for using Hyperplexity!</p>
                    <p>For support, contact: eliyahu@eliyahu.ai</p>
                    <p>Keep this email for your records.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Attach HTML body
        part = MIMEText(body_html, 'html', CHARSET)
        message.attach(part)
        
        # Send email via SES
        ses_client = boto3.client('ses')
        
        response = ses_client.send_raw_email(
            Source=SENDER,
            Destinations=[email_address, BCC_ADDRESS],
            RawMessage={
                'Data': message.as_string()
            }
        )
        
        message_id = response['MessageId']
        logger.info(f"Credit confirmation email sent successfully! Message ID: {message_id}")
        
        return {
            'success': True,
            'message_id': message_id,
            'message': f"Credit confirmation sent to {email_address}"
        }
        
    except Exception as e:
        logger.error(f"Failed to send credit confirmation email: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'message': f"Failed to send confirmation to {email_address}"
        }


def send_validation_results_email(email_address, excel_content, config_content, enhanced_excel_content, input_filename, config_filename, enhanced_excel_filename, session_id, summary_data, processing_time=None, reference_pin=None, metadata=None, preview_email=False, billing_info=None, config_id=None):
    """
    Send validation results via email with individual file attachments
    
          Args:
          email_address: Recipient email
          excel_content: Bytes content of the original Excel file
          config_content: Bytes content of the config JSON file
          enhanced_excel_content: Bytes content of the enhanced Excel file
          input_filename: Original input filename
          config_filename: Config filename
          enhanced_excel_filename: Enhanced Excel filename with timestamp
          session_id: Unique session identifier
          summary_data: Dictionary with validation summary info
          processing_time: Optional processing time in seconds
          reference_pin: Reference PIN for the validation
          metadata: Optional metadata including token usage
          billing_info: Optional billing information for receipt generation
          config_id: Configuration ID to reference instead of exposing filename
        
    Returns:
        dict: Response with status and message ID
    """
    try:
        # Create message with reference pin
        message = MIMEMultipart()
        message["From"] = SENDER
        message["To"] = email_address
        
        # Clean filename by removing excel_ prefix and VerifiedX suffix
        def clean_filename(filename):
            if not filename:
                return "Table"
            # Remove file extension
            base_name = filename.rsplit('.', 1)[0] if '.' in filename else filename
            # Remove excel_ prefix
            if base_name.startswith('excel_'):
                base_name = base_name[6:]
            # Remove VerifiedX suffix (where X is a number)
            import re
            base_name = re.sub(r'_Verified\d*$', '', base_name)
            return base_name or "Table"
        
        # Create subject line - use original filename with extension
        original_filename = input_filename or "Table"
        subject_type = "Hyperplexity Results"
        
        if reference_pin:
            message["Subject"] = f"🟩 {subject_type} - {original_filename} (Ref# {reference_pin})"
        else:
            message["Subject"] = f"🟩 {subject_type} - {original_filename}"
        
        # Extract summary data
        total_rows = summary_data.get('total_rows', 0)
        fields_validated = summary_data.get('fields_validated', [])
        confidence_distribution = summary_data.get('confidence_distribution', {})
        original_confidence_distribution = summary_data.get('original_confidence_distribution', {})
        
        # Extract token usage from metadata
        token_usage = None
        if metadata and 'token_usage' in metadata:
            token_usage = metadata['token_usage']
        
        # Validate enhanced Excel file first (before building email body)
        enhanced_excel_valid = False
        if enhanced_excel_content:
            # Validate the Excel content before attaching
            try:
                import openpyxl
                from io import BytesIO

                wb = openpyxl.load_workbook(BytesIO(enhanced_excel_content), read_only=True, data_only=True)

                # Check for required sheets
                required_sheets = ['Updated Values', 'Original Values', 'Details']
                missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]

                if missing_sheets:
                    logger.error(f"[EMAIL] Enhanced Excel missing required sheets: {missing_sheets}")
                    enhanced_excel_valid = False
                else:
                    # Check that Details sheet has content (more than just headers)
                    details_sheet = wb['Details']
                    details_row_count = 0
                    for row in details_sheet.iter_rows():
                        details_row_count += 1
                        if details_row_count > 10:  # Just check first 10 rows for efficiency
                            break

                    if details_row_count <= 1:  # Only header or empty
                        logger.error(f"[EMAIL] Details sheet has {details_row_count} rows - appears to be empty!")
                        enhanced_excel_valid = False
                    else:
                        # Check Updated sheet also has content
                        updated_sheet = wb['Updated Values']
                        updated_row_count = 0
                        for row in updated_sheet.iter_rows():
                            updated_row_count += 1
                            if updated_row_count > 2:  # At least header + 1 data row
                                break

                        if updated_row_count <= 1:
                            logger.error(f"[EMAIL] Updated sheet has {updated_row_count} rows - appears to be empty!")
                            enhanced_excel_valid = False
                        else:
                            # Check Original sheet also has content
                            original_sheet = wb['Original Values']
                            original_row_count = 0
                            for row in original_sheet.iter_rows():
                                original_row_count += 1
                                if original_row_count > 2:  # At least header + 1 data row
                                    break

                            if original_row_count <= 1:
                                logger.error(f"[EMAIL] Original sheet has {original_row_count} rows - appears to be empty!")
                                enhanced_excel_valid = False
                            else:
                                logger.info(f"[EMAIL] Enhanced Excel validation passed - Updated: {updated_row_count}+ rows, Original: {original_row_count}+ rows, Details: {details_row_count}+ rows")
                                enhanced_excel_valid = True

                wb.close()

            except Exception as e:
                logger.error(f"[EMAIL] Failed to validate enhanced Excel: {e}")
                enhanced_excel_valid = False

        # Now create email body with the validation result
        body_html = create_validation_results_email_body(
            session_id,
            total_rows,
            fields_validated,
            confidence_distribution,
            processing_time,
            reference_pin,
            token_usage,
            enhanced_excel_filename,
            input_filename,
            config_filename,
            preview_email,
            config_id,
            original_confidence_distribution,
            billing_info,
            enhanced_excel_valid=enhanced_excel_valid
        )

        # Attach HTML body
        part = MIMEText(body_html, 'html', CHARSET)
        message.attach(part)

        # Attach validated Excel file if validation passed
        if enhanced_excel_valid:
            part = MIMEApplication(enhanced_excel_content)
            part.add_header("Content-Disposition", f'attachment; filename="{enhanced_excel_filename}"')
            message.attach(part)
        else:
            logger.error(f"[EMAIL] Enhanced Excel validation failed for session {session_id} - will not send email, charge user, or report success")
            logger.warning(f"[EMAIL] This indicates a validation processing failure - returning error to prevent billing")
            # Return failure immediately - don't send email or charge user
            return {
                'success': False,
                'error': 'validation_failed',
                'message': f"Enhanced Excel validation failed - Details sheet appears empty or invalid. This indicates a processing error. User will not be charged."
            }

        # Attach original input file
        part = MIMEApplication(excel_content)
        part.add_header("Content-Disposition", f'attachment; filename="{input_filename}"')
        message.attach(part)
        
        # Note: Config JSON file attachment removed - only provide configuration code for reuse
        
        # Generate and attach receipt if there are charges
        if billing_info and billing_info.get('amount_charged', 0) > 0 and not preview_email:
            try:
                # Enhanced transaction details with API call counts and table name
                transaction_details = {
                    'rows_processed': billing_info.get('rows_processed', summary_data.get('total_rows', 0)),
                    'description': f"Full validation - {summary_data.get('total_rows', 0)} rows processed",
                    'session_id': session_id,
                    'perplexity_api_calls': billing_info.get('perplexity_api_calls', 0),
                    'anthropic_api_calls': billing_info.get('anthropic_api_calls', 0),
                    'qc_api_calls': billing_info.get('qc_api_calls', 0),  # Add QC calls to receipt
                    'columns_validated_count': billing_info.get('columns_validated_count', 0),
                    'table_name': billing_info.get('table_name', 'N/A'),
                    'input_filename': billing_info.get('table_name', 'N/A'),
                    'config_id': billing_info.get('config_id', 'N/A')
                }
                
                # Generate receipt (PDF or text fallback)
                receipt_bytes = generate_receipt_pdf_html(
                    session_id=session_id,
                    email=email_address,
                    amount=billing_info.get('amount_charged', 0),
                    transaction_details=transaction_details
                )
                
                # Detect file type based on content
                if receipt_bytes.startswith(b'%PDF'):
                    # It's a PDF file
                    receipt_filename = f"receipt_{session_id}.pdf"
                    content_type = "application/pdf"
                else:
                    # It's a text file
                    receipt_filename = f"receipt_{session_id}.txt"
                    content_type = "text/plain; charset=utf-8"
                
                # Attach receipt file
                receipt_part = MIMEApplication(receipt_bytes)
                receipt_part.add_header("Content-Disposition", f'attachment; filename="{receipt_filename}"')
                receipt_part.add_header("Content-Type", content_type)
                message.attach(receipt_part)
                
                receipt_type = "PDF" if receipt_bytes.startswith(b'%PDF') else "text"
                logger.info(f"{receipt_type} receipt attached to email for session {session_id}: ${billing_info.get('amount_charged', 0):.2f}")
                
            except Exception as e:
                logger.error(f"Failed to generate/attach PDF receipt for session {session_id}: {e}")
                # Fallback to HTML receipt if PDF generation fails
                try:
                    logger.info(f"Attempting HTML receipt fallback for session {session_id}")
                    transaction_details_fallback = {
                        'rows_processed': summary_data.get('total_rows', 0),
                        'description': f"Full validation - {summary_data.get('total_rows', 0)} rows processed",
                        'session_id': session_id
                    }
                    
                    receipt_html = generate_receipt(
                        session_id=session_id,
                        email=email_address,
                        amount=billing_info.get('amount_charged', 0),
                        raw_cost=billing_info.get('eliyahu_cost', 0),
                        multiplier=billing_info.get('multiplier', 1.0),
                        transaction_details=transaction_details_fallback
                    )
                    
                    receipt_filename = f"receipt_{session_id}.html"
                    receipt_part = MIMEApplication(receipt_html.encode('utf-8'))
                    receipt_part.add_header("Content-Disposition", f'attachment; filename="{receipt_filename}"')
                    receipt_part.add_header("Content-Type", "text/html; charset=utf-8")
                    message.attach(receipt_part)
                    
                    logger.info(f"HTML receipt fallback attached for session {session_id}")
                    
                except Exception as fallback_e:
                    logger.error(f"Both PDF and HTML receipt generation failed for session {session_id}: {fallback_e}")
                    # Don't fail the email if receipt generation fails
        
        # Send email via SES
        ses_client = boto3.client('ses')
        
        # Send to recipient with BCC
        destinations = [email_address]
        if BCC_ADDRESS and BCC_ADDRESS != email_address:
            destinations.append(BCC_ADDRESS)
            
        response = ses_client.send_raw_email(
            Source=SENDER,
            Destinations=destinations,
            RawMessage={
                'Data': message.as_string()
            }
        )
        
        message_id = response['MessageId']
        logger.info(f"Email sent successfully! Message ID: {message_id}")
        
        # Log email metadata to S3 (optional tracking)
        try:
            attachments_list = [enhanced_excel_filename, input_filename, config_filename]
            
            # Add receipt to attachments list if billing info was provided and charges applied
            if billing_info and billing_info.get('amount_charged', 0) > 0 and not preview_email:
                receipt_filename = f"receipt_{session_id}.html"
                attachments_list.append(receipt_filename)
            
            email_metadata = {
                'MessageID': message_id,
                'SessionID': session_id,
                'Recipient': email_address,
                'Timestamp': datetime.utcnow().isoformat(),
                'Subject': message["Subject"],
                'TotalRows': total_rows,
                'FieldsValidated': fields_validated,
                'ConfidenceDistribution': confidence_distribution,
                'Attachments': attachments_list,
                'BillingInfo': billing_info if billing_info else None,
                'PreviewEmail': preview_email
            }
            
            s3_client = boto3.client('s3')
            s3_client.put_object(
                Bucket='perplexity-cache',  # Using existing bucket
                Key=f'email_logs/{session_id}/email_metadata.json',
                Body=json.dumps(email_metadata, indent=2),
                ContentType='application/json'
            )
            logger.info("Email metadata saved to S3")
            
        except Exception as e:
            logger.warning(f"Failed to save email metadata: {e}")
            # Don't fail the email send if metadata save fails
        
        return {
            'success': True,
            'message_id': message_id,
            'message': f"Results sent successfully to {email_address}"
        }
        
    except ClientError as e:
        error_message = e.response['Error']['Message']
        logger.error(f"SES error: {error_message}")
        
        # Check for common SES errors
        if 'MessageRejected' in str(e):
            if 'Email address is not verified' in error_message:
                return {
                    'success': False,
                    'error': 'Email address not verified',
                    'message': 'The sender or recipient email is not verified in SES. Please verify email addresses.'
                }
            elif 'Sending rate exceeded' in error_message:
                return {
                    'success': False,
                    'error': 'Rate limit exceeded',
                    'message': 'Email sending rate limit exceeded. Please try again later.'
                }
        
        return {
            'success': False,
            'error': error_message,
            'message': f"Failed to send email: {error_message}"
        }
        
    except Exception as e:
        logger.error(f"Unexpected error sending email: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'message': f"Unexpected error: {str(e)}"
        }


def create_validation_results_email_body(session_id, total_rows, fields_validated, confidence_distribution, processing_time=None, reference_pin=None, token_usage=None, enhanced_excel_filename=None, input_filename=None, config_filename=None, preview_email=False, config_id=None, original_confidence_distribution=None, billing_info=None, enhanced_excel_valid=True):
    """Create clean validation results email body following Eliyahu.AI style guide"""
    
    # Format fields list
    if fields_validated:
        if len(fields_validated) <= 5:
            fields_html = ", ".join(fields_validated)
        else:
            fields_html = ", ".join(fields_validated[:5]) + f" and {len(fields_validated) - 5} more"
    else:
        fields_html = "No columns processed"
    
    # Format confidence distribution with Original -> Updated format
    confidence_html = ""
    total_validations = sum(confidence_distribution.values())
    for level in ["HIGH", "MEDIUM", "LOW"]:
        count = confidence_distribution.get(level, 0)
        if count > 0 or total_validations > 0:  # Show even if 0 for completeness
            percentage = (count / total_validations * 100) if total_validations > 0 else 0
            emoji = "🟢" if level == "HIGH" else "🟡" if level == "MEDIUM" else "🔴"
            
            # Show Original -> Updated format when original data is available
            if original_confidence_distribution:
                original_count = original_confidence_distribution.get(level, 0)
                original_total = sum(original_confidence_distribution.values())
                original_percentage = (original_count / original_total * 100) if original_total > 0 else 0
                confidence_html += f"<li>{emoji} <b>{level}:</b> {original_percentage:.1f}% (Original) -> {percentage:.1f}% (Updated)</li>"
            else:
                # Show updated results when no original data available 
                confidence_html += f"<li>{emoji} <b>{level}:</b> {percentage:.1f}% (Updated)</li>"
    
    if not confidence_html:
        confidence_html = "<li>No confidence data available</li>"
    
    # Format processing time
    time_info = ""
    if processing_time:
        if processing_time < 60:
            time_info = f"<p><b>Processing time:</b> {processing_time:.1f} seconds</p>"
        else:
            minutes = processing_time / 60
            time_info = f"<p><b>Processing time:</b> {minutes:.1f} minutes</p>"
    
    # API calls info - prefer billing_info (from actual run data) over token_usage
    api_calls_info = ""

    # DEBUG logging for API call sources
    logger.info(f"[EMAIL_API_DEBUG] billing_info exists: {billing_info is not None}")
    if billing_info:
        logger.info(f"[EMAIL_API_DEBUG] billing_info keys: {list(billing_info.keys())}")
        logger.info(f"[EMAIL_API_DEBUG] perplexity_api_calls: {billing_info.get('perplexity_api_calls', 0)}")
        logger.info(f"[EMAIL_API_DEBUG] anthropic_api_calls: {billing_info.get('anthropic_api_calls', 0)}")
        logger.info(f"[EMAIL_API_DEBUG] qc_api_calls: {billing_info.get('qc_api_calls', 0)}")

    if billing_info and (billing_info.get('perplexity_api_calls', 0) > 0 or
                        billing_info.get('anthropic_api_calls', 0) > 0 or
                        billing_info.get('qc_api_calls', 0) > 0):
        # Get API call counts from billing_info (authoritative source from DynamoDB runs)
        perplexity_calls = billing_info.get('perplexity_api_calls', 0)
        claude_calls = billing_info.get('anthropic_api_calls', 0)
        # NOTE: anthropic_api_calls already includes QC calls, don't add qc_api_calls again

        call_parts = []
        if perplexity_calls > 0:
            call_parts.append(f"<p><b>Perplexity Calls:</b> {perplexity_calls:,}</p>")
        if claude_calls > 0:
            call_parts.append(f"<p><b>Claude Calls:</b> {claude_calls:,}</p>")

        if call_parts:
            api_calls_info = "\n".join(call_parts)
            logger.info(f"[EMAIL_API_DEBUG] Using billing_info for API calls: {api_calls_info}")
    elif token_usage:
        # Fallback to token_usage if billing_info doesn't have call counts
        by_provider = token_usage.get('by_provider', {})
        perplexity_calls = by_provider.get('perplexity', {}).get('calls', 0)
        claude_calls = by_provider.get('anthropic', {}).get('calls', 0)

        call_parts = []
        if perplexity_calls > 0:
            call_parts.append(f"<p><b>Perplexity Calls:</b> {perplexity_calls:,}</p>")
        if claude_calls > 0:
            call_parts.append(f"<p><b>Claude Calls:</b> {claude_calls:,}</p>")

        if call_parts:
            api_calls_info = "\n".join(call_parts)
    
    # Cost info - get from billing_info which comes from DynamoDB runs table
    cost_info = ""
    if billing_info:
        # Use amount_charged which is the actual customer charge, not estimated_total_cost (Eliyahu internal cost)
        cost = billing_info.get('amount_charged', 0)
        logger.info(f"[EMAIL_COST_DEBUG] amount_charged: {cost}, will show cost: {cost > 0}")
        if cost > 0:
            cost_info = f"<p><b>Cost:</b> ${cost:.2f}</p>"
            logger.info(f"[EMAIL_COST_DEBUG] Cost info HTML: {cost_info}")
    
    # Reference pin removed from display per requirements
    
    # Preview email notice
    preview_notice = ""
    if preview_email:
        preview_notice = f"""
        <div style="background: #FFF3CD; border: 1px solid #FFEAA7; padding: 15px; border-radius: 6px; margin: 15px 0;">
            <p><b>📋 Preview Results</b></p>
            <p>This email contains validation results for a preview of your data ({total_rows} rows). To process your full table, please use the Hyperplexity Tool with your original table and Configuration ID: {config_id or 'N/A'}.</p>
        </div>
        """
    
    # Create email body following Eliyahu.AI style guide
    body_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Validation Complete - Hyperplexity Table Validation</title>
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                line-height: 1.6;
                color: #000000;
                max-width: 600px;
                margin: 0 auto;
                padding: 20px;
                background-color: #ffffff;
            }}
            .header {{
                background: #ffffff;
                color: #333333;
                padding: 30px;
                text-align: center;
                border-radius: 8px 8px 0 0;
                border-bottom: 3px solid #4CAF50;
            }}
            .content {{
                background: #ffffff;
                padding: 30px;
                border: 1px solid #E5E5E5;
                border-radius: 0 0 8px 8px;
            }}
            .summary {{
                background: #ffffff;
                border: 1px solid #E5E5E5;
                border-left: 4px solid #4CAF50;
                padding: 20px;
                border-radius: 8px;
                margin: 20px 0;
            }}
            .confidence {{
                background: #F8F9FA;
                border: 1px solid #E5E5E5;
                padding: 15px;
                border-radius: 8px;
                margin: 15px 0;
            }}
            .confidence ul {{
                margin: 10px 0;
                padding-left: 20px;
            }}
            .confidence li {{
                margin: 5px 0;
            }}
            .attachments {{
                background: #ffffff;
                border: 2px solid #4CAF50;
                padding: 20px;
                border-radius: 8px;
                margin: 25px 0;
                text-align: center;
            }}
            .attachments h3 {{
                color: #000000;
                margin-top: 0;
            }}
            .attachment-list {{
                list-style: none;
                padding: 0;
                margin: 15px 0;
            }}
            .attachment-list li {{
                background: #F8F9FA;
                border: 1px solid #E5E5E5;
                padding: 12px;
                margin: 8px 0;
                border-radius: 6px;
                border-left: 4px solid #4CAF50;
            }}
            .primary-file {{
                background: #4CAF50;
                color: white;
                font-weight: bold;
            }}
            .footer {{
                text-align: center;
                color: #666666;
                font-size: 14px;
                margin-top: 32px;
                padding-top: 20px;
                border-top: 1px solid #E5E5E5;
            }}
            .logo {{
                color: #4CAF50;
                font-weight: bold;
            }}
            a {{
                color: #000000;
                text-decoration: none;
                border-bottom: 2px solid #4CAF50;
            }}
            a:hover {{
                background-color: #4CAF50;
                color: #000000;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Hyperplexity Analysis Complete</h1>
            <p>Your validation results are ready</p>
        </div>
        
        <div class="content">
            <div class="summary">
                {preview_notice}
                <p><b>Total rows processed:</b> {total_rows:,}</p>
                <p><b>Columns validated:</b> {len(fields_validated)}</p>
                <p><small>({fields_html})</small></p>
                {api_calls_info}
                {time_info}
                {cost_info}
                
                <div class="confidence">
                    <p><b>Original and Updated Confidence Distribution:</b></p>
                    <ul>
                        {confidence_html}
                    </ul>
                </div>
            </div>
            
            <div class="attachments">
                <h3>📎 Attached Files</h3>
                {"<div style='background: #FFF3CD; border: 2px solid #FFB800; padding: 15px; margin: 15px 0; border-radius: 8px;'><b>⚠️ Important Notice:</b> The enhanced Excel file could not be generated due to a processing error. We have included your original file and you have not been charged for this validation. Please contact support for assistance.</div>" if not enhanced_excel_valid else ""}
                <ul class="attachment-list">
                    {"<li class='primary-file'>📊 <b>" + (enhanced_excel_filename or 'Validated_Results.xlsx') + "</b><br><small>Two worksheets: 'Updated' (with improved values) and 'Original' (fact-checked only). Both are color-coded by confidence level.</small></li>" if enhanced_excel_valid else ""}
                    <li>
                        📄 <b>{input_filename or 'Original_Table.xlsx'}</b><br>
                        <small>Your original uploaded table (unchanged)</small>
                    </li>
                    {"<li>🧾 <b>Receipt</b><br><small>Payment receipt for this validation</small></li>" if (billing_info and billing_info.get('amount_charged', 0) > 0 and not preview_email) else ""}
                </ul>
                <p><strong>Configuration Code:</strong> {config_id or 'N/A'}</p>
                <p><small>Save this code to keep the same settings for future validations.</small></p>
            </div>
            
            <p><b>Visit <a href="https://eliyahu.ai/hyperplexity">eliyahu.ai/hyperplexity</a></b> to process more tables or refine your configuration.</p>
            
            <p><b>Questions or need help?</b> Simply reply to this email and our team will assist you.</p>
        </div>
        
        <div class="footer">
            <p>Best regards,<br>
            The <a href="https://eliyahu.ai/hyperplexity">Eliyahu.AI</a> Team</p>
            
            <p><small>This email was sent because you requested validation results.</small></p>
        </div>
    </body>
    </html>
    """
    
    return body_html


def send_validation_code_email(email_address: str, validation_code: str):
    """
    Send email validation code to user.
    
    Args:
        email_address: Recipient email
        validation_code: 6-digit numerical code
        
    Returns:
        dict: Response with status and message ID
    """
    try:
        # Create message
        message = MIMEMultipart()
        message["From"] = SENDER
        message["To"] = email_address
        message["Subject"] = "Perplexity Validator - Email Verification Code"
        
        # Create email body with validation code
        body_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Email Verification - Perplexity Validator</title>
            <style>
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    line-height: 1.6;
                    color: #000000;
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                    background-color: #ffffff;
                }}
                .header {{
                    background: #000000;
                    color: white;
                    padding: 30px;
                    text-align: center;
                    border-radius: 8px 8px 0 0;
                }}
                .content {{
                    background: #ffffff;
                    padding: 30px;
                    border: 1px solid #E5E5E5;
                    border-radius: 0 0 8px 8px;
                }}
                .verification-code {{
                    background: #F8F9FA;
                    border: 2px solid #4CAF50;
                    color: #000000;
                    font-size: 32px;
                    font-weight: bold;
                    padding: 20px;
                    text-align: center;
                    border-radius: 8px;
                    letter-spacing: 8px;
                    margin: 20px 0;
                    font-family: 'Courier New', monospace;
                    position: relative;
                    transition: all 0.3s ease;
                }}
                .verification-code.hidden {{
                    background: #E5E5E5;
                    border-color: #E5E5E5;
                    color: #E5E5E5;
                    user-select: none;
                }}
                .verification-code.hidden::after {{
                    content: '●●●●●●';
                    position: absolute;
                    top: 50%;
                    left: 50%;
                    transform: translate(-50%, -50%);
                    color: #666666;
                    font-size: 24px;
                    letter-spacing: 12px;
                }}
                .code-notice {{
                    background: #FFFFFF;
                    border: 1px solid #00FF00;
                    color: #000000;
                    padding: 12px 20px;
                    border-radius: 8px;
                    margin: 15px 0;
                    text-align: center;
                    border-left: 4px solid #4CAF50;
                    font-size: 14px;
                }}
                .code-notice.hidden {{
                    display: none;
                }}
                .privacy-notice {{
                    background: #FFFFFF;
                    border: 1px solid #00FF00;
                    color: #000000;
                    padding: 25px;
                    border-radius: 8px;
                    margin: 25px 0;
                    border-left: 4px solid #4CAF50;
                    text-align: center;
                }}
                .privacy-link {{
                    color: #000000;
                    text-decoration: none;
                    border-bottom: 2px solid #4CAF50;
                    font-weight: bold;
                }}
                .privacy-link:hover {{
                    background-color: #4CAF50;
                    color: #000000;
                }}
                .acceptance-warning {{
                    background: #FFFFFF;
                    border: 2px solid #4CAF50;
                    color: #000000;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 15px 0;
                    text-align: center;
                    font-size: 14px;
                    border-left: 6px solid #00FF00;
                }}
                .privacy-link-container {{
                    text-align: center;
                    margin: 20px 0;
                }}
                .privacy-button {{
                    display: inline-block;
                    background: #000000;
                    color: white;
                    text-decoration: none;
                    padding: 15px 30px;
                    border-radius: 8px;
                    font-size: 16px;
                    font-weight: bold;
                    border: 2px solid #4CAF50;
                    transition: all 0.3s ease;
                }}
                .privacy-button:hover {{
                    background: #00FF00;
                    color: #000000;
                    text-decoration: none;
                }}

                .warning {{
                    background: #FFFFFF;
                    border: 1px solid #E5E5E5;
                    color: #000000;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 20px 0;
                    border-left: 4px solid #4CAF50;
                }}
                .footer {{
                    text-align: center;
                    color: #666666;
                    font-size: 14px;
                    margin-top: 32px;
                    padding-top: 20px;
                    border-top: 1px solid #E5E5E5;
                }}
                .logo {{
                    color: #4CAF50;
                    font-weight: bold;
                }}
            </style>

        </head>
        <body>
            <div class="header">
                <h1>🔐 Email Verification</h1>
                <p>Perplexity Validator Access Request</p>
            </div>
            
            <div class="content">
                <h2>Verify Your Email Address</h2>
                <p>You have requested access to the Perplexity Validator service. Use the verification code below to complete your registration:</p>
                
                <div class="verification-code">
                    {validation_code}
                </div>
                
                <div class="privacy-notice">
                    <p><strong>📋 Privacy Notice Acceptance Required</strong></p>
                    <p><strong>⚠️ IMPORTANT: Entering and submitting this verification code in the Perplexity Validator interface constitutes your explicit acceptance of our <a href="https://eliyahu.ai/privacy-notice" target="_blank" class="privacy-link">Privacy Notice</a> and consent to data processing.</strong></p>
                    
                    <div class="privacy-link-container">
                        <a href="https://eliyahu.ai/privacy-notice" target="_blank" class="privacy-button">
                            📖 Read Our Privacy Notice
                        </a>
                    </div>
                    
                    <p class="acceptance-warning"><strong>🔒 By proceeding with email verification, you acknowledge that you have read and agree to our Privacy Notice.</strong></p>
                </div>
                
                <div class="warning">
                    <strong>⏰ Important:</strong> This verification code will expire in <strong>10 minutes</strong>.
                    You have up to 3 attempts to enter the correct code.
                </div>
                
                <h3>How to Use Your Code:</h3>
                <ol>
                    <li>Copy the 6-digit verification code above</li>
                    <li>Return to the Perplexity Validator interface</li>
                    <li>Enter the code and click "Verify Email"</li>
                </ol>
                
                <p><strong>If you didn't request this verification,</strong> you can safely ignore this email. The code will expire automatically.</p>
                
                <p><strong>Need help?</strong> Reply to this email and our team will assist you.</p>
            </div>
            
            <div class="footer">
                <p>Best regards,<br>
                The <span class="logo">Eliyahu.AI</span> Team</p>
                
                <p><small>This is an automated security email for Perplexity Validator access verification.</small></p>
            </div>
        </body>
        </html>
        """
        
        # Attach HTML body
        part = MIMEText(body_html, 'html', CHARSET)
        message.attach(part)
        
        # Send email via SES
        ses_client = boto3.client('ses')
        
        response = ses_client.send_raw_email(
            Source=SENDER,
            Destinations=[email_address],
            RawMessage={
                'Data': message.as_string()
            }
        )
        
        message_id = response['MessageId']
        logger.info(f"Validation email sent successfully! Message ID: {message_id}")
        
        return {
            'success': True,
            'message_id': message_id,
            'message': f"Verification code sent to {email_address}"
        }
        
    except ClientError as e:
        error_message = e.response['Error']['Message']
        logger.error(f"SES error sending validation email: {error_message}")
        
        return {
            'success': False,
            'error': error_message,
            'message': f"Failed to send verification email: {error_message}"
        }
        
    except Exception as e:
        logger.error(f"Unexpected error sending validation email: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'message': f"Unexpected error: {str(e)}"
        }

def create_preview_email_body(markdown_table, total_rows, processing_time):
    """Create email body for preview mode results"""
    
    body_html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .header {{ background-color: #f8f9fa; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
            table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 0.9em; }}
            .logo {{ color: #007bff; font-weight: bold; }}
            pre {{ background-color: #f5f5f5; padding: 10px; border-radius: 5px; overflow-x: auto; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h2>Perplexity Validation Preview from <span class="logo">Eliyahu.AI</span></h2>
            <p>Preview of first row validation results</p>
        </div>
        
        <p>This is a preview showing validation results for the first row of your data.</p>
        
        <p><b>Total rows in file:</b> {total_rows}<br>
        <b>Processing time for first row:</b> {processing_time:.2f} seconds<br>
        <b>Estimated time for full validation:</b> {(total_rows * processing_time):.1f} seconds</p>
        
        <h3>Preview Results</h3>
        <pre>{markdown_table}</pre>
        
        <p>To validate the complete dataset, please use the normal mode.</p>
        
        <div class="footer">
            <p>Best regards,<br>
            The <a href="https://www.eliyahu.ai">Eliyahu.AI</a> Team</p>
        </div>
    </body>
    </html>
    """
    
    return body_html 