"""
AWS Lambda handler for the perplexity-validator-interface function.
Provides API interface for Excel table validation with preview capabilities.

Supports two main workflows:
1. Normal workflow: Upload files to S3, return immediate download link (placeholder zip)
2. Preview workflow: Process first row only, return Markdown table with time estimation
"""
import json
import boto3
import base64
import logging
import os
import time
import uuid
import tempfile
import zipfile
from datetime import datetime, timezone
from urllib.parse import unquote_plus
import io
import openpyxl
import csv
from io import StringIO
import secrets
import math

# Set up logging FIRST before any logger usage
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Import row key utilities and validation history loader conditionally
try:
    from row_key_utils import generate_row_key
    ROW_KEY_UTILS_AVAILABLE = True
    logger.info("Row key utilities imported successfully")
except ImportError as e:
    ROW_KEY_UTILS_AVAILABLE = False
    logger.warning(f"Row key utilities not available: {e}")
    
    # Row key generation is now handled by the imported function

try:
    from lambda_test_json_clean import load_validation_history_from_excel
    VALIDATION_HISTORY_AVAILABLE = True
    logger.info("Validation history loader imported successfully")
except ImportError as e:
    VALIDATION_HISTORY_AVAILABLE = False
    logger.warning(f"Validation history loader not available: {e}")
    
    # Fallback function using openpyxl instead of pandas
    def load_validation_history_from_excel(excel_path):
        """Load validation history from Excel using openpyxl (pandas-free implementation)."""
        logger.info(f"Using fallback validation history loader for: {excel_path}")
        try:
            # Import row key utilities if available
            # No sanitization needed for hash-based row keys
            
            # Load workbook with openpyxl
            logger.info(f"Loading workbook with openpyxl from: {excel_path}")
            workbook = openpyxl.load_workbook(excel_path, read_only=True)
            
            # Log available sheet names
            sheet_names = workbook.sheetnames
            logger.info(f"Available sheets: {sheet_names}")
            
            # Check for Details sheet
            if 'Details' not in sheet_names:
                logger.info("No Details worksheet found in Excel file")
                workbook.close()
                return {}
            
            # Load Details worksheet
            details_sheet = workbook['Details']
            logger.info("Found Details worksheet")
            
            # Get headers from first row
            headers = []
            for cell in details_sheet[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    headers.append('')
            
            logger.info(f"Details sheet headers: {headers}")
            
            # Create mapping for ID columns (handle ID: prefix for backwards compatibility)
            id_column_mapping = {}
            for header in headers:
                if header.startswith('ID:'):
                    # Map ID:ColumnName to ColumnName for backwards compatibility
                    clean_name = header[3:]  # Remove 'ID:' prefix
                    id_column_mapping[clean_name] = header
            
            logger.info(f"ID column mapping (for backwards compatibility): {id_column_mapping}")
            
            # Find column indices
            col_indices = {
                'row_key': None,
                'column': None,
                'value': None,
                'confidence': None,
                'quote': None,
                'sources': None,
                'timestamp': None
            }
            
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if 'row key' in header_lower:
                    col_indices['row_key'] = idx
                elif header_lower == 'column':
                    col_indices['column'] = idx
                elif 'validated value' in header_lower:
                    col_indices['value'] = idx
                elif header_lower == 'confidence':
                    col_indices['confidence'] = idx
                elif header_lower == 'quote':
                    col_indices['quote'] = idx
                elif header_lower == 'sources':
                    col_indices['sources'] = idx
                elif header_lower == 'timestamp':
                    col_indices['timestamp'] = idx
            
            logger.info(f"Column indices found: {col_indices}")
            
            # Check if we have minimum required columns
            if col_indices['row_key'] is None or col_indices['column'] is None:
                logger.warning("Required columns (Row Key, Column) not found in Details sheet")
                workbook.close()
                return {}
            
            # Convert to validation_history structure
            validation_history = {}
            row_count = 0
            
            # Iterate through rows (skip header)
            for row_idx, row in enumerate(details_sheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Extract values with safe indexing
                    row_key = str(row[col_indices['row_key']] or '') if col_indices['row_key'] < len(row) else ''
                    column = str(row[col_indices['column']] or '') if col_indices['column'] < len(row) else ''
                    
                    # Skip empty rows
                    if not row_key or not column:
                        continue
                    
                    # Get other values safely
                    value = str(row[col_indices['value']] or '') if col_indices['value'] and col_indices['value'] < len(row) else ''
                    confidence = str(row[col_indices['confidence']] or '') if col_indices['confidence'] and col_indices['confidence'] < len(row) else ''
                    quote = str(row[col_indices['quote']] or '') if col_indices['quote'] and col_indices['quote'] < len(row) else ''
                    sources_str = str(row[col_indices['sources']] or '') if col_indices['sources'] and col_indices['sources'] < len(row) else ''
                    timestamp = str(row[col_indices['timestamp']] or '') if col_indices['timestamp'] and col_indices['timestamp'] < len(row) else ''
                    
                    # Use the row key as-is (no conversion needed for hash-based keys)
                    sanitized_row_key = row_key
                    
                    # Initialize structures
                    if sanitized_row_key not in validation_history:
                        validation_history[sanitized_row_key] = {}
                    
                    if column not in validation_history[sanitized_row_key]:
                        validation_history[sanitized_row_key][column] = []
                    
                    # Parse sources
                    sources = []
                    if sources_str and sources_str != 'N/A' and sources_str.strip():
                        sources = [s.strip() for s in sources_str.split(';') if s.strip()]
                    
                    # Default timestamp if missing
                    if not timestamp or timestamp == 'N/A' or timestamp == 'nan':
                        timestamp = datetime.utcnow().isoformat()
                    
                    # Create history entry
                    history_entry = {
                        'timestamp': timestamp,
                        'value': value,
                        'confidence_level': confidence,
                        'quote': quote if quote != 'N/A' else '',
                        'sources': sources
                    }
                    
                    validation_history[sanitized_row_key][column].append(history_entry)
                    row_count += 1
                    
                    # Log first few entries for debugging
                    if row_count <= 3:
                        logger.info(f"Sample entry {row_count}: row_key='{sanitized_row_key}', column='{column}', value='{value}'")
                    
                except Exception as row_error:
                    logger.warning(f"Error processing row {row_idx}: {row_error}")
                    continue
            
            workbook.close()
            logger.info(f"Loaded validation history from Details worksheet for {len(validation_history)} row keys ({row_count} entries)")
            
            # Log sample keys for debugging
            if validation_history:
                sample_keys = list(validation_history.keys())[:3]
                logger.info(f"Sample validation history keys: {sample_keys}")
            
            return validation_history
            
        except Exception as e:
            logger.error(f"Error loading validation history from Excel: {e}")
            import traceback
            logger.error(traceback.format_exc())
        return {}

# Import multipart handling
try:
    from multipart import parse_form_data
except ImportError:
    # Fallback for basic parsing if multipart library not available
    parse_form_data = None

# Enhanced Excel functionality 
try:
    import xlsxwriter
    EXCEL_ENHANCEMENT_AVAILABLE = True
    logger.info("Enhanced Excel functionality available")
except ImportError:
    EXCEL_ENHANCEMENT_AVAILABLE = False
    logger.warning("Enhanced Excel functionality not available")

# Email sending functionality
try:
    from email_sender import send_validation_results_email, create_preview_email_body
    EMAIL_SENDER_AVAILABLE = True
    logger.info("Email sender functionality available")
except ImportError:
    EMAIL_SENDER_AVAILABLE = False
    logger.warning("Email sender functionality not available - will use download URLs")

# SQS and DynamoDB integration
SQS_INTEGRATION_AVAILABLE = False
SQS_IMPORT_ERROR = None

try:
    # Try importing each module separately to identify which one fails
    import_errors = []
    
    try:
        from sqs_service import send_preview_request, send_full_request, create_all_queues
        logger.info("sqs_service imported successfully")
    except Exception as e:
        import_errors.append(f"sqs_service: {str(e)}")
        logger.error(f"Failed to import sqs_service: {e}")
        
    try:
        from dynamodb_schemas import track_validation_call, create_call_tracking_table, create_token_usage_table, update_processing_metrics
        logger.info("dynamodb_schemas imported successfully")
    except Exception as e:
        import_errors.append(f"dynamodb_schemas: {str(e)}")
        logger.error(f"Failed to import dynamodb_schemas: {e}")
        
    try:
        from api_gateway_validation import validate_api_request, create_validation_error_response
        logger.info("api_gateway_validation imported successfully")
    except Exception as e:
        import_errors.append(f"api_gateway_validation: {str(e)}")
        logger.error(f"Failed to import api_gateway_validation: {e}")
    
    # If all imports succeeded, enable SQS integration
    if not import_errors:
        SQS_INTEGRATION_AVAILABLE = True
        logger.info("SQS and DynamoDB integration available - all imports successful")
    else:
        SQS_IMPORT_ERROR = "; ".join(import_errors)
        logger.error(f"SQS integration not available due to import errors: {SQS_IMPORT_ERROR}")
        
except Exception as e:
    SQS_IMPORT_ERROR = str(e)
    logger.error(f"Critical error during SQS imports: {e}")
    import traceback
    logger.error(traceback.format_exc())

# AWS clients
s3_client = boto3.client('s3')
lambda_client = boto3.client('lambda')

# Environment variables
S3_CACHE_BUCKET = os.environ.get('S3_CACHE_BUCKET', 'perplexity-cache')
S3_RESULTS_BUCKET = os.environ.get('S3_RESULTS_BUCKET', 'perplexity-results')
VALIDATOR_LAMBDA_NAME = os.environ.get('VALIDATOR_LAMBDA_NAME', 'perplexity-validator')

def generate_reference_pin() -> str:
    """Generate a 6-digit reference pin for the run."""
    return f"{secrets.randbelow(900000) + 100000:06d}"

def create_email_folder_path(email_address: str) -> str:
    """
    Create S3 folder path based on email address.
    If @ is allowed in S3 keys, use email directly.
    Otherwise, use domain/user structure.
    """
    if not email_address:
        return "no-email"
    
    try:
        # Clean the email address for use in S3 paths
        cleaned_email = email_address.lower().strip()
        
        # S3 allows @ in object keys, so we can use it directly
        # But let's replace @ with underscore for better compatibility
        # and avoid potential issues with some tools
        if '@' in cleaned_email:
            user, domain = cleaned_email.split('@', 1)
            # Use domain/user structure for better organization
            folder_path = f"{domain}/{user}"
        else:
            # If no @ found, use as-is
            folder_path = cleaned_email
        
        # Sanitize for S3 compatibility
        # Replace any remaining special characters
        folder_path = folder_path.replace(' ', '_').replace('+', 'plus')
        
        return folder_path
    except Exception as e:
        logger.warning(f"Error creating email folder path for {email_address}: {e}")
        return "email-error"

def parse_multipart_form_data(body, content_type, is_base64_encoded=False):
    """Parse multipart form data from the request body."""
    try:
        # Decode base64 if needed
        if is_base64_encoded:
            body = base64.b64decode(body)
        elif isinstance(body, str):
            body = body.encode('utf-8')
        
        # Extract boundary from content type
        boundary = None
        if content_type and 'boundary=' in content_type:
            boundary = content_type.split('boundary=')[1]
            if boundary.startswith('"') and boundary.endswith('"'):
                boundary = boundary[1:-1]
        
        if not boundary:
            raise ValueError("No boundary found in Content-Type header")
        
        # Simple multipart parser
        boundary_bytes = ('--' + boundary).encode('utf-8')
        end_boundary_bytes = ('--' + boundary + '--').encode('utf-8')
        
        files = {}
        form_data = {}
        
        # Split by boundary
        parts = body.split(boundary_bytes)
        
        for part in parts[1:]:  # Skip first empty part
            if end_boundary_bytes[2:] in part:  # End boundary
                break
            
            if not part.strip():
                continue
            
            # Find headers and content
            if b'\r\n\r\n' in part:
                headers_part, content = part.split(b'\r\n\r\n', 1)
                content = content.rstrip(b'\r\n')
            else:
                continue
            
            headers = headers_part.decode('utf-8', errors='ignore')
            
            # Parse Content-Disposition header
            name = None
            filename = None
            for line in headers.split('\r\n'):
                if line.lower().startswith('content-disposition:'):
                    # Extract name and filename
                    if 'name="' in line:
                        name = line.split('name="')[1].split('"')[0]
                    if 'filename="' in line:
                        filename = line.split('filename="')[1].split('"')[0]
            
            if name:
                if filename:  # File field
                    files[name] = {
                        'filename': filename,
                        'content': content  # Keep binary content as-is
                    }
                else:  # Regular form field (text only)
                    # Special handling for config_file field
                    if name == 'config_file':
                        try:
                            # Try to decode as UTF-8 first
                            decoded_content = content.decode('utf-8')
                            form_data[name] = decoded_content
                        except UnicodeDecodeError:
                            # If UTF-8 fails, try other encodings
                            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                                try:
                                    decoded_content = content.decode(encoding)
                                    form_data[name] = decoded_content
                                    logger.warning(f"Decoded config_file using {encoding} encoding")
                                    break
                                except UnicodeDecodeError:
                                    continue
                            else:
                                # If all encodings fail, try to extract JSON-like content
                                try:
                                    # Find JSON boundaries in the raw bytes
                                    start_idx = content.find(b'{')
                                    end_idx = content.rfind(b'}')
                                    if start_idx != -1 and end_idx != -1:
                                        json_bytes = content[start_idx:end_idx+1]
                                        decoded_content = json_bytes.decode('utf-8', errors='ignore')
                                        form_data[name] = decoded_content
                                        logger.warning("Extracted JSON content from config_file with errors ignored")
                                    else:
                                        # Last resort: store as base64
                                        form_data[name] = base64.b64encode(content).decode('ascii')
                                        logger.error(f"Could not decode config_file, stored as base64")
                                except Exception as e:
                                    logger.error(f"Failed to extract JSON from config_file: {e}")
                                    form_data[name] = str(content)  # Store raw representation
                    else:
                        # For other form fields, use standard UTF-8 decoding
                        try:
                            form_data[name] = content.decode('utf-8')
                        except UnicodeDecodeError:
                                logger.warning(f"Could not decode form field '{name}' as UTF-8. Trying latin-1.")
                                try:
                                    form_data[name] = content.decode('latin-1')
                                except Exception:
                                    logger.error(f"Could not decode form field '{name}'. Storing as hex.")
                                    form_data[name] = content.hex()
        
        # Log parsed data for debugging
        logger.info(f"Parsed form data fields: {list(form_data.keys())}")
        logger.info(f"Parsed files: {[f for f in files.keys()]}")
        
        # Additional validation for config_file
        if 'config_file' in form_data:
            config_content = form_data['config_file']
            # Try to validate it's JSON
            try:
                json.loads(config_content)
                logger.info("config_file contains valid JSON")
            except json.JSONDecodeError as e:
                logger.warning(f"config_file does not contain valid JSON: {e}")
                # Log first 200 chars for debugging
                logger.warning(f"config_file content preview: {config_content[:200]}...")
        
        return files, form_data
        
    except Exception as e:
        logger.error(f"Error parsing multipart data: {str(e)}")
        raise

def create_placeholder_zip():
    """Create a placeholder ZIP file indicating processing is in progress."""
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Add a README file
        readme_content = """
Processing in Progress
=====================

Your Excel file is currently being processed by the Perplexity Validator.

Processing Status: IN PROGRESS
Started: {timestamp}

Please check back in a few minutes. This file will be automatically 
replaced with your validation results once processing is complete.

If processing takes longer than expected, please contact support.
        """.format(timestamp=datetime.utcnow().isoformat() + 'Z').strip()
        
        zip_file.writestr('README.txt', readme_content)
        zip_file.writestr('STATUS.txt', 'PROCESSING')
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

def create_validation_result_zip(validation_results, session_id, total_rows):
    """Create a ZIP file with real validation results."""
    zip_buffer = io.BytesIO()
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        
        # Create validation results JSON
        results_data = {
            "session_id": session_id,
            "timestamp": timestamp,
            "validation_results": validation_results,
            "summary": {
                "total_rows": total_rows,
                "fields_validated": [],
                "confidence_distribution": {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
            }
        }
        
        # Analyze results for summary
        if validation_results:
            all_fields = set()
            confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
            
            for row_key, row_data in validation_results.items():
                for field_name, field_data in row_data.items():
                    if isinstance(field_data, dict) and 'confidence_level' in field_data:
                        all_fields.add(field_name)
                        conf_level = field_data.get('confidence_level', 'UNKNOWN')
                        if conf_level in confidence_counts:
                            confidence_counts[conf_level] += 1
            
            results_data["summary"]["fields_validated"] = list(all_fields)
            results_data["summary"]["confidence_distribution"] = confidence_counts
        
        # Add JSON file
        zip_file.writestr('validation_results.json', json.dumps(results_data, indent=2))
        
        # Create a summary report
        summary_report = f"""
Validation Results Summary
=========================

Session ID: {session_id}
Generated: {timestamp}
Total Rows Processed: {total_rows}

Fields Validated: {', '.join(results_data['summary']['fields_validated'])}

Confidence Distribution:
- HIGH: {results_data['summary']['confidence_distribution']['HIGH']} fields
- MEDIUM: {results_data['summary']['confidence_distribution']['MEDIUM']} fields  
- LOW: {results_data['summary']['confidence_distribution']['LOW']} fields

Processing Details:
==================
{json.dumps(validation_results, indent=2) if validation_results else 'No validation results'}
        """.strip()
        
        zip_file.writestr('SUMMARY.txt', summary_report)
        
        # Add completion marker
        zip_file.writestr('COMPLETED.txt', f'Validation completed at {timestamp}')
        
        # Create CSV report for easy viewing
        if validation_results:
            # Use Python's csv module for proper escaping
            csv_buffer = StringIO()
            csv_writer = csv.writer(csv_buffer)
            
            # Write header
            csv_writer.writerow(['Field', 'Value', 'Confidence', 'Sources', 'Quote'])
            
            for row_key, row_data in validation_results.items():
                for field_name, field_data in row_data.items():
                    if isinstance(field_data, dict) and 'confidence_level' in field_data:
                        value = str(field_data.get('value', ''))
                        confidence = field_data.get('confidence_level', 'UNKNOWN')
                        sources = ', '.join(field_data.get('sources', []))[:100] + '...' if len(field_data.get('sources', [])) > 0 else ''
                        quote = str(field_data.get('quote', ''))[:100] + '...' if len(str(field_data.get('quote', ''))) > 100 else str(field_data.get('quote', ''))
                        
                        csv_writer.writerow([field_name, value, confidence, sources, quote])
            
            zip_file.writestr('validation_results.csv', csv_buffer.getvalue())
            csv_buffer.close()
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

def safe_for_excel(value):
    """Convert value to Excel-safe format, handling control characters but NOT XML escaping."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Check for NaN without pandas
        if value != value:  # NaN check
            return ""
        # Check for infinity
        if isinstance(value, float) and (value == float('inf') or value == float('-inf')):
            return ""
    
    # Convert to string for processing
    value_str = str(value)
    
    # First, handle control characters that are illegal in XML
    # Replace all control characters except tab (9), newline (10), and carriage return (13)
    cleaned = []
    for char in value_str:
        code = ord(char)
        if code < 32 and code not in (9, 10, 13):
            # Replace illegal control characters with space
            cleaned.append(' ')
        elif code > 127 and code < 160:
            # Replace non-breaking spaces and other problematic high ASCII
            cleaned.append(' ')
        elif code == 8232 or code == 8233:
            # Replace Unicode line/paragraph separators with regular newlines
            cleaned.append('\n')
        else:
            cleaned.append(char)
    value_str = ''.join(cleaned)
    
    # IMPORTANT: Do NOT escape XML characters here!
    # xlsxwriter handles XML escaping internally.
    # Double-escaping causes corruption.
    
    # Handle Excel's cell content limit
    if len(value_str) > 32767:
        return value_str[:32700] + "..."
    
    return value_str

def create_enhanced_excel_with_validation(excel_file_content, validation_results, config_data, session_id, skip_history=False):
    """Create color-coded Excel file with validation results and history tracking.
    
    Args:
        excel_file_content: Original Excel file content
        validation_results: Validation results from the validator
        config_data: Configuration data
        session_id: Session ID for tracking
        skip_history: If True, skip loading existing Details sheet (for fallback mode)
    """
    if not EXCEL_ENHANCEMENT_AVAILABLE:
        logger.warning("Enhanced Excel not available, skipping Excel creation")
        return None
        
    try:
        # Create Excel buffer
        excel_buffer = io.BytesIO()
        
        # Load original Excel data
        workbook = openpyxl.load_workbook(io.BytesIO(excel_file_content))
        
        # Select the appropriate sheet (same logic as invoke_validator_lambda)
        if 'Results' in workbook.sheetnames:
            worksheet = workbook['Results']
            logger.info(f"Using 'Results' sheet as data source for enhanced Excel creation")
        elif len(workbook.sheetnames) > 0:
            worksheet = workbook[workbook.sheetnames[0]]
            logger.info(f"Using first sheet '{worksheet.title}' as data source for enhanced Excel creation")
        else:
            worksheet = workbook.active
            logger.info(f"Using active sheet: {worksheet.title}")
        
        # Get headers and data
        headers = [cell.value for cell in worksheet[1]]
        
        # Convert to list of dictionaries 
        rows_data = []
        for row_idx in range(2, worksheet.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers):
                if header:
                    cell_value = worksheet.cell(row=row_idx, column=col_idx + 1).value
                    row_data[header] = str(cell_value) if cell_value is not None else ""
            rows_data.append(row_data)
        
        # Get ID fields from config for proper row key generation
        id_fields = []
        for target in config_data.get('validation_targets', []):
            if target.get('importance', '').upper() == 'ID':
                # Support both 'name' and 'column' fields
                field_name = target.get('name') or target.get('column')
                if field_name:
                    id_fields.append(field_name)
        
        # If no ID fields found, try to use SimplifiedSchemaValidator to determine primary keys
        if not id_fields:
            try:
                from schema_validator_simplified import SimplifiedSchemaValidator
                validator = SimplifiedSchemaValidator(config_data)
                id_fields = validator.primary_key
                logger.info(f"Using primary keys from SimplifiedSchemaValidator: {id_fields}")
            except ImportError:
                logger.warning("SimplifiedSchemaValidator not available in deployment package")
            except Exception as e:
                logger.warning(f"Could not use SimplifiedSchemaValidator: {e}")
        
        # Generate row keys for each row
        row_keys = []
        for row_data in rows_data:
            if ROW_KEY_UTILS_AVAILABLE and id_fields:
                row_key = generate_row_key(row_data, id_fields)
            else:
                # Fallback to simple join
                key_columns = id_fields if id_fields else list(headers)[:3]
                row_key = "||".join([str(row_data.get(col, "")) for col in key_columns])
            row_keys.append(row_key)
        
        # Load existing Details entries from the original Excel (for history preservation)
        existing_details = []
        details_sheet_exists = False
        try:
            if 'Details' in workbook.sheetnames:
                details_sheet_exists = True
                details_worksheet = workbook['Details']
                
                # Read existing details headers
                details_headers = [cell.value for cell in details_worksheet[1]]
                
                # Read existing details data
                for row_idx in range(2, details_worksheet.max_row + 1):
                    detail_row = {}
                    for col_idx, header in enumerate(details_headers):
                        if header:
                            cell_value = details_worksheet.cell(row=row_idx, column=col_idx + 1).value
                            detail_row[header] = cell_value
                    
                    # Mark existing entries as "History"
                    if detail_row:
                        # If the row already has a 'New' column, preserve its value unless it's 'New'
                        if 'New' in detail_row and detail_row['New'] == 'New':
                            detail_row['New'] = 'History'
                        elif 'New' not in detail_row:
                            detail_row['New'] = 'History'
                        existing_details.append(detail_row)
                
                logger.info(f"Loaded {len(existing_details)} existing detail entries from Excel")
        except Exception as e:
            logger.warning(f"Could not load existing Details sheet: {e}")
        
        # Create Excel with xlsxwriter for advanced formatting
        with xlsxwriter.Workbook(excel_buffer, {'strings_to_urls': False, 'nan_inf_to_errors': True}) as workbook:
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'top',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            
            confidence_formats = {
                'HIGH': workbook.add_format({'bold': True, 'fg_color': '#C6EFCE', 'font_color': '#006100'}),
                'MEDIUM': workbook.add_format({'fg_color': '#FFEB9C', 'font_color': '#9C6500'}),
                'LOW': workbook.add_format({'italic': True, 'fg_color': '#FFC7CE', 'font_color': '#9C0006'})
            }
            
            # Create Results worksheet
            results_sheet = workbook.add_worksheet('Results')
            
            # Write headers
            for col_idx, col_name in enumerate(headers):
                results_sheet.write(0, col_idx, col_name, header_format)
                results_sheet.set_column(col_idx, col_idx, 20)  # Set column width
            
            # Write data rows with formatting
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row - try multiple key formats
                row_validation_data = None
                
                # First try the actual row key
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                # Then try position-based keys for backwards compatibility
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                for col_idx, col_name in enumerate(headers):
                    if not col_name:
                        continue
                        
                    cell_value = row_data.get(col_name, '')
                    validated_value = cell_value
                    confidence_level = None
                    quote = None
                    sources = None
                    
                    # Check if this column has validation results
                    if row_validation_data and isinstance(row_validation_data, dict):
                        if col_name in row_validation_data and isinstance(row_validation_data[col_name], dict):
                            field_data = row_validation_data[col_name]
                            validated_value = field_data.get('value', cell_value)
                            confidence_level = field_data.get('confidence_level')
                            quote = field_data.get('quote')
                            sources = field_data.get('sources', [])
                    
                    # Choose format based on confidence level (but not for ID fields)
                    cell_format = None
                    if col_name not in id_fields and confidence_level and confidence_level.upper() in confidence_formats:
                        cell_format = confidence_formats[confidence_level.upper()]
                    
                    # Write cell value
                    results_sheet.write(row_idx + 1, col_idx, safe_for_excel(validated_value or ''), cell_format)
                    
                    # Add comment if we have quote or sources
                    if quote or sources:
                        comment_parts = []
                        if quote and str(quote).strip() and quote != 'N/A':
                            comment_parts.append(f'Quote: "{quote}"')
                        if sources and isinstance(sources, list) and sources:
                            sources_text = ', '.join(sources[:3])  # Limit to first 3 sources
                            comment_parts.append(f'Sources: {sources_text}')
                        
                        if comment_parts:
                            comment_text = '\n\n'.join(comment_parts)
                            try:
                                results_sheet.write_comment(row_idx + 1, col_idx, comment_text, 
                                                          {'width': 300, 'height': 150})
                            except Exception as e:
                                logger.warning(f"Could not add comment: {e}")
            
            # Create Details worksheet with proper structure matching local version
            details_sheet = workbook.add_worksheet('Details')
            
            # Build detail headers dynamically to include ID fields
            detail_headers = ["Row Key", "Identifier"]
            
            # Add ID field columns without "ID:" prefix
            for id_field in id_fields:
                detail_headers.append(id_field)
            
            # Add the rest of the standard columns
            detail_headers.extend(["Column", "Original Value", "Validated Value", 
                            "Confidence", "Quote", "Sources", "Explanation", "Update Required", 
                            "Substantially Different", "Consistent with Model", "Model", "Timestamp", "New"])
            
            for col_idx, header in enumerate(detail_headers):
                details_sheet.write(0, col_idx, header, header_format)
            
            # Set column widths dynamically
            col_idx = 0
            details_sheet.set_column(col_idx, col_idx, 15)  # Row Key
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 30)  # Identifier
            col_idx += 1
            
            # ID field columns
            for _ in id_fields:
                details_sheet.set_column(col_idx, col_idx, 20)
                col_idx += 1
            
            # Standard columns
            details_sheet.set_column(col_idx, col_idx, 25)  # Column
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 30)  # Original Value
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 30)  # Validated Value
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 15)  # Confidence
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 50)  # Quote
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 40)  # Sources
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 60)  # Explanation
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 15)  # Update Required
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 20)  # Substantially Different
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 25)  # Consistent with Model
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 25)  # Model
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 20)  # Timestamp
            col_idx += 1
            details_sheet.set_column(col_idx, col_idx, 10)  # New
            
            detail_row = 1
            current_timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            
            # Track which row keys have been processed to avoid duplicates
            processed_row_keys = set()
            
            # First, write NEW validation results
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row
                row_validation_data = None
                
                # First try the actual row key
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                # Then try position-based keys for backwards compatibility
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                if row_validation_data and isinstance(row_validation_data, dict):
                    # Create identifier from ID fields
                    identifier_parts = []
                    for id_field in id_fields:
                        if id_field in row_data:
                            identifier_parts.append(f"{id_field}: {row_data[id_field]}")
                    identifier = ", ".join(identifier_parts) if identifier_parts else f"Row {row_idx + 1}"
                    
                    for field_name, field_data in row_validation_data.items():
                        if isinstance(field_data, dict) and 'confidence_level' in field_data:
                            # Write detail row with all columns
                            col_idx = 0
                            details_sheet.write(detail_row, col_idx, safe_for_excel(row_key))  # Row Key
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(identifier))  # Identifier
                            col_idx += 1
                            
                            # Write ID field values
                            for id_field in id_fields:
                                value = row_data.get(id_field, '')
                                details_sheet.write(detail_row, col_idx, safe_for_excel(str(value)))
                                col_idx += 1
                            
                            # Write standard columns
                            details_sheet.write(detail_row, col_idx, safe_for_excel(field_name))  # Column
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(row_data.get(field_name, ''))))  # Original value
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('value', ''))))  # Validated value
                            col_idx += 1
                            
                            confidence = field_data.get('confidence_level', '')
                            confidence_format = confidence_formats.get(confidence.upper()) if confidence else None
                            details_sheet.write(detail_row, col_idx, safe_for_excel(confidence), confidence_format)
                            col_idx += 1
                            
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('quote', ''))))
                            col_idx += 1
                            sources_text = ', '.join(field_data.get('sources', []))
                            details_sheet.write(detail_row, col_idx, safe_for_excel(sources_text))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('explanation', ''))))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('update_required', ''))))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('substantially_different', ''))))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('consistent_with_model_knowledge', ''))))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(field_data.get('model', ''))))  # Model
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, safe_for_excel(current_timestamp))
                            col_idx += 1
                            details_sheet.write(detail_row, col_idx, 'New')  # Mark all current validations as "New"
                            
                            # Track this combination as processed
                            processed_row_keys.add((row_key, field_name))
                            detail_row += 1
            
            # Then, append existing HISTORICAL details
            for existing_detail in existing_details:
                # Check if this is a duplicate of a new entry
                existing_row_key = existing_detail.get('Row Key', '')
                existing_column = existing_detail.get('Column', '')
                
                if (existing_row_key, existing_column) not in processed_row_keys:
                    # Write the historical entry
                    col_idx = 0
                    
                    # Row Key
                    details_sheet.write(detail_row, col_idx, safe_for_excel(str(existing_detail.get('Row Key', ''))))
                    col_idx += 1
                    
                    # Identifier
                    details_sheet.write(detail_row, col_idx, safe_for_excel(str(existing_detail.get('Identifier', ''))))
                    col_idx += 1
                    
                    # ID field values - extract from existing detail or try to reconstruct
                    for id_field in id_fields:
                        # Check for ID field with or without "ID:" prefix (for backwards compatibility)
                        value = ''
                        
                        # First try without prefix (new format)
                        if id_field in existing_detail:
                            value = existing_detail.get(id_field, '')
                        # Then try with prefix (old format)
                        elif f"ID:{id_field}" in existing_detail:
                            value = existing_detail.get(f"ID:{id_field}", '')
                        else:
                            # Try to extract from Row Key if not present
                            # This is a fallback for very old format Details sheets
                            if existing_row_key and '||' in existing_row_key:
                                # Try to parse the row key (assuming it follows the pattern)
                                key_parts = existing_row_key.split('||')
                                # This is approximate - we can't always reconstruct perfectly
                                id_field_index = id_fields.index(id_field) if id_field in id_fields else -1
                                if 0 <= id_field_index < len(key_parts):
                                    value = key_parts[id_field_index]
                        
                        details_sheet.write(detail_row, col_idx, safe_for_excel(str(value)))
                        col_idx += 1
                    
                    # Standard columns
                    standard_columns = ['Column', 'Original Value', 'Validated Value', 'Confidence', 
                                      'Quote', 'Sources', 'Explanation', 'Update Required', 
                                      'Substantially Different', 'Consistent with Model', 'Model', 'Timestamp', 'New']
                    
                    for col_name in standard_columns:
                        value = existing_detail.get(col_name, '')
                        
                        # Apply confidence formatting if applicable
                        if col_name == 'Confidence' and value and value.upper() in confidence_formats:
                            details_sheet.write(detail_row, col_idx, safe_for_excel(value), confidence_formats[value.upper()])
                        else:
                            details_sheet.write(detail_row, col_idx, safe_for_excel(str(value) if value is not None else ''))
                        
                        col_idx += 1
                    
                    detail_row += 1
            
            logger.info(f"Wrote {detail_row - 1} total detail entries (new + historical)")
            
            # Create Reasons worksheet
            reasons_sheet = workbook.add_worksheet('Reasons')
            reasons_headers = ["Row Key", "Field", "Explanation", "Update Required", "Substantially Different"]
            
            for col_idx, header in enumerate(reasons_headers):
                reasons_sheet.write(0, col_idx, header, header_format)
                reasons_sheet.set_column(col_idx, col_idx, 30)
            
            reasons_row = 1
            for row_idx, (row_data, row_key) in enumerate(zip(rows_data, row_keys)):
                # Get validation results for this row
                row_validation_data = None
                
                # First try the actual row key
                if row_key in validation_results:
                    row_validation_data = validation_results[row_key]
                # Then try position-based keys for backwards compatibility
                elif str(row_idx) in validation_results:
                    row_validation_data = validation_results[str(row_idx)]
                elif row_idx in validation_results:
                    row_validation_data = validation_results[row_idx]
                
                if row_validation_data and isinstance(row_validation_data, dict):
                    for field_name, field_data in row_validation_data.items():
                        if isinstance(field_data, dict) and 'explanation' in field_data:
                            reasons_sheet.write(reasons_row, 0, safe_for_excel(row_key))  # Use actual row key
                            reasons_sheet.write(reasons_row, 1, safe_for_excel(field_name))
                            reasons_sheet.write(reasons_row, 2, safe_for_excel(str(field_data.get('explanation', ''))))
                            reasons_sheet.write(reasons_row, 3, safe_for_excel(str(field_data.get('update_required', ''))))
                            reasons_sheet.write(reasons_row, 4, safe_for_excel(str(field_data.get('substantially_different', ''))))
                            
                            reasons_row += 1
        
        excel_buffer.seek(0)
        return excel_buffer
        
    except Exception as e:
        logger.error(f"Error creating enhanced Excel: {str(e)}")
        return None

def create_enhanced_result_zip(validation_results, session_id, total_rows, excel_file_content, config_data, reference_pin=None, input_filename='input.xlsx', config_filename='config.json', metadata=None):
    """Create enhanced ZIP file with color-coded Excel and comprehensive reports."""
    zip_buffer = io.BytesIO()
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        
        # Create validation results JSON
        results_data = {
            "session_id": session_id,
            "timestamp": timestamp,
            "reference_pin": reference_pin,
            "validation_results": validation_results,
            "summary": {
                "total_rows": total_rows,
                "fields_validated": [],
                "confidence_distribution": {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
            }
        }
        
        # Add token usage metadata if available
        if metadata:
            results_data["metadata"] = metadata
        
        # Analyze results for summary
        if validation_results:
            all_fields = set()
            confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
            
            for row_key, row_data in validation_results.items():
                for field_name, field_data in row_data.items():
                    if isinstance(field_data, dict) and 'confidence_level' in field_data:
                        all_fields.add(field_name)
                        conf_level = field_data.get('confidence_level', 'UNKNOWN')
                        if conf_level in confidence_counts:
                            confidence_counts[conf_level] += 1
            
            results_data["summary"]["fields_validated"] = list(all_fields)
            results_data["summary"]["confidence_distribution"] = confidence_counts
        
        # Add JSON file
        zip_file.writestr('validation_results.json', json.dumps(results_data, indent=2))
        
        # Create enhanced Excel file
        if excel_file_content and validation_results and EXCEL_ENHANCEMENT_AVAILABLE:
            try:
                excel_buffer = create_enhanced_excel_with_validation(
                    excel_file_content, validation_results, config_data, session_id
                )
                if excel_buffer:
                    zip_file.writestr('validation_results_enhanced.xlsx', excel_buffer.getvalue())
                    logger.info("Enhanced Excel file created successfully")
                else:
                    logger.warning("Enhanced Excel creation returned None")
            except Exception as e:
                logger.error(f"Error creating enhanced Excel: {str(e)}")
        
        # Create CSV report for easy viewing
        if validation_results:
            # Use Python's csv module for proper escaping
            csv_buffer = StringIO()
            csv_writer = csv.writer(csv_buffer)
            
            # Write header
            csv_writer.writerow(['Row', 'Field', 'Original_Value', 'Validated_Value', 'Confidence', 'Sources', 'Quote'])
            
            for row_key, row_data in validation_results.items():
                for field_name, field_data in row_data.items():
                    if isinstance(field_data, dict) and 'confidence_level' in field_data:
                        original = str(field_data.get('original_value', ''))
                        value = str(field_data.get('value', ''))
                        confidence = field_data.get('confidence_level', 'UNKNOWN')
                        sources = ', '.join(field_data.get('sources', []))[:100] + '...' if len(field_data.get('sources', [])) > 0 else ''
                        quote = str(field_data.get('quote', ''))[:100] + '...' if len(str(field_data.get('quote', ''))) > 100 else str(field_data.get('quote', ''))
                        
                        csv_writer.writerow([row_key, field_name, original, value, confidence, sources, quote])
            
            zip_file.writestr('validation_results.csv', csv_buffer.getvalue())
            csv_buffer.close()
        
        # Add original input and config files to the ZIP
        if excel_file_content:
            zip_file.writestr(f'original_files/{input_filename}', excel_file_content)
            logger.info(f"Added original input file to ZIP as: original_files/{input_filename}")
        
        if config_data:
            if isinstance(config_data, dict):
                config_json = json.dumps(config_data, indent=2)
            else:
                config_json = str(config_data)
            zip_file.writestr(f'original_files/{config_filename}', config_json)
            logger.info(f"Added config file to ZIP as: original_files/{config_filename}")
        
        # Create enhanced summary report
        pin_section = f"""
Reference Pin: {reference_pin}
""" if reference_pin else ""
        
        # Generate token usage section
        token_usage_section = ""
        if metadata and 'token_usage' in metadata:
            token_usage = metadata['token_usage']
            token_usage_section = f"""

Token Usage & Cost Analysis:
============================
Total Tokens: {token_usage.get('total_tokens', 0):,}
Total Cost: ${token_usage.get('total_cost', 0.0):.6f}
API Calls: {token_usage.get('api_calls', 0)} new, {token_usage.get('cached_calls', 0)} cached

Provider Breakdown:
"""
            # Add provider-specific details
            if 'by_provider' in token_usage:
                for provider, provider_data in token_usage['by_provider'].items():
                    if provider_data.get('calls', 0) > 0:
                        if provider == 'perplexity':
                            token_usage_section += f"""- Perplexity API: {provider_data.get('prompt_tokens', 0):,} prompt + {provider_data.get('completion_tokens', 0):,} completion = {provider_data.get('total_tokens', 0):,} tokens
  Cost: ${provider_data.get('input_cost', 0.0):.6f} input + ${provider_data.get('output_cost', 0.0):.6f} output = ${provider_data.get('total_cost', 0.0):.6f} total
  Calls: {provider_data.get('calls', 0)}
"""
                        elif provider == 'anthropic':
                            token_usage_section += f"""- Anthropic API: {provider_data.get('input_tokens', 0):,} input + {provider_data.get('output_tokens', 0):,} output
  Cache: {provider_data.get('cache_creation_tokens', 0):,} creation + {provider_data.get('cache_read_tokens', 0):,} read = {provider_data.get('total_tokens', 0):,} total tokens
  Cost: ${provider_data.get('input_cost', 0.0):.6f} input + ${provider_data.get('output_cost', 0.0):.6f} output = ${provider_data.get('total_cost', 0.0):.6f} total
  Calls: {provider_data.get('calls', 0)}
"""
            
            # Add model breakdown
            if 'by_model' in token_usage and token_usage['by_model']:
                token_usage_section += f"\nModel Breakdown:\n"
                for model, model_data in token_usage['by_model'].items():
                    api_provider = model_data.get('api_provider', 'unknown')
                    token_usage_section += f"- {model} ({api_provider}): {model_data.get('total_tokens', 0):,} tokens, ${model_data.get('total_cost', 0.0):.6f}, {model_data.get('calls', 0)} calls\n"
        
        summary_report = f"""
Enhanced Validation Results Summary
==================================

Session ID: {session_id}{pin_section}
Generated: {timestamp}
Total Rows Processed: {total_rows}

Fields Validated: {', '.join(results_data['summary']['fields_validated'])}

Confidence Distribution:
- HIGH: {results_data['summary']['confidence_distribution']['HIGH']} fields
- MEDIUM: {results_data['summary']['confidence_distribution']['MEDIUM']} fields  
- LOW: {results_data['summary']['confidence_distribution']['LOW']} fields{token_usage_section}

Files Included:
==============
- validation_results_enhanced.xlsx : Color-coded Excel with comments and multiple sheets
- validation_results.json          : Raw validation data in JSON format
- validation_results.csv           : Simple CSV format for basic analysis
- SUMMARY.txt                      : This summary file
- COMPLETED.txt                    : Processing completion marker
- original_files/{input_filename}  : Original input file
- original_files/{config_filename} : Original configuration file

Enhanced Excel Features:
=======================
✅ Color-coded confidence levels:
   - GREEN: HIGH confidence
   - YELLOW: MEDIUM confidence  
   - RED: LOW confidence
✅ Cell comments with quotes and sources
✅ Multiple worksheets: Results, Details, Reasons
✅ Comprehensive validation tracking
✅ Professional formatting with xlsxwriter

Processing Details:
==================
{json.dumps(validation_results, indent=2) if validation_results else 'No validation results'}
        """.strip()
        
        zip_file.writestr('SUMMARY.txt', summary_report)
        
        # Add completion marker
        zip_file.writestr('COMPLETED.txt', f'Enhanced validation completed at {timestamp}')
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

def upload_file_to_s3(file_content, bucket, key, content_type='application/octet-stream'):
    """Upload file content to S3."""
    try:
        s3_client.put_object(
            Bucket=bucket,
            Key=key,
            Body=file_content,
            ContentType=content_type
        )
        logger.info(f"Uploaded file to s3://{bucket}/{key}")
        return True
    except Exception as e:
        logger.error(f"Error uploading to S3: {str(e)}")
        return False

def generate_presigned_url(bucket, key, expiration=3600):
    """Generate a presigned URL for downloading a file from S3."""
    try:
        url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket, 'Key': key},
            ExpiresIn=expiration
        )
        return url
    except Exception as e:
        logger.error(f"Error generating presigned URL: {str(e)}")
        return None

def invoke_validator_lambda(excel_s3_key, config_s3_key, max_rows, batch_size, preview_first_row=False, preview_max_rows=5, sequential_call=None):
    """Invoke the core validator Lambda with Excel data."""
    logger.info(">>> ENTER invoke_validator_lambda <<<")
    logger.info(f">>> Parameters: excel_s3_key={excel_s3_key}, preview={preview_first_row}, sequential_call={sequential_call} <<<")
    
    try:
        logger.info(f"Starting invoke_validator_lambda - preview_first_row: {preview_first_row}")
        logger.info(f"Excel S3 key: {excel_s3_key}")
        logger.info(f"Config S3 key: {config_s3_key}")
        
        # Download Excel file from S3
        excel_response = s3_client.get_object(Bucket=S3_CACHE_BUCKET, Key=excel_s3_key)
        excel_content = excel_response['Body'].read()
        
        # Download config file from S3
        config_response = s3_client.get_object(Bucket=S3_CACHE_BUCKET, Key=config_s3_key)
        config_data = json.loads(config_response['Body'].read().decode('utf-8'))
        
        # Preprocess config to support both 'column' and 'name' fields
        if 'validation_targets' in config_data and isinstance(config_data['validation_targets'], list):
            for target in config_data['validation_targets']:
                if 'column' in target and 'name' not in target:
                    target['name'] = target['column']
                    # Keep 'column' for backward compatibility
        
        # Detect file type and load data accordingly
        # Check if this is a CSV file and convert to Excel format if needed
        original_file_type = "Excel"
        try:
            # Try to detect if this is a CSV file by attempting to decode as text
            text_content = excel_content.decode('utf-8')
            # Simple heuristic: if it contains commas and no Excel-specific markers, treat as CSV
            if ',' in text_content and not text_content.startswith(b'PK'.decode('utf-8')):
                original_file_type = "CSV"
                logger.info("Detected CSV file format - converting to Excel")
                
                # Parse CSV content
                csv_reader = csv.reader(io.StringIO(text_content))
                csv_rows = list(csv_reader)
                
                if not csv_rows:
                    raise ValueError("CSV file is empty")
                
                # Convert CSV to Excel format in memory
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "Data"
                
                # Write CSV data to Excel worksheet
                for row_idx, csv_row in enumerate(csv_rows, 1):
                    for col_idx, cell_value in enumerate(csv_row, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Save Excel workbook to bytes
                excel_buffer = io.BytesIO()
                workbook.save(excel_buffer)
                excel_content = excel_buffer.getvalue()
                excel_buffer.close()
                
                logger.info(f"Converted CSV with {len(csv_rows)} rows to Excel format")
            else:
                logger.info("Detected Excel file format")
        except UnicodeDecodeError:
            # If it can't be decoded as UTF-8, it's likely a binary Excel file
            logger.info("Detected binary Excel file format")
        
        # Now process as Excel file (whether original or converted from CSV)
        logger.info(f"Processing {original_file_type} file as Excel format")
        workbook = openpyxl.load_workbook(io.BytesIO(excel_content))
        
        # Log available sheet names
        logger.info(f"Available sheets in Excel: {workbook.sheetnames}")
        
        # Select the appropriate sheet
        if 'Results' in workbook.sheetnames:
            worksheet = workbook['Results']
            logger.info(f"Using 'Results' sheet as data source")
        elif len(workbook.sheetnames) > 0:
            worksheet = workbook[workbook.sheetnames[0]]
            logger.info(f"Using first sheet '{worksheet.title}' as data source")
        else:
            worksheet = workbook.active
            logger.info(f"Using active sheet: {worksheet.title}")
        
        # Read headers from first row
        headers = []
        for cell in worksheet[1]:
            header = cell.value
            if header:
                headers.append(str(header).strip())
            else:
                headers.append('')
        
        logger.info(f"Headers found: {headers}")
        logger.info(f"Total columns: {len(headers)}")
        
        # Process data rows
        rows = []
        total_rows = worksheet.max_row - 1  # Exclude header
        
        # Limit rows for preview mode or max_rows
        if preview_first_row:
            # For preview mode, we want to load all rows up to preview_max_rows
            # so the validator can process them sequentially and use caching
            max_process_rows = min(preview_max_rows, total_rows)
        else:
            max_process_rows = min(max_rows, total_rows) if max_rows else total_rows
        
        logger.info(f"Processing {max_process_rows} rows from {original_file_type} (total available: {total_rows})")
        
        # Get ID fields from config for row key generation
        id_fields = []
        for target in config_data.get('validation_targets', []):
            if target.get('importance', '').upper() == 'ID':
                # Support both 'name' and 'column' fields
                field_name = target.get('name') or target.get('column')
                if field_name:
                    id_fields.append(field_name)
        
        # If no ID fields found, try to use SimplifiedSchemaValidator to determine primary keys
        if not id_fields:
            try:
                from schema_validator_simplified import SimplifiedSchemaValidator
                validator = SimplifiedSchemaValidator(config_data)
                id_fields = validator.primary_key
                logger.info(f"Using primary keys from SimplifiedSchemaValidator: {id_fields}")
            except ImportError:
                logger.warning("SimplifiedSchemaValidator not available in deployment package")
            except Exception as e:
                logger.warning(f"Could not use SimplifiedSchemaValidator: {e}")
        
        logger.info(f"ID fields for row key generation: {id_fields}")
        
        # Process data rows using Excel format (works for both original Excel and converted CSV)
        for row_idx in range(2, min(2 + max_process_rows, worksheet.max_row + 1)):
            row_data = {}
            
            # Extract cell values
            for col_idx, header in enumerate(headers):
                if header:  # Skip empty headers
                    cell_value = worksheet.cell(row=row_idx, column=col_idx + 1).value
                    row_data[header] = str(cell_value) if cell_value is not None else ""
            
            # Debug: Log first row's data to see what we're getting
            if row_idx == 2:  # First data row
                logger.info(f"First row data extracted: {json.dumps(row_data)}")
                logger.info(f"Looking for ID fields: {id_fields}")
                for id_field in id_fields:
                    logger.info(f"  {id_field}: '{row_data.get(id_field, 'NOT FOUND')}'")
            
            # Generate row key using row_key_utils if available
            if ROW_KEY_UTILS_AVAILABLE and id_fields:
                row_key = generate_row_key(row_data, id_fields)
                logger.debug(f"Generated row key using row_key_utils: {row_key}")
            else:
                # Fallback to simple join if row_key_utils not available
                key_columns = id_fields if id_fields else headers[:3]
                row_key = "||".join([row_data.get(col, "") for col in key_columns])
                logger.debug(f"Generated row key using fallback method: {row_key}")
            
            row_data['_row_key'] = row_key
            rows.append(row_data)
        
        # Show sample row data to understand what keys will be generated
        if rows and len(rows) > 0:
            sample_row = rows[0]
            logger.info("Sample row data for first row:")
            for id_field in id_fields:
                value = sample_row.get(id_field, 'NOT FOUND')
                logger.info(f"  {id_field}: {value}")
        
        # Load validation history if available and we have Excel content
        validation_history = {}
        if not preview_first_row:  # Load validation history for all files (now all are Excel format)
            try:
                # Save Excel content to a temporary file for history extraction
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                    tmp_file.write(excel_content)
                    tmp_file_path = tmp_file.name
                
                # Use the appropriate validation history loader
                if VALIDATION_HISTORY_AVAILABLE:
                    # Use the imported function if available
                    logger.info("Using imported validation history loader (pandas-based)")
                    
                    # Before loading validation history, update the load function to use our ID fields
                    # This is a bit of a hack but ensures consistency
                    import lambda_test_json_clean
                    original_load_func = lambda_test_json_clean.load_validation_history_from_details
                    
                    def custom_load_validation_history_from_details(excel_path):
                        """Custom loader that uses our current ID fields"""
                        result = original_load_func(excel_path)
                        
                        # If the Details sheet has ID columns, the function will handle it
                        # If not, we need to check if we can remap the keys
                        if result and id_fields:
                            # Log what we're working with
                            logger.info(f"Validation history loaded with {len(result)} entries")
                            logger.info(f"Current ID fields for remapping: {id_fields}")
                        
                        return result
                    
                    # Temporarily replace the function
                    lambda_test_json_clean.load_validation_history_from_details = custom_load_validation_history_from_details
                
                    # Load validation history from Excel
                    original_validation_history = load_validation_history_from_excel(tmp_file_path)
                    
                    # Restore original function
                    lambda_test_json_clean.load_validation_history_from_details = original_load_func
                else:
                    # Use the fallback function defined in this file
                    logger.info("Using fallback validation history loader (openpyxl-based)")
                    original_validation_history = load_validation_history_from_excel(tmp_file_path)
                
                logger.info(f"Loaded validation history for {len(original_validation_history)} row keys from Excel")
                
                if original_validation_history:
                    # Log sample history keys for debugging
                    history_keys = list(original_validation_history.keys())[:3]
                    logger.info(f"Sample validation history keys from Excel: {history_keys}")
                    
                    # Since the new Details format includes ID fields, the keys should now match
                    # For old format files, we still try direct matching
                    validation_history = original_validation_history
                    
                    # Log matching summary
                    matched_count = 0
                    for row in rows:
                        payload_key = row.get('_row_key', '')
                        if payload_key and payload_key in validation_history:
                            matched_count += 1
                    
                    logger.info(f"Matched {matched_count} out of {len(rows)} rows with validation history")
                    
                    if matched_count == 0 and len(validation_history) > 0:
                        logger.warning("No rows matched with validation history - may be using different primary keys")
                        logger.info(f"Current primary keys: {id_fields}")
                
                # Clean up temp file
                os.unlink(tmp_file_path)
                
            except Exception as e:
                logger.warning(f"Could not load validation history: {e}")
                import traceback
                traceback.print_exc()
                validation_history = {}
        
        # Initialize aggregated metadata
        aggregated_metadata = {
            'total_rows': total_rows,
            'token_usage': {
                'total_tokens': 0,
                'total_cost': 0.0,
                'api_calls': 0,
                'cached_calls': 0,
                'by_provider': {
                    'perplexity': {
                        'prompt_tokens': 0,
                        'completion_tokens': 0,
                        'total_tokens': 0,
                        'calls': 0,
                        'input_cost': 0.0,
                        'output_cost': 0.0,
                        'total_cost': 0.0
                    },
                    'anthropic': {
                        'input_tokens': 0,
                        'output_tokens': 0,
                        'cache_creation_tokens': 0,
                        'cache_read_tokens': 0,
                        'total_tokens': 0,
                        'calls': 0,
                        'input_cost': 0.0,
                        'output_cost': 0.0,
                        'total_cost': 0.0
                    }
                },
                'by_model': {}
            }
        }
        
        # For sequential calls, we'll use validator response metadata instead of trying to predict cache hits

        # For preview mode, process rows sequentially
        if preview_first_row:
            # Intelligent sequential processing: send up to 5 rows, let validator find next non-cached row
            # This eliminates the need for users to specify --sequential-call numbers
            
            if sequential_call is not None:
                # Legacy support: user specified sequential call number
                total_rows_to_send = min(sequential_call, preview_max_rows, len(rows))
                logger.info(f"Legacy sequential call #{sequential_call}: sending {total_rows_to_send} rows to validator")
            else:
                # Intelligent mode: send up to preview_max_rows, let validator handle sequential processing
                total_rows_to_send = min(preview_max_rows, len(rows))
                logger.info(f"Intelligent preview mode: sending {total_rows_to_send} rows to validator for sequential processing")
                logger.info(f"Validator will use cached results and process next non-cached row")
            
            preview_rows = rows[:total_rows_to_send]
            
            # Create payload - let validator check its own cache
            payload = {
                "test_mode": True,
                "config": config_data,
                "validation_data": {
                    "rows": preview_rows
                },
                "validation_history": {}  # Empty - relying on validator's own cache checking
            }
            
            logger.info(f"Created payload with {len(preview_rows)} rows for preview validation")
            
            # LOG PAYLOAD INFO
            logger.info("=== PAYLOAD DEBUG INFO ===")
            logger.info(f"Sending {len(preview_rows)} rows to validator for cache checking")
            if sequential_call:
                logger.info(f"Sequential call #{sequential_call}: validator will check its own cache")
            
            # Log row keys being sent
            if preview_rows:
                logger.info(f"Row keys in payload:")
                for i, row in enumerate(preview_rows):
                    logger.info(f"  Row {i+1}: {row.get('_row_key', 'NO KEY')}")
            
            logger.info("=== END PAYLOAD DEBUG ===")
            
            try:
                # Track actual validation processing time
                validation_start_time = time.time()
                
                # Use synchronous invocation with monitoring for timeout
                response = lambda_client.invoke(
                    FunctionName=VALIDATOR_LAMBDA_NAME,
                    InvocationType='RequestResponse',
                    Payload=json.dumps(payload)
                )
                
                validation_processing_time = time.time() - validation_start_time
                logger.info(f"Validator Lambda processing took {validation_processing_time:.2f} seconds")
                
                response_payload = json.loads(response['Payload'].read().decode('utf-8'))
                logger.info(f"Validator response keys: {list(response_payload.keys()) if isinstance(response_payload, dict) else 'Not a dict'}")
                
                # Parse the actual validator response structure
                validation_results = None
                metadata = None
                if isinstance(response_payload, dict):
                    if 'body' in response_payload and isinstance(response_payload['body'], dict):
                        body = response_payload['body']
                        if 'data' in body and isinstance(body['data'], dict):
                            data = body['data']
                            if 'rows' in data and isinstance(data['rows'], dict):
                                validation_results = data['rows']
                                logger.info(f"Found validation results with {len(validation_results)} rows")
                        # Extract metadata including token usage
                        if 'metadata' in body:
                            metadata = body['metadata']
                            logger.info(f"Found metadata: {list(metadata.keys()) if isinstance(metadata, dict) else 'Not a dict'}")
                            # Merge ALL metadata fields, not just token_usage
                            for key, value in metadata.items():
                                if key == 'token_usage':
                                    # Special handling for token_usage to merge with aggregated
                                    aggregated_metadata['token_usage'] = value
                                    logger.info(f"Token usage - Total tokens: {value.get('total_tokens', 0)}, Total cost: ${value.get('total_cost', 0.0):.6f}")
                                else:
                                    # Copy other metadata fields directly
                                    aggregated_metadata[key] = value
                                    logger.info(f"Metadata {key}: {value}")
                    elif 'validation_results' in response_payload:
                        validation_results = response_payload['validation_results']
                        logger.info(f"Found direct validation_results")
                
                # Log the validation invocation time for debugging
                logger.info(f"Interface Lambda invocation time: {validation_processing_time:.2f}s")
                
                # Analyze cache pattern and determine which row was newly processed
                if metadata and 'token_usage' in metadata:
                    token_usage = metadata['token_usage']
                    total_api_calls = token_usage.get('api_calls', 0)
                    total_cached_calls = token_usage.get('cached_calls', 0)
                    
                    logger.info(f"Preview validation cache analysis:")
                    logger.info(f"  API calls: {total_api_calls}, Cached calls: {total_cached_calls}")
                    
                    # Determine which row was newly processed based on validation results
                    new_row_processed = None
                    if validation_results:
                        # Find the highest numbered row that has non-cached responses
                        for row_key in sorted(validation_results.keys(), key=lambda x: int(x) if x.isdigit() else -1, reverse=True):
                            row_data = validation_results[row_key]
                            if isinstance(row_data, dict) and '_raw_responses' in row_data:
                                # Check if this row has any non-cached responses
                                has_new_data = False
                                for resp_id, resp_data in row_data['_raw_responses'].items():
                                    if not resp_data.get('is_cached', True):
                                        has_new_data = True
                                        break
                                if has_new_data:
                                    new_row_processed = int(row_key) + 1  # Convert to 1-based
                                    break
                    
                    # If no new data found but we have API calls, assume first row was processed
                    if new_row_processed is None and total_api_calls > 0:
                        new_row_processed = 1
                    
                    logger.info(f"Identified newly processed row: {new_row_processed}")
                    
                    # Only validate cache pattern for legacy sequential calls
                    if sequential_call is not None:
                        logger.info(f"Legacy sequential call #{sequential_call} - validating expected pattern")
                        # Keep existing validation for backward compatibility
                        if sequential_call == 1:
                            cache_pattern_valid = (total_api_calls > 0)
                        else:
                            cache_pattern_valid = (total_api_calls > 0 and total_cached_calls > 0)
                        
                        if not cache_pattern_valid:
                            error_msg = f"Sequential call #{sequential_call} failed: Unexpected cache pattern (API: {total_api_calls}, Cached: {total_cached_calls})"
                            logger.error(error_msg)
                            return {
                                'status': 'cache_validation_failed',
                                'error': error_msg,
                                'total_rows': total_rows,
                                'sequential_call': sequential_call,
                                'actual_api_calls': total_api_calls,
                                'actual_cached_calls': total_cached_calls
                            }
                else:
                    # No metadata available
                    pass
                
                # For preview mode, return ALL validation results (not just "new" ones)
                # The preview should show validation results regardless of caching status
                total_processed_rows = len(validation_results) if validation_results else 0
                new_row_processed = 1 if total_processed_rows > 0 else None
                
                # Determine preview completion: we're done when we've processed 5 rows OR reached end of file OR no new row was processed
                preview_complete = (
                    new_row_processed is None or  # No new row processed (all cached)
                    new_row_processed >= preview_max_rows or  # Reached preview limit (5 rows)
                    new_row_processed >= len(rows)  # Reached end of file
                )
                
                logger.info(f"Preview mode: returning {total_processed_rows} validation results")
                logger.info(f"New row processed: {new_row_processed}, Preview complete: {preview_complete}")
                if validation_results:
                    logger.info(f"Validation result keys: {list(validation_results.keys())}")
                    for key, value in validation_results.items():
                        logger.info(f"  {key}: {type(value)} with {len(value) if isinstance(value, dict) else 0} fields")
                
                # Add total_rows info to response  
                result_response = {
                    'total_rows': total_rows,
                    'validation_results': validation_results,  # Return ALL results for preview
                    'metadata': aggregated_metadata,
                    'total_processed_rows': total_processed_rows,
                    'new_row_number': new_row_processed,
                    'preview_complete': preview_complete
                }
                
                if isinstance(response_payload, dict):
                    result_response.update(response_payload)
                
                return result_response
                
            except Exception as e:
                logger.warning(f"Validator Lambda timeout or error in preview mode: {str(e)}")
                return {
                    'status': 'timeout',
                    'total_rows': total_rows,
                    'validation_results': None,
                    'metadata': aggregated_metadata,
                    'note': f'Validation timed out (>25s), using demo data. Error: {str(e)}'
                }
        else:
            # For normal mode, process in batches
            all_validation_results = {}
            total_batches = (len(rows) + batch_size - 1) // batch_size
            
            logger.info(f"Processing {len(rows)} rows in {total_batches} batches of size {batch_size}")
            
            for batch_num in range(total_batches):
                start_idx = batch_num * batch_size
                end_idx = min(start_idx + batch_size, len(rows))
                batch_rows = rows[start_idx:end_idx]
                
                logger.info(f"Processing batch {batch_num+1}/{total_batches}, rows {start_idx+1}-{end_idx}")
                
                # Create payload for this batch
                batch_payload = {
                    "test_mode": False,
                    "config": config_data,
                    "validation_data": {
                        "rows": batch_rows
                    },
                    "validation_history": validation_history
                }
                
                # LOG BATCH PAYLOAD DEBUG INFO
                if batch_num == 0:  # Only log for first batch to avoid spam
                    logger.info("=== BATCH PAYLOAD DEBUG INFO ===")
                    logger.info(f"Batch payload has validation_history: {'validation_history' in batch_payload}")
                    logger.info(f"Validation history size: {len(validation_history)}")
                    logger.info(f"Batch rows count: {len(batch_rows)}")
                    
                    # Log sample row keys from batch
                    if batch_rows:
                        logger.info(f"Sample row keys from batch:")
                        for i, row in enumerate(batch_rows[:2]):
                            logger.info(f"  Row {i}: {row.get('_row_key', 'NO KEY')}")
                    
                    # Check if any batch rows match validation history
                    matches = 0
                    for row in batch_rows:
                        if row.get('_row_key') in validation_history:
                            matches += 1
                    logger.info(f"Rows with matching validation history: {matches}/{len(batch_rows)}")
                    logger.info("=== END BATCH PAYLOAD DEBUG ===")
                
                try:
                    # Invoke validator for this batch
                    response = lambda_client.invoke(
                        FunctionName=VALIDATOR_LAMBDA_NAME,
                        InvocationType='RequestResponse',
                        Payload=json.dumps(batch_payload)
                    )
                    
                    response_payload = json.loads(response['Payload'].read().decode('utf-8'))
                    logger.info(f"Batch {batch_num+1} validator response keys: {list(response_payload.keys()) if isinstance(response_payload, dict) else 'Not a dict'}")
                    
                    # Parse the batch validation results
                    batch_validation_results = None
                    batch_metadata = None
                    if isinstance(response_payload, dict):
                        if 'body' in response_payload and isinstance(response_payload['body'], dict):
                            body = response_payload['body']
                            if 'data' in body and isinstance(body['data'], dict):
                                data = body['data']
                                if 'rows' in data and isinstance(data['rows'], dict):
                                    batch_validation_results = data['rows']
                                    logger.info(f"Found validation results with {len(batch_validation_results)} rows in batch {batch_num+1}")
                            # Extract metadata including token usage from this batch
                            if 'metadata' in body:
                                batch_metadata = body['metadata']
                                if 'token_usage' in batch_metadata:
                                    batch_token_usage = batch_metadata['token_usage']
                                    # Aggregate token usage across batches
                                    agg_token_usage = aggregated_metadata['token_usage']
                                    agg_token_usage['total_tokens'] += batch_token_usage.get('total_tokens', 0)
                                    agg_token_usage['total_cost'] += batch_token_usage.get('total_cost', 0.0)
                                    agg_token_usage['api_calls'] += batch_token_usage.get('api_calls', 0)
                                    agg_token_usage['cached_calls'] += batch_token_usage.get('cached_calls', 0)
                                    
                                    # Aggregate by provider
                                    if 'by_provider' in batch_token_usage:
                                        for provider, provider_data in batch_token_usage['by_provider'].items():
                                            if provider in agg_token_usage['by_provider']:
                                                agg_provider = agg_token_usage['by_provider'][provider]
                                                for key, value in provider_data.items():
                                                    if isinstance(value, (int, float)):
                                                        agg_provider[key] = agg_provider.get(key, 0) + value
                                    
                                    # Aggregate by model
                                    if 'by_model' in batch_token_usage:
                                        for model, model_data in batch_token_usage['by_model'].items():
                                            if model not in agg_token_usage['by_model']:
                                                agg_token_usage['by_model'][model] = {}
                                            agg_model = agg_token_usage['by_model'][model]
                                            for key, value in model_data.items():
                                                if isinstance(value, (int, float)):
                                                    agg_model[key] = agg_model.get(key, 0) + value
                                                else:
                                                    agg_model[key] = value  # Non-numeric values like api_provider
                                    
                                    logger.info(f"Batch {batch_num+1} token usage - Tokens: {batch_token_usage.get('total_tokens', 0)}, Cost: ${batch_token_usage.get('total_cost', 0.0):.6f}")
                        elif 'validation_results' in response_payload:
                            batch_validation_results = response_payload['validation_results']
                            logger.info(f"Found direct validation_results in batch {batch_num+1}")
                    
                    # Merge batch results into all_validation_results
                    if batch_validation_results:
                        # Map numeric keys from batch to overall row indices
                        for batch_key, batch_result in batch_validation_results.items():
                            # Convert batch key to overall row index
                            if batch_key.isdigit():
                                overall_idx = start_idx + int(batch_key)
                                all_validation_results[str(overall_idx)] = batch_result
                            else:
                                # Use the key as-is
                                all_validation_results[batch_key] = batch_result
                    
                except Exception as e:
                    logger.error(f"Error processing batch {batch_num+1}: {str(e)}")
                    # Continue with next batch
                
                # Add a small delay between batches to avoid rate limiting
                if batch_num < total_batches - 1:
                    time.sleep(0.5)
            
            logger.info(f"Completed processing all batches. Total results: {len(all_validation_results)}")
            logger.info(f"Total aggregated token usage - Tokens: {aggregated_metadata['token_usage']['total_tokens']}, Cost: ${aggregated_metadata['token_usage']['total_cost']:.6f}")
            
            # Return combined results
            return {
                'total_rows': total_rows,
                'validation_results': all_validation_results,
                'metadata': aggregated_metadata,
                'status': 'completed'
            }
            
    except Exception as e:
        logger.error(f"Error invoking validator Lambda: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def create_markdown_table_from_results(validation_results, preview_row_count=3, config_s3_key=None):
    """Convert real validation results from validator Lambda to markdown table format.
    Always shows the first 3 rows in a transposed table format with confidence emojis.
    Fields are ordered by: ID fields first, then by search_group ascending, with a Search Group column at the end.
    ID fields show blue circles (🔵) with their actual values from the original Excel data.
    """
    if not validation_results:
        return "No validation results available."
    
    # Get confidence emoji mapping
    def get_confidence_emoji(confidence_level):
        confidence_map = {
            'HIGH': '🟢',
            'MEDIUM': '🟡', 
            'LOW': '🔴',
            'ID': '🔵',  # Blue circle for ID fields
            'UNKNOWN': '❓'
        }
        return confidence_map.get(confidence_level, '❓')
    
    # Load config to get field information and ordering
    field_config_map = {}
    all_config_fields = []  # Preserve original order from config
    
    if config_s3_key:
        try:
            config_response = s3_client.get_object(Bucket=S3_CACHE_BUCKET, Key=config_s3_key)
            config_content = json.loads(config_response['Body'].read().decode('utf-8'))
            
            for target in config_content.get('validation_targets', []):
                field_name = target.get('column')
                if field_name:
                    field_config_map[field_name] = target
                    all_config_fields.append(field_name)  # Preserve order from config
        except Exception as e:
            print(f"Warning: Could not load config from {config_s3_key}: {e}")
    

    
    # Get first few rows based on preview_row_count
    sorted_row_keys = list(validation_results.keys())[:preview_row_count]
    
    # Get all field names from validation results
    validation_field_names = set()
    for row_key in sorted_row_keys:
        if isinstance(validation_results[row_key], dict):
            validation_field_names.update(validation_results[row_key].keys())
    
    # Create comprehensive field list: all config fields (in order), plus any extra validation fields
    field_keys = []
    
    # First, add all config fields in their original order
    for field_name in all_config_fields:
        if field_name not in field_keys:  # Avoid duplicates
            field_keys.append(field_name)
    
    # Then add any validation fields not in config
    for field_name in validation_field_names:
        if field_name not in field_keys:
            field_keys.append(field_name)
    
    # Sort function: ID fields first (by search group), then other fields by search group
    # Within each group, preserve original order from config
    def get_field_sort_key(field_name):
        config = field_config_map.get(field_name, {})
        importance = config.get('importance', 'MEDIUM')
        search_group = config.get('search_group', 999)
        
        # Get original position in config for ordering within groups
        try:
            config_position = all_config_fields.index(field_name)
        except ValueError:
            config_position = 9999  # Put fields not in config at the end
        
        # ID fields come first (sort key 0), then by search group
        if importance == 'ID':
            return (0, search_group, config_position)
        else:
            return (1, search_group, config_position)
    
    # Sort field keys
    field_keys.sort(key=get_field_sort_key)
    
    # Create transposed table: fields as rows, data rows as columns
    table_lines = []
    
    # Add legend as a separate sentence before the table
    legend = "**Confidence Legend:** 🟢 High • 🟡 Medium • 🔴 Low • 🔵 ID/Input\n\n"
    
    # Create header row - add Search Group column
    header = "| Field"
    for i, row_key in enumerate(sorted_row_keys):
        row_number = i + 1
        header += f" | Row {row_number}"
    header += " | Search Group |"
    table_lines.append(header)
    
    # Create separator row
    separator = "|" + "-" * 26
    for _ in sorted_row_keys:
        separator += "|" + "-" * 31
    separator += "|" + "-" * 14 + "|"  # Search Group column
    table_lines.append(separator)
    
    # Create data rows (one for each field)
    for field_name in field_keys:
        # Truncate long field names
        if len(field_name) > 24:
            display_field = field_name[:21] + "..."
        else:
            display_field = field_name
        
        # Escape pipe characters in field name
        display_field = display_field.replace('|', '\\|')
        
        row_line = f"| {display_field:<24}"
        
        # Check if this is an ID field
        config = field_config_map.get(field_name, {})
        is_id_field = config.get('importance') == 'ID'
        
        # Add values for each data row
        for row_key in sorted_row_keys:
            row_results = validation_results[row_key]
            field_result = row_results.get(field_name, {})
            
            if isinstance(field_result, dict) and field_result:
                # Field was validated or is an ID field - use result
                confidence = field_result.get('confidence_level', 'UNKNOWN')
                value = field_result.get('value', '')
                
                # Get confidence emoji
                emoji = get_confidence_emoji(confidence)
                
                # Prepare value with emoji prefix
                if value:
                    display_value = f"{emoji} {value}"
                else:
                    display_value = f"{emoji} (empty)"
                
            else:
                # Field not in validation results - show N/A
                display_value = "N/A"
            
            # Truncate long values
            if len(str(display_value)) > 29:
                display_value = str(display_value)[:26] + "..."
            
            # Escape pipe characters
            display_value = display_value.replace('|', '\\|')
            
            row_line += f" | {display_value:<29}"
        
        # Add search group information
        search_group = config.get('search_group', '-')
        row_line += f" | {str(search_group):<12} |"
        
        table_lines.append(row_line)
    
    # Return legend + table
    return legend + '\n'.join(table_lines)

def create_markdown_table(validation_results):
    """Convert validation results to markdown table format (legacy function for backwards compatibility)."""
    if not validation_results or 'results' not in validation_results:
        return "No validation results available."
    
    results = validation_results['results']
    if not results:
        return "No validation results found."
    
    # Create markdown table
    table_lines = [
        "| Field   | Confidence | Value                     |",
        "|---------|------------|--------------------------|"
    ]
    
    for result in results:
        field = result.get('field', 'Unknown')
        confidence = result.get('confidence', 'UNKNOWN')
        value = result.get('value', '')
        
        # Truncate long values
        if len(str(value)) > 25:
            value = str(value)[:22] + "..."
        
        # Escape any pipe characters in the value
        value = str(value).replace('|', '\\|')
        
        table_lines.append(f"| {field:<7} | {confidence:<10} | {value:<25} |")
    
    return '\n'.join(table_lines)

def handle_status_check_request(request_data, context):
    """Handle status check requests via JSON POST body"""
    try:
        session_id = request_data.get('session_id')
        is_preview = request_data.get('preview_mode', False)
        
        logger.info(f"Status check for session: {session_id}, preview: {is_preview}")
        
        if is_preview:
            # Parse session ID to get timestamp and reference pin
            parts = session_id.split('_')
            if len(parts) >= 4 and parts[-1] == 'preview':
                # Session ID format: {date}_{time}_{reference_pin}_preview
                timestamp = f"{parts[0]}_{parts[1]}"  # Combine date and time
                reference_pin = parts[2]
                
                # Extract email folder from the session - for now use a broader search pattern
                # Since we don't have email info in status check, we'll search for the file
                email_folder = "default"  # fallback
                
                # Try multiple possible email folder patterns
                possible_keys = []
                
                # If email is provided, construct the exact path
                email_param = request_data.get('email', '')
                if email_param:
                    email_folder = create_email_folder_path(email_param)
                    possible_keys.append(f"preview_results/{email_folder}/{session_id}.json")
                    logger.info(f"Using email parameter for preview path: preview_results/{email_folder}/{session_id}.json")
                
                # Fallback patterns
                possible_keys.extend([
                    f"preview_results/default/{session_id}.json",
                    f"preview_results/{session_id}.json",
                    # Also try old format for backward compatibility
                    f"preview_results/default/{timestamp}_{reference_pin}_preview.json",
                    f"preview_results/{timestamp}_{reference_pin}_preview.json",  # Direct path
                ])
                
                # Also try to list all keys to see what's actually there
                logger.info(f"Searching for preview results with timestamp: {timestamp}, reference_pin: {reference_pin}")
                try:
                    list_response = s3_client.list_objects_v2(
                        Bucket=S3_RESULTS_BUCKET,
                        Prefix=f"preview_results/",
                        MaxKeys=50
                    )
                    if 'Contents' in list_response:
                        logger.info(f"Found {len(list_response['Contents'])} preview result files:")
                        for obj in list_response['Contents']:
                            logger.info(f"  - {obj['Key']}")
                            if f"{timestamp}_{reference_pin}" in obj['Key']:
                                logger.info(f"  ⭐ MATCH: {obj['Key']}")
                                possible_keys.insert(0, obj['Key'])  # Try this first
                    else:
                        logger.info("No preview result files found in S3")
                except Exception as e:
                    logger.warning(f"Failed to list S3 objects: {e}")
                
                preview_results_key = None
                preview_results = None
                
                # Try each possible key
                for key in possible_keys:
                    try:
                        results_response = s3_client.get_object(Bucket=S3_RESULTS_BUCKET, Key=key)
                        preview_results = json.loads(results_response['Body'].read().decode('utf-8'))
                        preview_results_key = key
                        logger.info(f"Found preview results at: {key}")
                        break
                    except s3_client.exceptions.NoSuchKey:
                        continue
                    except Exception as e:
                        logger.warning(f"Error checking key {key}: {e}")
                        continue
                
                if preview_results:
                    # Return the complete preview results
                    return {
                        'statusCode': 200,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps(preview_results)
                    }
                else:
                    # Preview still processing
                    return {
                        'statusCode': 200,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps({
                            'session_id': session_id,
                            'status': 'processing',
                            'preview_mode': True,
                            'message': 'Preview validation is still in progress'
                        })
                    }
            else:
                return {
                    'statusCode': 400,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({'error': 'Invalid preview session ID format'})
                }
        else:
            # Normal mode status check - check for results ZIP file
            parts = session_id.split('_')
            if len(parts) >= 2:
                timestamp = parts[0]
                reference_pin = parts[1]
                
                email_folder = "default"  # Could be enhanced
                results_key = f"results/{email_folder}/{timestamp}_{reference_pin}.zip"
                
                try:
                    s3_client.head_object(Bucket=S3_RESULTS_BUCKET, Key=results_key)
                    # File exists - processing complete
                    download_url = generate_presigned_url(S3_RESULTS_BUCKET, results_key, 86400)
                    
                    return {
                        'statusCode': 200,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps({
                            'session_id': session_id,
                            'status': 'completed',
                            'download_url': download_url,
                            'message': 'Processing completed successfully'
                        })
                    }
                except s3_client.exceptions.NoSuchKey:
                    # File doesn't exist - still processing
                    return {
                        'statusCode': 200,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps({
                            'session_id': session_id,
                            'status': 'processing',
                            'message': 'Validation is still in progress'
                        })
                    }
            else:
                return {
                    'statusCode': 400,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({'error': 'Invalid session ID format'})
                }
        
    except Exception as e:
        logger.error(f"Error in status check handler: {str(e)}")
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': 'Internal server error'})
        }

def handle_status_request(event, context):
    """Handle GET requests to check processing status."""
    try:
        # Extract session ID from path: /status/{session_id}
        path_parts = event.get('path', '').split('/')
        if len(path_parts) < 3:
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps({'error': 'Invalid status request path'})
            }
        
        session_id = path_parts[2]
        logger.info(f"Status check for session: {session_id}")
        
        # Check if this is a preview status request
        query_params = event.get('queryStringParameters') or {}
        is_preview = query_params.get('preview', 'false').lower() == 'true'
        email_param = query_params.get('email', '')
        
        if is_preview:
            # For preview mode, check S3 for stored results using same pattern as normal mode
            # Parse session ID to extract timestamp and reference pin
            parts = session_id.split('_')
            if len(parts) >= 4 and parts[-1] == 'preview':
                # Session ID format: {date}_{time}_{reference_pin}_preview
                timestamp = f"{parts[0]}_{parts[1]}"  # Combine date and time
                reference_pin = parts[2]
                
                # Try multiple possible email folder patterns
                possible_keys = []
                
                # If email is provided, construct the exact path
                if email_param:
                    email_folder = create_email_folder_path(email_param)
                    possible_keys.append(f"preview_results/{email_folder}/{session_id}.json")
                    logger.info(f"Using email parameter for preview path: preview_results/{email_folder}/{session_id}.json")
                
                # Fallback patterns
                possible_keys.extend([
                    f"preview_results/default/{session_id}.json",
                    f"preview_results/{session_id}.json",
                    # Also try old format for backward compatibility
                    f"preview_results/default/{timestamp}_{reference_pin}_preview.json",
                    f"preview_results/{timestamp}_{reference_pin}_preview.json",  # Direct path
                ])
                
                # Also try to list all keys to see what's actually there
                logger.info(f"GET Status: Searching for preview results with timestamp: {timestamp}, reference_pin: {reference_pin}")
                try:
                    list_response = s3_client.list_objects_v2(
                        Bucket=S3_RESULTS_BUCKET,
                        Prefix=f"preview_results/",
                        MaxKeys=50
                    )
                    if 'Contents' in list_response:
                        logger.info(f"Found {len(list_response['Contents'])} preview result files:")
                        for obj in list_response['Contents']:
                            logger.info(f"  - {obj['Key']}")
                            if f"{timestamp}_{reference_pin}" in obj['Key']:
                                logger.info(f"  ⭐ MATCH: {obj['Key']}")
                                possible_keys.insert(0, obj['Key'])  # Try this first
                    else:
                        logger.info("No preview result files found in S3")
                except Exception as e:
                    logger.warning(f"Failed to list S3 objects: {e}")
                
                preview_results = None
                
                # Try each possible key
                for key in possible_keys:
                    try:
                        results_response = s3_client.get_object(Bucket=S3_RESULTS_BUCKET, Key=key)
                        preview_results = json.loads(results_response['Body'].read().decode('utf-8'))
                        logger.info(f"Found preview results at: {key}")
                        break
                    except s3_client.exceptions.NoSuchKey:
                        continue
                    except Exception as e:
                        logger.warning(f"Error checking key {key}: {e}")
                        continue
                
                if preview_results:
                    # Return the complete preview results
                    return {
                        'statusCode': 200,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps(preview_results)
                    }
                else:
                    # Preview still processing
                    return {
                        'statusCode': 200,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps({
                            'session_id': session_id,
                            'status': 'processing',
                            'preview_mode': True,
                            'message': 'Preview validation is still in progress'
                        })
                    }
            else:
                return {
                    'statusCode': 400,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({'error': 'Invalid preview session ID format'})
                }
        else:
            # For normal mode, check S3 for results file
            try:
                # Parse session ID to extract timestamp and reference pin
                parts = session_id.split('_')
                if len(parts) >= 2:
                    # Handle different session ID formats
                    if len(parts) == 3:
                        # Format: YYYYMMDD_HHMMSS_PIN
                        timestamp = f"{parts[0]}_{parts[1]}"
                        reference_pin = parts[2]
                    else:
                        # Format: TIMESTAMP_PIN
                        timestamp = parts[0]
                        reference_pin = parts[1]
                    
                    # Check if results file exists in S3
                    possible_results_keys = []
                    
                    # If email is provided, we can construct the exact path
                    if email_param:
                        email_folder = create_email_folder_path(email_param)
                        # This should be the exact path
                        possible_results_keys.append(f"results/{email_folder}/{timestamp}_{reference_pin}.zip")
                        logger.info(f"Using email parameter to construct path: results/{email_folder}/{timestamp}_{reference_pin}.zip")
                    
                    # Fallback patterns if email not provided
                    possible_results_keys.extend([
                        f"results/{timestamp}_{reference_pin}.zip",  # Direct path (old format)
                        f"results/default/{timestamp}_{reference_pin}.zip",  # Default folder
                        f"results/status-check/{timestamp}_{reference_pin}.zip",  # Status check default
                    ])
                    
                    # Also try to list all results to find the right one
                    # Search with a more specific prefix to avoid hitting limits
                    search_prefixes = [
                        f"results/{timestamp[:8]}",  # Search by date prefix (YYYYMMDD)
                        f"results/",  # Fallback to general search
                    ]
                    
                    for search_prefix in search_prefixes:
                        try:
                            logger.info(f"Searching for results with prefix: {search_prefix}")
                            list_response = s3_client.list_objects_v2(
                                Bucket=S3_RESULTS_BUCKET,
                                Prefix=search_prefix,
                                MaxKeys=1000  # Increase limit
                            )
                            if 'Contents' in list_response:
                                logger.info(f"Found {len(list_response['Contents'])} objects with prefix {search_prefix}")
                                for obj in list_response['Contents']:
                                    if f"{timestamp}_{reference_pin}.zip" in obj['Key']:
                                        logger.info(f"Found matching results file: {obj['Key']}")
                                        possible_results_keys.insert(0, obj['Key'])  # Try this first
                                        break  # Found it, no need to continue
                                
                                # If we found a match, don't search other prefixes
                                if any(f"{timestamp}_{reference_pin}.zip" in key for key in possible_results_keys):
                                    break
                        except Exception as e:
                            logger.warning(f"Failed to list S3 results with prefix {search_prefix}: {e}")
                    
                    results_found = False
                    download_url = None
                    
                    # Try each possible key
                    for results_key in possible_results_keys:
                        try:
                            s3_client.head_object(Bucket=S3_RESULTS_BUCKET, Key=results_key)
                            # File exists - processing complete
                            download_url = generate_presigned_url(S3_RESULTS_BUCKET, results_key, 86400)
                            results_found = True
                            logger.info(f"Found results at: {results_key}")
                            break
                        except s3_client.exceptions.NoSuchKey:
                            continue
                        except Exception as e:
                            logger.warning(f"Error checking key {results_key}: {e}")
                            continue
                    
                    if results_found and download_url:
                        return {
                            'statusCode': 200,
                            'headers': {
                                'Content-Type': 'application/json',
                                'Access-Control-Allow-Origin': '*'
                            },
                            'body': json.dumps({
                                'session_id': session_id,
                                'status': 'completed',
                                'download_url': download_url,
                                'message': 'Processing completed successfully'
                            })
                        }
                    else:
                        # File doesn't exist - still processing
                        return {
                            'statusCode': 200,
                            'headers': {
                                'Content-Type': 'application/json',
                                'Access-Control-Allow-Origin': '*'
                            },
                            'body': json.dumps({
                                'session_id': session_id,
                                'status': 'processing',
                                'message': 'Validation is still in progress'
                            })
                        }
                else:
                    return {
                        'statusCode': 400,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps({'error': 'Invalid session ID format'})
                    }
                    
            except Exception as e:
                logger.error(f"Error checking status: {str(e)}")
                return {
                    'statusCode': 500,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({'error': 'Error checking status'})
                }
        
    except Exception as e:
        logger.error(f"Error in status handler: {str(e)}")
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': 'Internal server error'})
        }

def lambda_handler(event, context):
    """Handle HTTP requests to the interface Lambda."""
    logger.info("=== Interface Lambda Handler Start ===")
    logger.info(f"Event: {json.dumps(event, default=str)}")
    
    try:
        # Handle CORS preflight OPTIONS requests immediately
        if event.get('httpMethod') == 'OPTIONS':
            logger.info("Handling OPTIONS request for CORS")
            return {
                'statusCode': 200,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Access-Control-Allow-Methods': 'POST, GET, OPTIONS',
                    'Access-Control-Allow-Headers': 'Content-Type, Authorization'
                },
                'body': ''
            }
        
        # Check if this is an SQS event
        if 'Records' in event:
            logger.info("Detected SQS event")
            for record in event['Records']:
                if record.get('eventSource') == 'aws:sqs':
                    # Parse SQS message
                    try:
                        message_body = json.loads(record['body'])
                        logger.info(f"Processing SQS message: {json.dumps(message_body, default=str)}")
                        
                        # Transform SQS message to background processing format
                        # Handle both old format (preview_mode) and new format (request_type)
                        request_type = message_body.get('request_type', '')
                        is_preview = (request_type == 'preview') or message_body.get('preview_mode', False)
                        
                        background_event = {
                            "background_processing": True,
                            "preview_mode": is_preview,
                            "session_id": message_body.get('session_id'),
                            "timestamp": message_body.get('timestamp', datetime.utcnow().strftime('%Y%m%d_%H%M%S')),
                            "reference_pin": message_body.get('reference_pin'),
                            "excel_s3_key": message_body.get('excel_s3_key'),
                            "config_s3_key": message_body.get('config_s3_key'),
                            "results_key": message_body.get('results_key'),
                            "preview_max_rows": message_body.get('preview_max_rows', 5),
                            "email_folder": message_body.get('email_folder'),
                            "max_rows": message_body.get('max_rows', 1000),
                            "batch_size": message_body.get('batch_size', 10),
                            "sequential_call": message_body.get('sequential_call'),
                            "email": message_body.get('email'),
                            "email_address": message_body.get('email')  # For compatibility
                        }
                        
                        # Process using existing background handler
                        result = handle_background_processing(background_event, context)
                        logger.info(f"SQS message processed successfully: {result}")
                        
                    except Exception as e:
                        logger.error(f"Error processing SQS message: {str(e)}")
                        import traceback
                        logger.error(traceback.format_exc())
                        # Re-raise to let Lambda retry
                        raise
            
            # Return success for SQS batch
            return {
                'statusCode': 200,
                'body': json.dumps({'message': 'SQS messages processed successfully'})
            }
        
        # Check if this is a status polling request
        if event.get('httpMethod') == 'GET' and event.get('path', '').startswith('/status/'):
            return handle_status_request(event, context)
        
        # Check if this is a background processing request
        if event.get('background_processing'):
            return handle_background_processing(event, context)
        
        # Check if this is a status check request (JSON body with status_check=True)
        if event.get('httpMethod') == 'POST':
            # Get content type from headers
            headers = event.get('headers', {})
            content_type = headers.get('Content-Type') or headers.get('content-type', '')
            
            # Only process as JSON if content type indicates JSON
            if 'application/json' in content_type:
                try:
                    body = event.get('body', '{}')
                        
                    # Log the raw body for debugging
                    logger.info(f"Raw body type: {type(body)}, isBase64Encoded: {event.get('isBase64Encoded')}")
                    if body:
                        logger.info(f"Raw body length: {len(body)}")
                        logger.info(f"Raw body preview (first 100 chars): {repr(body[:100])}")
                
                    if event.get('isBase64Encoded'):
                        try:
                            body = base64.b64decode(body).decode('utf-8')
                        except UnicodeDecodeError:
                            # Try other encodings if UTF-8 fails
                            decoded_body = base64.b64decode(body)
                            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                                try:
                                    body = decoded_body.decode(encoding)
                                    logger.warning(f"Decoded base64 body using {encoding} encoding")
                                    break
                                except UnicodeDecodeError:
                                    continue
                            else:
                                # If all encodings fail, use latin-1 which never fails
                                body = decoded_body.decode('latin-1', errors='replace')
                                logger.error("Could not decode body with any standard encoding, using latin-1 with replacement")
                        
                        # Ensure body is not None
                        if body is None:
                            body = '{}'
                    
                    if body and body.strip():
                        try:
                            request_data = json.loads(body)
                        except json.JSONDecodeError as e:
                            logger.error(f"Failed to parse JSON body: {e}")
                            logger.error(f"Body type: {type(body)}, Body length: {len(body) if body else 0}")
                            logger.error(f"Body preview (first 500 chars): {repr(body[:500]) if body else 'None'}")
                            return {
                                'statusCode': 400,
                                'headers': {
                                    'Content-Type': 'application/json',
                                    'Access-Control-Allow-Origin': '*'
                                },
                                'body': json.dumps({
                                    'error': 'Invalid JSON in request body',
                                    'message': str(e),
                                    'body_type': str(type(body)),
                                    'body_length': len(body) if body else 0
                                })
                            }
                            
                        if request_data.get('status_check') and request_data.get('session_id'):
                            return handle_status_check_request(request_data, context)
                        
                        # Handle JSON action requests
                        action = request_data.get('action')
                        if action:
                            logger.info(f"Processing JSON action: {action}")
                            
                            # Handle email validation actions
                            if action == 'requestEmailValidation':
                                email = request_data.get('email', '').strip()
                                if not email:
                                    return {
                                        'statusCode': 400,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'missing_email',
                                            'message': 'Email address is required'
                                        })
                                    }
                                
                                # Try to create email validation request
                                try:
                                    from dynamodb_schemas import create_email_validation_request
                                    result = create_email_validation_request(email)
                                    
                                    return {
                                        'statusCode': 200,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps(result)
                                    }
                                except Exception as e:
                                    logger.error(f"Error creating email validation request: {e}")
                                    return {
                                        'statusCode': 500,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'internal_error',
                                            'message': 'Failed to create validation request'
                                        })
                                    }
                            
                            # Handle email code validation
                            elif action == 'validateEmailCode':
                                email = request_data.get('email', '').strip()
                                code = request_data.get('code', '').strip()
                                
                                if not email or not code:
                                    return {
                                        'statusCode': 400,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'missing_parameters',
                                            'message': 'Email and validation code are required'
                                        })
                                    }
                                
                                try:
                                    from dynamodb_schemas import validate_email_code
                                    result = validate_email_code(email, code)
                                    
                                    return {
                                        'statusCode': 200,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps(result)
                                    }
                                except Exception as e:
                                    logger.error(f"Error validating email code: {e}")
                                    return {
                                        'statusCode': 500,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'internal_error',
                                            'message': 'Failed to validate email code'
                                        })
                                    }
                            
                            # Handle email validation check
                            elif action == 'checkEmailValidation':
                                email = request_data.get('email', '').strip()
                                if not email:
                                    return {
                                        'statusCode': 400,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'missing_email',
                                            'message': 'Email address is required'
                                        })
                                    }
                                
                                try:
                                    from dynamodb_schemas import is_email_validated
                                    is_validated = is_email_validated(email)
                                    
                                    return {
                                        'statusCode': 200,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': True,
                                            'validated': is_validated,
                                            'message': 'Email is validated and ready to use' if is_validated else 'Email validation required'
                                        })
                                    }
                                except Exception as e:
                                    logger.error(f"Error checking email validation: {e}")
                                    return {
                                        'statusCode': 500,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'internal_error',
                                            'message': 'Failed to check email validation'
                                        })
                                    }
                            
                            # Handle combined check or send validation
                            elif action == 'checkOrSendValidation':
                                email = request_data.get('email', '').strip()
                                if not email:
                                    return {
                                        'statusCode': 400,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'missing_email',
                                            'message': 'Email address is required'
                                        })
                                    }
                                
                                try:
                                    from dynamodb_schemas import check_or_send_validation
                                    result = check_or_send_validation(email)
                                    
                                    return {
                                        'statusCode': 200,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps(result)
                                    }
                                except Exception as e:
                                    logger.error(f"Error in check or send validation: {e}")
                                    return {
                                        'statusCode': 500,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'internal_error',
                                            'message': 'Failed to process email validation'
                                        })
                                    }
                            
                            # Handle user stats lookup
                            elif action == 'getUserStats':
                                email = request_data.get('email', '').strip()
                                if not email:
                                    return {
                                        'statusCode': 400,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'missing_email',
                                            'message': 'Email address is required'
                                        })
                                    }
                                
                                try:
                                    from dynamodb_schemas import get_user_stats
                                    stats = get_user_stats(email)
                                    
                                    return {
                                        'statusCode': 200,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': True,
                                            'stats': stats
                                        })
                                    }
                                except Exception as e:
                                    logger.error(f"Error getting user stats: {e}")
                                    return {
                                        'statusCode': 500,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'internal_error',
                                            'message': 'Failed to get user stats'
                                        })
                                    }
                            
                            # Handle validateConfig action
                            elif action == 'validateConfig':
                                config_content = request_data.get('config', '')
                                if not config_content:
                                    return {
                                        'statusCode': 400,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'error': 'Missing config content',
                                            'valid': False
                                        })
                                    }
                                
                                # Parse and validate the config
                                try:
                                    if isinstance(config_content, str):
                                        config_data = json.loads(config_content)
                                    else:
                                        config_data = config_content
                                    
                                    # Basic validation - check for required fields
                                    if 'validation_targets' in config_data and isinstance(config_data['validation_targets'], list):
                                        return {
                                            'statusCode': 200,
                                            'headers': {
                                                'Content-Type': 'application/json',
                                                'Access-Control-Allow-Origin': '*'
                                            },
                                            'body': json.dumps({
                                                'valid': True,
                                                'message': 'Configuration is valid'
                                            })
                                        }
                                    else:
                                        return {
                                            'statusCode': 200,
                                            'headers': {
                                                'Content-Type': 'application/json',
                                                'Access-Control-Allow-Origin': '*'
                                            },
                                            'body': json.dumps({
                                                'valid': False,
                                                'message': 'Configuration must contain validation_targets array'
                                            })
                                        }
                                except json.JSONDecodeError as e:
                                    return {
                                        'statusCode': 200,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'valid': False,
                                            'message': f'Invalid JSON: {str(e)}'
                                        })
                                    }
                            
                            # Handle processExcel action
                            elif action == 'processExcel':
                                excel_base64 = request_data.get('excel_file', '')
                                config_base64 = request_data.get('config_file', '')
                                email_address = request_data.get('email', 'test@example.com')
                                preview = request_data.get('preview', False)
                                async_mode = request_data.get('async', False)
                                preview_max_rows = request_data.get('preview_max_rows', 5)
                                max_rows = request_data.get('max_rows', 1000)
                                batch_size = request_data.get('batch_size', 10)
                                
                                # Validate email is authenticated
                                try:
                                    from dynamodb_schemas import is_email_validated
                                    if not is_email_validated(email_address):
                                        return {
                                            'statusCode': 403,
                                            'headers': {
                                                'Content-Type': 'application/json',
                                                'Access-Control-Allow-Origin': '*'
                                            },
                                            'body': json.dumps({
                                                'success': False,
                                                'error': 'email_not_validated',
                                                'message': 'Email address must be validated before processing. Please request and enter a validation code first.'
                                            })
                                        }
                                except Exception as e:
                                    logger.error(f"Error checking email validation: {e}")
                                    return {
                                        'statusCode': 500,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'success': False,
                                            'error': 'validation_check_failed',
                                            'message': 'Unable to verify email validation status'
                                        })
                                    }
                                
                                if not excel_base64 or not config_base64:
                                    return {
                                        'statusCode': 400,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'error': 'Missing excel_file or config_file'
                                        })
                                    }
                                
                                # Decode base64 files
                                try:
                                    excel_content = base64.b64decode(excel_base64)
                                    config_content = base64.b64decode(config_base64)
                                except Exception as e:
                                    return {
                                        'statusCode': 400,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'error': f'Invalid base64 encoding: {str(e)}'
                                        })
                                    }
                                
                                # Generate session ID and reference PIN
                                session_id = str(uuid.uuid4())
                                timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                                reference_pin = generate_reference_pin()
                                email_folder = create_email_folder_path(email_address)
                                
                                # Upload files to S3
                                excel_s3_key = f"uploads/{email_folder}/{timestamp}_{reference_pin}_excel_test.xlsx"
                                config_s3_key = f"uploads/{email_folder}/{timestamp}_{reference_pin}_config_test.json"
                                
                                try:
                                    # Upload Excel file
                                    if not upload_file_to_s3(
                                        excel_content, 
                                        S3_CACHE_BUCKET, 
                                        excel_s3_key,
                                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                    ):
                                        raise Exception("Failed to upload Excel file to S3")
                                    
                                    # Upload config file
                                    if not upload_file_to_s3(
                                        config_content, 
                                        S3_CACHE_BUCKET, 
                                        config_s3_key,
                                        'application/json'
                                    ):
                                        raise Exception("Failed to upload config file to S3")
                                    
                                    logger.info(f"JSON action files uploaded - Excel: {excel_s3_key}, Config: {config_s3_key}")
                                    
                                    # Track in DynamoDB if available
                                    if SQS_INTEGRATION_AVAILABLE:
                                        track_validation_call(
                                            session_id=session_id,
                                            email=email_address,
                                            reference_pin=reference_pin,
                                            request_type='preview' if preview else 'full',
                                            excel_s3_key=excel_s3_key,
                                            config_s3_key=config_s3_key
                                        )
                                    
                                    # Handle preview mode
                                    if preview:
                                        if async_mode:
                                            # Async preview - send to SQS
                                            preview_session_id = f"{timestamp}_{reference_pin}_preview"
                                            
                                            if SQS_INTEGRATION_AVAILABLE:
                                                message_id = send_preview_request(
                                                    session_id=preview_session_id,
                                                    excel_s3_key=excel_s3_key,
                                                    config_s3_key=config_s3_key,
                                                    email=email_address,
                                                    reference_pin=reference_pin,
                                                    preview_max_rows=preview_max_rows,
                                                    email_folder=email_folder,
                                                    max_rows=max_rows,
                                                    batch_size=batch_size,
                                                    sequential_call=None,
                                                    async_mode=True
                                                )
                                                
                                                if message_id:
                                                    logger.info(f"Preview request sent to SQS: {message_id}")
                                                    
                                                    # Track async preview request
                                                    try:
                                                        from dynamodb_schemas import track_user_request
                                                        track_user_request(
                                                            email=email_address,
                                                            request_type='preview'
                                                        )
                                                    except Exception as e:
                                                        logger.warning(f"Failed to track async preview request: {e}")
                                                    
                                                    return {
                                                        'statusCode': 200,
                                                        'headers': {
                                                            'Content-Type': 'application/json',
                                                            'Access-Control-Allow-Origin': '*'
                                                        },
                                                        'body': json.dumps({
                                                            'status': 'processing',
                                                            'session_id': preview_session_id,
                                                            'reference_pin': reference_pin,
                                                            'message': 'Preview processing started',
                                                            'async_mode': True
                                                        })
                                                    }
                                            
                                            # Fallback if SQS not available
                                            return {
                                                'statusCode': 500,
                                                'headers': {
                                                    'Content-Type': 'application/json',
                                                    'Access-Control-Allow-Origin': '*'
                                                },
                                                'body': json.dumps({
                                                    'error': 'SQS integration not available'
                                                })
                                            }
                                        else:
                                            # Sync preview - process immediately
                                            start_time = time.time()
                                            validation_results = invoke_validator_lambda(
                                                excel_s3_key, config_s3_key, preview_max_rows, 
                                                batch_size, True, preview_max_rows, None
                                            )
                                            processing_time = time.time() - start_time
                                            
                                            # Return preview results
                                            if validation_results and 'validation_results' in validation_results:
                                                # Track user request in USER_TRACKING_TABLE
                                                try:
                                                    from dynamodb_schemas import track_user_request
                                                    # Extract token usage from validation results if available
                                                    metadata = validation_results.get('metadata', {})
                                                    token_usage = metadata.get('token_usage', {})
                                                    
                                                    track_user_request(
                                                        email=email_address,
                                                        request_type='preview',
                                                        tokens_used=token_usage.get('total_tokens', 0),
                                                        cost_usd=token_usage.get('total_cost', 0.0),
                                                        perplexity_tokens=token_usage.get('perplexity_tokens', 0),
                                                        perplexity_cost=token_usage.get('perplexity_cost', 0.0),
                                                        anthropic_tokens=token_usage.get('anthropic_tokens', 0),
                                                        anthropic_cost=token_usage.get('anthropic_cost', 0.0)
                                                    )
                                                except Exception as e:
                                                    logger.warning(f"Failed to track user request: {e}")
                                                
                                                markdown_table = create_markdown_table_from_results(
                                                    validation_results['validation_results'], 3, config_s3_key
                                                )
                                                return {
                                                    'statusCode': 200,
                                                    'headers': {
                                                        'Content-Type': 'application/json',
                                                        'Access-Control-Allow-Origin': '*'
                                                    },
                                                    'body': json.dumps({
                                                        'status': 'preview_completed',
                                                        'reference_pin': reference_pin,
                                                        'markdown_table': markdown_table,
                                                        'total_rows': validation_results.get('total_rows', 1),
                                                        'total_processed_rows': validation_results.get('total_processed_rows', 1),
                                                        'processing_time': processing_time
                                                    })
                                                }
                                            else:
                                                return {
                                                    'statusCode': 200,
                                                    'headers': {
                                                        'Content-Type': 'application/json',
                                                        'Access-Control-Allow-Origin': '*'
                                                    },
                                                    'body': json.dumps({
                                                        'status': 'preview_completed',
                                                        'reference_pin': reference_pin,
                                                        'total_rows': 0,
                                                        'total_processed_rows': 0,
                                                        'processing_time': processing_time,
                                                        'message': 'No validation results'
                                                    })
                                                }
                                    
                                    # Handle full processing
                                    else:
                                        results_key = f"results/{email_folder}/{timestamp}_{reference_pin}.zip"
                                        
                                        if SQS_INTEGRATION_AVAILABLE:
                                            message_id = send_full_request(
                                                session_id=session_id,
                                                excel_s3_key=excel_s3_key,
                                                config_s3_key=config_s3_key,
                                                email=email_address,
                                                reference_pin=reference_pin,
                                                results_key=results_key,
                                                max_rows=max_rows,
                                                batch_size=batch_size,
                                                email_folder=email_folder
                                            )
                                            
                                            if message_id:
                                                logger.info(f"Full processing request sent to SQS: {message_id}")
                                                
                                                # Track full processing request
                                                try:
                                                    from dynamodb_schemas import track_user_request
                                                    track_user_request(
                                                        email=email_address,
                                                        request_type='full'
                                                    )
                                                except Exception as e:
                                                    logger.warning(f"Failed to track full processing request: {e}")
                                                
                                                return {
                                                    'statusCode': 200,
                                                    'headers': {
                                                        'Content-Type': 'application/json',
                                                        'Access-Control-Allow-Origin': '*'
                                                    },
                                                    'body': json.dumps({
                                                        'status': 'processing_started',
                                                        'reference_pin': reference_pin,
                                                        'message': 'Processing started. Results will be sent to your email.'
                                                    })
                                                }
                                        
                                        # Fallback response
                                        return {
                                            'statusCode': 200,
                                            'headers': {
                                                'Content-Type': 'application/json',
                                                'Access-Control-Allow-Origin': '*'
                                            },
                                            'body': json.dumps({
                                                'status': 'processing_started',
                                                'reference_pin': reference_pin,
                                                'message': 'Processing started (fallback mode)'
                                            })
                                        }
                                        
                                except Exception as e:
                                    logger.error(f"Error processing JSON action: {str(e)}")
                                    return {
                                        'statusCode': 500,
                                        'headers': {
                                            'Content-Type': 'application/json',
                                            'Access-Control-Allow-Origin': '*'
                                        },
                                        'body': json.dumps({
                                            'error': f'Processing failed: {str(e)}'
                                        })
                                    }
                            
                            # Handle checkStatus action
                            elif action == 'checkStatus':
                                # Delegate to existing status check handler
                                return handle_status_check_request(request_data, context)
                            
                            # Handle diagnostics action
                            elif action == 'diagnostics':
                                # Return diagnostic information
                                diagnostics = {
                                    'sqs_integration_available': SQS_INTEGRATION_AVAILABLE,
                                    'sqs_import_error': SQS_IMPORT_ERROR,
                                    'environment': {
                                        'S3_CACHE_BUCKET': S3_CACHE_BUCKET,
                                        'S3_RESULTS_BUCKET': S3_RESULTS_BUCKET,
                                        'VALIDATOR_LAMBDA_NAME': VALIDATOR_LAMBDA_NAME
                                    },
                                    'boto3_version': boto3.__version__,
                                    'python_version': os.sys.version,
                                    'lambda_function_version': context.function_version if hasattr(context, 'function_version') else 'N/A',
                                    'memory_limit': context.memory_limit_in_mb if hasattr(context, 'memory_limit_in_mb') else 'N/A'
                                }
                                
                                return {
                                    'statusCode': 200,
                                    'headers': {
                                        'Content-Type': 'application/json',
                                        'Access-Control-Allow-Origin': '*'
                                    },
                                    'body': json.dumps(diagnostics, indent=2)
                                }
                            
                            else:
                                return {
                                    'statusCode': 400,
                                    'headers': {
                                        'Content-Type': 'application/json',
                                        'Access-Control-Allow-Origin': '*'
                                    },
                                    'body': json.dumps({
                                        'error': f'Unknown action: {action}'
                                    })
                                }
                            
                except json.JSONDecodeError:
                    pass  # Not JSON, continue with other handlers
                except Exception as e:
                    logger.error(f"Error processing JSON request: {str(e)}")
                    pass  # Continue with other handlers
        
        # Add print statements for debugging
        print(f"[INTERFACE] Lambda handler started")
        print(f"[INTERFACE] Event type: {event.get('httpMethod', 'background' if event.get('background_processing') else 'unknown')}")
        
        logger.info(f"Received event: {json.dumps(event, default=str)}")
        
        # Check if this is a background processing call (self-invoke)
        if event.get('background_processing'):
            print(f"[INTERFACE] Background processing mode detected")
            logger.info("Background processing mode detected")
            return handle_background_processing(event, context)
        
        # Parse request for main API calls
        http_method = event.get('httpMethod', 'POST')
        headers = event.get('headers', {})
        body = event.get('body', '')
        is_base64_encoded = event.get('isBase64Encoded', False)
        
        # Don't decode base64 here - let individual handlers do it
        # This was causing issues with multipart parsing
        
        # Parse query parameters
        query_params = event.get('queryStringParameters') or {}
        preview_first_row = query_params.get('preview_first_row', 'false').lower() == 'true'
        async_mode = query_params.get('async', 'false').lower() == 'true'
        max_rows = int(query_params.get('max_rows', '1000'))
        batch_size = int(query_params.get('batch_size', '10'))
        
        # Parse sequential call parameter
        sequential_call = query_params.get('sequential_call')
        sequential_call_num = None
        if sequential_call:
            try:
                sequential_call_num = int(sequential_call)
                if sequential_call_num < 1:
                    raise ValueError("Sequential call must be >= 1")
                logger.info(f"Sequential call #{sequential_call_num} requested - will process rows 1-{sequential_call_num}")
            except ValueError as e:
                logger.warning(f"Invalid sequential_call parameter: {sequential_call} - {e}")
                sequential_call_num = None
        
        logger.info(f"Parameters - preview_first_row: {preview_first_row}, async_mode: {async_mode}, max_rows: {max_rows}, batch_size: {batch_size}, sequential_call: {sequential_call_num}")
        
        # For preview mode, get max rows to process (default 5, or sequential call number)
        preview_max_rows = 5
        if preview_first_row:
            if sequential_call_num is not None:
                preview_max_rows = min(sequential_call_num, 5)  # Use sequential call number, capped at 5
                logger.info(f"Sequential call: processing {preview_max_rows} rows for call #{sequential_call_num}")
            else:
                preview_max_rows = min(int(query_params.get('preview_max_rows', '5')), 5)  # Cap at 5
        
        # Get content type
        content_type = headers.get('Content-Type') or headers.get('content-type', '')
        logger.info(f"Content-Type header: {content_type}")
        
        # Handle file upload (multipart/form-data)
        if 'multipart/form-data' in content_type:
            logger.info("Detected multipart/form-data request")
            try:
                files, form_data = parse_multipart_form_data(body, content_type, is_base64_encoded)
                
                # Extract email from form data (with default for testing)
                email_address = form_data.get('email', 'eliyahu@eliyahu.ai')
                logger.info(f"Email address: {email_address}")
                
                # Validate email is authenticated
                try:
                    from dynamodb_schemas import is_email_validated
                    if not is_email_validated(email_address):
                        return {
                            'statusCode': 403,
                            'headers': {
                                'Content-Type': 'application/json',
                                'Access-Control-Allow-Origin': '*'
                            },
                            'body': json.dumps({
                                'success': False,
                                'error': 'email_not_validated',
                                'message': 'Email address must be validated before processing. Please request and enter a validation code first.'
                            })
                        }
                except Exception as e:
                    logger.error(f"Error checking email validation: {e}")
                    return {
                        'statusCode': 500,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps({
                            'success': False,
                            'error': 'validation_check_failed',
                            'message': 'Unable to verify email validation status'
                        })
                    }
                
                # Validate request using API Gateway validation if available
                if SQS_INTEGRATION_AVAILABLE:
                    # Combine files and form_data into single request_data dict
                    request_data = {
                        'action': 'processExcel',
                        'files': files,
                        'form_data': form_data
                    }
                    is_valid, validation_result = validate_api_request(request_data)
                    if not is_valid:
                        logger.warning(f"Request validation failed: {validation_result}")
                        return create_validation_error_response(validation_result)
                    
                    # Initialize infrastructure if needed
                    try:
                        create_all_queues()
                        create_call_tracking_table()
                        create_token_usage_table()
                        logger.info("SQS and DynamoDB infrastructure initialized")
                    except Exception as e:
                        logger.warning(f"Infrastructure initialization warning: {e}")
                
                # Validate required files (fallback validation)
                excel_file = files.get('excel_file')
                config_file = files.get('config_file')
                
                # If config_file is not in files, check form_data (web interface sends it as text)
                if not config_file and form_data.get('config_file'):
                    # Create a file-like structure from the form data
                    config_content = form_data.get('config_file')
                    if isinstance(config_content, str):
                        config_file = {
                            'filename': 'config.json',
                            'content': config_content.encode('utf-8') if isinstance(config_content, str) else config_content
                        }
                        logger.info("Config file found in form_data, converting to file structure")
                
                if not excel_file:
                    return {
                        'statusCode': 400,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps({
                            'error': 'Missing required file: excel_file'
                        })
                    }
                
                if not config_file:
                    return {
                        'statusCode': 400,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps({
                            'error': 'Missing required file: config_file'
                        })
                    }
                
                # Generate unique session ID and reference pin
                timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                reference_pin = generate_reference_pin()
                # Use timestamp_referencePin format for session ID (not UUID) so status handler can parse it
                session_id = f"{timestamp}_{reference_pin}"
                
                # Create email-based folder structure
                email_folder = create_email_folder_path(email_address)
                
                # Upload files to S3 with new structure
                excel_s3_key = f"uploads/{email_folder}/{timestamp}_{reference_pin}_excel_{excel_file['filename']}"
                config_s3_key = f"uploads/{email_folder}/{timestamp}_{reference_pin}_config_{config_file['filename']}"
                
                # Upload Excel file
                if not upload_file_to_s3(
                    excel_file['content'], 
                    S3_CACHE_BUCKET, 
                    excel_s3_key,
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                ):
                    raise Exception("Failed to upload Excel file to S3")
                
                # Upload config file
                if not upload_file_to_s3(
                    config_file['content'], 
                    S3_CACHE_BUCKET, 
                    config_s3_key,
                    'application/json'
                ):
                    raise Exception("Failed to upload config file to S3")
                
                logger.info(f"Files uploaded - Excel: {excel_s3_key}, Config: {config_s3_key}")
                
                # Check for async preview mode
                if preview_first_row and async_mode:
                    # Async preview mode - send to SQS priority queue
                    logger.info("Async preview mode - sending to SQS preview queue")
                    
                    # Generate unique session ID for this preview 
                    preview_session_id = f"{timestamp}_{reference_pin}_preview"
                    
                    # Track in DynamoDB if available
                    if SQS_INTEGRATION_AVAILABLE:
                        track_validation_call(
                            session_id=preview_session_id,
                            email=email_address,
                            reference_pin=reference_pin,
                            request_type='preview',
                            excel_s3_key=excel_s3_key,
                            config_s3_key=config_s3_key
                        )
                    
                    # Send to SQS preview queue
                    try:
                        if SQS_INTEGRATION_AVAILABLE:
                            message_id = send_preview_request(
                                session_id=preview_session_id,
                                excel_s3_key=excel_s3_key,
                                config_s3_key=config_s3_key,
                                email=email_address,
                                reference_pin=reference_pin,
                                preview_max_rows=preview_max_rows,
                                email_folder=email_folder,
                                max_rows=max_rows,
                                batch_size=batch_size,
                                sequential_call=sequential_call_num,
                                async_mode=True
                            )
                            
                            if message_id:
                                logger.info(f"Preview request sent to SQS: {message_id}")
                            else:
                                raise Exception("Failed to send message to SQS preview queue")
                        else:
                            # Fallback to old lambda invocation if SQS not available
                            background_payload = {
                                "background_processing": True,
                                "preview_mode": True,
                                "session_id": preview_session_id,
                                "timestamp": timestamp,
                                "reference_pin": reference_pin,
                                "excel_s3_key": excel_s3_key,
                                "config_s3_key": config_s3_key,
                                "preview_max_rows": preview_max_rows,
                                "email_folder": email_folder,
                                "max_rows": max_rows,
                                "batch_size": batch_size,
                                "sequential_call": sequential_call_num
                            }
                            
                            function_name = context.function_name if hasattr(context, 'function_name') else 'interface-validator'
                            lambda_client.invoke(
                                FunctionName=function_name,
                                InvocationType='Event',
                                Payload=json.dumps(background_payload)
                            )
                            logger.info("Fallback: Background preview processing triggered via lambda")
                    
                    except Exception as e:
                        logger.error(f"CRITICAL: Failed to trigger preview processing: {str(e)}")
                        return {
                            'statusCode': 500,
                            'headers': {
                                'Content-Type': 'application/json',
                                'Access-Control-Allow-Origin': '*'
                            },
                            'body': json.dumps({
                                'error': 'Failed to start preview processing',
                                'message': str(e),
                                'session_id': preview_session_id
                            })
                        }
                    
                    # Return immediately with session ID for polling
                    response_body = {
                        "status": "processing",
                        "session_id": preview_session_id,
                        "reference_pin": reference_pin,
                        "message": "Preview processing started in background",
                        "preview_mode": True,
                        "async_mode": True,
                        "poll_url": f"/status/{preview_session_id}?preview=true",
                        "total_rows": 0  # Will be updated when processing completes
                    }
                    
                    return {
                        'statusCode': 200,
                        'headers': {
                            'Content-Type': 'application/json',
                            'Access-Control-Allow-Origin': '*'
                        },
                        'body': json.dumps(response_body)
                    }
                
                # Process based on workflow type
                if preview_first_row:
                    # Preview workflow - process first row synchronously
                    start_time = time.time()
                    
                    # Track initial DynamoDB call for sync preview
                    if SQS_INTEGRATION_AVAILABLE:
                        try:
                            track_validation_call(
                                session_id=session_id,
                                email=email_address,
                                reference_pin=reference_pin,
                                request_type='preview',
                                excel_s3_key=excel_s3_key,
                                config_s3_key=config_s3_key,
                                async_mode=False,  # This is sync preview
                                trigger_source='api_gateway',
                                trigger_method='sync'
                            )
                            logger.info(f"Tracked sync preview call in DynamoDB: {session_id}")
                        except Exception as e:
                            logger.warning(f"Failed to track initial preview call: {e}")
                    
                    try:
                        # Try to invoke validator Lambda for preview
                        validation_results = invoke_validator_lambda(
                            excel_s3_key, config_s3_key, max_rows, batch_size, True, preview_max_rows, sequential_call_num
                        )
                        
                        processing_time = time.time() - start_time
                        
                        # Convert results to markdown table
                        if validation_results and 'validation_results' in validation_results and validation_results['validation_results']:
                            # Parse real validation results
                            real_results = validation_results['validation_results']
                            total_rows = validation_results.get('total_rows', 1)
                            
                            # Convert to markdown format - always show first 3 rows
                            markdown_table = create_markdown_table_from_results(real_results, 3, config_s3_key)
                            
                            # Extract cost and time metadata from validation results
                            metadata = validation_results.get('metadata', {})
                            token_usage = metadata.get('token_usage', {})
                            
                            # NEW: Extract validation structure metrics
                            validation_metrics = metadata.get('validation_metrics', {})
                            
                            # Calculate per-row cost and time estimates
                            total_cost = token_usage.get('total_cost', 0.0)
                            total_tokens = token_usage.get('total_tokens', 0)
                            api_calls = token_usage.get('api_calls', 0)
                            cached_calls = token_usage.get('cached_calls', 0)
                            
                            # Update DynamoDB with completion metrics for sync preview
                            if SQS_INTEGRATION_AVAILABLE:
                                try:
                                    from dynamodb_schemas import update_processing_metrics
                                    
                                    # Map validation response data to DynamoDB schema fields
                                    processed_rows = validation_results.get('total_processed_rows', 1)
                                    by_provider = token_usage.get('by_provider', {})
                                    
                                    # First update status using update_call_status
                                    from dynamodb_schemas import update_call_status
                                    update_call_status(session_id, 'completed')
                                    
                                    # Then update metrics
                                    metrics_update = {
                                        # Timestamps
                                        'completed_processing_at': datetime.utcnow().isoformat() + 'Z',
                                        
                                        # Row counts
                                        'total_rows': total_rows,
                                        'processed_rows': processed_rows,
                                        'preview_max_rows': preview_max_rows,
                                        
                                        # Timing metrics
                                        'processing_time_seconds': processing_time,
                                        'validation_time_seconds': processing_time,
                                        'avg_time_per_row_seconds': processing_time / processed_rows if processed_rows > 0 else 0,
                                        
                                        # Cost metrics
                                        'total_cost_usd': total_cost,
                                        'avg_cost_per_row_usd': total_cost / processed_rows if processed_rows > 0 else 0,
                                        
                                        # Token metrics
                                        'total_tokens': total_tokens,
                                        'avg_tokens_per_row': total_tokens / processed_rows if processed_rows > 0 else 0,
                                        'total_api_calls': api_calls + cached_calls,
                                        'total_cached_calls': cached_calls,
                                        
                                        # Provider-specific metrics
                                        'perplexity_api_calls': by_provider.get('perplexity', {}).get('calls', 0),
                                        'perplexity_total_tokens': by_provider.get('perplexity', {}).get('total_tokens', 0),
                                        'perplexity_cost_usd': by_provider.get('perplexity', {}).get('total_cost', 0.0),
                                        'anthropic_api_calls': by_provider.get('anthropic', {}).get('calls', 0),
                                        'anthropic_total_tokens': by_provider.get('anthropic', {}).get('total_tokens', 0),
                                        'anthropic_cost_usd': by_provider.get('anthropic', {}).get('total_cost', 0.0),
                                        
                                        # Models used
                                        'perplexity_models_used': list(token_usage.get('by_model', {}).keys()),
                                        
                                        # Cache hit rate
                                        'cache_hit_rate': cached_calls / (api_calls + cached_calls) if (api_calls + cached_calls) > 0 else 0
                                    }
                                    
                                    # NEW: Add validation structure metrics if available
                                    if validation_metrics:
                                        # Add to metrics update
                                        metrics_update.update({
                                            'validated_columns_count': validation_metrics.get('validated_columns_count', 0),
                                            'search_groups_count': validation_metrics.get('search_groups_count', 0),
                                            'high_context_search_groups_count': validation_metrics.get('high_context_search_groups_count', 0),
                                            'claude_search_groups_count': validation_metrics.get('claude_search_groups_count', 0)
                                        })
                                    
                                    update_processing_metrics(session_id, metrics_update)
                                    logger.info(f"Updated DynamoDB metrics for sync preview: {session_id}")
                                    
                                except Exception as e:
                                    logger.error(f"Failed to update DynamoDB metrics: {e}")
                                    # Don't fail the request, just log the error
                            
                            # NEW: Extract batch timing information for sync preview
                            metadata = validation_results.get('metadata', {})
                            batch_timing = metadata.get('batch_timing', {})
                            total_processed_rows = validation_results.get('total_processed_rows', 1)
                            
                            if batch_timing:
                                # Use batch timing data for timing estimates
                                time_per_batch = batch_timing.get('average_batch_time_seconds', 0)
                                total_batches = math.ceil(total_rows / 5)
                                estimated_total_processing_time = time_per_batch * total_batches
                                
                                # Calculate per-row cost estimates (costs are per-row, not per-batch)
                                per_row_cost = total_cost / total_processed_rows if total_processed_rows > 0 else 0
                                estimated_total_cost = per_row_cost * total_rows
                                
                                per_row_tokens = total_tokens / total_processed_rows if total_processed_rows > 0 else 0
                                estimated_total_tokens = per_row_tokens * total_rows
                                
                                logger.info(f"SYNC PREVIEW - Using batch timing: {time_per_batch:.1f}s/batch, {total_batches} batches = {estimated_total_processing_time:.0f}s total")
                                logger.info(f"SYNC PREVIEW - Using per-row costs: ${per_row_cost:.6f}/row × {total_rows} rows = ${estimated_total_cost:.6f} total")
                            else:
                                # Fallback to old method
                                validator_processing_time = metadata.get('processing_time', processing_time)
                                estimated_total_processing_time = validator_processing_time * math.ceil(total_rows / 5)
                                estimated_total_cost = (total_cost / total_processed_rows) * total_rows if total_processed_rows > 0 else 0
                                estimated_total_tokens = (total_tokens / total_processed_rows) * total_rows if total_processed_rows > 0 else 0
                                
                                logger.info(f"SYNC PREVIEW - Using fallback timing: {validator_processing_time:.1f}s * {math.ceil(total_rows / 5)} batches = {estimated_total_processing_time:.0f}s total")
                            
                            response_body = {
                                "status": "preview_completed",
                                "session_id": session_id,
                                "reference_pin": reference_pin,
                                "markdown_table": markdown_table,
                                "total_rows": total_rows,
                                "total_processed_rows": total_processed_rows,
                                "new_row_number": validation_results.get('new_row_number', 1),
                                "preview_complete": validation_results.get('preview_complete', True),
                                "preview_processing_time": processing_time,
                                "estimated_total_processing_time": estimated_total_processing_time,  # NEW: batch-based estimate
                                "cost_estimates": {
                                    "preview_cost": total_cost,  # Total cost for this preview 
                                    "estimated_total_cost": estimated_total_cost,  # NEW: batch-based estimate
                                    "preview_tokens": total_tokens,  # Total tokens for this preview
                                    "estimated_total_tokens": estimated_total_tokens,  # NEW: batch-based estimate
                                    "api_calls": api_calls,
                                    "cached_calls": cached_calls
                                },
                                "token_usage": token_usage,  # Include full token usage details
                                # NEW: Add validation structure metrics to response  
                                "validation_metrics": validation_metrics if validation_metrics else {}
                            }
                        elif validation_results and validation_results.get('status') == 'timeout':
                            # Handle timeout scenario with realistic demo data
                            total_rows = validation_results.get('total_rows', 1)
                            
                            markdown_table = """| Field                  | Confidence | Value                     |
|------------------------|------------|--------------------------|
| Product Name           | HIGH       | FAP-2286                  |
| Developer              | HIGH       | Clovis Oncology           |
| Target                 | MEDIUM     | FAP                       |
| Indication             | LOW        | Solid tumors              |"""
                            
                            response_body = {
                                "status": "preview_completed",
                                "session_id": session_id,
                                "reference_pin": reference_pin,
                                "markdown_table": markdown_table,
                                "total_rows": total_rows,
                                "total_processed_rows": 0,
                                "new_row_number": None,
                                "preview_complete": False,
                                "preview_processing_time": processing_time,
                                "estimated_total_processing_time": total_rows * processing_time * 10,  # Estimate longer time
                                "note": validation_results.get('note', 'Validation timed out, showing demo data')
                            }
                        elif validation_results and 'total_rows' in validation_results:
                            # Got response but no validation results - might be empty or error
                            total_rows = validation_results.get('total_rows', 1)
                            
                            # Check if this was a validator issue vs empty results
                            if 'note' in validation_results:
                                note_msg = validation_results['note']
                            else:
                                note_msg = "Validator completed but returned no results"
                            
                            # Provide helpful demo data instead of error message
                            markdown_table = """| Field                  | Confidence | Value                     |
|------------------------|------------|--------------------------|
| Product Name           | HIGH       | [Demo] FAP-2286           |
| Developer              | HIGH       | [Demo] Clovis Oncology    |
| Target                 | MEDIUM     | [Demo] FAP                |
| Indication             | LOW        | [Demo] Solid tumors       |"""
                            
                            response_body = {
                                "status": "preview_completed",
                                "session_id": session_id,
                                "reference_pin": reference_pin,
                                "markdown_table": markdown_table,
                                "total_rows": total_rows,
                                "total_processed_rows": 0,
                                "new_row_number": None,
                                "preview_complete": False,
                                "preview_processing_time": processing_time,
                                "estimated_total_processing_time": total_rows * processing_time * 5,  # Estimate 5x longer
                                "note": f"{note_msg} - Showing demo data structure"
                            }
                        else:
                            # Complete fallback to realistic demo data
                            markdown_table = """| Field                  | Confidence | Value                     |
|------------------------|------------|--------------------------|
| Product Name           | HIGH       | [Demo] FAP-2286           |
| Developer              | HIGH       | [Demo] Clovis Oncology    |
| Target                 | MEDIUM     | [Demo] FAP                |
| Indication             | LOW        | [Demo] Solid tumors       |
| Therapeutic Radionuclide| HIGH      | [Demo] Lu-177            |
| Development Stage      | MEDIUM     | [Demo] Phase 2           |"""
                            total_rows = 100
                            
                            estimated_total_time = total_rows * processing_time * 10  # Estimate much longer for full processing
                            
                            response_body = {
                                "status": "preview_completed",
                                "session_id": session_id,
                                "reference_pin": reference_pin,
                                "markdown_table": markdown_table,
                                "total_rows": total_rows,
                                "total_processed_rows": 0,
                                "new_row_number": None,
                                "preview_complete": False,
                                "preview_processing_time": processing_time,
                                "estimated_total_processing_time": estimated_total_time,
                                "note": "No validation response - showing demo data structure"
                            }
                        
                    except Exception as e:
                        logger.error(f"Error in preview processing: {str(e)}")
                        # Return demonstration data instead of error
                        processing_time = time.time() - start_time
                        
                        # Update DynamoDB with error status
                        if SQS_INTEGRATION_AVAILABLE:
                            try:
                                from dynamodb_schemas import update_call_status, update_processing_metrics
                                # First update status
                                update_call_status(session_id, 'error', error_message=str(e))
                                
                                # Then update metrics
                                metrics_update = {
                                    'processing_time_seconds': processing_time,
                                    'completed_processing_at': datetime.utcnow().isoformat() + 'Z'
                                }
                                update_processing_metrics(session_id, metrics_update)
                                logger.info(f"Updated DynamoDB with error status for: {session_id}")
                            except Exception as db_error:
                                logger.warning(f"Failed to update DynamoDB with error: {db_error}")
                        
                        markdown_table = """| Field   | Confidence | Value                     |
|---------|------------|--------------------------|
| Name    | HIGH       | Preview Error            |
| Email   | LOW        | Processing failed        |
| Status  | LOW        | Please try again         |"""
                        
                        response_body = {
                            "status": "preview_completed",
                            "session_id": session_id,
                            "reference_pin": reference_pin,
                            "markdown_table": markdown_table,
                            "total_rows": 1,
                            "total_processed_rows": 0,
                            "new_row_number": None,
                            "preview_complete": False,
                            "preview_processing_time": processing_time,
                            "estimated_total_processing_time": processing_time,
                            "warning": f"Preview processing encountered an error: {str(e)}"
                        }
                else:
                    # Normal workflow - return immediate response and trigger background processing
                    start_time = time.time()
                    
                    try:
                        # Create placeholder ZIP file immediately with new naming structure
                        results_key = f"results/{email_folder}/{timestamp}_{reference_pin}.zip"
                        placeholder_zip = create_placeholder_zip()
                        
                        if not upload_file_to_s3(
                            placeholder_zip,
                            S3_RESULTS_BUCKET,
                            results_key,
                            'application/zip'
                        ):
                            raise Exception("Failed to upload placeholder file to S3")
                        
                        # Check if email sending is available
                        if EMAIL_SENDER_AVAILABLE and email_address:
                            print(f"[BACKGROUND] Email sender available, preparing to send to {email_address}")
                            # Email will be sent after processing
                            message = f"Processing started. Results will be emailed to {email_address} when complete."
                            include_download_url = False
                        else:
                            # Fallback to download URL
                            download_url = generate_presigned_url(S3_RESULTS_BUCKET, results_key, 86400)
                            if not download_url:
                                raise Exception("Failed to generate download URL")
                            message = f"Processing started. Download URL will be available when complete."
                            include_download_url = True
                        
                        # Track in DynamoDB if available
                        if SQS_INTEGRATION_AVAILABLE:
                            track_validation_call(
                                session_id=session_id,
                                email=email_address,
                                reference_pin=reference_pin,
                                request_type='full',
                                excel_s3_key=excel_s3_key,
                                config_s3_key=config_s3_key,
                                results_s3_key=results_key
                            )
                        
                        # Send to SQS standard queue for full processing
                        try:
                            if SQS_INTEGRATION_AVAILABLE:
                                message_id = send_full_request(
                                    session_id=session_id,
                                    excel_s3_key=excel_s3_key,
                                    config_s3_key=config_s3_key,
                                    email=email_address,
                                    reference_pin=reference_pin,
                                    results_key=results_key,
                                    max_rows=max_rows,
                                    batch_size=batch_size,
                                    email_folder=email_folder
                                )
                                
                                if message_id:
                                    logger.info(f"Full processing request sent to SQS: {message_id}")
                                else:
                                    raise Exception("Failed to send message to SQS standard queue")
                            else:
                                # Fallback to old lambda invocation if SQS not available
                                background_payload = {
                                    "background_processing": True,
                                    "session_id": session_id,
                                    "timestamp": timestamp,
                                    "reference_pin": reference_pin,
                                    "excel_s3_key": excel_s3_key,
                                    "config_s3_key": config_s3_key,
                                    "results_key": results_key,
                                    "max_rows": max_rows,
                                    "batch_size": batch_size,
                                    "email_address": email_address,
                                    "email_folder": email_folder
                                }
                                
                                function_name = context.function_name if hasattr(context, 'function_name') else 'interface-validator'
                                lambda_client.invoke(
                                    FunctionName=function_name,
                                    InvocationType='Event',
                                    Payload=json.dumps(background_payload)
                                )
                                logger.info("Fallback: Background processing triggered via lambda")
                        
                        except Exception as e:
                            logger.error(f"CRITICAL: Failed to trigger full processing: {str(e)}")
                            # Don't fail completely, but log the issue prominently
                        
                        processing_time = time.time() - start_time
                        
                        response_body = {
                            "status": "processing_started",
                            "session_id": session_id,
                            "reference_pin": reference_pin,
                            "message": message,
                            "processing_time": processing_time,
                            "note": "File will be updated with enhanced Excel results when processing completes"
                        }
                        
                        if include_download_url:
                            response_body['download_url'] = download_url
                        
                    except Exception as e:
                        logger.error(f"Error in normal mode processing: {str(e)}")
                        raise Exception(f"Normal mode processing failed: {str(e)}")
                
                return {
                    'statusCode': 200,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps(response_body)
                }
                
            except Exception as e:
                logger.error(f"Error processing multipart upload: {str(e)}")
                return {
                    'statusCode': 500,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({
                        'error': 'File upload processing failed',
                        'message': str(e)
                    })
                }
        
        else:
            # Handle non-multipart requests (for testing)
            logger.info(f"Non-multipart request detected. Content-Type: {content_type}")
            logger.info(f"Request body type: {type(body)}, is_base64_encoded: {is_base64_encoded}")
            if body:
                logger.info(f"Body preview (first 200 chars): {str(body)[:200]}")
            
            # Check if body contains JSON data (GPT or web interface might send JSON instead of multipart)
            if content_type and 'application/json' in content_type and body:
                logger.info("Detected JSON content type, checking for JSON-style request")
                try:
                    # First check if body is already decoded
                    if isinstance(body, str):
                        json_body = json.loads(body)
                    else:
                        # Body might be bytes, decode it
                        if is_base64_encoded:
                            decoded_body = base64.b64decode(body)
                            json_body = json.loads(decoded_body.decode('utf-8'))
                        else:
                            json_body = json.loads(body.decode('utf-8'))
                    
                    logger.info(f"JSON body keys: {list(json_body.keys()) if isinstance(json_body, dict) else 'not a dict'}")
                    
                    # Check if this is a web interface request with file data
                    if isinstance(json_body, dict) and any(key in json_body for key in ['excel_file', 'config_file', 'email']):
                        logger.info("Detected web interface JSON request, redirecting to JSON action handler")
                        
                        # The JSON action handler expects an 'action' field
                        if 'action' not in json_body:
                            json_body['action'] = 'processExcel'
                        
                        # Also need to handle preview parameters from query string
                        if preview_first_row:
                            json_body['preview'] = True
                            json_body['preview_max_rows'] = preview_max_rows
                        
                        # Redirect to the JSON action handler
                        # Create a modified event for the JSON handler
                        json_event = {
                            'httpMethod': 'POST',
                            'headers': headers,
                            'body': json.dumps(json_body),
                            'isBase64Encoded': False,
                            'queryStringParameters': query_params
                        }
                        
                        # Process through the JSON handler section
                        return lambda_handler(json_event, context)
                except Exception as e:
                    logger.error(f"Error parsing potential JSON body: {e}")
            
            if preview_first_row:
                # Return test preview data
                start_time = time.time()
                time.sleep(0.1)  # Simulate processing
                processing_time = time.time() - start_time
                
                markdown_table = """| Field   | Confidence | Value                     |
|---------|------------|--------------------------|
| Name    | HIGH       | John Smith               |
| Email   | MEDIUM     | john.smith@example.com   |
| Phone   | LOW        | (555) 123-4567           |"""
                
                response_body = {
                    "status": "preview_completed",
                    "markdown_table": markdown_table,
                    "total_rows": 100,
                    "preview_rows_processed": preview_max_rows,
                    "preview_processing_time": processing_time,
                    "estimated_total_processing_time": 100 * processing_time
                }
            else:
                # Return test normal data
                response_body = {
                    "status": "processing_started",
                    "download_url": "https://example-bucket.s3.amazonaws.com/Still_Processing.zip",
                    "password": "temp123",
                    "message": "Processing started. File will be available at the provided URL once complete."
                }
            
            return {
                'statusCode': 200,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps(response_body)
            }
        
    except Exception as e:
        logger.error(f"Error processing request: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'error': 'Internal server error',
                'message': str(e)
            })
        }

def handle_background_processing(event, context):
    """Handle background processing for both normal mode and preview mode validation."""
    try:
        print(f"[BACKGROUND] Starting background processing")
        logger.info("Starting background processing")
        
        # Check if this is preview mode
        is_preview = event.get('preview_mode', False)
        
        # Extract parameters from event
        session_id = event['session_id']
        timestamp = event['timestamp']
        reference_pin = event.get('reference_pin', '000000')
        excel_s3_key = event['excel_s3_key']
        config_s3_key = event['config_s3_key']
        max_rows = event.get('max_rows', 1000)
        batch_size = event.get('batch_size', 10)
        email_folder = event.get('email_folder', 'default')
        
        # Quick S3 bucket access verification
        try:
            print(f"[BACKGROUND] Verifying S3 bucket access: {S3_RESULTS_BUCKET}")
            s3_client.head_bucket(Bucket=S3_RESULTS_BUCKET)
            print(f"[BACKGROUND] ✅ S3 bucket access verified")
        except Exception as s3_check_error:
            print(f"[BACKGROUND] ⚠️ S3 bucket access issue: {s3_check_error}")
            logger.warning(f"S3 bucket access issue: {s3_check_error}")
        
        if is_preview:
            # Preview mode processing
            preview_max_rows = event.get('preview_max_rows', 5)
            sequential_call_num = event.get('sequential_call')
            logger.info(f"Background preview processing for session {session_id}")
            print(f"[BACKGROUND] Preview mode - max rows: {preview_max_rows}")
            if sequential_call_num:
                print(f"[BACKGROUND] Sequential call #{sequential_call_num} - validating cache integrity")
        else:
            # Normal mode processing
            results_key = event['results_key']
            email_address = event.get('email_address', 'eliyahu@eliyahu.ai')
            logger.info(f"Background normal processing for session {session_id}")
            print(f"[BACKGROUND] Normal mode - Email: {email_address}")
        
        # Invoke validator Lambda to get real results
        if is_preview:
            print(f"[BACKGROUND] Starting preview validation for session {session_id}")
            print(f"[BACKGROUND] Calling invoke_validator_lambda with:")
            print(f"[BACKGROUND]   excel_s3_key: {excel_s3_key}")
            print(f"[BACKGROUND]   config_s3_key: {config_s3_key}")
            print(f"[BACKGROUND]   max_rows: {max_rows}")
            print(f"[BACKGROUND]   batch_size: {batch_size}")
            print(f"[BACKGROUND]   preview_mode: True")
            print(f"[BACKGROUND]   preview_max_rows: {preview_max_rows}")
            if sequential_call_num:
                print(f"[BACKGROUND]   sequential_call: {sequential_call_num}")
            
            # Preview mode: process the requested number of preview rows
            print(f"[BACKGROUND] ✅ SIMPLIFIED PREVIEW: Processing {preview_max_rows} rows")
            print(f"[BACKGROUND] ✅ Removing all sequential logic - just simple preview")
            
            validation_results = invoke_validator_lambda(
                excel_s3_key=excel_s3_key,
                config_s3_key=config_s3_key,
                max_rows=preview_max_rows,  # Process requested preview rows
                batch_size=preview_max_rows,  # Process all preview rows in one batch
                preview_first_row=True,
                preview_max_rows=preview_max_rows
                # NO sequential_call parameter - removed completely
            )
            
            print(f"[BACKGROUND] Preview validation completed, got results: {validation_results is not None}")
            print(f"[BACKGROUND] Raw validation_results type: {type(validation_results)}")
            print(f"[BACKGROUND] Raw validation_results: {validation_results}")
            
            if validation_results:
                print(f"[BACKGROUND] Validation results keys: {list(validation_results.keys())}")
                for key, value in validation_results.items():
                    print(f"[BACKGROUND]   {key}: {type(value)} = {value}")
                    
                # No sequential call validation needed - simple 3-row preview
        else:
            print(f"[BACKGROUND] Starting normal validation for session {session_id}, Email: {email_address}")
            logger.info(f"Background processing for session {session_id}")
            # Get sequential_call from event, or None for normal mode
            sequential_call_num = event.get('sequential_call')
            validation_results = invoke_validator_lambda(
                excel_s3_key, config_s3_key, max_rows, batch_size, False, 5, sequential_call_num  # False = normal mode
            )
        
        # Check if we got validation results - be more careful about the check
        has_results = (validation_results and 
                      'validation_results' in validation_results and 
                      validation_results['validation_results'] is not None)
        
        if has_results:
            real_results = validation_results['validation_results']
            result_count = len(real_results) if isinstance(real_results, dict) else 0
            print(f"[BACKGROUND] ✅ Got validation results with {result_count} rows")
            print(f"[BACKGROUND] Real results type: {type(real_results)}")
            print(f"[BACKGROUND] Real results content: {real_results}")
            
            if result_count == 0:
                print(f"[BACKGROUND] ⚠️ Validation results dict is empty, using fallback")
                has_results = False
        
        if has_results:
            total_rows = validation_results.get('total_rows', 1)
            
            # Extract token usage and cost information
            metadata = validation_results.get('metadata', {})
            token_usage = metadata.get('token_usage', {})
            
            # NEW: Extract validation structure metrics
            validation_metrics = metadata.get('validation_metrics', {})
            if token_usage:
                logger.info(f"Token usage summary - Total tokens: {token_usage.get('total_tokens', 0)}, Total cost: ${token_usage.get('total_cost', 0.0):.6f}")
                logger.info(f"API calls: {token_usage.get('api_calls', 0)} new, {token_usage.get('cached_calls', 0)} cached")
            
            if is_preview:
                # Preview mode - store results in S3 for status polling
                logger.info("Got preview validation results, storing for polling")
                print(f"[BACKGROUND] Preview mode: got {len(real_results)} validation results")
                
                # Simple preview analysis: just process the 3 rows we got
                token_usage = metadata.get('token_usage', {})
                total_api_calls = token_usage.get('api_calls', 0)
                total_cached_calls = token_usage.get('cached_calls', 0)
                total_cost = token_usage.get('total_cost', 0.0)
                total_tokens = token_usage.get('total_tokens', 0)
                
                print(f"[BACKGROUND] ✅ SIMPLE PREVIEW ANALYSIS:")
                print(f"[BACKGROUND] Token usage: {total_api_calls} API calls, {total_cached_calls} cached calls, ${total_cost:.6f} total cost")
                
                # Simple analysis: we processed 3 rows, that's it
                total_rows_processed = len(real_results)  # Should be 3
                expected_api_calls = total_rows_processed * 5  # About 5 API calls per row
                
                print(f"[BACKGROUND] Simple math:")
                print(f"[BACKGROUND]   Rows processed: {total_rows_processed}")
                print(f"[BACKGROUND]   Expected API calls: {expected_api_calls} (5 per row)")
                print(f"[BACKGROUND]   Actual API calls: {total_api_calls}")
                print(f"[BACKGROUND]   Actual cached calls: {total_cached_calls}")
                print(f"[BACKGROUND]   Total cost: ${total_cost:.6f}")
                print(f"[BACKGROUND]   Total tokens: {total_tokens}")
                
                # Report actual API call counts from validator (don't "correct" them)
                total_calls_reported = total_api_calls + total_cached_calls
                print(f"[BACKGROUND] ✅ Actual API call counts from validator:")
                print(f"[BACKGROUND]   Expected calls: {expected_api_calls} (5 per row × {total_rows_processed} rows)")
                print(f"[BACKGROUND]   Actual new calls: {total_api_calls}")
                print(f"[BACKGROUND]   Actual cached calls: {total_cached_calls}")
                print(f"[BACKGROUND]   Total calls: {total_calls_reported}")
                
                if total_calls_reported != expected_api_calls:
                    print(f"[BACKGROUND] ℹ️ Call count difference may indicate partial caching or different search group count")
                
                # DEBUG: NEW BATCH TIMING ANALYSIS
                print(f"[BACKGROUND] ============ BATCH TIMING ANALYSIS ============")
                
                # Check all metadata keys
                print(f"[BACKGROUND] Available metadata keys: {list(metadata.keys())}")
                
                # NEW: Extract batch timing information
                batch_timing = metadata.get('batch_timing', {})
                validator_processing_time = metadata.get('processing_time', 0.0)  # Keep for fallback
                
                if batch_timing:
                    print(f"[BACKGROUND] ✅ Found batch timing data!")
                    print(f"[BACKGROUND] Batch timing keys: {list(batch_timing.keys())}")
                    
                    total_batches = batch_timing.get('total_batches', 0)
                    batch_size = batch_timing.get('batch_size', 5)
                    avg_batch_time = batch_timing.get('average_batch_time_seconds', 0)
                    total_batch_time = batch_timing.get('total_batch_time_seconds', 0)
                    avg_row_time = batch_timing.get('average_time_per_row_seconds', 0)
                    
                    print(f"[BACKGROUND] Batch metrics:")
                    print(f"[BACKGROUND]   Total batches: {total_batches}")
                    print(f"[BACKGROUND]   Batch size: {batch_size}")
                    print(f"[BACKGROUND]   Average batch time: {avg_batch_time:.2f}s")
                    print(f"[BACKGROUND]   Total batch time: {total_batch_time:.2f}s")
                    print(f"[BACKGROUND]   Average row time: {avg_row_time:.2f}s")
                    
                    # Use batch timing for calculations
                    processing_time = total_batch_time
                    time_per_batch = avg_batch_time
                    time_per_row = avg_row_time
                    processing_time_sources = [f"batch_timing.total_batch_time_seconds: {total_batch_time:.1f}s"]
                    
                    print(f"[BACKGROUND] ✅ Using batch timing: {processing_time:.1f}s total, {time_per_batch:.1f}s per batch")
                    
                else:
                    print(f"[BACKGROUND] ⚠️ No batch timing data found, falling back to legacy processing_time")
                    
                    # Fallback to old method
                if validator_processing_time > 0:
                        processing_time = validator_processing_time
                        # Estimate batch timing from total time (assuming 3 rows processed)
                        processed_rows = total_rows_processed if total_rows_processed > 0 else 3
                        time_per_row = processing_time / processed_rows
                        time_per_batch = time_per_row * 5  # 5 rows per batch
                        processing_time_sources = [f"metadata.processing_time: {validator_processing_time:.1f}s (fallback)"]
                        
                        print(f"[BACKGROUND] ✅ Using fallback timing: {processing_time:.1f}s total")
                        print(f"[BACKGROUND] Estimated batch timing: {time_per_batch:.1f}s per batch")
                else:
                        print(f"[BACKGROUND] ❌ No timing data found at all!")
                        processing_time = 0.0
                        time_per_batch = 20.0  # Fallback estimate
                        time_per_row = 4.0    # Fallback estimate  
                        processing_time_sources = ["No timing data found - using fallback estimates"]
                
                print(f"[BACKGROUND] ============ 🎯 FINAL BATCH TIMING RESULT ============")
                print(f"[BACKGROUND] ✅ Total processing time: {processing_time:.1f}s")
                print(f"[BACKGROUND] 🚀 Time per batch (5 rows): {time_per_batch:.1f}s")
                print(f"[BACKGROUND] → Time per row (derived): {time_per_row:.1f}s")
                print(f"[BACKGROUND] 📊 Sources: {processing_time_sources}")
                
                # BATCH-BASED cost calculation: calculate per-row and per-batch costs
                if total_rows_processed > 0:
                    per_row_cost = total_cost / total_rows_processed
                    per_row_tokens = total_tokens / total_rows_processed  
                    # per_row_time now comes from batch timing analysis above
                    
                    print(f"[BACKGROUND] ✅ COST & TOKEN CALCULATIONS (PER-ROW):")
                    print(f"[BACKGROUND]   Total cost for {total_rows_processed} rows: ${total_cost:.6f}")
                    print(f"[BACKGROUND]   → Per-row cost: ${per_row_cost:.6f}")
                    print(f"[BACKGROUND]   → Per-row tokens: {per_row_tokens:.0f}")
                    print(f"[BACKGROUND] 🚀 BATCH TIMING METRICS (PARALLELIZATION):")
                    print(f"[BACKGROUND]   → Per-row time: {time_per_row:.1f}s")
                    print(f"[BACKGROUND]   → Time per batch (5 rows parallel): {time_per_batch:.1f}s")
                else:
                    # Fallback if no rows processed
                    per_row_cost = 0.02
                    per_row_tokens = 200
                    time_per_row = 4.0    # Use the time_per_row from timing analysis
                    print(f"[BACKGROUND] Using fallback estimates (per-row costs, batch timing)")
                
                # Create simple preview table
                print(f"[BACKGROUND] Creating preview table for {total_rows_processed} rows")
                markdown_table = create_markdown_table_from_results(real_results, preview_max_rows, config_s3_key)
                
                # NEW: Calculate batch processing estimates using BATCH TIMING
                import math
                total_batches = math.ceil(total_rows / 5)
                
                print(f"[BACKGROUND] 🎯 BATCH TIMING ESTIMATES:")
                print(f"[BACKGROUND]   → Total rows to process: {total_rows}")
                print(f"[BACKGROUND]   → Batches needed (5 rows each): {total_batches}")
                print(f"[BACKGROUND]   → Time per batch: {time_per_batch:.1f}s")
                
                # Use actual batch timing for estimates
                estimated_batch_time_seconds = total_batches * time_per_batch
                estimated_batch_time_minutes = estimated_batch_time_seconds / 60
                
                print(f"[BACKGROUND]   → Estimated processing time: {estimated_batch_time_seconds:.0f}s ({estimated_batch_time_minutes:.1f} min)")
                print(f"[BACKGROUND] 💰 COST ESTIMATES (PER-ROW BASED):")
                print(f"[BACKGROUND]   → Cost per row: ${per_row_cost:.6f}")
                print(f"[BACKGROUND]   → Estimated total cost: ${per_row_cost * total_rows:.6f}")
                
                # NEW: Batch-oriented preview results
                preview_results = {
                    "status": "preview_completed",
                    "session_id": session_id,
                    "reference_pin": reference_pin,
                    "markdown_table": markdown_table,
                    "total_rows": total_rows,
                    "total_processed_rows": total_rows_processed,  # Rows we processed in this preview
                    "new_row_number": total_rows_processed,  # Last row processed
                    "preview_complete": True,  # Preview is always complete for 3-row mode
                    "preview_processing_time": processing_time,  # Actual time from validator
                    
                    # NEW: Batch timing estimates
                    "estimated_total_processing_time": estimated_batch_time_seconds,
                    "estimated_total_time_minutes": round(estimated_batch_time_minutes, 1),
                    "estimated_batches": total_batches,
                    "time_per_batch_seconds": time_per_batch,
                    
                    # Keep per-row metrics for compatibility but add batch metrics
                    "per_row_processing_time": time_per_row,  # Time per row (for compatibility)
                    
                    "cost_estimates": {
                        "preview_cost": total_cost,  # Total cost for this preview
                        "estimated_total_cost": per_row_cost * total_rows,  # Per-row based estimate
                        "preview_tokens": total_tokens,  # Total tokens for this preview
                        "estimated_total_tokens": per_row_tokens * total_rows,  # Per-row based estimate
                        "api_calls": total_api_calls,
                        "cached_calls": total_cached_calls,
                        
                        # Per-row metrics (costs & tokens)
                        "per_row_cost": per_row_cost,  # Cost per row
                        "per_row_tokens": per_row_tokens,  # Tokens per row
                        "per_row_time": time_per_row,  # Time per row
                        
                        "rows_processed": total_rows_processed  # Number of rows in this preview
                    },
                    "token_usage": token_usage,
                    
                    # NEW: Add validation structure metrics to response  
                    "validation_metrics": validation_metrics if validation_metrics else {}
                }
                
                # Store preview results in S3 using the same pattern as normal mode
                # Use the exact session_id to ensure consistency
                preview_results_key = f"preview_results/{email_folder}/{session_id}.json"
                print(f"[BACKGROUND] Storing preview results to S3: {preview_results_key}")
                
                # ADD: Update DynamoDB with comprehensive preview metrics
                if SQS_INTEGRATION_AVAILABLE:
                    try:
                        from dynamodb_schemas import update_processing_metrics
                        
                        # Map validation response data to DynamoDB schema fields
                        metrics_update = {
                            # Processing status
                            'status': 'completed',
                            'processing_completed_at': datetime.now(timezone.utc).isoformat(),
                            
                            # Row processing metrics
                            'total_rows': total_rows,
                            'processed_rows': total_rows_processed,
                            'new_rows_processed': total_rows_processed,  # For preview, all rows are "new"
                            
                            # Timing metrics (with correct units)
                            'processing_time_seconds': processing_time,
                            'validation_time_seconds': processing_time,  # Same as processing time
                            'avg_time_per_row_seconds': time_per_row,  # Use batch-derived per-row time
                            
                            # Cost and token metrics (with correct units)
                            'total_cost_usd': total_cost,
                            'total_tokens': total_tokens,
                            'total_api_calls': total_api_calls,
                            'total_cached_calls': total_cached_calls,
                            'avg_cost_per_row_usd': per_row_cost,
                            'avg_tokens_per_row': per_row_tokens,
                            
                            # Batch timing metrics (timing only)
                            'estimated_total_batches': total_batches,
                            'time_per_batch_seconds': time_per_batch,
                            
                            # Preview-specific estimates (per-row costs, batch timing)
                            'preview_per_row_cost_usd': per_row_cost,
                            'preview_per_row_tokens': per_row_tokens,
                            'preview_per_row_time_seconds': time_per_row,
                            'preview_estimated_total_cost_usd': per_row_cost * total_rows,  # Per-row based
                            'preview_estimated_total_tokens': per_row_tokens * total_rows,  # Per-row based
                            'preview_estimated_total_time_hours': estimated_batch_time_seconds / 3600,  # Batch timing based
                        }
                        
                        # NEW: Add validation structure metrics if available
                        if validation_metrics:
                            metrics_update.update({
                                'validated_columns_count': validation_metrics.get('validated_columns_count', 0),
                                'search_groups_count': validation_metrics.get('search_groups_count', 0),
                                'high_context_search_groups_count': validation_metrics.get('high_context_search_groups_count', 0),
                                'claude_search_groups_count': validation_metrics.get('claude_search_groups_count', 0)
                            })
                        
                        # Add API provider-specific metrics if available
                        if 'by_provider' in token_usage:
                            for provider, provider_data in token_usage['by_provider'].items():
                                if provider.lower() == 'perplexity':
                                    metrics_update.update({
                                        'perplexity_api_calls': provider_data.get('api_calls', 0),
                                        'perplexity_cached_calls': provider_data.get('cached_calls', 0),
                                        'perplexity_prompt_tokens': provider_data.get('prompt_tokens', 0),
                                        'perplexity_completion_tokens': provider_data.get('completion_tokens', 0),
                                        'perplexity_total_tokens': provider_data.get('total_tokens', 0),
                                        'perplexity_cost_usd': provider_data.get('cost', 0.0),
                                        'perplexity_models_used': list(provider_data.get('models', {}).keys())
                                    })
                                elif provider.lower() == 'anthropic':
                                    metrics_update.update({
                                        'anthropic_api_calls': provider_data.get('api_calls', 0),
                                        'anthropic_cached_calls': provider_data.get('cached_calls', 0),
                                        'anthropic_input_tokens': provider_data.get('input_tokens', 0),
                                        'anthropic_output_tokens': provider_data.get('output_tokens', 0),
                                        'anthropic_cache_tokens': provider_data.get('cache_tokens', 0),
                                        'anthropic_total_tokens': provider_data.get('total_tokens', 0),
                                        'anthropic_cost_usd': provider_data.get('cost', 0.0),
                                        'anthropic_models_used': list(provider_data.get('models', {}).keys())
                                    })
                        
                        # Update DynamoDB with comprehensive metrics
                        update_success = update_processing_metrics(session_id, metrics_update)
                        
                        if update_success:
                            print(f"[BACKGROUND] ✅ DynamoDB updated with comprehensive preview metrics")
                            logger.info(f"DynamoDB updated for session {session_id} with {len(metrics_update)} fields")
                        else:
                            print(f"[BACKGROUND] ❌ Failed to update DynamoDB metrics")
                            logger.error(f"Failed to update DynamoDB metrics for session {session_id}")
                            
                    except Exception as db_error:
                        print(f"[BACKGROUND] ❌ DynamoDB update error: {db_error}")
                        logger.error(f"Error updating DynamoDB metrics: {db_error}")
                else:
                    print(f"[BACKGROUND] ⚠️ SQS integration not available, skipping DynamoDB update")
                
                # More robust S3 storage with detailed error handling
                s3_write_success = False
                s3_error = None
                for attempt in range(3):  # Try up to 3 times
                    try:
                        print(f"[BACKGROUND] S3 write attempt {attempt + 1}/3")
                        s3_client.put_object(
                            Bucket=S3_RESULTS_BUCKET,  # Use same bucket as normal results
                            Key=preview_results_key,
                            Body=json.dumps(preview_results, indent=2),
                            ContentType='application/json'
                        )
                        logger.info(f"Preview results stored at {preview_results_key}")
                        print(f"[BACKGROUND] ✅ Successfully stored preview results to S3 on attempt {attempt + 1}")
                        s3_write_success = True
                        break
                    except Exception as e:
                        s3_error = str(e)
                        logger.error(f"Attempt {attempt + 1} failed to store preview results: {e}")
                        print(f"[BACKGROUND] ❌ Attempt {attempt + 1} failed to store preview results: {e}")
                        if attempt < 2:  # Don't sleep on last attempt
                            time.sleep(1)  # Wait 1 second before retry
                
                if not s3_write_success:
                    print(f"[BACKGROUND] 🚨 CRITICAL: All S3 write attempts failed - {s3_error}")
                    logger.error(f"CRITICAL: All S3 write attempts failed - {s3_error}")
                
                return {
                    'statusCode': 200,
                    'body': json.dumps({
                        'status': 'preview_completed',
                        'session_id': session_id,
                        'results_stored': s3_write_success,
                        's3_error': s3_error if not s3_write_success else None
                    })
                }
            
            else:
                # Normal mode - create enhanced ZIP as before
                logger.info("Got real validation results, creating enhanced ZIP")
                
                # Get original files for enhanced Excel creation
                excel_response = s3_client.get_object(Bucket=S3_CACHE_BUCKET, Key=excel_s3_key)
                excel_content = excel_response['Body'].read()
                
                config_response = s3_client.get_object(Bucket=S3_CACHE_BUCKET, Key=config_s3_key)
                config_data = json.loads(config_response['Body'].read().decode('utf-8'))
                
                # Extract filenames from S3 keys
                input_filename = excel_s3_key.split('/')[-1].replace(f'{timestamp}_{reference_pin}_excel_', '')
                config_filename = config_s3_key.split('/')[-1].replace(f'{timestamp}_{reference_pin}_config_', '')
                
                # Create enhanced result ZIP with real validation data
                enhanced_zip = create_enhanced_result_zip(
                    real_results, session_id, total_rows, excel_content, config_data,
                    reference_pin=reference_pin, input_filename=input_filename, config_filename=config_filename,
                    metadata=metadata
                )
                
                # ADD: Update DynamoDB with comprehensive full validation metrics
                if SQS_INTEGRATION_AVAILABLE:
                    try:
                        from dynamodb_schemas import update_processing_metrics
                        
                        # Extract metrics from validation response
                        token_usage = metadata.get('token_usage', {})
                        processing_time = metadata.get('processing_time', 0.0)
                        total_cost = token_usage.get('total_cost', 0.0)
                        total_tokens = token_usage.get('total_tokens', 0)
                        total_api_calls = token_usage.get('api_calls', 0)
                        total_cached_calls = token_usage.get('cached_calls', 0)
                        
                        # Calculate per-row metrics
                        processed_rows = len(real_results)
                        per_row_cost = total_cost / processed_rows if processed_rows > 0 else 0.0
                        per_row_tokens = total_tokens / processed_rows if processed_rows > 0 else 0
                        per_row_time = processing_time / processed_rows if processed_rows > 0 else 0.0
                        
                        # Count validation confidence levels
                        confidence_counts = {"high": 0, "medium": 0, "low": 0}
                        validation_targets_count = 0
                        
                        for row_data in real_results.values():
                            for field_name, field_data in row_data.items():
                                if isinstance(field_data, dict) and 'confidence_level' in field_data:
                                    validation_targets_count += 1
                                    conf_level = field_data.get('confidence_level', '').lower()
                                    if conf_level in confidence_counts:
                                        confidence_counts[conf_level] += 1
                        
                        # Map validation response data to DynamoDB schema fields  
                        metrics_update = {
                            # Processing status
                            'status': 'completed',
                            'processing_completed_at': datetime.now(timezone.utc).isoformat(),
                            
                            # Row processing metrics
                            'total_rows': total_rows,
                            'processed_rows': processed_rows,
                            'new_rows_processed': processed_rows,
                            'validation_targets_count': validation_targets_count,
                            
                            # Timing metrics (with correct units)
                            'processing_time_seconds': processing_time,
                            'validation_time_seconds': processing_time,
                            'avg_time_per_row_seconds': per_row_time,
                            
                            # Cost and token metrics (with correct units) 
                            'total_cost_usd': total_cost,
                            'total_tokens': total_tokens,
                            'total_api_calls': total_api_calls,
                            'total_cached_calls': total_cached_calls,
                            'avg_cost_per_row_usd': per_row_cost,
                            'avg_tokens_per_row': per_row_tokens,
                            
                            # Quality metrics
                            'high_confidence_count': confidence_counts['high'],
                            'medium_confidence_count': confidence_counts['medium'],
                            'low_confidence_count': confidence_counts['low'],
                            'validation_accuracy_score': confidence_counts['high'] / validation_targets_count if validation_targets_count > 0 else 0.0,
                            
                            # File information
                            'original_excel_filename': input_filename,
                            'original_config_filename': config_filename,
                            'results_file_size_bytes': len(enhanced_zip),
                        }
                        
                        # NEW: Add validation structure metrics if available
                        if validation_metrics:
                            metrics_update.update({
                                'validated_columns_count': validation_metrics.get('validated_columns_count', 0),
                                'search_groups_count': validation_metrics.get('search_groups_count', 0),
                                'high_context_search_groups_count': validation_metrics.get('high_context_search_groups_count', 0),
                                'claude_search_groups_count': validation_metrics.get('claude_search_groups_count', 0)
                            })
                        
                        # Add API provider-specific metrics if available
                        if 'by_provider' in token_usage:
                            for provider, provider_data in token_usage['by_provider'].items():
                                if provider.lower() == 'perplexity':
                                    metrics_update.update({
                                        'perplexity_api_calls': provider_data.get('api_calls', 0),
                                        'perplexity_cached_calls': provider_data.get('cached_calls', 0),
                                        'perplexity_prompt_tokens': provider_data.get('prompt_tokens', 0),
                                        'perplexity_completion_tokens': provider_data.get('completion_tokens', 0),
                                        'perplexity_total_tokens': provider_data.get('total_tokens', 0),
                                        'perplexity_cost_usd': provider_data.get('cost', 0.0),
                                        'perplexity_models_used': list(provider_data.get('models', {}).keys())
                                    })
                                elif provider.lower() == 'anthropic':
                                    metrics_update.update({
                                        'anthropic_api_calls': provider_data.get('api_calls', 0),
                                        'anthropic_cached_calls': provider_data.get('cached_calls', 0),
                                        'anthropic_input_tokens': provider_data.get('input_tokens', 0),
                                        'anthropic_output_tokens': provider_data.get('output_tokens', 0),
                                        'anthropic_cache_tokens': provider_data.get('cache_tokens', 0),
                                        'anthropic_total_tokens': provider_data.get('total_tokens', 0),
                                        'anthropic_cost_usd': provider_data.get('cost', 0.0),
                                        'anthropic_models_used': list(provider_data.get('models', {}).keys())
                                    })
                        
                        # Update DynamoDB with comprehensive metrics
                        update_success = update_processing_metrics(session_id, metrics_update)
                        
                        if update_success:
                            print(f"[BACKGROUND] ✅ DynamoDB updated with comprehensive full validation metrics")
                            logger.info(f"DynamoDB updated for session {session_id} with {len(metrics_update)} fields")
                        else:
                            print(f"[BACKGROUND] ❌ Failed to update DynamoDB metrics")
                            logger.error(f"Failed to update DynamoDB metrics for session {session_id}")
                            
                    except Exception as db_error:
                        print(f"[BACKGROUND] ❌ DynamoDB update error: {db_error}")
                        logger.error(f"Error updating DynamoDB metrics: {db_error}")
                
                # Upload enhanced results to replace placeholder
                if upload_file_to_s3(
                    enhanced_zip,
                    S3_RESULTS_BUCKET,
                    results_key,
                    'application/zip'
                ):
                    logger.info(f"Enhanced results uploaded to {results_key}")
                    
                    # Send email if available and address provided
                    email_sent = False
                    if EMAIL_SENDER_AVAILABLE and email_address:
                        print(f"[BACKGROUND] Email sender available, preparing to send to {email_address}")
                        # Extract summary data for email
                        all_fields = set()
                        confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
                        
                        for row_key, row_data in real_results.items():
                            for field_name, field_data in row_data.items():
                                if isinstance(field_data, dict) and 'confidence_level' in field_data:
                                    all_fields.add(field_name)
                                    conf_level = field_data.get('confidence_level', 'UNKNOWN')
                                    if conf_level in confidence_counts:
                                        confidence_counts[conf_level] += 1
                        
                        summary_data = {
                            'total_rows': total_rows,
                            'fields_validated': list(all_fields),
                            'confidence_distribution': confidence_counts
                        }
                        
                        print(f"[BACKGROUND] Summary prepared: {len(all_fields)} fields, {total_rows} rows")
                        
                        # Calculate processing time (you might want to track this more accurately)
                        processing_time_seconds = metadata.get('processing_time', 0.0)
                        
                        # Create enhanced Excel filename with timestamp
                        timestamp_str = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                        base_name = input_filename.rsplit('.', 1)[0] if '.' in input_filename else input_filename
                        enhanced_excel_filename = f"{base_name}_Validated_{timestamp_str}.xlsx"
                        
                        print(f"[BACKGROUND] Enhanced Excel filename: {enhanced_excel_filename}")
                        
                        # Create enhanced Excel content
                        enhanced_excel_buffer = create_enhanced_excel_with_validation(
                            excel_content, real_results, config_data, session_id
                        )
                        enhanced_excel_content = enhanced_excel_buffer.getvalue() if enhanced_excel_buffer else excel_content
                        
                        # Send email with individual files
                        print(f"[BACKGROUND] Calling send_validation_results_email with individual files...")
                        email_result = send_validation_results_email(
                            email_address=email_address,
                            excel_content=excel_content,  # Original Excel file
                            config_content=json.dumps(config_data, indent=2).encode('utf-8'),
                            enhanced_excel_content=enhanced_excel_content,  # Enhanced Excel file
                            input_filename=input_filename,
                            config_filename=config_filename,
                            enhanced_excel_filename=enhanced_excel_filename,
                            session_id=session_id,
                            summary_data=summary_data,
                            processing_time=processing_time_seconds,
                            reference_pin=reference_pin,
                            metadata=metadata
                        )
                        
                        print(f"[BACKGROUND] Email result: {email_result}")
                        
                        if email_result['success']:
                            logger.info(f"Email sent successfully to {email_address}")
                            email_sent = True
                            
                            # Track user request completion with final token usage
                            try:
                                from dynamodb_schemas import track_user_request
                                track_user_request(
                                    email=email_address,
                                    request_type='full',  # This is background processing for full requests
                                    tokens_used=total_tokens,
                                    cost_usd=total_cost,
                                    perplexity_tokens=token_usage.get('by_provider', {}).get('perplexity', {}).get('total_tokens', 0),
                                    perplexity_cost=token_usage.get('by_provider', {}).get('perplexity', {}).get('cost', 0.0),
                                    anthropic_tokens=token_usage.get('by_provider', {}).get('anthropic', {}).get('total_tokens', 0),
                                    anthropic_cost=token_usage.get('by_provider', {}).get('anthropic', {}).get('cost', 0.0)
                                )
                                logger.info(f"User request tracking updated for {email_address}")
                            except Exception as e:
                                logger.warning(f"Failed to track user request completion: {e}")
                            
                            # Track email delivery in DynamoDB
                            if SQS_INTEGRATION_AVAILABLE:
                                try:
                                    from dynamodb_schemas import track_email_delivery
                                    track_email_delivery(
                                        session_id=session_id,
                                        email_sent=True,
                                        delivery_status='delivered',
                                        message_id=email_result.get('message_id', ''),
                                        bounce_reason=''
                                    )
                                    logger.info(f"Email delivery tracked in DynamoDB for session {session_id}")
                                except Exception as e:
                                    logger.error(f"Failed to track email delivery in DynamoDB: {e}")
                        else:
                            logger.error(f"Failed to send email: {email_result['message']}")
                            
                            # Track email failure in DynamoDB
                            if SQS_INTEGRATION_AVAILABLE:
                                try:
                                    from dynamodb_schemas import track_email_delivery
                                    track_email_delivery(
                                        session_id=session_id,
                                        email_sent=False,
                                        delivery_status='failed',
                                        message_id='',
                                        bounce_reason=email_result.get('message', 'Unknown error')
                                    )
                                    logger.info(f"Email failure tracked in DynamoDB for session {session_id}")
                                except Exception as e:
                                    logger.error(f"Failed to track email failure in DynamoDB: {e}")
                    else:
                        print(f"[BACKGROUND] Email NOT sent - Available: {EMAIL_SENDER_AVAILABLE}, Address: {email_address}")
                        
                        # Track email not sent in DynamoDB
                        if SQS_INTEGRATION_AVAILABLE:
                            try:
                                from dynamodb_schemas import track_email_delivery
                                track_email_delivery(
                                    session_id=session_id,
                                    email_sent=False,
                                    delivery_status='not_attempted',
                                    message_id='',
                                    bounce_reason='Email sender not available or no email address'
                                )
                                logger.info(f"Email not sent tracked in DynamoDB for session {session_id}")
                            except Exception as e:
                                logger.error(f"Failed to track email not sent in DynamoDB: {e}")
                    
                    # Clean up upload files (optional)
                    try:
                        s3_client.delete_object(Bucket=S3_CACHE_BUCKET, Key=excel_s3_key)
                        s3_client.delete_object(Bucket=S3_CACHE_BUCKET, Key=config_s3_key)
                        logger.info("Cleaned up upload files")
                    except Exception as e:
                        logger.warning(f"Failed to clean up upload files: {e}")
                    
                    return {
                        'statusCode': 200,
                        'body': json.dumps({
                            'status': 'background_completed',
                            'session_id': session_id,
                            'enhanced_file_uploaded': True,
                            'email_sent': email_sent
                        })
                    }
                else:
                    logger.error("Failed to upload enhanced results")
                    return {
                        'statusCode': 500,
                        'body': json.dumps({
                            'status': 'background_failed',
                            'error': 'Failed to upload enhanced results'
                        })
                    }
        else:
            logger.warning("No validation results from validator Lambda")
            print(f"[BACKGROUND] ⚠️ No validation results returned from validator")
            if validation_results:
                print(f"[BACKGROUND] Available keys in response: {list(validation_results.keys())}")
                if 'validation_results' in validation_results:
                    print(f"[BACKGROUND] validation_results key exists but content: {validation_results['validation_results']}")
                    print(f"[BACKGROUND] validation_results type: {type(validation_results['validation_results'])}")
            else:
                print(f"[BACKGROUND] validation_results is None or empty")
            
            # For preview mode, still store a response so polling doesn't hang
            if is_preview:
                print(f"[BACKGROUND] Creating fallback preview response")
                fallback_preview_results = {
                    "status": "preview_completed",
                    "session_id": session_id,
                    "reference_pin": reference_pin,
                    "markdown_table": "| Field | Confidence | Value |\n|-------|------------|-------|\n| No results | N/A | Validation returned no data |",
                    "total_rows": 0,
                    "total_processed_rows": 0,
                    "new_row_number": None,
                    "preview_complete": True,
                    "preview_processing_time": 1.0,
                    "estimated_total_processing_time": 0,
                    "note": "Validation completed but returned no results"
                }
                
                preview_results_key = f"preview_results/{email_folder}/{session_id}.json"
                print(f"[BACKGROUND] Storing fallback preview results to S3: {preview_results_key}")
                
                # More robust S3 storage with detailed error handling
                s3_write_success = False
                s3_error = None
                for attempt in range(3):  # Try up to 3 times
                    try:
                        print(f"[BACKGROUND] S3 write attempt {attempt + 1}/3")
                        s3_client.put_object(
                            Bucket=S3_RESULTS_BUCKET,
                            Key=preview_results_key,
                            Body=json.dumps(fallback_preview_results, indent=2),
                            ContentType='application/json'
                        )
                        logger.info(f"Fallback preview results stored at {preview_results_key}")
                        print(f"[BACKGROUND] ✅ Successfully stored fallback preview results to S3 on attempt {attempt + 1}")
                        s3_write_success = True
                        break
                    except Exception as e:
                        s3_error = str(e)
                        logger.error(f"Attempt {attempt + 1} failed to store fallback preview results: {e}")
                        print(f"[BACKGROUND] ❌ Attempt {attempt + 1} failed to store fallback preview results: {e}")
                        if attempt < 2:  # Don't sleep on last attempt
                            time.sleep(1)  # Wait 1 second before retry
                
                if not s3_write_success:
                    print(f"[BACKGROUND] 🚨 CRITICAL: All S3 write attempts failed - {s3_error}")
                    logger.error(f"CRITICAL: All S3 write attempts failed - {s3_error}")
                
                return {
                    'statusCode': 200,
                    'body': json.dumps({
                        'status': 'preview_completed',
                        'session_id': session_id,
                        'results_stored': s3_write_success,
                        'note': 'No validation results available',
                        's3_error': s3_error if not s3_write_success else None
                    })
                }
            
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'status': 'background_completed',
                    'session_id': session_id,
                    'enhanced_file_uploaded': False,
                    'note': 'No validation results available'
                })
            }
    
    except Exception as e:
        logger.error(f"Error in background processing: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # For preview mode, still try to store an error response so polling doesn't hang
        if event.get('preview_mode', False):
            try:
                session_id = event.get('session_id', 'unknown')
                timestamp = event.get('timestamp', 'unknown')
                reference_pin = event.get('reference_pin', '000000')
                email_folder = event.get('email_folder', 'default')
                
                error_preview_results = {
                    "status": "preview_error",
                    "session_id": session_id,
                    "reference_pin": reference_pin,
                    "markdown_table": "| Field | Confidence | Value |\n|-------|------------|-------|\n| Error | N/A | Background processing failed |",
                    "total_rows": 0,
                    "total_processed_rows": 0,
                    "new_row_number": None,
                    "preview_complete": True,
                    "preview_processing_time": 0,
                    "estimated_total_processing_time": 0,
                    "error": str(e),
                    "note": "Background processing encountered an error"
                }
                
                preview_results_key = f"preview_results/{email_folder}/{session_id}.json"
                print(f"[BACKGROUND] 🚨 Storing error preview results to S3: {preview_results_key}")
                
                s3_client.put_object(
                    Bucket=S3_RESULTS_BUCKET,
                    Key=preview_results_key,
                    Body=json.dumps(error_preview_results, indent=2),
                    ContentType='application/json'
                )
                print(f"[BACKGROUND] ✅ Error response stored to prevent polling timeout")
                
            except Exception as s3_error:
                print(f"[BACKGROUND] 🚨 CRITICAL: Could not store error response to S3: {s3_error}")
                logger.error(f"CRITICAL: Could not store error response to S3: {s3_error}")
        
        return {
            'statusCode': 500,
            'body': json.dumps({
                'status': 'background_failed',
                'error': str(e)
            })
        } 