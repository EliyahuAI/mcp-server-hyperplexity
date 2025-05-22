import os
import pandas as pd
import openpyxl
import yaml
import logging
import asyncio
import argparse
import time
import glob
from datetime import datetime, timedelta
from tab_sanity.validator import Validator
from tab_sanity.schema_validator import SchemaValidator
from tab_sanity.parallel_validator import ParallelValidator
from tab_sanity.config import Config, ValidationTarget, ValidationType, ColumnImportance
from tab_sanity.cache import PerplexityCache
from typing import List, Dict, Tuple, Optional
import json
import hashlib
import httpx
import traceback
import sys

# Configure logging - this is the initial setup which will be replaced with file-specific logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(stream=open(os.devnull, 'w', encoding='utf-8'))  # Create with utf-8 but redirect to null initially
    ]
)

# Ensure all logging handlers use UTF-8 encoding
for handler in logging.root.handlers:
    if isinstance(handler, logging.StreamHandler):
        handler.setStream(sys.stdout)

# Store for tracking cache operations
recent_cache_logs = []

# Create debug stringify function to handle Unicode characters safely
def safe_str(s):
    """Convert any string to ASCII-safe representation."""
    if s is None:
        return "None"
    try:
        # Replace problem Unicode characters with ASCII approximations
        s = str(s)
        s = s.replace('\u2265', '>=')  # Replace ≥ with >=
        s = s.replace('\u2264', '<=')  # Replace ≤ with <=
        s = s.replace('\u2212', '-')   # Replace minus sign with hyphen
        s = s.replace('\u2192', '->')  # Replace right arrow with ->
        s = s.replace('\u2013', '-')   # Replace en dash with hyphen
        s = s.replace('\u2014', '--')  # Replace em dash with double hyphen
        s = s.replace('\u00d7', 'x')   # Replace multiplication sign with x
        s = s.replace('\u03b1', 'alpha') # Replace α with alpha
        s = s.replace('\u03b2', 'beta')  # Replace β with beta
        s = s.replace('\u03b3', 'gamma') # Replace γ with gamma
        s = s.replace('\u03bc', 'micro') # Replace μ with micro
        return s
    except:
        return "[Unicode String]"

# Override the logger to handle Unicode
old_info = logging.info
def safe_info(msg, *args, **kwargs):
    try:
        old_info(msg, *args, **kwargs)
    except UnicodeEncodeError:
        old_info(safe_str(msg), *args, **kwargs)
logging.info = safe_info

old_warning = logging.warning
def safe_warning(msg, *args, **kwargs):
    try:
        old_warning(msg, *args, **kwargs)
    except UnicodeEncodeError:
        old_warning(safe_str(msg), *args, **kwargs)
logging.warning = safe_warning

old_error = logging.error
def safe_error(msg, *args, **kwargs):
    try:
        old_error(msg, *args, **kwargs)
    except UnicodeEncodeError:
        old_error(safe_str(msg), *args, **kwargs)
logging.error = safe_error

# Patch SchemaValidator directly to handle the too many values to unpack error
def patch_schema_validator():
    """Apply a direct patch to the SchemaValidator class to fix the unpack error."""
    try:
        from tab_sanity.schema_validator import SchemaValidator as SV
        original_method = SV.validate_row_multiplex
        
        async def fixed_validate_row_multiplex(self, row, targets):
            """Fixed version that properly handles the 8-value tuple returned by our enhanced query."""
            try:
                # Group columns by search_group for multiplexing
                grouped_columns = self._group_columns_by_search_group(targets)
                
                results = {}
                multiplex_groups = 0
                individual_validations = 0
                
                for group_id, group_targets in grouped_columns.items():
                    logging.info(f"Multiplex validating search group {group_id} with {len(group_targets)} columns")
                    
                    if len(group_targets) > 1:
                        # Use multiplex validation for this group
                        multiplex_groups += 1
                        
                        # Build multiplex prompt
                        prompt, website_url = self._build_multiplex_prompt(row, group_targets)
                        
                        # Try to get cached result first
                        prompt_hash = self._get_hash(prompt)
                        cached_result = self.cache.get("perplexity_schema_raw", self.model_name, prompt, website_url)
                        
                        if cached_result:
                            logging.info(f"Using cached multiplex result with hash {prompt_hash[:10]}")
                            response_json = json.dumps(cached_result)
                        else:
                            logging.info(f"No cached multiplex result found with hash {prompt_hash[:10]}, querying API...")
                            # Use query_perplexity but handle extra return values
                            response = await self._query_perplexity(prompt, website_url, None)
                            
                            # Handle different return value lengths
                            if isinstance(response, tuple):
                                if len(response) >= 8:  # Our enhanced version with cache tracking
                                    response_json = response[0]
                                elif len(response) >= 6:  # Original version
                                    response_json = response[0]
                                else:
                                    logging.error(f"Unexpected response format with {len(response)} values")
                                    response_json = ""
                            else:
                                response_json = response
                        
                        # Process results
                        try:
                            # Parse JSON response
                            parsed_data = json.loads(response_json)
                            logging.info(f"Processing array response with {len(parsed_data)} items")
                            
                            for item in parsed_data:
                                try:
                                    column = item.get("column", "")
                                    answer = item.get("answer", "")
                                    quote = item.get("quote", "")
                                    confidence_level = item.get("confidence", "MEDIUM")
                                    sources = item.get("sources", [])
                                    update_required = item.get("update_required", False)
                                    
                                    # Map confidence level to numeric value
                                    confidence_map = {"LOW": 0.5, "MEDIUM": 0.8, "HIGH": 0.95}
                                    numeric_confidence = confidence_map.get(confidence_level, 0.8)
                                    
                                    # Get main source if available
                                    main_source = sources[0] if sources else ""
                                    
                                    # Store result, with empty cache tracking fields
                                    results[column] = (answer, numeric_confidence, sources, confidence_level, quote, main_source, "", "")
                                    
                                    # Log for debugging
                                    answer_preview = answer[:30] + "..." if len(answer) > 30 else answer
                                    quote_preview = quote[:30] + "..." if len(quote) > 30 else quote
                                    if quote:
                                        logging.info(f"Processed multiplex result for {column}: {answer_preview} | Quote: {quote_preview} (confidence: {confidence_level})")
                                    else:
                                        logging.info(f"Processed multiplex result for {column}: {answer_preview} (confidence: {confidence_level})")
                                except Exception as item_error:
                                    logging.error(f"Error processing multiplex result item: {item_error}")
                            
                        except json.JSONDecodeError:
                            logging.error(f"Error parsing JSON response: {response_json[:100]}...")
                    
                    else:
                        # Use single-column validation for individual columns
                        for target in group_targets:
                            individual_validations += 1
                            try:
                                result = await self.validate_row(row, target)
                                results[target.column] = result
                            except Exception as e:
                                logging.error(f"Error validating single column {target.column}: {e}")
                
                return results, multiplex_groups, individual_validations
            
            except Exception as e:
                logging.error(f"Error in schema multiplex validation: {e}")
                logging.error(traceback.format_exc())
                return {}, 0, 0
        
        # Apply the patch
        SV.validate_row_multiplex = fixed_validate_row_multiplex
        logging.info("Applied direct patch to SchemaValidator.validate_row_multiplex")
        
        # Also patch the determine_next_check_date method
        original_next_check = SV.determine_next_check_date
        
        async def fixed_determine_next_check_date(self, row, validation_results):
            """Fixed version that handles the error in next check date calculation."""
            try:
                return await original_next_check(self, row, validation_results)
            except ValueError as e:
                if "too many values to unpack" in str(e):
                    logging.warning(f"Value unpacking error in next check date calculation: {e}")
                    # Return a default date
                    now = datetime.now()
                    default_date = now + timedelta(days=90)
                    return default_date, ["Error calculating next check date - using default (90 days)"]
                raise
            except Exception as e:
                logging.error(f"Error in determine_next_check_date: {e}")
                # Return a default date
                now = datetime.now()
                default_date = now + timedelta(days=90)
                return default_date, ["Error calculating next check date - using default (90 days)"]
        
        # Apply the patch
        SV.determine_next_check_date = fixed_determine_next_check_date
        logging.info("Applied direct patch to SchemaValidator.determine_next_check_date")
        
        return True
    except Exception as e:
        logging.error(f"Failed to patch SchemaValidator: {e}")
        logging.error(traceback.format_exc())
        return False

# Apply the SchemaValidator patch right away
patch_success = patch_schema_validator()
if patch_success:
    logging.info("Successfully applied SchemaValidator patches at module level for multiplex and next_check_date")
else:
    logging.warning("Failed to apply SchemaValidator patches for multiplex and next_check_date - some errors may occur")

# Define and apply the patch for SchemaValidator.validate_row
def patch_sv_validate_row_method():
    # Ensure imports are local to the patching function if not already globally available
    from tab_sanity.schema_validator import SchemaValidator as SV_ValidateRowPatch
    from tab_sanity.config import ValidationTarget as VT_ValidateRowPatch, ColumnImportance as CI_ValidateRowPatch
    # Use typing directly for these type hints
    from typing import Tuple, Optional, List
    import pandas as pd_ValidateRowPatch
    from datetime import datetime as dt_ValidateRowPatch, timedelta as td_ValidateRowPatch
    import traceback as tb_ValidateRowPatch
    import logging as logging_ValidateRowPatch # Use a local alias for logging

    original_sv_validate_row = getattr(SV_ValidateRowPatch, 'validate_row', None)
    if not original_sv_validate_row:
        logging_ValidateRowPatch.error("Original SchemaValidator.validate_row not found for patching.")
        return

    async def patched_sv_validate_row_impl(self, row: pd_ValidateRowPatch.Series, target: VT_ValidateRowPatch) -> Tuple[Optional[str], float, List[str], str, str, str]:
        if target.importance == CI_ValidateRowPatch.IGNORED:
            current_value = row.get(target.column)
            return current_value, 1.0, [], "HIGH", "", ""

        if not self.config.recheck.force_recheck:
            last_check_val = row.get("Last Check")
            next_check_val = row.get("Next Check Date")
            last_check = pd_ValidateRowPatch.to_datetime(last_check_val, errors='coerce')
            next_check = pd_ValidateRowPatch.to_datetime(next_check_val, errors='coerce')

            if pd_ValidateRowPatch.notna(next_check):
                if dt_ValidateRowPatch.now() < next_check:
                    logging_ValidateRowPatch.info(f"Skipping validation for {target.column} - next check not due until {next_check.strftime('%Y-%m-%d') if pd_ValidateRowPatch.notna(next_check) else 'N/A'}")
                    current_value = row.get(target.column)
                    return current_value, 1.0, [], "HIGH", "", ""
            elif pd_ValidateRowPatch.notna(last_check):
                days_since_last_check = (dt_ValidateRowPatch.now() - last_check).days
                min_days = self.config.recheck.min_days_between_checks
                if days_since_last_check < min_days:
                    logging_ValidateRowPatch.info(f"Skipping validation for {target.column} - last checked {days_since_last_check} days ago (min: {min_days})")
                    current_value = row.get(target.column)
                    return current_value, 1.0, [], "HIGH", "", ""
        
        try:
            prompt, website_url = self._build_prompt(row, target)
            # _query_perplexity (via shim) should return 8 values
            full_result_tuple = await self._query_perplexity(prompt, website_url, None)
            
            # Use typing.Tuple directly here
            if not isinstance(full_result_tuple, Tuple) or len(full_result_tuple) < 6:
                 logging_ValidateRowPatch.error(f"Unexpected result format or length from _query_perplexity for {target.column}: {full_result_tuple}")
                 return None, 0.0, [], "LOW", "", ""

            # Explicitly take the first 6 elements for the return, matching original signature
            answer, numeric_confidence, sources, confidence_level, quote, main_source = full_result_tuple[:6]
            
            if not answer or (isinstance(answer, str) and answer.strip() == ""):
                logging_ValidateRowPatch.warning(f"Empty answer received for {target.column}.") # Removed prompt from log to reduce noise
                return None, 0.0, [], "LOW", "", ""
            
            return answer, numeric_confidence, sources, confidence_level, quote, main_source
        except Exception as e:
            logging_ValidateRowPatch.error(f"Exception in patched SchemaValidator.validate_row for column '{target.column}': {e}")
            logging_ValidateRowPatch.error(f"Traceback: {tb_ValidateRowPatch.format_exc()}")
            return None, 0.0, [], "LOW", "", ""

    SV_ValidateRowPatch.validate_row = patched_sv_validate_row_impl
    logging_ValidateRowPatch.info("Applied direct patch to SchemaValidator.validate_row method (version 3)")

patch_sv_validate_row_method()

# Patch SchemaValidator to make compatible with our modified _query_perplexity
original_query_perplexity = SchemaValidator._query_perplexity

async def compatible_query_perplexity(self, prompt, context_url=None, message_history=None):
    """Make the original method compatible with our enhanced version."""
    if hasattr(self, '_enhanced_query'):
        # Our enhanced version with cache tracking
        return await self._enhanced_query(prompt, context_url, message_history)
    else:
        # Original version without cache tracking
        result = await original_query_perplexity(self, prompt, context_url, message_history)
        # Add empty cache tracking fields
        if isinstance(result, tuple):
            if len(result) == 6:
                return result + ("", "")
        return result

# Store original method
SchemaValidator._original_query_perplexity = original_query_perplexity
# Apply the patched version
SchemaValidator._query_perplexity = compatible_query_perplexity

# This is the enhanced version we'll use that includes cache file tracking
async def enhanced_query_perplexity(self, prompt, context_url=None, message_history=None):
    """Enhanced version that tracks cache files."""
    api_key = self.api_key
    if not api_key:
        logging.error("No API key provided for Perplexity API")
        return "No API key provided", 0.0, [], "LOW", "", "", "", ""
    
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "authorization": f"Bearer {api_key}"
    }
    endpoint = "https://api.perplexity.ai/chat/completions"
    
    # Rate limiting
    await self.rate_limiter.acquire()
    
    try:
        # Determine if this is a multiplex validation
        is_multiplex = prompt.count("Column:") > 1
        
        # Prepare system message content
        system_content = "You are a data validation expert. Return your answer in valid JSON format."
        if is_multiplex:
            system_content += " Return an array of objects, one for each column."
        
        # Add context URL to system message if provided
        if context_url:
            system_content += f" Use this URL for information: {context_url}"
        
        # Build messages array with system message first
        messages = [
            {"role": "system", "content": system_content},
            {"role": "user", "content": prompt}
        ]
        
        # Get the appropriate schema
        schema = self._multiplex_schema if is_multiplex else self._single_column_schema
        
        # Calculate prompt hash for cache lookup
        prompt_hash = self._get_hash(prompt)
        cache_file = ""
        prompt_file = ""
        
        # Build the API request with correct response_format structure
        payload = {
            "model": self.model_name,
            "messages": messages,
            "temperature": 0.1,
            "response_format": {
                "type": "json_schema",
                "json_schema": {
                    "schema": schema
                }
            }
        }
        
        # Make API call
        async with httpx.AsyncClient(timeout=60.0) as client:
            try:
                response = await client.post(endpoint, headers=headers, json=payload, timeout=60.0)
                
                # Log HTTP status for debugging
                if logging.getLogger().level <= logging.DEBUG:
                    logging.debug(f"Response status: {response.status_code}")
                    logging.debug(f"Response headers: {dict(response.headers)}")
                
                response.raise_for_status()
                data = response.json()
                
                # Log response data in debug mode
                if logging.getLogger().level <= logging.DEBUG:
                    logging.debug(f"Response data: {json.dumps(data, indent=2)}")
                
                # Extract content
                full_response = data["choices"][0]["message"]["content"]
                
                # Extract cache file info from recent logs
                for i in range(len(recent_cache_logs) - 1, -1, -1):
                    log = recent_cache_logs[i]
                    if "Cache STORED:" in log:
                        parts = log.split("Cache STORED: ")
                        if len(parts) > 1:
                            cache_file = parts[1].strip().rstrip(".")
                            break
                
                # Look for prompt file
                for i in range(len(recent_cache_logs) - 1, -1, -1):
                    log = recent_cache_logs[i]
                    if "Prompt saved to" in log:
                        parts = log.split("Prompt saved to ")
                        if len(parts) > 1:
                            prompt_file = parts[1].strip()
                            break
                
                if is_multiplex:
                    # For multiplex, return raw JSON result for parsing by validate_row_multiplex
                    try:
                        parsed_data = json.loads(full_response)
                        # Store in cache if it's valid JSON
                        if isinstance(parsed_data, list):
                            # Cache info will be captured by the log handler
                            self.cache.set("perplexity_schema_raw", self.model_name, prompt, parsed_data, context_url)
                            logging.info(f"Cached multiplex result with {len(parsed_data)} items")
                    except json.JSONDecodeError:
                        logging.error(f"Failed to parse JSON response for caching: {full_response[:100]}...")
                    
                    return full_response, 0.8, [], "MEDIUM", "", "", cache_file, prompt_file
                else:
                    # Parse normal response format
                    try:
                        result = json.loads(full_response)
                        answer = result.get("answer", "")
                        quote = result.get("quote", "")
                        confidence_level = result.get("confidence", "MEDIUM")
                        sources = result.get("sources", [])
                        
                        # Map confidence level to numeric value
                        confidence_map = {"LOW": 0.5, "MEDIUM": 0.8, "HIGH": 0.95}
                        numeric_confidence = confidence_map.get(confidence_level, 0.8)
                        
                        # Get main source if available
                        main_source = sources[0] if sources else ""
                        
                        return answer, numeric_confidence, sources, confidence_level, quote, main_source, cache_file, prompt_file
                    except json.JSONDecodeError:
                        logging.error(f"Failed to parse JSON response: {full_response[:100]}...")
                        return "", 0.5, [], "LOW", "", "", "", ""
            
            except Exception as e:
                logging.error(f"HTTP error: {e}")
                # Log the detailed error response if available
                try:
                    if hasattr(e, 'response') and hasattr(e.response, 'text'):
                        error_content = e.response.text
                        logging.error(f"Error response content: {error_content[:200]}...")
                except:
                    pass
                return f"API error: {str(e)}", 0.0, [], "LOW", "", "", "", ""
                
    except Exception as e:
        logging.error(f"Error in _query_perplexity: {e}")
        if logging.getLogger().level <= logging.DEBUG:
            logging.error(f"Error details: {traceback.format_exc()}")
        return "", 0.5, [], "LOW", "", "", "", ""

# Store the enhanced query method
SchemaValidator._enhanced_query = enhanced_query_perplexity

# Enhanced fixed multiplex prompt
def get_fixed_multiplex_prompt(context, columns_to_validate):
    """Generate a fixed multiplex prompt with explicit JSON formatting instructions."""
    return f"""You are a data validation expert. Your task is to validate multiple conference fields:

Context:
{context}

I need you to validate the following columns:

{columns_to_validate}

Provide your answer as VALID JSON ONLY. Format your response as a JSON array with objects for each column.
DO NOT include any markdown formatting, text explanations, or code blocks.
Return RAW JSON directly like this (with no backticks or markdown):

[
  {{
    "column": "Column Name 1",
    "answer": "validated value 1",
    "quote": "direct quote from source",
    "sources": ["url1", "url2"],
    "confidence": "HIGH",
    "update_required": true,
    "explanation": "brief explanation"
  }},
  {{
    "column": "Column Name 2",
    "answer": "validated value 2",
    "quote": "direct quote from source",
    "sources": ["url1", "url2"],
    "confidence": "HIGH", 
    "update_required": false,
    "explanation": "brief explanation"
  }}
]

IMPORTANT: Your response must ONLY be the JSON array with no other text."""

# Patch the validator's methods to use our improved prompts
def patch_validator_methods():
    """Apply monkey patches to fix formatting issues."""
    # Store the original methods
    original_basic_method = Validator._build_basic_multiplex_prompt
    original_template_method = Validator._build_multiplex_prompt
    
    # Create patched methods
    def patched_build_basic_multiplex_prompt(self, context, columns_to_validate):
        logging.info("Using patched basic multiplex prompt formatter")
        return get_fixed_multiplex_prompt(context, columns_to_validate)
    
    def patched_build_multiplex_prompt(self, row, targets):
        """Use our fixed implementation directly, bypassing the template logic."""
        # Build context
        context = ""
        for pk in self.config.primary_key:
            context += f"{pk}: {row.get(pk, '')}\n"
        
        # Build column details
        columns_to_validate = ""
        for target in targets:
            column_info = self.column_config.get(target.column, {})
            description = column_info.get('description', '')
            format_info = column_info.get('format', '')
            notes = column_info.get('notes', '')
            examples = self._format_examples(target.column)
            
            columns_to_validate += f"Column: {target.column}\n"
            columns_to_validate += f"Current Value: {row.get(target.column, '')}\n"
            columns_to_validate += f"Description: {description}\n"
            columns_to_validate += f"Format: {format_info}\n"
            if notes:
                columns_to_validate += f"Notes: {notes}\n"
            if examples:
                columns_to_validate += f"Examples:\n{examples}\n"
            columns_to_validate += "\n---\n\n"
        
        logging.info("Using fixed direct multiplex prompt formatter")
        return get_fixed_multiplex_prompt(context, columns_to_validate), row.get(self.config.website_column, '')
    
    # Apply the patches
    Validator._build_basic_multiplex_prompt = patched_build_basic_multiplex_prompt
    Validator._build_multiplex_prompt = patched_build_multiplex_prompt
    logging.info("Applied validator method patches")

# Context-aware validator for improved results
class ContextAwareValidator(SchemaValidator):
    def __init__(self, parent_validator, enhanced_prompt, website_url):
        # Share schema validation functionality from parent
        self.config = parent_validator.config
        self.api_key = parent_validator.api_key
        self.enable_next_check_date = parent_validator.enable_next_check_date
        self.column_config = parent_validator.column_config
        self.model_name = parent_validator.model_name
        self.prompts = parent_validator.prompts
        self.cache = parent_validator.cache  # Use the same cache instance
        self.rate_limiter = parent_validator.rate_limiter
        
        # Copy schema definitions
        self._single_column_schema = parent_validator._single_column_schema
        self._multiplex_schema = parent_validator._multiplex_schema
        
        # Store the enhanced prompt
        self._enhanced_prompt = enhanced_prompt
        self._website_url = website_url
        
        # Copy any other needed attributes from parent
        self._get_hash = parent_validator._get_hash if hasattr(parent_validator, '_get_hash') else lambda p: hashlib.md5(p.encode()).hexdigest()
        
        # Use the same query function (which could be the patched one)
        self._query_perplexity = parent_validator._query_perplexity
        if hasattr(parent_validator, '_enhanced_query'):
            self._enhanced_query = parent_validator._enhanced_query
        
        logging.info("Created ContextAwareValidator with shared cache")
    
    def _build_multiplex_prompt(self, row, targets):
        """Override to return our enhanced prompt."""
        return self._enhanced_prompt, self._website_url
        
    async def validate_row_multiplex(self, row, targets):
        """Override the parent method to add error handling."""
        try:
            # Call the parent's method
            return await super().validate_row_multiplex(row, targets)
        except ValueError as e:
            logging.error(f"Error in ContextAwareValidator.validate_row_multiplex: {e}")
            logging.error(traceback.format_exc())
            # Return empty results to avoid crashing
            return {}, 0, 0
        except Exception as e:
            logging.error(f"Unexpected error in ContextAwareValidator.validate_row_multiplex: {e}")
            logging.error(traceback.format_exc())
            # Return empty results to avoid crashing
            return {}, 0, 0

# Enhanced ParallelValidator with context-aware features
class EnhancedParallelValidator(ParallelValidator):
    """A parallel validator with enhanced multiplex validation."""
    
    def __init__(self, config: Config, api_key: str, enable_next_check_date: bool = True, max_concurrent_rows: int = 5):
        super().__init__(config, api_key, enable_next_check_date, max_concurrent_rows)
        # Force model to sonar-pro which supports json_schema
        self.model_name = "sonar-pro"
        # Add hash function used for cache lookups
        self._get_hash = lambda prompt: hashlib.md5(prompt.encode()).hexdigest()
        # Add counters for tracking validation statistics
        self.total_multiplex_groups = 0
        self.total_individual_validations = 0
        logging.info(f"Enhanced parallel validator using model: {self.model_name}")
        
        # Organize validation targets by search_group for better multiplex grouping
        self.search_groups = {}
        
        # Load column configuration directly from YAML file
        table_dir = os.path.dirname(config.spreadsheet)
        config_path = os.path.join(table_dir, "column_config.yml")
        with open(config_path, 'r', encoding='utf-8') as f:
            config_data = yaml.safe_load(f)
            column_config = config_data.get('columns', {})
        
        # First pass - identify all search groups
        all_groups = set()
        for target in config.validation_targets:
            # Try to find column_info by looking up in column_config directly
            column_info = column_config.get(target.column, {})
            if not column_info:
                # Also try normalized lookup
                norm_target = normalize_column_name(target.column)
                for col_name, col_info in column_config.items():
                    if normalize_column_name(col_name) == norm_target:
                        column_info = col_info
                        break
            
            search_group = column_info.get('search_group', 0)
            all_groups.add(search_group)
        
        # Ensure we have sequential search groups starting from 0
        all_groups = sorted(list(all_groups))
        self.group_shift = 0
        if all_groups and all_groups[0] > 0:
            # Remember the shift amount
            self.group_shift = all_groups[0]
            logging.info(f"Shifting search groups by {self.group_shift} to start from 0")
        
        # Initialize search groups
        for group in all_groups:
            normalized_group = group - self.group_shift
            self.search_groups[normalized_group] = []
        
        # Second pass - add targets to search groups
        for target in config.validation_targets:
            # Try to find column_info by looking up in column_config directly
            column_info = column_config.get(target.column, {})
            if not column_info:
                # Also try normalized lookup
                norm_target = normalize_column_name(target.column)
                for col_name, col_info in column_config.items():
                    if normalize_column_name(col_name) == norm_target:
                        column_info = col_info
                        break
                    
            search_group = column_info.get('search_group', 0)
            normalized_group = search_group - self.group_shift
            self.search_groups[normalized_group].append(target)
        
        # Log search group assignments
        for group_id, targets in self.search_groups.items():
            target_names = [t.column for t in targets]
            logging.info(f"Search group {group_id}: {target_names}")
            
    def _group_columns_by_search_group(self, targets: List[ValidationTarget]) -> Dict[int, List[ValidationTarget]]:
        """Override to use our normalized search groups."""
        grouped_columns = {}
        for target in targets:
            if target.importance == ColumnImportance.IGNORED:
                continue
                
            # Use our precomputed normalized search groups
            found = False
            for group_id, group_targets in self.search_groups.items():
                if target in group_targets:
                    if group_id not in grouped_columns:
                        grouped_columns[group_id] = []
                    grouped_columns[group_id].append(target)
                    found = True
                    break
            
            # Fallback if not found in any group (shouldn't happen)
            if not found:
                # Get the target column name - this allows us to handle
                # columns that might have different instances but same name
                target_name = target.column
                
                # Try to find a matching group by column name
                for group_id, group_targets in self.search_groups.items():
                    group_names = [t.column for t in group_targets]
                    if target_name in group_names:
                        if group_id not in grouped_columns:
                            grouped_columns[group_id] = []
                        grouped_columns[group_id].append(target)
                        found = True
                        logging.info(f"Matched column {target_name} to search group {group_id} by name")
                        break
            
            # Final fallback - add to group 0 if still not found
            if not found:
                search_group = 0
                if search_group not in grouped_columns:
                    grouped_columns[search_group] = []
                grouped_columns[search_group].append(target)
                logging.warning(f"Column {target.column} not found in any search group, defaulting to group 0")
                
        return grouped_columns

    async def validate_row(self, row, target):
        """Override validate_row to ensure it returns 6 values, not 8."""
        try:
            # super().validate_row() now calls the patched SchemaValidator.validate_row,
            # which correctly handles the 8-tuple from _query_perplexity and returns 6 values.
            result_6_tuple = await super().validate_row(row, target)
            
            # Ensure it's a 6-tuple as expected from the patched method
            if isinstance(result_6_tuple, tuple) and len(result_6_tuple) == 6:
                return result_6_tuple
            elif isinstance(result_6_tuple, tuple) and len(result_6_tuple) == 8: 
                # This case should ideally not happen if the patch is working, but as a fallback, take first 6
                logging.warning(f"EnhancedParallelValidator.validate_row received 8-tuple from super, expected 6. Taking first 6 for {target.column}")
                return result_6_tuple[:6]
            else:
                logging.error(f"Unexpected result format from super().validate_row for {target.column}: {result_6_tuple}. Returning defaults.")
                # Return fallback values consistent with a 6-tuple signature
                return row.get(target.column), 0.0, [], "LOW", "", ""

        except ValueError as e:
            logging.error(f"ValueError in EnhancedParallelValidator.validate_row for {target.column}: {e}")
            logging.error(traceback.format_exc())
            return row.get(target.column), 0.0, [], "LOW", "", ""
        except Exception as e:
            logging.error(f"Unexpected error in EnhancedParallelValidator.validate_row for {target.column}: {e}")
            logging.error(traceback.format_exc())
            return row.get(target.column), 0.0, [], "LOW", "", ""

    async def validate_row_multiplex(self, row, targets):
        """Override validate_row_multiplex to update cache file tracking."""
        try:
            results, mplex_groups, indiv_validations = await super().validate_row_multiplex(row, targets)
            
            # Update result tuples to include cache file information
            for col, result in results.items():
                if isinstance(result, tuple):
                    if len(result) == 6:  # Original format
                        results[col] = result + ("", "")  # Add empty cache and prompt file
                    elif len(result) == 8:  # Already has cache info
                        continue  # No change needed
                    else:  # Unexpected length
                        logging.warning(f"Unexpected result length {len(result)} for column {col}")
                        # Pad with empty values if shorter than expected
                        if len(result) < 6:
                            results[col] = result + tuple("" for _ in range(8 - len(result)))
            
            return results, mplex_groups, indiv_validations
        except ValueError as e:
            if "too many values to unpack" in str(e):
                logging.error(f"Error in validate_row_multiplex: {e}")
                logging.error(f"Traceback: {traceback.format_exc()}")
                # Return empty results to avoid crashing
                return {}, 0, 0
            else:
                raise
        except Exception as e:
            logging.error(f"Unexpected error in validate_row_multiplex: {e}")
            logging.error(traceback.format_exc())
            # Return empty results to avoid crashing
            return {}, 0, 0
    
    async def _process_row(self, row: pd.Series, row_key: str, row_idx: int) -> Tuple[str, Dict, int, int, Optional[datetime], List[str]]:
        """Process a single row with enhanced context-aware validation."""
        logging.info(f"Starting validation for row {row_idx}: {row_key}")
        
        # Group columns by search_group for multiplexing
        grouped_columns = self._group_columns_by_search_group(self.config.validation_targets)
        
        # Initialize context accumulator for this row
        row_context = f"Previous validations for this conference:\n\n"
        
        # Initialize results dictionary
        row_results = {}
        mplex_groups = 0
        indiv_validations = 0
        
        # Process each group using multiplex validation
        for group_id, targets in grouped_columns.items():
            logging.info(f"Processing search group {group_id} with {len(targets)} columns")
            
            # Skip empty groups
            if not targets:
                continue
            
            # Use multiplex validation for groups with multiple columns
            if len(targets) > 1:
                mplex_groups += 1
                logging.info(f"Using multiplex validation for {len(targets)} columns in group {group_id}")
                try:
                    # First get the raw prompt for this multiplex validation
                    base_prompt, website_url = self._build_multiplex_prompt(row, targets)
                    
                    # Add accumulated context to the prompt
                    if len(row_context) > 200:  # Only if we have meaningful context
                        enhanced_prompt = f"{base_prompt}\n\nAdditional Context - Previous validation results for this conference:\n{row_context}\n\nNote: Use the above context to maintain consistency across fields."
                        logging.info("Using enhanced prompt with previous validation context")
                    else:
                        enhanced_prompt = base_prompt
                        
                    # Add general notes if they're not already included
                    general_notes = self._get_general_notes()
                    if general_notes and "general notes" not in enhanced_prompt.lower():
                        enhanced_prompt += f"\n\nGeneral Notes for All Validations:\n{general_notes}"
                        logging.info(f"Added general notes ({len(general_notes)} chars)")
                    
                    # Create a context-aware validator that uses our enhanced prompt
                    context_validator = ContextAwareValidator(self, enhanced_prompt, website_url)
                    
                    # Use context-aware validator for multiplex validation
                    try:
                        group_results, group_mplex, group_indiv = await context_validator.validate_row_multiplex(row, targets)
                        logging.info(f"Successfully processed multiplex group {group_id}")
                    except Exception as ctx_e:
                        logging.error(f"Context-aware validation failed: {ctx_e}")
                        # Fall back to standard validation
                        try:
                            group_results, group_mplex, group_indiv = await self.validate_row_multiplex(row, targets)
                            logging.info(f"Successfully processed with standard validation")
                        except Exception as multi_e:
                            logging.error(f"Standard validation also failed: {multi_e}")
                            # Initialize empty results dict for this group
                            group_results = {}
                            group_mplex = 0
                            group_indiv = 0
                    
                    # Update counters
                    mplex_groups += group_mplex
                    indiv_validations += group_indiv
                    
                    # If group_results is empty (failed multiplex), fall back to individual validation
                    if not group_results:
                        logging.warning(f"No results from multiplex validation, falling back to individual validation")
                        for target in targets:
                            indiv_validations += 1
                            try:
                                # Fallback to validating each column individually
                                column = target.column
                                logging.info(f"Fallback: Individually validating {column}")
                                answer, confidence, citations, confidence_level, quote, source_url = await self.validate_row(row, target)
                                
                                # Add to context and store results
                                row_context += f"Column '{column}': {answer} (Confidence: {confidence_level})\n"
                                group_results[column] = (answer, confidence, citations, confidence_level, quote, source_url)
                                
                                logging.info(f"Fallback successful for {column}: {answer[:30]}... (Confidence: {confidence_level})")
                            except Exception as inner_e:
                                logging.error(f"Individual fallback failed for {target.column}: {inner_e}")
                                # Use empty fallback value
                                group_results[target.column] = ("Unable to validate", 0.0, [], "LOW", "", "")
                    
                    # Add to overall row results
                    row_results.update(group_results)
                    
                    # Update context with results for future validations in this row
                    for column, result in group_results.items():
                        if isinstance(result, tuple) and len(result) >= 4:
                            value = result[0]
                            confidence_level = result[3]
                            row_context += f"Column '{column}': {value} (Confidence: {confidence_level})\n"
                
                except Exception as e:
                    logging.error(f"ERROR in multiplex validation: {e}")
                    logging.error(traceback.format_exc())
            
            else:
                # For single column groups, use standard validation
                indiv_validations += 1
                try:
                    column = targets[0].column
                    logging.info(f"Individually validating {column}")
                    
                    # Call validate_row directly
                    answer, confidence, citations, confidence_level, quote, source_url = await self.validate_row(row, targets[0])
                    
                    # Add to context accumulator
                    row_context += f"Column '{column}': {answer} (Confidence: {confidence_level})\n"
                    
                    # Store result
                    row_results[column] = (answer, confidence, citations, confidence_level, quote, source_url)
                    
                    logging.info(f"Validated {column}: {answer[:30]}... (Confidence: {confidence_level})")
                except Exception as e:
                    logging.error(f"Error in single column validation: {e}")
                    logging.error(traceback.format_exc())
        
        # Final pass: Find columns with low confidence and try one more validation
        low_confidence_columns = []
        for target in self.config.validation_targets:
            column = target.column
            if column in row_results:
                result = row_results[column]
                if isinstance(result, tuple) and len(result) >= 4:
                    # Check confidence level - retry if not HIGH
                    confidence_level = result[3]
                    if confidence_level != "HIGH":
                        low_confidence_columns.append(target)

        if low_confidence_columns:
            logging.info(f"Performing final pass on {len(low_confidence_columns)} columns with non-HIGH confidence")
            # Create one consolidated prompt with all low confidence columns
            # to benefit from cross-column context
            try:
                # Include all the context we've built so far
                base_prompt, website_url = self._build_multiplex_prompt(row, low_confidence_columns)
                
                # Add all the context we've accumulated so far
                enhanced_prompt = f"{base_prompt}\n\nAdditional Context - Here are the validation results so far:\n{row_context}\n\nPlease focus on improving confidence in these columns."
                
                # Add general notes again for emphasis
                general_notes = self._get_general_notes()
                if general_notes:
                    enhanced_prompt += f"\n\nGeneral Notes for All Validations:\n{general_notes}"
                
                # Create a context-aware validator for the final pass
                context_validator = ContextAwareValidator(self, enhanced_prompt, website_url)
                
                # Attempt the final validation
                try:
                    final_results, final_mplex, final_indiv = await context_validator.validate_row_multiplex(row, low_confidence_columns)
                    logging.info(f"Successfully completed final validation pass")
                    
                    # Update counters
                    mplex_groups += final_mplex
                    indiv_validations += final_indiv
                    
                    # Update results but only if the new confidence is higher
                    for column, new_result in final_results.items():
                        if isinstance(new_result, tuple) and len(new_result) >= 4: # new_result could be 6-tuple or 8-tuple
                            original_result = row_results.get(column) # original_result is expected to be 8-tuple
                            if isinstance(original_result, tuple) and len(original_result) >= 4:
                                orig_confidence_level = original_result[3]
                                new_confidence_level = new_result[3]
                                
                                confidence_rank = {"LOW": 1, "MEDIUM": 2, "HIGH": 3}
                                orig_rank = confidence_rank.get(orig_confidence_level.upper(), 0)
                                new_rank = confidence_rank.get(new_confidence_level.upper(), 0)
                                
                                if new_rank > orig_rank:
                                    logging.info(f"Final pass improved {column} confidence from {orig_confidence_level} to {new_confidence_level}")
                                    # Ensure new_result is an 8-tuple before assignment
                                    if len(new_result) == 6:
                                        row_results[column] = new_result + ("", "") # Pad to 8-tuple
                                    elif len(new_result) == 8:
                                        row_results[column] = new_result
                                    else:
                                        logging.warning(f"Unexpected tuple length {len(new_result)} for {column} in final pass, not updating.")
                                elif new_rank == orig_rank and new_result[0] != original_result[0] and new_result[0] is not None and str(new_result[0]).strip() != "":
                                    # If confidence is same, but answer is different and not empty, prefer new answer
                                    logging.info(f"Final pass updated answer for {column} with same confidence {new_confidence_level}. Old: '{original_result[0]}', New: '{new_result[0]}'")
                                    if len(new_result) == 6:
                                        row_results[column] = new_result + ("", "") 
                                    elif len(new_result) == 8:
                                        row_results[column] = new_result
                                else:
                                    logging.info(f"Final pass did not improve {column} confidence ({orig_confidence_level} vs {new_confidence_level}) or answer was not better.")
                            else:
                                # No valid original result, use the new one (padded to 8-tuple if necessary)
                                logging.info(f"Final pass added result for {column} (no valid original)")
                                if len(new_result) == 6:
                                    row_results[column] = new_result + ("", "")
                                elif len(new_result) == 8:
                                    row_results[column] = new_result
                                else:
                                    logging.warning(f"Unexpected tuple length {len(new_result)} for {column} in final pass (no original), not adding.")
                except Exception as e:
                    logging.error(f"Error in final validation pass: {e}")
                    logging.error(traceback.format_exc())
            except Exception as e:
                logging.error(f"Error setting up final validation pass: {e}")
                logging.error(traceback.format_exc())
        
        # Calculate next check date
        next_check_date = None
        next_check_reason = []
        if self.enable_next_check_date and row_results:
            try:
                next_check_date, next_check_reason = await self.determine_next_check_date(row, row_results)
                logging.info(f"Next check date: {next_check_date}")
            except Exception as e:
                logging.error(f"Failed to determine next check date: {e}")
                next_check_date = datetime.now() + timedelta(days=90)
                next_check_reason = [f"Error calculating: {str(e)}"]
        
        logging.info(f"Completed validation for row {row_idx} with {len(row_results)} column results")
        return row_key, row_results, mplex_groups, indiv_validations, next_check_date, next_check_reason

    async def validate_dataframe(self, df: pd.DataFrame):
        """Validate all rows in parallel with context-aware features."""
        # Generate row keys first, to be used consistently throughout
        row_keys = {}
        for i in range(len(df)):
            try:
                pk_values = []
                for pk in self.config.primary_key:
                    pk_value = self._get_primary_key_value(df.iloc[i], pk)
                    pk_values.append(str(pk_value) if not pd.isna(pk_value) else "")
                
                row_key = "|".join(pk_values)
                if not row_key:
                    row_key = f"Row_{i}"
                row_keys[i] = row_key
            except Exception as e:
                logging.error(f"Error generating row key for row {i}: {e}")
                row_keys[i] = f"Row_{i}"
            
        # If we have no primary key values at all, force all keys to be row indices
        if all(key.startswith("Row_") for key in row_keys.values()):
            logging.warning("All primary keys are missing or empty, using row index as primary key for all rows")
            for i in range(len(df)):
                row_keys[i] = f"Row_{i}"
        
        # Create a semaphore to limit concurrent tasks
        semaphore = asyncio.Semaphore(self.max_concurrent_rows)
        results = {}
        next_check_dates = {}
        
        # Prepare copies of the dataframe
        updated_df = df.copy()
        
        # Create 'Last Check' and 'Next Check Date' columns if they don't exist
        if 'Last Check' not in updated_df.columns:
            updated_df['Last Check'] = pd.NaT
        if 'Next Check Date' not in updated_df.columns:
            updated_df['Next Check Date'] = pd.NaT
        
        # Process rows in batches for better progress reporting
        total_rows = len(df)
        processed_rows = 0
        max_batch_size = self.max_concurrent_rows
        
        # Calculate total batches for logging
        total_batches = (total_rows + max_batch_size - 1) // max_batch_size
        current_batch = 1
        
        # Split into batches
        for batch_start in range(0, total_rows, max_batch_size):
            batch_end = min(batch_start + max_batch_size, total_rows)
            batch_size = batch_end - batch_start
            
            logging.info(f"Processing batch {current_batch}/{total_batches} with {batch_size} rows")
            
            # Create tasks for this batch
            tasks = []
            for i in range(batch_start, batch_end):
                row = df.iloc[i]
                row_key = row_keys[i]
                
                # Create a task that will release the semaphore when done
                tasks.append(self._process_row_with_semaphore(semaphore, row, row_key, i))
            
            # Wait for all tasks in this batch to complete
            batch_results = await asyncio.gather(*tasks)
            
            # Process batch results
            for row_key, row_results, mplex_groups, indiv_validations, next_check_date, next_check_reason in batch_results:
                results[row_key] = row_results
                if next_check_date:
                    next_check_dates[row_key] = (next_check_date, next_check_reason)
                
                # Track validation statistics
                self.total_multiplex_groups += mplex_groups
                self.total_individual_validations += indiv_validations
            
            processed_rows += batch_size
            current_batch += 1
        
        # Update 'Last Check' and 'Next Check Date' columns
        now = datetime.now()
        
        # Store the row_keys for use in formatting later
        self._row_keys = row_keys
        
        for i, row_key in row_keys.items():
            # Only update rows that were successfully validated
            if row_key in results:
                # Update Last Check to now
                updated_df.at[i, 'Last Check'] = now
                
                # Update Next Check Date if available
                if row_key in next_check_dates:
                    next_date, _ = next_check_dates[row_key]
                    updated_df.at[i, 'Next Check Date'] = next_date
        
        return results, updated_df

    async def _process_row_with_semaphore(self, semaphore, row, row_key, row_idx):
        """Process a row while respecting the semaphore limit."""
        async with semaphore:
            return await self._process_row(row, row_key, row_idx)

    def _get_primary_key_value(self, row, pk):
        """Extract primary key values consistently using normalized column matching."""
        # Direct match
        if pk in row:
            return row[pk]
        
        # Normalized match
        pk_norm = normalize_column_name(pk)
        for col in row.index:
            if normalize_column_name(col) == pk_norm:
                return row[col]
        
        # Not found
        logging.warning(f"Primary key column '{pk}' not found in row")
        return "[Unknown]"

def clean_filename(filename):
    """Remove timestamp from filename if present."""
    # Remove any _YYYYMMDD pattern at the end
    base_name = os.path.splitext(filename)[0]
    if '_' in base_name and base_name.split('_')[-1].isdigit():
        base_name = '_'.join(base_name.split('_')[:-1])
    return base_name + '.xlsx'

def get_latest_excel_file(directory):
    """Find the most recent Excel file in the directory."""
    excel_files = glob.glob(os.path.join(directory, "*.xlsx"))
    if not excel_files:
        raise FileNotFoundError("No Excel files found in the directory")
    
    # Sort by modification time, newest first
    latest_file = max(excel_files, key=os.path.getmtime)
    return latest_file

# Add a function to directly find cache files in the cache directory
def find_cache_files(cache_dir, column_names, primary_key_value):
    """Find relevant cache files and prompt files in the cache directory."""
    cache_files = {}
    
    try:
        # Get all json files in the cache directory
        files = glob.glob(os.path.join(cache_dir, "*.json"))
        prompt_files = glob.glob(os.path.join(cache_dir, "*_prompt.txt"))
        
        # For each file, check if it contains column names or the primary key
        for file_path in files:
            filename = os.path.basename(file_path)
            # Skip prompt files as we'll handle them separately
            if filename.endswith("_prompt.json"):
                continue
                
            # Try to read the file to check its contents
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    # Check if the content contains column names or primary key
                    matches = False
                    for col in column_names:
                        if f'"column": "{col}"' in content or f'"Column: {col}' in content:
                            matches = True
                            cache_files[col] = filename
                    
                    # Also check if it contains the primary key value
                    if primary_key_value and primary_key_value in content:
                        # This file is relevant to this row
                        for col in column_names:
                            if col not in cache_files:
                                cache_files[col] = filename
            except:
                pass
                
        # Find matching prompt files
        prompt_lookup = {}
        for file_path in prompt_files:
            filename = os.path.basename(file_path)
            file_id = filename.split('_')[0]  # Get the ID part
            
            # See if we can find a matching cache file
            for col, cache_file in cache_files.items():
                if cache_file.startswith(file_id):
                    prompt_lookup[col] = filename
                    
            # If we couldn't match by ID, try content matching
            if not prompt_lookup:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        # Check if content contains column names
                        for col in column_names:
                            if f'Column: {col}' in content and col not in prompt_lookup:
                                prompt_lookup[col] = filename
                except:
                    pass
    
    except Exception as e:
        logging.error(f"Error finding cache files: {e}")
    
    return cache_files, prompt_lookup

async def test_parallel_multiplex(input_file, api_key, concurrent_rows, max_rows=None, debug=False):
    """Test parallel validation with enhanced multiplex capabilities."""
    start_time = time.time()
    
    # Configure logging to save logs alongside the output file
    output_dir = os.path.dirname(input_file)
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(output_dir, f"{base_name}_parallel_multiplex_{concurrent_rows}_{timestamp}.log")
    
    # Remove existing file handlers and add our new one
    for handler in logging.root.handlers[:]:
        if isinstance(handler, logging.FileHandler):
            logging.root.removeHandler(handler)
    
    # Add file handler that writes to the same directory as the output file
    file_handler = logging.FileHandler(log_file, encoding='utf-8')  # Add UTF-8 encoding
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logging.root.addHandler(file_handler)
    
    logging.info(f"Testing parallel multiplex validation with {concurrent_rows} concurrent rows")
    logging.info(f"Logging to {log_file}")
    
    # Apply patches to validator methods
    patch_validator_methods()
    
    # Add a log handler to capture cache information
    class CacheLogHandler(logging.Handler):
        def emit(self, record):
            msg = self.format(record)
            if "Cache HIT" in msg or "Cache STORED" in msg or "Prompt saved to" in msg:
                recent_cache_logs.append(msg)
                if len(recent_cache_logs) > 100:
                    recent_cache_logs.pop(0)
    
    # Add cache log handler
    cache_handler = CacheLogHandler()
    cache_handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(message)s')
    cache_handler.setFormatter(formatter)
    logging.getLogger().addHandler(cache_handler)
    
    # Set debug logging level if requested
    if debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logging.debug("Debug logging enabled")
    
    # Load the spreadsheet directly (no temporary copy)
    try:
        logging.info(f"Loading data from {input_file}")
        df = pd.read_excel(input_file)
        logging.info(f"Loaded {len(df)} rows from {input_file}")
    except Exception as e:
        logging.error(f"Error loading file: {e}")
        return None, None, None
    
    # Limit number of rows if requested
    if max_rows and max_rows < len(df):
        df = df.iloc[0:max_rows].copy()
        logging.info(f"Limited to {len(df)} rows for testing")
    
    # Create configuration
    config = Config.from_table_directory(input_file, "Main View")
    if config is None:
        logging.error("Column config needs to be filled in. Please edit column_config.yml and try again.")
        return
    
    config.recheck.force_recheck = True
    
    # Use this for column validation
    # VALIDATION: Check for mismatches between config columns and input file columns
    logging.info("Validating column configuration against input file...")
    
    # Get raw columns from both sources for detailed comparison
    excel_columns = list(df.columns)
    config_columns = [target.column for target in config.validation_targets]
    
    # Print detailed information for debugging
    logging.info("=== COLUMN VALIDATION ===")
    logging.info(f"Excel file has {len(excel_columns)} columns: {excel_columns}")
    logging.info(f"Config has {len(config_columns)} columns: {config_columns}")
    
    # Create normalized maps for better matching
    excel_columns_norm = [normalize_column_name(col) for col in excel_columns]
    config_columns_norm = [normalize_column_name(col) for col in config_columns]
    
    # Create maps for lookup
    excel_columns_map = {normalize_column_name(col): col for col in excel_columns}
    config_columns_map = {normalize_column_name(col): col for col in config_columns}
    
    # Display normalized column names for debugging
    logging.info(f"Normalized Excel columns: {excel_columns_norm}")
    logging.info(f"Normalized Config columns: {config_columns_norm}")
    
    # Create a validation target lookup by normalized name
    validation_targets_by_norm = {normalize_column_name(target.column): target for target in config.validation_targets}
    
    # Find exact column matches
    exact_matches = [col for col in excel_columns if col in config_columns]
    logging.info(f"Found {len(exact_matches)} exact column matches")
    
    # Find columns that match after normalization
    normalized_matches = []
    for excel_col in excel_columns:
        excel_norm = normalize_column_name(excel_col)
        if excel_norm in config_columns_map and excel_col not in exact_matches:
            normalized_matches.append((excel_col, config_columns_map[excel_norm]))
    
    logging.info(f"Found {len(normalized_matches)} additional columns that match after normalization:")
    for excel_col, config_col in normalized_matches:
        logging.info(f"  - '{excel_col}' matches '{config_col}' after normalization")
    
    # Find columns in config but not in file (even after normalization)
    missing_in_file = []
    for config_col in config_columns:
        config_norm = normalize_column_name(config_col)
        if config_norm not in excel_columns_map:
            missing_in_file.append(config_col)
    
    # Find columns in file but not in config (even after normalization)
    missing_in_config = []
    for excel_col in excel_columns:
        if excel_col in ["Last Check", "Next Check Date"]:
            continue
        excel_norm = normalize_column_name(excel_col)
        if excel_norm not in config_columns_map:
            missing_in_config.append(excel_col)
    
    # Report detailed findings
    if exact_matches:
        logging.info(f"EXACT MATCHES ({len(exact_matches)}):")
        for col in exact_matches:
            target = next((t for t in config.validation_targets if t.column == col), None)
            importance = target.importance.value if target else "N/A"
            logging.info(f"  - '{col}' (Importance: {importance})")
    
    if normalized_matches:
        logging.warning(f"FUZZY MATCHES ({len(normalized_matches)}):")
        for excel_col, config_col in normalized_matches:
            target = next((t for t in config.validation_targets if t.column == config_col), None)
            importance = target.importance.value if target else "N/A"
            logging.warning(f"  - Excel: '{excel_col}' ≈ Config: '{config_col}' (Importance: {importance})")
    
    if missing_in_file:
        error_msg = f"COLUMNS IN CONFIG BUT MISSING IN EXCEL ({len(missing_in_file)}):\n"
        for col in missing_in_file:
            target = next((t for t in config.validation_targets if t.column == col), None)
            importance = target.importance.value if target else "N/A"
            error_msg += f"  - '{col}' (Importance: {importance})\n"
        logging.error(error_msg)
    
    if missing_in_config:
        warning_msg = f"COLUMNS IN EXCEL BUT MISSING IN CONFIG ({len(missing_in_config)}):\n"
        for col in missing_in_config:
            warning_msg += f"  - '{col}'\n"
        logging.warning(warning_msg)
    
    # Check for primary key column mismatch using normalized comparison
    missing_primary_key = []
    for pk in config.primary_key:
        pk_norm = normalize_column_name(pk)
        found = False
        
        # First check for exact match
        if pk in excel_columns:
            found = True
            
        # Then check for normalized match
        if not found:
            for excel_col in excel_columns:
                if normalize_column_name(excel_col) == pk_norm:
                    found = True
                    # Log the mapping for debugging
                    logging.info(f"Primary key '{pk}' fuzzy-matched to '{excel_col}'")
                    break
                    
        if not found:
            missing_primary_key.append(pk)
    
    if missing_primary_key:
        error_msg = f"CRITICAL ERROR: Primary key columns missing in Excel file:\n"
        for pk in missing_primary_key:
            error_msg += f"  - '{pk}'\n"
        error_msg += f"\nPrimary key in config: {config.primary_key}\n"
        error_msg += f"Available columns in file: {excel_columns}\n"
        logging.error(error_msg)
        logging.error("Cannot proceed with validation due to missing primary key columns.")
        logging.error("Please ensure your Excel file contains all primary key columns specified in column_config.yml.")
        return None, None, None
    
    # Check if there are critical columns that have fuzzy matches
    critical_fuzzy_matches = []
    for excel_col, config_col in normalized_matches:
        target = next((t for t in config.validation_targets if t.column == config_col), None)
        if target and target.importance == ColumnImportance.CRITICAL:
            critical_fuzzy_matches.append((excel_col, config_col))
    
    if critical_fuzzy_matches:
        warning_msg = f"WARNING: The following CRITICAL columns have formatting differences:\n"
        for excel_col, config_col in critical_fuzzy_matches:
            warning_msg += f"  - Excel: '{excel_col}' ≈ Config: '{config_col}'\n"
        logging.warning(warning_msg)
        logging.warning("These critical columns will be validated using fuzzy matching, which may cause issues.")
    
    # Trim any whitespace from the API key
    api_key = api_key.strip()
    logging.info(f"API key length: {len(api_key)}")
    
    # Set up cache
    table_dir = os.path.dirname(input_file)
    cache_dir = os.path.join(table_dir, ".perplexity_cache")
    os.makedirs(cache_dir, exist_ok=True)
    
    # List the existing cache files for reference
    cache_files = []
    prompt_files = []
    try:
        for f in os.listdir(cache_dir):
            if f.endswith(".json") and not f.endswith("_prompt.json"):
                cache_files.append(f)
            elif f.endswith("_prompt.txt"):
                prompt_files.append(f)
        logging.info(f"Found {len(cache_files)} cache files and {len(prompt_files)} prompt files in {cache_dir}")
        if debug:
            logging.debug(f"Cache files: {cache_files[:5]}...")
            logging.debug(f"Prompt files: {prompt_files[:5]}...")
    except Exception as e:
        logging.warning(f"Could not list cache directory contents: {e}")
    
    # Initialize the enhanced parallel validator
    logging.info(f"Initializing enhanced parallel validator with {concurrent_rows} concurrent rows")
    validator = EnhancedParallelValidator(config, api_key, max_concurrent_rows=concurrent_rows)
    
    # Run validation
    try:
        results, updated_df = await validator.validate_dataframe(df)
        
        # Calculate stats
        total_time = time.time() - start_time
        avg_time_per_row = total_time / len(df) if len(df) > 0 else 0
        
        # Log results
        logging.info(f"\n=== VALIDATION SUMMARY ===")
        logging.info(f"Total rows processed: {len(df)}")
        logging.info(f"Total time: {total_time:.2f} seconds")
        logging.info(f"Average time per row: {avg_time_per_row:.2f} seconds")
        
        # Expected sequential time (rough estimate)
        est_sequential_time = avg_time_per_row * len(df) * (1.0 / concurrent_rows) * 0.8
        logging.info(f"Estimated sequential time: {est_sequential_time:.2f} seconds")
        
        if est_sequential_time > 0:
            speedup = est_sequential_time / total_time
            logging.info(f"Estimated speedup: {speedup:.2f}x")
        
        # Save results to Excel with formatting
        output_file = os.path.join(output_dir, f"{base_name}_parallel_multiplex_{concurrent_rows}_{timestamp}.xlsx")
        
        # Save dataframe first with original values
        original_df = df.copy()
        updated_df.to_excel(output_file, index=False)
        
        # Now add formatting
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        ws.title = "Main View"
        
        # Create a dictionary mapping confidence levels to colors
        colors = {
            "HIGH": "90EE90",  # Light green
            "MEDIUM": "FFD700",  # Gold
            "LOW": "FFB6C1",  # Light pink
            "N/A": "D3D3D3"  # Light gray for no result
        }
        
        # Create a bidirectional mapping between normalized and actual column names 
        norm_to_actual = {}  # Maps normalized column names to actual column names
        for col in excel_columns:
            norm = normalize_column_name(col)
            if norm not in norm_to_actual:
                norm_to_actual[norm] = col
            else:
                # If there's a conflict, prefer the exact match with config
                if col in config_columns:
                    norm_to_actual[norm] = col
        
        # Create a reverse mapping
        actual_to_norm = {v: k for k, v in norm_to_actual.items()}
        
        # Generate row_keys with enhanced handling
        row_keys = {}
        for i in range(len(df)):
            try:
                pk_values = []
                for pk in config.primary_key:
                    pk_value = get_primary_key_value(df.iloc[i], pk)
                    pk_values.append(str(pk_value) if not pd.isna(pk_value) else "")
                
                row_key = "|".join(pk_values)
                if not row_key:
                    row_key = f"Row_{i}"  # Fallback if all primary keys are empty
                row_keys[i] = row_key
            except Exception as e:
                logging.error(f"Error generating row key for row {i}: {e}")
                row_keys[i] = f"Row_{i}"  # Fallback
        
        # If we have no primary key values at all, force all keys to be row indices
        if all(key.startswith("Row_") for key in row_keys.values()):
            logging.warning("All primary keys are missing or empty, using row index as primary key for all rows")
            for i in range(len(df)):
                row_keys[i] = f"Row_{i}"
        
        # Apply formatting to Main View - COMPLETELY REWRITTEN FOR RELIABILITY
        logging.info("Applying formatting to Main View...")
        
        # Use validator's stored row keys if available
        formatted_row_keys = validator._row_keys if hasattr(validator, '_row_keys') else row_keys
        
        # Debug log to check what results we have
        logging.debug(f"Results keys: {list(results.keys())[:5]}")
        logging.debug(f"Row keys (sample): {[formatted_row_keys.get(i) for i in range(min(5, len(df)))]}")
        
        # Apply formatting to Main View - COMPLETELY REWRITTEN FOR RELIABILITY
        logging.info("Applying formatting to Main View...")
        cells_formatted = 0
        for row_idx in range(len(df)):
            try:
                # Get current row key
                row_key = formatted_row_keys.get(row_idx, f"Row_{row_idx}")
                
                # Get results for this row
                row_results = results.get(row_key, {})
                
                # Debug output for row key mapping issues
                if not row_results and row_idx < 5:  # Only log for first few rows to avoid spam
                    logging.warning(f"No validation results for row {row_idx} ({row_key})")
                    available_keys = list(results.keys())[:10]  # Show at most 10 keys
                    logging.debug(f"Available result keys: {available_keys}")
                    continue
                
                # Process each column in this row
                for col_idx, col_name in enumerate(df.columns):
                    # Skip Last Check and Next Check Date columns
                    if col_name in ["Last Check", "Next Check Date"]:
                        continue
                    
                    # Get the cell in the Excel worksheet (add 2 because of headers and 1-indexing)
                    excel_row = row_idx + 2
                    excel_col = col_idx + 1
                    
                    try:
                        cell = ws.cell(row=excel_row, column=excel_col)
                        
                        # Get normalized column name for matching
                        col_name_norm = normalize_column_name(col_name)
                        
                        # Check if we have validation results for this column (using normalized keys)
                        result_col = None
                        for result_key in row_results.keys():
                            if normalize_column_name(result_key) == col_name_norm:
                                result_col = result_key
                                break
                                
                        if result_col:
                            result = row_results[result_col]
                            
                            # Only process valid result tuples
                            if isinstance(result, tuple) and len(result) >= 4:
                                # Get result components
                                value = result[0]
                                confidence_level = result[3]
                                
                                # Always update the cell value
                                cell.value = value
                                
                                # Get original value for comparison
                                original_value = original_df.iloc[row_idx][col_name] if col_name in original_df.columns else None
                                
                                # Check if there's a change
                                changed = False
                                if original_value is not None and value is not None:
                                    str_original = str(original_value).strip() if not pd.isna(original_value) else ""
                                    str_value = str(value).strip() if not pd.isna(value) else ""
                                    changed = str_original != str_value
                                
                                # ALWAYS apply confidence-based coloring
                                color = colors.get(confidence_level, colors["N/A"])
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color=color,
                                    end_color=color,
                                    fill_type="solid"
                                )
                                
                                # Make changed cells bold
                                if changed:
                                    cell.font = openpyxl.styles.Font(bold=True)
                                
                                # Add comment with quotes and sources
                                if len(result) >= 6 and (result[4] or result[5]):
                                    quote = result[4] or ""
                                    source = result[5] or ""
                                    comment_text = ""
                                    if quote:
                                        comment_text += f"Quote: {quote}\n\n"
                                    if source:
                                        comment_text += f"Source: {source}"
                                    
                                    if comment_text:
                                        cell.comment = openpyxl.comments.Comment(comment_text, "Validation")
                                        
                                logging.info(f"Formatted cell {excel_row}:{excel_col} ({col_name}) with confidence {confidence_level}")
                                cells_formatted += 1
                    
                    except Exception as cell_e:
                        logging.error(f"Error formatting cell {excel_row}:{excel_col} ({col_name}): {cell_e}")
            
            except Exception as row_e:
                logging.error(f"Error processing formatting for row {row_idx}: {row_e}")
        
        logging.info(f"Formatted {cells_formatted} cells in Main View")
        
        # Create and format Detailed View - COMPLETELY REWRITTEN FOR RELIABILITY
        logging.info("Creating Detailed View...")
        ws_details = wb.create_sheet("Detailed View")
        headers = ["Row", "Primary Key", "Column", "Original Value", "Answer", "Change Required", 
                  "Confidence", "Raw Confidence", "Sources", "Quote", "Source URL", 
                  "Column Type", "Column Importance", "Cache File", "Prompt File", "Timestamp"]
        ws_details.append(headers)
        
        # Use validator's stored row keys if available
        formatted_row_keys = validator._row_keys if hasattr(validator, '_row_keys') else row_keys
        
        # Create maps for looking up columns by normalized names
        excel_columns_normalized = {normalize_column_name(col): col for col in df.columns}
        config_columns_normalized = {}
        for target in config.validation_targets:
            normalized = normalize_column_name(target.column)
            config_columns_normalized[normalized] = target.column
            
        # Create validation target lookup by normalized name
        validation_targets_by_norm = {normalize_column_name(target.column): target for target in config.validation_targets}
        
        # Collect detailed rows with better error handling
        detailed_rows = []
        for row_idx in range(len(original_df)):
            try:
                # Get the row data and key
                row = original_df.iloc[row_idx]
                row_key = formatted_row_keys.get(row_idx, f"Row_{row_idx}")
                
                # Get validation results for this row - debug for first few rows
                row_results = results.get(row_key, {})
                if not row_results and row_idx < 5:
                    logging.warning(f"No validation results found for row {row_idx} with key '{row_key}'")
                    logging.debug(f"Available results keys: {list(results.keys())[:10]}")
                
                # Get cache files info
                row_columns = list(row_results.keys()) + list(original_df.columns)
                cache_files, prompt_files = find_cache_files(cache_dir, row_columns, row_key)
                
                # Add all columns to detailed view
                for col_name in original_df.columns:
                    if col_name in ["Last Check", "Next Check Date"]:
                        continue
                    
                    # Get original value
                    original_value = row[col_name] if col_name in row.index else None
                    
                    # Convert to string with proper handling of NaN
                    original_value_str = str(original_value) if not pd.isna(original_value) else ""
                    
                    # Get normalized column name for matching
                    col_name_norm = normalize_column_name(col_name)
                    
                    # Find matching result column using normalized names
                    result_col = None
                    result_value = None
                    for result_key, result_data in row_results.items():
                        result_key_norm = normalize_column_name(result_key)
                        if result_key_norm == col_name_norm:
                            result_col = result_key
                            result_value = result_data
                            break
                    
                    # Debug output to help diagnose issues
                    if row_idx < 2 and col_name in ["Indication", "Therapeutic Radionuclide"]:
                        logging.debug(f"Column: {col_name}, Normalized: {col_name_norm}")
                        if result_col:
                            logging.debug(f"  Matched to result key: {result_col}")
                            if isinstance(result_value, tuple):
                                logging.debug(f"  Result type: tuple, length: {len(result_value)}")
                                if len(result_value) > 0:
                                    logging.debug(f"  Result value (first element): {result_value[0]}")
                            else:
                                logging.debug(f"  Result type: {type(result_value)}")
                        else:
                            logging.debug("  No matching result found")
                    
                    # Check if we have validation result for this column (using normalized matching)
                    if result_col and result_value:
                        if isinstance(result_value, tuple) and len(result_value) >= 4:
                            # Extract result components
                            value = result_value[0]
                            confidence = result_value[1] if isinstance(result_value[1], (int, float)) else 0.0
                            citations = result_value[2] if isinstance(result_value[2], list) else []
                            confidence_level = result_value[3]
                            quote = result_value[4] if len(result_value) > 4 else ""
                            source_url = result_value[5] if len(result_value) > 5 else ""
                            
                            # Get cache files for this column
                            cache_file = cache_files.get(result_col, cache_files.get(col_name, ""))
                            prompt_file = prompt_files.get(result_col, prompt_files.get(col_name, ""))
                            
                            # Determine if change is required
                            value_str = str(value) if not pd.isna(value) else ""
                            change_required = "Yes" if original_value_str != value_str else "No"
                            
                            # Get column metadata from validation targets using normalized name
                            target = validation_targets_by_norm.get(col_name_norm)
                            col_type = target.type.value if target else ""
                            col_importance = target.importance.value if target else ""
                        else:
                            # Invalid result tuple, use defaults
                            value = str(result_value) if result_value is not None else ""
                            confidence = 0.0
                            citations = []
                            confidence_level = "N/A"
                            quote = ""
                            source_url = ""
                            cache_file = ""
                            prompt_file = ""
                            change_required = "No"
                            target = validation_targets_by_norm.get(col_name_norm)
                            col_type = target.type.value if target else ""
                            col_importance = target.importance.value if target else ""
                    else:
                        # No validation result for this column
                        value = ""
                        confidence = 0.0
                        citations = []
                        confidence_level = "N/A"
                        quote = ""
                        source_url = ""
                        cache_file = ""
                        prompt_file = ""
                        change_required = "No"
                        target = validation_targets_by_norm.get(col_name_norm)
                        col_type = target.type.value if target else ""
                        col_importance = target.importance.value if target else ""
                    
                    # Add to detailed rows
                    detailed_rows.append({
                        "Row": row_idx + 1,
                        "Primary Key": row_key,
                        "Column": col_name,
                        "Original Value": original_value_str,
                        "Answer": str(value) if not pd.isna(value) else "",
                        "Change Required": change_required,
                        "Confidence": confidence_level,
                        "Raw Confidence": confidence,
                        "Sources": "; ".join(citations) if citations else "",
                        "Quote": quote,
                        "Source URL": source_url,
                        "Column Type": col_type,
                        "Column Importance": col_importance,
                        "Cache File": cache_file,
                        "Prompt File": prompt_file,
                        "Timestamp": datetime.now().isoformat()
                    })
            
            except Exception as row_e:
                logging.error(f"Error creating detailed view for row {row_idx}: {row_e}")
                logging.error(traceback.format_exc())
        
        # Add rows to worksheet
        for drow in detailed_rows:
            try:
                row_data = []
                for header in headers:
                    row_data.append(drow.get(header, ""))
                ws_details.append(row_data)
            except Exception as append_e:
                logging.error(f"Error adding row to detailed view: {append_e}")
        
        logging.info(f"Added {len(detailed_rows)} rows to Detailed View")
        
        # Format column widths
        for i, header in enumerate(headers):
            col_letter = openpyxl.utils.get_column_letter(i + 1)
            if header in ["Cache File", "Prompt File"]:
                ws_details.column_dimensions[col_letter].width = 40
            elif header in ["Original Value", "Answer", "Quote", "Primary Key"]:
                ws_details.column_dimensions[col_letter].width = 30
            elif header in ["Sources", "Source URL"]:
                ws_details.column_dimensions[col_letter].width = 50
            else:
                ws_details.column_dimensions[col_letter].width = 15
        
        # Format the Detailed View with coloring by confidence
        for row in ws_details.iter_rows(min_row=2, max_row=len(detailed_rows)+1):
            try:
                confidence_level = row[6].value  # "Confidence" column
                
                # Color the Answer column (index 4) based on confidence
                if confidence_level == "HIGH":
                    row[4].fill = openpyxl.styles.PatternFill(
                        start_color=colors["HIGH"], end_color=colors["HIGH"], fill_type="solid"
                    )
                elif confidence_level == "MEDIUM":
                    row[4].fill = openpyxl.styles.PatternFill(
                        start_color=colors["MEDIUM"], end_color=colors["MEDIUM"], fill_type="solid"
                    )
                elif confidence_level == "LOW":
                    row[4].fill = openpyxl.styles.PatternFill(
                        start_color=colors["LOW"], end_color=colors["LOW"], fill_type="solid"
                    )
                
                # Highlight rows where changes are required
                if row[5].value == "Yes":  # "Change Required" column
                    for cell in row:
                        cell.font = openpyxl.styles.Font(bold=True)
            
            except Exception as format_e:
                logging.error(f"Error formatting detailed view row: {format_e}")
        
        # Save the workbook
        try:
            wb.save(output_file)
            logging.info(f"Results saved to {output_file}")
        except Exception as save_e:
            logging.error(f"Error saving workbook: {save_e}")
        
        logging.info(f"Parallel validation completed in {total_time:.2f} seconds")
        logging.info(f"Average time per row: {total_time:.2f} seconds")
        
        # Calculate validation efficiency
        total_columns = len(config.validation_targets)
        total_validations = validator.total_multiplex_groups + validator.total_individual_validations
        efficiency = 0.0
        if total_columns > 0 and total_validations > 0:
            # Efficiency = columns validated / API calls made
            efficiency = (total_columns * len(df)) / total_validations * 100.0
        
        batch_count = (len(df) + concurrent_rows - 1) // concurrent_rows
        logging.info(f"Processed in {batch_count} parallel batches")
        
        # Find the earliest next check date
        earliest_date = None
        for row_key, results_dict in results.items():
            row_idx = next((i for i, k in row_keys.items() if k == row_key), None)
            if row_idx is None:
                continue
            
            next_check = updated_df.loc[row_idx, 'Next Check Date'] if 'Next Check Date' in updated_df.columns else None
            if next_check and (earliest_date is None or next_check < earliest_date):
                earliest_date = next_check
        
        # Print summary
        print("\nValidation Run Summary")
        print(f"Total rows processed: {len(df)}")
        print(f"Multiplex validation groups: {validator.total_multiplex_groups}")
        print(f"Individual column validations: {validator.total_individual_validations}")
        print(f"Validation efficiency: {efficiency:.1f}% (higher is better)")
        print(f"\nRecommended Next Run Date: {earliest_date.strftime('%Y-%m-%d') if earliest_date else 'Not determined'}")
        print(f"Note: This is based on the earliest recommended check date across all rows.")
        
        return total_time, len(df), output_file
        
    except Exception as e:
        logging.error(f"Error during validation: {e}")
        logging.error(traceback.format_exc())
        return None, None, None

async def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Test parallel validation with improved multiplex')
    parser.add_argument('--input', '-i', type=str, help='Input Excel file')
    parser.add_argument('--concurrent', '-c', type=int, default=5, help='Number of concurrent rows to process')
    parser.add_argument('--rows', '-r', type=int, help='Maximum number of rows to process')
    parser.add_argument('--api-key', '-k', type=str, help='Perplexity API key (defaults to PERPLEXITY_API_KEY env var)')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    args = parser.parse_args()
    
    # Set debug logging if requested
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logging.debug("Debug logging enabled")
    else:
        # Ensure logging is at INFO level if --debug is not set
        logging.getLogger().setLevel(logging.INFO)
    
    # Get API key from argument or environment variable
    api_key = args.api_key
    if not api_key:
        api_key = os.environ.get("PERPLEXITY_API_KEY")
        if not api_key:
            api_key = input("Enter your Perplexity API key: ")
    
    # Find input file if not specified
    input_file = args.input
    if not input_file:
        # Find the most recent Excel file in the tables directory
        tables_dir = os.path.join("..", "tables", "CongressesMasterList")
        try:
            input_file = get_latest_excel_file(tables_dir)
            logging.info(f"Using most recent Excel file: {input_file}")
        except Exception as e:
            logging.error(f"Error finding Excel file: {e}")
            input_file = input("Enter path to input Excel file: ")
    
    if not os.path.exists(input_file):
        logging.error(f"Input file does not exist: {input_file}")
        return
        
    # Run the parallel multiplex test
    total_time, rows_processed, output_file = await test_parallel_multiplex(
        input_file, api_key, args.concurrent, args.rows, args.debug
    )
    
    if total_time and rows_processed:
        logging.info(f"Processed {rows_processed} rows in {total_time:.2f} seconds")
        logging.info(f"Average time per row: {total_time/rows_processed:.2f} seconds")
        logging.info(f"Results saved to {output_file}")
    else:
        logging.error("Validation failed or was interrupted")
        
if __name__ == "__main__":
    asyncio.run(main())
