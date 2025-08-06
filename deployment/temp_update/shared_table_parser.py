#!/usr/bin/env python3
"""
Consolidated S3 Table Parser
Unified parser for CSV/Excel files from S3 used by both interface and validation code
"""

import boto3
import csv
import json
import logging
import os
import tempfile
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)

# Try to import openpyxl
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel files won't be supported")

class S3TableParser:
    """Unified S3 table-to-JSON parser used by both interface and validation code"""
    
    def __init__(self):
        self.s3_client = boto3.client('s3')
        self.logger = logging.getLogger(__name__)
    
    def parse_s3_table(self, bucket: str, key: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Parse Excel/CSV from S3 into standard JSON format
        
        Args:
            bucket: S3 bucket name
            key: S3 object key
            sheet_name: Excel sheet name (optional, uses first sheet if not specified)
            
        Returns:
            Structured table data with metadata
        """
        try:
            # Download file from S3
            self.logger.info(f"Downloading table from S3: s3://{bucket}/{key}")
            response = self.s3_client.get_object(Bucket=bucket, Key=key)
            file_content = response['Body'].read()
            
            # Determine file type from key
            filename = key.split('/')[-1]
            if filename.endswith('.csv'):
                return self._parse_csv_content(file_content, filename)
            elif filename.endswith(('.xlsx', '.xls')):
                if not OPENPYXL_AVAILABLE:
                    raise ImportError("openpyxl is required for Excel file support")
                return self._parse_excel_content(file_content, filename, sheet_name)
            else:
                raise ValueError(f"Unsupported file format: {filename}")
                
        except Exception as e:
            self.logger.error(f"Failed to parse S3 table {bucket}/{key}: {str(e)}")
            raise
    
    def _parse_csv_content(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Parse CSV content into structured format"""
        try:
            # Decode content
            text_content = file_content.decode('utf-8')
            
            # Parse CSV
            reader = csv.reader(text_content.splitlines())
            rows = list(reader)
            
            if not rows:
                raise ValueError("Empty CSV file")
            
            # Extract headers and data
            headers = rows[0]
            data_rows = rows[1:]
            
            # Clean headers
            clean_headers = [str(header).strip() if header else f"Column_{i+1}" 
                           for i, header in enumerate(headers)]
            
            # Convert to structured format
            structured_data = []
            for row in data_rows:
                if any(cell.strip() for cell in row):  # Skip empty rows
                    row_dict = {}
                    for i, header in enumerate(clean_headers):
                        cell_value = row[i] if i < len(row) else ""
                        row_dict[header] = str(cell_value).strip() if cell_value else ""
                    structured_data.append(row_dict)
            
            return {
                'filename': filename,
                'total_rows': len(structured_data),
                'total_columns': len(clean_headers),
                'column_names': clean_headers,
                'data': structured_data,
                'metadata': {
                    'file_type': 'csv',
                    'source': 's3',
                    'sample_rows': min(5, len(structured_data))
                }
            }
            
        except Exception as e:
            self.logger.error(f"Failed to parse CSV content: {str(e)}")
            raise
    
    def _parse_excel_content(self, file_content: bytes, filename: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Parse Excel content into structured format"""
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_file_path = temp_file.name
            
            try:
                # Load workbook
                workbook = load_workbook(temp_file_path, read_only=True)
                
                # Select worksheet
                if sheet_name:
                    if sheet_name not in workbook.sheetnames:
                        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
                    worksheet = workbook[sheet_name]
                else:
                    worksheet = workbook.active
                    sheet_name = worksheet.title
                
                # Read data
                all_rows = list(worksheet.iter_rows(values_only=True))
                
                if not all_rows:
                    raise ValueError("Empty Excel worksheet")
                
                # Extract headers and data
                headers = all_rows[0]
                data_rows = all_rows[1:]
                
                # Clean headers
                clean_headers = [str(header).strip() if header is not None else f"Column_{i+1}" 
                               for i, header in enumerate(headers)]
                
                # Convert to structured format
                structured_data = []
                for row in data_rows:
                    if row and any(cell is not None for cell in row):  # Skip empty rows
                        row_dict = {}
                        for i, header in enumerate(clean_headers):
                            cell_value = row[i] if i < len(row) else None
                            row_dict[header] = str(cell_value).strip() if cell_value is not None else ""
                        structured_data.append(row_dict)
                
                workbook.close()
                
                return {
                    'filename': filename,
                    'total_rows': len(structured_data),
                    'total_columns': len(clean_headers),
                    'column_names': clean_headers,
                    'data': structured_data,
                    'metadata': {
                        'file_type': 'excel',
                        'sheet_name': sheet_name,
                        'source': 's3',
                        'sample_rows': min(5, len(structured_data))
                    }
                }
                
            finally:
                # Clean up temp file
                os.unlink(temp_file_path)
                
        except Exception as e:
            self.logger.error(f"Failed to parse Excel content: {str(e)}")
            raise
    
    def get_table_sample(self, bucket: str, key: str, max_rows: int = 10) -> Dict[str, Any]:
        """Get a sample of the table data without loading the entire file"""
        try:
            full_data = self.parse_s3_table(bucket, key)
            
            # Return sample
            sample_data = full_data['data'][:max_rows]
            
            return {
                'filename': full_data['filename'],
                'total_rows': full_data['total_rows'],
                'total_columns': full_data['total_columns'],
                'column_names': full_data['column_names'],
                'sample_data': sample_data,
                'metadata': full_data['metadata']
            }
            
        except Exception as e:
            self.logger.error(f"Failed to get table sample: {str(e)}")
            raise
    
    def analyze_table_structure(self, bucket: str, key: str) -> Dict[str, Any]:
        """Analyze table structure for AI config generation"""
        try:
            # Get sample data
            sample = self.get_table_sample(bucket, key, max_rows=20)
            
            # Analyze each column
            column_analysis = {}
            for col_name in sample['column_names']:
                # Get values for this column
                values = [row.get(col_name, '') for row in sample['sample_data']]
                non_empty_values = [v for v in values if v.strip()]
                
                column_analysis[col_name] = {
                    'data_type': self._infer_data_type(non_empty_values),
                    'non_null_count': len(non_empty_values),
                    'unique_count': len(set(non_empty_values)),
                    'sample_values': list(set(non_empty_values))[:5],  # Up to 5 unique samples
                    'fill_rate': len(non_empty_values) / len(values) if values else 0.0
                }
            
            return {
                'basic_info': {
                    'filename': sample['filename'],
                    'shape': [sample['total_rows'], sample['total_columns']],
                    'column_names': sample['column_names'],
                    'sample_rows_analyzed': len(sample['sample_data'])
                },
                'column_analysis': column_analysis,
                'domain_info': self._infer_domain_info(sample),
                'metadata': sample['metadata']
            }
            
        except Exception as e:
            self.logger.error(f"Failed to analyze table structure: {str(e)}")
            raise
    
    def _infer_data_type(self, values: List[str]) -> str:
        """Infer data type from sample values"""
        if not values:
            return "Unknown"
        
        # Check for numbers
        numeric_count = 0
        for value in values:
            try:
                float(value)
                numeric_count += 1
            except ValueError:
                pass
        
        if numeric_count / len(values) > 0.8:
            return "Numeric"
        
        # Check for dates (simple heuristic)
        date_keywords = ['date', 'time', '/', '-', '20', '19']
        date_count = sum(1 for value in values if any(kw in value.lower() for kw in date_keywords))
        
        if date_count / len(values) > 0.5:
            return "Date"
        
        return "Text"
    
    def _infer_domain_info(self, sample: Dict[str, Any]) -> Dict[str, Any]:
        """Infer domain information from table structure"""
        column_names = [name.lower() for name in sample['column_names']]
        
        # Domain keywords
        domain_scores = {}
        
        # Biotech/pharma keywords
        biotech_keywords = ['target', 'indication', 'phase', 'compound', 'drug', 'protein', 'gene']
        biotech_score = sum(1 for col in column_names if any(kw in col for kw in biotech_keywords))
        if biotech_score > 0:
            domain_scores['biotech'] = biotech_score
        
        # Competitive intelligence keywords
        ci_keywords = ['company', 'competitor', 'product', 'market', 'revenue', 'stage']
        ci_score = sum(1 for col in column_names if any(kw in col for kw in ci_keywords))
        if ci_score > 0:
            domain_scores['competitive_intelligence'] = ci_score
        
        # Financial keywords
        financial_keywords = ['price', 'cost', 'revenue', 'profit', 'budget', 'finance']
        financial_score = sum(1 for col in column_names if any(kw in col for kw in financial_keywords))
        if financial_score > 0:
            domain_scores['financial'] = financial_score
        
        # Determine likely domain
        if domain_scores:
            likely_domain = max(domain_scores, key=domain_scores.get)
            confidence = domain_scores[likely_domain] / len(column_names)
        else:
            likely_domain = "general"
            confidence = 0.1
        
        return {
            'likely_domain': likely_domain,
            'domain_scores': domain_scores,
            'confidence': confidence
        }

# Global instance for easy imports
s3_table_parser = S3TableParser()