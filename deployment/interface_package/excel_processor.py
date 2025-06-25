"""
Excel processor for validating data and updating Excel files with validation results.

This module provides functionality to:
1. Load an Excel or CSV file and a config file
2. Generate a JSON version of the data
3. Update the Excel with validation results, color-coded by confidence
4. Add quotes as comments
5. Create a second sheet with detailed information for each cell
"""

import os
import pandas as pd
import json
import yaml
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import Dict, List, Any, Tuple, Optional, Union
import logging
from row_key_utils import generate_row_key

# Configure logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Define colors for confidence levels
COLOR_HIGH = "C6EFCE"  # Green
COLOR_MEDIUM = "FFEB9C"  # Yellow
COLOR_LOW = "FFC7CE"  # Red
COLOR_NEUTRAL = "FFFFFF"  # White

class ExcelProcessor:
    """Process Excel files and validation results."""
    
    def __init__(self, input_file: str, config_file: str = None, row_filter: Union[List[int], str] = None):
        """
        Initialize the Excel processor.
        
        Args:
            input_file: Path to the input Excel or CSV file
            config_file: Path to the configuration file (YAML or JSON)
            row_filter: Optional list of row indices to process or a filter expression
        """
        self.input_file = Path(input_file)
        self.config_file = self._find_config_file(config_file)
        self.config = self._load_config()
        self.df = self._load_data()
        self.filtered_df = self._apply_row_filter(row_filter)
        self.details_data = []  # Will store detailed validation results

    def _find_config_file(self, config_file: str = None) -> Optional[Path]:
        """
        Find the configuration file, trying both config.yml and column_config.yml.
        
        Args:
            config_file: Path provided by the user
            
        Returns:
            Path to the configuration file, or None if not found
        """
        if config_file:
            path = Path(config_file)
            if path.exists():
                return path
                
        # Try different config file names in the input directory
        input_dir = self.input_file.parent
        
        # First, check for column_config.yml (used by lambda function)
        column_config_path = input_dir / "column_config.yml"
        if column_config_path.exists():
            logger.info(f"Using lambda-compatible config file: {column_config_path}")
            return column_config_path
            
        # Then, check for config.yml
        config_path = input_dir / "config.yml"
        if config_path.exists():
            logger.info(f"Using config file: {config_path}")
            return config_path
            
        return None

    def _load_config(self) -> Dict:
        """Load configuration from file or use defaults."""
        if not self.config_file or not Path(self.config_file).exists():
            logger.warning("No config file provided or file doesn't exist. Using defaults.")
            return {
                "primary_key": ["id"],
                "columns_to_validate": [],
                "output_directory": str(self.input_file.parent),
                "detailed_view": True
            }
        
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                if Path(self.config_file).suffix.lower() == '.json':
                    config_data = json.load(f)
                else:  # Assume YAML
                    config_data = yaml.safe_load(f)
                
                # Handle two different config formats: standard config and column_config
                if "columns" in config_data:
                    # This is a column_config.yml format used by lambda function
                    return self._convert_column_config_format(config_data)
                else:
                    # This is our standard config.yml format
                    return config_data
                    
        except Exception as e:
            logger.error(f"Failed to load config file: {e}")
            return {
                "primary_key": ["id"],
                "columns_to_validate": [],
                "output_directory": str(self.input_file.parent),
                "detailed_view": True
            }

    def _convert_column_config_format(self, column_config: Dict) -> Dict:
        """
        Convert column_config.yml format to our standard config format.
        
        Args:
            column_config: Configuration in column_config.yml format
            
        Returns:
            Configuration in standard format
        """
        config = {
            "primary_key": column_config.get("primary_key", ["id"]),
            "columns_to_validate": [],
            "column_importance": {},
            "output_directory": str(self.input_file.parent),
            "detailed_view": True,
            "column_formats": {},
            "general_notes": column_config.get("general_notes", "")
        }
        
        # Process columns
        columns = column_config.get("columns", {})
        for column_name, column_data in columns.items():
            # Add to columns_to_validate if not ignored
            if column_data.get("importance") != "ignored":
                config["columns_to_validate"].append(column_name)
            
            # Set importance
            importance_map = {
                "critical": "HIGH",
                "important": "HIGH",
                "interesting": "MEDIUM",
                "nice_to_have": "LOW",
                "ignored": "IGNORED"
            }
            importance = column_data.get("importance", "ignored")
            config["column_importance"][column_name] = importance_map.get(importance, "MEDIUM")
            
            # Set column format
            config["column_formats"][column_name] = {
                "format": column_data.get("format", "String"),
                "description": column_data.get("description", ""),
                "notes": column_data.get("notes", ""),
                "examples": column_data.get("examples", [])
            }
        
        return config

    def _load_data(self) -> pd.DataFrame:
        """Load data from Excel or CSV file."""
        try:
            if self.input_file.suffix.lower() in ['.xlsx', '.xls']:
                return pd.read_excel(self.input_file)
            elif self.input_file.suffix.lower() == '.csv':
                return pd.read_csv(self.input_file)
            else:
                raise ValueError(f"Unsupported file format: {self.input_file.suffix}")
        except Exception as e:
            logger.error(f"Failed to load data file: {e}")
            raise

    def _apply_row_filter(self, row_filter: Union[List[int], str] = None) -> pd.DataFrame:
        """
        Apply filtering to the DataFrame.
        
        Args:
            row_filter: Either a list of row indices or a query string (e.g. "column > value")
            
        Returns:
            Filtered DataFrame
        """
        if row_filter is None:
            return self.df  # Return the full DataFrame if no filter is specified
        
        try:
            if isinstance(row_filter, list):
                # Filter by row indices
                return self.df.iloc[row_filter]
            elif isinstance(row_filter, str):
                # Filter by query string
                return self.df.query(row_filter)
            else:
                logger.warning(f"Unsupported row filter type: {type(row_filter)}. Using full dataset.")
                return self.df
        except Exception as e:
            logger.error(f"Error applying row filter: {e}. Using full dataset.")
            return self.df

    def to_json(self, output_file: str = None) -> str:
        """
        Convert the Excel/CSV data to JSON.
        
        Args:
            output_file: Optional path to save the JSON file
            
        Returns:
            JSON string representation of the data
        """
        # Convert DataFrame to records format (list of dictionaries)
        records = self.filtered_df.to_dict('records')
        
        # Convert to JSON
        json_data = json.dumps(records, indent=2, default=str)
        
        # Save to file if output_file is provided
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(json_data)
            logger.info(f"Saved JSON data to {output_file}")
        
        return json_data

    def update_with_validation(self, validation_results: Dict, output_file: str = None) -> None:
        """
        Update the Excel file with validation results.
        
        Args:
            validation_results: Dictionary of validation results from lambda function
            output_file: Path to save the updated Excel file
        """
        # If no output file is specified, create one in the same directory as the input file
        if not output_file:
            stem = self.input_file.stem
            output_file = self.input_file.parent / f"{stem}_validated.xlsx"

        # Create a workbook and add the main sheet
        wb = openpyxl.Workbook()
        main_sheet = wb.active
        main_sheet.title = "Validated Data"
        
        # Add column headers
        for col_num, col_name in enumerate(self.df.columns, 1):
            cell = main_sheet.cell(row=1, column=col_num, value=col_name)
            cell.font = Font(bold=True)
            
        # Process the validation results and update cells
        data = validation_results.get('data', {})
        
        # Set up detailed view data
        details_rows = []
        details_cols = ["Row Key", "Column", "Original Value", "Validated Value", 
                        "Confidence", "Needs Update", "Quote", "Sources", "Explanation"]
        
        # Process primary key to build row keys
        primary_key = self.config.get('primary_key', ['id'])
        
        # Populate data and apply styles
        for row_idx, df_row in enumerate(self.filtered_df.iterrows(), 2):
            df_row_idx, row_data = df_row
            
            # Build row key
            row_key = generate_row_key(row_data, primary_key)
            
            # Check if we have validation results for this row
            if row_key not in data:
                logger.warning(f"No validation results found for row key: {row_key}")
                
                # Add row values to Excel without validation formatting
                for col_idx, col_name in enumerate(self.df.columns, 1):
                    orig_value = row_data[col_name]
                    cell = main_sheet.cell(row=row_idx, column=col_idx, value=orig_value)
                    cell.fill = PatternFill(start_color=COLOR_NEUTRAL, end_color=COLOR_NEUTRAL, fill_type="solid")
                continue
            
            # Get validation results for this row
            row_results = data[row_key]
            
            # Add row values to Excel
            for col_idx, col_name in enumerate(self.df.columns, 1):
                # Get the original value from DataFrame
                orig_value = row_data[col_name]
                cell = main_sheet.cell(row=row_idx, column=col_idx, value=orig_value)
                
                # Check if this column has validation results
                column_validated = False
                if 'validation_results' in row_results and col_name in row_results['validation_results']:
                    val_result = row_results['validation_results'][col_name]
                    
                    # Get validation details
                    validated_value = val_result.get('answer', '')
                    confidence = val_result.get('confidence', 0)
                    confidence_level = val_result.get('confidence_level', '')
                    quote = val_result.get('quote', '')
                    sources = val_result.get('sources', [])
                    update_required = val_result.get('update_required', False)
                    explanation = val_result.get('explanation', '')
                    
                    # Update the cell value if update is required
                    if update_required and validated_value:
                        cell.value = validated_value
                    
                    # Apply cell formatting based on confidence
                    if confidence_level == "HIGH":
                        cell.fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid")
                    elif confidence_level == "MEDIUM":
                        cell.fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type="solid")
                    elif confidence_level == "LOW":
                        cell.fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type="solid")
                    
                    # Add comment with quote and sources
                    comment_text = ""
                    if quote:
                        comment_text += f"Quote: {quote}\n\n"
                    if sources:
                        comment_text += f"Sources: {', '.join(sources)}\n\n"
                    if update_required:
                        comment_text += f"Update required: Yes\n"
                    else:
                        comment_text += f"Update required: No\n"
                    
                    if comment_text:
                        cell.comment = Comment(comment_text, "Validation")
                    
                    column_validated = True
                    
                    # Add to details data for the detailed view sheet
                    details_rows.append({
                        "Row Key": row_key,
                        "Column": col_name,
                        "Original Value": str(orig_value),
                        "Validated Value": validated_value,
                        "Confidence": confidence,
                        "Needs Update": "Yes" if update_required else "No",
                        "Quote": quote,
                        "Sources": "; ".join(sources),
                        "Explanation": explanation
                    })
                
                # If column was not validated, use neutral styling
                if not column_validated:
                    cell.fill = PatternFill(start_color=COLOR_NEUTRAL, end_color=COLOR_NEUTRAL, fill_type="solid")
        
        # Add a legend for the color coding
        legend_row = len(self.filtered_df) + 3
        main_sheet.cell(row=legend_row, column=1, value="Color Legend:").font = Font(bold=True)
        
        # High confidence
        legend_cell = main_sheet.cell(row=legend_row+1, column=1, value="High Confidence")
        legend_cell.fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid")
        
        # Medium confidence
        legend_cell = main_sheet.cell(row=legend_row+2, column=1, value="Medium Confidence")
        legend_cell.fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type="solid")
        
        # Low confidence
        legend_cell = main_sheet.cell(row=legend_row+3, column=1, value="Low Confidence")
        legend_cell.fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type="solid")
        
        # No validation
        legend_cell = main_sheet.cell(row=legend_row+4, column=1, value="Not Validated")
        legend_cell.fill = PatternFill(start_color=COLOR_NEUTRAL, end_color=COLOR_NEUTRAL, fill_type="solid")
        
        # Auto-adjust column widths
        for col in main_sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 10)  # Add padding
            main_sheet.column_dimensions[column].width = adjusted_width
        
        # Create the detailed view sheet if requested
        if self.config.get('detailed_view', True) and details_rows:
            self._create_detailed_sheet(wb, details_rows, details_cols)
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Updated Excel file saved to {output_file}")

    def _create_detailed_sheet(self, workbook: openpyxl.Workbook, 
                               details_rows: List[Dict], 
                               details_cols: List[str]) -> None:
        """
        Create a detailed view sheet with all validation information.
        
        Args:
            workbook: The openpyxl workbook to add the sheet to
            details_rows: List of dictionaries with detailed validation data
            details_cols: List of column names for the detailed view
        """
        # Create detailed view sheet
        detailed_sheet = workbook.create_sheet(title="Detailed View")
        
        # Add headers
        for col_idx, col_name in enumerate(details_cols, 1):
            cell = detailed_sheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        
        # Add data rows
        for row_idx, row_data in enumerate(details_rows, 2):
            for col_idx, col_name in enumerate(details_cols, 1):
                cell = detailed_sheet.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                
                # Apply conditional formatting based on confidence or update status
                if col_name == "Confidence":
                    confidence = float(row_data.get("Confidence", 0))
                    if confidence >= 0.9:
                        cell.fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid")
                    elif confidence >= 0.7:
                        cell.fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type="solid")
                
                elif col_name == "Needs Update":
                    if row_data.get("Needs Update") == "Yes":
                        cell.fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid")
        
        # Create a table for easier filtering and sorting
        table_range = f"A1:{get_column_letter(len(details_cols))}{len(details_rows) + 1}"
        table = Table(displayName="ValidationDetails", ref=table_range)
        
        # Add a style to the table
        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Add the table to the worksheet
        detailed_sheet.add_table(table)
        
        # Auto-adjust column widths
        for col in detailed_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        max_length = max(max_length, 10)
            adjusted_width = min(max(max_length + 2, 15), 50)  # Add padding but cap at 50
            detailed_sheet.column_dimensions[column].width = adjusted_width
        
        # Set wrap text for quote and sources columns
        for col_name in ["Quote", "Sources", "Explanation"]:
            if col_name in details_cols:
                col_idx = details_cols.index(col_name) + 1
                for row in range(2, len(details_rows) + 2):
                    cell = detailed_sheet.cell(row=row, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

def process_file(input_file: str, config_file: str, validation_json: str, output_file: str = None, row_filter: Union[List[int], str] = None) -> None:
    """
    Process a file with validation results.
    
    Args:
        input_file: Path to the input Excel or CSV file
        config_file: Path to the configuration file (YAML or JSON)
        validation_json: Path to the validation results JSON file or JSON string
        output_file: Path to save the updated Excel file
        row_filter: Optional list of row indices to process or a filter expression
    """
    # Initialize the processor
    processor = ExcelProcessor(input_file, config_file, row_filter)
    
    # Load validation results
    if os.path.exists(validation_json):
        with open(validation_json, 'r', encoding='utf-8') as f:
            validation_results = json.load(f)
    else:
        try:
            validation_results = json.loads(validation_json)
        except json.JSONDecodeError:
            logger.error("Invalid JSON for validation results")
            raise ValueError("Invalid JSON for validation results")
    
    # Update Excel with validation results
    processor.update_with_validation(validation_results, output_file)
    
    # Return the path to the updated file
    if not output_file:
        stem = Path(input_file).stem
        output_file = Path(input_file).parent / f"{stem}_validated.xlsx"
    
    logger.info(f"Processing complete. Output file: {output_file}")

def end_to_end_process(
    input_file: str, 
    config_file: str = None, 
    output_file: str = None, 
    api_key: str = None,
    row_filter: Union[List[int], str] = None,
    use_lambda: bool = True
) -> None:
    """
    End-to-end processing from Excel to validated Excel.
    
    Args:
        input_file: Path to the input Excel or CSV file
        config_file: Path to the configuration file
        output_file: Path to save the updated Excel file
        api_key: API key for Perplexity API
        row_filter: Optional list of row indices to process or a filter expression
        use_lambda: Whether to use Lambda function for validation (default: True)
    """
    import tempfile
    import subprocess
    import time
    import requests
    
    # Initialize the processor
    processor = ExcelProcessor(input_file, config_file, row_filter)
    
    # Create a temporary JSON file
    with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as temp_json:
        temp_json_path = temp_json.name
        json_data = processor.to_json(temp_json_path)
        logger.info(f"Created temporary JSON file: {temp_json_path}")
    
    # Get validation results
    validation_results = {}
    
    try:
        if use_lambda:
            # Check if we should use the Lambda ARN or the Function URL
            lambda_arn = os.environ.get('LAMBDA_FUNCTION_ARN')
            lambda_url = os.environ.get('LAMBDA_FUNCTION_URL')
            
            if lambda_arn:
                # Use boto3 to invoke the Lambda function by ARN
                logger.info(f"Using Lambda function ARN: {lambda_arn}")
                try:
                    import boto3
                    
                    # Create a Lambda client
                    lambda_client = boto3.client('lambda', region_name='us-east-1')
                    
                    # Read the input file
                    with open(temp_json_path, 'r') as f:
                        payload = f.read()
                    
                    # Invoke the Lambda function
                    response = lambda_client.invoke(
                        FunctionName=lambda_arn,
                        InvocationType='RequestResponse',
                        Payload=payload
                    )
                    
                    # Process the response
                    if response['StatusCode'] == 200:
                        import json
                        payload_bytes = response['Payload'].read()
                        validation_results = json.loads(payload_bytes.decode('utf-8'))
                        logger.info("Received validation results from Lambda ARN")
                    else:
                        logger.error(f"Lambda error: {response['StatusCode']}")
                        raise ValueError(f"Lambda error: {response['StatusCode']}")
                except ImportError:
                    logger.error("boto3 not installed. Please install it with: pip install boto3")
                    raise ImportError("boto3 not installed. Please install it with: pip install boto3")
                except Exception as e:
                    logger.error(f"Error invoking Lambda by ARN: {e}")
                    raise
            elif lambda_url:
                # Use HTTP request to invoke the Lambda function URL
                logger.info(f"Using Lambda function URL: {lambda_url}")
                
                # Send request to Lambda
                with open(temp_json_path, 'r') as f:
                    payload = f.read()
                    
                response = requests.post(
                    lambda_url, 
                    data=payload,
                    headers={
                        "Content-Type": "application/json",
                        "x-api-key": api_key or os.environ.get('PERPLEXITY_API_KEY', '')
                    },
                    timeout=180  # Longer timeout for Lambda processing
                )
                
                if response.status_code == 200:
                    validation_results = response.json()
                    logger.info("Received validation results from Lambda URL")
                else:
                    logger.error(f"Lambda error: {response.status_code} - {response.text}")
                    raise ValueError(f"Lambda error: {response.status_code}")
            else:
                # No Lambda configuration found
                logger.error("Neither LAMBDA_FUNCTION_ARN nor LAMBDA_FUNCTION_URL environment variables are set.")
                logger.error("Please set one of these environment variables or use --local flag.")
                raise ValueError("Lambda function configuration not found")
        else:
            # Use local validation script
            logger.info("Using local validation script")
            
            # Check if validator.py exists
            validator_script = Path("validator.py")
            if not validator_script.exists():
                logger.warning("validator.py not found. Looking in src directory...")
                validator_script = Path("src/validator.py")
                
            if not validator_script.exists():
                logger.error("validator.py not found. Cannot proceed with validation.")
                raise FileNotFoundError("validator.py not found")
            
            # Run validation script
            cmd = [
                "python", str(validator_script), 
                "--input", temp_json_path,
                "--output", temp_json_path + ".results"
            ]
            
            if api_key:
                cmd.extend(["--api-key", api_key])
                
            logger.info(f"Running command: {' '.join(cmd)}")
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode != 0:
                logger.error(f"Validation error: {result.stderr}")
                raise ValueError(f"Validation error: {result.stderr}")
            
            # Read validation results
            with open(temp_json_path + ".results", 'r') as f:
                validation_results = json.load(f)
                logger.info("Loaded validation results")
    except Exception as e:
        logger.error(f"Error during validation: {e}")
        raise
    finally:
        # Clean up temporary files
        try:
            os.unlink(temp_json_path)
            if os.path.exists(temp_json_path + ".results"):
                os.unlink(temp_json_path + ".results")
        except:
            pass
    
    # Process validation results
    if validation_results:
        process_file(input_file, config_file, json.dumps(validation_results), output_file, row_filter)
        logger.info("End-to-end processing complete")
    else:
        logger.error("No validation results to process")

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Process Excel/CSV files with validation results")
    parser.add_argument("input_file", help="Path to the input Excel or CSV file")
    parser.add_argument("--config", "-c", help="Path to the configuration file")
    parser.add_argument("--validation", "-v", required=True, help="Path to validation results JSON or JSON string")
    parser.add_argument("--output", "-o", help="Path to save the updated Excel file")
    parser.add_argument("--rows", "-r", help="Row indices to process (e.g. '0,1,2' or '0-5')")
    
    args = parser.parse_args()
    
    # Parse row filter
    row_filter = None
    if args.rows:
        if ',' in args.rows:
            # Comma-separated list of indices
            row_filter = [int(idx) for idx in args.rows.split(',')]
        elif '-' in args.rows:
            # Range of indices
            start, end = args.rows.split('-')
            row_filter = list(range(int(start), int(end) + 1))
        else:
            # Single index
            row_filter = [int(args.rows)]
    
    process_file(args.input_file, args.config, args.validation, args.output, row_filter) 