#!/usr/bin/env python3
"""
Demo Test Orchestrator Script

This script runs end-to-end testing of all demo files in the demos/ directory.
It coordinates preview validation, full validation, and result verification for each demo.

Usage:
    # Run all demos in dev environment
    python test_all_demos.py --demos-dir ./demos --email eliyahu@eliyahu.ai --environment dev

    # Run with custom options
    python test_all_demos.py --demos-dir ./demos --email test@example.com \
        --environment dev --output-dir ./test_results --stop-on-error

    # Skip preview step (only run full validation)
    python test_all_demos.py --demos-dir ./demos --email eliyahu@eliyahu.ai \
        --environment dev --skip-preview

Requirements:
    - Uses dev environment throughout
    - Test email: eliyahu@eliyahu.ai (default)
    - Avoids sending emails (uses preview_email=false flag)
    - Tests all demos in the demos/ directory
"""

import sys
import argparse
import time
import logging
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime

# Import supporting modules
from demo_api_client import DemoAPIClient
from demo_test_reporter import create_report, add_demo_result, save_report, print_report
from demo_session_manager import DemoSessionManager
import requests


class DemoTestOrchestrator:
    """Orchestrates end-to-end testing of all demo files"""

    def __init__(
        self,
        demos_dir: Path,
        email: str,
        environment: str,
        output_dir: Path,
        stop_on_error: bool = True,
        skip_preview: bool = False,
        skip_validation: bool = False
    ):
        """
        Initialize the test orchestrator

        Args:
            demos_dir: Path to demos directory
            email: Test email address
            environment: Environment name (dev, test, staging, prod)
            output_dir: Output directory for test results
            stop_on_error: Stop on first error if True
            skip_preview: Skip preview step if True
            skip_validation: Skip validation step if True
        """
        self.demos_dir = Path(demos_dir)
        self.email = email
        self.environment = environment
        self.output_dir = Path(output_dir)
        self.stop_on_error = stop_on_error
        self.skip_preview = skip_preview
        self.skip_validation = skip_validation

        # Initialize API client - uses demo management API
        self.api_client = DemoAPIClient(email=email)

        # Initialize report
        self.report = create_report(
            email=email,
            environment=environment,
            test_description=f"Automated end-to-end testing of all demos in {demos_dir}"
        )

        # Setup logging
        self._setup_logging()

    def _setup_logging(self):
        """Setup logging configuration"""
        # Create output directory
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Log file path
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = self.output_dir / f"test_all_demos_{timestamp}.log"

        # Configure logging
        logging.basicConfig(
            level=logging.INFO,
            format='[%(asctime)s] [%(levelname)s] %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler(sys.stdout)
            ]
        )

        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Log file: {log_file}")

    def log(self, message: str, level: str = "INFO"):
        """Log a message"""
        if level == "INFO":
            self.logger.info(message)
        elif level == "WARNING":
            self.logger.warning(message)
        elif level == "ERROR":
            self.logger.error(message)
        elif level == "DEBUG":
            self.logger.debug(message)

    def discover_demos(self) -> list:
        """
        Discover all demo folders in the demos directory

        Returns:
            List of demo dictionaries with folder info
        """
        self.log("=" * 70)
        self.log("DEMO DISCOVERY")
        self.log("=" * 70)
        self.log(f"Scanning demos directory: {self.demos_dir}")

        if not self.demos_dir.exists():
            self.log(f"Demos directory does not exist: {self.demos_dir}", "ERROR")
            return []

        # Find all subdirectories
        demo_folders = sorted([d for d in self.demos_dir.iterdir() if d.is_dir()])
        self.log(f"Found {len(demo_folders)} demo folders")

        demos = []
        for demo_folder in demo_folders:
            demo_info = self._analyze_demo_folder(demo_folder)
            if demo_info:
                demos.append(demo_info)

        valid_count = sum(1 for d in demos if d['valid'])
        self.log(f"Valid demos: {valid_count}/{len(demos)}")
        self.log("")

        return demos

    def _analyze_demo_folder(self, demo_folder: Path) -> Optional[Dict[str, Any]]:
        """
        Analyze a demo folder and validate structure

        Args:
            demo_folder: Path to demo folder

        Returns:
            Dictionary with demo info or None if invalid
        """
        demo_name = demo_folder.name
        self.log(f"Analyzing: {demo_name}")

        demo_info = {
            'name': demo_name,
            'folder': demo_folder,
            'valid': False,
            'errors': []
        }

        # Find data file (excluding output files)
        data_files = [
            f for f in demo_folder.glob("*.xlsx")
            if 'output' not in f.name.lower() and 'hyperplexity' not in f.name.lower()
        ]
        data_files.extend([
            f for f in demo_folder.glob("*.csv")
            if 'output' not in f.name.lower()
        ])

        if not data_files:
            demo_info['errors'].append("No data file found")
        elif len(data_files) > 1:
            demo_info['errors'].append(f"Multiple data files found: {[f.name for f in data_files]}")
            demo_info['data_file'] = data_files[0]
        else:
            demo_info['data_file'] = data_files[0]

        # Find config file
        config_files = list(demo_folder.glob("*.json"))

        if not config_files:
            demo_info['errors'].append("No config file found")
        elif len(config_files) > 1:
            demo_info['errors'].append(f"Multiple config files found: {[f.name for f in config_files]}")
            demo_info['config_file'] = config_files[0]
        else:
            demo_info['config_file'] = config_files[0]

        # Find description file
        desc_files = list(demo_folder.glob("*.md"))

        if desc_files:
            demo_info['description_file'] = desc_files[0]
            try:
                with open(desc_files[0], 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                    # Extract first line as display name
                    first_line = content.split('\n')[0]
                    if first_line.startswith('# '):
                        demo_info['display_name'] = first_line[2:].strip()
                    else:
                        demo_info['display_name'] = demo_name
            except Exception as e:
                demo_info['display_name'] = demo_name
                demo_info['errors'].append(f"Error reading description: {e}")
        else:
            demo_info['display_name'] = demo_name

        # Determine if valid
        demo_info['valid'] = (
            len(demo_info['errors']) == 0
            and 'data_file' in demo_info
            and 'config_file' in demo_info
        )

        # Log status
        if demo_info['valid']:
            self.log(f"  [VALID] {demo_info['display_name']}")
            self.log(f"    Data: {demo_info['data_file'].name}")
            self.log(f"    Config: {demo_info['config_file'].name}")
        else:
            self.log(f"  [INVALID] {demo_name}", "WARNING")
            for error in demo_info['errors']:
                self.log(f"    - {error}", "WARNING")

        return demo_info

    def list_available_demos(self):
        """
        List all available demos from the API and map to local demos

        Returns:
            Dictionary mapping display_name to API demo name (S3 folder name)
        """
        self.log("=" * 70)
        self.log("AVAILABLE DEMOS CHECK")
        self.log("=" * 70)

        try:
            result = self.api_client.list_demos()

            if result.get('success'):
                demos = result.get('demos', [])
                self.log(f"Found {len(demos)} available demos on server")

                # Create mapping from display_name to API name (S3 folder name)
                demo_name_map = {}
                for demo in demos:
                    display_name = demo.get('display_name', demo.get('name'))
                    api_name = demo.get('name')  # This is the S3 folder name the API expects
                    demo_name_map[display_name] = api_name
                    self.log(f"  - {display_name} -> {api_name}")
                self.log("")
                return demo_name_map
            else:
                self.log(f"Failed to list demos: {result.get('error', 'Unknown error')}", "WARNING")
                self.log("")
                return {}
        except Exception as e:
            self.log(f"Error listing demos: {str(e)}", "WARNING")
            self.log("")
            return {}

    def test_demo(self, demo: Dict[str, Any]) -> Dict[str, Any]:
        """
        Test a single demo (preview + full validation + download)

        Args:
            demo: Demo info dictionary

        Returns:
            Dictionary with test results
        """
        demo_name = demo['display_name']
        result = {
            'status': 'pending',
            'demo_path': str(demo['folder']),
            'session_id': None,
            'preview_time': 0.0,
            'preview_cost': 0.0,
            'preview_results': None,
            'full_validation_time': 0.0,
            'full_cost': 0.0,
            'full_validation_results': None,
            'download_success': False,
            'download_path': None,
            'error': None,
            'error_details': None
        }

        try:
            self.log("=" * 70)
            self.log(f"TESTING DEMO: {demo_name}")
            self.log("=" * 70)
            self.log(f"Folder: {demo['folder']}")
            self.log(f"Data file: {demo['data_file'].name}")
            self.log(f"Config file: {demo['config_file'].name}")
            self.log("")

            # Step 1: Run preview validation (unless skipped)
            if not self.skip_preview:
                preview_result = self._run_preview(demo)
                result.update(preview_result)

                if not preview_result.get('success', False):
                    result['status'] = 'failed'
                    result['error'] = 'Preview validation failed'
                    result['error_details'] = preview_result.get('error')
                    return result

                # Save session_id from preview for use in validation
                if 'session_id' in preview_result:
                    demo['session_id'] = preview_result['session_id']
                    result['session_id'] = preview_result['session_id']

            # Step 2: Run full validation (unless skipped)
            if not self.skip_validation:
                validation_result = self._run_full_validation(demo)
                result.update(validation_result)

                if not validation_result.get('success', False):
                    result['status'] = 'failed'
                    result['error'] = 'Full validation failed'
                    result['error_details'] = validation_result.get('error')
                    return result

                # Save session ID
                if 'session_id' in validation_result:
                    result['session_id'] = validation_result['session_id']

                # Step 3: Download and verify results
                download_result = self._download_and_verify(demo, validation_result)
                result.update(download_result)

                if not download_result.get('success', False):
                    result['status'] = 'failed'
                    result['error'] = 'Download/verification failed'
                    result['error_details'] = download_result.get('error')
                    return result

            # Success!
            result['status'] = 'passed'
            self.log(f"[SUCCESS] Demo test completed: {demo_name}")
            self.log("")

        except Exception as e:
            self.log(f"[ERROR] Demo test failed with exception: {str(e)}", "ERROR")
            result['status'] = 'failed'
            result['error'] = str(e)
            import traceback
            result['error_details'] = traceback.format_exc()

        return result

    def _run_preview(self, demo: Dict[str, Any]) -> Dict[str, Any]:
        """Run preview validation for a demo using the demo management API"""
        self.log("STEP 1: Loading Demo and Running Preview")
        self.log("-" * 70)

        start_time = time.time()

        try:
            # Step 1: Load demo from S3 (uses demo management API)
            # Use the api_name (S3 folder name) from the API mapping
            demo_api_name = demo.get('api_name')
            if not demo_api_name:
                raise Exception(f"No API name mapping for demo '{demo['display_name']}'")

            session_id, demo_info = self.api_client.call_demo_api(demo_api_name, email=self.email)

            self.log(f"Demo loaded. Session ID: {session_id}")

            # Step 2: Trigger preview validation
            preview_result = self.api_client.trigger_preview(
                session_id=session_id,
                email=self.email,
                preview_max_rows=5,
                wait_for_completion=True
            )

            elapsed_time = time.time() - start_time

            if not preview_result.get('success'):
                return {
                    'success': False,
                    'preview_time': elapsed_time,
                    'error': 'Preview validation failed'
                }

            # Extract preview data
            preview_data = preview_result.get('preview_data', {})

            # Extract cost estimate
            preview_cost = 0.0
            if 'estimated_total_cost' in preview_data:
                preview_cost = preview_data['estimated_total_cost']

            self.log(f"Preview completed in {elapsed_time:.1f}s")
            self.log(f"Estimated cost: ${preview_cost:.3f}")
            if 'total_rows' in preview_data:
                self.log(f"Total rows: {preview_data['total_rows']}")
            self.log("")

            return {
                'success': True,
                'session_id': session_id,
                'preview_time': elapsed_time,
                'preview_cost': preview_cost,
                'preview_results': preview_data
            }

        except Exception as e:
            elapsed_time = time.time() - start_time
            self.log(f"Preview error: {str(e)}", "ERROR")
            return {
                'success': False,
                'preview_time': elapsed_time,
                'error': str(e)
            }

    def _run_full_validation(self, demo: Dict[str, Any]) -> Dict[str, Any]:
        """Run full validation for a demo using the demo management API"""
        self.log("STEP 2: Full Validation")
        self.log("-" * 70)

        start_time = time.time()

        try:
            # Get session_id from demo result (or load demo again if needed)
            session_id = demo.get('session_id')

            if not session_id:
                # Load demo from S3 if we don't have a session_id yet
                # Use the api_name (S3 folder name) from the API mapping
                demo_api_name = demo.get('api_name')
                if not demo_api_name:
                    raise Exception(f"No API name mapping for demo '{demo['display_name']}'")

                session_id, demo_info = self.api_client.call_demo_api(demo_api_name, email=self.email)
                self.log(f"Demo loaded. Session ID: {session_id}")

            # Trigger full validation
            validation_result = self.api_client.trigger_full_validation(
                session_id=session_id,
                email=self.email,
                max_rows=None,  # Process all rows
                wait_for_completion=True
            )

            elapsed_time = time.time() - start_time

            if not validation_result.get('success'):
                return {
                    'success': False,
                    'full_validation_time': elapsed_time,
                    'error': 'Full validation failed'
                }

            # Extract actual cost
            full_cost = validation_result.get('total_cost', 0.0)

            # Get download URL
            download_url = validation_result.get('download_url')

            self.log(f"Full validation completed in {elapsed_time:.1f}s")
            self.log(f"Actual cost: ${full_cost:.3f}")
            self.log(f"Processed rows: {validation_result.get('processed_rows', 'unknown')}")

            if download_url:
                self.log(f"Download URL available")
            else:
                self.log("Warning: No download URL in response", "WARNING")

            self.log("")

            return {
                'success': True,
                'session_id': session_id,
                'full_validation_time': elapsed_time,
                'full_cost': full_cost,
                'full_validation_results': validation_result,
                'download_url': download_url
            }

        except Exception as e:
            elapsed_time = time.time() - start_time
            self.log(f"Validation error: {str(e)}", "ERROR")
            return {
                'success': False,
                'full_validation_time': elapsed_time,
                'error': str(e)
            }

    def _download_and_verify(self, demo: Dict[str, Any], validation_result: Dict[str, Any]) -> Dict[str, Any]:
        """Download result file from URL and verify integrity"""
        self.log("STEP 3: Download and Verify Results")
        self.log("-" * 70)

        download_url = validation_result.get('download_url')

        if not download_url:
            self.log("No download URL available", "ERROR")
            return {
                'success': False,
                'download_success': False,
                'error': 'No download URL in validation result'
            }

        try:
            # Determine output filename (overwrite existing output file)
            demo_folder = demo['folder']
            output_files = list(demo_folder.glob("*_Output.xlsx"))
            output_files.extend(list(demo_folder.glob("*_Hyperplexity_Output.xlsx")))

            if output_files:
                output_path = output_files[0]
                self.log(f"Overwriting existing output file: {output_path.name}")
            else:
                # Create new output filename
                data_file_name = demo['data_file'].stem
                output_path = demo_folder / f"{data_file_name}_Output.xlsx"
                self.log(f"Creating new output file: {output_path.name}")

            # Download file from URL
            self.log(f"Downloading from URL...")
            response = requests.get(download_url, stream=True)
            response.raise_for_status()

            # Write file
            with open(output_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)

            file_size = output_path.stat().st_size
            self.log(f"Download complete: {file_size:,} bytes")

            # Verify file integrity using openpyxl
            import openpyxl

            wb = openpyxl.load_workbook(output_path, read_only=True)
            sheet_count = len(wb.sheetnames)
            first_sheet = wb[wb.sheetnames[0]]
            row_count = first_sheet.max_row
            wb.close()

            self.log(f"File validation: {sheet_count} sheets, {row_count} rows in first sheet")
            self.log("")

            return {
                'success': True,
                'download_success': True,
                'download_path': str(output_path),
                'file_size': file_size,
                'sheet_count': sheet_count
            }

        except Exception as e:
            self.log(f"Download/verification error: {str(e)}", "ERROR")
            return {
                'success': False,
                'download_success': False,
                'error': str(e)
            }

    def run_all_tests(self):
        """Run tests for all discovered demos"""
        self.log("=" * 70)
        self.log("DEMO TEST ORCHESTRATOR")
        self.log("=" * 70)
        self.log(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log(f"Environment: {self.environment}")
        self.log(f"Email: {self.email}")
        self.log(f"Demos directory: {self.demos_dir}")
        self.log(f"Output directory: {self.output_dir}")
        self.log(f"Stop on error: {self.stop_on_error}")
        self.log(f"Skip preview: {self.skip_preview}")
        self.log(f"Skip validation: {self.skip_validation}")
        self.log("")

        # Step 1: Discover demos
        demos = self.discover_demos()

        if not demos:
            self.log("No demos found. Exiting.", "ERROR")
            return 1

        # Filter to valid demos only
        valid_demos = [d for d in demos if d['valid']]

        if not valid_demos:
            self.log("No valid demos found. Exiting.", "ERROR")
            return 1

        self.log(f"Testing {len(valid_demos)} valid demos")
        self.log("")

        # Step 2: List available demos on server and get name mapping
        demo_name_map = self.list_available_demos()

        if not demo_name_map:
            self.log("Failed to get API demo mapping. Cannot proceed.", "ERROR")
            return 1

        # Step 3: Add API names to local demos
        for demo in valid_demos:
            display_name = demo['display_name']
            if display_name in demo_name_map:
                demo['api_name'] = demo_name_map[display_name]
            else:
                self.log(f"WARNING: No API mapping found for '{display_name}'", "WARNING")
                demo['api_name'] = None

        # Step 4: Run tests for each demo
        for idx, demo in enumerate(valid_demos, 1):
            demo_name = demo['display_name']

            if not demo.get('api_name'):
                self.log(f"[{idx}/{len(valid_demos)}] SKIPPING {demo_name} - No API mapping", "WARNING")
                continue

            self.log(f"[{idx}/{len(valid_demos)}] Starting test: {demo_name}")

            # Run test
            test_result = self.test_demo(demo)

            # Add result to report
            add_demo_result(self.report, demo_name, test_result)

            # Check if we should stop on error
            if test_result['status'] == 'failed' and self.stop_on_error:
                self.log(f"Test failed and stop-on-error is enabled. Stopping.", "ERROR")
                self.log(f"Error: {test_result['error']}", "ERROR")
                break

        # Step 4: Generate and save report
        self.log("=" * 70)
        self.log("GENERATING TEST REPORT")
        self.log("=" * 70)

        # Save reports in all formats
        text_report_path = save_report(self.report, self.output_dir, format="text")
        json_report_path = save_report(self.report, self.output_dir, format="json")
        html_report_path = save_report(self.report, self.output_dir, format="html")

        self.log(f"Text report: {text_report_path}")
        self.log(f"JSON report: {json_report_path}")
        self.log(f"HTML report: {html_report_path}")
        self.log("")

        # Print report to console
        print_report(self.report)

        # Determine exit code
        summary = self.report['summary']
        if summary['failed'] > 0:
            self.log(f"Tests completed with {summary['failed']} failure(s)", "ERROR")
            return 1
        else:
            self.log(f"All tests passed! ({summary['passed']}/{summary['total']})")
            return 0


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Demo Test Orchestrator - End-to-end testing of all demo files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Run all demos in dev environment (default)
  python test_all_demos.py

  # Run with custom options
  python test_all_demos.py --demos-dir ./demos --email test@example.com \
      --environment dev --output-dir ./test_results --stop-on-error

  # Skip preview step (only run full validation)
  python test_all_demos.py --skip-preview

  # Skip validation step (only run preview)
  python test_all_demos.py --skip-validation

  # Continue on errors
  python test_all_demos.py --no-stop-on-error
        """
    )

    parser.add_argument(
        '--demos-dir',
        type=str,
        default='./demos',
        help='Path to demos directory (default: ./demos)'
    )

    parser.add_argument(
        '--email',
        type=str,
        default='eliyahu@eliyahu.ai',
        help='Test email address (default: eliyahu@eliyahu.ai)'
    )

    parser.add_argument(
        '--environment',
        type=str,
        default='dev',
        choices=['dev', 'test', 'staging', 'prod'],
        help='Environment to test (default: dev)'
    )

    parser.add_argument(
        '--output-dir',
        type=str,
        default='./test_results',
        help='Output directory for test results (default: ./test_results)'
    )

    parser.add_argument(
        '--stop-on-error',
        action='store_true',
        default=True,
        help='Stop on first error (default: True)'
    )

    parser.add_argument(
        '--no-stop-on-error',
        dest='stop_on_error',
        action='store_false',
        help='Continue testing even if errors occur'
    )

    parser.add_argument(
        '--skip-preview',
        action='store_true',
        help='Skip preview step (only run full validation)'
    )

    parser.add_argument(
        '--skip-validation',
        action='store_true',
        help='Skip validation step (only run preview)'
    )

    args = parser.parse_args()

    # Validate arguments
    if args.skip_preview and args.skip_validation:
        print("[ERROR] Cannot skip both preview and validation")
        return 1

    # Create orchestrator
    orchestrator = DemoTestOrchestrator(
        demos_dir=Path(args.demos_dir),
        email=args.email,
        environment=args.environment,
        output_dir=Path(args.output_dir),
        stop_on_error=args.stop_on_error,
        skip_preview=args.skip_preview,
        skip_validation=args.skip_validation
    )

    # Run tests
    try:
        return orchestrator.run_all_tests()
    except KeyboardInterrupt:
        print("\n[INFO] Testing interrupted by user")
        return 130
    except Exception as e:
        print(f"\n[ERROR] Fatal error: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
