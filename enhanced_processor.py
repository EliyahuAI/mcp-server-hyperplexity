#!/usr/bin/env python
"""
Enhanced Excel processor that adds missing columns and ensures examples are passed.
"""

import os
import sys
import json
import yaml
import pandas as pd
import numpy as np
import logging
from pathlib import Path
from typing import Dict, List, Union, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configure logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Add src directory to path if needed
src_dir = Path("src")
if src_dir.exists():
    sys.path.append(str(src_dir.absolute()))

# Import our modules
from src.excel_processor import ExcelProcessor, COLOR_HIGH, COLOR_MEDIUM, COLOR_LOW, COLOR_NEUTRAL

class EnhancedExcelProcessor(ExcelProcessor):
    """An enhanced version of ExcelProcessor that handles all column config fields."""
    
    def __init__(self, input_file: str, config_file: str = None, row_filter: Union[List[int], str] = None):
        """
        Initialize with input file and optional config file.
        
        Args:
            input_file: Path to the Excel or CSV file
            config_file: Path to the configuration file
            row_filter: Optional list of row indices to process or a filter expression
        """
        # Load raw config first
        self.raw_config = None
        if config_file and Path(config_file).exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    if Path(config_file).suffix.lower() == '.json':
                        self.raw_config = json.load(f)
                    else:  # Assume YAML
                        self.raw_config = yaml.safe_load(f)
            except Exception as e:
                logger.error(f"Failed to load raw config file: {e}")
        
        # Now proceed with normal initialization
        super().__init__(input_file, config_file, row_filter)
    
    def update_with_validation(self, validation_results: Dict, output_file: str = None) -> None:
        """
        Update the Excel file with validation results, adding missing columns from validation targets.
        
        Args:
            validation_results: Dictionary of validation results from lambda function
            output_file: Path to save the updated Excel file
        """
        # If no output file is specified, create one in the same directory as the input file
        if not output_file:
            stem = self.input_file.stem
            output_file = self.input_file.parent / f"{stem}_validated.xlsx"

        # Create a workbook and add the main sheet
        wb = openpyxl.Workbook()
        main_sheet = wb.active
        main_sheet.title = "Validated Data"
        
        # Get all columns from validation targets in the raw config
        all_columns = list(self.df.columns)  # Start with columns in the DataFrame
        
        # Add columns from validation targets that aren't in the original DataFrame
        if self.raw_config and 'columns' in self.raw_config:
            for col_name in self.raw_config['columns']:
                if col_name not in all_columns:
                    all_columns.append(col_name)
                    logger.info(f"Added missing column from config: {col_name}")
        
        # Add column headers
        for col_num, col_name in enumerate(all_columns, 1):
            cell = main_sheet.cell(row=1, column=col_num, value=col_name)
            cell.font = Font(bold=True)
            
        # Process the validation results and update cells
        data = validation_results.get('data', {})
        
        # Set up detailed view data
        details_rows = []
        details_cols = ["Row Key", "Column", "Original Value", "Validated Value", 
                        "Confidence", "Needs Update", "Quote", "Sources", "Explanation"]
        
        # Process primary key to build row keys
        primary_key = self.config.get('primary_key', ['id'])
        
        # Populate data and apply styles
        for row_idx, df_row in enumerate(self.filtered_df.iterrows(), 2):
            df_row_idx, row_data = df_row
            
            # Build row key
            row_key_parts = []
            for key in primary_key:
                if key in row_data:
                    value = row_data[key]
                    if value is None or pd.isna(value) or value == '':
                        value = '[empty]'
                    row_key_parts.append(str(value))
                else:
                    row_key_parts.append('[missing]')
            row_key = '|'.join(row_key_parts)
            
            # Check if we have validation results for this row
            if row_key not in data:
                logger.warning(f"No validation results found for row key: {row_key}")
                
                # Add row values to Excel without validation formatting
                for col_idx, col_name in enumerate(all_columns, 1):
                    orig_value = row_data.get(col_name, "")
                    cell = main_sheet.cell(row=row_idx, column=col_idx, value=orig_value)
                    cell.fill = PatternFill(start_color=COLOR_NEUTRAL, end_color=COLOR_NEUTRAL, fill_type="solid")
                continue
            
            # Get validation results for this row
            row_results = data[row_key]
            
            # Add row values to Excel
            for col_idx, col_name in enumerate(all_columns, 1):
                # Get the original value from DataFrame
                orig_value = row_data.get(col_name, "")
                cell = main_sheet.cell(row=row_idx, column=col_idx, value=orig_value)
                
                # Check if this column has validation results
                column_validated = False
                if 'validation_results' in row_results and col_name in row_results['validation_results']:
                    val_result = row_results['validation_results'][col_name]
                    
                    # Get validation details - try both 'answer' and 'value' fields
                    validated_value = val_result.get('answer', val_result.get('value', ''))
                    confidence = val_result.get('confidence', 0)
                    confidence_level = val_result.get('confidence_level', '')
                    quote = val_result.get('quote', '')
                    sources = val_result.get('sources', [])
                    update_required = val_result.get('update_required', False)
                    explanation = val_result.get('explanation', '')
                    
                    # Always show validated value when it exists, regardless of update_required flag
                    if validated_value and validated_value != "":
                        cell.value = validated_value
                    
                    # Apply cell formatting based on confidence
                    if confidence_level == "HIGH":
                        cell.fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid")
                    elif confidence_level == "MEDIUM":
                        cell.fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type="solid")
                    elif confidence_level == "LOW":
                        cell.fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type="solid")
                    
                    # Add comment with quote and sources
                    comment_text = ""
                    if quote:
                        comment_text += f"Quote: {quote}\n\n"
                    if sources:
                        comment_text += f"Sources: {', '.join(sources)}\n\n"
                    if update_required:
                        comment_text += f"Update required: Yes\n"
                    else:
                        comment_text += f"Update required: No\n"
                    
                    if comment_text:
                        cell.comment = Comment(comment_text, "Validation")
                    
                    column_validated = True
                    
                    # Add to details data for the detailed view sheet
                    details_rows.append({
                        "Row Key": row_key,
                        "Column": col_name,
                        "Original Value": str(orig_value),
                        "Validated Value": validated_value,
                        "Confidence": confidence,
                        "Needs Update": "Yes" if update_required else "No",
                        "Quote": quote,
                        "Sources": "; ".join(sources),
                        "Explanation": explanation
                    })
                
                # If column was not validated, use neutral styling
                if not column_validated:
                    cell.fill = PatternFill(start_color=COLOR_NEUTRAL, end_color=COLOR_NEUTRAL, fill_type="solid")
        
        # Add a legend for the color coding
        legend_row = len(self.filtered_df) + 3
        main_sheet.cell(row=legend_row, column=1, value="Color Legend:").font = Font(bold=True)
        
        # High confidence
        legend_cell = main_sheet.cell(row=legend_row+1, column=1, value="High Confidence")
        legend_cell.fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid")
        
        # Medium confidence
        legend_cell = main_sheet.cell(row=legend_row+2, column=1, value="Medium Confidence")
        legend_cell.fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type="solid")
        
        # Low confidence
        legend_cell = main_sheet.cell(row=legend_row+3, column=1, value="Low Confidence")
        legend_cell.fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type="solid")
        
        # No validation
        legend_cell = main_sheet.cell(row=legend_row+4, column=1, value="Not Validated")
        legend_cell.fill = PatternFill(start_color=COLOR_NEUTRAL, end_color=COLOR_NEUTRAL, fill_type="solid")
        
        # Auto-adjust column widths
        for col in main_sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 10)  # Add padding
            main_sheet.column_dimensions[column].width = adjusted_width
        
        # Create the detailed view sheet if requested
        if self.config.get('detailed_view', True) and details_rows:
            self._create_detailed_sheet(wb, details_rows, details_cols)
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Updated Excel file saved to {output_file}")

def process_file(input_file: str, config_file: str, validation_json: str, output_file: str = None, row_filter: Union[List[int], str] = None) -> None:
    """
    Process a file with validation results using the enhanced processor.
    
    Args:
        input_file: Path to the input Excel or CSV file
        config_file: Path to the configuration file
        validation_json: Path to validation results JSON or JSON string
        output_file: Path to save the updated Excel file
        row_filter: Optional list of row indices or filter expression
    """
    # Initialize the processor
    processor = EnhancedExcelProcessor(input_file, config_file, row_filter)
    
    # Load validation results
    if os.path.exists(validation_json):
        with open(validation_json, 'r', encoding='utf-8') as f:
            validation_results = json.load(f)
    else:
        try:
            validation_results = json.loads(validation_json)
        except json.JSONDecodeError:
            logger.error("Invalid JSON for validation results")
            raise ValueError("Invalid JSON for validation results")
    
    # Update Excel with validation results
    processor.update_with_validation(validation_results, output_file)
    
    # Return the path to the updated file
    if not output_file:
        stem = Path(input_file).stem
        output_file = Path(input_file).parent / f"{stem}_validated.xlsx"
    
    logger.info(f"Processing complete. Output file: {output_file}") 