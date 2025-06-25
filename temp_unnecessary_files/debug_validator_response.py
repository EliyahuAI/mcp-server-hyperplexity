#!/usr/bin/env python3
"""
Debug script to see exactly what the validator Lambda returns
"""
import boto3
import json
import openpyxl
import io

# Test the validator Lambda directly with our test data
def test_validator_directly():
    """Call the validator Lambda directly to see its response"""
    
    lambda_client = boto3.client('lambda')
    
    # Read our test files
    with open('test_cases/simple_config.json', 'r') as f:
        config_data = json.load(f)
    
    # Read Excel file and extract data
    wb = openpyxl.load_workbook('test_cases/real_excel.xlsx')
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"📋 Excel headers: {headers}")
    
    # Get first data row
    row_data = {}
    for col_idx, header in enumerate(headers):
        if header:
            cell_value = ws.cell(row=2, column=col_idx + 1).value
            row_data[header] = str(cell_value) if cell_value is not None else ""
    
    print(f"📊 First row data: {row_data}")
    
    # Create row key
    id_fields = []
    for target in config_data.get('validation_targets', []):
        if target.get('importance') == 'ID':
            id_fields.append(target['column'])
    
    key_columns = id_fields[:3] if len(id_fields) >= 3 else headers[:3]
    row_key = "||".join([row_data.get(col, "") for col in key_columns])
    row_data['_row_key'] = row_key
    
    print(f"🔑 Row key: {row_key}")
    
    # Create payload exactly like interface lambda does
    payload = {
        "test_mode": True,  # Preview mode
        "config": config_data,
        "validation_data": {
            "rows": [row_data]
        },
        "validation_history": {}
    }
    
    print(f"📤 Sending payload to validator:")
    print(f"   - test_mode: {payload['test_mode']}")
    print(f"   - config keys: {list(payload['config'].keys())}")
    print(f"   - validation_data keys: {list(payload['validation_data'].keys())}")
    print(f"   - rows count: {len(payload['validation_data']['rows'])}")
    print(f"   - first row keys: {list(payload['validation_data']['rows'][0].keys())}")
    
    try:
        # Call validator Lambda
        response = lambda_client.invoke(
            FunctionName='perplexity-validator',
            InvocationType='RequestResponse',
            Payload=json.dumps(payload)
        )
        
        print(f"🔄 Lambda response status: {response['StatusCode']}")
        
        # Parse response
        response_payload = json.loads(response['Payload'].read().decode('utf-8'))
        
        print(f"📥 Raw validator response:")
        print(json.dumps(response_payload, indent=2, default=str))
        
        # Check specific fields
        if isinstance(response_payload, dict):
            print(f"\n🔍 Response analysis:")
            print(f"   - Type: {type(response_payload)}")
            print(f"   - Keys: {list(response_payload.keys())}")
            
            if 'validation_results' in response_payload:
                val_results = response_payload['validation_results']
                print(f"   - validation_results type: {type(val_results)}")
                if val_results:
                    print(f"   - validation_results keys: {list(val_results.keys()) if isinstance(val_results, dict) else 'Not a dict'}")
                else:
                    print(f"   - validation_results is empty/None")
            else:
                print(f"   - NO 'validation_results' key found")
                
            if 'body' in response_payload:
                print(f"   - Found 'body' key, checking inside...")
                body = response_payload['body']
                if isinstance(body, str):
                    try:
                        body = json.loads(body)
                        print(f"   - Parsed body keys: {list(body.keys())}")
                    except:
                        print(f"   - Body is string, not JSON")
                elif isinstance(body, dict):
                    print(f"   - Body keys: {list(body.keys())}")
        
    except Exception as e:
        print(f"❌ Error calling validator: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    print("🔍 DEBUGGING VALIDATOR LAMBDA RESPONSE")
    print("=" * 60)
    test_validator_directly() 