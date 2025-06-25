#!/usr/bin/env python3
"""
Script to create proper test files that match the actual validator configuration
"""
import json
import openpyxl
import os

def create_real_test_excel():
    """Create an Excel file with columns matching the config file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers that match the config
    headers = [
        "Product Name",
        "Developer", 
        "Target",
        "Indication",
        "Therapeutic Radionuclide",
        "Diagnostic Radionuclide",
        "Modality Type",
        "Development Stage",
        "Key Trial ID",
        "Projected Launch",
        "FDA-EMA Designation",
        "Strategic Partners",
        "Recent News"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add sample data row
    sample_data = [
        "FAP-2286",
        "Clovis Oncology",
        "FAP",
        "Solid tumors",
        "Lu-177",
        "Ga-68",
        "Theranostic peptide",
        "Phase 2",
        "NCT04939610",
        "2026",
        "Fast Track (FDA, 2023)",
        "Lantheus",
        "- 12/15/2024: Phase 2 interim results positive (HIGH)\n- 11/30/2024: FDA Fast Track designation granted (HIGH)"
    ]
    
    # Write data row
    for col, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Save file
    os.makedirs('test_cases', exist_ok=True)
    wb.save('test_cases/real_excel.xlsx')
    print("✅ Created real_excel.xlsx with proper column structure")

def create_simple_test_config():
    """Create a simplified config for testing basic validation"""
    
    config = {
        "general_notes": "Test configuration for radiopharmaceutical validation",
        "default_model": "sonar-pro",
        "validation_targets": [
            {
                "column": "Product Name",
                "description": "Official code or INN of the radiopharmaceutical",
                "importance": "ID",
                "format": "String",
                "notes": "Use sponsor's current nomenclature",
                "examples": ["FAP-2286", "225Ac-PSMA-617"],
                "search_group": 0
            },
            {
                "column": "Developer",
                "description": "Lead company developing the product",
                "importance": "ID", 
                "format": "String",
                "notes": "Name mergers or major co-developers",
                "examples": ["Novartis", "POINT Biopharma"],
                "search_group": 0
            },
            {
                "column": "Target",
                "description": "Molecular target or receptor",
                "importance": "CRITICAL",
                "format": "String",
                "notes": "Use gene/protein symbol",
                "examples": ["FAP", "PSMA", "SSTR2"],
                "search_group": 1
            },
            {
                "column": "Indication",
                "description": "Main disease or tumor type",
                "importance": "CRITICAL",
                "format": "String", 
                "notes": "Concise but specific",
                "examples": ["mCRPC", "GEP-NETs"],
                "search_group": 1
            }
        ]
    }
    
    # Save config
    os.makedirs('test_cases', exist_ok=True)
    with open('test_cases/simple_config.json', 'w') as f:
        json.dump(config, f, indent=2)
    print("✅ Created simple_config.json for basic testing")

if __name__ == "__main__":
    create_real_test_excel()
    create_simple_test_config()
    print("\n📁 Test files created in test_cases/")
    print("   - real_excel.xlsx (matches config columns)")
    print("   - simple_config.json (simplified for testing)") 